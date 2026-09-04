# =============================================================================
# app.py  -  Kombinierter Generator: Suche + Fahrzeugwäsche  ->  eine app.html
# =============================================================================
# Die erzeugte suche.html ist vollstaendig eigenstaendig.
# Alle PDF-Dokumente werden komprimiert in die einzelne HTML-Datei eingebettet.
# Starten:  streamlit run app.py
# =============================================================================

from __future__ import annotations

# Native Bibliotheken auf einen Thread begrenzen. Das reduziert Startspitzen
# und vermeidet instabile OpenMP-/BLAS-Konstellationen auf kleinen Cloud-VMs.
import os
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")
os.environ.setdefault("ARROW_NUM_THREADS", "1")
os.environ.setdefault("MALLOC_ARENA_MAX", "2")

import faulthandler
try:
    faulthandler.enable(all_threads=True)
except Exception:
    pass

def _boot_log(message: str) -> None:
    print(f"[NFC-BOOT] {message}", flush=True)

_boot_log("01 Python-Skript gestartet")

import streamlit as st
_boot_log("02 Streamlit importiert")

import importlib
import importlib.util
import json
import base64
import unicodedata
import re
import datetime
import time
import hashlib
import io
import zlib
from pathlib import Path
from typing import List


class _LazyPandas:
    """Importiert pandas erst bei der ersten tatsächlichen Dateiverarbeitung."""
    _module = None

    def _load(self):
        if self._module is None:
            _boot_log("PANDAS-START: pandas wird jetzt importiert")
            self._module = importlib.import_module("pandas")
            _boot_log(f"PANDAS-OK: pandas {getattr(self._module, '__version__', '?')} importiert")
        return self._module

    def __getattr__(self, name):
        return getattr(self._load(), name)


pd = _LazyPandas()
_boot_log("03 Standardimporte bereit; pandas wird verzögert geladen")

st.set_page_config(page_title="NFC Generator v50", layout="wide")
_boot_log("04 Seitenkonfiguration gesetzt")

APP_CACHE_VERSION = "waschen-tanken-dashboard-2026-09-04-v50-reisekosten-vorschau"
EXTRA_CACHE_VERSION = "extra-parser-2026-09-04-v50-reisekosten-vorschau"
APP_DISPLAY_VERSION = "50"
APP_DISPLAY_NAME = "NFC Generator"



# =============================================================================
# EXCEL-ENGINE — calamine ist 5-10x schneller als openpyxl beim Lesen.
# Wenn 'python-calamine' installiert ist, nutzen wir es; sonst Fallback.
# Schreiben bleibt openpyxl (calamine kann nur lesen).
# =============================================================================
# Nur die Verfügbarkeit prüfen; kein nativer Import während des App-Starts.
_HAS_CALAMINE = importlib.util.find_spec("python_calamine") is not None
EXCEL_READ_ENGINE = "calamine" if _HAS_CALAMINE else "openpyxl"
_boot_log(f"05 Excel-Engine vorbereitet: {EXCEL_READ_ENGINE}")

EXCLUDED_DRIVER_NAMES = (
    "Ch.Holtz", "Paasch", "Meyer", "Ihde", "Spedition M+S Express 4", "Spedition M+S Express 3",
    "Spedition M+S Express 2", "Spedition M+S Express 1", "Spedition Meyer 1", "Spedition Meyer 2",
    "Spedition Meyer 3", "Spedition Meyer 4", "Spedition Meyer 5", "Spedition Meyer 6",
    "Spedition Meyer 7 (36er)", "Spedition Meyer 8", "Spedition Meyer Sz.", "Paasch & Reinke 1",
    "Paasch & Reinke 2", "Paasch & Reinke 3", "deVries", "Spedition Ihde", "Insellogistik 1",
    "Insellogistik 2", "Zippel Logistik T1", "Zippel Logistik T2", "Zippel Logistik T3",
    "Ch. Holtz T1", "Ch. Holtz T2", "Ch. Holtz T3", "T&D Spedition", "Kudex 1", "Kudex 2", "FP Fleischwerk",
)






def read_upload_bytes(uploaded_file) -> bytes:
    if uploaded_file is None:
        return b""
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    data = uploaded_file.read()
    if data is None:
        return b""
    if isinstance(data, str):
        return data.encode("utf-8")
    return data


def upload_signature(uploaded_file) -> str:
    """Liefert eine eindeutige Signatur fuer ein UploadedFile.

    Streamlit-UploadedFile-Objekte haben eine ``file_id`` bzw. ``_file_urls``,
    die sich nur dann aendert, wenn der Nutzer eine neue Datei hochlaedt.
    Wir benutzen Name + Size + file_id als billigen Schluessel und sparen uns
    das vollstaendige Auslesen + SHA1 ueber 10-50 MB pro Re-Run.
    Fallback ist die alte SHA1-Methode, falls kein file_id verfuegbar ist.
    """
    if uploaded_file is None:
        return ""
    name = getattr(uploaded_file, "name", "") or ""
    size = getattr(uploaded_file, "size", 0) or 0
    fid = getattr(uploaded_file, "file_id", None) or getattr(uploaded_file, "_file_urls", None) or ""
    if fid:
        return f"{name}|{size}|{fid}"
    # Fallback: SHA1 ueber Inhalt
    payload = read_upload_bytes(uploaded_file)
    digest = hashlib.sha1()
    digest.update(name.encode("utf-8", errors="ignore"))
    digest.update(b"|")
    digest.update(payload)
    return digest.hexdigest()


def uploads_signature(uploaded_files) -> str:
    digest = hashlib.sha1()
    for uploaded_file in uploaded_files or []:
        digest.update(upload_signature(uploaded_file).encode("utf-8", errors="ignore"))
        digest.update(b"|")
    return digest.hexdigest()


def combine_signatures(*parts: str) -> str:
    digest = hashlib.sha1()
    for part in parts:
        digest.update((part or "").encode("utf-8", errors="ignore"))
        digest.update(b"|")
    return digest.hexdigest()


def is_excluded_driver(name: str) -> bool:
    name = str(name)
    return any(excluded in name for excluded in EXCLUDED_DRIVER_NAMES)


def get_cached_export_b64(cache_key: str, source_value: str, builder) -> str:
    cache = st.session_state.setdefault("_export_b64_cache", {})
    cached = cache.get(cache_key)
    if cached and cached.get("source") == source_value:
        return cached.get("b64", "")

    payload = builder(source_value) or b""
    encoded = base64.b64encode(payload).decode("ascii") if payload else ""
    cache[cache_key] = {"source": source_value, "b64": encoded}
    return encoded




# =============================================================================
# CLOUD-STABILE ASSET-VERWALTUNG
# Große PDF-, HTML- und JavaScript-Daten liegen nicht mehr als Python-Literale
# in dieser Datei. Das senkt den RAM-Bedarf des Python-Parsers massiv.
# Die erzeugte suche.html bleibt weiterhin eine einzige, eigenständige Datei.
# =============================================================================
_ASSET_DIR = Path(__file__).resolve().parent / "nfc_assets"
_STATIC_ASSET_PATH = _ASSET_DIR / "static_payload.json.zlib"
_PDF_ASSET_PATH = _ASSET_DIR / "embedded_pdfs.json.zlib"
_ASSET_META_PATH = _ASSET_DIR / "asset_meta.json"
_STATIC_PAYLOAD_CACHE = None
_PDF_DOCUMENTS_CACHE = None
_ASSET_META_CACHE = None


def _read_zlib_json(path: Path, label: str):
    if not path.is_file():
        raise FileNotFoundError(
            f"{label} fehlt: {path}. Bitte den Ordner 'nfc_assets' zusammen "
            "mit suche_druck_quell_datei.py ins GitHub-Repository hochladen."
        )
    packed = path.read_bytes()
    try:
        raw = zlib.decompress(packed)
        return json.loads(raw.decode("utf-8"))
    except Exception as exc:
        raise RuntimeError(f"{label} ist beschädigt oder unlesbar: {path}") from exc


def _static_payload() -> dict:
    global _STATIC_PAYLOAD_CACHE
    if _STATIC_PAYLOAD_CACHE is None:
        value = _read_zlib_json(_STATIC_ASSET_PATH, "Statische NFC-Daten")
        if not isinstance(value, dict):
            raise RuntimeError("Statische NFC-Daten haben ein ungültiges Format.")
        _STATIC_PAYLOAD_CACHE = value
    return _STATIC_PAYLOAD_CACHE


def _static_payload_text(key: str) -> str:
    value = _static_payload().get(key)
    if not isinstance(value, str):
        raise KeyError(f"Statischer NFC-Baustein fehlt: {key}")
    return value


_REINIGUNGSNACHWEIS_PDF_B64 = (
    "eNrcfQV4VMcXL20JEpziungCZHNtDQo0IQRSnAQPsiQbQRLYJAQNxaVAcXeneLDg7looBAsQ3B2K867M3N27uXNnSPf/3ve99gsr555zfnfkzG/mnp2p1Cwg"
    "0IfWs56Vbt0+c96T1lG62C5dPX/5xdO3mT02PCHMZtd59ezbMzzC27N2bU9bTLggZpwuC+nb06bzbWaNtMV5+taNTYiJ19Gevg2jw+N07XUcf2ELXQcnVTaD"
    "al1rvLV7bKSnZEMn2G7hpMCp+vL0bWGLi02wh/EagqRen/j6wfHWeJv4qX6wUWcQ7fBvLTojNOkbGMvDE64IpHUm6YJARkdT4C2ro1n52jZNu3S1hUmXB/Xg"
    "fRp0tMMSXzphwbZ4/h4loT94rQteg3icAYE63xBbn3ipAHwb28Kjrf6xfXgdSk+Jf2aO1pvNFpOR0xksBj1jMhgsNH85X5Ax8baY+DgdDWD6xcTExgtF2kEo"
    "KDsvy1BShgwlJZeKp69/Y51vk1h7D2t3T98wK19FDj0jqV5dP4WeKYOeULyevsEJXeLFjyH2BJsg8PRtYu0hiPmG4W+Ns4m14Otnj7Z2bxzC111MWGx4dEyk"
    "zrd1dIxfTFw0/EKqrwBbXJg9umd8rF1nBvUUbY+LrxtltetYxtO3kRV8YAz8p9bR4fFRcTqLS9mYVbE6TEuuJJgysMDu1sg40UcQ30Sjw/xiIrvbdBRfGXFh"
    "QgVYKIOnr2BD+ODD0LykrrVnA1t0ZFS8zsSY+Qt7R4qAdBzH33tjax/pE2M0GiSPraWLOYrXbQM0GQP/oZHNKpYJy/IFGm/r0Yo3Ian4S03Ih7chOtUxFEUJ"
    "7hT9zAJuuL2OMQk37/w/y7LSn1m4SpALn4RX4c9gMMp/0vXSO+lfaM1g5oDUaDSJfyaGEf/EzzTfv8QrpeuF7wR9M+9HuEb6hlJeD95bOMmy8J0zahkX5Xjv"
    "jJfh9YU/QS68Cr6U9yLcpSiFWhQF0FDi+/9f/ncuOUctkv4vaHSAzUgMjd/XyZkMndzHP7Z7+Pf1dPEO0F2dNhnlrk67xkGa/t7OLgN0Z483WRQ9XhA69XgT"
    "SY83ufR4vo+jezzNEHR5xydHN1BrJI5rHf/DruLcvOB36leqN0bhX+FbZ9/CZ+EPIhde+YFR/l7EJIYrAZnC5v/VjssHaqfOwX5352CdO0dIdA9bXBNbYovY"
    "HtaYZsHf2UM4zR7COsZC2uDaQbjv6CAqIHG9xGyhFb3E6NxLjBbWuZdQynGR5gjHRY6Rewml7CUGo9hL+AGB4oyCO2UvMTh6icG5LjOyIEAAnapTZHagXCW6"
    "AaDRYiSI7R5rD+5pDeOvDLD1jg6ztajvz9d3dHxcM5u9bmyPnrExYvHw5cazO3vP2O4CYY2wdo+zCfXYnf9OJxRuvC3Axte6TbjbmEjelZlhWOEe4uLtNmsP"
    "zz5znnbI3ibkSJGO/fNsuZT/4IjCzUqN8ig8asWGeos3VJ71gG1d79CIPD8UCAjZm6NVyPoFC4ZMXdQm/4DD+/Utpuh3JU7bUsh6Yxt7wta29di+03bu/LDn"
    "9Uuff+fVjIu4N5+9f3JX/61bpg259S1XF8/M/Dem6YhFr2YtWjTGt4eHx6FI/57DjnaY0KHth6VPFyx4f9DLe+W402t3Fmn1qMSU0rOCXy09tXDB+2Er/vpr"
    "e3mvGgea9sxVtfyk4NAvS8ctWvD+ff7RAZ+X2Ndditd9sCR51n8/2/d8t65dd29KKFSo7A9770w6fmV/sindUqnoEHvo+w33P0S/GBxW82jxpAGf2/dq3NfW"
    "OCpi5ZE/FvlaDYe+Ve3ardtwf2v8PDOV/Llj+LH5E4vOvPDw4cNeHi2HF1taLP/Iz8dK7Bs3tHZosfmzgjx3lilwu8j+j3GpC6pP+MPvAzOsSvng8Z0Cjn0+"
    "MXG7fuHLvvFVRi6iQ0JSKxYNqR9Z5Hqu0zWPNB1+c9/U+QnlvM4MLxs3bX7WAZOXXW6btOifkftff47/bPng30z/4ESOlFtbR3u27pD7+NAuQcf0n81eAe3K"
    "Vp2aOHVTnpyRHvRLi+Xg13PHZ+1YlysppeDQOVl1h3fn7LSljvfX01laDzD9kupjfrzkpsfBfEWHeo7K33VukYAsdW6cqmVZfvrl+RTLj+FfF3TrNc63ZGTN"
    "RkGUb/XGIR0q9N0zZ1CjmeNf5/Kpsq9Boz75nxXy8Bg5qWuvBmMeBLbJ1ifkxZlx55/8kWu1+fSQFdXKeVgf2l5+uDDsbGzXrmkH7NP6HXyzZnGNiQN/KNAp"
    "Jc77a0T+lqmvNlYrN8xcKcftzx9GTqZaB94p8jb1+Mj1I6fknH9mxO43uRvuep28IiVlxqVObTvM7Rhq/aRb+nREtpAqi5559Sk19bWt/4xbz4eaYq+/+Pm3"
    "awcmdfgw86Zl9LcBs+185cWYW2W/aHldwW/J7Q/V/PYvaH3560r/4n2/1njc3H9zZLMNufsfy340703Wu3HggvKRSZvetEzsmKtCi+DgMWOY4Rf2V6W3Fp45"
    "qwK3bXrKm64p74Is3avculX07wa3Qlu16l3uhX+CIe3Mt0F+Zx/nTl1w8vDA6ncm/fr2cXp6vi5xZQ5NrXcv9fehX9rERmR71+rCwuTkSQN+n/l+S+VSJ71W"
    "j/Gp/rpp9TfjjnNvG898kXdSHJNa61VQo+s7/mpyqvGuzzeTn70r13Re6AevHgvfz8/SpOSNmRNuxW/28q61cc5S609lfmR/fn1i7fAcfqEH9v/iv/FoZJOY"
    "4UVHFh612POvYf7W6OiH2U8VWLFqyZKK9drkvJsy8nDJHSNO9cozO752SGqlgCN3693+5f6driGhbZd+rhnceefs2cey3s/5wKvd813zdrf98i37tm0XIj/k"
    "LHrDM0J3Z/feMwMLzp/UokXTEZuSS5wZlXtk5MqSAyfn/lrq8Z3jAVX2NerLPN3jFdwuaWf1ngUqFp4y7XgP+3jqLfftYKnlo4/mHBxVNnX3t1tf7z+KHeP7"
    "5/KcttgNtW4YzF/fNSxX9MaYv7reOpectOtlrxwfHobP2Xjzfly5jlb/bQtrBjTzqOQRk7pmYmn/vKPylwmf9Eft6TmCP99u3iG4f6WKbWe9uP/ee2FQUD1b"
    "wS1Fp+3akLtBuz6WFo8bvt5G/1Uzp/GD54DyTMeihY6fqpu+4MbjmssiLz6s+TAuZPqoPDVKpdx//a7GoekBHf2+FDluSXo1d9T+0Il3QlomjGu9+sWe8Aon"
    "C30Zwbzb49Vj5+TOkzbe7J+4b9GJda1CB9kelB540NdorLVrkf/p21mfhBUsb436bUXD308UWu59tNix5NIH857aReWxjztz5WGhToXm/fhu9JBcfpa1fm2L"
    "9/1gaLHV++fXVbNdTspSblCvgjNuFJhzt2i2KhG7Th9v6HO0ceFWS3JWs5f4bdn68lMnD/Fv1vBo4cCl94dXijha2Fx4bjlvn4ZtJrZuUW71Cq91FS9v+rya"
    "rVLlyInH7yv/+3zP1S9pUcv/frD+ly1Dn9T99OxptQKFk5btXva+pK7S8ju565TJ8Y3bsn6Jf++42737tLpdMdVShdr25dWbyQcjo6MLL8p2uXx+pk6dOpV9"
    "K1evVr36vQKfti48uPTI0kLL7ty5P2XKlN+nf10wrteoC+l/rV69+sG1B0dXLNhhTdx1IHTuzrmrQ7eue9vxQKBvRI4vX/POCyu+Md+aWpsftT3ZY3vr19+y"
    "bb0/fPnQureTj9+Z1+HVs56Hbsa/WFn/8rnjE8q0Cmkdfjdizp6EpOGT8yYHRhWelpDnQ+DXDoeCjnw8U4UqOfesZ6tWOytNHlS+Uindx7RKrXfNvmS/k/Xj"
    "vfXTTu/b4Ntp9+fqpRu9/Hf/xvfnRlo9Pjzvs7bI+BOVPvvf7/y89czyE0edr161qv1l4uWb662LzgeEm/JfTtLlq/NnqV/b1dzUyHeWZ8dPLVsdv7F3yfvI"
    "CrvWM4vWb7B9CQ0dk/tqZGifhInzvdZMnjI9fu/zLMkPPy1vk5ijXP+L50uYFz6MiNpdvF3nEdPaLrZ9+pvJt+nJiCmlDmT/d3GlNnnC5xWfle3WmN0J/tub"
    "NJhdKOpbx78DjgSejJ/x5vHul9M/l67TY2SLz0//7lMq1/2L1/qd2rDrvHHwrhuPhyR/rLneY9GQzbu7JTR9nNwoZs6jx21n3hl2LXbhyDPNuuuPn6z74e6K"
    "naVGH5/4Pk/BP254x64v0aCt5fhP3vUe9vUy5Dt0d2GbCSGFTtQ3fQ5f8iVPnaklk25ce3G1yJtXMd3frdq/qHqJTcE7i16+22VXjf3r3njmOtGj56Aq67h8"
    "bcvfeR8z73nd20/+PjPy7IZqQ781CFy6tF+Xl1knRo1Pudig+vU/DgVW3/0ir6X18vPNuvocMpxoO7F/K/vTXBvMbUMbzv958fvOVMKwPVX7fHnX9lvWb4dL"
    "7aCvfCs2enL/Pw7NOnQo6VbX7W8bv0w8kpayf22THw9X+do2dHfo8J++Zr9+PZ9fzb9vTX5w7+eQkJTQ0MdV1jb+dDIiuNlnQ6ExAZOfduwTGDCq2506XTt+"
    "Wdbv6bjN54aVTch66srgnWWa5k/1f+K3p8dvL/cMGzPGM/Xs2Sn7XwXGDyu8v9fp2xWHba37wMvz5uJJkyrdmlWs15o5R9K8vb6eXnEoKvSI75vyNczFVuRN"
    "Ti3dcmfxSkXOr2vX7HPYD7n2rQi+cjN9efeA0SsfnCo58kq7j54hYwYxa1/7Rq/z77/z0727X1q0L/lW5BKb8rf8o8jzuYWKrxz374GXBSauWL9+/a9e1fVr"
    "J3xKu3Dmp1FHUp6fS/Jjvb5aV6xcmT7S98ypmRWKvB4cdvfu3WnTInd/uOif8HFE0saxI/NufLPzzKDYcsubeKxtG5fj4NEHPe2v6vePrjh//rLFG9vVP9Ku"
    "fuTszZuPHKk67ZZvXa/qadteT5u47eqsua2Nbed12nO9YI4Ka3o+uZR24Wqt1s9fXNv5IaJt26H1S/mdv1Hn4PGv469//vLlSOyDP2d83bgmKbTftbQvff2H"
    "zDkVd7bTro9coUo+V7f0LVP7ScKn3v9cvTE2xV5v/cAb3k3O9zkz8+mnJjVr9Ouwo0zB2XMqrflQ+dP4lHvUh9DQ0F8SB8d9sjaqcjD3r43qH+TKplQtl/5l"
    "x5TFy2cm/Nqo5ME3D6zvC+z5I7Jb/zzHh51uXuJojqzWsqsuBTZr3vzuv+8vJH1qz04v9ejPkx9vX1le/VJkeNKLg6NrpqcXn78vW5a+3+7fX/l4aP65cz5N"
    "PV05z8H80TUGfFhcf+I5zy+bWq0qN2LkjJw/G1v+9qVqttX/lCgxu93+lrOeDO9ULnBwoUXHuL87F8q2+t83UwZWm94pKip4nG3B+RLdCg6slyeww9CHa7e1"
    "s+29OP3YsXMf41o/v5t7VdE8I5bvG3DhRtSxgt2ebXjXbuXTft0ev96Z3iju4Bu/k80OXqp5PXnroM/X88z/559/Ikbt8vz66MGc923O7P2VYuPSFr5LOdYh"
    "f3Bo9W/bHvyZo+z8mzFjZs+OXneu2OnerWMP5j7wJVuFFcNmDP6peP0js+dfGRF5ZlHEGePzc7ah69Y9X772se/N9ANFxzwaW75LoC6w8Iay40uv2spaLDNN"
    "v/+dNXXkjMafzvQ6MnDg/C1Pd0ToujQdHNs97UzPV77j/5h+9d8L9LCC+fXLNx+fNTLf4qWP7Yn5fp/X2q9R62lZap5tlOtu8xFPQ4rbg/4tcLJdWNHogfV2"
    "nRqT4nXo727/RLaLz1ba1nvZhFc7xnWOWNN7ygdmc4OQlbUGtW6/KGy9z7oXZ+zVWfZ8lz3Th0+ePnzX9HXP/8j37EjigiVLBgzb3bfotTsbzq+uWq1O+gEj"
    "1bTpnBTvh/TBFRWWLfVkY75+7bSlbNv16ZeOFt9RKDvV4+O6vmVKfHvYPcvJWXevLSvb7ff9nQOfTS+/x7zl6bvK+l457h1rcPBk2+hjcw2GhNvp1x8vO7oz"
    "yadvSqXy5bd79y4961L/Rp26VGu883lS8vChjV96f7FcPJc37vH4hrPeF1r99s+yDVt9iLsfY9rXMt++htTyObPbX9pY9Lznjvo+PtXslI9Pj0chWwY0HpDz"
    "oE/2vNd0y/feH1q8+GWPgYbLwQcKbOy8L+ckS0EP/cSsjdd7MqEL1+9MSfl8anaJZ3Z7q+2DGmRr165d3uO6LNtuzPxxUMdWrerFHz9+PL3guRdv3lwbn7Ju"
    "7xlfNnHw8BZtfjtewPBqaK7O9XptGDUl2+Djnw43G5KraJlNe9OG9Fyep3iXWZPO3WzZnBrScFroxJxXq1b9K6heszDb0jnbU+aVy9ln5ii20Zayv9b3ql5m"
    "cO/oW+XK0GkhMffyekybEjLBP2DRX1unhzw0Re3b03259e6jR3+u/ynL4QGLtxZvdb3IjDEDqftXF7ytVKnSnSd/1pn1076P5Ze9XvI0fPXC5ovvzci9Nm7L"
    "+SJXjx9MDwqa6bO77YVsn173/q1RYODANyt27C044G3qem5s25Jl5gXkDQw8bLrQbvT6pEHhp72fX0nYlZD2onSNkOff9iW0qTGv9eHle3P/9FM184+js530"
    "+FLYGJZ0Ja313Foe4x5Frs85/vOLoeEl70/v1P1N0acXOu1MvKz/t/RyfTefS5cTb7UybJ7bq9exZvtf3W3ysdq7IavbpB+8MHX27B3H381aO/jt2eLt/YqH"
    "38l3NXfud536v2u0/dju3X+/zv3UsOld/tLjym558OjZy5fW1JSR3/J2HZwzi9/Nzx9affzY/2OTyITAsTc+rHo4esiQH2cWGNy89oBLdRP//vpDv5oNosMG"
    "baHpXqdPDvjm29T78bMHo3Z41Ljc5XnfmXce9yhUuHCPtH5dnkwauKT8Be/KPyel9X9x89Rff308lqPWRvtd73Ps63uXS5Qs27pT+G9v6j9+fuN53z/Ojx3T"
    "b82ezutOF1x1cXHu0XVKLdmR7WOnQRc7Rt7vP2fWHw33HrzRPVtTU5XhUb7r2+56SWWddm7EtpTQwdYh3ao02lRj4MeFa9cW7nb94JrBg9OfPx32b/lKlZZe"
    "7mxs97vfzF/SQueaz1w/uTC4aXr1sQ3mP2wfsv5Lvz9WH6uU9GfdPuYmZbIOXvQsYtuh8TN3rin87rUudV32uF2pHtn99374emBrgdif6nZPeFAiZ+vOET02"
    "1D5ceNHSzQurjE9ZMGHUqGwr/Ro3XnNuxuZZl2b+e+7qnGnT5k+YcCXwVP7D/XRbeuQY36FlrwG13/VveK6I13Kf2kvHB4/+oc3CTnsa7x33T4h13eia3IuJ"
    "204GtWjYrt3CMYNzBtUIuVC8zdg+B3ouLlQ8ZizDjJn+NeLI1MPHm73zpX6xWDwKja2+s9CGuQff7vNf9tfhGkW8Ns4twUZMXPrmrwJFpy/798v1LUMfjznX"
    "OaR+cvO7rSsUqzXUZ9zyC+PH59MN+LHChUuBM3+YX7J5ttQSBQsut0Qs39/AeCWypGHKl1xzzwdO/fAwMjimbIU6e6Im3NsyL6Xel/XLHlEXt6SnzJ+ddOrS"
    "1f4/13/a82CeV6u3JOQ6pfvlbOem2W2V9fomJwP/3HZ7z9VFvy8OfjDsXoXL2are3FCM21++5o6kj01nhKXlqsl57KztnzD392m3Uw8fHDBsxZNbh37fcqna"
    "lmJ7X3is27xs9s9rk74t9jD39Ytvt+Pl4x3p/U5VXzup4e4P03/Kl694im/79h+2vj1SbN6zlkc+dEtdO3Foenp62pTLvUauPTs/+cApn2tRyz/tteyY9SeT"
    "Ne/vSZG1FqfmnFSqyKuBHsMTDtD3B+9irFnebHh2Zs8qq2fC2aufa/jUKV/+xJUrg/5cMyiu+9X8v/z6wnfU9KKPS82r3Tim3px8EcPDf6tKT65S8sCEWvma"
    "dh3TKvqWJXngjiVLjXvmP1h8rEn5Zl+t855unLDiU76h9/Rb5te4OoZN/jZ2btqIjT/lOflkRdimA/sfJx8puzipJh0Qm73t+gtHVj8dF0Z5lDr3/MrjSFvc"
    "SlvapCuthvZ90yD5ut+opTWsiQ27V6zeon2jddlHHvFp22tT0crFzkanlVsXOKJFh7TV1WJKtLF/ft/fGDF0yLka49qtG9bj91dP89bxLBgwqGZsp4bXfvrF"
    "cO/z3PyLipecfcWvf+H943/86Neh29LXbb8V/+NJg/we2S8snzX5wYl96S9r72gYfO952SG1sqwf+OViWHiOoy/eztnUV/eue7PT2/MU7Riws2PX9IsFEsdP"
    "qrwx4kt0iwrvD7fy7F43//Vi1qfRp/K2ajp15tKlixK5Bz917mJP9njoVXrxtWvlw17H/TaDStudNmPSpIFvr8eVWVti/7wT0386ZC7YLXV6rR+fzUoteW3d"
    "nCr2fQn/lgy7lCsxafYPY/fsufP4ZMPpV172rDJ00IPxSyvcq/Nm6GFu49OrZ/go22v4LznHRRv6GfbPmb/584pGhfe22LxyWdWKr/6+OmdG5ctFPQad21Sl"
    "W51RE4+tzOZT69lPu1J/bu9RtlHT0CJFig59XeWHvcfYb2VWn86b/c3hC6t+tB8btjTvXymexSr4rtu7tX2/V0uaJs6p6TU+/LDPtVxtqcQxA8qZaHpD8Zh1"
    "NfSXT1QbkhLQreTW2nUr/XBp9M+JxyrF5Zk79dg1v7FvLs48HHI+vc+IRStbPd++1GPNg9TY/rHXVnV622fQvRUbL+3O/qzS2Xzjp2adT9N9Oxzu8XBI4tot"
    "z6deq//iRd9LOT8VoTp/XjQ0vGXVqlVbnk/7UPjl/eBGusr3f7s4t2uNHOcXryp/oF7M2UdvH3do+eiW1fTvkTlfLwVdzX57aOP+DVtsu39m/tZnhya0aV1v"
    "+IfUj8HHvFeu7+4353KfxV9z7/0aPSI7R+erf/pyNvrzof7pjSuWq/Vz/mIRvc72pb5uXVn55xLH80aFh4fv6D0pZGu9aoExG1aujPYJXPNo9Jcek0/nvrxw"
    "+L5z81NWZu/dpVm94KctG589njfhWw1z7aned17dzX6qp8/ZC5+adOh/4pUpbf3Enh8mzNzx+nWlXC83PD4zO7XD3aQTnT5O6nSqab4DYabOuo2Rf+5pNPiG"
    "b3KFh7lDcwRUyp74vPedPOElG16sXHb+jopX7y5gjvQzr11717Rq/i+j++6opu/wcXGlh3snjp15qtyzlLSFg/7dNyx7L3uXTgHFC066vKR4+6Lv5x97/a5W"
    "7gED7p8eWL7CspkTl2RbmGivXadOp7aX+1Sc1OjXbTVKxVwfW3lfxTefkweV3nG7ysqjFy7sOPE+T46465cuzfX6tv5Ski56ceXXN/f9ZG/mtfrse2u+/V/q"
    "vN8xt9T9chtvRydX8+7zZ4Er9VNKsGyuqW/GBZ1kwzeMPvO0f+jjbD7JFUp1+9u/y7t5rwdX9P99ar87Myrc+nPMuXPJtfs9/nhzady9RvXb/H7u/OvZ8dbt"
    "6fNMG/cMfzOuf/rDw92OT8hdwnxjZ+nxxQosHnvn36d0apbkGlnu129qqDSvXKVKnaxzmtdPvsEcOvip1+qNRxZP+21S2u3KjYvNfB93O7Dp9RwX6xrPd0y4"
    "nTSqUeHsg/ZcXxtzI+LZtV3XTp9893TtwPWruteM6BEVP7Bd1tDStQaZ9nf7bX7Q1Y0bS1ZM+3VJ3jGXDOmBva7m8Dhk+1gn+erYgI9vbPf0dfamhf/Q6Xiw"
    "dUfeAVOuvE+Y+irpzNWDd+xlkx8v3TdsZBX7kbd5tky8farwzPd93iUWGZ9Su/zs1gd63ck9Mfe6PwOu3IoqEnFkbPHndwZER3u/75u84ezC7Offz11gWTX3"
    "6s3DeSvv+Lnuo0ctbDMvepwZPvBTxaTd3Zr2LnW5ZNcvxfRvj548+dwY4bOfPrn6UP1tR88fqbLibpZSo/NE5C3z15/9Dywu2K7L1X6do89Nf9gsdfu9lMCo"
    "ufY605qc/bCq1dyKbescPTHqz51rZk2a2PhoTsOsmXMrZT/S9NKlPOdbegTZ25+dybz9xZLrca51pa+2PvDi4JSkpdn8OnacOWnWZV3o6kejpjZJyNn/Ztnm"
    "N3N/6nx63ZV7v+Y9/elps/aN2Ad3Is+UCH8WFvHkVvrfsbWuURV3+x24fWtSH7Onx7QtZQ+nbM/Wc3nupF4V1ybf2Fozzv7oxozmLVumNR+6uXh505eVW5iT"
    "fb/+MuVfw9UWo6zGYvOnRf/db/LruYEl8+4fnWXGlbhr017Ze48ZPmxY/MPybVdXaURNunpg4/3nsaOrRMT+dHts7qAsh5tM234qvcDti5OH9u6dGHQ9e1zo"
    "0/qvJ/6ReGLUvfJ/dNgReuei/4KV0cuGjSzYYPOTBcEVZmXLtifxyOatpXcFfrpfqxnz8VjBJbvXDi/Sel3yhtBGjRqd10+6Gzt8Idt6yE1b925Pfr2Zcvdr"
    "w2u9gs+fb7whz947R1M+Nnj78HyDpT9MevpgYVz1BnEluJ4fY4KDe9tWrl1bfNfN32voA3ps6XV75vtOe18GD/K62Ov61favCt260C2xz4xHp6i+DVo8T1xX"
    "+92DVYftZSICJltPXt6Y51myLWLa2FLR50M7JF/4tvnVsms/jq5f61HS9JO+Xh9WzA4f/cuze+dLbJyeJ7ZUGfvd8MjIrQezZnm9afO4Jf9e6xIWNrVKyckV"
    "S5wsWfPCb/1u1O37ZlaBS68ub/RZUjo26sGJMrsue+f2+dJjTfclmx7/M7BgZb+lP/w5/8QcjyrTJkzwW9SpWbc69g5RUQxd0PytsDH5U9YK22Y/3RU5K+/B"
    "d2M7bh/5x4YDBzoFLWO7GGc2vBizg028euvons0Xj9f/c3+1Gz+GXn//Ovr06enbL/qs6xHot/rt6fnLmDZf+5t+SR6zOHfQsWuxQT0TJt67tM1j06HRFYs9"
    "udv6zA/tR37yeTneq0/O2D0j7+hmmGKHHqCebG35Zlwuz0em1vN3jxrWsmR65U7b3wwLKXe+BjO+fHPbn81H5PCte6LVlaQ3N1oWr31iUlbjmiup60ZQf1/J"
    "HXqjqG/7HEeozpFt+w5scseevrJFWg1DgxvvH12o7+1tCQri66Wp4W3vrlENkv+c91fJpaYNrZ9Ylg17+Hdo/wtfPq076rcnsvmp3W3T8k3f0vnU8znJW/8t"
    "+/Tjemu+Pq963X6XPVtgv9o5EnvnDV3/toGpUt1F+RYbOw9p3r37l0f5S5c+VXJ/i+lpRdbH/PokrdXuub1q5F4y8OOmfRdr5H42pvDIoCZpcU+2799f0P7i"
    "ZPHYsvOXD1uR36fbmRtnDryrNrZ90cKVLoe2+Ofq6mEVhj1rtypvmTc+vr4j53hx3LnzkxtfTe90vdnFMgtOPZn69sTIXe2zvHzdec3IkRVnv36bJ2bRkLDZ"
    "6Sc/f9gSFDQp/fcFlXS6l0cnnujb93nfv38+Wzdh0YcGpUbG78t5NP32v0MPMC2C48rUCyxd51nNqEW98n4d3KTk7FWle9a4Ea3b2Ov2ymuhQcc8PJZev76h"
    "qP7ancqph8ed9nl5Pteus9unjPVfq18zs8gXmm63ybogzvzTshJldIsX17DVtdur3ll/51mjFi1apdLdplKTN997kvShxavVheKb9+yZvHhxTXN+e8XyzL1V"
    "1iebhvc7deqM6fqdfB5jI17fex9T+nHzp9tuGwybPzYs/c+j3iX8sjWlJ+6+na1o167t4veXLDDqQfzUJ5UnVvTYPip12La4KiHbt+9uUPDjluiwo2lLf/ut"
    "6qMxwbMrd/+n+aDXiS+7VCuzbMWiKcV/6pXu87hqyzLdKp0vX6FC0S51759u1LPn+jfeky5MMjS813Z4cnLVm687N+rw/tdbRV4YO01O61PmQMm9XyvpC90b"
    "Pumc8dmVxfV9X47wrRJUdtWZA0Nm1Nuz3jfKbi/olVr8+d2W89Njrj9pYm8xI5dvx4O9jvQ19570ZmfV61TriYUutY2tVX9NYP78+f9effm3ipvrni5yoXjs"
    "m32eF5LmxnfomhBdfV3NVo+PvriysWzE6s0V5qxbPYCfZ5R++TpqXlDQ6PCuv/56olul6NRmK+eMH5h89EFimbXT047F7isRG+V1udZs46nmgxdnb+A7Z+OL"
    "ind7bjXks5Qpd7Wmp+fJp1ffzn3dPH+xYlc2WWpO+SXh8Mdfh/7aa9XdtkdDt5cf/TnlcvN7z6iAt7N7p2x+snl76Wdck59HrFyebeOxe29Wtc5zvfDMmTOP"
    "Pf5l+qa1XyYkhEZTp1oui63g/dO6G7+MKhM0fXzJyIslpl/bsnZAswcrTsY/Pe6VL7JI9cQ3b3JX95k68H7X+S8f/z6/AzdTv6tE/tm7Sl4YNq7rgslHHj2d"
    "8nDkP2NDh/s9bvOpP73nS+8Zqx/Omzfg+oJDS5r7bLuYEBbhmzx63Jr7HZtPjEpd23DMT4/6ZwmwPphUyzi9zXLThSmmCXmy7H0/Z87HslmHTf0j6q+oHsdO"
    "rhpkXLB69YDLF8+ffR06bs3LG1u+FG88tPEg+9RBb7xe7dg2rUP/BpeH5oky/d5204Msy4tMn/MhreCOtV8ezRpxoUaXqmtZS1C2FUsC/um2vnmVSWeLX4nL"
    "1qf/SEO3Qx0v9CoWd7Pz0Df3X96mKiUsW1Vixz9D/Gu2PHq20rZiY1OvNH6WVih1Qomr9ql/sUPznAxKf/Hv42OVO+XpUqMblTbl+uTR3zwqVq7cr8/OnX3P"
    "l6zo49PhY4jdzliaBOgmfHvgvenWPytL5ir67HBy7ONW+QZ9+3Vbnfk/VEu9sMqPD8GpJQcMzHmPsSQ9vfLz/r5t6g8YuHhpHt2rI2WbtmjR4vqFUimev2Xb"
    "XNIUkz3rsAbzj584sWPAB6b23aIFy3+qPefz5x/Xjok8U3FryUs7Qs58qJnYr3poVNMSDx48KFDWd1C+H6at6bDj1/MTL1zrffrAgaEBR8uUKbDk0qrWJa+V"
    "e/UyffLtpKResZE/WCzDs+e73jblRfW1qX/v33/7w+t813c0CApK6JnS0cTtOHXKd66519w/ypoDAsLbtOunD+2TddW2S1c+HjySyi5cv/79x7e/dnk0ucrb"
    "WmWLFr3/sPiIka/8IpY92fizh0d5r8h0/8gt1ysk999/06fpvAE7oyYWbn/7TH/T83c5PD1vRfyc9d2S7SVNz/yfD4xa5e31/+a/2jm+/fCqZun6QkoDyBSA"
    "yQ3OqZogp4CmOIuZcsoq6MX/T+ktFouRYYVsCJCiqvjKwhg51shJ0rAenvAijrboGYOYw2rQ2W2erT1jPH397PHREVaQQesXH28Ni7KFCwmzIbE9hTRWOZs3"
    "OsYaHx0b45Ro0cBmDbfZeXA6/4C6Uj5vZByPz0LpTQYpGcgivKMp3p+BMQjveTgwTzcg1rM5f7VJb2AksVG4lNUbTYKeWURYlYfoHyJmAoumQiI8paQimr+U"
    "1nMsVAzhb1MXKWUS8yAoXX3P9jovnbeugy7kN896IQhXZqUr4SKapvWMfJHZrDe44CG4RITMZIRsZvVCdiNn5tUBYoA00Fvnw+q8YqUXu7eO03n1ED4YdV5d"
    "pC+78y+MRedl9eYN6bzivXX8zQgvzreYWWgMY9ZzqshcypDciUqVsazwjYE18m3UxU0Lbx1L6bxs0s1F8/fKv8QoPkXyLwb+NUH6Fgjht3GZKQgaWUdqGEMp"
    "g8FdXoxILzo3ueC7Ia3uQVmA4aAAdaB9gaYImlmU9GKXWptN8QkYAFeGCXomWSNRrULdVEliUiRB2TUX/2UpPgA53PDhiDO4dmn8NQgoLGPUU2bVnhMsFCjn"
    "WgTxoJjBty6IM4/DQOuN6rGFdpsLo95sJgkS/8GFRW9R9dAbtK5Y0Fpj3Fx8JoueY/63xWfmhLGNMMb+hwZp0LOMSvfgKQDHmHQhYeJ44++uu2LNqDCTMVxm"
    "3gtHq3pR3lMzt7VCVmzo/+ubAv0pgxsfPnKbaHBXFOW2BmhixAb4v74tkwHpxn19iae5rHt90IyeYly8GPUM8UDjoNL+jeuK1xr1JoHuGA2CS1ov/jEi9Y6o"
    "KgidZYxSahZ+ZiaLKUbPBztnObgPhG1hXLU4pGaT3swoLmAsfIma0O5ZnlyzsljwpdTneAcmFq2Pk0s3D6pE/B7WiNO9SWJxMuEkhdDVpdAxwnQ9vmbgPIDj"
    "50H8iAnnAZRY+w5GQHARbJzKaYk0yaBM8rREbDD17HECYzeLA78PYwR83uR4BeJQysRI71RnL98NiGP4ylcB1FhwK3Z8o+zOx8DTlXpRAKAw/zDzNK+bzS1A"
    "+OHPkhGGKmOTeCH0AHihawURXKRaQTQn9g1XIPWVtRMDXuFn10KJBrUWCb5X1F5Gjvv9MI2snjJgyyvT1g1C/1YY50cdimLBqNNMahfucGVi9Yb/3Y3wg42Z"
    "RbZuq7J2YCVa3eKbEYKkgbBJg2EIOgHDkGubJrlKDQpLqRcyaNShrIlziTqgRFybsU76GpZbF/CKDEqZBsxxIvdxBcwAh4zsGIYnCnxjdhMAI89bVUoMSYS+"
    "34OJzuiB0jOcEfQxhmKMbnJlYQSiQtAQlSQFDMNCYJaGSnE4VTAQSSgRECepPEYLYmm9z1koTetFVak7OUnlAVxNVWZroq5ox0kqj+6qiMW7gWUnChmLeBMK"
    "MaenWQcjIZBJBQGFMhVTFgRwK6JReJXl0LZCV6RpUCbTNGVBIU3LclXYEoWDQpnCKUsSXVryBarGtaUSyxKGE/4/I8MJC9sZFp/ltiyVilNr5vsNK5Y4q2fN"
    "6OZMYL+5CAW9xu0fGx8f2wOzzB0YGxvvWOYmvy+e8eiV8wmeeQgFbOQtmFxiDiOtqPG3yP8bJP4rvecjLx+PWAv83Mhb+Nm1uOZGS2t3JrCYRIvrcj7CohuQ"
    "JkgvPGWhJXsWIzQeAxRjlAZipBdwKY+LAhbjJTB82JauiJRehNHEQInDCEeJS4B8WxdXAB1WWvDXcLKHaKARo/wY6UBMOXCAb3kcNMDBv+Vg6QjG/fiycQbs"
    "VFjCF6C0GkpXtXasdH5nQ0JVqIWfwatUKEAFyj+UZU1OS4HS7bkHAMPQQsjPAEDnPgesMIipNVnYNJD1KzYs0DIiFS3SUb1uQsmTZT4cqaFUtJZEBVSnPuDU"
    "XIWFRo6TgceBjwC5TW73ghC82BQGWkmV79IloR1lFwFdExiXOlk/6YoMReXkXok4zoFYLHqlC8loF9CxrOAVwnJfV5Bim4EWZi2OGnDFBsscBi5E1zZIjyGE"
    "LtNQ7uEq0RBeBAJeFLAK786lDroBMODqLuot1lGbhu9r5zCQiZ8TlG0Mfq1TRF8AJAzgilJc1BVYtGmGeunaEFAOsUrvdkfR0u6LCIzwpNClnjPdjgyu7cjM"
    "X2TW8dMD1h3mja5Pjmi9mef+enOG4Vdc+3B6C5dDhJmkQZjwx4M3UBBus8O3LYEoJh58YbPHgXdhQBRlj44AX8W7eEiEHuzhbrtPmhKeMrjcaBdbnAsEYUCH"
    "76Mj493mnRWr1cW9+26ONekzWA+3wpuLc6ko+DlargunKnQbJpNFaLf/s1s208I6h4v5duCGElzr1RoTLt4jy991TIzNHmXt3kX8ghFuOs6lAQYCM9YoOz/4"
    "GIRVr4RI6SrWcVVvcFWs3VF8PVws+du6W8NtMW67bZYzCnMAl9uOjoH+bNExzhjcGCMYszipcTjlXSU6SjcCFI7dFi3erg/N0zCi+uhii3BfJ+e7mQtSxSoe"
    "79ZtvmijMI/NUCzua+G0WW/KaF8YgCNAsfKFnaExxsByDZQv6gHehUfa7MLKVxi4IsoWI3EMRbMOlusooYv7WpBJb1FpQtL4wDqFfBiSuoHX7gkxkSogE/g+"
    "nCjHbBMXk2F4kUceqNIoAbZRKILjjy0hDLyLcsT+bnKPctsoIDxGpVnVKuVc4kawjEgYkcK7R8dEOiqLcXvfpqUnfTIih/to2Hjgq627Eq57YBhMejojDFu8"
    "S61mXITPvEujRU8LizwKp4I5RvgYBiqGzrxLNoNLE6s3ZnDpGjPEFZrgntYYcXWmcd2gAGGTqkbWmEhhgBGWYwPqeX//6guPRm1NyWwWmiQ5FNoNUBidsI5p"
    "Ul3jMhn5huCAwwcxWnpswUjzAeGTwOIt4rzQB0wsGTAhZaSJofQYzyS9aygMRrRsql3mKlMNM82JKZyumDOumP83JzyPYjM4gXcFbj5Gujnw4lIy0dK3YYpC"
    "VCoKzzN04C0trM64Cz7DUsKCjAI+PyQzRgvsY7y7zP65DybLUMJas2sp43oD45bewKp2BEpYSSOHwroBSsbZJ0Rj5AS2rYFGcCKkNvPzVAtfjjx34SckPEWi"
    "nZ/mgScHjEV4Nsd3HuE5PN8+TCbHRS53xWncFYE15DKl9MSeYxST6gBpQcPqvE7kWDJ2bmrkrl2HAI7VG0wZXSvKE5aogROaAPDCtwcjtrgMGsVFYA1iVhkm"
    "OQMtJOg6EAsrLPzIIcQRA0gD4PiC6gnmS9KLHXxrlT6CZ6wJSqGYZ8vwb3rHZhA4F3omb8AkPtfk+IJ3njL5e/OWZJzwITDAGS69AJwx0ouwYmU0uwuVNFVx"
    "RYXr5EY31C/jUr8m8UqO5ZwWsPgAzbNMmQRRVObu2cWVidGbTa6uVAfMzDowSG3V1QHjHutSQ3I1rnOPcYs4TCqMC2mPMhVljFRdt3gy8/NW1TowOdKehFX0"
    "lqBPRIGOzvdKmo+O/UCXiZaujncLJvhQWAUV7SYHDD+Z1q47GHb5JqTnB0rgQywu7ahr0uiVeGPI5weMXmWgaOiSqZN582Z+EpLBfGN3mefrSXgU72JeamX4"
    "9XmXB1JOS/2ZQpdhacXC6Y1kYzDDNxwjT785IT8MjIW0QZGA4tIazBqtgcAa8imDUW/hFS1KHvZfn8E4F+d/AMcPIpw5Izr1EjXTegd5MhiFlDHt/mXRKlG8"
    "NRRoIUcoYxOtK5VolEsR9nQpq0y7ZWnxSZFKzzC4zwWn5kLnPvtG4RcVrvabSCXn9IzNmFmPGVbOGJOQNEXUY4UfYDgNESwBbaa1FlhI7KEKijMJSX0cp1yu"
    "bykVFOyG0kQD/Xhc8SgUymG7jBCVXUbh/wKZ7xVCCpYzZGEJnWMdi2ImRic9F3STT46ipPCh9hjSbT5YNR/Y5S7aPS3DqEy15EQKZnSNABSY+AQCygWmI1HK"
    "aZHL3CojR8o0LvHnHa64VDqZMmkSpL8bKWGC6Jq/qCkE02hJCKbRTmKeQ9AoVcADJSE/WJldMi4Znn8wKGVIHCSpSByc8zFpWuy16rpwlASezeL6iLPcjMQM"
    "458kleKfc6qn+MsxhC6sU0kKZlIZM0ER2tpSUEfiuC+t4sijvlwLCCEsZ4QYFiVKbEZbBqWBkMIbQoiV3E3qBwwj/EujBwGGgLghTCFJGyWQNtdJnVMwDcoY"
    "5DLryoifncoUjDUIvxZzcsFolozW2iLOFgovJ/72LkPZMAb4rNZPpWgy6ctAiWMaWdkYzOJUFPhg+falUTJa65PallDP46TMPQGrWfkcRHwA1lClUDLlxiQ9"
    "unBxQxrrOU7wpRKTaZRMDhSqUjlOIKRgYsI5eoSzmG8VNFoZtBmHsnKQgcUna5s4i0V5gUXPOAzQ4jDqGvDVfcMYpiqUQ5iqFBS0VKkgwCmKWU0CC1lNBotY"
    "VQZKUFWmaF6wfGjF3avpgXtXE8E7V5NJnRAQA5oSFxEkYkCbtWbdtAH/pEDLHPzNq0GsXmfiaBGWCvgZlYEhXK+X3IC1KQxqggVdLXNI1DQ/8giNlgg3pEOS"
    "J5EOYWCTrHihrSFRM5RR7OhEqGFUkPyAcVIbNsnSjIY5NG5G2KWBEDaIRw4/+FZiIRh/teyhgRtYgfGTAQeBADgSxxlN2AxFMDiiraFBg7GLDDXg3JIfiXNj"
    "YNMEy00a5pC4WZaROjQJbsj3Qe9nCYIJwxDMVbXsoWMgZSKNJqp0geF9oyZ/ogw191NVhI8ABBl65qeqKq8YC0LUvE9VU2blghBJQdRVIWuFquoEBOqi+Iek"
    "rk4/1P3C5TZBiJptqmrKCwiCED3XVNUFVS61M2lsN2eYTaKksPJQclhFSDmoB6RcEXVkMmN2nZOi9EGhocSwYFByZ3LDCmkSJMyGYbHMBm1Lm9awFjGSEtEa"
    "wQcRp2E4LKdB28IRGjLEoJsLbgjYDGPAsxmUKRyVIcMLgovghITHMEY8j0HawpAYQsBSSINO8G3ChGcwaGMY+kIIWer6ohcC7mLGcxeUKRxxIcQrRW/BCRFr"
    "IXhIhrSFoyxkiMGoIXZuEr7CEjx1QRvDkRUSzKpkhRa2F0CxFUmIoivqqo55HC9EExZ1ZafZFMUgKYu6rtOUhmLQpAWh7JhXUIwGbZG1UbwFGFAnLgjfDpZN"
    "MUjqoq7rRHQpRoO8qGubwAZJDIq6qItAHaoLQSUhhFIlIISOQIPkKuqaUhGpy0AJqAsVFIUvPc55wUTjESpL41mKljkMUTEZ9Qzh+gtw4+AXWqgZPFfRMoel"
    "K0S4IceQPEGaoQWbJWAsaGtY0kKEGhINyY/MNbRgcwS8RcMcjrqQwQZsw+EH30oMBOxFyx6OwJABB8EAOALEQwu2kYDDoK1haQwZakA9JD8y+9CCbSJgMhrm"
    "sGSGCDfkH6D3syTBxEzAZ7TsYSkNAXI1SsNaaD1jQFAaIERQGoQqnMaJQiSlQSjLUypRiqA0CF15diNKUZQGpQwnGrKyKqVxaCMoDTSgSmlQviEFF6UISoPQ"
    "lbmwKEVSGoQ2aANSw4MDvsmlmlFSWJMoOawtpBxUCFKuCENOJMekLFSUPig3lBiWDEquYDt8LCYlOxY82dGwhuE6BpNQJGRcR/RCRHU4Ck91NKxhmQ4Ratj1"
    "RUcERIejCYgO0hiW5xBhhiFHdENCcziCRBINaziWQwYahDrZDb59ECR5aJnDcRwy2CAcSH7wFIcjyL9AG8MyHDLMILKLbkgIDmcgIDhoa1h+Q4QajilSlyeh"
    "N5yRgN5omMOyGwLcquzGRAtDFILdSEIUu1FXhRM/UYhmN+rK8vRLlKLYjbquPAcSpUh2g1CGExFZWZ3dyNoodgMMqLMbhG9Iy0Upit2o68rMWJSi2Y26NmgD"
    "YrtDLdkghLAeEWJYVSgxqAyU2Dn4oBdvENqgwBBSWCIIsYLUsNLvcUkeNHEmPKvRMoehNfwtGU2EtEZyQ/S4iTPjeY2WOSyxIcINe73kieChE2chYDZoa1hq"
    "Q4QaxhvJD8mjJwNFwG00zOHIDRlsEOkcfrCtxEATsBstezh6QwYcBAXgCP8YysAQ8Bu0NSzBIUMNgrvkh+RhlIElYDga5rAUhwg3HFhA7yd5JGXgCDiOlj0s"
    "ySFArkpyDOLPchAkRxKiSI66KpzyiUI0yVFXlqdeohRFctR15RmQKEWSHIQynIjIyuokR9ZGkRxgQJ3kIHxDas5IZ4mokxx1XZkfi1I0yVHXBm1AangoloOS"
    "wppEyWFtIeWgQpByRRhCUx2UPig3lBiWDEquYDvShiNEbMeATxjWNIdhO7QwnJGRHckLEdkx4POFNc1hyQ4JbNj5JUcEXMdAkC6sYQ3LdUhAw6gjuSGiOgTZ"
    "wlrmcFSHCDUIdw43+CZCkCysaQ/HdIhwg5gA/OCJjpEgV1jDGpboEIEG8V1yQ8JzjASpwlrmsDyHBDYcWkC/J6E5RoJMYU17WJqDB67KcvggT5tQLEcSoliO"
    "uqpjAsgL0SxHXdlpGkabkCxHXddpLkSb0CwHoeyYkNAmDZYja6NYDjCgznIQvh30nDYhWY66rhND5qVolqOuDdqA1O5QLAclhTWJksPaQspBhSDlihiEZjko"
    "fVBuKDEsGZTcmeUwFulXMiRPqoz45GFNcxiWQ1EaJ5IoaQ5wQ/SsyohPIdY0h6U5RLhB9weeCJ5WGQkSiTWsYXkOEWoQeIAfkudVRoJ0Yi1zOKJDBlsKeU5+"
    "8K2EIKlY0x6O6ZABlwIDdIR/ZmUkSC3WsIalOmSopSgP/JA8tTISJBhrmcNyHSLcYISBvZ/kuZWJIM1Y0x6W7BAgV2U7vEUORXZEGYrrqCrC6Z8gQzMdVVV5"
    "DiYIUTxHVVOeBwlCJMtRV4VzEaiqznGgLoriSOrqDEfdL+TlghDFb1Q1ZWYsCNHsRlXXBE4UM8pHl2XMwkFKQeUh5aCK0HKpHtByRdRBZuEg9aVCQ4pBwSDl"
    "CnJjMonp2yTkxoTPOdY0p01uhHZiJOQ2khcibmPCpxxrmsNxGyLYkIxIjgiojYkg41jDGo7aEIGGVERyQ8JsTAQJx1rmMMyGDDUgIg43+CZCkG+saQ9DbMhw"
    "g5gA/OB5jYkg3VjDGo7XkIEGPERyQ0JrTATZxlrmcLSGCDZkIaDfE7EagmRjTXs4VkMAXI3UMBaDMBSpsxogRNAahCqc74lCJLFBKMuzLlGKoDYIXXnqI0pR"
    "5AalDOcfsrIqvXFoI/gNNKBKcFC+IRsXpQiKg9CVCbEoRZIchDZoA1K7QyzhIKWwJlFyWFtIOagQpFwRg5BLOEh9UG4oMSwZlFzBcvjKpQiTjU34ZGNNcxiW"
    "Y2SEgw/IaI7khojmmPHpxprmsDSHCDfs/pInAp5jJkg41rCG5TlEqGHgkfyQEB0zQcqxljkc0SGDDUKeww++lRAkHWvawzEdMuAgMABHeKpjJkg71rCGpTpk"
    "qEGUl/yQcB0zQeKxljks1yHCDUcY0PtJyI6ZIPVY0x6W7BAgV2U7JuHcKgTZEWUorqOq6Jj+0RpMR1XVaQ5GI3mOqqbTPIhGsxx1VcdchNbgOFAXRXFM4PAv"
    "NYaj7tfBy2kkv1HVdGLGtAa7UdUFVS61MxS5QUlh5aHksIqQclAPSLki6qDJDUofFBpKDAsGJVeQG47Ss2RJOGZ8yrGWNQy14Uj2+wR9TvRClINjxicca1nD"
    "EhvuO86okBwRpOCYCdKN0cawtIYIMww0ohuSDBwLQbKxhjUcqeG+Y6tchxts+7AQpBprmcNRGu47drEFfvD5NxaCRGO0MSyhIcMMorrohiT9xkKQZqxhDUtn"
    "iFDDEUXq8iTZNxaCJGMtc1gyQ4Bblczw1cyifkgFhCg6o64K53miEE1o1JXl2ZYoRVEadV15yiNKkaQGoQznHbKyOq2RtVG8BhhQJzYI35CFi1IUtVHXlYmw"
    "KEWTG3Vt0AbEdociNwghrEeEGFYVSgwqAyV2Dj5oXoPQBgWGkMISQYgVpIZx2TxY67wafGqxpjkMreH7BUP4QyrghmjFxoLPLdY0hyU2RLhhr2cUuwdrwSZI"
    "LtawhqU2RKhhvGGUmwdrwSbILtYyhyM3ZLBBpGNc9g7Wwk2QXqxpD0dvyICDoMAodg/WPOGHIMFYwxyW4ZDBBtGdUe4erImbIMVYyx6W5BABh0ML47J7sCZy"
    "giRjTYNYnkMAXZXnsBa1HfdBYGadz3TNwHPUVUF0lIRonqOuDGMU6zjFToXnqOvCQMFatE5GQCmD3spaNE9GcGijeA4woM5zEL5By2WdjgLLyHPUdWHbkaRo"
    "nqOuDdoAo9ii1zURByWFNYmSw9pCykGFIOWKQIROxEHpg3JDiWHJoOQKwkMxZGyHpghyjJHGMFyHIt+gWPBBRHRoiiC9GGkMS3Oo79ihmKEYEo5DUySZxQhb"
    "WIZDfccWxQzFENEbmiJJKkYZw5Eb6jv2KGYohozZ0BRJPjHSGo7XUN+xSbHohYTUkKQSI2xhGQ31HbsUMxRDSGdIsohRxrBchvqObYrFHk5EZEhOh0Rbw7IY"
    "KnP7FAtBHLlPMRCiWAyjsU8xEKJZDKO1TzGQolgMo7VPMZAiWQyjtU+xk7I6i2Ew+xRDA+oshtHapxhIUSyG0dqnGEjRLIbR2KeYoZA/B1cXwTqkNH4IjhCC"
    "SqA0fgLOUBq//1bXBEVEafzyW13oTFVoMyvuNE3EVmh80rCmPW3CQvMt30L4iyjghoyz0Pi0YU17ONpCBhz0cuCJhLnQBInDGuZw5IUMNggwwA8Rf6EJUoe1"
    "7GEoDCFuKbY5+SFoKATJw5oGMUSGELkUEaAjAi5DE6QPa5jD0RlC2FJAB36IGA1NkECsZQ9HasiAg9EEhgAyXkOQQqxpEEdtSKCrUhtKY79iIERRG0pjv2Ig"
    "RFMbSmu/YiBFURtKa79iIEVSG0prv2InZXVqQ2H2K4YG1KkNpbVfMZCiqA2ltV8xkKKpDaWxXzFoeAh2g5SCmkTKQW2h5VKFoOWKQIQkO0h9qdyQYlAySLmC"
    "9QiniBNudkPT+CRiTXsY1mPkTRCSHskL2emYDD6JWNMelvSQ4IYkRXJEcjomQ5BErGEOy3lIUEOKIrkhOh2TIUgi1rKHozxEsAFBcbghaCUEScSaBnGMhwg4"
    "CAzADz7jhmYIkog1zGEJDxFqwE8kNyQ5NzRDkESsZQ/Ld0hwQ3YCOj9LFEwIkog1DWLpDh65GtuhzYyeQ/1iCggRbAehCqeDohDJdhDK8pxMlCLYDkJXnheJ"
    "UhTbQSnDuYmsrMp2HNoItgMNqLIdlG/I00Upgu0gdGWmLEqRbAehDdqA1O5QbAclhTWJksPaQspBhSDliiiEZjsofVBuKDEsGZRcwXb4LmAkPQucwacVa9rD"
    "sB3+X46U7khuCOkOPrNY0x6W7hABhwFA8kTEdwiSizXMYfkOEWwYeyQ/RISHJcgv1rKHIzxkuEHYc/jBNxSWIMVY0yCO8ZAhB9EBOCKgPCxBlrGGOSzlIYMN"
    "Yr3kh4jzsASJxlr2sJyHCDgcaEAIICI9LEGusaZBLOkhgK7Keozy5vUZSY/R6eycDJxHVRFOBgUZmvGoqsoTMvkYGRW+o6opz4kcZ7mosB11VTgtcTpQRYXr"
    "GJ1PD1ChOpK6OtNR9wsJuuN4kYw8R1VTZsjiCR9IlqOqC6pcamcZx3koduzwq6KrKoRjufOpBBkaDMotkKoahqOtKEQ3KJRpWa5qXB4SHdv7qzQ5lG1Zrmpb"
    "Hrac9uBXaZRI41CubhwOLc4b5as0W9m6Cl+Elyh2i1Zp2sCHettGwjcja1QO007bzmds+ijDUKxqWo6j0sbwyL6BMi7LVa1rS4X4KsRZT+Gvlyelt1gsRoal"
    "jfxFFPiTv+L44jSbePpgFiVhfNAGF3G0RVzntRj0UuwXxhI5bAvjiV98vDUsyhaua6/zDYntqevg6RvSt6dN59vMGhkdY42PjhVGn4Qu8eKXDWzWcJsdDjW+"
    "9YONusg4Hp9Fws//z1rEcqB4f3wrEt7zcHyDelgjbQZdQCx/O71A3xfERuFSVig+Xs/sOjqJppz5LkXrORYqimNTpADCIoBQG6fUXJmVroSLaL4qGPkis1lv"
    "cMFDcIkImckI2cwTHh4yJ68hREKkgd46H1bnFSu92L11nM6rh/DBqPPqIn3ZnX9hLDovqzdvSOcV780TOfHF+RYzC41hzMJOaSrIXMqQ3IlKlbGs8I2BNQqt"
    "XOmmhbeOpXReNunmovl75V9iFJ8i+RcD/5ogfQuE8Nu4zBQEjawjNYwCuXKXFyPSi85NLvhuSKt7UBZgOChAHWhfoCmCZhYlvdil1mZTfAIGwJVhgp5J1khU"
    "q1A3VRJjMhOVXXPxX2HvRoPDDR+OxANAFV0afw0CCssY9ZRZtecECwXKuRZBPChm8K0L4szjMNDitpwqOGi3uTCKew3hg8R/cCFMqtU89AatKxa01hg3F59J"
    "nOL8T4vPzImTR7IY+x8apEH8WWGG7sFTAE74KUWYON74u+uuRNJEGC4z74WjVb0o76mZ21ohKzb0//VNgf6UwY0PH7lNNLgrinJbA+RZq9AA/9e3ZTIg3biv"
    "L1HCJMWtPvhJlUDXFV6MeoZ4oFGugAjXGsV5jcFoEGdDevFPnhAY9QoZo5RKk14ophhxwuYkB/eBsC2MqxaH1GwSNr5xvoCxWKRfkyHcszy5ZmWx4EupD+Yr"
    "SH2cXLp5UCXg/HKpRpzuTRJT0pHashRCV5dCxwjT0iKbNA/g+HmQMIsG8wBKsSRGdBFsnMppiTTJoEzytERsMPXscQJjN4sDvw9jBHze5HgF4lDKxEjvVGcv"
    "3w2IY8QVkwyAGgtuxY5vlN35GHi6Ui8KABTmH/yU16ubzS1A+OHPkhGGKmOTeCH0AHihawURXKRaQTQn9g1XIPWVtRMDXuFn10KJBrUWCb5X1F5Gjvv9MI2s"
    "sMEDrrwybd0g9G+FcX7UoSgWjDrNpHbhDlcmVvz53//oRvjBxswiW7dVWTuwEq1u8S3ktrEGwiYNhiHoBAxDrm2a5Co1KCylXsigUYeyJs4l6oAScW3GOulr"
    "WG5dwCsyKGUaMMeJ3McVMAMcMrJjGJ4o8I3ZTQCMZvHIJlcASCL0/R5MdEYPlJ7hjKCPMRRjdJMrCyMQFYKGqCQpYBgWAjPtus4LGIgklAiI8/I2HKMFcYa1"
    "bzCtF1Wl7uS8SAsHcDVVma2JuqKdjMukCMTi3cCyAxsVizehEHN6mnUwEgKZVBBQKFMxZUEAt5S0J6+TV1kObSt0RZoGZTJNUxYU0rQsV4UtUTgolCmcsiTR"
    "pSVfoGpcWyqxLGE44f8zMhxf/RkXn+W2LJWKU2vm+w0rljirceQGkX3XR5Oua9z+sfHxsT0wy9yBsbHxjmVu8vviGY9eOZ/gmYdQwEbegskl5jDSihp/i/y/"
    "QeK/0ns+8vLxiLXAz428+cqU1txoae3OBBaTaHFdzkdYdAPSBOmFpyy0ZM9ihMZjgGKM0kCM9AIu5XFRwGK8BIYP29IVkdKLMJoYKHEY4ShxCZBv6+IKoMNK"
    "C/4aTvYQDTRilB8jHYgpBw7wLY+DBjj4txwsHcG4H182zoCdCkv4ApRWQ+mq1o6Vzu9sSKgKtRiFh0IZKhSgAuUfyrImp6VA6fbcA4BhaCHkZwCgc58DVhjE"
    "1JosbBrI+hUbFmgZkYoW6aheN6EUnkizqigVrSVRAdWpDzg1V2GhkeNk4HHgI0Buk9u9IAQvNoWBVlLlu3RJaEfZRUDXBMalTtZPuiJDUTm5VyKOcyAWi17p"
    "QjLaBXQsK3iFsNzXFaTYZqCV29K5YoNlDgMXomsbpMcQQpdpKPdwlWgILwIBLwpYhXfnUgfdABhwdRf1FuuoTcP3tXMYyMTPCco2Br/WKaIvABIGcEUpLuoK"
    "LNo0Q710bQgoh1ild7ujaGn3RQRGeFJIa2w/+D3mDa7tyMxfZNbx0wPWHeaNrk+OaL2Z5/56c4bhV1z7cHoLl0OEmaRBmPDHgzdQEG6zw7ctgSgmHnxhs8eB"
    "d2FAFGWPjgBfxbt4SIQe7OFuu0+aEp4yuNxoF1ucCwRhQIfvoyPj3eadFavVxb37bo416TNYD7fCm4tzqSj4OVquC6cqdBsmk0Vot/+zWzbTwjqHi/l24IYS"
    "XOvVGhMu3iPL33VMjM0eZe3eRfyCEW46zqUBBgIz1ig7P/gYhFWvhEjpKtZxVW9wVazdUXw9XCz527pbw20xbrttljMKcwCX246Ogf5s0THOGNwYIxizOKlx"
    "OOVdJTpKNwIUjt0WLd6uD83TMKL66GKLcF8nN4gnPTkjVazi8W7d5os2ir9Sci0W97Vw4YfHGe0LA3AEKFa+sDM0xhhYroHyRT3Au/BIm11Y+QoDV0TZYiSO"
    "oWjWwXIdJXRxXwsyiUeHuTYhaXxgnUI+DEndwGv3hJhIFZAJfB9OlGO2iYvJMLzIIw9UaZQA2ygUwfHHlhAG3kU5Yn83uUe5bRQQHqPSrGqVci5xI1hGJIxI"
    "4d2jYyIdlcW4vW/T0pM+GZHDfTRsPPDV1l0J1z0wDOA0BCUMW7xLrWZchM+8S6NFT0s/EHdyKphjhI9hoGLozLtkM7gE2eYZ7lNRjOIKjTKjXesnG9+BRm1N"
    "yWwWmiQ5FNoNUBidsI5pUl3jMhn5huCAwwcxWnpswUjzAeGTwOIt4rzQB0wsGTAhZaSJofQYzyS9aygMRrRsql3mKlMNs5DbqoI544r5f3PC8yg2gxN4V+Dm"
    "Y6SbAy8uJRMtfRumKESlovA8Qwfe0sLqjLvgMywlLMgo4PNDMmO0wD7Gu8vsn/tgsgwlrDW7ljKuNzBu6Q2sakeghJU0ciisG6BknH1CNEZOYNsaaAQnTr9K"
    "4LmLcGawUVgrczzNA08OGOfdZllKa78WDvubRk1ryGVK6Yk9xygm1QHSgobVeZ3IsWTs3NTIXbsOARyrN5gyutb4NSXwAn79qF1cBuxvKTWtQcwqwyRnoIUE"
    "XQdiYYWFHzmEOGIAaQAcX1A9wXxJerGDb63SR/CMNUEpFPNsGf5N79gMAudCz+QNmMTnmhxf8M5TJn9v3pKMEz4EBjjDpReAM0Z6EVasjGZ3oZKmKq6ocJ3c"
    "6Ib6ZVzq1yReybHO52fwAZpnmTIJoqjM3bOLK/AjEaUr1QEzsw4MUlt1dcC4x7rUkFhOc6Uvs8Yt4jCpMC6kPcpUlDFSdd3iSfzplVodmBxpT8IqekvQJ6JA"
    "R+d7Jc1Hx36gy0RLV8e7BRN8KKyCinaTA0b4dSX7HcdMAR9icWlHXYLDGDSMIZ8fMHqVgaKhS6ZO5s2bxc1SXcw3dpd5vp6ER/Eu5qVWhl+fd3kg5bTUnyl0"
    "GZZWLJzeSDYGg58WcpyQHwbGQtqgeeoY/vf1WtaQTxmMeguvaFHysP/6DMa5OP8DOH4QETZscUWn9VN0UGvSbxK1+xfB9sla1lCghRyhjE20rlSiUS5F2NOl"
    "rDLtlqXFJ0UqPcPgPhecmgud++wbxWOxXOw3kUrO6RmbMbMeM6ycMSYhaYqox4IfpsIhgiWgzSSbXWvaQxUUZxKPfueUy/UtpYKC3VCaaKAfjysehUI5bJcR"
    "orLLKPxfIPO9QkjBcoYsLKFzrGNRzMTopOeCbvLJUZQUPtQeQ7rNB6vmA7vcRbunZRiVqZacSMGMrhGAAhOfQEC5wHQkSjktcplbZeRImcYl/rzDFRfh3hac"
    "kRImiOo7GSCEYBotCRE7GSBUAQ+UhMi9ChDKkDhIUsReBAhdOEoCz2ZxfcTlp/ooVRD/JCni1/gIXVinkhT5c3uEtrYU1JE47kurOAbXHSNQQljOCDEsSpTY"
    "jLYMSgMhhTeEECu5G+N0UBCtsaMTAXFDmEKSNkogba6TOqdgGpQxyGXWlRE/O3XZfsnJBaNZMgR7TSJtofBy4m/vMpQNY4DPav1UiiaTvgyUOKax33GILPQh"
    "7siELhmC7SwRllDP46TMPQGrWfkcRHwA1lClUDLlRto8ytUNaaznOMSJYDRKJgcKTv1ILxAnOI0Dv4AUtb0LShm0GYey6vYtDm3E3izQgOreLAjfMIapCuUQ"
    "xqFP7AKVCgKcopjVJLCQ1WSwiFVloARVZYrmJW9vo7h7NT1w72oieOdqMuftDzna5bxQrU2/8MePaprT3v2Q42dUBsLdD4Ebst0PCRZ0tczhNj8kww3pEK04"
    "MFQLNsmKF9oabu9DMtQwKtDK80K1YJMszWiYw2x9SAgbxCPa5bhQLdwEx49q2sPsfEgIHAQCWnFeqAZshuD0UQ1ruI0PCVEDzk0rTwvVgk1w+KiWOdy+h2S4"
    "Id+nXc4K1QJOcPaopj3ctockyFXpAsP7Rk3+RBlq7qeqCB8BCDL0zE9VVV4xFoSoeZ+qpszKBSGSgqirQtYKVdUJCNRF8Q9JXZ1+qPuFy22CEDXbVNWUFxAE"
    "IXquqaoLqpxWnLbpOptESWHloeSwipByUA9IuSLqIPd2RuqDQkOJYcGg5M7khrUwZMyGwZ80iralTWtYC/lJo4IPIk7D4A8aRdvCERoyxKCbsxaGhM0wBOeM"
    "okzhqAwZXhBcWAtDxGMYgmNGkbYwJIYQsBTSWAtDxmAYglNG0cYw9IUQstT1RS8E3IXgkFGUKRxxIcQrRW/WwpCxFoKHZEhbOMpChhiMGmLnJuErLMFTF7Qx"
    "HFkhwaxKVmiNI0aBEEVXaI0jRoEQTVhorSNGgRRFWWitI0aBFElaaK0jRp2U1WkLjTliFBpQJy601hGjQIqiLrTWEaNAiiYvtMYRo6wFecSougjUoboQVBJC"
    "KFUCQugINEiuoq4pFZG6DJSAulBBUfjS40xkR4yy+BNGNc1hiIrJqGcI11+AG6ITRln8AaOa5rB0hQg35BiSJ4IDRlmC80U1rGFJCxFqSDQkPyTni7IEx4tq"
    "mcNRFzLYgG04/OBbCcHpopr2cASGDDgIBsAR/nRRluBwUQ1rWBpDhhpQD8kPyeGiLMHZolrmsGSGCDfkH6D3k5wtyhIcLappD0tpCJCrURrWonG0KBAiKA1C"
    "FU7jtI8WRSjLUyqto0URuvLsRvNoUZQynGhoHy3q0EZQGmhAldKgfEMKrnW0KEJX5sLaR4sitEEbkBoeHPBNLtWMksKaRMlhbSHloEKQckUYciI5JmWhovRB"
    "uaHEsGRQcgXbMTgdwILp6fiDRbWsYbiOwSQUCRnXEb0QUR0Of6yoljUs0yFCDbu+QT52BgOa4FBRtDEszyHCDEOOwXHgDQY0QSKJhjUcyyEDDUKdwemoHQxq"
    "giQPLXM4jkMGG4QDyQ+e4nAE+RdoY1iGQ4YZRHaD43QhDGiCw0Q1rGH5DRFqOKZIXZ6E3nAER4lqmcOyGwLcquzGRAtDFILdSEIUu1FXhRM/UYhmN+rK8vRL"
    "lKLYjbquPAcSpUh2g1CGExFZWZ3dyNoodgMMqLMbhG9Iy0Upit2o68rMWJSi2Y26NmgDBvkELpUlG4QQ1iNCDKsKJQaVgRI7Bx/04g1CGxQYQgpLBCFWkBrW"
    "+ZQw7fVYDn+AqKY5DK3hb8loIqQ1khuix00c/vxQTXNYYkOEG/Z61nG6GQY2wfGhGtaw1IYINYw3rNO5adqwDQSnh2qZw5EbMtgg0rHOJ7JhcBMcHqppD0dv"
    "yICDoAAc4R9DGQjODtWwhiU4ZKhBcGedDpHDwCY4OlTLHJbiEOGGAwvo/SSPpAwEJ4dq2sOSHALkqiTHIP4sB0FyJCGK5KirwimfKESTHHVleeolSlEkR11X"
    "ngGJUiTJQSjDiYisrE5yZG0UyQEG1EkOwjek5ox0log6yVHXlfmxKEWTHHVt0AZYx2GIKiwHJYU1iZLD2kLKQYUg5YowhKY6KH1QbigxLBmUXMF2pA1HiNiO"
    "AZ8wrGkOw3ZoYTgjIzuSFyKyY8DnC2uaw5IdEtiw80uOCLiOgSBdWMMaluuQgIZRR3JDRHUIsoW1zOGoDhFqEO4cbvBNhCBZWNMejukQ4QYxAfjBEx0jQa6w"
    "hjUs0SECDeK75IaE5xgJUoW1zGF5DglsOLSAfk9Cc4wEmcKa9rA0Bw9cleXwQZ5GHZAOhCiWo67qmADSGmekI5SdpmE08pR0hK7TXIhGn5OOUnZMSGiNk9Id"
    "2iiWAwyosxyEbwc9p5GnpSN0nRgyrXFeOkIbtAGp3aFYDkoKaxIlh7WFlIMKQcoVMQjNclD6oNxQYlgyKLkzy2Es0q9kSJ5UGfHJw5rmMCyHojROJFHSHOCG"
    "6FmVEZ9CrGkOS3OIcIPuDzwRPK0yEiQSa1jD8hwi1CDwAD8kz6uMBOnEWuZwRIcMthTynPzgWwlBUrGmPRzTIQMuBQboCP/MykiQWqxhDUt1yFBLUR74IXlq"
    "ZSRIMNYyh+U6RLjBCAN7P8lzKxNBmrGmPSzZIUCuynZ4ixyK7IgyFNdRVYTTP0GGZjqqqvIcTBCieI6qpjwPEoRIlqOuCuciUFWd40BdFMWR1NUZjrpfyMsF"
    "IYrfqGrKzFgQotmNqq4JnChmlI8uy5iFg5SCykPKQRWh5VI9oOWKqIPMwkHqS4WGFIOCQcoV5MZkEtO3SciNCZ9zrGlOm9wI7cRIyG0kL0TcxoRPOdY0h+M2"
    "RLAhGZEcEVAbE0HGsYY1HLUhAg2piOSGhNmYCBKOtcxhmA0ZakBEHG7wTYQg31jTHobYkOEGMQH4wfMaE0G6sYY1HK8hAw14iOSGhNaYCLKNtczhaA0RbMhC"
    "QL8nYjUEycaa9nCshgC4GqlhLAZhKFJnNUCIoDUIVTjfE4VIYoNQlmddohRBbRC68tRHlKLIDUoZzj9kZVV649BG8BtoQJXgoHxDNi5KERQHoSsTYlGKJDkI"
    "bdAGpHaHWMJBSmFNouSwtpByUCFIuSIGIZdwkPqg3FBiWDIouYLl8JVLESYbm/DJxprmMCzHyAgHH5DRHMkNEc0x49ONNc1haQ4Rbtj9JU8EPMdMkHCsYQ3L"
    "c4hQw8Aj+SEhOmaClGMtcziiQwYbhDyHH3wrIUg61rSHYzpkwEFgAI7wVMdMkHasYQ1LdchQgygv+SHhOmaCxGMtc1iuQ4QbjjCg95OQHTNB6rGmPSzZIUCu"
    "ynZMwrlVCLIjylBcR1XRMf2jNZiOqqrTHIxG8hxVTad5EI1mOeqqjrkIrcFxoC6K4pjA4V9qDEfdr4OX00h+o6rpxIxpDXajqguqXGpnKHKDksLKQ8lhFSHl"
    "oB6QckXUQZMblD4oNJQYFgxKriA3HKVnyZJwzPiUYy1rGGrDkez3Cfqc6IUoB8eMTzjWsoYlNtx3nFEhOSJIwTETpBujjWFpDRFmGGhENyQZOBaCZGMNazhS"
    "w33HVrkON9j2YSFINdYyh6M03HfsYgv84PNvLASJxmhjWEJDhhlEddENSfqNhSDNWMMals4QoYYjitTlSbJvLARJxlrmsGSGALcqmeGrmUX9kAoIUXRGXRXO"
    "80QhmtCoK8uzLVGKojTquvKUR5QiSQ1CGc47ZGV1WiNro3gNMKBObBC+IQsXpShqo64rE2FRiiY36tqgDYjtDkVuEEJYjwgxrCqUGFQGSuwcfNC8BqENCgwh"
    "hSWCECtIDeOyebDWeTX41GJNcxhaw/cLhvCHVMAN0YqNBZ9brGkOS2yIcMNezyh2D9aCTZBcrGENS22IUMN4wyg3D9aCTZBdrGUOR27IYINIx7jsHayFmyC9"
    "WNMejt6QAQdBgVHsHqx5wg9BgrGGOSzDIYMNojuj3D1YEzdBirGWPSzJIQIOhxbGZfdgTeQEScaaBrE8hwC6Ks9hLWo77oPAzDqf6ZqB56irgugoCdE8R10Z"
    "xijWcYqdCs9R14WBgrVonYyAUga9lbVonozg0EbxHGBAnecgfIOWyzodBZaR56jrwrYjSdE8R10btAFGsUWvayIOSgprEiWHtYWUgwpByhWBCJ2Ig9IH5YYS"
    "w5JByRWEh2LI2A5NEeQYI41huA5FvkGx4IOI6NAUQXox0hiW5lDfsUMxQzEkHIemSDKLEbawDIf6ji2KGYohojc0RZJUjDKGIzfUd+xRzFAMGbOhKZJ8YqQ1"
    "HK+hvmOTYtELCakhSSVG2MIyGuo7dilmKIaQzpBkEaOMYbkM9R3bFIs9nIjIkJwOibaGZTFU5vYpFoI4cp9iIESxGEZjn2IgRLMYRmufYiBFsRhGa59iIEWy"
    "GEZrn2InZXUWw2D2KYYG1FkMo7VPMZCiWAyjtU8xkKJZDKOxTzFDIX8Ori6CdUhp/BAcIQSVQGn8BJyhNH7/ra4JiojS+OW3utCZqtBmVtxpmoit0PikYU17"
    "2oSF5lu+hfAXUcANGWeh8WnDmvZwtIUMOOjlwBMJc6EJEoc1zOHICxlsEGCAHyL+QhOkDmvZw1AYQtxSbHPyQ9BQCJKHNQ1iiAwhcikiQEcEXIYmSB/WMIej"
    "M4SwpYAO/BAxGpoggVjLHo7UkAEHowkMAWS8hiCFWNMgjtqQQFelNpTGfsVAiKI2lMZ+xUCIpjaU1n7FQIqiNpTWfsVAiqQ2lNZ+xU7K6tSGwuxXDA2oUxtK"
    "a79iIEVRG0prv2IgRVMbSmO/YtDwEOwGKQU1iZSD2kLLpQpByxWBCEl2kPpSuSHFoGSQcgXrEU4RJ9zshqbxScSa9jCsx8ibICQ9khey0zEZfBKxpj0s6SHB"
    "DUmK5IjkdEyGIIlYwxyW85CghhRFckN0OiZDkESsZQ9HeYhgA4LicEPQSgiSiDUN4hgPEXAQGIAffMYNzRAkEWuYwxIeItSAn0huSHJuaIYgiVjLHpbvkOCG"
    "7AR0fpYomBAkEWsaxNIdPHI1tkObGT2H+sUUECLYDkIVTgdFIZLtIJTlOZkoRbAdhK48LxKlKLaDUoZzE1lZle04tBFsBxpQZTso35Cni1IE20HoykxZlCLZ"
    "DkIbtAGp3aHYDkoKaxIlh7WFlIMKQcoVUQjNdlD6oNxQYlgyKLmC7fBdwEh6FjiDTyvWtIdhO/y/HCndkdwQ0h18ZrGmPSzdIQIOA4DkiYjvECQXa5jD8h0i"
    "2DD2SH6ICA9LkF+sZQ9HeMhwg7Dn8INvKCxBirGmQRzjIUMOogNwREB5WIIsYw1zWMpDBhvEeskPEedhCRKNtexhOQ8RcDjQgBBARHpYglxjTYNY0kMAXZX1"
    "GOXN6zOSHqPT2TkZOI+qIpwMCjI041FVlSdk8jEyKnxHVVOeEznOclFhO+qqcFridKCKCtcxOp8eoEJ1JHV1pqPuFxJ0x/EiGXmOqqbMkMUTPpAsR1UXVLnU"
    "zjKO81Ds2OFXRVdVCMdy51MJMjQYlFsg/T/tXXk81N+7NwzZshRlCRNabDOf2WcslSai7FtZojEzGMtgZmxpQUXShkiFUEqptClaLS2IQiVt2osiVKQU93xm"
    "Rt/W+/vd+7r3de8fzqteM2d7znOe85z38zwfnznnj4THrK2g8u8K9TfS3+v/SPy7SfzneP8/qNzfaH+v/yPt72brhzP4/6CUfyU+Vv9n4mOm5ceD8v+gtt+p"
    "/8FfHGvy02nRf1Bt0Rh/1u2/sk/564p+h+kfjp3/XfX/Rnis+o+kv+Oo8GD4v+6NvxH/Xv9H6v95LYyvMM7KyrI4TB6fy6KHwd/C/YNlY7msAIDKWIos9D2h"
    "SEQimHgA6nsZfNykoIbzTxmR8FsZFov/vYyE+62MQIJ+KyNiSb+XkX/vSyb8Pi6V8ht/WCKJ8lsZmfB7GZX6W18cnvR7GZFA/a2MAmF/L8P92heLxeF/GJfP"
    "pbNDWVxZ2P66slewYNljXMLD+SggPJSLLMaWExCOwgq+z5kjy+PTuXzBKmGBKMl42RkzrBytZcFao+D1g6m4xUWwUBganU8PDQ+UxTjRA1k8FE5AAIWxtEZ5"
    "Y6lwZhkK40APA1Xm5iiMVZg/i8lkMa0BL8ISYZ33bFo4h8/i8FE0LosJPtn0UJ4BaowCcAsE/8bUB8AAJOIDhbFjcQL5QSgcAQ9mirEWUHWN8ufD7M2mR0SE"
    "shl0Pjucg2HgIugGv7QXkhUpp5iYb3dwVJi/mJiYNvhkwh3ElMWUEsTEDolRdkdKCApAq6uiVguFrcJ+bhXF5ZjCLU1xLDyZySQQTbAEf7wJgUEmmtCZOMgE"
    "Ryaw/CF/JgPyh6nInRaRMxCSo/N+GxRN5/FYXHgWPNBOplLUQQ/uMO+6PO7cRpu8A3KGbR/X0NjCDmxGOAe0EJvqH8D0F2OHgdXB8KIDjWLDQmEKMf5spr85"
    "KEDFsJn8IAtdMpakiwpisQOD+KJMNJsVMz881kIXdqpACUpQGsAODbXQ5YRzWLooQIvDs9AN4vMjTDGYmJgYdAweHc4NxOCA0sGj6c6RNY+gA2EzLXTtiRAF"
    "DRAVhceS0XgqlQa2HgBQ4CNQQB5LQMH1VKrQz6dQCCgChYjGU6jA3QYRIgVHI5AA2Aj+nAVCL/iCbhwFjQchLxUgHBW+ghK2uHhBNRWHpwluJMXDj0ZgcKKg"
    "8ATgtgPEJ4MsET4FH4vGEoR/TiFSSTQcGYwOweRwoBsBcEEUPN4EzOJgsCeQ4cGEH1iasBgraoUV9YJ/nQ2oYEVEiaJBqDThoGQRD2M8kYW3psK/wxLwTBDN"
    "gUATzokknCKRKpoyHiUUwZhIRCKiAnYEIiOKRIgTiRSPEopYKHICQA3BEhBpeDxwjIFnTARkyKIrLnEUEopIAHn4HGuIIDCfRBACQGQqjYAHZOAAAkQ+BAKQ"
    "PAkMD7+FAtjEw2yQKILhhJ9EmrAcL2pHEPVDCcmQRGQpomFINOGwJBEbRBFb8OHWMJt41I+KQ4BwHiIF8rInkCEgJuHPZclkEo0A/BUKPH2S8BdMVDAsrDAU"
    "4HdA8BMSCEgR0IVAORZnRwCRAxFWQNAcGAcaKAPlwt+w4eDb58G4RIj6Qx5IG9YJYXu7ManisYANnAf8y3x4OmPFcOSHpcLVFDSJCB+OgEcD44LC4wAbZKBL"
    "ZOAs4cl2eLBYJCAeoD9ooCiADfgIavgpIbz4QIoE8EmAf1yFE0odzJoK0xfO2kYkBS97+JY6QFCkdQQ7oTYJz7wmULEeeOBmYClEu7F2AvnjcXZCeYIdIqz3"
    "ELW3+4Welz2OJBgRCICKJlHxNByJiMZBwt8H4CH4QDGicCagA0mwYMB/JMJbElZ7sCWBMwIsCBXICeg6kCIREryZhYatKx4+2AAHhwzwquDt8KAcwhIE0qfi"
    "yTQcvCfgh6fw4pIIgiMbKVQ4T4C5+yFPRpOpFA88YBKIzQ4HlBIiCe83IQHYwZFIYPXg+0ywYPZY1Nik8CBahMX9T54Mxid7CCeLFX3igRQg4SYWzo5Aw0HA"
    "KxQwQABqBj8ehh10guCkBiq8jmCdcCKpUalUDzxw8MhAeKJiWA7wJsIR4BuU4TwIUwS/Hgf8AGUGXgHAErAaeLBK8KqBYUEZkA+snUAdgNpAWCBcgBEEGBbh"
    "E8zhE0YBm1TSWB6sLoQFcqQK2+PscPClJWT4Xl8CmDYQLyBHIcE3+FEEEIOlAu3FUwRiwJGwqF9mDcQgUD+iUB1pOBzcDT5ThCCAUVhZsQBSBKMJ3kME5XAe"
    "7DEgYTv4YFoIRjgAoCSYeTwoht9rgy9ahLcoUAYS4Z882GJ4HEXU3G4M62ENphBJHgTYBQY4+70cIApBcC0xWEMcCL8IYO1Fk6VSgQwBDlPwVDtAEtQDmQCE"
    "IEI4GiwjAnxdEeiHx+EEuoTHiZCEiBNtOZxwzjYiCQAcIsI7ECYLYhDYUoE8AQAdAcA7ngwvASQYhgjYhWDYp1Lhy7KJsCwosKUCo5BhtAazwMO37ALRg6WA"
    "66l44QVeRPhKcSAzLATjFh4ACuBaoEgUGqwRFBxJKHuAXwTRuzAEgYbj/smDpaSQyR5gV4BZUe3gPAngG54ExiGTaQT4ETLYEXjBjoL7C6eFJ8Ht8T/kYUCj"
    "eginSxR94rzshRPCieaHpRFhlYTFC0IlCEwIBmAKFlZBMDAsL4CAeJHcACKCnQG2OgzEonJYELDxJZBgQ0ARbFkqFs6DrSK4DAnIFQY0oLN42DqDcQFY2cHP"
    "f+C7BWDcB8pFABsUSxWeLSlYXThqwYryFHhHgR0Hm3y4PZFoR6AAhIHANADbALwF06DAwA3sF4UCm0GAXPA6ATkQ4Pszfp6215iT5B9KZ4ToYoAPBPtCc4B7"
    "iZj5k6fnH879g6fHELh56Ggc7L4NwI2KAkWFifsChN9CBU0ZwGvls5iRTHYgG3jhruFRXAYLdstjrUU+GSMazY7gMwROGYcVw2OEM1k8jKg9T9Ae9pMxcHTA"
    "YTEtQwPDuWx+UBibYc9isulhvPAAfgydy7IMBD55EZMDnPWIwAi+CdGEZMIPYnNC2JzAwGgWlwd4+q2CGRPE4sRqAoAnmUBUEwjnBlFN8URTsG1IFBIAGC8u"
    "PTTUUjg1Ww4jNApEBh/BpPeLpDT771IKovOC0EwQe8BCiobbFAezYgEJnkBORSxBAKMhhqi7FhAqcPjVfbsF/HNg4gGoMDqHHQBaMeihgQG8IDqOSGLCRJei"
    "HAkSXVeDny89YFK4+mNJw91RKY/6oK/Wxb6mV9R3LLd+woigM+eJiYmniricJfTaGaG/c8kIpbPDhCsp7gUzuT+YzQGccRgs2wWxxrFhEaZsNtOUSYYoLBKF"
    "bkIlMBkmAKyIJlQ6k2lCDyABVSYTAlhYPF9Ayw+sA4tL54dz/dggXtsrmFEgLYjOX+jkxoR9/iJGFDc0Vp/HCg3QE8zU4pf4AfM9PBDNN2pOn8USSWzGyRKH"
    "t3baw9Wb1PhnZsW0NZPEr7FIys/7Q3gRLIaHcI1ZQNPREJvHDuTQ+VFcVqz9DwMJKGP+S9GPkJnv5Lgijfb7h921//MTElI0/FcU/9mIIrpWU/2iSks+dnfJ"
    "3X0zf4aXpGH2U7MOh+69rMLWV0aVxguEdA3+Fd3vqisi+6HckNdPjllioppNvi9/TdmtkU/KcanIC8i62fH02eTEICbDlM/mh7LYgigOHcEM+EFpxcScKD9t"
    "FxzjD+HjdwmDZk7SsCa2rvOU6yxCUPTVpid5Sm6AkiTXQUkSoQUS4ghxcSW54q9dChGb3ksmYN8tmMbkQROlZAyTbZI/TUQg5CTFoEXY6ZCOlIQ7UkJezdXV"
    "Ds0ID0PRcE6WKFuaJcpF8NY8EasGTYGbyMorgiYoWjg3IpwriMSxcpAMXCM1WdzdFdKeOhH4dASAnEQiFiJ6gSwZZPGCLMELWvhTc6w6NFVIVckxgsWxtEU5"
    "RtDCjVF2djSsFqQp5ElFVCUAMZQrixvNZrCgJITuj/NASIpJJCEUxUC5rHgSAiGWdzZYxc5efMebxNTTU1ou+i51mCzBLdz5mK0SufLldL1biTpFxFy7vZMW"
    "mfdyjC69Yh69u24xW9+LcfNNtOy2E6wHKxHmm5+OVEviY8pmdZihTE8pMlJyNJvsczUunVqDV15wyA0iW1dM+LYrboEcb4n7XgmG8kqtSY8OL8W8WnG0ozin"
    "dnKmg02f19CTlfsyFt8JPdtQI2lYl3Ex6vb2wPIlEw8F7wqDIOO77vRNwa2up8VitXytaPkOGXslEvrLvFtOvB1JU++dcvYuWSarlGmysz5l19C0S+jm14yE"
    "bI1dcQXkTNJeYr1NvXEOQ7JrwW6i9uFK6QOuoVq5Qbw6WW1qo5TtB3EJhBhibxIiEEiEAckDUWpNRiBGkeKQmGCxtfSQYEkTVKgWSgsf7thatiHenqem1fTy"
    "3b5iQyhcStpICiEpOQGBQDIgOkQdy0PiySZj1ojLNxGqIy8UVhmMSHVMYNUxsTVxwaIZLC6k/70rIllN1DWcwYv4qa9AC7RQSCVIAbAqCzogEWs7fREIyACu"
    "mIHUhVBjdCSQ37/p/9NUHAFpwU0VkFOQKh/ylOxvT7yexx2ylo1WW1isGeTIhxTgakVYBEgJcakqSF1KZqyzBPKHQSVkIfw/VUjkTHl9CEv1ZxAgvAmRwSSb"
    "gKiXCtCcRTIJYPpjQYhNZoLQ/pdNJZEkjhBLouwIZekUksU0+z/3aPDP3tnvrU0cCAzVuGkUwJ7XNPC1levxuCFr9puhALWO+TSLIcnYgiDt2eqpR8rfrfk6"
    "gGvdVOH3OjSB89Z75Q5v6PVkRfT8ComrEbtGRlAyh4nWdwqkN9laJdSezdBxv4b0Us0Lb33bHrc72/T8gknbCJKqhFUfLYvxOzDoUGcX2TAL22Oq168lL1We"
    "0PjwXT666+oEagXeB3V/0juxJXg3dMTsRTMMHDNWolu/THsUiF7R6RRLdT671vT+htXb9LZlDPgv3fWypoSXnH+lPbKrV82HGZ/+TpGMic6J9WC+Dp/forQq"
    "A3F+/ZTg8DB15vr2Ud3snsreE80+GpLrVgUtI3ovPn6FbTMlTupV0JaCy/QcU7uhKv24l/uDnvSqf2vb1oY2sZRnRil5+GVuP12Yu1Fcw/x0VWfYhyu019HG"
    "QVP3vX91PW1D1pRO5XhVQ65+opLF9KuMMMyMhL07vQ8ZFH9b+Obpp5CXyGFvp5lhRoVJy7ADj/xLVuIkHtpdKzbmQk4JSzKlkp51Hj15QHpNmSfhqBK31zDv"
    "aOzhZ/hu1/L6na2zSwzJhQ6lXxtXL909K8zq6SnDCbnIC5LQkkIsGoWuf1an5PagMGOK40V9Ys4IRpL9ULPC/nnHBIxGZ+hA9cZnt0a4MbcmlxjdXrLjqLyc"
    "VEq7GBJxMfP0pFRPKVcoScoRSkKShVCsMsuIUdeoZW+ye50zN/Rjri/O+WjHr3DsiJ0Jdo0A+rR+gmMXV/AffsgLYPm/isnArQWYTMGS8RDAZBAagZwo+z+K"
    "/0niv8OyOAzL4gCWwZa4bnqqqSL3+hQDj4q4A9pDIYOf27dya/fnTNY4ImFoeZ+ujV4o7XJBm+98cRmye0XS6arhL4qds4rTOkf8P8gv75fws9Fo4CHW8LJe"
    "qDtwGh1WEDuWKj+9mzT1YeuUhbSY3Y9t1DfE7dlL3efp4PvEQW7hjvdR8+9GW5xes0/3Yc+Eq+E6t7nWm9qcr5WfWfE+87PerDsJ4pOcojZGFAR1vT6/bvbK"
    "XMM3BgerMCP10crq1u7HMIw3q/I/qh5bsG9exzySz4KNbiFLbEbPKYfvu/NmS4323ZXt52+wLxXoiNlnWpntRMe13kTk7zE4Z31PJZYBZeOdc30zw768JCxa"
    "GnvSIHTZyooozSzznTHT49tPLXp7v+matGwbsr6o4+iOk6eds+9n7K49cc7p8sjpfXYSWctMMoId/V8+MfzELtDMRd6pzBS733LOV0FawVH6qlzZZwO75npk"
    "99sPJyKd4qNnBaV/dtpQEUNsbyp8eveWwetFV1PcPPVrAk+ETZAhPjh7Q63CuviJXNr9pwf2BN9YPqVVadLlT9Ud6rw49ycRp3rfNF4biZ6W6hAgy95gotRT"
    "a11k+GLNptuTcpdYmyxV87/TUPJl9VDNPXrwlOvEZQWZe0+wSjA9vTmvv64+a/PxUGHk18yyNW8OOKjxNrDvzK/6vC5jvaPSdbeTFqZDjwxYynENtPzsipSl"
    "2dN7W9Ljx8zSTIEBn/TdLElDUuADKMjPOI2Q+u9ZgD9atb8amR/t4VfDRVGmCV92MjL7eEu6L2YeW1UbD8X+YA9DoeB/06jN/8Fqkv4tqwnvbhN4u5mAvS4w"
    "n3+wKjdIX+b0KmaTTLaUKbsHjMavkTCaKBbiMXD9iPYFw8Qjxfub7ffOUcl8kaksz77zYlMJ5xXZTXrKsxm53ayA2eZ4iyKNXUXnmgKZi15f5E9HFnJqS2de"
    "NjO+7175baED8snm2N3ktZf5ZlcW0nvDLLn7HiyjudZnO7zUeoI0fVzyCN1StnRH2+uPfkiunnnITbkLVESp1KbMkq0HFKM95zxFZgx8ujds7S9ZyuHc3mBd"
    "cBJjGOxX7HB/6PbXxnt7zQ/GWGnKVn/OqaRebyyMrDJMmhR568XgjGfO08PaVB+/cN0cvmjV1C2+VWfl7yjExE1PXHGj59M0rd1bTp/YON8jb2uIpaHCSEhu"
    "2juNvbNv+n3oamZqt72rSV2s9jbj5kBZzSxJ+4r0QvKMna++JTz8eP4aOf7iKYRJ4NI+zOvak2mRLxdIrIBOF5+ImuOtt9rsltt+08prjQrT8Kp3gokrEqPX"
    "YEaW3GcPyc+Z2+u2FZEWb13LyP+i03N5DSZindGgR9LHA9f174b66sebXji41ExNmX0yd2trBH80jTPzXG2p1TVXyezcheVXrUt9B75uSFMP2ktUfLkyOItV"
    "Xvq2cvRsxZPdixaeuNGtikkVX6X6rixX3v5y/Q3EAM5i66eZBLkR3bQVPZ99fQP9nFW8H+h7DLd4Fp7pmLIrQOZ0SzL6y4RVq7Pidi8IWElaPziwNxAEAG48"
    "Pq6Qzefx3cJDWCBGLmRE00M9VdKgJJWU7yozQbwgSSUWFPHFEQisIqxMy5NtECwJpDgAZigxRUpO1FJGCYEsiEVGQ3DLsb1xCQJ7A/tLLyRKuX+L+eNLuokH"
    "NsXfnqiAG+k5gBw14y1D+rRe0Es7eLJJXHp++fZXHnKFFmpT4McFEBXCQVQ8kYiHHyaSiGQvSCIBgRCXEatMNOJXKqBDCpKUAoCRHAQxi6/IUNIfWM/Gbdy0"
    "eZHUmUsjy5+Lh86NPPOroYz/OZYQZaTlxWmW2CmQCpyZID/Rlc5BWXNBlM7mMcL/s4BDHpKFq+TkJdxcLbHa0DShSVQVNQSFKFseL4rNCQQGeCzQgYANJZBw"
    "JNiokiEyRBVloaj/TeZExlhCXvEH5uxYdCBFRNHvYdFWOCzaCMKiRLE3DVeObP/W923WnmuxB+jIhlEFA7rhcOGGJr+Soh15CqmYtPQ6nfPHsrdZSE9zT6rO"
    "3kIKVj2Syu7c9sB4AW5tiWxTzEti57sSpvdIxZMtea/FUYWza4mPDiiRu3Q/ydjv2HVqHa8KvfKg08LZBaf62Dp9x1IMVniOKKqoEOd/vsd2mboIRX4f1KjY"
    "b3ImKRuluVhZwXqSj8U+leSKxVYBlCIH+1kNWWej5BsdUoL38GtRqpIIfb+Jl1f7HB3R+qAjl9Ujlv7YfCB5t+EkX4U2xWnNj+q/TKyzr61f56Vma/yCwtNY"
    "7lJh9+7qhzOn2x8FH7yBOrus7bIPUYn5XkJP2XWu/xaWpPKyCKPb9SEyiVYunRpLerysl91fFp07I04paK3zvpx1UencHPfR9uc0/+M97MBZluevbH0fPMFo"
    "rW3nUxM1gsI56oSywAunsrTrxE9csAps4iI2bPdzXyxlfKSsp/ijflmMxJy1G5nOW9M3yVOzFx+zOPoFubNwjtomi76LU0mF6AFGudD+xULRv4Rkv0YnUwRW"
    "D87DKz9m7qR/MHH7ZrklFW1MiLe4ktw0N+JezdqdzrE/WbL3Kn3Hrk11s9CjrX/YwGvmJlhUlPxmT4BKoIb1HS2rqs7nnMBNQiltRu20IU5yvGztsfrAnLmn"
    "i/U+G9rMeeKcp2M1L7b63FqHTKf63bOygx7b33g/OK01rrT3TbXO/bkDOZ7SryKaUj++W5h26Hr36KbVKk+t29Nf5VZovyBYj9JbY4dWUajPXXo0/QI6eVVi"
    "XLMXJaVbkjUN95KnfzZ7HO/j1m53Ldhw8ZqbNP5yyQTe3vqWmLKUg9/Kj68L5lzaYHhjV+Gu0qL8JJsJ8uQr4gznHQ9cVoZrB9azpeffuTkwfETTN23bsyfS"
    "ceZozclrg/OLO2dYbl2vhOh5oVZdYLp5JM/FSaM3peMps8n983TU49MGvXtmcVrOEqxLF+k6vvZXjXlkpfeIblKDNdh6uXJ1nLhU6aklt6u7b2pJPfWJZ9ke"
    "U39rX7rv2qx5DvtPlW+9ubL2FO1A9C1fiPGuZOpR4mw+s9SL+NwbYzrHQZbeBtnTdo86oK9Or5xjaoxSXwJFBW9hLi5qlmmZcrcgcKZlWOCQRER0xYyZHp/T"
    "Z86qXWwLJUmuBsgXIEI+5GZkw9Urclfuqx873++gnFX85rz6r8gX+78JLpogNBaAy6QfwEUUd0CoX2BPTRGHFRR8B77/U1T+V/CXev525bO6QrmL9HMd5syA"
    "1Qcu5oYp3ToITZhpjvz4JARFfZxF6DlYxJK2eUBOxq3d0ii7qKDqAyG0oCqnyWhP8cfrK+8pzuxYXW9f1MvwTn816FnSivji8PkexchLoWm5V05iTsld16vP"
    "1dvQNodpS2S6tr9f+DnUaqPq8jOfjmJ8i4gO6ynr2XsIec327TSLJkLYV0bm11sR5xZ1RncGOwTYFvR8HVK72Ltr6pmSmMO9ZyiNlbcN8ZejFdvmKcdcH7m2"
    "bKrJvFsdaeZVam3zVCx3Xt8+13iVWIjL1/wGpzMyC61Un0rquX857hjulpP1cdXZoaCnutNWPNyYTktTPY2eRDvuP63gYoP+FnTvpX0t9p0Vc7E3Z1pcEXN6"
    "8/6ePorZc4eXj46IPdThYS75+XZCUeeRssHhkr0+14NKFzyAYt5fJvBD3MOupgy/x5nnRypYTf7U7aQ4PNKGxM4813zZnv5s/oOyWiOOn5G65qtnx3o3+yxo"
    "0qnirdpa1HUhLOn8DNY6aYjdoyqEv3go7t9y/ScLXf+/IeAfAe5HBFxaM6/AfGG0UckrTkl69pbsujP3XvzJo87oJvt0hL5l3j9aqCwDcY8+5EoPmJ1XXGHR"
    "dvUpZnAtflWnzpJ5ag52Ow+F3N55TyXtQUib3VBDsyf+VmEBxUdvwwL94ZSng5lSs84X9tj5BS4bJZZiW7vjlW/Po5Eda9Zp5p5l1ikeK9JavYs09LQ56aQT"
    "7cQT9wsJ2+s8XRwW8TuWXbNauvvMF+iQ2cOF5++gjSspagzjOK8zj9yPWMx/vEHiZn/QCE/stteZDn3xHelVrwNHHu6X/VBsrXzSedqaSTc3l/k9V6urmHf1"
    "wvPPbYsjm5nLXjJrFlf5xrRYXPPOV70Z/2379teIxdLO94eZH0uV+4jb+wcvrjZSeHRil4XTI6fabOaWsFJ53p7ZUmbImOYdPeFWjsfxaD/rt2tPH6TrnL+b"
    "dHQoIbzmOVsVehW20nRBhpFUjknZQ0L1xWJDm9cpdR0U+XKkPL7h+TqniP4T2os7NjsOpqs/9Fv/pf9Vy92GA+9O7RmcHmuQ2Pc8YmkMQil5n+niHWutO+/X"
    "BEfXH5x0ffo0RauLrOMFyeiE856L7XfTdMJO3V335GvfnTr6tom0LZOD82r7p5uvWyY1YtbWrke1kvPTipz4BfetICFrQbmj2Mta/rdnDinc5I99zogUbekL"
    "t8uNiKFSShn7zEKC885+qzKbkBtdI39HjbrG9v7z7dPmOEkH37V7lkrP631780rtPE7ECgpA4FcAgW+IEHiyk1lo6qY+fQm88dZPjluGVyV1a/5/RWDirwhM"
    "9Po/5O1fPf0ZzOhqVTl7JP+pxuGoEg/N9kPHTwZn+GXHDWScUTHRH7pUpVG2/GL1rqa4XQ9Dd42QdlXnTNk394Vx3tniYifeSdmazxofLttiYi+0dp9/Pffj"
    "7hYNxfcF9CLj1BkNzlLKUz/PUMptvYr5mHrZcpLCyEztW2HEpu5zxOXzStY7Xxx8ty2n37A9b4VSSPA99D3Gh5cqPgZXknWrN04LanpLyT/olO2tRxumjM5f"
    "aWZa6IyZUW/zEDvqxH3Xt7IwpkktPJH+Ye4lTmer5vH90qQXIbjhkZgS390+jx5ht9ltYY1OSbxueWM21rVBjXbDBslZ1cacctJA6cUptOkZNy+ux5IPJcGr"
    "tHtf0XUeseoUbGRsd+L4fbJFjY8yxVwP611IWo863nuxR6swRuydqTW+KL+7rEsKG3AqIPxEambB6iwFxek3LId8+7Z0NVVW3cEvi21unoW96QRxeVra94Nr"
    "qjJvHnmoyEwZHJZ22BA9MKd3NMN4ZwxablDqYc7pPBkJ028fNWSubfXz6lptM1HGqFx741pN5HtvVXNjJYmy9KI8yzMuQS0TlTXULx2ruu+4/MI7g0M9KZsj"
    "fWZv2rxh7wwVhlW75Nd4+tZQazczNZ/6Y0lla+uO682dtkfD+13RXbdyQ9ej59Y/2P9lR373ulTuhxXz6qecXr90oXe/Pp1aZNfRfrVE933+JW874+tC+A+A"
    "mH+Af8Rv8P8P0v8RyP87SL90KDFdPud9KH2kLlmrzXud1v62ixX7l8kvs71eUWjQZJJXsfRz37pUC/Oh8DvBOqcMi9wz+gITd/qWu8kP77JOeC7uLOZapSBR"
    "vWFJR2+S++rnH1epvVWc+XXt6PSS6sMdDzQzS9WbmqFoNdPMvBuHAiUmUjXQCfzHYX2rJ0djut44WaZRLEJfPpnzDvOA6UbyVb7ob3W6ZfqpBPNNam/v57Yt"
    "OnjHjeUd0rJ0dZVtytu3i/0y5M22392VPFjoeWNhH+Vw7bE1EnxmS5WLWeHx+NkhTxXYO4bE1eRnTzWbNXp0n1L+J9b1go9vKi9sv7HgWc5xWZugd+4Hpxcv"
    "3zrS0LaqJT4ZlavwoSS2+er9eei6REk5H/8Sw+phj9dqcRO2KJd1v3xwK2LEupFRwuw37yorZp73vbRe2kT72I3nlbf2Kd91OReRlpw3jL4kO9jqbqQ334jm"
    "012YRJ97M7Ln2KcdXS+u15h/PRxz5vRDVLVdV+oJxROYNbWrM5i50uXrOhqzv6UeW1xOW9XadST5m/o9kz0XxN5VnVroobfSRjVdwUExMBUV7y9dXe/r3j8o"
    "ddVZb0l4b8pFFZR13MXt4sfrULPzsjZumOkuGeDSq7bm3uKzJV4qXTXGR/bI1frM/1A4MIsZF7lJ1iv/W+Vnq7gJEdpPgyXKydlbln6zuuQ2Y9NaCQvFezPx"
    "es1LY/ZXYJMkggDSM2F1S8z8P/VV//6M44enLQVJiBmQxnc9lpHATvzpmQ2E+acOidVFolpijtvP3fv40b0ssdy4Qd382a64+ELNrIm7L7Hee+pk6EKJd34g"
    "II7BJl6BEmugxEtQ4nkk6oJj5YltdomkwA2cXsunYSPIe7oT8i55ZyxqLrq83Fp3BZS4A1qzb9X/U7n9sNmRSYiERbigoelX5ts7IhvtP57PJyzY9EX3ZOQd"
    "iVHmN8dVnXOwC+4pu9rpbE6b5lZn3FLzSM/rTvrtJtvHEEqzLBzFXqIVJbFhbSjnXllrWMbqmNii4IxKKfebTz4fDn1bP8D46jSy356aMirzkrReyq9VmeHp"
    "9jpP5mBOTBMqhWy1b5dGLvHaTI106iccpJuezTLjJHtqEm4OLz28e+J1nB0JvfauL6lwz9I5r7uMLodtW7414pDa4iVbd1cVTeMppz1zzdZS260l53li6ky1"
    "1X7rQnOe5H/0VVbZuXP086K8FMJIpDcUGtnN8x6Nmvh8lb1kq56+D4OgMGi0wL0l6tO9j2EFNlAckRJ5Z3dPj+K9PEWOeKM7k9KuOl33YOyUlQ+mPXU7Q8i1"
    "Yu0M0vWbIP94lp+48S3nuSnmNct36O5qk1S23pn6IdHj0QvjyLAp+6/p9Zs4v/kSzAuxhxzPVV+ktwXTPgTntbzFScTGTRnQrN6c/HmEUut/WMl90+bVJbM9"
    "9nSFvI/sn1HxjcX1oIfyCoPgB+Pwt0TPCXehpAm3ZBFArSc0gq91gr+zCh6jI5BJEypBUTmU2FUEEJ/xQbnIV6Gp6YqSS+XcD7eIAWF1q4GnM/Y4kUwkYUGs"
    "mZgDJWZBtkD9JI0UJMQ1wMKrPE4iu/q6tuOPP0Fa6BP7o149lVX/8x8Efn8zIkHstzEKlL8XUQVFRK9C+CV67X94l0IK2mDhNgRRt0I9CP6j2vc24shJSKW+"
    "gsTtKVdkHxMtp8ruPKnT/qenMdnh/OkYWYP6rc/ILmmaygaPFzZ5lse6dCmc7/vyah7+8Ko1V0OV3HG1cUsmaxUbfHqUftG096vr9FjXzvML5x1/USRndm/B"
    "9hTs+WzFlwO+/U/OB257umPTOtsjHkeXmMWmstrjN3Ze3cU9r7RB89P2iY833NZdcswHsXFpffaiIXsHnawnE1g7jOsDj8XmaCDuZfKgZI2Id+SLnhW0h1td"
    "znnvEes4jOQdrkrQvO4ma6ejsK6ORc5a3YCmq5tPz3uSUypx9mvwUcU3dBVTYzlnH1bwg+KtQ88tm77IUvtVZUZz5tsEUIu+PMfNeGg7s2nfudxJO98tvKkY"
    "LhmoSbAzn+2wOXj/wINXZw/MTX/jOjwpOujc2+iKfMO7zxPVXf37yERjWnzIYbeGXY+pzeUnQl6pXD7y6Nn14EtS9nN6butVSCkXib9c3rxvlTl+UZLvoiHz"
    "dq5kAmUz9tGxI0df3Gvd+jBWc7HyrLQn81JNGy3tMqctGo18eBkblPl+t+mRLJZ8uqF+nd72bXeKC5IksSAeMAH/DYClUBa9Q4ORUdjdedlGo8TKviXQiK79"
    "v/8ODRbCAceeOPYOzViW4AXZ/4y+ypCikKr0GFWskeCvZoAX3T/x4khzdUK5sHgR4Rwmi/svn5lceT20s/7zS5ObQTFd7HSnhwfTJ/k4VpW27DR4nn3vs+cn"
    "zYCEjdrZ+Yal+5M/M0M8Khy2XX4jb7zAO0suwid18fnc11mWQx5iE2+hZrpxLCZdPbnnNN9r6EFIvlPQtrxpD7225SgcTKN1JOXlq99csfytdFrZ7q0GOloV"
    "CskyC4+Iu3v7ym5907hxIEx/ojqlc8VwdU1zj4TRwOv0iNcrKm9rJO/ZMZkwKYj3ZI8ybnnZ1UN8675yKOD92xV9hxIWxNL2NJT697eescO67Ndf8E2vDdca"
    "V33/sHFAd+eGDNo6ZUyYcVZmonl/6mHErLTbtrdC/K8GGc7eXzdNnVesFSqX1BV6+5hsg9aHAzNCPWsqTuxR4XIIz1Q1NSfujZUww4tFzmg4UoyorPc+o0Zz"
    "KLrlnfSV6cs3i79s2utamj3cszYyN9p8Tw1HIqzbk9mrI0Etlf+itGnPvNFK8+PSjY4Sj+1nEMoal1XOXx9C8t71FOLdff9p7pO8xpSWyoyba29f9gkY1j4i"
    "2S90mhNToMR1/9X3eBT/ASBJpDhQ0z88UZH5wc/+E+r+7JZPSPiTWx0jOaH77CzKUONAqJy66pzZt0/K3owZSA5aGbX2pe3loC+1ARdaI2K+2s3c8Ala7KbD"
    "vJG+UqzsjFf5i0M796SQupbnHtPdcuGVymg1UbOd6Xe4RXPO03mpJWqNqa8LcAHl6KmX+qn2Gt9CeRcHxUgRz+/4xd1r41bj7D6uP6g1v8I6KQF/Jcj43KSX"
    "X3H5qX57mkf2pIX4IH35e5w0+yUN1FVznRrEL3sMDckUx8Yon3yRvvFlg+HppyeuHJ9+oWh9MmXv4QpTcaqYom8bVuX0yFnTQwVn+1wd3n8wdGxnrk9SifN6"
    "6Lgm7bZrjcWKqtP7L6wzfKtG6M7ADYuVs5/7aZgez3SuX+qSk9VM0i7b9YgZnCzny/A8leRwYOPXI59oVFqks+2Da9fmpzcstyqjebO/LlI2xw6uNFNMyFOt"
    "8vbtK1FFzZAc6LH6aj1tJL5y0N7YUn/f8rK9H94+rxiIXFtne2Hb1Lmd7elbCEtJyY/v3Dg940qm5ydeNR6hMgyFb6qU7LJojQ+aQxr83B1rOHz8S998Pvbd"
    "rnZNG47a8aH8ZqP6y1uHzVfcvPZCp/VAsFYt/lZbdUN5p26m3VE/wpZedvyDlY2sAewp91T3JIc6Tn3tA8f0c753rVzdtFcF2X3yEb9v8Kg82pikuCJNeeGm"
    "HsaTvaWEexpXjllVns19VcxFPjpdMb1j27VA+C1az5lRYuNpPI2n8TSextN4Gk/jaTyNp/E0nsbTeBpP42k8jafxNJ7G03gaT+NpPI2n8TSextN4Gk/jaTyN"
    "p/E0nsbTv5sGPRFi6w9er8qf6fpOklakKF29/MASc3T/bEpe1WHLLNOoLv5GhZhnOjbGaCaH47ApJbmKz9jmpl2es7chcFDlQrubR66p+KWDp0ylvVUW1ZqU"
    "eXIS5+/TeBVGw6+eF9WhP8pXfng8sLe2dKZhsRSVfL3G0iZxEq+/uKHkCkdmIOxCw5d2T+wXTRt0c2m/rXHgNx/oQlcm2s404tX8D7Uvm1hu+MijlCNrM128"
    "06YdjSf7d2WttMvdHKakFlZYY8Q4mdX2dk2ji9SsOZhJVhgpTOKLK1M1nbVv6O6CVqUm5qnews2u+OSR56Cd8vqkpFmzj9KFyHl31O1N1yPoL1f6RWduUnpv"
    "NW/D73cyCA7bHztU39LahRUqeNGYF8SOQGHgl4H97EVnZqIwC1g8BurPZ/ZjrP9WYSU8m98aJTi+3wU+e/8/PahfeMEAfFWAawSLgcK4/43yPxcDCG4swKOw"
    "ssIbEEhk0dUQWPgKBPg247EKyq93MmAJWBwR/6f7ElDC+xJwEGrsvgQsYB7jxGVFo4SXI6B+vjCBgCUQqKILE/4DBZtIVg=="
)

_REISEKOSTEN_XLSX_B64 = (
    "eNrsu3N0ZV3TL7pjO+nYtp100LFt2zY7tm3b6di20bFtO+mc9PN+ON8d7x133L/PqT1W7TnXrKpZC7N+VWuvLScJAooKAAdAAgAAQgBH9sT0GxAAIAoFAKAA"
    "IIE1f9jaOBnbOOkqudsZO2rTuVlbEeSDAlPkAoAB/5f+j6aB5DUbLEbE4CvqB1yVriq1hIby8lBSdHFq5VNChTkIjHP9pUFLTbyQnjXX3/xKpDiDKp1+YDWf"
    "Vzk3kRezD/mNgzjEmFdU9vLwwn4y7tZa2mG6CLrgOUICUJngZ7gqBmEk1cO2W5F4yUBl2yvRVLOFqNWFJIt6MEqvP6NFi783xT6bVhoDWAOwDg7snuWprpDb"
    "WMSygjFibOyhY9jQctDfn8ZGBzjk1ys3z+SSfiJWyscul2ZyfwAzXc+XyZd2eliptxUiIrGwhpY1GocGh93gKe72iI1UTWwBnaHLzbsU3/U8ciLRZpk4ZYXF"
    "mx7qlTL6bdq3ubSk341wpm2/NCNlc1zi/oTvxK5le0ywJMGKGsQduAfR0EFsrb0dHV4ZeFNv1d1feBg7U2nSWydj9Hx+MY5/Ocs3Jdb2Un8Wwk/5HBXdBRdK"
    "45uWFU7mROQoKhI+9Xz5dUzoAIL04L37u9qT/br5EQjfY1GLZAnh6Mib1ulnhUKD6yxEfEaJ4QgtnF3eTPIGHk4nXOVPRkyvTX8Xp1XLdUbWTeLLDyTh38SX"
    "/XyDfsrscLa2vOzRq8eXFbjfibsNgeGg9tlTcz9Oi/+0mKmGWJXRIGJ+t70DMuGO6jzoH3Aueg0tCengZESbxsNRrZ0W6mLNzDI8rybFevjWNLGUe4P/ddk/"
    "P0EAcv8jZjDihk0SfcWMza+QAPMVM3QdjK0c6en+8v8bK/4v/aXK+FkJEEZE4S2wT8L9Tu41ZaHCta5QOTnEMnwGaZPmPHEyEvm1bvaqb9qlwjPDRpfrma0J"
    "o1Z7lW5FqbBvKtEVGtC34Y033tQ56K4FlpQKIjWz9SUqNpSjK/fvVrdC3xzNJahR4inovjc6JDm2AIlZCMMrDvp5lTnh4B46RPKCKlmeQ6W3dfk/b/yQgcg2"
    "qNbEPHoVDB9kS0z53IWLZrKYgAwnjhKkMWLMYIpQ0zunbFb45UGzZF5TCt0YuMc/icA/VToS1q5F6dKjbbpgM1Ec3IYpAz1mIFEM+QiloRWPG8+DY3fl2th6"
    "9Yw6+/yE5K23Ue9hv38XR+vlj4+q+2JF02GvRZ+2VKJG95kQxJbWQebtoY/R5wcJ3u3wCv/vlpRYu4apFigAAAINACB87XGzone1dbA0sLW1/AvBhSqajrts"
    "qD5XUp8BwzVlhuOM4J6CLoB+kg0SY2CYDe9RSE09NE0HUs7Yy3PfLWJsxxUnLpYzIPWlq/Uux/0zhDUeZp3Q2MRf1WF0uPMENIdQxhXVL+Zc9VzFhGgmUFMc"
    "xNp3SWJ3/ZM13Vq2LWr0HLrYgB6jJakkKsG1ZHsJBxp6dwGMSprWkTr/PLWVekRQciZSoYO0C65o26NUpl8ikIOcGF6thYxhblqHgWVYQrGw1HK6xQ2xcBp4"
    "OSAphnV0MNSH7kCIMOnyDxp+6SJ4Ia7UQ2A6RnqP3xojq37mIPju8++xxeudRJVmDvlLdimlAKyXJLrr8DnZBMnplrseet3FsmlHYBYZ1y1J310tDq7q3kEw"
    "HHyw4IAaqaoDtztc8xsf21oPJx/iTYxUFJd6hG9CHXHLJliYaBYZ8sJJtH/B/egiDAlLNg+tP05hCVBuxm3Bne9amXengU27ncTVfMLYKydSWQJdRKkfMC4u"
    "5r8SodiPOGutrpkG0fZII/LXFwbPpNQ7mlrNMTYI4hXjvN50I/JNUUIUY9fDdQNo4IS4EQ6rFYlQ1q/WICPHkRhFVb0/KBSrxforqiTUAULsg5kJsuUSeYfV"
    "xGZUkB+1cN1lQjf2UxqIiZbC0R+RVm7WJl7xnYbrHpPhLsuly9/42uBpUPR/nNQhg99f5XtCnoF288c687iwbsPnL7v/CceZxnXuyGblj7N1Lz1S/8ibb8bZ"
    "ws4v3YD3L/2Eq0T0wXDUKncSxYlaYeXVG80oguiwVolFSWbhHPk8PceT7PY1S2+0LenYDR9ok8KUPjg3pqYrlbhLb0lYy14UlFNJf9Ckmceckqh15cj95BHO"
    "UmVXTVi5EPVpbNtGuEwcyhnuRo+xYLJGmM3Zi0j7WDK0GifQ/YZveHW3G8OkrtXAPv7isemJmEGd82o5Vn+Bfr5Rbl6zxguZsS11Eegd5uiLFukpoDCx514y"
    "eZWmVJEwCWYMosN0jVNeucj5rYZcIrGOCowisuHi1SNVvj4u5H7uhyF3Cc1ORrU5ITrRiwfXulKEY3op6i2zHc+ywvQGcptE44SKZo6i203hHYp/YLWUobjf"
    "CIp3u9kNOLhpdyyljhIxBSeXv/YdiYM8bisST7KNDJEpfpmutoaeyAObmrIZ0YgTRfNiIhnhAloLBRpGzhzJrCTRAvKgtoL73vM5Sl58ZatMBzFoafMbsU0S"
    "WZQWxffEOCEUsAYOrAunNHhyxbuhSFD6yFDb1J22bulRFUzvLla5t9ItzlOIkK6I6Nc/eVfuh43eJAhzD0opGKaMC1IfczL8QA7r/Gfwc1dF5ZtyqRRuXVun"
    "91kVPY0dSFgC4eDRi1dv2YvhIqpVbOHg/b/FSkr/ZEM2Y9m/TFnmG6qWQP02SmyTUdmTay7YOPdxUJc953FXGOsrEcuo5/7QxSa/7kyMY2aPMTG1weH5LoU/"
    "te7Jl/l9WhTN8s+iBjYw8KxlnwG78uEh97/F9j/mp6rQX9g+AQIAYAEggb4C0b/g/X8PR/8J9UBfUA/0fzrUdSTISg4wIA5e+W8RTmwUV7WFWrVBCC9XVOMT"
    "FM4XW2UkTjyXbG3K1Tnrl6Xs2m/GPvnwMK1tmC6Fu8IuuJUQSCogHkfSY6tSu/HyaG/4EfyuXqhyLdXj99e2d6/zHlJx9BJqOhEXQheFpnH0pSfFjMgqHng6"
    "3i2+zZvHGZYb18koqJY/NG+L22mGQCXPNtClgny7xRrQPza83drsKscKiNnscD5+lgrcHwmalQvBUu/l8IlQRxYQyDYdWTJ6vt4R+xZnGi1OxGlzS5p6nqP/"
    "LWOlEsrQyRKv0mZ9fYKEltWhnQxUHbfWaweR/qdJLIo71KGJMgbeqpvdfEiR/Z+wFHJ2NS+gf3cnpafaptp/QZrBF+Bh/DekOZoZGzs50v/zxfj3ZspMWIte"
    "I/hXWYH4xDjE6WQ4qK6sLrzRXn66k75tHwk8BOagsOh6tT0wKtFqv1SPiT3xNN0zvftixML5826MDexAXAKKZS65fOXb/fDTw3EYMr6QIrQ1NsUxJo+3js97"
    "LYdG8x0rP4GYmZfAGeJqJiO9JUMfeVMQuCjVWSTk80nTRDO80Io0U7qxOZehSGmwDmQrdl+tchFJh1or+0wgzIDfN6C7p0YLb1lmB8gpyOa1ZmJz6YjQqDmq"
    "TG4KUpHDmrUhrrpZtacw2UAS8hDX4OmWlJNX/MTI+CkPzSDQ0fP4O8MXRzZTwU2xmAeEvmsoRzs5e5Cr0nNvypuTqKQN/T3XPhN2NS26lCtl9Rm+kwcs/f06"
    "8StpvscUOvSiPtxoZSQIZQG+O6/JkkRa9Cw/HyqOV/Yrkld7wws6+I58mXBIBCwjJ3AkDPha4CqYtj6r9yv4Nqwj6u9hci4FXd5XaL+zIrq56pHwVFQ+FYxL"
    "b1q2zSnzkD1xu1b20OXWuPRsk73SDXpnGd3dfYDoZ8tT5eOFSc1ZDepIRQaJQ2RQSujwrhbgnmtVkETC1t/0pOndhbXI9sicExh1yAtqEY6O1aKN4Qe2c3DB"
    "DWMwyVfmE2VNMYEhFMhUoYeojbOvxV/5IFHPkHuX6UXYOJB+V53z/RN2/mGZKpo2AQyIHdC3QNBNhLGQsVhil34UCl1zVELmnm/VjMb4M5Cv6Ct7aSvko2+S"
    "Sxtnh5uJLxcfzC1oEhW5n2bVvO9c50rXRXvFvB3eS9pbLkueefnoBWMdzjOqIbSJ8kCZT3avFM5438d4rEQHZomdBAnIdlMtpgzApHlDzjlGzVElwmIbUcgC"
    "t4she69OIXF/H5C8w43gG7kpiIvRIsjsYe+ppaK090LpldBgxxkPg9BdyAGjpAkZLW+VYOaFMnL+oLWMWZmVUM9z1OgZCNXPG+RpgHFz1a5dkBldkL2GNEys"
    "2scwj80HofLNuTSPZZQVumCkGcFwVgPoJbuMtgHMHhtK+2YZ5wAK8qS+0yD88Cbw4I7tnG6BK0m0+xhlrVKpd0/rOCo/0wIw122i4auROtU5hwTt7SfwUELw"
    "tuhkIPpDs8twClCHzxnsjzsvdhv0TRPOMIqefdyrTo75Ti/RXIIyy/wUj8lbutWvNbw4f0tHega9OpCAy5TRf+t3BzWeteq2qDtV1cjHZyGG0N9OTB5gyRKd"
    "iS+i5cZn4xKH/+29vIzWLv3nedm8LjTHd/oEuOnXFUuntJ2UG3po54uWrVaYWgviZ+z7f19Rp5hx6D4GRGFf0n2Q84BC69wLEUQtyEstUW5pf1aUOYGbK+kq"
    "6kH1bE+6dfMAYYR9uuERkfsGfK5sR5FpAO5Ti7BvCAf3Pc/ZWi4cjhPL1LSVf1x8W1mGxn5QZZBGzS4Nl+GncaC+eZlIYr1c0u55pJDbm/+3e4C9n2CSMGUC"
    "vTtW6P2P+erFzU7FYnLT99G4DxrtP+Iyjz1uWbSAB/p/C3j3ojnQnF8hKv77/2uYYvonTClp2q4yoPrkNHwivmSRwEOBglYVq2xQVc03XN5aZYLUYQ1PsIFx"
    "XXRfcaiSkWCx+uMM3zxlXa3Xt3wf6fMEtw6kSgl8JbzJ3adCOTAR9/Hqxi4aYE5OwqRWgpJ0AP1RXbPZrXme5MPE+fMhNo9FMl6zPQvCb4dTzMSrMQh5qzIp"
    "eWK3YBPUpMLPGE3smwBqcAI/k3x8un51iF3pg4Zdi7IXeV4Da1iwv/qCrUMdK+0aIfZAinnTsJomLzmy7Rk7Ln7U9L14byOThbV3tZGcJHeMw7eJPxGKt4R+"
    "EsU5LmtIyLaMLUkUgbwqO8iyty8fxUrpVdCVI35DkZ4QUd2Qn3sovefr83ajka5SR2k4Qj/2+Pt/2VUtKgrbjhNJc1WXkmN3x1nNxbU6nbzIGEnekltVI2mc"
    "2uxQFF66OHyQqkZI6V21lvLAnZG+8B8h2Iv5VwgvrvKBnU+OrjK3rvouEL3RSPRw9kL9snqABLODOU5GvhKmfBlJqyArGh29Q2fnFf51RmZTtJDgaBfTg1iU"
    "Gj4krr8ebtriBxzDNwMlc3ElkaVZ0JA779Jm47yprPiq7Eltsi2C1i1k4ugUocGs9ktb/23S5ypj1Re4bpQ2O/K9DAVew1kzVX0p/e7cDSZ4ryHfiM/BoIVq"
    "hnmTl3JdFUIzHZeqvQZ3+p0SztRQdNR1iGoF2mKGFR47ry/P2s7+eJ4jUyjFXu0OOqSpV3gUUWt9e6aJ48me57dzOwUamDLd7OMMTJ7+R/mdWcbdgmwCSXYh"
    "HXmU7IKvn8Gq3DZER54YhLHz7+0ZtpYy/CIoaR3JirPktuiWY+iVH9iLpWf+FqusIHQa8eOhTXBmWxC/CA8KrfFbhDOMT4flmaRPqGmF1nRyE+xXn0B4g46I"
    "xnMtbNl9lae/7u3/XIiV6Ws2ayKoXwtxRnEeuMl9drIWbqAtlqioJB9fr0VmPoRM+Lz4Ysu19BeCKX42owX4RERA0eSOnf1IGvPc5F1JSUnKRf94/8bkEr1v"
    "95sH97jJVfan1/M7zh/XLLRfrjk6NqYHXo+XNzVHKfS+C7Ldn7/jYl/JqfHxanw/XN/u35/vOWsrIrxcXO+WWW/zLA9cmqCHQxCyQf64sHe3vru4Or6Ejnjj"
    "t4KucNQ7Gx2NbMaNxa3GBRBHEVfFehXj8sD3hvUYXy9JTqAgLxniukSBzkf8eX+KtsJZ4ap3NLoSOxC7GOmNCw5r/WO3tUEwloz+gj0sPxtTUJ23tpKO85PG"
    "kxqbUH0FaT7CYSZiPxQOQn8/ZOpFhp5uKPqj3PLIRWXAXhEmLcfTzec7fDRLdIOgys1qTHRmaqqlLgldcYtadGdqmyW5ltTs2XeOOUj1gegmwfrnfdjesJ8n"
    "Tw7pGXKuF+tTh3fLtqJYohqEFRwt0X91q2vIFDYbDZO45yD1DwTFXhAzIVN9QtDnUfe6LlfrxnDNccyDah1EZaSkln5VCG31z3isGaHhf9Cwnhnv5Xy8rsux"
    "855m2tsVAqk/pCUzP74qoIwavBu+E/ZXcHUlXvKuujU8mB3NbhaMFawWBIhGidIiTg848MZlchjiwpMXfKaB9wnFu+EN1fxsM+lDpng/SeFX7SeNNbq9NnpM"
    "o2nOm2zlyyNEGbMyq8hz5it1wbvw+63qXQ5M1a9ZddW3yqZjScajVmMM78HWtW5ljm2BOPZitDIWrHFh1NL/sQKdwWlBJiNJnkvUQb2rb+ybiYs8WubzTjt0"
    "jk2e9gab4PuMcnWc+4QdeLhzgPYjoK5uUc6NefAMszLPdXSmb2yV/EeApiJh4t0HtpGFUZuw0FZR7wY3wa9wmXSl+ZuhzKUfVSaHRwqE2F5eCzHoRzF3veKy"
    "zy0sgrrtT2VJ8ks9I8AQ4wQaacFJHk8DKTeY64BJQ4sKimpGl/DdyNqnUmCNmTYFpKqpn8n24Baog0oV2lHUcwX5btujBG8A0hTzPdIxjfJQukvm3GMDkIbQ"
    "MtqZTF1Xh22irrVHD4Luy9hwiN9itIanuRTlNqjChrRyFNGGfhSTKTz0jPXIijBDVxhharj1QgTJoWjEkqhGJESDyLjlTATJFCNioAZh/DTysXGQA8iY6mGj"
    "i2fPcRSSUszIfmr5BMzIdfkHcaGokk6lYz+/+sRIoaTAdUICIkSIkkZGzdQK48c3xIxX6VBJMFFCwiR/R2/EFlDF70sVhP4aIIovDSX5UhAmmSWgZlT/EYkX"
    "1npOifmvCb/mFgon5VLaiYRJpqD8n6ZJi6Gzv3xWwMP4n7IH/+mq3394w1SPnLBmuyeab1WJ+f8tWkIJ0g72cJ31tRcMM8w/UqhQLf8H2pQZnqn4X+0vIWim"
    "L3kiYmr7dMGKNXY8ia+z8yX25ZUCzWwetCkLMqnx3wP6zzMXWTlgTpWMgwYzPfWlntHaIDPg9x+yEFhh8f8YQ5QUsheOsGaZjrAO+Efc5syAA9l70D8RM2zZ"
    "qtg5v1gtn9qASvqGW17h79F/mS0VVsuvUbMTubEglVGsc3rT/hHWZSAwJnaPFlH8d4Rl8K/+QjzpazgY5R3HW7iw4lHTP7bXVWnCn+0f0WAUDv/hSjKq4X81"
    "Ypn+mTHumPqbtHvYIAvaeCumZh0FURg/z98TVGxG4qFE+dV756DOjysNxTMlSC4NJfsWMCFGU1ymNjUJS/J3KuJZVZp0gy9fZlEl77+s8yCj1Qnp1hEMokQI"
    "qOyyoEk+t7k9yygN/JUTS/s7WeGzvcrAf01dOPHPISDH3P2VTV8V/2vE5Mu0sCLPl1A+ZfJfoci/Jo3/2QsVNhpu/HGIz+DkeXV7liX47d7yDxMBtleVXU1e"
    "S+oygFYlryjGle/T4TWPduxMUb7rs4xyDL+Gro+dEL99Cxiu+7zH7+55dZUyvP23Z67xunTNzz78jYAqP4lV1CWf2eKW/Pori2XybpvwXe+2iB3BKUPL1sil"
    "b/BmurzbxwiQnudLw6oIkOn7bwFm9Oc9EJ5hdCvscuUjGrqV46yfMMQn0lt2MxT86qi+vkqQUJopH7ZNP/uaUvOr8s08IVBt13V+3ESrhu9GVfjtvvwTpT//"
    "KTMT46Xh/RkQRWoEDwHJc+V6q2RU25+EkGz2QeCQpxe0GhaW6KfBg3uT+xn0pIaQWPAo/P5JApP6st9He8lKv+uCKeZskqKwelFSNQKmGfy2g83lp6KflNiK"
    "KjXgdyiAO93+NHCWPpyKFtiR2/XJkWNQNf6agWe1T07fnsldSF7nme++vWeDsDgg1MnuxTxVj0cXr+X91m5aaGueOu6VyD3CS/LnLHbzEJFUhQTLh82y8ImS"
    "XujcgrWv+CLyWOCHg7Sd0cxYwm+8FTcjbprwyfws8gteh9owzQyL7TdPpaezJ8l+fO+FzEKr+Xzagf+G83NV65hVDsQ/1Ej4l+Nl0pnogg6EyBDN2psaDaUX"
    "m6K2BqEIw1NcJvovD7jkHRfICyoWi/Mnz4hRpjw+Jjwxw2RUzIILUtDCyMJg7Aya0OLKyd3DnCdNh82d7Hj1miLB4+coJpYzcHhrVtt8ly45XOq5Y/Ivzay1"
    "5O/699O6TY7Nivgh+Ant8ZtPIQXWGZ9dklZxtOTwgjmYWY34KxY/2dRPuhKEbTm62S9KHDw8209cNEXC0lYDKA5ln132bC6rdloKYtdnZltUpakoriJYVTe+"
    "1d8jvDshM4CGMYgZEcWjib8F1ilYTnKMhWnFS+iTeOoucClix9O0VXXthbBHf/blNUXGYcvErulh+huHdoBLTtv382WR5WSDBJMhoEbrBGsWPAEKlkIZH/vM"
    "UR6XEUqi3EF3WMz4L0H6pLNy3FcrOKiYesPyR6WwYHoXGS6a9LoXHYF4tvPpn1HiSK9F6NXFxhZDBOdLWZj0ZuF5GM5QqkNJttWLVLZJ0wxaCxnC4sASJz0i"
    "/F2Gf2X23NCOXz3ToSrMGbtMFgvZD54JUJDeUg9FbieOUPkFAfpU86OxidJRGukigdGFpk7g42E0wI0AeYZZss2DDLLjoOeIGmKEmJxSFV6nyrblYHCpvSOm"
    "HN9B4sxyOZDNNdZ2t+x03hnLui5cSZLOjIoEC7AQyq+0rTNQysWqctKSTr2NspPDJhwHOCRIN+W6eISNzzQY6gdG77Ybq8i0kR8xCqfriOQEwMSx1jJvmCYI"
    "j6XKiFgmmUbFoseb8wlwImEIHSv1KFJMog8stDvEhXR6DnnkRBZ/mWuqaOKXkpyb+pMNtBDHWFJB4fTQXlNFI/PH+yctZuqGIHQCh5e6nxOK3p3LSTUbX5cf"
    "wBVSRPot24GcfKuLwYbnl/jmNqk0icAB5lY1EOhAMO7DUBBWTMOhn90DxOBwo1S3sxy2a0YojCe6phQcpscYPhHgNuOtRmOwJuy3py5a8FvZhZKtTrui0W3G"
    "3HN7ZXNWgTOVNRmZvKOzJvj7sGNehE6AY/rOY68AKmuyCONZgwCHzao4+IlYtIaPtL8vCnMHs8lPr17YK3RGEML9thDBH5i+GMzVjLNN3DPp3RFhTWhuECmf"
    "FHxabki43KJv8fhNH79e7KlGIxK8ToG3xKKXjp7rTbX0wTs6/R8/J2ZcV1+ydrlMJ66xU3ferUy0Tm5GoCorHhy6q0Z14/ttu3piLV3fP5zv00NVa7eB9LXX"
    "9fj2VTjXRa7vn1/RjGwCtZQ/yXf/fbnZF7J26w0OACxh/33f4p9y08nM2Nr4X/yfB2Ln6pK22OyIXejHvgSP2cf64k2RI36kkvYZHIxWpMLRC+GrxgPadGJj"
    "0TbCDjRyuWYa9wq5zXjhTUDIuSn3Sy3Ss8X8fwRD9iU0NJysyMFf6+GbBj44rjCn7Gq0fVMZnPgDC3mswZIqDgBxLPs/0kWqtm26XY1p+RV2nEIxZcWpU5Ex"
    "tbbcPjst3bQHSQghH0sccJv91cqsOpXiCWkM3ZNMSYJMiemQK7dLQiV++CH42fRJLDq0ykswSIMk7oKObGJGmYnJbcUGllQcQ5cnUx/oGQ6Vjv/ULC/UgUxb"
    "XrIKbX977tppyZ91oOKWygWZ0AC6xr/t/+5eCjxQq3Sdi5Y4SU5hrUW98+CFokh+Tnt+libbRb/FgKMKJ4sEdyhF3anh6GwpswVI1Rhybt+QaWYqMQVwwApY"
    "UkvGmVQ1q5hPm33jfzxjupBp1alufQb+SCJNy7hwad+cMK4TghvxqaLdWRqLWjVZ7MWOHVxKuW58NCmZt7sGN7qJIvWybVwpbXUDCceUHpY0u5jRoS/5HhLJ"
    "sfbYtwToUpnagUEV6ueZY4lAkl1U36XrRWg5BlKzNVCbRS0zFxHhCBS72jYioPg9cc3VG0TuQpkoKZlOH5+Xm2YBhm+CUW8icPUQcaXb7Ras0+3mcX/8MluQ"
    "E43H993jvqDowGp9yFxp/fl0bfrP+6vh/fAvWx3el76YFOH3Gj6v2/HpnDe/WGmYMB2fl/vmI1ldn4eTqy3emx52MDzUHe90xDCbhB8NwygzkxhDSZmwAy8u"
    "94ZSWbgc4Kt4c9po3D1rihNeLWz2Ua6y/IQ3QLZSeBz1l8MG+aUhd5Mc9j+5M3+xoeuArC90QH+7lJ0bmdaSCm/hjUTw6uDmcNZayzKSkzYVOFnuQkERndAO"
    "VgwijkcOumf+TjZQsOEtgdt+OSNnNyoKvVFSGlW2ROLNZQdsXmeeQbadUiZratdomJHlWTqZJntG+5VvExIy3EFukU8w7sHxyHPcxJ3vBaLB+cBULJPv/1EK"
    "FNZSpbp2EsIQ3SCMm2s+igZ4w8h0h+W5uG909ReHRaCGR3uz2OpPRYzQfr8c4BjtYJpnoTPgizWElI90ab3tdW3JIHxlfwYqIW97wyONbpelB3U13w9tgET8"
    "MCLi6eoQaMmIUWTrG1DCpLiSKZIxey2un6RYuBiv5+UeZttBvCOdfqwz8Ea+1FXLb74m2SCNRRsXXTAO6qqc6fXtEsxJHEfXTUEvwii/z79i03ZzKCbzEsJc"
    "U1UUsRGfD5HkiAzmVWPp90/DOwLR+Na/1PHhahfXY4KArOUtpNyBvwfr8RyOKUmclhBNMtYSyDlHkwGUilNgPaQGbd2RJnjm8b1OWcigsT/Exb2II2RqNLWM"
    "hcnxKnI2nMxOxYS64mqo0qCIa9aq2gKNOJ0yNtrCMak3Z7y8d3c9feaA4DiqZwwO47jxnOFNTyR49dJFJKdV+FeXHPooBoL4JXHVU30M1ZD4L7W0oLhhUY9i"
    "nkFRMBtp6UfclqxBcWGXRpTlyY4nOp10uDkY+hOc78nRNLRk5JKz+tRmvyseAEEclmDEGBk7lvF+S0U2NIwR2z5USn3BH2WjVn3OO5PFfDTUdtQGF12FtYoQ"
    "TISL35Lr4ODuKBCAktMJt0Z0qwLU6SWAsE9sCQfHsED7t8jAlmedaTGjxErDp+gqU7kRQUYkKXMoiB73dgixasOW+7d9cMtzp2MlWj4Y0AQ+f1pyW1vfjdBO"
    "uc+h0Y/DTlUbJg9y0DdzCZumWe2QxRwxpeqqaBDjHR8mbVC4+ihLjenGUg0y/Mkpy59GaYlWw7RACLTLoGGYe1mFbpFjDKFKHGQ5h2oRYm4EyQer9tlhBEVb"
    "N8/49Mc5/mQ7ooFL+M2gXM6LRSiww6NQR2CEPhYf1qUYqnz7azzeiJMQRR7M3thkEC8WpO+1DZztfgUvF0NwQymhAxZj/ebZ0fcIkyXUKeFO9sLLiimnmXdm"
    "w8+67vwP8yNRo+AsPQ6Y9NWKqUe5O6dbf6/PVMFvlUrSqIoHZvxGTC0qwg+IKT4qQREpjDOQxumn3CYitxr9NWpQee9BNjwxdDFNPIgZjhVhbKn0RGu128vH"
    "vjWKlpsp7g1O3ZG55hOnafNzd8fDekeHm8mNo5fILM3lOue6+YdUrPhHt/DTRzO2BYzhVHbLQXQdr2sATidYDSKZlPjb6TahFmFfor6uGP+t5lF6Fbbvv/0N"
    "Z8tmeLnwq+khCwDA/QutHJ3crYwd/+LUOs6aTYHIfafCJ1CdMn45cg/bwOPq1GFww1yTBmWZUpmmMIahJyqdTk9Z52LXveer8kd8NmYt3BAIMyiOqWpF6rjd"
    "YSbz0ZQ+ODr1ElL2OYFQFzvqIrMwqtX0LflIBtzg1iVcGTW0vpEHyvwNrydF9YVcB4b35wZB5O9ob/bmCzEAog4H6pF8w/G4g9mk9RLs9HvQRW9WGX6LCWhJ"
    "rxQ8up+1jOB77NkIrfCs8ZJQNsaP4/ZvMIhLsHgpMKO3G0GYeYwOcg1ktUK9ja9Xptn4qYYQctpluPPIKuV4DYHKgE/PjLCRrHUtunpgvvuSWcye5I8TVHUX"
    "vIPs9QRpSwioXNWAQov6ubGZbmgqwdMRv4/Aa2SbG2gPNP0TTTVee5br5OodkTy9F0l6rPDVx3jELXc+zPCsaIqBCq8L2Z9XCHOyfRk5Wx3yU6gHCdd7R0X4"
    "nR9XOb57ARAnFyXxb88vbrWon1paOMB3BJ2esULuAYbQhMr0A8cx+ABs8Q/mEInTIQEcpqeTBEt1F0QVHkh+Ns5XEjm7379ipMnqz8bOWI3v3V3uO67abLd2"
    "6hzqjQp6F55cqEua+FlqBe/xeT/pp3ue+Lf0eW9qZwneiXjS4LXkRwKLLYePUoCYcIHZ8dA/cfuFimSlpNE5zzCkKjBAWy8OyGkh/LJjxLGEUgR58YkWc60w"
    "NHBDLuKPk0ZKHYWVI/0OuH0IhLYRemhiqdtZu9hwWeAR8ZIN6ewJwwBtkLBcAYAhR7BfN3IssEHvFR3JmiKJCvatoe/XBHAOlZT75KBYcXtWk2bYFbhaGHA9"
    "QcIu25h2RCTDT2dgduMgttiK0kNJkdlghxrW5ucIgV0BZBgRK6zqxPvS50lwQx3CSvYdvpg0VE20lE47NCLh74AgQ75Ufpw6ezSAmaM1i925Oeg0hBnbcHdc"
    "989uQ77wntlocBSjKa9pp2tYRtRickI6P0QjChFHSEK2viZSkl6O7EIveDK91MhA6OKbQB1Bunk7vfFg6sa55IP3iPw5hoxmhUXIT5359JxUHkcz934tk57A"
    "FNc6YW4otCMjHeZXh/wmJ5v87cbnMEEthOcwDbO157BwrGcse7pI3PTdu+DuIFhuREvU6M9igygN5YeBZa2KchPt6Qiz7bjCF+VZFTMVGSHFGJQDuxLXDn1U"
    "Y3+SuMCOhXrcI3jT8G7Z5lHKas+q4rJ4wsRO+67o8bRnRIfLfYYp0T2UzJzivMFaRT3PDQm7oEUnbiA+lbUSMMc6u9+qJHCyY7QdM0oUyGQtCwA4jrQgl5Pc"
    "tKPVlngpKmqWe7uXx2vkc6AB4wlktRYsTXSadLOGlZXkheJxIeuYKd2Uy2V42FuqA8TOB35aTYdidYeO5eGsH7UtBwTBWUi/ZgHQUXULUI58/Lr8RnZ5GWr6"
    "Cgsa0idDAPSDjORs4CJufgSMPNM1SDJw9dwSFy132l1EvelhwmjWPka7E+TZxPiYKhZG/X1gH9x+UvI3jZ5xtPcpNsFvyUYiqGkvgRGzyyMLBvAOFdewjoCa"
    "WngmTlMHNpsdY1isYQUmOtxJaEYUHT4JNC3CAEZF59vbhmXu3yf6gN8ihfxtLr14btlWx/FolOnh13dp8jHM1SnBLD6DGWGb8GcatIW1afhkDfqjsrAXSmz8"
    "e6ylRiDAfdYL5sd+fFxg92ZAWlPUtRrnC2M82C4XvbuMjFx+x5G0/sF/6uYjERxNQigdiylxLmnHlzcLNNTnuJXOWfPqQkYnLlvs+Pl2JfFB+OqrKC4xAsOH"
    "d0AtyfMiLtx6b6pE10342yL3tXTyTmkS7YYT6fmhfgiENRjJl9OGYE72i80Bo0eA/kpww70RlFmUgbG2yB7Ebf5NBarUgFEw5N1eawu6bJwgwAptxaz4LQ7Z"
    "h2oD7ULfuS1tzNRCeE9l3AGHfVladQ/ZtFd3kcMGoIV70CLetizsaGkgQBv7zaUskzVA5IBK9BopoDgRTaKTaAQVSIet2LBWv3aEJs1RX0y16HrUZNSBNbF+"
    "TvdDuTx8VJSwn4pCe2pxZk57iL8xfNNxVy5ic/402WKqG525UN1IQoIzCvS70lQiqk10hNRUF0ujFG1waq1BpsbzlPriqCvLltDhhKD4ICpJf/+hJaDMFcAC"
    "olO2MKm0veOC+TbEsYeTlle3X5v0zjHPrWTRMKJVWsWqWtSp6HNYvoSqW4sZtSRXaFBQCN+/1jCrZ0/AyadLiMcLxlWeDBltMWivozMnp2rPWw4RAMuRZfRS"
    "lMbT4EEbQPthSDP/OzuTTcGc0h53h4d/c6/TNlea2+DPhgjVwOCVgYUldT+RJjaDWsPQ24ZDmtkFUU+s9P68/lB1CaqVxjy66czAQpRoX8vrWMFkAz2Lxixv"
    "Y3JmIoFc/9FuSp/0mDcKhTKMDzQtYWNSWaGW8g8QlfUuuRyJRomFP6DyNsWSFFvGk9P7ZNzdmJqmZvnqCqjcHpZ6Y/CzcnSUzV0kYw3ZOlNzxO0Ig/g3qA+/"
    "5hF3adSXQuQGf98ecioY5L8gEeCr9h4pldk2MelaWXd3Il+zMf24iT6X/CaloOPSISQT2AGyZQm0MGrxBIZhzh9+LaK/v1a7wNM3kJ0uSd62tYkT3r6new6l"
    "0UZAMloaczwaAxma18L95yc7l7zzcdSqdBzBx2Y/t3NvR35tz3Q8wtQIhUOiVUOQO/b9qyRop4LvFtbG2tnmYlSh01rbgnlzu0JNVY6khqhyddTmwYX78LNE"
    "V4IYEgvWcsXxD8i3pJ6aWlhEcCfnzfZWVg0F1dZTYALNBA+snkWuJVV2dhFZc8ol3WbN0LIsU6s4kUfnAFnlBRk+Jz3iRgJ9x76m9g1WteEN8QkOQZjcrKHO"
    "pbkkSWllkoKqUgVR9nYFVVvL+LyyKuxOR+cKTA1o16rdR9PwyvKbYisM9yp8wiWX2Oa4+Cirwhp/tx5zIKRbE9L9jWL3gs9/m840dbX9AQEBAGS+uqj/kc6Y"
    "6TsYGyk6OZjbmP6T1SSqTMpgMyJ2qZ5yRz7OKI4UQ1KhxufOICIS/oRsQGa/ujRaWtX2PMOfOzT22GFsORjkgkcrP0cL4h3M+Zj0ujM1tbv/Vb0NlHSoAFsw"
    "vaEDSqGQZh+HUNC+djIlALrKEFcnAY0j0i0nKloQgw+ugU1QPdNfylB8MlwnA1OZdjkGDnZBU1Sellc0ZrhBYZRGhyitPdjeuZnSYqIlFCtBPcGw+onsh/Ye"
    "Mljj6/dhxjLjxcJQAMXchDdgwhKKew07VaQ3BDGdViv1cpwvENUnauqI3ypQXZe4mqsX1teAqLL/J+9sxRUfi1xSpZR/Qi1RhQbkuVrOch/Iq92dhYfeHMJg"
    "mUq9hoEWwdu0FFolTjCQk5rKFEq4mzUe7RC35UjNeVviK1PQXXfV00ddex+KnuYwVIm4zJuQ650Ihw2qdY6iYuiMFUyQWxAQYIuxuAgoKVW3HnlMH2gyctSv"
    "t4nVH8RvWzoYG+7Os0hGOzQSWVKjMEmhRsWklf9ZC428v802cchIHgRMhiNe6WAMN+YrQ3rHK8UoT/vBbINa42+HidAWg8UxUA7AXls/zQX6toxRiQKrpJ0k"
    "0tSTcpv38Ay6VPWcOM/SdMCFU4Civ07hmKKZYx+pBBzIMaIdm6snxdOjjee9UZtIh9aGy1IOESmi2OhO1bDCU2SMX8y/lYndf/t9YsMgU270jcGl2xui+s+e"
    "0UFw1gD2Mo/8aq4rv3T3hdnwyxF3a7vftGeWzpW1i8uTTt0vT8RV85irVnLq7G9dfx7vjmhC74wcmLD7gyY4pQ2gUszOJlFYucA8NchNr8MWJwm9krRTs17r"
    "3Nc5K+YEqzfH2hOpCt6xxkB1FRTuVL2lqdM7GtTyqHZWavZsvRvHV23oSxHcwrnvcYzoBea8A5SFh5bnAIPP0J+pRQfhlKhR8Wm0cLyWr933km55s4pmqMP7"
    "n4XWmNms+FdernQwbSc6EVenJsCmUqQqB/xLK+3vHJeM5TjMZclXJu0D3UsFezPEQvQTmqWMH/gYH/920cSUk9OkwQMABnn/9YKEkYO+69/18p+Nfz220rZ0"
    "zBD79tGh2fjpt3tHlZ9xWBTOYyz5W8XChKnMyORoD1VkJN+fH2QkRuqg++oE/L6Q3GMF/syFig14oAD7cm/d8cr7rb+f5vdlH03MXGFNAntwdZB4bHTBflPN"
    "5+NZ6d6phs7z6HLMPMn5WYx+9afdTY1vlzeGaeebV6o5B3/803VdjWyEweaZy50e1tT7uzQXAopVVuHKBP01Y/tFNPMnXkj2oLYFhgt7jmke31SJvhrsOK2H"
    "M89naMHkO8Us/UAS2tMJQeTeOElqsXqrRXzRLwdKsT2Y16bxq5/04fR6Oymj+31q1Yc5A8k0nd0EJq3Z9EvUSy7DFnXtHPunPMKvOEcV0/HTjWn4Rua6S41E"
    "XhS7FC+RB7fI8e2ydwiL/KBXV9+NNFguz5IP5LUl3zHJHg0x9NJt9+QKnqO3+pF8xiIsxDWN5HU/LVnyVX8/hDqLqJ0y7IU7V1enLWHKz3cWT5nxS79Rf7LV"
    "eQKZNPX8eLDRmqaocz10+rbjf+ek0AE1p86eJZVPUJmDxEMY6DyYGR7xMV7qw/V8dqsrMkzwju4TTFf3vFXj8bRrrarl/aG6Lsy58j5UQemuCxe6NX7EV9R7"
    "pCU514sKgqhjkjpF6Utz8z20Pwyi6I8/x9yr3SGG2sXhnyOnSPl7dGGq4Cv9+B3zO88W8ou8mgwy40SNtsflJcpk+433mV3BRgD3didMrXTl4MOSbo6p+00O"
    "Q5a5V16NaSqJvdsnfLPypkUnDJ9G9tCsG+qgW+kNYXsn5kCay9GwNy5zcM7rRTrYW4cOM99bR6orA6uj1IE0quhggZuts/W3+80NMvPDrYH3aijNruZAXJ5x"
    "EdZxDil6eNuP7Zt5CTrYOY7JTdI9xAAjzcV7m3XTTLMn741fDB2+fOAHd/OPjHrShgwR4Qzzu2a/l7TtMw7mB/H3zoxoshv6YnYdMKdVM7cftqdZfT1SIhbb"
    "kUT1FWA/+o4QMLbKBNKpgNpLfrK6glWuiACdUQJVgFYi104T6eVp0xtC3g/Q9wmkoyIOJyE5l+92uUUYIhRTnlcWqbt09P3B+NMaWo8YjNwMwKKxpcURHSRt"
    "6nfdfnRrnC6mlEmXtscQzzYXh70d89pqGm5Fqm1YiGEZxbwvOW+AaAEqST0YIGn8jWFyooVd/BQmUmuDFTVM+DRfiDXAVjNRQb12JufIcMrgYu3C03/HBWIC"
    "a+XJ4mBtKpd+DDuFwH5ejQ9Nm+Pouh0DYcP1Hqf5/m7p4mRe5HNF99Ne+HkB9xR4gW2ql4yX5nFVj68ZM6dwPRhdBZt94X3TQdi4jbXdFC/vepTinXHNmctq"
    "oDL6yE41zE0125ohBk70YT1Fdk5WHtkGY9rS9PtMYeBFF1fxOgT30OY5nXEyoi4xX7Tt+zZzTiu4eCs5oBaZBhvQI+GvFNqDDKI3iQggXKh9SuZEVj9SS+RM"
    "2l/9fiWRa82PUCyqKmhHeiqDwi3Jsq3a8rmq/Jobx5YPNLRNIZ+LhtByoNH2HYaOoEOFTjR0R03mZKu6EKqXpB1v/3sIOCAV085esrXy2g+LfR0C380BScMg"
    "l8E/euh7Fzro3hrGBOd1SEBiED/BtUk0/gvP5jY9gJ3/9yD9M2uYIqw2oqijdvv0u7vu9KfWs3qp5U07aDC5cypUmUoifqnlD2sQmb5vNOsLMF09YW8eeZYR"
    "7t82ZPwaoVAgxcR1vKjw9uTM9ji7gaOfXm5oJ2oxhqkA1Zh9ONghCTHLib47to5hVwLaMF6DV0BGvfRW0X6Mjw/2y5g6IOo1T4s2z1CsPytbKoV1ZfUYHLd5"
    "I76D6ErOyRpAqZ4fEf7UrqSepoR4p6qtQEXPcMPs6RCixjvaGUCwKxynud1Dpoe+LX5ogdHi7Pbb3QURBDPfHR41YIP5hTesAYR8xztgCysXhqF8p6wvdQSa"
    "DK+yTn3Rkccf1sxy99RzFSO3LvOrl+BGXkZxXTLRvKhlLZQJs/+R0iYBKQEqj0J+fcARpzk5XcKeenDDM/w6QmFdwt6W+t558ANPftPX8lMaiGqqT+XYWnXX"
    "jq9x8zFb3pbYRiN4S9UjA6TMGqYbe+lifi35CDO6GTNoy64lafM7F0ENzmWnOYqELXXnh6vH25MTdLa0vUlUFVFk7hOSIEYFSPBa7RsnUWT9CFlRtaeeUQJW"
    "trKvMUg1Bxw7gLiqge/3bqTak5X9ENa0SEserdIPw8LdSKqp1G33JLk9thU45BYFEZbq37Yo2FMoxkYLCWG4CqPlFyh84qmzmw8e5988eEFBD4gxIOgRTWha"
    "oQecrabRiFQHMZpTD8kdY5PGC7TwkOyUGEV6VG8oGYfw53ESRd9j9tAFqEhN2oVTE2MoiCFwA5UlHOIxpD5BRgWoqQ1sTWLZkMwZcEJoAsbX2+IK4ZDmVxa3"
    "ZbBZOXh6XMzsBK3UbHMhIatjLVeHvjmJ0w4bZbKPtQoOOJeRusi1CskmSO6vvOuOO4ZVDnmVoziGPRafETDfjbMFleyfVMo7SiCfw0OQSw3hv08ceWU4C0on"
    "IT3dkyQSKwUGBgtqz2sLt0UfVo4rB0YGG6RLl3yAZzqlJvNfXrs4jVsXJOfSziEHn/jBDjj/NBeQtHL3nEHW49XXoaLVR9RklMih21Myh1IfOvllFiYyZ5wG"
    "63kTg55Lh5fHSkICOW9//EcdQpTjGzIyqMAfuXn36XK2Du3GnUZ569QqJzqrXY4i/RZgqCRH+/FxeR7VPOAiNh0l64aZVoL6lllFELm/nFjOKo2w3WhK3l3l"
    "/EmzJRigbEUIfNA2UQTtZ1GuXpg4ABSuQCgHR+K/JDzFxZEj2Tdh8RW3G5dJbDDiH3ne2nON64zxHsOThesuW83UOXzPxvgQNWfrUWkj5ut1Up4T9bhHjBgK"
    "+6SQqzZ1/VUYgy5O0wLlVb+VJWf+qrSYkePbI9Gh1jzsr4aSq/RkA7JokvvJJfGTCqiNvUaZyJXfqsRFmexMnOpY7ACXGadEsOqJt+6Isb1MZ9m83pS9JKaJ"
    "sSSmxVqADEMY/HsZR7W0s1OqsTqkE5UFLH1sVVllVwk96Ptn8mJIiqX4JKfmS+Y5WvL0y81TykhnCthPbYj8utb9D/XyQ5yCSmMeSvk/5+fRfUxANmgiFOOh"
    "zRV/eG9W8yullU6b4X6q++cvyHEPN7rVvWuoY2cXWGx6F5B9l+W92XtZoJf8iBCBdI57rof4WTrZqMyiS1rZsRqIVjSZIs0kI6kElbAviZG3IX2La+5s0FbK"
    "qLpj/aaS4Iz0sdweXgX750c9hXrZexQ8F5Nj34Nc3o1dlQ8iQtrFYcacymGGTtMQseiKeYYKXqs4HIpxIPvqBiDbJNqPJr++ntbYBddBAewWnjK22fQIW5iC"
    "O1TudzINapinMbNJOjVmJ2HItILSBUazjAhFMuKCTICe6zyKOfU8Soyl+TUwGLjUdB6ssEyQ6pKSOtyoQ2mj4rjnMM08urLLrt9K2Ji5oswaJ3VqtyCkS+il"
    "i3uc6GBKyvUKSTXKrXvswVbkQHbRzQgVU9vs5QARHZcUS47qnFvgG0FF5hyzM16di6LwJFtx+Nyemhl9bp31YUspuBIaqpnFd0EoBbGKUkrhKmIhNeW7uP6M"
    "BKqkDwvlpmGQSef9RIzjCQ12/LKmRDixVNomnJXAfbTy65yCvmkbwc27JuxOyHuqH/zWYzzi5t2/QJVi5Pbm3MpP6n8l5Mg5mbTvB6wMrWt74GDekt5zNMoT"
    "lNO9zCxEgJLa8HqqY4RyZkOX4s0baZ3YH3hQ436lM+4GJa34UoZ4XIs+QiF7NN55Gr7IwjLW6OvbJt3PNIivSOFa/rkZwZLRFKPyfYB6YN8O6U2VEzIVl0xL"
    "dPrwpuABq4zSiQG97/Y6So62yP6nk9sVYbGUbUuUBkqV9xyb8/UvpaeUjxcmuWP46BNP6dhubNo67GAPtuEM4eijF4rEZxoc0eGSLPbCGQtAY6AFXKjTBVuV"
    "Hd0iwqs9HUua5lbzBSTGSqXWadQPb1VEUfUoaxRwtOP0KZnL/ldNKmvG6NY5ogJVo3xCQFdzYxBEbr4hY6mzLTUBlu1Oo3U/6kYE4VbiLiwZXS6QBKw+2rLV"
    "w57trWNCZX7tdqn+kIOH9LgW8aOwqSQ90Qv/075j/B0BZHbPtZHqx1kA5dGmUpVvCgRNDNF3MlfLAK+A4v4CimhhVQMaVX+OtROgrKl+UEGPeBPm7pGtwhXr"
    "r0RW7qdPKniEeipqKa3x8gbNvtEAy2jJLA+E3DKMKTthMYMn++Gmp84zWNtDvqaCgthRTFtKqQF/DRfIOIZTLLLrRivofPkexhI8c6okuNF1ki82biq6e6j8"
    "zd4OTzndRyL4b9cus8AOFXY9b4eZZkziAVFo3qh+otTbQSjwK+LDvFN5LVtE4k4srT88Rx1Ht5Mf0ycYxNv04Uqw/NJo/XoWZmvwESMmOD93J+deYNrFOQ23"
    "RxwByRmyE3dSoixu37P9Ygu9bsdfG0aBtn/uaLcv0Z0xWOgaBgf3Tx2Y9XA9hnL8nsGBB/+hMC2XVj0h6/gKGlanm/LsLg5Hz7lHfPOrAqPtXUBdO3g77ur+"
    "9GZE3v3X9DpLZx7Ko19SNMKpBl32bpGTSIBDx9GP7usXYzGsYdQnSlWTlHxM9wSRtwwt3/5Ekg2JsU8N8OzuEpJ3iY59w7sPBKL31XbKluOZ9e9ZohC0RO+C"
    "0DSdqvxzU3Q7Kdd2juLPrFqq57mQeuU83tcUmAFH1PC/0+vv4g3UFwq/JRh5iY0Zq8Ofn3Si/Vwoqut9TwTloQ5q2JCfEDDoQVok7PZTvbwNUcN++WFiqRHe"
    "KnbKfwasb93fCPC2ltxei1x2Xo5LjzauqC3wp0bTghJ9pJsIvxgif0rjuzp7yYHhg3Gm9RnWDOOF0cLkU0cK5mldZG15mzXluBB/Fac5WY3KCbFP1NPGlFsV"
    "k9y0k6hgwoWLXhNXz+gm3YY76GWJIAsCR5cL0GxKzWrG/Wl3c3mn73LD929fwHgQK2LUBAEAyL9KWuz/RznrYm0l9J8V7Vd7QdVado0B8c9388/ecaqboiLn"
    "9AYaljg16kLxWmkQa/xiqLSBInDyaORmUJ+Py/NCemwCNVKMJLHpo+GjuwRPDxOlovgeXRXwiJk0dHW5Iebs+vqE8Ba5ovRTmco5vTiJo4sMrDI8wfITpQGe"
    "Wt0K1kXi39dUh3Nhejlx0GOY461G7BbVo0TbSfC42ZDWxfY6JJTj7CZifSwwuFsn6YgbURizZvo/TjMyUeICp9QXA6bTXDwXCOsUhBXCNfG46HDvPazvnCAH"
    "RvRNB/g1SEw9vnGZa66+9BkUfZS7ZeGcF8I4D44WIauX2ZIISQcdBk1G8lNhhPYgoThXVQuahnP3Yka4UyqpeFrFejCgKdDF1Y0x4VnySByj6rH6KRMoezWO"
    "hhPOCVqul7YPDcanOitI5Dz6obWffoArPvJwt1IFBWuBuOhQcS554b+Iq9sT1kfvK+K24G08OHbqvjWMgWyt4eafaZB4fHteiDbR0ppw4ocA91lDUu+r+Hmp"
    "xz8O1EGrN7JaF18OmWXJX3amuB35juPWHcu3nX6mRNoDlWHZWJnYTUmZaJLqdWzGUVhWRRqkzJDsb+oBPCjUlakL3F4SfCSe54CdEFm+DWpxEDLzizGlngVM"
    "dMY7dpBZWe+jN6Qp7Mzz2gz7lzRP7av1HLYze1wLItN3ZMwY+DSaViAWPBRYgq20yAkFFqyjupFKlKiiBOVGlMihF8oI4nC/WbNM3RWwGK97rAp72fJQ+BkY"
    "kT1uKdK3IaYeEFYez6XX5VOIawwiG9aVHtmWSJ5MQcPJMCMRcqL22MNEBkseuF0zLyxvFeExkurBWSby7U/QQdQoZY9qKRfJJS5Ic5qLvOusFUJgu9HTg1X8"
    "MA2op/KeLQWwhm7TincdzHM4duPtNUKydbPeDfcHwVynZUgg49nmeruNo1h/yv49UXH9q61Cu58KHGbyv+Y3NBjmPAlpUxWtq8CmNxRgXCH1V2BL33bybGZ8"
    "PzZyl5yALGc2eyB33Un0Xof2SPjhHk8u9CKxMJpL4+4/S4A8Ky+EZ7YUQD8IN5w/W9Av3KlxZ7KzbGyEnbnNNPyeXoN+LhEihONrAMSAVoEfJntY+eLO2EvR"
    "u9DTJHzaEMkoAsXP/HT0gLvjdWaCPuELtv63NQn9H3+gJATQsrJ1hWABAH+3/3gua21sZK5Pb26tb2rMSGdhZ2z6ufy5DUCUEBEXAQABAQHWvz6AzzXADwAw"
    "0F/6y4H/EijIXw4GCgoCCg4GDv7PBgEF+bVBgINDwkBCQf+lrxYsDDTs385fI/9SBQYDAQGDhgCHgP7/TZ89ACRIgBvAGASICACMBASCBPQ5AMAHAIDA/nHv"
    "v/4lCwQMAgoGDvHlBsyXQCPil/sgIF9Og315/DXq/TUOAEUCQyZkFABHkdeHILJHZfKLy4MkFqzrRVOYvSZhNnDwh4L+ho6BiUVKRk5BScXCysbOwcn1Q0hY"
    "RFRMXEJRSVlFVU1dw9DI2MTUzNzC0cnZxdXN3SMgMCg4JDQsPD4hMSk5JTUtPb+gsKj4f7X35fFUdd//R5IGY6aMNyFTmWUINwmZQ5Ihl1BkTBS6nDJnLMoc"
    "lXSTJJkqcjPPmTOl65rn7jVe7vQ7+jyfp+fx6fO8Pq/f39/l5WWf4+519lpnrfdaa99z9ka9zHv1rriktKz8/YePNbV19Q2NTc0t3T29fV/7BwaHsGPjE5NT"
    "0zOzc/jlldW19Q3C5ta2XDQALc2/6bdysUJy7dq+B/TbctHsurX9AdbddIKyew5qmtI7XGc7Ind3L/vpB0+LqvcJyZvhOC77dO3nFFbAiuC3Rfsp2f8mWMj/"
    "l2R/CvZLriGAgZYGunm0rAAcIN8kv0ePPoPTUwH7VvU9VGAMFblWoInUfMlr4LJq4/ugK8yTmHpS+tBp5eZmPsv7tVSAbpickntqEYcaMwygcOLyJ4LP9D+g"
    "AtU1gzHmBHmf7ABkEEAFmsz6iwjSddaVdgLlYKlksWer0X4VYiPxc/+kgAyhOcUfHt4qYiPln11AvDL0Ctu68O2gvlrByZWTFQVl0kJNK8YibUXnX449Eus6"
    "v/Ap/dHmw4z+NBvPWnoSl+W6LXzFUEMIxEqKKIPVvmjn8R4sWOdFEskmudn7oy5SxAWpwGdlNCH9FkmEp5G0GkoFgDFINDe8NCGTFEkJ84GGi8Exo6hAlEc2"
    "2NkAJ4iDyzrSJLjWOoIKPEPgeanAvltUgMcUEvcLOTWB1I2JwhAM0ISvHSThoY4N7B8XMB0LcqzQogJ3roN4RThelZuevDeXCoSiR9moAAtGgwo8jQVxZlRg"
    "ItYBYtVKTu3YZoUmGJYTuqgAczM57TbyGCGOCoS1gniLEqTBa8d7h/9grjkW58TH/W9ZCgmPmWAQP3Y07hZpFzjGO41ZlU2g7EOQ0uLhox/gxbJUIN3q58hf"
    "QLLmgVgCFSjW8CdY4rmgc8VUoP5CJ16jMF743+pwga6ac5EKZN0HR/kQJCMYXmXzj/Fnb/4cfxR8xhlGOojelOeiwOKIClSgS5xwD+o2TQWcz4A424TBOyPD"
    "1rZIFfzkh48fzgreLmB0NBfKpEn2lnSjHIxzpARq/qGccTTO9GSaNNr5AL/+nDJykhCei0q9/4q/ylcXJsa/gvDCYrpvHB1fkv88IdP8GP7aIteo+WvwXN4F"
    "9mG7+QrbDYOrbpc8/ZF61STN+07KmY3E03LrCtMKn6L7aqQueGWfWXRbhZ8pCTJm9Dpb+tXDIlmywn8qzhZZbaXFQBztsoJ0jZ443UkFFpg9oDucHE1JSE50"
    "mYyZir46TU/iUz6Nrm6Er5oej6Ropn8yqdGliEcxE+XAGSM7SypAzyp7nqAKzuVYJ5BC1QNTQgdgBz7DB5OowOlrFNktF7dwbAcxjQp0z8IcsK+oQLh78yZk"
    "Od1XguWHKb3ZJY3JuWgXtEAQdMfl4zaoAH8+K6UXVpIDttcRVqjAgQM/KCft3NxOXDbNxsJaXgs/ri5cG1Jjq/YAm8qZqUCIKxVYFsVdch+POoIhXkKQuOEr"
    "6hLSQ+D6AzWlRcbxQtIxaZCwaARbVCSeZUKXiMajcW8+UYHx0il8295aiyiURIAKB2IOTkcFOov/ZLh75TtkaU8+ogly4FSganAPvMQRbK8mjFOB/fMYyklz"
    "1OJxorYSYoi2EUawtKMC+KGV1ImlF6hoJK4XfEnigMRvg1P2QV7UlJnU8T6SEipBBUabqUDaTBGlG1Gigq3SIAxD3Fq3uYkuwoinlejHMXFo3OvK7dFNP2if"
    "u+Y/1aiSQ9YdRW9xQh6CM4JTzkxmBdXBV7khxe8xIXV0ShNEuOq8ohAsXI3ZuCz/Det+jWZyyjpiMc7arVD01tCG+4pV2fS16UNfurJCPmVQ9vOcJSd4nIWv"
    "mJagybwcSEjhuz1TfJA6whvd8PX1qhBIAfdYqQCDnyc498wV3FKFhi3q30oX0hqc1VPrIK7zWHnMZ16H1TS6UsZoSFz4zbBn64VhePl7nSohjtXr66VT06Ja"
    "HLeODgev+eXf8jJi7uyIxCwLuoDr9GWQhxhow5jBCc2/2tr77Lovr8Bau8gpDInPbNiDwlROPh9jQmZTD0SFvkNK0xPfbZuaOWRqh/iskfyilGgjFOVe1ieT"
    "6rAEhuHxn/9GNCAMR7xJKlTgBmzrktCimmICSeKKdEvKpBAu8OuNyTcmuOhSdnLoBaNKTAViQ9s6Oni0nyhq/CIS5+eusQXTotSA7+xiiHb+Hba4quJxmwIO"
    "38fHpb662SimNhnePXG/TbZfcXUZrOZU2wPdn5qNr3CcAQQdfIXhVXspXZhD6KtxdiuXCCfHhZ6NzHG2egXIvC1/vyLE+F5KNO7zhB/As1oHqfZx8LdjGl7x"
    "XhBKLEE+0dRHyhjr30U+Jr+kVBA7anek1n55+pCYWR6+QV/cCL2u/1Hf0g0TqJ7cOGcr8tgfnjyhMDO96Ww98HCyXXmhJ+k6oiGipRWLUO94yUXZ+2n0vAqM"
    "fw0kHaynAo0S2tD15PyFXhLmIa8bTZamqCBkUCT2/nVMB5wvA0K9ESUq8OAVhHGjLxBlw34ECPYIxrNoUhwinQrUSWwhwS6HEBK4+f32u7GOnMxbzZJUALEh"
    "F6SajdPJ3o0yiU8vp+y7DBnroyNQyOAlcX7SUJnLJogkUIF4y7btqGcAetwP9opbBrFwCHaX1Y2zcWYJw1YfIBTAiFLYtqAoItQNxRnz1S3IFhrpIFBeqZxK"
    "4Qri3Zrwg1CJ4qpBIT+AOz7mWjdCsDAKNSpZ95OEYU4rGj1X3m5q6D6O9XDM3GycLXsNXUx7H2Jg7htsS+yjJSU+urKYsu8uUY/kSn7qiK9F2uMr3Oh1Vb3W"
    "by5qfGcWbSdaplFeKHxBlsxneoxzq6wxE8Qq2Ij8NlRgULsbXEVakO49mkQGPSDLyhAEFBu0uNJq2e6u9+WR2CpqWjjJI9IRGWGIpfHXdkT0iGM09ltttWrf"
    "Ms/BnLBHDLrCInid7336lNWN9CS7la7VweB7SCjYPaCHYuFb9XJKtLUYIfDGKD8DFegVgCBA+zgHeRJ+L4sRji3UWrGhMCkpF20jai4EzDNp830kHnDsUXnj"
    "hhOEUEOVKRQm+OWPV6u74TN/fuDgBOECRiMRnPFANICG6I1gPirwIYEslbYhwCtNONcGI9+GENcpmM7GmOHavsULB+1krXDoJrR+ukpU4TqPHDy66BDi9nlX"
    "7RN7JOrMif1vKLWIotrgwe/S/Wpi6IOREVSgyLeukqjZRRmc9rx/dKTK1yZwWgWuv+UAH42Qi9ddNEMTtGAU48J6Qv+s+Cg6gqSllhBVjHCIDonNZXaMevxw"
    "LJer3fjh2FJ6oF7p17atY6ia7wpjhlXKC+gDHfpGuVKB1/tOqosobCxrsEeAgVkq9NeITyHePYc/k/NSqMDeG1Sgue8YiNMabDQZegDiouCkDIRMVTpcp+rD"
    "HOSQ7lBuIp/ei5Oua/a4QwVaSsBNB1ybOat5vSxfXuEHehLr/utUQEkTRmEVpQJJrvMQ8Ol9MM4u2QKxB8HNE/kUUYMZKAZp/SCVNYArKtXmnxHsH52pwOYF"
    "yJ6hGFLN/hzDXPcM/imHCiyuBUEAT3fIPbjxHAgZDBwymHvEg8H1ZpRGvTEn8F7+GJxwNu7Bpi+iiJzdBIWx2ZKKjYDvA8MYNTfIVHWyKTJzAvCpHCjL2DLl"
    "ALMPLSlgs1mqMh4SQiPniUgRpDFtIrI4UulIjXlKhZlMeglW5LXFop9vD9gTYKeZ7t1xV1HuW0N2cYFxhGi6UmxB27yeG+e2cpWGb6w6YqdJUTww8t1FcNWs"
    "S7cWvSm69e8jxLA5IW2ih8KzCpeHHH/iyuBzEj98zfwYetVwbY5CJ0phvY0Ee0zIjPnN6NGJbkvtbUtEQSAxkzf/VW1fEtbNE5wupMReoyhsOYH0kVzkMHDW"
    "t6MSNwKWOWLIt0lpBNrdTxrQ+mmI2KmXdQRtIWKpms+GIqWq4QI8yQ3rKBVZSGlBlDHXdQxAKJNeGK3BPfORf62h6HJZit6boXYJCclIujPmOlpOt4QdS28/"
    "A6uhOHGtYXZUOhJNEKuCssUhQr3wYvoDWDqp4x6JA5cdV3xv6sOSqS7rvfRoP1HOiV4zUxOLd3pCFXT1SKsHo1+5aGMeq15b1DludWlkda6jOIVywJlokdRx"
    "NxgGjj6DImCSxDYIlmJK9eZHvSIhkFuIwVAMC2NzsvnYzHVlFtyqE7qyP0tVnoT84DNiM7UPL1S3dzG1l9i/CbuCInGi1l96VtETQAhwtzvmJVEOaBHN0HBc"
    "mD9UOtBxoUlpCB7dEyr5rCazl1i35IJgk2gmLBoHxY9mAT8TNuQZmysWwWsU3XkYiS1jvZXwlHhJe5xbQ3gmnTGbJLSIipFeOKn3DDNNRgVTgROu8gki12Md"
    "RVXG+QtJ7L4PiHxQ4eHMqwu5eFT+FJogWLuZjGHryHFBt2xVQXlYSEIRrKgXvm5agiCLR6PbF1YilMFQjcyHy8gkRj4NzmRRmmbj/V0JCt5Jj5IrWsySVtwp"
    "ac0rxxpVcRL+bUGWZq9nxgJyj488vpi/JZzSROYnQLgtDCU423BW+BPOdpyQJOqAbGrWP4M6gsT3ot6XORRFeeUXSYnO+uSIJV6tSUaR88BVR8tEyBtaGctr"
    "0fQm5PNkOASGJnbFOGL2Pht7kA+zpThPaJxNaV/wv50CYob7s+hnBHhF1l5DseHDpvxyFpwgJLr46hWtaxb3LOS6u6A0TGENElgheCi/3wjXQWE/ANlYfw+C"
    "CRx9k6VG+HT1BogrXN+HL78nWPzAsTLjdK9wxUXuuxJvLpd4PL9zlr770yHKdzulTvRVE97gTjiU5zsVcvjrUlj3vTbjZ0a8vx2gqP3+ffxBfGXy/J2IUyYt"
    "g7MNbmuG/Q3oFxd07Q4hPS3vxun1ovS0smIy8qyvhysv63JOfMgVyy91ylZIj/ZAfhJgxlVw06v6zSu/bDEacvWpVxK9ZTwkvIV51kuIA0fjssuezL/wvzpM"
    "kYAS0rGfMOVNLvn4WLoWQTgnTQXYEPpDYDViLVVN0QN/oB5exJwAVR77bSU+EhJVBAyuxGZ66gl8V87Ogn8FTLgL9Qn9kAXyE2xwPetXcD8W9UbhTN/8i7Xx"
    "nJUfbU/k+R+bKntYqxb/bkl3aMZewvLQydBdlru/nGz48SLZKtXB2++du7TG0Zwco/rT3n5mPus+eQznn+4bOr78xdRD1CNWHIcII6QrSrw2Mjittlzhujzz"
    "4eElhaG+a0nfCLFgNR/ag2nLMgKNC0aTjPRgJNb+9W4ojTuMZqQC2FPQuSCL7cLvKPmposFeP6OTlH3ZeNWx72Mcyd/H9i5dEom7I9FS07ibRXO3H8d30mly"
    "xVOCH1GLpELOUgTDUrI9CXefz27E+Z14fsRD9s2LRCGGB4fetXPHTj/6cv1udCP3+EYPG230RavJ99znvfPNk9lyTF+aJqhf2JT1KT/mV1yc3GU0c0O3FrMP"
    "zz5i4mGlg48byvTLKDGbel54a9izTIMfQRDqX0yvhHQPEjQQlKO5CIIw82IyCj5oBwXYaihnwSuuX4E8jO5gj4YwpUeO8gUGaBz85t/s/nz4lf8zfmePBQ8L"
    "46MpGuZL95jYlYXavB7ROht4BQ9iypbCiYIG2MdiJPVKvE7BWExazrtud2lkGoe5/QiloJUzgMPiVLNKo6HlY9hzVPeWrfbdJk7LsIvfnOEe+jGx50c0LDVc"
    "68fMn3+4PHmtMr555qp5DRQl0cXz41lHCotypPTalVrb/Xi1iXDwQsqH9nzlbMJR0YU4fMq6eWhDF8mmrCdHdcUVZULpQLNNN3PZbNgzgAWGgSuzdlC9TXfY"
    "B3R8Dy5AELJlzmFFjoe7IRiDB+AMnlWCLnZ8SbrpjRcSG4gG6ThjvMwZv6BqS4mWcT4mi0jlVwicmW7jwnUmZlJvEFRN6T8rrH0Pqo1nHd665HC90vG0grxo"
    "TE1DyeGCQy2ZL83jCsRfZApx+YmL3kZdSSird94UuIQ9jtQwUVKTe2pVYzRSKvbFeEuqYKbb7UpwIvIkFUgEn4IDAv3bWRP7jmMOSojJ2NWan8ENqvdHJ1cS"
    "GLaD23nIQchc6oGNtffksbzuVKBDjwqsuvUngtVtceW1MAh6LI5DufCp9IpCWqQuTiMBO7A6Yl4/FFRap8R14Jj8JYl3MJWtEP/CccQLhWv6CczvtZvrLW/q"
    "UwFP0RkNPp9G1riK0k340WB4cKsAnCCLZWS2CB1LaASjs1jS3VLTu5kj40MkH+336fF8pGHvTW+3r/CFIpybdAwPktgsx8Mba37EwlnVLjFwpe8f6cAzdja7"
    "N5qWD71dfEAxoxnpuS/5YxxviTVjKI9TAEn74tMx3FWsOi/kxx+L0W48urIb1vKBUD7rqBstr68S68ODe6ymw4OW/irewjy35to72UPPE/PYvFypBFnwvdvg"
    "USW6FrNAUnipZsvjnXdgWYFMaFGiSFrbrU/2uxV0hGhwgNVGlH3+uuuCPeCVOMnPAkfx9DE2NZY8Yj3jPZdP+LTLNVWWp17qYe7U0WuQUY2nHQc/Y6LgNPDR"
    "h3ZcXcGirrd4PHPfkJx4amZvXbvgpuzUPX3voDb8elRom0SAgP1u2vtwIyECFhdC+1FawPVVIP87x61zUCEVzNwkK5qg8PXHil72rJ0AYvXhZ9ySb1EAM2Nv"
    "jvEbRcciCf97Fo+v3TtfIijtcKgr2o+3m82z2a5hIJ4uL75G9WKbr9O4wUDPtf0Ssshyjxvloi65eJGJYDv/q2MbCyg3bLvGsCm5BHRxs4lauGwQY8PN1hRe"
    "AzTDfjQ2MZ3GKe+qgBPEIsPWEqLRpahqk10kRbxF3a2sIwS3Z9YuWw153l3txPn8XL0XTq17JIYzuFVEuQs7sfj80cHZcQ/pFRjWjDkO3Gv1CI1V7d6QHtGw"
    "H18t1CRLSGD9n0UE3G2WJOqfq9PWKbh2I0HRX6xBT4teb4jxUoOx7bxB461WaabZr61XQwtiIo/lGn2XSbwZWmHdl1zo1FdQVQFjpDQLHO+F7DmVCpQyFRAl"
    "Sf42rFjMISOGIek4zs7T75TEGd5MC58OiOtKjXJqpne96Y5OsDtO2CRqQbdSulKTKI9U66l0CMAlK2OL3va9yYn/olSO1BNKyWKx8pwKeurEgMxVFjwLC4I3"
    "IMKSAGH81+SB2Y/218208DFYucQuwrT+Bbh/TDh3YQZqNXYdaRE1tpWp+urKQCxvzdfMOZOB8xkTsfMfZVD7dMxzuYzPhbsUBCRZo7rM7ziaS6Ie6fKNR6Cs"
    "LEJMbfH6i1qPm188xMh5Ms5eDeqvBXHGl8CxWwubRG//inWQgB0lIg7MCagVeRjzM7VghctCZg98q3v7MFl7mZTy/pxSES+HnjBfkN520DgJOjadpwJdZzsh"
    "Z34jsZoP3sMw/BC2OoKPf/iDdHz9eXsAM4k9tJ5HKocorMv0OK9WFZVQomI1dl+GdvFtfK83FbjHij85K1L7TctG+qLqSl5IrpWO2UMJC9SFTPXn51MU2ELC"
    "zotIxPmMfLHs4fNiogL5dvxTKkjJj2/PFYpdEh7n1jF5NKzTm3FLXxIRDkNlQ5+YyJ5BrNOX/muig+lLMGxGl/JkHb0l7lK4uh9kg8NRJMUZJBoaP2TyIW7S"
    "JHZETTos7A3Jxu95bc41mHu8jbzAFChxtixblo9jWPIm/eEbCG4oh5HLEhEVXc/HYyjsCHI4+wWMVd7CuLGeIu9wuuyjwk9vHr+5ekxdRa/Vpqq9sDmSpCU5"
    "Bm6povlBLFRzDpyDjw6CbAmbH7dPaEInTLVNyigTcJwtfPNgcKFJBxXYlbIZDxXiV5BGVEDXdDybJAYS99OTeBXHoAit9PMEnLhfLAcqV31gBCkYKSroQkI2"
    "hU56hYsimIGHk/lN4kHcUyqwIA6OcWOgxKmr5+dkGTOlJlqcCC5B46CD0tQHonO+SCqAEwdTcyiMT/CnqEAGCt9IYYeRw8WoQKgL6Rx8alYNSvnDaaD8R+xZ"
    "wmv4RCTlAJSVH4GZ1W5CbmBGBWRrKYy5+FOzRNOfU2i0kIL68hGEE+P1zSKfPjUuqow1Lzp+jLcSrOrj9OO9yNIiIEfLbGuq7WrAbR1+tvdFDrf2iOLeo+mD"
    "L3J8SV2Zy9/4z75vP5EVp6LfHjQ9DwsvuYnbiPAl00zL8fKKFZAa5oZZpEB9sWcflE6WdVIH//5l8qSPC4GPBgBe7gYAof9YPO5fC6f+WkLu57KpdQ9NTWoP"
    "szRijEhME9/DyzJ6Urs+HTt6W14VDnS7GbjfBYw3FkWuy893+0kbbTWMVFODSQlvbckljINPV1pcmMSN7jxJfnHu4xiuKqjq9u6x0M5wl1gdpturb1u8vJIL"
    "B6JYmDetrUyeRa/zzbEssDnafR3YbxKnzS+Ze71gr6wvR/Hp4Oj9VsnCB6aMFQ//cEk0qq72iOa+v/i94Ut/qZGVxD03IbR171MU3XBp+9C9H/4fXkrYht+l"
    "zX1NaN1MCtQft7MnGzI7yjWPpnNddY+cSV1deJ/eMzMh8sTd8EkoL5mJP3qWxUj1aJyfPr2v16HTfBk/zj3uG7DK7GbOlbYqS7WPKBbsXxIFWiLuTLEBZLTk"
    "103EMuK339ST57gi0FBLhOZP5f75Tf2/VPvXx89/KjcsoTV6lyyL9kYIleWa16vhcIZuVyW50ZudzGKJc74lK4ebGbY29nyLlZFp/Zhym0mRbymYFZ4ZE6Op"
    "KX2Shc9XQeoiXcrK9BTtmTx2huOyb0PouZRj5/tw9Qf6TkgUG2gr6Ivtn7LuESiaSVPn4GaL/Gooe/vVYhsn2yyvVYNH/6IM9cpnQgO7aM7hbEPmMdLZdhL8"
    "pHFSDzJ2a/jGS1Csc3PK81RLcv4u8a7i66sttwOM3GkaaTBZvNm/lZ/nLVM4Fmp9gw5F/yW/t4+rp6+zj7mzr+9PNew4lpW67Orpc7XJvumq2BUxh732Ypeb"
    "nJpo352liT3W5n7r0CHlEKqSjJyst1zb2AnsWQd7p9GrUfaXd3WzGrad6n5QHSZsjpe8EqYf6nL0cAS+y49ud0CEusyTZ/XncoQ/l9xR3Ssnc0emTRORen1w"
    "T9KoA6v9XhmHK50z9tXzby3Hgib5vDnaToEU+tOZDeaYpJyjTNEsnx+VXVdleik4oFR3QFj6M86hrv4Is4xN+Hs6YkiNkYWF4EBfV6OIWNvA/rq9aworG4zx"
    "MvJdN/B407N6RgbH2sWxPEYzOYIyPLtqZXcDv9WS/YQq0Am11oDtBYt/asnR18f9nI+X940/Wz8tRDi+3ZhWliHsx9lKrbeBzsWknleuIqcylQ9bCxxuVx15"
    "UmcncqEv6Hbs+7amMlL76/LpxPZ9bgaGTJ4fBRjMRgW7uqo3KOQtbkM2k9AOwxn7hj3D7ZP1qj7HCw3NVMALOezXmHQeuDp7HBDixHxMPD3ZYPrxc71PrPgT"
    "D3NteAmrqX5ZeRC6xMAGeUQGSag+TAzoflLewpaquSbEKkHEgkPNmWuvP9POT73OhVXXJZJTUOOYHK7n+fWtY565y8/XZE++p7+xPKc+0P77hUU4n73Ob4Ra"
    "M/+kh58wxB5vakAryxKGoaNq2W7NTb46UrZqfQZvVUK54+HZEyNcoXvi4Xf0raC0u1jHKzcj3ZqXTlsk4YD5N/S6fE6GSIHRacJ16i5j3uUoq9Q9F89i36I/"
    "DWrnTRq9t2jO7rM6MA+Yu618vc3IxtH75vk5bZEyqbP4g477J/zehOUYnylwonZu8KLxdGyhST3S4Tdtr/AGFsqF8hqhhxOPv41OuuhTmtExermZEI6k6Bax"
    "YgrsLBNsktaS1F+4s78/kL4xugH+VnwzUe3mbTPY+Cfx5X+Kn1DtWSPDEPEjaZ0prd7PqFVYgSWctsf45OEGgcPPmdofDlhcFtxaT231GkSTB78v/VBJ403m"
    "SQucSedkSJwW7C6NRK9sEEVTJeL3pKS+jeSWXLhVKD38WQSW+txblTOH/ZV67JGehqE53bY3J1S2ah6FKmQVT7I807VBhXi+FHo24ZTiUVHmlnFJWAa5qSs9"
    "HdCtjQgXlQmdMxE+fxOFUu9BpCVt6Ve+vbA09/pzWOa35q41ldndsx7XLyLdz3/M1bijV/H5W2YD58LSb9/WGRZ4bL5tBtP/pAeFP8zAbde2GRygRPriS1N6"
    "ks3qZ7RyO+qHYG+nTtjpv7strFkJYtac90Q4TA3zL0FqeMQTzv8a8oaiFMHlvuofm/OTtIYeTbRHHXxkzMNrg/ElB8yWUd6SiVUxXRxvmBq+vigo7Ut4WHRT"
    "ZeuIxen5SidJwTMmZcbh3i+FTFub3xCDHZVOODrJ+GwKYb6sPNdtPLLtDah2PdrH76OO3qpH0rP51LAtr5nW65Aq1aQQ5RJe+w2bq3naGPtuvyX+3gzGnhac"
    "eAm1lqFflj/Ed3B31HJxcPXcFtrpy/YCtey7Oz+NwsMqyvOYo8/2SO2XQi6s0p4O9NjVbKQXMoPJCOd82qpCW9u72Gq1XscSJ0oz/2x3xHOklwuLN5rf8ra9"
    "rN5c7lrYru+Hnnokx1jr5aseZyyqPGV6x9hVzVvgkY6oI1b3+Tc1pLfc5KE9Ao133iXwyKq0KjlbXU+hf5LHcWDXx0WmVNZzNMWb+3GF2UepzX5HU9IyqeEm"
    "XhnMv5WqK0CRsw1qSUK3nBv66+TlaOTs6+Dk4Osgbehw2dldz/OK18+XFzsMrtXKsNzB5kcNDr/ra7gZO/madlZw9VDfLvUQRUbx3FuFlW4pI4y0K/gOn4CG"
    "zwLvubKTwqoW7zjIDW6tSmFDbeVpJJ8nWgiKPsTdo5OavxzCdOMcT20712UOPWeN5IMnnhHbAmR5dSyGZtqzDG72pvMF3PzCO/GMjT+BEnwyTEnYmUOc0C7n"
    "JYI0QpnGBfLwOEQ+DeDPdPmyi6iTqXBTJ5bgk1WVcDmX9muoNexBmW1j+uK+Quq5m7Gb18XUZaqkSD8S5n+vAiI/vNcdEv/JLgBgBfbSQCr4w6a9fJz/2KXq"
    "/5ajhwiZiHALlWGvx9ATmRJvGfUO+Wkef51jIsk+NTBgc9ZzWCYu0amYvB5TproRqWmXeGMp4YaUXkrjq1WNU7xW+QzhNJoCYnYH8ta5AyW+UnYFNMiJXLSS"
    "bOrcx2A91W8p7iVWqOjuN3Kn7GDs9NmWc9+dZuw9tGRPkAAO+bs3xESXrl2SnX8TmMDbgBkQ8D7TxGdpn2afjGWaURJivXQ/JKFp6lKK+ilvBpHzT15XO7CM"
    "SRfSXUqD0wc6j71+rq74pkdk6eFxghXfiK6QruNwmSbvhqwthQXtLiBAfFV1fuSN8hM1jkwfxravqree7ck7hfJfHvJ469Tg+e3drYN1I+IYmXK9yYLUE5wP"
    "aR6NYV44eObnvju8HHWmyOQ0PJ3DyZjt9NWueWWlVNvMKrIK296OtafT66HLfIVLMYo+aj7jbJGG0VGdEw38jURZzNpJ2nwDEniv65toOZP1fTqjI+l1bGNA"
    "E7ETQz1P2fc7E62cR3KlQPanSruNPX8xUQdv7/+z0F+UmVQNmSEUgGnXmSKW4r65i0WcumBawh848Nb++u2rTqqsEgYRTTzNWVe93nFlsOfc6+Hfs6qzjAKD"
    "jjj7fZq3lbUX5kWKSF3NFevM6LyIzqtzJK+NxYhFVSuNhjTqLzyoK6UQNI87zsi+05ZIDEfpu6cRwx86lQ3QtCcuJHVHDNCgFDydXGzccFKpKUzqrQFWdnrf"
    "ZLf2ZMRQXMR0Gus//1D0CGlGrp7/Qdz92oL/kK3Hyd7cRNSZkdVPxXuUR+zA8qN79h/n2jgZatyc9/Lg2v52g5LKDoGe8wUMypVG9HOcfE8TUSGBp5pqbUPi"
    "nvd9jqFhjLx4n/yoQYWrnmev/BWRPVGcaqLfvp04GjmwVmeKqVO0flFKXMngzknMczvPhF0NdXWHz28+FaSbjhu0f65Xo3LVY1CQ7JLCsgkO0EewBtZ9CC3p"
    "Fv/OuipqXTN9ymtV/r6SLlhMSRMgqjr1929mcfVdO8CaUZI4TkqgW7uvw3vpxIeREIz9ytBK4X1EbQ9BcqY132ozzzAmq/ehmnogmevfxkyz6zjw33YG3Em/"
    "2SdwJ4O/bxP2V1KAcvm/bBq2s+POzZB+UST9f2yNtLPz3zcw+SuxMPzs/F+2M9nJZ+f2Fb/Imum/b2axk8vO1eV/EQ3bf19rfieXnYsG/iJvwd8uIbiTwc51"
    "nH6RgNiOVZ12dt25ZsIvWpT5/QoKOznsfIH8Fx1R+O+vk+/ksvO5/V+EP/2PT/H/i9F/f9j4F4Xr/P7R451D2Tnr84s0HP6nOaCdDHfOdPyijMv/07zHToY7"
    "pw7+wtDxf55I2Ml0Z6X9i6qd/qnu3slnZ6X6i5qd/6lu3clnZ8n3i5qu/FMBuJPPzpLpF3Ve/acCaiefnbXHL+pw+c9KZGfvnTn+L8p1/YeMfyebv+fJf6WX"
    "136y+XvWvLP733OYv5KBx1+6/5HRnDOg27P9Py7oRx5qHfLePvp/VDh/oQ=="
)

def _load_embedded_pdf_documents() -> dict:
    global _PDF_DOCUMENTS_CACHE
    if _PDF_DOCUMENTS_CACHE is None:
        value = _read_zlib_json(_PDF_ASSET_PATH, "Eingebettete PDF-Daten")
        if not isinstance(value, dict):
            raise RuntimeError("Eingebettete PDF-Daten haben ein ungültiges Format.")

        # Zusätzliches Formblatt direkt in dieser Python-Datei. Dadurch muss für
        # den neuen Tab keine bestehende nfc_assets-Datei ersetzt werden.
        value = dict(value)
        value["reinigung"] = {
            "z": [_REINIGUNGSNACHWEIS_PDF_B64],
            "filename": "Reinigungsnachweis LKW.pdf",
            "title": "Reinigungs- und Fahrernachweis",
        }
        _PDF_DOCUMENTS_CACHE = value
    return _PDF_DOCUMENTS_CACHE


def _load_asset_meta() -> dict:
    global _ASSET_META_CACHE
    if _ASSET_META_CACHE is None:
        if not _ASSET_META_PATH.is_file():
            raise FileNotFoundError(
                f"Asset-Metadaten fehlen: {_ASSET_META_PATH}. Bitte den Ordner "
                "'nfc_assets' vollständig hochladen."
            )
        value = json.loads(_ASSET_META_PATH.read_text(encoding="utf-8"))
        if not isinstance(value, dict):
            raise RuntimeError("Asset-Metadaten haben ein ungültiges Format.")
        _ASSET_META_CACHE = value
    return _ASSET_META_CACHE


def _asset_installation_error() -> str:
    missing = [
        str(path.relative_to(Path(__file__).resolve().parent))
        for path in (_STATIC_ASSET_PATH, _PDF_ASSET_PATH, _ASSET_META_PATH)
        if not path.is_file()
    ]
    if not missing:
        return ""
    return "Fehlende Programmdateien: " + ", ".join(missing)


# =============================================================================
# EINGEBETTETE HTML-TEMPLATES  (Base64 -> werden beim Start dekodiert)
# =============================================================================

# _SUCHE_B64 wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.
# _DRUCK_B64 wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# Hinweis: SUCHE_HTML_TEMPLATE wird unten in get_suche_template() einmalig
# aufgebaut und gecached. Alle _patch_*-Funktionen bleiben Pure Functions.



def _patch_suche_template_layout(template: str) -> str:
    """Macht die Ergebnis-Tabelle robuster bei langen Notizen und vielen Touren."""
    replacements = [
        (
            '.cell{display:flex; flex-direction:column; gap:6px; min-height:84px; width:100%; justify-content:flex-start}',
            '.cell{display:flex; flex-direction:column; gap:8px; min-height:96px; width:100%; min-width:0; justify-content:flex-start}'
        ),
        (
            '.cell-stack{display:flex; flex-direction:column; gap:6px; min-height:100%}',
            '.cell-stack{display:flex; flex-direction:column; gap:8px; min-height:100%; min-width:0}'
        ),
        (
            '.cell-slot{display:flex; flex-wrap:wrap; gap:6px; align-items:flex-start; align-content:flex-start; min-height:26px}',
            '.cell-slot{display:flex; flex-wrap:wrap; gap:6px; align-items:flex-start; align-content:flex-start; min-height:26px; width:100%; min-width:0}'
        ),
        (
            '.cell-slot.note-slot{min-height:24px}',
            '.cell-slot.note-slot{min-height:32px}'
        ),
        (
            '.cell-notiz{\n  display:flex; flex-wrap:wrap; gap:6px; margin-top:0;\n}',
            '.cell-notiz{\n  display:flex; flex-wrap:wrap; gap:6px; margin-top:0; width:100%; min-width:0;\n}'
        ),
        (
            '.notiz-badge{\n  display:inline-block;\n  background:#fef3c7;\n  border:1px solid #fbbf24;\n  color:#92400e;\n  border-radius:4px;\n  padding:3px 8px;\n  font-size:10px;\n  font-weight:700;\n  line-height:1.25;\n  max-width:100%;\n  overflow:hidden;\n  text-overflow:ellipsis;\n  white-space:nowrap;\n}',
            '.notiz-badge{\n  display:inline-flex;\n  align-items:flex-start;\n  background:#fef3c7;\n  border:1px solid #fbbf24;\n  color:#92400e;\n  border-radius:4px;\n  padding:4px 8px;\n  font-size:10px;\n  font-weight:700;\n  line-height:1.3;\n  max-width:100%;\n  min-width:0;\n  overflow:hidden;\n  text-overflow:clip;\n  white-space:normal;\n  overflow-wrap:anywhere;\n  word-break:break-word;\n}'
        ),
        (
            '.tour-inline{display:flex; flex-wrap:wrap; gap:6px; align-content:flex-start; min-height:26px}',
            '.tour-inline{display:flex; flex-wrap:wrap; gap:6px; row-gap:8px; align-content:flex-start; min-height:26px; width:100%; min-width:0}'
        ),
        (
            '.tour-btn{\n  display:inline-flex;\n  align-items:center;\n  min-height:24px;\n  background:var(--chip-tour-bg);\n  border:1px solid var(--chip-tour-bd);\n  color:var(--chip-tour-tx);\n  padding:4px 8px;\n  border-radius:var(--radius-pill);\n  font-weight:800;\n  font-size:10px;\n  cursor:pointer;\n  line-height:1.2;\n  letter-spacing:.12px;\n  box-shadow:none;\n}',
            '.tour-btn{\n  display:inline-flex;\n  align-items:center;\n  min-height:24px;\n  max-width:100%;\n  min-width:0;\n  flex:0 1 auto;\n  background:var(--chip-tour-bg);\n  border:1px solid var(--chip-tour-bd);\n  color:var(--chip-tour-tx);\n  padding:4px 8px;\n  border-radius:var(--radius-pill);\n  font-weight:800;\n  font-size:10px;\n  cursor:pointer;\n  line-height:1.2;\n  letter-spacing:.12px;\n  box-shadow:none;\n}'
        ),
        (
            '.phone-col{display:grid; grid-template-columns:1fr; gap:6px; align-content:start; min-height:108px}',
            '.phone-col{display:grid; grid-template-columns:1fr; gap:6px; align-content:start; min-height:108px; min-width:0}'
        ),
        (
            '<col style="width:15%">\n            <col style="width:33%">\n            <col style="width:18%">\n            <col style="width:10%">\n            <col style="width:24%">',
            '<col style="width:13%">\n            <col style="width:38%">\n            <col style="width:21%">\n            <col style="width:8%">\n            <col style="width:20%">'
        ),
    ]
    for old, new in replacements:
        template = template.replace(old, new)
    return template


def _patch_suche_template_header(template: str) -> str:
    """Header etwas kompakter und linksbuendig (vorher Inline-Patch)."""
    return (
        template
        .replace(
            ".header{\n  padding:10px 12px;\n  background:linear-gradient(180deg,#f0f2f7 0%,#e6eaf1 100%);\n  border-bottom:1px solid var(--grid);\n  display:flex; align-items:center; justify-content:center;\n  position:relative;\n}",
            ".header{\n  padding:8px 18px;\n  background:linear-gradient(180deg,#f0f2f7 0%,#e6eaf1 100%);\n  border-bottom:1px solid var(--grid);\n  display:flex; align-items:center; justify-content:flex-start;\n  position:relative;\n}"
        )
        .replace(
            ".brand-logo{height:46px; width:auto}",
            ".brand-logo{height:32px; width:auto; margin-left:4px}"
        )
    )


def _patch_suche_template_weniger_luftig(template: str) -> str:
    """Macht die obere Übersicht wieder etwas kompakter, ohne ungleich zu wirken."""
    replacements = [
        (
            '.table-section{\n  padding:10px 12px 16px;\n  overflow:visible;\n}',
            '.table-section{\n  padding:8px 12px 12px;\n  overflow:visible;\n}'
        ),
        (
            'tbody td{\n  padding:10px 10px;\n  vertical-align:top;\n  font-weight:650;\n  border-bottom:1px solid var(--grid);\n  border-right:1px solid var(--grid);\n  background:#fff;\n  overflow:hidden;\n}',
            'tbody td{\n  padding:8px 9px;\n  vertical-align:top;\n  font-weight:650;\n  border-bottom:1px solid var(--grid);\n  border-right:1px solid var(--grid);\n  background:#fff;\n  overflow:hidden;\n}'
        ),
        (
            'tbody tr+tr td{border-top:2px solid var(--row-sep)}',
            'tbody tr+tr td{border-top:1px solid var(--row-sep)}'
        ),
        (
            '.cell{display:flex; flex-direction:column; gap:8px; min-height:96px; width:100%; min-width:0; justify-content:flex-start}',
            '.cell{display:flex; flex-direction:column; gap:4px; min-height:68px; width:100%; min-width:0; justify-content:flex-start}'
        ),
        (
            '.cell-stack{display:flex; flex-direction:column; gap:8px; min-height:100%; min-width:0}',
            '.cell-stack{display:flex; flex-direction:column; gap:4px; min-height:100%; min-width:0}'
        ),
        (
            '.cell-top{font-weight:800; min-height:18px; line-height:1.25}',
            '.cell-top{font-weight:800; min-height:16px; line-height:1.15}'
        ),
        (
            '.cell-sub{color:var(--muted); min-height:18px; line-height:1.25}',
            '.cell-sub{color:var(--muted); min-height:14px; line-height:1.15}'
        ),
        (
            '.cell-slot.note-slot{min-height:32px}',
            '.cell-slot.note-slot{min-height:22px}'
        ),
        (
            '.tour-inline{display:flex; flex-wrap:wrap; gap:6px; row-gap:8px; align-content:flex-start; min-height:26px; width:100%; min-width:0}',
            '.tour-inline{display:flex; flex-wrap:wrap; gap:6px; row-gap:6px; align-content:flex-start; min-height:24px; width:100%; min-width:0}'
        ),
        (
            '.tour-btn{\n  display:inline-flex;\n  align-items:center;\n  min-height:24px;\n  max-width:100%;\n  min-width:0;\n  flex:0 1 auto;\n  background:var(--chip-tour-bg);\n  border:1px solid var(--chip-tour-bd);\n  color:var(--chip-tour-tx);\n  padding:4px 8px;\n  border-radius:var(--radius-pill);\n  font-weight:800;\n  font-size:10px;\n  cursor:pointer;\n  line-height:1.2;\n  letter-spacing:.12px;\n  box-shadow:none;\n}',
            '.tour-btn{\n  display:inline-flex;\n  align-items:center;\n  min-height:22px;\n  max-width:100%;\n  min-width:0;\n  flex:0 1 auto;\n  background:var(--chip-tour-bg);\n  border:1px solid var(--chip-tour-bd);\n  color:var(--chip-tour-tx);\n  padding:3px 8px;\n  border-radius:var(--radius-pill);\n  font-weight:800;\n  font-size:10px;\n  cursor:pointer;\n  line-height:1.15;\n  letter-spacing:.12px;\n  box-shadow:none;\n}'
        ),
        (
            '.phone-col{display:grid; grid-template-columns:1fr; gap:6px; align-content:start; min-height:108px; min-width:0}',
            '.phone-col{display:grid; grid-template-columns:1fr; gap:5px; align-content:start; min-height:92px; min-width:0}'
        ),
    ]
    for old, new in replacements:
        template = template.replace(old, new)
    return template




def _patch_suche_template_aufklappbare_hinweise(template: str) -> str:
    # v16: Mehr-Buttons in der oberen Übersicht entfernt
    """Macht die Hinweise in der oberen Tour-Übersicht sichtbar aufklappbar."""
    original_css = """.ts-stack{
  min-height:46px;
  display:flex;
  flex-direction:column;
  justify-content:flex-start;
  gap:4px;
  width:100%;
}
.ts-main{
  font-weight:750;
  white-space:nowrap;
  overflow:hidden;
  text-overflow:ellipsis;
}
.ts-main.name{
  font-weight:850;
  font-size:10.5px;
}
.ts-sub{
  min-height:21px;
  font-size:8.5px;
  font-weight:550;
  font-style:italic;
  color:#8b7355;
  line-height:1.3;
  white-space:normal;
  overflow:hidden;
  display:-webkit-box;
  -webkit-box-orient:vertical;
  -webkit-line-clamp:2;
}
.ts-sub.empty{
  visibility:hidden;
}
.tour-row{ cursor:pointer; }
.tour-row:hover td{ background:#eff6ff; }"""
    enhanced_css = """.ts-stack{
  min-height:46px;
  display:flex;
  flex-direction:column;
  justify-content:flex-start;
  gap:4px;
  width:100%;
}
.ts-main{
  font-weight:750;
  white-space:nowrap;
  overflow:hidden;
  text-overflow:ellipsis;
}
.ts-main.name{
  font-weight:850;
  font-size:10.5px;
}
.ts-note-wrap{
  display:flex;
  flex-direction:column;
  align-items:flex-start;
  gap:3px;
  width:100%;
  min-width:0;
  padding-right:0;
}
.ts-sub{
  min-height:14px;
  width:100%;
  font-size:8.5px;
  font-weight:550;
  font-style:italic;
  color:#8b7355;
  line-height:1.25;
  white-space:nowrap;
  overflow:hidden;
  text-overflow:ellipsis;
  display:block;
  cursor:default;
}
.ts-note-toggle{
  display:none !important;
}
.ts-note-wrap.empty{
  padding-right:0;
}
.ts-sub.empty{
  visibility:hidden;
}
.tour-row{ cursor:pointer; }
.tour-row:hover td{ background:#eff6ff; }
@media print{
  .ts-note-toggle{ display:none !important; }
  .ts-note-wrap{ padding-right:0 !important; }
  .ts-sub{
    white-space:normal !important;
    overflow:visible !important;
    text-overflow:clip !important;
    display:block !important;
  }
}"""

    template = template.replace(original_css, enhanced_css)
    template = template.replace(""".ts-note-wrap{
  display:flex;
  flex-direction:column;
  align-items:flex-start;
  gap:3px;
  width:100%;
  min-width:0;
}""", """.ts-note-wrap{
  display:flex;
  flex-direction:column;
  align-items:flex-start;
  gap:3px;
  width:100%;
  min-width:0;
  padding-right:0;
}""")
    template = template.replace(""".ts-sub{
  min-height:14px;
  width:100%;
  font-size:8.5px;
  font-weight:550;
  font-style:italic;
  color:#8b7355;
  line-height:1.25;
  white-space:nowrap;
  overflow:hidden;
  text-overflow:ellipsis;
  display:block;
}""", """.ts-sub{
  min-height:14px;
  width:100%;
  font-size:8.5px;
  font-weight:550;
  font-style:italic;
  color:#8b7355;
  line-height:1.25;
  white-space:nowrap;
  overflow:hidden;
  text-overflow:ellipsis;
  display:block;
}
""")
    template = template.replace(""".ts-note-toggle{
  display:inline-flex;
  align-items:center;
  justify-content:center;
  min-height:18px;
  padding:1px 7px;
  border:1px solid #cfd7e3;
  border-radius:999px;
  background:#ffffff;
  color:#52627c;
  font-size:8px;
  font-weight:800;
  line-height:1;
  cursor:pointer;
}""", """.ts-note-toggle{
  display:none !important;
}""")
    template = template.replace(""".ts-note-wrap.empty .ts-note-toggle{
  display:none;
}""", """.ts-note-wrap.empty{
  padding-right:0;
}
.ts-note-wrap.empty .ts-note-toggle{
  display:none;
}""")
    template = template.replace("""@media print{
  .ts-note-toggle{ display:none !important; }
  .ts-sub{
    white-space:normal !important;
    overflow:visible !important;
    text-overflow:clip !important;
    display:block !important;
  }
}""", """@media print{
  .ts-note-toggle{ display:none !important; }
  .ts-note-wrap{ padding-right:0 !important; }
  .ts-sub{
    white-space:normal !important;
    overflow:visible !important;
    text-overflow:clip !important;
    display:block !important;
  }
}""")

    original_js = """    const td3=document.createElement('td');
    const td3w=document.createElement('div'); td3w.className='ts-stack';
    const td3m=document.createElement('div'); td3m.className='ts-main name'; td3m.textContent=name;
    const td3s=document.createElement('div'); td3s.className='ts-sub';
    const _nz = kundenNotizen[csb];
    const _parts=[];
    if(_nz){
      if(_nz.c) _parts.push(_nz.c);
      if(_nz.d) _parts.push(_nz.d);
    }
    if(_parts.length){
      td3s.textContent=_parts.join(' · ');
    } else {
      td3s.classList.add('empty');
      td3s.textContent='-';
    }
    td3w.append(td3m, td3s);
    td3.appendChild(td3w);

"""
    enhanced_js = """    const td3=document.createElement('td');
    const td3w=document.createElement('div'); td3w.className='ts-stack';
    const td3m=document.createElement('div'); td3m.className='ts-main name'; td3m.textContent=name;
    const td3noteWrap=document.createElement('div'); td3noteWrap.className='ts-note-wrap';
    const td3s=document.createElement('div'); td3s.className='ts-sub';
    const _nz = kundenNotizen[csb];
    const _parts=[];
    if(_nz){
      if(_nz.c) _parts.push(_nz.c);
      if(_nz.d) _parts.push(_nz.d);
    }
    if(_parts.length){
      const _noteText = _parts.join(' · ');
      td3s.textContent = _noteText;
      td3s.title = _noteText;
    } else {
      td3noteWrap.classList.add('empty');
      td3s.classList.add('empty');
      td3s.textContent='-';
    }
    td3noteWrap.append(td3s);
    td3w.append(td3m, td3noteWrap);
    td3.appendChild(td3w);

"""
    template = template.replace(original_js, enhanced_js)
    template = template.replace("""      td3s.textContent = _noteText;
      td3s.title = _noteText;
      td3btn.addEventListener('click', (ev)=>{
        ev.preventDefault();
        ev.stopPropagation();
        const expanded = td3noteWrap.classList.toggle('expanded');
        td3btn.textContent = expanded ? 'Weniger' : 'Mehr';
      });
""", """      td3s.textContent = _noteText;
      td3s.title = _noteText;
""")
    return template




def _patch_suche_template_tour_summary_collapsible(template: str) -> str:
    template = template.replace(""".tour-summary-head{
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:10px;
  padding:6px 8px;
  border-bottom:1px solid var(--grid);
  background:linear-gradient(180deg,#ffffff 0%, #f7f9fe 100%);
}""", """.tour-summary-head{
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:10px;
  padding:6px 8px;
  border-bottom:1px solid var(--grid);
  background:linear-gradient(180deg,#ffffff 0%, #f7f9fe 100%);
}
.tour-summary-head-main{
  display:flex;
  align-items:center;
  gap:8px;
  min-width:0;
  flex:1 1 auto;
  cursor:pointer;
}
.tour-summary-head-main:hover .tour-summary-title{
  color:#1d4ed8;
}
.tour-summary.collapsed .tour-summary-head{
  border-bottom:none;
}
.tour-summary.collapsed .tour-summary-tablewrap{
  display:none;
}
.tour-summary.collapsed #btnCopyTour,
.tour-summary.collapsed #btnPrintTour{
  display:none;
}""")

    template = template.replace(""".tour-summary-actions{ display:flex; align-items:center; gap:6px; }
.print-btn{""", """.tour-summary-actions{ display:flex; align-items:center; gap:6px; flex-wrap:wrap; justify-content:flex-end; }
.summary-toggle-btn{
  min-width:128px;
  background:linear-gradient(180deg,#ffe27a 0%,#cbd5e1 100%) !important;
  border-color:#d4a514 !important;
  color:#5f4200 !important;
  box-shadow:0 1px 0 rgba(255,255,255,.55) inset, 0 1px 2px rgba(15,23,42,.08) !important;
}
.summary-toggle-btn:hover{
  background:linear-gradient(180deg,#ffea97 0%,#f6d458 100%) !important;
  border-color:#c99700 !important;
  color:#4d3400 !important;
}
.summary-toggle-btn:active{
  background:linear-gradient(180deg,#efc845 0%,#e2b62f 100%) !important;
}
.print-btn{""")

    template = template.replace("""  .tour-summary-head{
    border:none !important;
    background:#fff !important;
    padding:0 0 6mm 0 !important;
  }

  .print-btn{ display:none !important; }
""", """  .tour-summary-head{
    border:none !important;
    background:#fff !important;
    padding:0 0 6mm 0 !important;
  }
  .tour-summary-head-main{
    cursor:default !important;
  }
  #tourSummary.collapsed .tour-summary-tablewrap{
    display:block !important;
  }

  .print-btn{ display:none !important; }
""")

    template = template.replace("""        <div class="tour-summary-head">
          <div>
            <div class="tour-summary-title" id="tourSummaryTitle"></div>
            <div class="tour-summary-meta" id="tourSummaryMeta"></div>
          </div>
          <div class="tour-summary-actions">
            <button class="print-btn" id="btnCopyTour" title="Tour als Tabelle (Outlook) kopieren">Kopieren</button>
            <button class="print-btn" id="btnPrintTour" title="Tour-Übersicht drucken (A4)">Drucken</button>
          </div>
        </div>
""", """        <div class="tour-summary-head">
          <div class="tour-summary-head-main" id="tourSummaryHeadMain" title="Übersicht anzeigen oder ausblenden">
            <div>
              <div class="tour-summary-title" id="tourSummaryTitle"></div>
              <div class="tour-summary-meta" id="tourSummaryMeta"></div>
            </div>
          </div>
          <div class="tour-summary-actions">
            <button class="print-btn summary-toggle-btn" id="btnToggleTourSummary" type="button" aria-expanded="false" title="Übersicht anzeigen oder ausblenden">Übersicht anzeigen</button>
            <button class="print-btn" id="btnCopyTour" title="Tour als Tabelle (Outlook) kopieren">Kopieren</button>
            <button class="print-btn" id="btnPrintTour" title="Tour-Übersicht drucken (A4)">Drucken</button>
          </div>
        </div>
""")

    template = template.replace("""function closeTourSummary(){
  $('#tourSummary').style.display='none';
  $('#tourSummaryTitle').textContent='';
  $('#tourSummaryMeta').textContent='';
  $('#tourSummaryBody').innerHTML='';
}
""", """function setTourSummaryCollapsed(collapsed){
  const wrap = $('#tourSummary');
  if(!wrap) return;
  wrap.classList.toggle('collapsed', !!collapsed);
  const btn = $('#btnToggleTourSummary');
  if(btn){
    btn.textContent = collapsed ? 'Übersicht anzeigen' : 'Übersicht ausblenden';
    btn.setAttribute('aria-expanded', collapsed ? 'false' : 'true');
  }
}
function toggleTourSummary(ev){
  if(ev){
    ev.preventDefault();
    ev.stopPropagation();
  }
  const wrap = $('#tourSummary');
  if(!wrap || wrap.style.display==='none') return;
  setTourSummaryCollapsed(!wrap.classList.contains('collapsed'));
}
function closeTourSummary(){
  $('#tourSummary').style.display='none';
  $('#tourSummaryTitle').textContent='';
  $('#tourSummaryMeta').textContent='';
  $('#tourSummaryBody').innerHTML='';
  setTourSummaryCollapsed(true);
}
""")

    template = template.replace("""  wrap.style.display='block';
}
""", """  wrap.style.display='block';
  setTourSummaryCollapsed(true);
}
""")

    template = template.replace("""$('#btnCopyTour').addEventListener('click', onCopyTour);

$('#btnPrintTour').addEventListener('click', ()=>{
  if($('#tourSummary').style.display==='none'){ return; }
  window.print();
});
""", """$('#btnCopyTour').addEventListener('click', onCopyTour);
$('#btnToggleTourSummary').addEventListener('click', toggleTourSummary);
$('#tourSummaryHeadMain').addEventListener('click', toggleTourSummary);
$('#tourSummaryHeadMain').addEventListener('keydown', (ev)=>{
  if(ev.key === 'Enter' || ev.key === ' '){
    toggleTourSummary(ev);
  }
});
$('#tourSummaryHeadMain').tabIndex = 0;

$('#btnPrintTour').addEventListener('click', ()=>{
  if($('#tourSummary').style.display==='none'){ return; }
  setTourSummaryCollapsed(false);
  window.print();
});
""")
    return template




def _patch_suche_template_tour_summary_collapsible_fix(template: str) -> str:
    """Macht den Aufklapper der oberen Übersicht robust und standardmäßig zugeklappt."""
    if '.tour-summary.collapsed .tour-summary-tablewrap{display:none !important;}' not in template:
        template = template.replace(
            '</style>',
            '.tour-summary.collapsed .tour-summary-tablewrap{display:none !important;}\n'
            '.tour-summary.collapsed #btnCopyTour,\n'
            '.tour-summary.collapsed #btnPrintTour{display:none !important;}\n'
            '.summary-toggle-btn{min-width:156px;font-weight:900;background:linear-gradient(180deg,#ffe27a 0%,#cbd5e1 100%) !important;border-color:#d4a514 !important;color:#5f4200 !important;box-shadow:0 1px 0 rgba(255,255,255,.55) inset, 0 1px 2px rgba(15,23,42,.08) !important;}\n'
            '.summary-toggle-btn:hover{background:linear-gradient(180deg,#ffea97 0%,#f6d458 100%) !important;border-color:#c99700 !important;color:#4d3400 !important;}\n'
            '.summary-toggle-btn:active{background:linear-gradient(180deg,#efc845 0%,#e2b62f 100%) !important;}\n'
            '.tour-summary-head-main{cursor:pointer;}\n'
            '</style>',
            1,
        )

    if 'id="tourSummaryTableWrap"' not in template:
        template = template.replace(
            '<div class="tour-summary-tablewrap">',
            '<div class="tour-summary-tablewrap" id="tourSummaryTableWrap">',
            1,
        )

    if 'id="btnToggleTourSummary"' not in template:
        template = template.replace(
            '<div class="tour-summary-actions">',
            '<div class="tour-summary-actions">\n'
            '            <button class="print-btn summary-toggle-btn" id="btnToggleTourSummary" type="button" '
            'onclick="toggleTourSummary(event); return false;" aria-expanded="false" '
            'aria-controls="tourSummaryTableWrap" title="Übersicht anzeigen oder ausblenden">Übersicht anzeigen</button>',
            1,
        )

    template = template.replace(
        'id="btnToggleTourSummary" type="button"',
        'id="btnToggleTourSummary" type="button" onclick="toggleTourSummary(event); return false;" aria-controls="tourSummaryTableWrap"',
    )
    template = template.replace(
        'id="tourSummaryHeadMain" title="Übersicht anzeigen oder ausblenden"',
        'id="tourSummaryHeadMain" title="Übersicht anzeigen oder ausblenden" onclick="toggleTourSummary(event)" role="button" aria-controls="tourSummaryTableWrap"',
    )

    template = re.sub(
        r"wrap\.style\.display='block';\\n(\s*setTourSummaryCollapsed\(true\);\\n)?",
        "wrap.style.display='block';\\n  ensureTourSummaryToggleBindings();\\n  setTourSummaryCollapsed(true);\\n",
        template,
        count=1,
    )

    template = template.replace(
        "$('#btnCopyTour').addEventListener('click', onCopyTour);",
        "ensureTourSummaryToggleBindings();\\n$('#btnCopyTour').addEventListener('click', onCopyTour);",
        1,
    )

    template = template.replace(
        "$('#btnPrintTour').addEventListener('click', ()=>{\\n  if($('#tourSummary').style.display==='none'){ return; }\\n  window.print();\\n});",
        "$('#btnPrintTour').addEventListener('click', ()=>{\\n  const wrap = $('#tourSummary');\\n  if(!wrap || getComputedStyle(wrap).display==='none'){ return; }\\n  setTourSummaryCollapsed(false);\\n  window.print();\\n});",
        1,
    )

    helper = """
function setTourSummaryCollapsed(collapsed){
  const wrap = $('#tourSummary');
  if(!wrap) return;
  wrap.classList.toggle('collapsed', !!collapsed);
  const btn = $('#btnToggleTourSummary');
  if(btn){
    btn.textContent = collapsed ? 'Übersicht anzeigen' : 'Übersicht ausblenden';
    btn.setAttribute('aria-expanded', collapsed ? 'false' : 'true');
  }
}
function ensureTourSummaryToggleBindings(){
  const btn = $('#btnToggleTourSummary');
  if(btn && !btn.dataset.toggleBound){
    btn.dataset.toggleBound = '1';
    btn.onclick = function(ev){ toggleTourSummary(ev); return false; };
  }
  const head = $('#tourSummaryHeadMain');
  if(head && !head.dataset.toggleBound){
    head.dataset.toggleBound = '1';
    head.onclick = function(ev){ toggleTourSummary(ev); };
    head.tabIndex = 0;
    head.addEventListener('keydown', function(ev){
      if(ev.key === 'Enter' || ev.key === ' '){
        toggleTourSummary(ev);
      }
    });
  }
}
function toggleTourSummary(ev){
  if(ev){
    ev.preventDefault();
    ev.stopPropagation();
  }
  const wrap = $('#tourSummary');
  if(!wrap || getComputedStyle(wrap).display === 'none') return false;
  setTourSummaryCollapsed(!wrap.classList.contains('collapsed'));
  return false;
}
document.addEventListener('DOMContentLoaded', function(){
  ensureTourSummaryToggleBindings();
  setTourSummaryCollapsed(true);
});
"""
    if helper not in template:
        template = template.replace('</script>', helper + '\n</script>', 1)

    template = template.replace(
        "ensureTourSummaryToggleBindings();\\n$('#btnCopyTour').addEventListener('click', onCopyTour);",
        "ensureTourSummaryToggleBindings();\n$('#btnCopyTour').addEventListener('click', onCopyTour);"
    )
    template = template.replace(
        "$('#btnPrintTour').addEventListener('click', ()=>{\n  if($('#tourSummary').style.display==='none'){ return; }\n  window.print();\n});",
        "$('#btnPrintTour').addEventListener('click', ()=>{\n  const wrap = $('#tourSummary');\n  if(!wrap || getComputedStyle(wrap).display==='none'){ return; }\n  setTourSummaryCollapsed(false);\n  window.print();\n});",
        1,
    )

    return template





def _patch_suche_template_search_all_inputs(template: str) -> str:
    """Die Suche soll bei jeder Eingabe alle Felder durchsuchen,
    auch bei kurzen oder rein numerischen Werten wie 3-, 4- oder 5-stelligen Eingaben."""
    old = """function onSmart(){
  const qRaw=$('#smartSearch').value.trim();
  closeTourSummary();

  if(!qRaw){ renderTable([]); return; }

  if(/^\\d{1,3}$/.test(qRaw)){
    const n=qRaw.replace(/^0+(\\d)/,'$1');
    const r=allCustomers.filter(k=>(k.touren||[]).some(t=>(t.tournummer||'').startsWith(n)));
    renderTable(r);
    return;
  }

  if(/^\\d{4}$/.test(qRaw)){
    const n=qRaw.replace(/^0+(\\d)/,'$1');
    const tr=allCustomers.filter(k=>(k.touren||[]).some(t=>(t.tournummer||'')===n));
    const cr=allCustomers.filter(k=>(k.csb_nummer||'')===n);
    const r=dedupByCSB([...tr,...cr]);

    if(tr.length){
      renderTourSummary(tr, n);
    }
    renderTable(r);
    return;
  }

  const q=normDE(qRaw);
  const r=allCustomers.filter(k=>{
    const fb=k.fachberater||'';
    const text=(k.name+' '+k.strasse+' '+k.ort+' '+k.csb_nummer+' '+k.sap_nummer+' '+fb+' '+(k.schluessel||'')+' '+(k.fb_phone||'')+' '+(k.market_phone||'')+' '+(k.market_email||'')+' '+((kundenNotizen[k.csb_nummer]||{}).c||'')+' '+((kundenNotizen[k.csb_nummer]||{}).d||''));
    return normDE(text).includes(q);
  });
  renderTable(r);
}"""

    new = """function onSmart(){
  const qRaw=$('#smartSearch').value.trim();
  closeTourSummary();

  if(!qRaw){ renderTable([]); return; }

  const q = normDE(qRaw);
  const n = normalizeDigits(qRaw);
  const isNumeric = /^\\d+$/.test(qRaw);
  const normalizeRahmentourCode = (value) => String(value||'').toUpperCase().replace(/\\s+/g,'');
  const getRahmentourList = (tournummer, dayLabel) => {
    const key = normalizeDigits(tournummer);
    const raw = rahmentourIndex[key];
    const day = String(dayLabel||'').trim();
    const rows = Array.isArray(raw) ? raw : (raw ? [raw] : []);
    const exact = rows
      .filter(item => item && typeof item === 'object' && String(item.day||'').trim() === day)
      .map(item => normalizeRahmentourCode(item.sap))
      .filter(Boolean);
    const fallback = rows
      .map(item => {
        if(item && typeof item === 'object') return normalizeRahmentourCode(item.sap);
        return normalizeRahmentourCode(item);
      })
      .filter(Boolean);
    return (exact.length ? exact : fallback).filter((v, i, a) => a.indexOf(v) === i);
  };
  const rahmenQuery = normalizeRahmentourCode(qRaw);

  let r;

  if(isNumeric && n){
    // Strict numeric search: match only numbers with same digit count
    const nLen = n.length;
    r = allCustomers.filter(k => {
      if(normalizeDigits(k.csb_nummer) === n) return true;
      if(normalizeDigits(k.sap_nummer) === n) return true;
      if(normalizeDigits(k.schluessel) === n) return true;
      return (k.touren||[]).some(t => {
        const tn = normalizeDigits(t.tournummer);
        return tn && tn.length === nLen && tn === n;
      });
    });

    // Show tour summary for exact tour matches
    const tr = allCustomers.filter(k => (k.touren||[]).some(t => normalizeDigits(t.tournummer) === n));
    if(tr.length) renderTourSummary(tr, n);
    r = dedupByCSB(tr.length ? [...tr, ...r] : r);
  } else {
    // Text search: broad matching
    r = allCustomers.filter(k=>{
      const fb = k.fachberater || '';
      const tourText = (k.touren||[]).map(t => {
        const tourNum = t.tournummer || '';
        return [tourNum, t.liefertag || ''].filter(Boolean).join(' ');
      }).join(' ');
      const text = (
        (k.name||'') + ' ' +
        (k.strasse||'') + ' ' +
        (k.ort||'') + ' ' +
        (k.csb_nummer||'') + ' ' +
        (k.sap_nummer||'') + ' ' +
        fb + ' ' +
        (k.schluessel||'') + ' ' +
        (k.fb_phone||'') + ' ' +
        (k.market_phone||'') + ' ' +
        (k.market_email||'') + ' ' +
        ((kundenNotizen[k.csb_nummer]||{}).c||'') + ' ' +
        ((kundenNotizen[k.csb_nummer]||{}).d||'') + ' ' +
        tourText
      );
      return normDE(text).includes(q);
    });

    if(rahmenQuery){
      const rr = allCustomers.filter(k => (k.touren||[]).some(t => getRahmentourList(t.tournummer, t.liefertag).some(rahmen => rahmen === rahmenQuery)));
      if(rr.length){
        renderTourSummary(rr, rahmenQuery);
        r = dedupByCSB([...rr, ...r]);
      }
    }
  }

  renderTable(dedupByCSB(r));
}"""

    if old in template:
        return template.replace(old, new, 1)
    return template




def _patch_suche_template_kisoft_rahmentour(template: str) -> str:
    """Blendet die Kisoft-Rahmentour in der Tour-Uebersicht sichtbar ein."""
    template = template.replace(
        ".tour-summary-meta{ display:none !important; }",
        ".tour-summary-meta{display:block !important;font-size:11px;font-weight:850;color:#7c5b00;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}"
    )

    template = template.replace(
        """  // Wochentag(e) für diese Tour
  const daySet = new Set();
  for(const k of list){
    for(const t of (k.touren||[])){
      if(String(t.tournummer||'').trim() === String(tour).trim()){
        if(t.liefertag) daySet.add(String(t.liefertag).trim());
      }
    }
  }
  const dayOrder = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"];
  const days = Array.from(daySet).sort((a,b)=>dayOrder.indexOf(a)-dayOrder.indexOf(b));
  const dayLabel = days.length ? days.join("/") : "";

  $('#tourSummaryTitle').textContent = dayLabel ? `${tour} – ${dayLabel}` : `${tour}`;
  $('#tourSummaryMeta').textContent  = "";
""",
        """  const normalizeRahmentourCode = (value) => String(value||'').toUpperCase().replace(/\\s+/g,'');
  const queryTourDigits = normalizeDigits(tour);
  const queryRahmen = normalizeRahmentourCode(tour);
  const getRahmentourList = (tournummer, dayLabel) => {
    const tourNum = normalizeDigits(tournummer);
    const raw = rahmentourIndex[tourNum];
    const day = String(dayLabel||'').trim();
    const rows = Array.isArray(raw) ? raw : (raw ? [raw] : []);
    const exact = rows
      .filter(item => item && typeof item === 'object' && String(item.day||'').trim() === day)
      .map(item => normalizeRahmentourCode(item.sap))
      .filter(Boolean);
    const fallback = rows
      .map(item => {
        if(item && typeof item === 'object') return normalizeRahmentourCode(item.sap);
        return normalizeRahmentourCode(item);
      })
      .filter(Boolean);
    return (exact.length ? exact : fallback).filter((v, i, a) => a.indexOf(v) === i);
  };
  const matchesTour = (t) => {
    const tourNum = normalizeDigits(t && t.tournummer);
    const rahmenListe = getRahmentourList(t && t.tournummer, t && t.liefertag);
    if(queryTourDigits && tourNum === queryTourDigits) return true;
    if(queryRahmen && rahmenListe.includes(queryRahmen)) return true;
    return false;
  };
  const extractTournummer = (k) => {
    for(const t of (k.touren||[])){
      if(matchesTour(t)) return normalizeDigits(t.tournummer) || String(t.tournummer||'').trim();
    }
    return '';
  };

  const summaryTournummer = list.map(extractTournummer).find(Boolean) || normalizeDigits(tour) || String(tour||'').trim();

  // Wochentag(e) für diese Tour oder Rahmentour
  const daySet = new Set();
  for(const k of list){
    for(const t of (k.touren||[])){
      if(matchesTour(t)){
        if(t.liefertag) daySet.add(String(t.liefertag).trim());
      }
    }
  }
  const dayOrder = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"];
  const days = Array.from(daySet).sort((a,b)=>dayOrder.indexOf(a)-dayOrder.indexOf(b));
  const dayLabel = days.length ? days.join("/") : "";

  const summaryRahmentourSet = new Set();
  for(const k of list){
    for(const t of (k.touren||[])){
      if(matchesTour(t)){
        getRahmentourList(t.tournummer, t.liefertag).forEach(v => summaryRahmentourSet.add(v));
      }
    }
  }
  if(!summaryRahmentourSet.size){
    getRahmentourList(summaryTournummer, dayLabel).forEach(v => summaryRahmentourSet.add(v));
  }
  const summaryRahmentour = Array.from(summaryRahmentourSet).join(', ');

  $('#tourSummaryTitle').textContent = dayLabel ? `${summaryTournummer} – ${dayLabel}` : `${summaryTournummer}`;
  $('#tourSummaryMeta').textContent  = summaryRahmentour ? `Kisoft Rahmentouren: ${summaryRahmentour}` : "";
""",
        1,
    )

    template = template.replace(
        "const lfa = (a.lf_map && a.lf_map[tour]) ? a.lf_map[tour] : '';",
        "const lfa = (a.lf_map && a.lf_map[summaryTournummer]) ? a.lf_map[summaryTournummer] : '';",
        1,
    )
    template = template.replace(
        "const lfb = (b.lf_map && b.lf_map[tour]) ? b.lf_map[tour] : '';",
        "const lfb = (b.lf_map && b.lf_map[summaryTournummer]) ? b.lf_map[summaryTournummer] : '';",
        1,
    )
    template = template.replace(
        "const lf   = (k.lf_map && k.lf_map[tour]) ? String(k.lf_map[tour]).trim() : '';",
        "const lf   = (k.lf_map && k.lf_map[summaryTournummer]) ? String(k.lf_map[summaryTournummer]).trim() : '';",
        1,
    )
    return template





def _patch_suche_template_rahmentour_list_in_rows(template: str) -> str:
    """Zeigt in der Ergebnisliste alle Kisoft-Rahmentouren je Tour an."""
    old = """    const _rh = rahmentourIndex[tnum];
    if(_rh){
      const rhSpan = document.createElement('span');
      rhSpan.className = 'rahmen';
      rhSpan.textContent = _rh;
      b.appendChild(rhSpan);
    }
"""
    new = """    const _rhRaw = rahmentourIndex[tnum];
    const _day = String(t.liefertag || '').trim();
    const _rhRows = Array.isArray(_rhRaw) ? _rhRaw : (_rhRaw ? [_rhRaw] : []);
    const _rhExact = _rhRows
      .filter(item => item && typeof item === 'object' && String(item.day || '').trim() === _day)
      .map(item => String(item.sap || '').trim())
      .filter(Boolean);
    const _rhFallback = _rhRows
      .map(item => item && typeof item === 'object' ? String(item.sap || '').trim() : String(item || '').trim())
      .filter(Boolean);
    const _rhList = (_rhExact.length ? _rhExact : _rhFallback).filter((v, i, a) => a.indexOf(v) === i);
"""
    if old in template:
        return template.replace(old, new, 1)
    return template





def _patch_suche_template_sonderliste_marktkauf(template: str) -> str:
    """Sends Kunden-Liste data to parent via postMessage for standalone panel."""

    # Add JS helper functions and postMessage sender after allCustomers is built
    template = template.replace(
        "  allCustomers = Array.from(map.values());\n}",
        """  allCustomers = Array.from(map.values());
}

const KUNDEN_LISTE_GROUP_ORDER = ['SuL','Malchow','Neumünster','Direkt','Marktkauf','Gemischt','Ohne Rahmentour'];

function normalizeRahmentourCodeGlobal(value){
  return String(value||'').toUpperCase().replace(/\\s+/g,'').trim();
}
function getRahmentourListGlobal(tournummer, dayLabel){
  const key = normalizeDigits(tournummer);
  const raw = rahmentourIndex[key];
  const day = String(dayLabel||'').trim();
  const rows = Array.isArray(raw) ? raw : (raw ? [raw] : []);
  const exact = rows
    .filter(item => item && typeof item === 'object' && String(item.day||'').trim() === day)
    .map(item => normalizeRahmentourCodeGlobal(item.sap))
    .filter(Boolean);
  const fallback = rows
    .map(item => item && typeof item === 'object' ? normalizeRahmentourCodeGlobal(item.sap) : normalizeRahmentourCodeGlobal(item))
    .filter(Boolean);
  return (exact.length ? exact : fallback).filter((v, i, a) => a.indexOf(v) === i);
}
function getCustomerRahmentourCodes(k){
  const out = [];
  (k.touren||[]).forEach(t => {
    getRahmentourListGlobal(t.tournummer, t.liefertag).forEach(code => {
      if(code && out.indexOf(code) === -1) out.push(code);
    });
  });
  return out;
}
function classifyKundenCodes(codes){
  if(!codes.length) return 'Ohne Rahmentour';
  const hasZ = codes.some(code => code.includes('Z'));
  const hasM = codes.some(code => code.includes('M'));
  const hasN = codes.some(code => code.includes('N'));
  const count = (hasZ?1:0) + (hasM?1:0) + (hasN?1:0);
  if(count > 1) return 'Gemischt';
  if(hasZ) return 'SuL';
  if(hasM) return 'Malchow';
  if(hasN) return 'Neumünster';
  return 'Direkt';
}

function hasMarktkaufTour(k){
  return (k.touren||[]).some(t => {
    const num = String(t.tournummer||'').replace(/\\D/g,'');
    return /^\\d88\\d$/.test(num);
  });
}
function hasMalchowTour(k){
  return (k.touren||[]).some(t => {
    const num = String(t.tournummer||'').replace(/\\D/g,'');
    return /^\\d777\\d$/.test(num);
  });
}
function hasNMSTour(k){
  return (k.touren||[]).some(t => {
    const num = String(t.tournummer||'').replace(/\\D/g,'');
    return /^\\d222\\d$/.test(num);
  });
}

function sendKundenDataToParent(){
  if(!allCustomers || !allCustomers.length) return;
  const groups = {};
  KUNDEN_LISTE_GROUP_ORDER.forEach(g => { groups[g] = []; });
  const ordered = [...allCustomers].sort((a, b) => {
    const c = String(a.name||'').localeCompare(String(b.name||''), 'de');
    return c !== 0 ? c : String(a.sap_nummer||'').localeCompare(String(b.sap_nummer||''), 'de');
  });
  ordered.forEach(k => {
    const codes = getCustomerRahmentourCodes(k);
    let group = classifyKundenCodes(codes);
    if(hasMarktkaufTour(k)) group = 'Marktkauf';
    else if(hasMalchowTour(k) && (group === 'Direkt' || group === 'Ohne Rahmentour')) group = 'Malchow';
    else if(hasNMSTour(k) && (group === 'Direkt' || group === 'Ohne Rahmentour')) group = 'Neumünster';
    const touren = (k.touren||[])
      .map(t => {
        const num = normalizeDigits(t.tournummer) || String(t.tournummer||'').trim();
        const day = String(t.liefertag||'').trim();
        return day ? (num + ' (' + day + ')') : num;
      })
      .filter(Boolean)
      .filter((v, i, a) => a.indexOf(v) === i)
      .join(', ');
    if(!groups[group]) groups[group] = [];
    groups[group].push({
      sap: normalizeDigits(k.sap_nummer) || '',
      csb: normalizeDigits(k.csb_nummer) || '',
      name: k.name || '-',
      ort: [k.postleitzahl||'', k.ort||''].filter(Boolean).join(' '),
      bereich: k.bereich || '-',
      touren: touren || '-',
      rahmentouren: codes.length ? codes.join(', ') : '-'
    });
  });
  try {
    window.parent.postMessage({type:'kunden-liste-data', groups: groups, order: KUNDEN_LISTE_GROUP_ORDER}, '*');
  } catch(e){}
}
// Auto-send after a short delay to let iframe fully initialize
setTimeout(sendKundenDataToParent, 500);
window.onKundenListe = function(){ sendKundenDataToParent(); };
window.onMkListe = window.onKundenListe;
""",
        1,
    )

    return template



def _patch_suche_template_perf_searchindex(template: str) -> str:
    """Vor-normalisierter Suchindex pro Kunde + DocumentFragment-Rendering.

    Spart O(n) Stringkonkatenation + normDE() pro Tastendruck (#5),
    und O(n) DOM-Reflows beim Tabellen-Render (#6).
    """
    # 1) buildData: am Ende _search pro Kunde befuellen
    needle_build = "  allCustomers = Array.from(map.values());\n}"
    repl_build = (
        "  allCustomers = Array.from(map.values());\n"
        "  // Vor-normalisierter Suchindex pro Kunde — vermeidet Konkatenation pro Tastendruck\n"
        "  for(const c of allCustomers){\n"
        "    const csb = c.csb_nummer || '';\n"
        "    const noteC = (kundenNotizen[csb] && kundenNotizen[csb].c) || '';\n"
        "    const noteD = (kundenNotizen[csb] && kundenNotizen[csb].d) || '';\n"
        "    const tourText = (c.touren||[]).map(t => ((t.tournummer||'') + ' ' + (t.liefertag||''))).join(' ');\n"
        "    c._search = normDE([\n"
        "      c.name||'', c.strasse||'', c.ort||'', csb,\n"
        "      c.sap_nummer||'', c.fachberater||'', c.schluessel||'',\n"
        "      c.fb_phone||'', c.market_phone||'', c.market_email||'',\n"
        "      noteC, noteD, tourText\n"
        "    ].join(' '));\n"
        "  }\n"
        "}"
    )
    if needle_build in template:
        template = template.replace(needle_build, repl_build, 1)

    # 2) onSmart Text-Branch: lange Konkatenation durch k._search ersetzen
    # Wir suchen das gesamte "const text = ( ... );  return normDE(text).includes(q);"
    pattern_textsearch = re.compile(
        r"const text = \(\s*\n"
        r"(?:[^;]+\n)+?"
        r"\s*\);\s*\n"
        r"\s*return normDE\(text\)\.includes\(q\);",
        re.MULTILINE
    )
    template = pattern_textsearch.sub(
        "return (k._search || '').includes(q);",
        template,
        count=1
    )

    # 3) renderTable: DocumentFragment statt einzelnem appendChild
    needle_rt = (
        "  body.innerHTML='';\n"
        "  if(list.length){\n"
        "    list.forEach(k=>body.appendChild(rowFor(k)));\n"
        "    tbl.style.display='table';"
    )
    repl_rt = (
        "  body.innerHTML='';\n"
        "  if(list.length){\n"
        "    const frag = document.createDocumentFragment();\n"
        "    list.forEach(k=>frag.appendChild(rowFor(k)));\n"
        "    body.appendChild(frag);\n"
        "    tbl.style.display='table';"
    )
    if needle_rt in template:
        template = template.replace(needle_rt, repl_rt, 1)

    return template




def _patch_suche_template_multi_keys(template: str) -> str:
    """Unterstuetzt mehrere Schluesselnummern je Kundennummer."""
    normalize_needle = """function normalizeDigits(v){
  if(v == null) return '';
  let s = String(v).trim().replace(/\\.0$/,'');
  s = s.replace(/[^0-9]/g,'').replace(/^0+(\\d)/,'$1');
  return s;
}"""
    normalize_repl = normalize_needle + """
function normalizeKeyList(value){
  const source = Array.isArray(value)
    ? value
    : String(value == null ? '' : value).split(/[\\s,;|/]+/);
  const out = [];
  for(const item of source){
    const key = normalizeDigits(item);
    if(key && !out.includes(key)) out.push(key);
  }
  return out;
}
function customerKeys(customer){
  const direct = normalizeKeyList(customer && customer.schluessel);
  if(direct.length) return direct;
  const csb = customer && customer.csb_nummer ? customer.csb_nummer : '';
  return normalizeKeyList(keyIndex[csb]);
}
function customerKeyText(customer){
  return customerKeys(customer).join(' / ');
}"""
    if normalize_needle in template:
        template = template.replace(normalize_needle, normalize_repl, 1)

    template = template.replace(
        "rec.schluessel   = normalizeDigits(rec.schluessel) || (keyIndex[csb]||'');",
        "rec.schluessel   = normalizeKeyList(rec.schluessel).length ? normalizeKeyList(rec.schluessel) : normalizeKeyList(keyIndex[csb]);",
        1,
    )

    template = template.replace(
        "if(normalizeDigits(k.schluessel) === n) return true;",
        "if(customerKeys(k).includes(n)) return true;",
        1,
    )

    template = template.replace(
        "c.sap_nummer||'', c.fachberater||'', c.schluessel||'',",
        "c.sap_nummer||'', c.fachberater||'', customerKeyText(c),",
        1,
    )

    old_row = """  const key=(k.schluessel||'')||(keyIndex[csb]||'');
  keySlot.appendChild(key ? el('span','badge-key',key) : makePlaceholder('Kein Schlüssel'));"""
    new_row = """  const keys = customerKeys(k);
  if(keys.length){
    keys.forEach(key => keySlot.appendChild(el('span','badge-key',key)));
  } else {
    keySlot.appendChild(makePlaceholder('Kein Schlüssel'));
  }"""
    if old_row in template:
        template = template.replace(old_row, new_row, 1)

    old_key_search = """  const r=[];
  for(const k of allCustomers){
    const key=(k.schluessel||'')||(keyIndex[k.csb_nummer]||'');
    if(key===n) r.push(k);
  }"""
    new_key_search = """  const r=[];
  for(const k of allCustomers){
    if(customerKeys(k).includes(n)) r.push(k);
  }"""
    if old_key_search in template:
        template = template.replace(old_key_search, new_key_search, 1)

    return template

def _patch_suche_template_optik(template: str) -> str:
    """Kleinere Optik-Korrekturen (#8 Webfont, #10 Body-Weight)."""
    # #8: Webfont-Cocktail abspecken — nur Inter Tight (3 Weights) + JetBrains Mono (1 Weight)
    old_fonts = (
        '<link href="https://fonts.googleapis.com/css2?'
        'family=Inter:wght@500;600;700;800;900'
        '&family=Inter+Tight:wght@500;600;700;800;900'
        '&family=JetBrains+Mono:wght@500;600;700&display=swap" rel="stylesheet">'
    )
    new_fonts = (
        '<link rel="preconnect" href="https://fonts.googleapis.com">'
        '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>'
        '<link href="https://fonts.googleapis.com/css2?'
        'family=Inter+Tight:wght@600;700;900'
        '&family=JetBrains+Mono:wght@600&display=swap" rel="stylesheet">'
    )
    if old_fonts in template:
        template = template.replace(old_fonts, new_fonts, 1)

    # #10: body font-weight 650 → 600 (lesbarer, ruhiger) — auch ohne Semikolon
    template = re.sub(r"font-weight\s*:\s*650\b", "font-weight:600", template)

    return template



def _patch_druck_template_cleanup(template: str) -> str:
    """Entfernt Debug-console.log-Spam aus dem Druck-Template (#7).

    Im Live-Einsatz unnoetig und unprofessionell.
    """
    # Alle console.log / console.warn / console.error Statements entfernen
    template = re.sub(
        r"^\s*console\.(?:log|warn|error|info|debug)\([^;]*\);?\s*\n",
        "",
        template,
        flags=re.MULTILINE,
    )
    return template

DRUCK_HTML_TEMPLATE_PATCHES = (_patch_druck_template_cleanup,)


@st.cache_resource(show_spinner=False)

def _patch_suche_template_kundenart_absetzer_rampe(template: str) -> str:
    """Zeigt je Kunde eine SAP-basierte Kennzeichnung Absetzer/Rampenkunde in der Suche."""
    # Daten-Konstante einfuegen
    template = template.replace(
        "const rahmentourIndex  = {  };",
        "const rahmentourIndex  = {  };\nconst kundenArtIndex    = {  };",
        1,
    )

    # Kleine Badge-Optik fuer die Ergebnisliste
    css = """
.kundenart-chip{
  display:inline-flex;
  align-items:center;
  min-height:24px;
  max-width:100%;
  padding:4px 8px;
  border-radius:5px;
  border:1px solid #8fb7ff;
  background:#eef6ff;
  color:#1d4ed8;
  font-size:10px;
  font-weight:900;
  line-height:1.1;
  white-space:nowrap;
  overflow:hidden;
  text-overflow:ellipsis;
}
.kundenart-chip.rampe{
  border-color:#86efac;
  background:#ecfdf5;
  color:#166534;
}
.kundenart-chip.kombi{
  border-color:#fbbf24;
  background:#fffbeb;
  color:#92400e;
}
"""
    if ".kundenart-chip" not in template:
        template = template.replace("</style>", css + "\n</style>", 1)

    helper = """function makeKundenArtChip(value){
  const label = String(value||'').trim();
  if(!label) return null;
  const s = document.createElement('span');
  const n = normDE(label);
  s.className = 'kundenart-chip' + (n.includes('rampe') && n.includes('absetzer') ? ' kombi' : (n.includes('rampe') ? ' rampe' : ''));
  s.textContent = label;
  s.title = 'Kundenart: ' + label;
  return s;
}
"""
    if "function makeKundenArtChip" not in template:
        template = template.replace(
            "function makeEmptyChip(label, value='-'){",
            helper + "\nfunction makeEmptyChip(label, value='-'){",
            1,
        )

    # Kundenart bereits beim Datenaufbau an den Kunden haengen
    template = template.replace(
        "        rec.sap_nummer   = normalizeDigits(rec.sap_nummer);",
        "        rec.sap_nummer   = normalizeDigits(rec.sap_nummer);\n        rec.kunden_art   = kundenArtIndex[rec.sap_nummer] || '';",
        1,
    )

    # In der CSB/SAP-Spalte anzeigen
    template = template.replace(
        "  idSlot1.appendChild(makeIdChip('CSB', csb));\n  idSlot2.appendChild(makeIdChip('SAP', sap));\n  c1.append(idSlot1,idSlot2);",
        "  idSlot1.appendChild(makeIdChip('CSB', csb));\n  idSlot2.appendChild(makeIdChip('SAP', sap));\n  c1.append(idSlot1,idSlot2);\n  const artChip = makeKundenArtChip(k.kunden_art || '');\n  if(artChip){ const artSlot = el('div','cell-slot compact'); artSlot.appendChild(artChip); c1.appendChild(artSlot); }",
        1,
    )

    # Kundenart in den Suchindex aufnehmen, damit Suche nach Absetzer/Rampe funktioniert
    template = template.replace(
        "      noteC, noteD, tourText",
        "      noteC, noteD, c.kunden_art||'', tourText",
        1,
    )
    return template

def _patch_suche_template_enter_search(template: str) -> str:
    """Enter im Suchfeld loest nur die zum fokussierten Feld passende Suche aus.
    Vorher rief der Enter-Handler onSmart() UND onKey() auf – bei leerem Exakt-Feld
    leerte onKey() die gerade gefuellte Trefferliste sofort wieder."""
    old = (
        "    if(e.key === 'Enter'){\n"
        "      const a = document.activeElement;\n"
        "      if(a && (a.id==='smartSearch' || a.id==='keySearch')){\n"
        "        onSmart();\n"
        "        onKey();\n"
        "      }\n"
        "    }"
    )
    new = (
        "    if(e.key === 'Enter'){\n"
        "      const a = document.activeElement;\n"
        "      if(a && a.id==='smartSearch'){ e.preventDefault(); onSmart(); }\n"
        "      else if(a && a.id==='keySearch'){ e.preventDefault(); onKey(); }\n"
        "    }"
    )
    return template.replace(old, new)

def get_suche_template() -> str:
    """Baut das Suche-Template einmalig (nach Re-import gecached).
    Spart bei jedem Streamlit-Re-Run die ~11 String-Patches uebers ~700 KB Template."""
    tpl = base64.b64decode(_static_payload_text("_SUCHE_B64")).decode("utf-8")
    for patch in (
        _patch_suche_template_layout,
        _patch_suche_template_header,
        _patch_suche_template_weniger_luftig,
        _patch_suche_template_aufklappbare_hinweise,
        _patch_suche_template_tour_summary_collapsible,
        _patch_suche_template_tour_summary_collapsible_fix,
        _patch_suche_template_search_all_inputs,
        _patch_suche_template_enter_search,
        _patch_suche_template_kisoft_rahmentour,
        _patch_suche_template_rahmentour_list_in_rows,
        _patch_suche_template_sonderliste_marktkauf,
        _patch_suche_template_perf_searchindex,
        _patch_suche_template_kundenart_absetzer_rampe,
        _patch_suche_template_multi_keys,
        _patch_suche_template_optik,
    ):
        tpl = patch(tpl)
    return tpl


@st.cache_resource(show_spinner=False)
def get_druck_template() -> str:
    """Baut das Druck-Template einmalig (nach Re-import gecached)."""
    tpl = base64.b64decode(_static_payload_text("_DRUCK_B64")).decode("utf-8")
    for patch in DRUCK_HTML_TEMPLATE_PATCHES:
        tpl = patch(tpl)
    return tpl


# =============================================================================
# DRUCK – Konstanten & Hilfsfunktionen
# =============================================================================

BEREICH  = "Alle Sortimente Fleischwerk"
DAYS_DE  = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag"]
TOUR_COLS = {
    "Montag": "Mo", "Dienstag": "Die", "Mittwoch": "Mitt",
    "Donnerstag": "Don", "Freitag": "Fr", "Samstag": "Sam",
}
DAY_SHORT_TO_DE = {
    "Mo": "Montag",  "Di": "Dienstag", "Die": "Dienstag",
    "Mi": "Mittwoch","Mit": "Mittwoch","Mitt": "Mittwoch",
    "Do": "Donnerstag","Don": "Donnerstag","Donn": "Donnerstag",
    "Fr": "Freitag", "Sa": "Samstag",  "Sam": "Samstag",
}
SORT_PRIO = {"21": 0, "1011": 1, "22": 2, "41": 3, "65": 4, "0": 5, "91": 6}


def norm_val(x) -> str:
    if x is None:
        return ""
    if isinstance(x, float) and pd.isna(x):
        return ""
    s = str(x).replace("\u00a0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def normalize_time(s) -> str:
    if isinstance(s, (datetime.time, pd.Timestamp)):
        return s.strftime("%H:%M") + " Uhr"
    s = norm_val(s)
    if not s:
        return ""
    if re.fullmatch(r"\d{1,2}:\d{2}", s):
        return s + " Uhr"
    if re.fullmatch(r"\d{1,2}", s):
        return s.zfill(2) + ":00 Uhr"
    return s


def safe_time(val) -> str:
    raw = norm_val(val)
    if re.fullmatch(r"(Montag|Dienstag|Mittwoch|Donnerstag|Freitag|Samstag)", raw):
        return ""
    return normalize_time(val)


def canon_group_id(label: str) -> str:
    s = norm_val(label).lower()
    m = re.search(r"\b(1011|21|41|65|0|91|22)\b", s)
    if m:
        return m.group(1)
    if "bio" in s and "gefl" in s:
        return "41"
    if "wiesenhof" in s or "gefl" in s:
        return "1011"
    if "frischfleisch" in s or "veredlung" in s or "schwein" in s or "p" in s and "k" in s:
        return "65"
    if "fleisch" in s or "wurst" in s or "heidemark" in s:
        return "21"
    if "avo" in s or "gew" in s:
        return "0"
    if "werbe" in s:
        return "91"
    if any(x in s for x in ("pfeiffer","gmyrek","siebert","bard","mago")):
        return "22"
    return "?"


def detect_neue_bspalten(columns: list) -> dict:
    """
    Erkennt neues B-Spalten-Format: '{Tag} {Gruppe} B_{Feld}'
    Feld = Zeit | Sortiment | Sort | Tag
    Bestelltag steht als Datenwert in der B_Tag-Spalte (nicht im Spaltennamen).
    Gibt dict zurück: {(liefertag, gruppe): {"zeit": col, "sort": col, "tag_col": col}}
    """
    rx = re.compile(
        r"^(Mo|Die|Di|Mitt|Mit|Mi|Don|Donn|Do|Fr|Sam|Sa)\s+(\d+)\s+B_(Zeit|Sortiment|Sort|Tag)$",
        re.IGNORECASE,
    )
    groups: dict = {}
    order:  list = []
    for col in columns:
        m = rx.match(str(col).strip())
        if not m:
            continue
        day = DAY_SHORT_TO_DE.get(m.group(1))
        grp = m.group(2)
        feld = m.group(3).lower()
        if not day:
            continue
        key = (day, grp)
        if key not in groups:
            groups[key] = {}
            order.append(key)
        fk = "sort" if feld in ("sortiment", "sort") else feld
        groups[key][fk] = col
    return groups


def detect_bspalten(columns: List[str]) -> dict:
    rx_b = re.compile(
        r"^(Mo|Die|Di|Mitt|Mit|Mi|Don|Donn|Do|Fr|Sam|Sa)\s+"
        r"(?:(Z|L)\s+)?(.+?)\s+B[_ ]\s*(Mo|Die|Di|Mitt|Mit|Mi|Don|Donn|Do|Fr|Sam|Sa)$",
        re.IGNORECASE,
    )
    rx_no_b = re.compile(
        r"^(Mo|Die|Di|Mitt|Mit|Mi|Don|Donn|Do|Fr|Sam|Sa)\s+"
        r"(Z|L)\s+(.+?)\s+(Mo|Die|Di|Mitt|Mit|Mi|Don|Donn|Do|Fr|Sam|Sa)$",
        re.IGNORECASE,
    )
    mapping: dict = {}
    for c in columns:
        if re.search(r"\sB[_ ]\s*", c, re.IGNORECASE):
            continue
        m = rx_no_b.match(c.strip())
        if m:
            dd = DAY_SHORT_TO_DE.get(m.group(1))
            zl = m.group(2).upper()
            gt = m.group(3).strip()
            bd = DAY_SHORT_TO_DE.get(m.group(4))
            if dd and bd:
                key = (dd, gt, bd)
                mapping.setdefault(key, {})
                if zl == "Z":   mapping[key]["zeit"] = c
                elif zl == "L": mapping[key]["l"]    = c
    for c in columns:
        m = rx_b.match(c.strip())
        if m:
            dd = DAY_SHORT_TO_DE.get(m.group(1))
            zl = (m.group(2) or "").upper()
            gt = m.group(3).strip()
            bd = DAY_SHORT_TO_DE.get(m.group(4))
            if dd and bd:
                key = (dd, gt, bd)
                mapping.setdefault(key, {})
                if zl == "Z":
                    if "zeit" not in mapping[key]: mapping[key]["zeit"] = c
                elif zl == "L":
                    if "l" not in mapping[key]:    mapping[key]["l"]    = c
                else:
                    mapping[key]["sort"]       = c
                    mapping[key]["group_text"] = gt
    return mapping


def detect_neue_triplets(columns: list) -> list:
    """
    Erkennt neue Spaltenstruktur: Montag_Zeit, Montag_Sort, Montag_Tag (ggf. mit .1, .2 Suffixen).
    Gibt eine Liste von Dicts zurueck:
      {"liefertag": "Montag", "zeit_col": "Montag_Zeit", "sort_col": "Montag_Sort", "tag_col": "Montag_Tag"}
    """
    TAGE = "Montag|Dienstag|Mittwoch|Donnerstag|Freitag|Samstag"
    FELDER = "Zeit|Sort|Tag"
    # Suffix wie .1 .2 usw. (pandas doppelte Spaltennamen)
    rx = re.compile(
        r"^(" + TAGE + r")_(" + FELDER + r")(?:[.](\d+))?$",
        re.IGNORECASE,
    )
    groups: dict = {}
    order:  list = []
    for col in columns:
        m = rx.match(str(col).strip())
        if not m:
            continue
        day   = m.group(1).capitalize()
        field = m.group(2).lower()          # zeit | sort | tag
        suf   = int(m.group(3)) if m.group(3) else 0
        key   = (day, suf)
        if key not in groups:
            groups[key] = {}
            order.append(key)
        groups[key][field] = col
    result = []
    for key in order:
        g = groups[key]
        if "zeit" in g or "sort" in g or "tag" in g:
            result.append({
                "liefertag": key[0],
                "zeit_col":  g.get("zeit"),
                "sort_col":  g.get("sort"),
                "tag_col":   g.get("tag"),
            })
    return result


def detect_triplets(columns: List[str]) -> dict:
    rx = re.compile(
        r"^(Mo|Die|Di|Mitt|Mit|Mi|Don|Donn|Do|Fr|Sam|Sa)\s+(.+?)\s+"
        r"(Zeit|Zeitende|Bestellzeitende|Uhrzeit|Sort|Sortiment|Tag|Bestelltag)$",
        re.IGNORECASE,
    )
    found: dict = {}
    for c in columns:
        m = rx.match(c.strip())
        if not m: continue
        dd  = DAY_SHORT_TO_DE.get(m.group(1))
        if not dd: continue
        gt  = m.group(2).strip()
        ek  = m.group(3).lower()
        key = "Sort" if ek in ("sort","sortiment") else "Tag" if ek in ("tag","bestelltag") else "Zeit"
        found.setdefault(dd, {}).setdefault(gt, {})[key] = c
    return found


def detect_ds_triplets(columns: List[str]) -> dict:
    rx = re.compile(
        r"^DS\s+(.+?)\s+zu\s+(Mo|Die|Di|Mitt|Mit|Mi|Don|Donn|Do|Fr|Sam|Sa)\s+(Zeit|Sort|Tag)$",
        re.IGNORECASE,
    )
    tmp: dict = {}
    for c in columns:
        m = rx.match(c.strip())
        if not m: continue
        dd = DAY_SHORT_TO_DE.get(m.group(2))
        if dd:
            k = f"DS {m.group(1)} zu {m.group(2)}"
            tmp.setdefault(dd, {}).setdefault(k, {})[m.group(3).capitalize()] = c
    return tmp


def load_logo_data_uri() -> str:
    candidates = []
    try:
        here = Path(__file__).resolve().parent
        candidates.append(here / "Logo_NORDfrische Center (NFC).png")
    except Exception:
        pass
    candidates.append(Path.cwd() / "Logo_NORDfrische Center (NFC).png")
    candidates.append(Path("/mnt/data/Logo_NORDfrische Center (NFC).png"))
    for p in candidates:
        try:
            if p.exists() and p.is_file():
                return "data:image/png;base64," + base64.b64encode(p.read_bytes()).decode("ascii")
        except Exception:
            continue
    return ""


def logo_file_to_data_uri(f) -> str:
    if not f:
        return ""
    return f"data:{f.type or 'image/png'};base64," + base64.b64encode(f.getvalue()).decode("ascii")


# =============================================================================
# SUCHE – Hilfsfunktionen
# =============================================================================

BLATTNAMEN = ["DIREKT", "MK", "HUPA_NMS", "HUPA_MALCHOW"]

# Blattnamen robust finden:
# Excel kann bei Blattnamen Gross-/Kleinschreibung, Leerzeichen oder alte Namen enthalten.
# Darum wird nicht mehr nur exakt verglichen, sondern normalisiert gesucht.
BLATT_ALIASE = {
    "DIREKT": ["DIREKT", "Direkt", "Direkt 1 - 99"],
    "MK": ["MK", "Hupa MK 882"],
    "HUPA_NMS": ["HUPA_NMS", "HuPa_NMS", "Hupa 2221-4444"],
    "HUPA_MALCHOW": ["HUPA_MALCHOW", "HuPa_Malchow", "Hupa 7773-7779"],
}


def normalize_sheet_name_py(name: str) -> str:
    s = str(name or "")
    s = (s.replace("\u00A0", " ").replace("\ufeff", "").strip())
    s = re.sub(r"\s+", " ", s)
    return s.casefold().replace(" ", "").replace("_", "").replace("-", "")


def find_existing_sheet_name(available_names, *candidates) -> str:
    by_norm = {}
    for real_name in available_names or []:
        by_norm.setdefault(normalize_sheet_name_py(real_name), real_name)
    for candidate in candidates:
        real = by_norm.get(normalize_sheet_name_py(candidate))
        if real:
            return real
    return ""


def find_customer_sheet_names(available_names) -> list:
    found = []
    seen = set()
    for target in BLATTNAMEN:
        real = find_existing_sheet_name(available_names, *BLATT_ALIASE.get(target, [target]))
        if real and real not in seen:
            found.append(real)
            seen.add(real)
    return found

# Spaltennamen in der Quelldatei koennen je nach Version unterschiedlich heissen.
# Aktuelles Format laut Datei: CSB | SAP | Name | Strasse | Plz | Ort | Mo | Die | Mitt | Don | Fr | Sam
# Altes Format war unter anderem: Nr | SAP-Nr. | ...
SPALTEN_ALIASE = {
    "csb_nummer":   ["CSB", "Nr", "CSB-Nr", "CSB-Nr.", "CSB Nummer", "CSB-Nummer", "Kunden Nr", "Kunden-Nr", "Kundennummer"],
    "sap_nummer":   ["SAP", "SAP-Nr.", "SAP-Nr", "SAP Nr", "SAP Nummer", "SAP-Nummer"],
    "name":         ["Name", "Marktname", "Kundenname", "Marktname / Kundenname"],
    "strasse":      ["Strasse", "Straße", "Str.", "Strasse / Nr", "Straße / Nr"],
    "postleitzahl": ["Plz", "PLZ", "Postleitzahl"],
    "ort":          ["Ort"],
    "fachberater":  ["Fachberater", "Berater"],
}


def normalize_header_py(value: str) -> str:
    x = str(value or "")
    x = x.replace("\u00A0", " ").replace("\ufeff", "").strip().lower()
    x = (x.replace("ä", "ae").replace("ö", "oe")
           .replace("ü", "ue").replace("ß", "ss"))
    x = unicodedata.normalize("NFD", x)
    x = "".join(ch for ch in x if unicodedata.category(ch) != "Mn")
    x = re.sub(r"[^a-z0-9]+", "", x)
    return x


def find_column_index(columns, aliases) -> int | None:
    by_norm = {normalize_header_py(col): idx for idx, col in enumerate(columns)}
    for alias in aliases:
        idx = by_norm.get(normalize_header_py(alias))
        if idx is not None:
            return idx
    return None


def first_existing_column(columns, aliases) -> str:
    by_norm = {normalize_header_py(col): col for col in columns}
    for alias in aliases:
        col = by_norm.get(normalize_header_py(alias))
        if col is not None:
            return col
    return ""


def row_get_first(row, aliases, default=""):
    # Kompatibel mit pd.Series (hat .index) und dict (.keys())
    keys = row.index if hasattr(row, "index") else row.keys()
    for col in aliases:
        if col in keys:
            return row.get(col, default)
    by_norm = {normalize_header_py(col): col for col in keys}
    for alias in aliases:
        real = by_norm.get(normalize_header_py(alias))
        if real is not None:
            return row.get(real, default)
    return default
LIEFERTAGE_MAPPING = {
    "Montag": "Mo", "Dienstag": "Die", "Mittwoch": "Mitt",
    "Donnerstag": "Don", "Freitag": "Fr", "Samstag": "Sam",
}


def normalize_digits_py(v) -> str:
    if pd.isna(v):
        return ""
    s = str(v).strip().replace(".0", "")
    s = "".join(ch for ch in s if ch.isdigit())
    if not s:
        return ""
    s = s.lstrip("0")
    return s if s else "0"


def norm_de_py(s: str) -> str:
    if not s:
        return ""
    x = (s.replace("\u200b","").replace("\u200c","")
          .replace("\u200d","").replace("\ufeff","")
          .replace("\u00A0"," ").replace("\u2013","-").replace("\u2014","-").lower()
          .replace("\xe4","ae").replace("\xf6","oe").replace("\xfc","ue").replace("\xdf","ss"))
    x = unicodedata.normalize("NFD", x)
    x = "".join(ch for ch in x if unicodedata.category(ch) != "Mn")
    x = re.sub(r"\(.*?\)", " ", x)
    x = re.sub(r"[./,;:+*_#|]", " ", x)
    x = re.sub(r"-", " ", x)
    x = re.sub(r"[^a-z\s]", " ", x)
    return " ".join(x.split())


def build_key_map(df: pd.DataFrame) -> dict:
    """Liest alte und neue Schluesseldateien robust ein.

    Neues Format:
      Titelzeile
      Knd-Nr. | Schluessel-Nr.

    Das alte Format mit Kundennummer in Spalte 1 und Schluessel in Spalte 6
    bleibt weiterhin kompatibel. Mehrere Schluessel je Kunde werden erhalten.
    """
    if df is None or df.empty:
        return {}

    raw = df.dropna(how="all").reset_index(drop=True)
    if raw.empty or raw.shape[1] < 2:
        st.warning("Schluesseldatei enthaelt keine auswertbaren Spalten.")
        return {}

    csb_aliases = [
        "Knd-Nr.", "Knd-Nr", "Knd Nr", "Kundennummer", "Kunden-Nr.",
        "Kunden-Nr", "Kunden Nr", "CSB", "CSB-Nr.", "CSB-Nr",
        "CSB Nummer", "CSB-Nummer",
    ]
    key_aliases = [
        "Schluessel-Nr.", "Schluessel-Nr", "Schluessel Nr",
        "Schluesselnummer", "Schluessel", "Key", "Key-Nr.", "Key Nr",
    ]

    csb_norms = {normalize_header_py(v) for v in csb_aliases}
    key_norms = {normalize_header_py(v) for v in key_aliases}
    header_row = None
    csb_col = None
    key_col = None

    # Ueberschriften koennen wegen einer Titelzeile erst in Zeile 2 stehen.
    for row_idx in range(min(len(raw), 20)):
        values = raw.iloc[row_idx].tolist()
        normalized = [normalize_header_py(v) if not pd.isna(v) else "" for v in values]
        row_csb = next((idx for idx, val in enumerate(normalized) if val in csb_norms), None)
        row_key = next((idx for idx, val in enumerate(normalized) if val in key_norms), None)
        if row_csb is not None and row_key is not None and row_csb != row_key:
            header_row = row_idx
            csb_col = row_csb
            key_col = row_key
            break

    if csb_col is None or key_col is None:
        # Rueckwaertskompatibilitaet zum bisherigen 6-Spalten-Format.
        csb_col = 0
        key_col = 5 if raw.shape[1] > 5 else 1
        data = raw
    else:
        data = raw.iloc[header_row + 1:].reset_index(drop=True)

    collected: dict[str, list[str]] = {}
    for row in data.itertuples(index=False, name=None):
        csb = normalize_digits_py(row[csb_col] if len(row) > csb_col else "")
        key = normalize_digits_py(row[key_col] if len(row) > key_col else "")
        if not csb or not key:
            continue
        keys = collected.setdefault(csb, [])
        if key not in keys:
            keys.append(key)

    # Einzelschluessel bleiben Strings; nur Mehrfachzuordnungen werden Listen.
    # So bleiben auch aeltere HTML-Staende weitgehend kompatibel.
    return {csb: keys[0] if len(keys) == 1 else keys for csb, keys in collected.items()}


def build_berater_map(df: pd.DataFrame) -> dict:
    out = {}
    for row in df.itertuples(index=False, name=None):
        v = ("" if len(row) < 1 or pd.isna(row[0]) else str(row[0])).strip()
        n = ("" if len(row) < 2 or pd.isna(row[1]) else str(row[1])).strip()
        t = ("" if len(row) < 3 or pd.isna(row[2]) else str(row[2])).strip()
        if not t:
            continue
        for k in {norm_de_py(f"{v} {n}"), norm_de_py(f"{n} {v}")}:
            if k and k not in out:
                out[k] = t
    return out


def build_berater_csb_map(df: pd.DataFrame) -> dict:
    out = {}
    for row in df.itertuples(index=False, name=None):
        fach = str(row[0]).strip() if len(row) > 0 and not pd.isna(row[0]) else ""
        csb = normalize_digits_py(row[8]) if len(row) > 8 and not pd.isna(row[8]) else ""
        tel = str(row[14]).strip() if len(row) > 14 and not pd.isna(row[14]) else ""
        mail = str(row[23]).strip() if len(row) > 23 and not pd.isna(row[23]) else ""
        if csb:
            out[csb] = {"name": fach, "telefon": tel, "email": mail}
    return out


def format_lf(v) -> str:
    if pd.isna(v): return ""
    s = str(v).strip().replace(".0", "")
    if not s: return ""
    if s.isdigit(): return f"LF{int(s)}"
    s2 = s.replace(" ", "").upper()
    return s2 if s2.startswith("LF") else s


def build_winter_map(excel_file_obj) -> dict:
    """Liest die Ladefolge aus der Quelldatei.

    Neues Format laut aktueller Datei:
      Blatt: T-B-Druck Quelle
      Spalten: Datum | Tour | LA.F | CSB | NAME | STRASSE | ORT | PLZ | SAP-Nr.

    Fuer aeltere Dateien bleibt "Mo-Sa Winter" als Rueckfall erhalten.
    Ergebnis: {CSB: {Tour: LF...}}
    """
    out = {}

    try:
        excel_file_obj.seek(0)
    except Exception:
        pass

    try:
        book = pd.ExcelFile(excel_file_obj, engine=EXCEL_READ_ENGINE)
        available = list(book.sheet_names)
        sheet = find_existing_sheet_name(
            available,
            "T-B-Druck Quelle",
            "T-B Druck Quelle",
            "T B Druck Quelle",
            "T_B_Druck_Quelle",
            "TBDruckQuelle",
            "Mo-Sa Winter",
            "Mo Sa Winter",
        )
        if not sheet:
            return out
        dfw = pd.read_excel(book, sheet_name=sheet, header=0)
    except Exception:
        return out

    cols = list(dfw.columns)
    tour_col = first_existing_column(cols, ["Tour", "Tournr", "Tour Nr", "Tour-Nr", "Tournummer"])
    lf_col   = first_existing_column(cols, ["LA.F", "LAF", "LA F", "Ladefolge", "Lade Folge", "LF"])
    csb_col  = first_existing_column(cols, ["CSB", "Kunde", "Kundennummer", "Kunden Nr", "Kunden-Nr", "Nr"])

    # Fallback auf feste Positionen aus dem Screenshot:
    # A Datum, B Tour, C LA.F, D CSB
    if not tour_col and len(cols) > 1:
        tour_col = cols[1]
    if not lf_col and len(cols) > 2:
        lf_col = cols[2]
    if not csb_col and len(cols) > 3:
        csb_col = cols[3]

    if not (tour_col and lf_col and csb_col):
        return out

    # Vektorisiert statt iterrows — pro Zeile fielen sonst 3 Funktionsaufrufe an,
    # bei 10k+ Zeilen war das mit 200-500ms der teuerste einzelne Schritt.
    sub = dfw[[csb_col, tour_col, lf_col]]

    def _vec_norm_digits(series: pd.Series) -> pd.Series:
        s = series.astype(str)
        s = s.str.replace(r"\.0$", "", regex=True)
        s = s.str.replace(r"\D", "", regex=True)
        s = s.str.lstrip("0")
        s = s.where(s != "", pd.NA)
        return s

    csb_s  = _vec_norm_digits(sub[csb_col])
    tour_s = _vec_norm_digits(sub[tour_col])

    valid = csb_s.notna() & tour_s.notna()
    if not valid.any():
        return out
    csb_arr  = csb_s[valid].tolist()
    tour_arr = tour_s[valid].tolist()
    lf_raw   = sub.loc[valid, lf_col].tolist()

    for kd, tour, raw in zip(csb_arr, tour_arr, lf_raw):
        lf = format_lf(raw)
        if lf:
            out.setdefault(kd, {})[tour] = lf
    return out

def to_data_url_suche(f) -> str:
    mime = f.type or ("image/png" if f.name.lower().endswith(".png") else "image/jpeg")
    return f"data:{mime};base64," + base64.b64encode(read_upload_bytes(f)).decode("utf-8")


# =============================================================================
# CACHED WRAPPER  – arbeiten mit rohen Bytes, sodass @st.cache_data greift.
# Vermeidet, dass dieselben Excel/CSV-Dateien bei jedem Streamlit-Re-Run
# erneut geparst werden. winter_map/notizen_map liefen vorher 2x pro Run.
# =============================================================================

@st.cache_data(show_spinner=False, max_entries=8)
def cached_winter_map(excel_bytes: bytes) -> dict:
    if not excel_bytes:
        return {}
    return build_winter_map(io.BytesIO(excel_bytes))


@st.cache_data(show_spinner=False, max_entries=8)
def cached_key_map(key_bytes: bytes) -> dict:
    if not key_bytes:
        return {}
    # header=None ist absichtlich: Die neue Datei hat eine Titelzeile oberhalb
    # der eigentlichen Spaltennamen. build_key_map erkennt die Kopfzeile selbst.
    df = pd.read_excel(io.BytesIO(key_bytes), sheet_name=0, header=None, engine=EXCEL_READ_ENGINE)
    return build_key_map(df)


@st.cache_data(show_spinner=False, max_entries=8)
def cached_berater_map(berater_bytes: bytes) -> dict:
    if not berater_bytes:
        return {}
    bf = pd.read_excel(io.BytesIO(berater_bytes), sheet_name=0, header=None, engine=EXCEL_READ_ENGINE)
    bf = bf.rename(columns={0: "Vorname", 1: "Nachname", 2: "Nummer"}).dropna(how="all")
    return build_berater_map(bf)


@st.cache_data(show_spinner=False, max_entries=8)
def cached_berater_csb_map(bcsb_bytes: bytes) -> dict:
    if not bcsb_bytes:
        return {}
    try:
        bcf = pd.read_excel(io.BytesIO(bcsb_bytes), sheet_name=0, header=0, engine=EXCEL_READ_ENGINE)
    except Exception:
        bcf = pd.read_excel(io.BytesIO(bcsb_bytes), sheet_name=0, header=None, engine=EXCEL_READ_ENGINE)
    return build_berater_csb_map(bcf)


@st.cache_data(show_spinner=False, max_entries=8)
def cached_lieferhinweis_map(csv_bytes: bytes) -> dict:
    if not csv_bytes:
        return {}
    return build_lieferhinweis_csv(io.BytesIO(csv_bytes))


@st.cache_data(show_spinner=False, max_entries=8)
def cached_rahmentour_map(csv_bytes: bytes) -> dict:
    if not csv_bytes:
        return {}
    return build_rahmentour_map(io.BytesIO(csv_bytes))


@st.cache_data(show_spinner=False, max_entries=8)
def cached_kundenart_map(csv_bytes: bytes) -> dict:
    if not csv_bytes:
        return {}
    return build_kundenart_map(io.BytesIO(csv_bytes))


# =============================================================================
# HTML-ERZEUGUNG: SUCHE
# =============================================================================

def build_kundenart_map(csv_file) -> dict:
    """Liest die Kundenart-CSV mit SAP NR., LKW-Absetzer und LKW-Rampe.
    Ergebnis: {sap_nummer: "Absetzer" | "Rampenkunde" | "Absetzer + Rampenkunde"}.
    """
    if csv_file is None:
        return {}
    try:
        csv_file.seek(0)
        raw = csv_file.read()
        if isinstance(raw, str):
            raw = raw.encode("utf-8", errors="ignore")
    except Exception:
        return {}

    if not raw:
        return {}

    def _read_csv_bytes(payload: bytes):
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
            try:
                return pd.read_csv(io.BytesIO(payload), sep=None, engine="python", encoding=enc, dtype=str)
            except Exception:
                continue
        try:
            return pd.read_csv(io.BytesIO(payload), sep=";", encoding="utf-8-sig", dtype=str)
        except Exception:
            return pd.DataFrame()

    def _norm_header_local(value: str) -> str:
        return norm_de_py(str(value or "")).replace("-", " ").replace(".", " ").strip()

    def _is_marked(value) -> bool:
        if value is None or pd.isna(value):
            return False
        s = str(value).strip().lower()
        if not s or s in ("nan", "none", "0", "nein", "no", "false", "falsch", "-"):
            return False
        return True

    df = _read_csv_bytes(raw)
    if df.empty:
        return {}

    norm_cols = {_norm_header_local(col): col for col in df.columns}

    sap_col = None
    for key, col in norm_cols.items():
        if "sap" in key and ("nr" in key or "nummer" in key or key.strip() == "sap"):
            sap_col = col
            break
    if sap_col is None:
        sap_col = df.columns[0]

    absetzer_col = None
    rampe_col = None
    for key, col in norm_cols.items():
        if "absetzer" in key:
            absetzer_col = col
        if "rampe" in key:
            rampe_col = col

    result = {}
    for _, row in df.iterrows():
        sap = normalize_digits_py(row.get(sap_col, ""))
        if not sap:
            continue
        labels = []
        if absetzer_col is not None and _is_marked(row.get(absetzer_col, "")):
            labels.append("Absetzer")
        if rampe_col is not None and _is_marked(row.get(rampe_col, "")):
            labels.append("Rampenkunde")
        if labels:
            result[sap] = " + ".join(labels)
    return result


def build_lieferhinweis_csv(csv_file) -> dict:
    """Liest Lieferhinweis-CSV: ';'-getrennt, gequotet.
    Felder: [0]=SAP-Nr, [1]=CSB-Nr, [2]=Name, [3]=Strasse, [4]=PLZ,
            [5]=Ort, [6]=Art/Rollcontainer, [7]=Lieferhinweis, ...
    Key im Ergebnis: CSB-Nr (normalisiert, führende Nullen entfernt).
    Gibt {csb: {'d': lieferhinweis}} zurück; Ladehilfsmittel wird nicht angezeigt."""
    if csv_file is None:
        return {}
    import csv as _csv
    import io as _io
    try:
        csv_file.seek(0)
        raw = csv_file.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8", errors="replace")
    except Exception:
        return {}
    result = {}
    reader = _csv.reader(_io.StringIO(raw), delimiter=";", quotechar='"')
    for row in reader:
        if len(row) < 2:
            continue
        csb = normalize_digits_py(row[1]) if len(row) > 1 else ""
        if not csb:
            continue
        art      = row[6].strip() if len(row) > 6 else ""
        hinweis  = row[7].strip() if len(row) > 7 else ""
        if art or hinweis:
            entry = {}
            if art:
                pass
            if hinweis:
                entry["d"] = hinweis
            result[csb] = entry
    return result


def build_rahmentour_map(csv_file) -> dict:
    """Liest Rahmentourprofil-CSV: ';'-getrennt, gequotet.
    Erwartet bevorzugt die Kopfzeilen "SAP Rahmentour", "CSB Tournummer" und "Wochentag".
    Gibt {csb_tournr: [{"sap": "...", "day": "Montag"}, ...]} zurück.
    Buchstaben wie M/N/Z bleiben erhalten; Mehrfachzuordnungen bleiben erhalten."""
    if csv_file is None:
        return {}
    import csv as _csv
    import io as _io
    try:
        csv_file.seek(0)
        raw = csv_file.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8-sig", errors="replace")
    except Exception:
        return {}

    def _norm_header(value: str) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip().lower())

    def _norm_sap(value) -> str:
        s = str(value or "").strip().replace(".0", "")
        s = re.sub(r"\s+", "", s).upper()
        return s

    result = {}
    reader = _csv.reader(_io.StringIO(raw), delimiter=";", quotechar='"')

    try:
        header = next(reader)
    except StopIteration:
        return result

    header_norm = [_norm_header(col) for col in header]
    sap_idx = header_norm.index("sap rahmentour") if "sap rahmentour" in header_norm else 0
    csb_idx = header_norm.index("csb tournummer") if "csb tournummer" in header_norm else 1
    day_idx = header_norm.index("wochentag") if "wochentag" in header_norm else 2

    for row in reader:
        if not row:
            continue
        sap = _norm_sap(row[sap_idx] if len(row) > sap_idx else "")
        csb = normalize_digits_py(row[csb_idx]) if len(row) > csb_idx else ""
        day = str(row[day_idx]).strip() if len(row) > day_idx else ""
        if not (csb and sap):
            continue
        result.setdefault(csb, [])
        entry = {"sap": sap, "day": day}
        if entry not in result[csb]:
            result[csb].append(entry)
    return result


def generate_suche_html(excel_file, key_file, logo_file,
                         berater_file, berater_csb_file,
                         lieferhinweis_csv=None, rahmentour_csv=None,
                         kundenart_csv=None) -> str:
    if logo_file is None:
        raise ValueError("Bitte Logo hochladen.")

    logo_data_url = to_data_url_suche(logo_file)
    excel_bytes = read_upload_bytes(excel_file)

    with st.spinner("Lese Schluesseldatei ..."):
        key_map = cached_key_map(read_upload_bytes(key_file))

    berater_map: dict = {}
    if berater_file is not None:
        with st.spinner("Lese Fachberater-Telefonliste ..."):
            berater_map = cached_berater_map(read_upload_bytes(berater_file))

    berater_csb_map: dict = {}
    if berater_csb_file is not None:
        with st.spinner("Lese Fachberater-CSB-Zuordnung ..."):
            berater_csb_map = cached_berater_csb_map(read_upload_bytes(berater_csb_file))

    with st.spinner("Lese Ladefolgen (T-B-Druck Quelle) ..."):
        winter_map = cached_winter_map(excel_bytes)

    tour_dict: dict = {}

    def kunden_sammeln(df: pd.DataFrame):
        column_index = {str(col): idx for idx, col in enumerate(df.columns)}
        if not column_index:
            return

        day_columns = [
            (tag, spaltenname, column_index.get(spaltenname))
            for tag, spaltenname in LIEFERTAGE_MAPPING.items()
            if spaltenname in column_index
        ]
        field_columns = {field: find_column_index(df.columns, aliases) for field, aliases in SPALTEN_ALIASE.items()}
        csb_idx = field_columns.get("csb_nummer")
        n_cols = len(df.columns)

        # Lokale Refs sparen Attribut-Lookups in der Hot-Loop
        _norm_dig = normalize_digits_py
        _kmap = key_map
        _bmap = berater_csb_map
        _setdef = tour_dict.setdefault

        # Felder die normalisiert werden muessen
        field_items = [(f, i) for f, i in field_columns.items() if i is not None]

        for row in df.itertuples(index=False, name=None):
            # Erst pruefen ob ueberhaupt ein gueltiger Liefertag in der Zeile ist
            # — vermeidet das Bauen des Entry-Dicts wenn der Kunde "leer" ist.
            valid_days = []
            for tag, _, day_idx in day_columns:
                if day_idx is None or day_idx >= n_cols:
                    continue
                v = row[day_idx]
                # Schneller String-Check ohne unnoetige Konversion
                if v is None or (isinstance(v, float) and v != v):  # NaN
                    continue
                tournr_raw = str(v).strip()
                if not tournr_raw:
                    continue
                # isdigit nach . entfernen
                if not tournr_raw.replace(".", "", 1).isdigit():
                    continue
                valid_days.append((tag, _norm_dig(tournr_raw)))

            if not valid_days:
                continue

            # Entry EINMAL pro Zeile bauen — vorher 1-3x (pro Liefertag) identisch.
            base_entry = {}
            for field, idx in field_items:
                if idx < n_cols:
                    v = row[idx]
                    base_entry[field] = "" if v is None else str(v).strip()
                else:
                    base_entry[field] = ""

            csb = _norm_dig(row[csb_idx] if csb_idx is not None and csb_idx < n_cols else "")
            base_entry["csb_nummer"] = csb
            base_entry["sap_nummer"] = _norm_dig(base_entry.get("sap_nummer", ""))
            base_entry["postleitzahl"] = _norm_dig(base_entry.get("postleitzahl", ""))
            base_entry["schluessel"] = _kmap.get(csb, "")

            if csb:
                bcsb = _bmap.get(csb)
                if bcsb and bcsb.get("name"):
                    base_entry["fachberater"] = bcsb["name"]

            # Pro Liefertag eine flache Kopie + tournummer/liefertag setzen
            for tag, tournr in valid_days:
                entry = base_entry.copy()
                entry["liefertag"] = tag
                _setdef(tournr, []).append(entry)

    with st.spinner("Verarbeite Kundendatei ..."):
        try:
            excel_book = pd.ExcelFile(io.BytesIO(excel_bytes), engine=EXCEL_READ_ENGINE)
            alle_blaetter = list(excel_book.sheet_names)
        except Exception:
            excel_book = None
            alle_blaetter = []

        vorhandene = find_customer_sheet_names(alle_blaetter)
        zu_lesen = vorhandene or alle_blaetter or BLATTNAMEN

        if excel_book is not None:
            for blatt in zu_lesen:
                try:
                    kunden_sammeln(pd.read_excel(excel_book, sheet_name=blatt))
                except (ValueError, KeyError):
                    continue
        else:
            for blatt in zu_lesen:
                try:
                    kunden_sammeln(pd.read_excel(io.BytesIO(excel_bytes), sheet_name=blatt))
                except (ValueError, KeyError):
                    continue

    if not tour_dict:
        blaetter_info = ", ".join(alle_blaetter) if alle_blaetter else "unbekannt"
        raise ValueError(
            f"Keine gueltigen Kundendaten gefunden. "
            f"Verfuegbare Blaetter: {blaetter_info}. "
            f"Erwartet: {', '.join(BLATTNAMEN)}"
        )

    sorted_tours = dict(sorted(
        tour_dict.items(),
        key=lambda kv: int(kv[0]) if str(kv[0]).isdigit() else 0,
    ))

    notizen_map: dict = cached_lieferhinweis_map(read_upload_bytes(lieferhinweis_csv))
    rahmen_map:  dict = cached_rahmentour_map(read_upload_bytes(rahmentour_csv))
    kundenart_map: dict = cached_kundenart_map(read_upload_bytes(kundenart_csv))

    # separators=(",", ":") — spart bei grossen Maps 10-20% Output-Groesse
    # und ist auch beim json.dumps selbst etwas schneller.
    _dump = lambda d: json.dumps(d, ensure_ascii=False, separators=(",", ":"))

    return (
        get_suche_template()
        .replace("const tourkundenData   = {  }",
                 f"const tourkundenData   = {_dump(sorted_tours)}")
        .replace("const keyIndex         = {  }",
                 f"const keyIndex         = {_dump(key_map)}")
        .replace("const beraterIndex     = {  }",
                 f"const beraterIndex     = {_dump(berater_map)}")
        .replace("const beraterCSBIndex  = {  }",
                 f"const beraterCSBIndex  = {_dump(berater_csb_map)}")
        .replace("const winterIndex      = {  }",
                 f"const winterIndex      = {_dump(winter_map)}")
        .replace("const kundenNotizen    = {  }",
                 f"const kundenNotizen    = {_dump(notizen_map)}")
        .replace("const rahmentourIndex  = {  }",
                 f"const rahmentourIndex  = {_dump(rahmen_map)}")
        .replace("const kundenArtIndex    = {  }",
                 f"const kundenArtIndex    = {_dump(kundenart_map)}")
        .replace("__LOGO_DATA_URL__", logo_data_url)
        .replace("</style>", ".header{display:none !important;} .page{padding-top:0 !important;} .container{margin-top:0 !important;} </style>")
    )


# =============================================================================
# WOCHENAUSLASTUNG: Aggregation aus der Quelldatei
# =============================================================================
# Liest die vier Depot-Blaetter und zaehlt pro Wochentag:
#   - Kunden/Tag = gefuellte (gueltige) Tournummer-Zellen
#   - Touren/Tag = distinkte Tournummern
# Rueckgabe-Struktur (kompakt, wird per Instanz in INSTANCES eingebettet):
#   {"days":[...], "depots":[...], "kunden":{depot:[6 Werte]}, "touren":{depot:[6 Werte]}}
WOCHE_DEPOT_LABELS = [
    ("DIREKT", "Direkt"),
    ("MK", "Marktkauf"),
    ("HUPA_NMS", "Neumünster"),
    ("HUPA_MALCHOW", "Malchow"),
]
WOCHE_DAYS = ["Mo", "Di", "Mi", "Do", "Fr", "Sa"]
# Reihenfolge passend zu WOCHE_DAYS; Schluessel = Spaltenname in LIEFERTAGE_MAPPING
WOCHE_DAY_COLS = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag"]
# Kunden, die als eigene Gruppe (statt unter ihrem Depot) gefuehrt werden.
# (Anzeige-Label, Suchbegriff im Namen [klein]). Reihenfolge = Anzeige-Reihenfolge am Ende.
WOCHE_SONDERKUNDEN = [("Picnic", "picnic")]


def compute_woche_data(excel_file) -> dict:
    try:
        excel_bytes = read_upload_bytes(excel_file)
    except Exception:
        return {}
    if not excel_bytes:
        return {}

    try:
        book = pd.ExcelFile(io.BytesIO(excel_bytes), engine=EXCEL_READ_ENGINE)
        available = list(book.sheet_names)
    except Exception:
        return {}

    depot_labels = [label for _, label in WOCHE_DEPOT_LABELS]
    sonder_labels = [lab for lab, _ in WOCHE_SONDERKUNDEN]
    # Akkumulatoren fuer alle moeglichen Gruppen
    kunden = {g: [0] * len(WOCHE_DAYS) for g in depot_labels + sonder_labels}
    touren_sets = {g: [set() for _ in WOCHE_DAYS] for g in depot_labels + sonder_labels}
    cust = {}  # ckey -> {"bits":int, "grp":label, "name":str}  (distinkte Kunden, Tage vereinigt)
    seen_depot = set()

    for target, label in WOCHE_DEPOT_LABELS:
        sheet = find_existing_sheet_name(available, *BLATT_ALIASE.get(target, [target]))
        if not sheet:
            continue
        try:
            df = pd.read_excel(book, sheet_name=sheet)
        except (ValueError, KeyError):
            continue
        seen_depot.add(label)

        cols_list = list(df.columns)
        pos = {str(c): idx for idx, c in enumerate(cols_list)}
        name_pos = pos.get("Name")
        csb_pos = pos.get("CSB")
        sap_pos = pos.get("SAP")
        ort_pos = pos.get("Ort")
        day_pos = [pos.get(LIEFERTAGE_MAPPING.get(sn)) for sn in WOCHE_DAY_COLS]
        n_cols = len(cols_list)

        for row in df.itertuples(index=False, name=None):
            # Zielgruppe bestimmen: Sonderkunde (z. B. Picnic) oder Depot
            nm_raw = ""
            if name_pos is not None and name_pos < n_cols:
                nm_raw = str(row[name_pos]).strip()
            grp = label
            nm_low = nm_raw.lower()
            for lab, needle in WOCHE_SONDERKUNDEN:
                if needle in nm_low:
                    grp = lab
                    break
            row_bits = 0
            for c, dp in enumerate(day_pos):
                if dp is None or dp >= n_cols:
                    continue
                v = row[dp]
                if v is None or (isinstance(v, float) and v != v):
                    continue
                s = str(v).strip()
                if not s or not s.replace(".", "", 1).isdigit():
                    continue
                kunden[grp][c] += 1
                touren_sets[grp][c].add(normalize_digits_py(s))
                row_bits |= (1 << c)

            if row_bits:
                csb_val = ""
                if csb_pos is not None and csb_pos < n_cols and row[csb_pos] is not None:
                    csb_val = normalize_digits_py(row[csb_pos])
                sap_val = ""
                if sap_pos is not None and sap_pos < n_cols and row[sap_pos] is not None:
                    sap_val = normalize_digits_py(row[sap_pos])
                ort_val = ""
                if ort_pos is not None and ort_pos < n_cols and row[ort_pos] is not None and not (isinstance(row[ort_pos], float) and row[ort_pos] != row[ort_pos]):
                    ort_val = str(row[ort_pos]).strip()
                # Kunde eindeutig per CSB (sonst Name) – Tage vereinigen
                ckey = csb_val if csb_val else ("n:" + nm_low)
                ent = cust.get(ckey)
                if ent is None:
                    cust[ckey] = {"bits": row_bits, "grp": grp, "name": nm_raw,
                                  "csb": csb_val, "sap": sap_val, "ort": ort_val}
                else:
                    ent["bits"] |= row_bits
                    if not ent.get("sap"):
                        ent["sap"] = sap_val
                    if not ent.get("ort"):
                        ent["ort"] = ort_val

    # Anzeige-Reihenfolge: vorhandene Depots, dann Sonderkunden mit Daten
    out_depots = [d for d in depot_labels if d in seen_depot]
    for lab in sonder_labels:
        if any(kunden[lab]):
            out_depots.append(lab)

    if not out_depots:
        return {}

    touren = {g: [len(s) for s in touren_sets[g]] for g in out_depots}
    kunden = {g: kunden[g] for g in out_depots}

    # Rhythmus-Aggregation (distinkte Kunden, Tage vereinigt)
    grp_index = {g: i for i, g in enumerate(out_depots)}
    patterns = {}
    kunden_liste = []
    for ent in cust.values():
        bits = ent["bits"]
        if not bits:
            continue
        key = str(bits)
        patterns[key] = patterns.get(key, 0) + 1
        gi = grp_index.get(ent["grp"], 0)
        kunden_liste.append([ent["name"], gi, bits, ent.get("csb", ""), ent.get("sap", ""), ent.get("ort", "")])
    # Kompakt + stabil sortiert (nach Name)
    kunden_liste.sort(key=lambda x: x[0].lower())

    return {
        "days": WOCHE_DAYS,
        "depots": out_depots,
        "kunden": kunden,
        "touren": touren,
        "patterns": patterns,
        "kunden_liste": kunden_liste,
    }


# =============================================================================
# HTML-ERZEUGUNG: DRUCK
# =============================================================================

SHEETS_DRUCK = {
    "direkt":  "DIREKT",
    "mk":      "MK",
    "nms":     "HUPA_NMS",
    "malchow": "HUPA_MALCHOW",
}


def generate_druck_html(up, logo_up, fcsb_file=None, lieferhinweis_csv=None) -> str:
    logo_uri = logo_file_to_data_uri(logo_up) or load_logo_data_uri()
    all_data: dict = {}

    _SHEETS_ALT = {
        "direkt":  "Direkt 1 - 99",
        "mk":      "Hupa MK 882",
        "nms":     "Hupa 2221-4444",
        "malchow": "Hupa 7773-7779",
    }

    # Excel einmal komplett in den Speicher lesen + ExcelFile-Handle wiederverwenden
    excel_bytes = read_upload_bytes(up)
    excel_buf = io.BytesIO(excel_bytes)
    try:
        excel_book = pd.ExcelFile(excel_buf, engine=EXCEL_READ_ENGINE)
        available_sheets = list(excel_book.sheet_names)
    except Exception:
        excel_book = None
        available_sheets = []

    for area_key, sheet_name in SHEETS_DRUCK.items():
        with st.spinner(f"Verarbeite: {sheet_name} ..."):
            df = None
            if excel_book is not None:
                candidates = [sheet_name, _SHEETS_ALT.get(area_key, "")]
                if area_key == "direkt":
                    candidates += BLATT_ALIASE["DIREKT"]
                elif area_key == "mk":
                    candidates += BLATT_ALIASE["MK"]
                elif area_key == "nms":
                    candidates += BLATT_ALIASE["HUPA_NMS"]
                elif area_key == "malchow":
                    candidates += BLATT_ALIASE["HUPA_MALCHOW"]
                real_sheet = find_existing_sheet_name(available_sheets, *candidates)
                if real_sheet:
                    try:
                        df = pd.read_excel(excel_book, sheet_name=real_sheet)
                    except Exception:
                        df = None
            if df is None:
                continue

        cols    = df.columns.tolist()
        trip    = detect_triplets(cols)
        neue    = detect_neue_triplets(cols)   # neues Format: Montag_Zeit / Montag_Sort / Montag_Tag
        bmap    = detect_bspalten(cols)
        nbmap   = detect_neue_bspalten(cols)   # neues B-Format: "Die 1001 B_Zeit" etc.
        ds_trip = detect_ds_triplets(cols)
        data: dict = {}

        # Indizes pro Liefertag einmalig vorbauen (vorher in jeder Zeile neu gefiltert).
        bmap_by_day:  dict = {}
        nbmap_by_day: dict = {}
        for k in bmap:
            bmap_by_day.setdefault(k[0], []).append(k)
        for k in nbmap:
            nbmap_by_day.setdefault(k[0], []).append(k)
        # Liste der neuen Triplets pro Liefertag
        neue_by_day: dict = {}
        for nt in neue:
            neue_by_day.setdefault(nt["liefertag"], []).append(nt)

        # Spalten-Indizes EINMAL vor der Schleife berechnen — vorher wurde
        # pro Zeile ein dict(zip(cols, row_tuple)) und 7x row_get_first() (mit
        # eigener normalize_header_py-Schleife je Aufruf) ausgefuehrt. Das war
        # bei 10k+ Zeilen der groesste Hot-Path in generate_druck_html.
        col_to_idx = {c: i for i, c in enumerate(cols)}
        col_to_idx_norm = {normalize_header_py(c): i for i, c in enumerate(cols)}

        def _alias_idx(aliases):
            for a in aliases:
                if a in col_to_idx:
                    return col_to_idx[a]
            for a in aliases:
                idx = col_to_idx_norm.get(normalize_header_py(a))
                if idx is not None:
                    return idx
            return None

        idx_csb   = _alias_idx(SPALTEN_ALIASE["csb_nummer"])
        idx_sap   = _alias_idx(SPALTEN_ALIASE["sap_nummer"])
        idx_name  = _alias_idx(SPALTEN_ALIASE["name"])
        idx_str   = _alias_idx(SPALTEN_ALIASE["strasse"])
        idx_plz   = _alias_idx(SPALTEN_ALIASE["postleitzahl"])
        idx_ort   = _alias_idx(SPALTEN_ALIASE["ort"])
        idx_fach  = _alias_idx(SPALTEN_ALIASE["fachberater"])
        idx_tour  = {d: col_to_idx.get(TOUR_COLS[d]) for d in DAYS_DE}

        # Spalten-Indizes fuer Triplet-/B-Spalten vorbauen — sonst wird r.get()
        # pro Zeile fuer jede Spalte aufgerufen.
        trip_idx: dict = {}      # {day: [(gt, sort_idx, zeit_idx, tag_idx)]}
        for d in DAYS_DE:
            if d in trip:
                bucket = []
                for gt, f_cols in trip[d].items():
                    bucket.append((
                        gt,
                        col_to_idx.get(f_cols.get("Sort")),
                        col_to_idx.get(f_cols.get("Zeit")),
                        col_to_idx.get(f_cols.get("Tag")),
                    ))
                trip_idx[d] = bucket

        neue_idx: dict = {}      # {day: [(sort_idx, zeit_idx, tag_idx)]}
        for d in DAYS_DE:
            bucket = []
            for nt in neue_by_day.get(d, ()):
                bucket.append((
                    col_to_idx.get(nt.get("sort_col")) if nt.get("sort_col") else None,
                    col_to_idx.get(nt.get("zeit_col")) if nt.get("zeit_col") else None,
                    col_to_idx.get(nt.get("tag_col"))  if nt.get("tag_col")  else None,
                ))
            if bucket:
                neue_idx[d] = bucket

        bmap_idx: dict = {}      # {day: [(sort_idx, zeit_idx, l_idx, fallback_tag)]}
        for d in DAYS_DE:
            bucket = []
            for bk in bmap_by_day.get(d, ()):
                bf = bmap[bk]
                bucket.append((
                    col_to_idx.get(bf.get("sort", "")),
                    col_to_idx.get(bf.get("zeit", "")),
                    col_to_idx.get(bf.get("l")) if bf.get("l") else None,
                    bk[2],
                ))
            if bucket:
                bmap_idx[d] = bucket

        nbmap_idx: dict = {}     # {day: [(sort_idx, zeit_idx, tag_idx)]}
        for d in DAYS_DE:
            bucket = []
            for nbk in nbmap_by_day.get(d, ()):
                nbf = nbmap[nbk]
                bucket.append((
                    col_to_idx.get(nbf.get("sort", "")),
                    col_to_idx.get(nbf.get("zeit", "")),
                    col_to_idx.get(nbf.get("tag", "")),
                ))
            if bucket:
                nbmap_idx[d] = bucket

        ds_trip_idx: dict = {}   # {day: [(sort_idx, zeit_idx, tag_idx)]}
        for d in DAYS_DE:
            if d in ds_trip:
                bucket = []
                for k_ds, f_cols in ds_trip[d].items():
                    bucket.append((
                        col_to_idx.get(f_cols.get("Sort")),
                        col_to_idx.get(f_cols.get("Zeit")),
                        col_to_idx.get(f_cols.get("Tag")),
                    ))
                ds_trip_idx[d] = bucket

        # Lokale Refs sparen Attribute-Lookups in der Hot-Loop
        _norm_val = norm_val
        _safe_time = safe_time
        _canon_group_id = canon_group_id
        _SORT_PRIO = SORT_PRIO

        def _gv(row, idx):
            """Sicherer Zugriff auf Tuple-Index, gibt '' zurueck wenn idx None."""
            if idx is None:
                return ""
            try:
                return row[idx]
            except IndexError:
                return ""

        for row in df.itertuples(index=False, name=None):
            knr = _norm_val(_gv(row, idx_csb))
            if not knr: continue
            bestell: list = []
            for d_de in DAYS_DE:
                day_items: list = []
                if d_de in trip_idx:
                    for gt, si, zi, ti in trip_idx[d_de]:
                        s   = _norm_val(_gv(row, si))
                        t   = _safe_time(_gv(row, zi))
                        tag = _norm_val(_gv(row, ti))
                        if s or t or tag:
                            day_items.append({
                                "liefertag": d_de, "sortiment": s,
                                "bestelltag": tag, "bestellschluss": t,
                                "prio": _SORT_PRIO.get(_canon_group_id(s), 50),
                            })
                # Neues Format: Montag_Zeit / Montag_Sort / Montag_Tag
                if d_de in neue_idx:
                    for si, zi, ti in neue_idx[d_de]:
                        s   = _norm_val(_gv(row, si)) if si is not None else ""
                        t   = _safe_time(_gv(row, zi)) if zi is not None else ""
                        tag = _norm_val(_gv(row, ti)) if ti is not None else ""
                        if s or t or tag:
                            day_items.append({
                                "liefertag": d_de, "sortiment": s,
                                "bestelltag": tag, "bestellschluss": t,
                                "prio": _SORT_PRIO.get(_canon_group_id(s), 50),
                            })
                if d_de in bmap_idx:
                    for si, zi, li, fallback_tag in bmap_idx[d_de]:
                        s = _norm_val(_gv(row, si))
                        z = _safe_time(_gv(row, zi))
                        tag = _norm_val(_gv(row, li)) if li is not None else fallback_tag
                        if not tag: tag = fallback_tag
                        if s or z:
                            day_items.append({
                                "liefertag": d_de, "sortiment": s,
                                "bestelltag": tag, "bestellschluss": z,
                                "prio": _SORT_PRIO.get(_canon_group_id(s), 50),
                            })
                # Neues B-Format: "Die 1001 B_Zeit" / "Die 1001 B_Sortiment" / "Die 1001 B_Tag"
                if d_de in nbmap_idx:
                    for si, zi, ti in nbmap_idx[d_de]:
                        s   = _norm_val(_gv(row, si))
                        z   = _safe_time(_gv(row, zi))
                        tag = _norm_val(_gv(row, ti))
                        if s or z:
                            day_items.append({
                                "liefertag": d_de, "sortiment": s,
                                "bestelltag": tag, "bestellschluss": z,
                                "prio": _SORT_PRIO.get(_canon_group_id(s), 50),
                            })
                if d_de in ds_trip_idx:
                    for si, zi, ti in ds_trip_idx[d_de]:
                        s   = _norm_val(_gv(row, si))
                        t   = _safe_time(_gv(row, zi))
                        tag = _norm_val(_gv(row, ti))
                        if s or t or tag:
                            day_items.append({
                                "liefertag": d_de, "sortiment": s,
                                "bestelltag": tag, "bestellschluss": t,
                                "prio": 5.5,
                            })
                day_items.sort(key=lambda x: x["prio"])
                bestell.extend(day_items)

            data[knr] = {
                "plan_typ":    "",
                "bereich":     BEREICH,
                "kunden_nr":   knr,
                "sap_nummer":  _norm_val(_gv(row, idx_sap)),
                "name":        _norm_val(_gv(row, idx_name)),
                "strasse":     _norm_val(_gv(row, idx_str)),
                "plz":         _norm_val(_gv(row, idx_plz)),
                "ort":         _norm_val(_gv(row, idx_ort)),
                "fachberater": _norm_val(_gv(row, idx_fach)),
                "tours":       {d: _norm_val(_gv(row, idx_tour[d])) for d in DAYS_DE},
                "bestell":     bestell,
            }

        all_data[area_key] = data
        st.success(f"✓ {sheet_name}: {len(data)} Kunden verarbeitet")

    # Ladefolge aus Blatt T-B-Druck Quelle — gecached, vermeidet zweite Berechnung
    ladefolge_map: dict = {}
    try:
        ladefolge_map = cached_winter_map(excel_bytes)
    except Exception:
        pass

    json_data      = json.dumps(all_data,      ensure_ascii=False, separators=(",", ":"))
    ladefolge_json = json.dumps(ladefolge_map, ensure_ascii=False, separators=(",", ":"))

    # Notizen aus Lieferhinweis-CSV (gecached)
    notizen_map: dict = cached_lieferhinweis_map(read_upload_bytes(lieferhinweis_csv))
    notizen_json = json.dumps(notizen_map, ensure_ascii=False, separators=(",", ":"))

    return (
        get_druck_template()
        .replace("__DATA_JSON__",      json_data)
        .replace("__LADEFOLGE_JSON__", ladefolge_json)
        .replace("__KUNDEN_NOTIZEN_JSON__", notizen_json)
        .replace("__LOGO_DATAURI__",   logo_uri or "")
    )


# =============================================================================
# HTML KOMBINIEREN  →  app.html
# =============================================================================

# Die Sa-/So-Auswertung wird direkt aus TIMEREC_DATA aufgebaut.

def parse_fahrer_excel(dateien: list) -> str:
    """Verarbeitet mehrere Touren-Excel-Dateien → JSON [{name, years:{year:{krank,urlaub,ausgleich,arbeit,arbeit_samstag,touren,eintraege:[]}}}]"""
    import json as _json
    from io import BytesIO
    import datetime as _dt
    import pandas as _pd
    import re as _re

    WOCHENTAGE = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"]

    def ist_ausgeschlossen(name):
        return is_excluded_driver(name)

    def fmt_zeit(val):
        try:
            if _pd.isna(val): return "n.A."
        except: pass
        if isinstance(val, str):
            val = val.strip()
            if val in ["0:00","00:00","00:00:00"]: return "00:00"
            if ":" in val:
                p = val.split(":")
                if len(p) >= 2:
                    try: return f"{int(p[0]):02d}:{int(p[1]):02d}"
                    except: pass
        elif isinstance(val, float) and val > 0:
            h = int(val * 24); m = int((val * 1440) % 60)
            return f"{h:02d}:{m:02d}"
        elif isinstance(val, (_dt.datetime, _pd.Timestamp)):
            return val.strftime("%H:%M")
        elif isinstance(val, _dt.time):
            return val.strftime("%H:%M")
        return "n.A."

    def fmt_tour(val):
        """Tournummer / Tourtext robust aus Excel übernehmen.
        Auch Texte wie 'z.b.v. / Sonderaufgaben' werden unverändert als Tour angezeigt.
        """
        try:
            if _pd.isna(val):
                return ""
        except Exception:
            pass
        txt = str(val).strip()
        if txt.lower() in ("nan", "none", "nat"):
            return ""
        # Excel-Zellumbrüche und doppelte Leerzeichen sauber im HTML anzeigen
        txt = _re.sub(r"\s+", " ", txt)
        return txt

    # fahrer_map: name → { year → { eintraege:[], krank, urlaub, ausgleich } }
    fahrer_map = {}

    for datei in dateien:
        try:
            datei.seek(0)
            df = _pd.read_excel(BytesIO(datei.read()), sheet_name="Touren", header=None)
            df = df.iloc[5:].reset_index(drop=True)
        except Exception:
            continue

        current_year = _dt.datetime.now().year
        for row in df.itertuples(index=False, name=None):
            datum_raw = row[14] if len(row) > 14 else None
            datum = _pd.to_datetime(datum_raw, errors="coerce")
            if _pd.isna(datum):
                continue

            kw = int(datum.strftime("%W")) + 1  # Sonntag-Start wie original
            jahr = datum.year
            if datum.month == 1 and kw >= 52:
                jahr -= 1
            if datum.year != current_year:
                continue

            wd = WOCHENTAGE[datum.weekday()]
            datum_str = f"{wd}, {datum.strftime('%d.%m.%Y')}"
            uhrzeit = fmt_zeit(row[8] if len(row) > 8 else None)
            tour = fmt_tour(row[15] if len(row) > 15 else None)
            # Sonderfälle: Bei z.b.v. / Sonderaufgaben steht der Tourtext in manchen
            # Exporten nicht in der normalen Tournummer-Spalte P. Dann gezielt
            # nach diesem Text in den nahen Tour-/Bemerkungsspalten suchen.
            if not tour:
                for _idx in (0, 1, 2, 12, 13, 15, 16, 17, 18):
                    if len(row) <= _idx:
                        continue
                    _cand = fmt_tour(row[_idx])
                    if _cand and _re.search(r"z\.?\s*b\.?\s*v|sonderaufgaben|sonder", _cand, _re.I):
                        tour = _cand
                        break
            lkw_raw = row[11] if len(row) > 11 and _pd.notna(row[11]) else ""
            lkw_str = str(lkw_raw).strip()
            lkw = str(int(float(lkw_str))) if lkw_str.replace(".", "").replace("-", "").isdigit() else (lkw_str if lkw_str not in ("nan", "None", "") else "")

            paare = []

            def _cell_text(v):
                try:
                    if _pd.isna(v):
                        return ""
                except Exception:
                    pass
                s = str(v).strip()
                if s.lower() in ("nan", "none", "nat"):
                    return ""
                return _re.sub(r"\s+", " ", s)

            def _add_pair(nachname, vorname):
                nachname = _cell_text(nachname)
                vorname = _cell_text(vorname)
                if not nachname or not vorname:
                    return
                key = (nachname.lower(), vorname.lower())
                if key not in {(a.lower(), b.lower()) for a, b in paare}:
                    paare.append((nachname, vorname))

            def _add_full_name(value):
                full = _cell_text(value)
                if not full:
                    return
                if _re.search(r"z\.?\s*b\.?\s*v|sonderaufgaben|sonder|krank|urlaub|ausgleich", full, _re.I):
                    return
                # Standard: "Nachname, Vorname"
                if "," in full:
                    parts = [p.strip() for p in full.split(",") if p.strip()]
                    if len(parts) >= 2:
                        _add_pair(parts[0], " ".join(parts[1:]))
                    return
                # Fallback für eine einzelne Namenszelle: beide üblichen Reihenfolgen
                # zulassen; der spätere JavaScript-Abgleich ist ebenfalls robust.
                parts = full.split()
                if len(parts) >= 2 and not any(ch.isdigit() for ch in full):
                    _add_pair(parts[-1], " ".join(parts[:-1]))
                    _add_pair(parts[0], " ".join(parts[1:]))

            # Normale Exportstruktur: D/E und G/H sind Nachname/Vorname.
            if len(row) > 4:
                _add_pair(row[3], row[4])
            if len(row) > 7:
                _add_pair(row[6], row[7])
            # Robuster Fallback: z.b.v.-Zeilen haben den Fahrer manchmal in nur einer Zelle.
            for _idx in (3, 4, 6, 7):
                if len(row) > _idx:
                    _add_full_name(row[_idx])

            # Falls die Tourenzeile keinen Fahrer enthält, bleibt sie trotzdem als
            # Planungs-Eintrag erhalten. Die 10H-Zuordnung kann sie dann über
            # Anfangsdatum, LKW und Startzeit finden.
            if not paare and (tour or lkw or uhrzeit != "n.A."):
                paare.append(("Unzugeordnet", "Planung"))

            for nachname, vorname in paare:
                if not nachname or not vorname:
                    continue
                if ist_ausgeschlossen(nachname):
                    continue
                if nachname in ("0", "nan") or vorname in ("0", "nan"):
                    continue
                name = f"{nachname}, {vorname}"
                yr = str(jahr)
                if name not in fahrer_map:
                    fahrer_map[name] = {}
                if yr not in fahrer_map[name]:
                    fahrer_map[name][yr] = {"eintraege": []}
                fahrer_map[name][yr]["eintraege"].append({
                    "kw": kw,
                    "datum": datum_str,
                    "tour": tour,
                    "zeit": uhrzeit,
                    "lkw": lkw,
                    "samstag": wd == "Samstag" or (wd == "Freitag" and uhrzeit >= "18:00" and uhrzeit != "n.A."),
                })

    # Aggregate
    result = []
    for name, years in sorted(fahrer_map.items()):
        years_out = {}
        for yr, data in sorted(years.items()):
            eintr = data["eintraege"]
            tour_cnt = {}
            for e in eintr:
                t = e["tour"].lower()
                if "krank" not in t and "urlaub" not in t and "ausgleich" not in t and e["tour"]:
                    tour_cnt[e["tour"]] = tour_cnt.get(e["tour"], 0) + 1
            lkw_cnt = {}
            for e in eintr:
                lv = e.get("lkw","").strip()
                if lv and lv not in ("nan","None","","0") and not e["tour"].lower().strip() in ("ausgleich",) and not any(k in e["tour"].lower() for k in ["krank","urlaub","ausgleich"]):
                    lkw_cnt[lv] = lkw_cnt.get(lv, 0) + 1
            years_out[yr] = {
                "krank":          sum(1 for e in eintr if "krank"     in e["tour"].lower()),
                "urlaub":         sum(1 for e in eintr if "urlaub"    in e["tour"].lower()),
                "ausgleich":      sum(1 for e in eintr if "ausgleich" in e["tour"].lower()),
                "arbeit":         sum(1 for e in eintr if e["zeit"] != "n.A."),
                "arbeit_samstag": sum(1 for e in eintr if e["samstag"] and e["zeit"] != "n.A."),
                "touren":         dict(sorted(tour_cnt.items(), key=lambda x: (int(x[0]) if x[0].isdigit() else 9999, x[0]))),
                "lkw":            dict(sorted(lkw_cnt.items(), key=lambda x: -x[1])),
                "eintraege":      sorted(eintr, key=lambda x: (x["kw"],)),
            }
        result.append({"name": name, "years": years_out})
    return _json.dumps(result, ensure_ascii=False)


# =============================================================================
# SPEDITEURE — Auswertung der Tourenpläne nach externen Speditionen.
# Sucht in den Touren-Excel die Spediteur-Namen (inkl. Abwandlungen) und
# wertet pro Spedition nach Jahr/Monat/Untername aus: was wurde wann mit welcher
# Tournummer gefahren.
# =============================================================================
SPEDITEUR_KATALOG = [
    ("8001", "Spedition Meyer 1"), ("8002", "Spedition Meyer 2 (36er)"), ("8003", "Spedition Meyer 3"),
    ("8004", "Spedition Meyer 4"), ("8005", "Spedition Meyer 5"), ("8006", "Spedition Meyer 6"),
    ("8007", "Spedition Meyer 7"), ("8008", "Spedition Meyer 8"), ("8009", "Spedition Meyer SZ"),
    ("8010", "Spedition Meyer 9"), ("80101", "Spedition Meyer 10"), ("80102", "Spedition Meyer 11"),
    ("8011", "Paasch & Reinke 1"), ("8012", "Paasch & Reinke 2"), ("8013", "Paasch & Reinke 3"),
    ("8015", "Ch. Holtz T1"), ("8016", "Ch. Holtz T2"), ("8017", "Ch. Holtz T3"),
    ("8019", "deVries - 1"), ("8020", "deVries - 2"), ("8022", "Spedition Ihde"),
    ("8024", "Zippel Logistik T2"), ("8025", "Zippel Logistik T1"), ("8026", "Zippel Logistik T3"),
    ("8027", "Zippel Logistik T4"), ("8029", "Zippel Logistik T5"),
    ("8030", "Insellogistik 1"), ("8031", "Insellogistik 2"), ("8032", "Insellogistik 3"), ("8033", "Insellogistik 4"),
    ("8034", "T & D"), ("8035", "Sped. Maas"), ("8036", "Nordfrost"), ("8037", "Emons"), ("8038", "Thermotraffic"),
    ("8039", "Kudex 1"), ("8040", "Kudex 2"), ("8045", "Kudex 3"), ("8046", "Kudex 4"),
    ("8041", "Pfenning 1"), ("8042", "Pfenning 2"), ("8043", "Pfenning 3"), ("8044", "Pfenning 4"),
]
_SPED_GENERIC = {"spedition", "sped", "logistik", "gmbh", "kg", "co", "transporte", "transport", "cocg", "und"}
_SPED_COMPANY_ALIASES = {"paasch": "paaschreinke", "reinke": "paaschreinke"}

# Oberkategorie (Spedition) je Katalog-Nummer. Die Katalog-Namen sind Unterkategorien.
_SPED_GRUPPEN = {
    "Spedition Meyer": {"8001","8002","8003","8004","8005","8006","8007","8008","8009","8010","80101","80102"},
    "Paasch & Reinke": {"8011","8012","8013"},
    "Ch. Holtz":       {"8015","8016","8017"},
    "deVries":         {"8019","8020"},
    "Spedition Ihde":  {"8022"},
    "Zippel Logistik": {"8024","8025","8026","8027","8029"},
    "Insellogistik":   {"8030","8031","8032","8033"},
    "Kudex":           {"8039","8040","8045","8046"},
    "Pfenning":        {"8041","8042","8043","8044"},
}
_SPED_NR_GRUPPE = {nr: g for g, nrs in _SPED_GRUPPEN.items() for nr in nrs}


def _sped_gruppe(nr: str, name: str) -> str:
    """Liefert die Ober-Spedition zu einer Katalog-Nummer.
    Einzeleinträge ohne Familie (T&D, Maas, Nordfrost, ...) bilden ihre eigene Gruppe."""
    return _SPED_NR_GRUPPE.get(str(nr), name)


def _sped_tokens(text: str) -> list:
    s = str(text or "").lower()
    s = s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\([^)]*\)", " ", s)   # Zusätze wie "(36er)" entfernen
    toks = [t for t in re.split(r"[^a-z0-9]+", s) if t]
    return [t for t in toks if t not in _SPED_GENERIC]


def _sped_parse(text: str):
    """Zerlegt einen Namen in (compact, company, variant).

    variant ist die Tour-Kennung am Ende (z.B. '1', '10', 't2', 'sz')."""
    toks = _sped_tokens(text)
    if not toks:
        return "", "", ""
    if re.fullmatch(r"(t\d+|sz|\d+)", toks[-1]):
        variant = toks[-1]
        comp_toks = toks[:-1]
    else:
        variant = ""
        comp_toks = toks[:]
    company = "".join(comp_toks)
    company = _SPED_COMPANY_ALIASES.get(company, company)
    return company + variant, company, variant


def _sped_build_index():
    by_compact, by_company, company_disp = {}, {}, {}
    for nr, name in SPEDITEUR_KATALOG:
        compact, company, variant = _sped_parse(name)
        entry = {"nr": nr, "name": name, "company": company, "variant": variant}
        by_compact[compact] = entry
        by_company.setdefault(company, []).append(entry)
        disp = re.sub(r"\s*[-–]?\s*(T\d+|SZ|\d+)\s*$", "", name, flags=re.I)
        disp = re.sub(r"\s*\([^)]*\)\s*", "", disp).strip()
        company_disp.setdefault(company, disp or name)
    return by_compact, by_company, company_disp


_SPED_BY_COMPACT, _SPED_BY_COMPANY, _SPED_COMPANY_DISP = _sped_build_index()
_SPED_BY_NR = {nr: {"nr": nr, "name": name, **dict(zip(("compact", "company", "variant"), _sped_parse(name)))}
               for nr, name in SPEDITEUR_KATALOG}


def _sped_clean_number(value) -> str:
    """Normiert Nummernzellen aus Excel, zum Beispiel 8035.0 -> 8035."""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = str(value or "").strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    return s if re.fullmatch(r"\d+", s) else ""


def _sped_has_marker(text: str) -> bool:
    """Erkennt, ob ein Text wirklich nach Spediteur/Firma aussieht.

    Wichtig gegen Fehlzuordnungen: Der Nachname 'Maas' allein darf nicht
    automatisch als 'Sped. Maas' gewertet werden.
    """
    s = str(text or "").lower()
    s = s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return bool(re.search(r"\b(sped|spedition|logistik|transport|transporte|gmbh|kg)\b", s))


def _sped_match(text: str, *, allow_bare_single_company: bool = False):
    # Exakte Katalog-Nummer hat Vorrang. Damit wird nicht über Namen geraten,
    # wenn in der Datei bereits eine eindeutige Spediteur-Nummer steht.
    nr = _sped_clean_number(text)
    if nr and nr in _SPED_BY_NR:
        e = _SPED_BY_NR[nr]
        return {"nr": e["nr"], "name": e["name"], "company": e["company"], "variant": e["variant"]}

    toks = _sped_tokens(text)
    # Einzelne blanke Nachnamen/Firmennamen sind zu unsicher: 'Maas' kann
    # ein Fahrer-Nachname sein. Erlaubt ist das nur, wenn die Zelle einen
    # Firmenmarker wie 'Sped.' enthält oder bewusst freigeschaltet wird.
    if len(toks) == 1 and not _sped_has_marker(text) and not allow_bare_single_company:
        return None

    compact, company, variant = _sped_parse(text)
    if not compact:
        return None
    if compact in _SPED_BY_COMPACT:
        return _SPED_BY_COMPACT[compact]
    ents = _SPED_BY_COMPANY.get(company)
    if ents:
        if variant:
            for e in ents:
                if e["variant"] == variant:
                    return e
            return None  # Firma bekannt, Variante unbekannt -> nicht raten
        if len(ents) == 1:
            return ents[0]
    return None


def _sped_match_pair(left, right):
    """Matcht die beiden Fahrer-/Namensspalten zusammen.

    Wenn links eine Nummer steht, ist diese bindend. Eine fremde Fahrer-Nummer
    darf nicht durch einen zufälligen Nachnamen rechts überschrieben werden
    (Beispiel: Michael Maas != Sped. Maas 8035).
    """
    ltxt = str(left or "").strip()
    rtxt = str(right or "").strip()
    lnr = _sped_clean_number(left)
    rnr = _sped_clean_number(right)

    if lnr:
        if lnr in _SPED_BY_NR:
            return _sped_match(lnr)
        # Nummer vorhanden, aber keine Spediteur-Katalognummer: nicht über
        # den Namen raten. Genau dadurch wurden Fahrer mit gleichem Nachnamen
        # bisher fälschlich als Spedition erkannt.
        return None
    if rnr:
        if rnr in _SPED_BY_NR:
            return _sped_match(rnr)
        return None

    # Erst die Paar-Zelle prüfen, damit 'Sped.' + 'Maas' sauber erkannt wird.
    combo = (ltxt + " " + rtxt).strip()
    if combo:
        ent = _sped_match(combo)
        if ent:
            return ent

    # Einzelzellen nur mit strenger Erkennung. Bare Ein-Wort-Namen wie 'Maas'
    # werden hier absichtlich nicht angenommen.
    for val in (ltxt, rtxt):
        ent = _sped_match(val)
        if ent:
            return ent
    return None


def parse_spediteure_excel(dateien: list) -> str:
    """Liest mehrere Touren-Excel und liefert JSON {katalog, fahrten}.

    Pro Treffer: Spediteur-Nr/Name, Jahr, Monat, Datum, Wochentag, Tour, LKW, Zeit.
    """
    import json as _json
    from io import BytesIO
    import datetime as _dt
    import pandas as _pd

    WOCHENTAGE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]

    def _zeit(val):
        try:
            if _pd.isna(val):
                return ""
        except Exception:
            pass
        if isinstance(val, str):
            v = val.strip()
            if ":" in v:
                p = v.split(":")
                try:
                    return f"{int(p[0]):02d}:{int(p[1]):02d}"
                except Exception:
                    return ""
            return ""
        if isinstance(val, float) and val > 0:
            total_min = int(round(val * 24 * 60))
            h = (total_min // 60) % 24
            m = total_min % 60
            return f"{h:02d}:{m:02d}"
        if isinstance(val, (_dt.datetime, _pd.Timestamp)):
            return val.strftime("%H:%M")
        if isinstance(val, _dt.time):
            return val.strftime("%H:%M")
        return ""

    def _tour(val):
        try:
            if _pd.isna(val):
                return ""
        except Exception:
            pass
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        t = str(val).strip()
        if t.lower() in ("nan", "none", "nat"):
            return ""
        if re.fullmatch(r"\d+\.0", t):
            t = t[:-2]
        return re.sub(r"\s+", " ", t)

    def _lkw(val):
        try:
            if _pd.isna(val):
                return ""
        except Exception:
            pass
        s = str(val).strip()
        if s.lower() in ("nan", "none", ""):
            return ""
        return str(int(float(s))) if s.replace(".", "").replace("-", "").isdigit() else s

    seen = set()
    fahrten = []
    used_nr = set()

    for datei in dateien:
        try:
            datei.seek(0)
            df = _pd.read_excel(BytesIO(datei.read()), sheet_name="Touren", header=None)
            df = df.iloc[5:].reset_index(drop=True)
        except Exception:
            continue
        quelle = getattr(datei, "name", "") or ""

        for row in df.itertuples(index=False, name=None):
            datum = _pd.to_datetime(row[14] if len(row) > 14 else None, errors="coerce")
            if _pd.isna(datum):
                continue

            # Spediteur-Erkennung aus den Fahrer-/Namenspaaren (D/E, G/H).
            # Die Nummernspalte ist bindend: Steht dort eine fremde Fahrer-Nummer,
            # darf ein gleicher Nachname nicht als Spedition gewertet werden.
            ent = None
            for a, b in ((3, 4), (6, 7)):
                left = row[a] if len(row) > a else ""
                right = row[b] if len(row) > b else ""
                ent = _sped_match_pair(left, right)
                if ent:
                    break
            if not ent:
                continue

            tour = _tour(row[15] if len(row) > 15 else "")
            lkw = _lkw(row[11] if len(row) > 11 else "")
            zeit = _zeit(row[8] if len(row) > 8 else None)
            iso = datum.strftime("%Y-%m-%d")
            key = (ent["nr"], iso, tour, lkw, zeit)
            if key in seen:
                continue
            seen.add(key)
            used_nr.add(ent["nr"])
            fahrten.append({
                "nr": ent["nr"],
                "name": ent["name"],
                "gruppe": _sped_gruppe(ent["nr"], ent["name"]),
                "jahr": str(datum.year),
                "monat": f"{datum.month:02d}",
                "datum": datum.strftime("%d.%m.%Y"),
                "iso": iso,
                "wd": WOCHENTAGE[datum.weekday()],
                "tour": tour,
                "lkw": lkw,
                "zeit": zeit,
                "quelle": quelle,
            })

    katalog = [{"nr": nr, "name": name, "gruppe": _sped_gruppe(nr, name)}
               for nr, name in SPEDITEUR_KATALOG if nr in used_nr]
    fahrten.sort(key=lambda x: (x["iso"], (int(x["nr"]) if x["nr"].isdigit() else 0)))
    return _json.dumps({"katalog": katalog, "fahrten": fahrten}, ensure_ascii=False)


def parse_versp_abfahrt_excel(dateien: list) -> str:
    """Liest die reguläre Abfahrtszeit je Tour direkt aus dem Tourenplan.

    Grundlage im Tourenplan:
      Spalte A (Index 0) = Tournummer
      Spalte I (Index 8) = Abfahrtszeit

    Die Verspätungstabelle braucht die Abfahrt zur Tournummer, nicht zur
    Kundenzeile. Deshalb wird ein flacher Index aufgebaut:
      {"__by_tour": {"1006": "20:00", ...}}

    Wichtig für die Tourenpläne:
    - Tage stehen oft als Blocküberschrift, nicht sauber je Zeile in einer Datumsspalte.
    - Abendtouren für Montag stehen im Sonntag-Block. Für die Verspätung wird daher
      zuerst exakt nach Tournummer gesucht, unabhängig vom Tagesblock.
    - Zeiten können als 20:00, 20.00, 7.00, 09:00, Excel-Zeitwert oder 0 stehen.
      0 beziehungsweise 0:00 ist eine gültige Uhrzeit und darf nicht als leer gelten.
    - Es werden nur echte Tourenplan-Blätter gelesen. Kundendaten-/Tourkundenlisten
      werden bewusst ignoriert, weil dort ebenfalls Tournummern und Uhrzeiten stehen
      und sonst Lieferzeiten statt Abfahrtszeiten übernommen würden.
    """
    import json as _json
    from io import BytesIO
    import datetime as _dt
    import re as _re
    import pandas as _pd
    from collections import Counter

    WOCHENTAGE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
    WD_BY_LOWER = {w.lower(): w for w in WOCHENTAGE}
    DAY_HEADER_RE = _re.compile(
        r"\b(Montag|Dienstag|Mittwoch|Donnerstag|Freitag|Samstag|Sonntag)\b\s*,?\s*\d{1,2}\.",
        _re.IGNORECASE,
    )

    def _is_empty(val) -> bool:
        try:
            if _pd.isna(val):
                return True
        except Exception:
            pass
        return str(val).strip().lower() in ("", "nan", "none", "nat")

    def _zeit(val):
        if _is_empty(val):
            return ""

        if isinstance(val, (_dt.datetime, _pd.Timestamp)):
            return val.strftime("%H:%M")
        if isinstance(val, _dt.time):
            return val.strftime("%H:%M")

        if isinstance(val, (int, float)):
            try:
                f = float(val)
            except Exception:
                return ""
            if f < 0:
                return ""
            # Excel speichert Uhrzeiten als Tagesbruchteil: 0,8333 = 20:00.
            # 0,0 ist 00:00 und bleibt gültig.
            if 0 <= f < 1:
                total_min = int(round(f * 24 * 60)) % (24 * 60)
                h = total_min // 60
                m = total_min % 60
                return f"{h:02d}:{m:02d}"
            # Zahl als Stunde, zum Beispiel 7 oder 15.
            if float(f).is_integer() and 0 <= int(f) <= 23:
                return f"{int(f):02d}:00"
            # Zahl als HHMM, zum Beispiel 700, 1530 oder 2000.
            iv = int(round(f))
            h = iv // 100
            m = iv % 100
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
            return ""

        v = str(val).strip()
        if not v:
            return ""
        low = v.lower().replace(" ", "")
        if low in ("n.a.", "n.a", "na", "nan", "none", "elternzeit"):
            return ""

        m = _re.match(r"^\s*(\d{1,2})\s*[:\.]\s*(\d{1,2})(?:\s*:\s*\d{1,2})?\s*$", v)
        if m:
            h = int(m.group(1))
            mi = int(m.group(2))
            if 0 <= h <= 23 and 0 <= mi <= 59:
                return f"{h:02d}:{mi:02d}"
            return ""

        digits = "".join(ch for ch in v if ch.isdigit())
        if len(digits) in (1, 2):
            h = int(digits)
            if 0 <= h <= 23:
                return f"{h:02d}:00"
        if len(digits) in (3, 4):
            h = int(digits[:-2])
            mi = int(digits[-2:])
            if 0 <= h <= 23 and 0 <= mi <= 59:
                return f"{h:02d}:{mi:02d}"
        return ""

    def _tour_code(val):
        """Nur echte Tournummern übernehmen, keine Zahlen aus Namen/Freitext ziehen."""
        if _is_empty(val):
            return ""
        if isinstance(val, (int, float)):
            try:
                f = float(val)
            except Exception:
                return ""
            if f.is_integer() and f >= 0:
                return str(int(f))
            return ""
        s = str(val).strip()
        if _re.fullmatch(r"\d+", s):
            return s
        if _re.fullmatch(r"\d+\.0+", s):
            return s.split(".", 1)[0]
        return ""

    def _tour_code_from_text(val):
        """Nur für Sonderfälle wie 'NMS-Shuttle Tour 5045' als Fallback."""
        if _is_empty(val):
            return ""
        s = str(val).strip()
        m = _re.search(r"\bTour\s+(\d{3,6})\b", s, flags=_re.IGNORECASE)
        if m:
            return m.group(1)
        return ""

    def _weekday_from_header_row(row) -> str:
        """Erkennt den Tagesblock im Tourenplan.

        Wichtig: In echten Excel-Tourenplänen ist die Tagesüberschrift oft kein
        Text wie "Montag, 1. Juni 2026", sondern ein echter Datumswert mit
        Zellformat. Beim Lesen mit pandas/calamine kommt dann zum Beispiel
        Timestamp(2026-06-01) an. Genau deshalb waren zuletzt alle Abfahrten leer:
        das Blatt wurde als "ohne Tagesblock" verworfen.
        """
        cells = list(row[: min(len(row), 10)])

        # 1) Echte Excel-/pandas-Datumswerte erkennen.
        for x in cells:
            if _is_empty(x):
                continue
            try:
                if isinstance(x, (_pd.Timestamp, _dt.datetime)):
                    if getattr(x, "year", 0) >= 2020:
                        return WOCHENTAGE[int(x.weekday())]
                elif isinstance(x, _dt.date):
                    if getattr(x, "year", 0) >= 2020:
                        return WOCHENTAGE[int(x.weekday())]
            except Exception:
                pass

            # 2) Excel-Seriendatum als Zahl, falls der Reader nicht in Datum wandelt.
            try:
                if isinstance(x, (int, float)) and 40000 <= float(x) <= 60000:
                    d = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(x))
                    if d.year >= 2020:
                        return WOCHENTAGE[int(d.weekday())]
            except Exception:
                pass

        joined = " ".join(str(x).strip() for x in cells if not _is_empty(x))

        # 3) Anzeigeformat mit ausgeschriebenem Wochentag.
        m = DAY_HEADER_RE.search(joined)
        if m:
            return WD_BY_LOWER.get(m.group(1).lower(), "")

        # 4) Datums-Text ohne Wochentag, zum Beispiel 01.06.2026.
        m = _re.search(r"\b(\d{1,2})[\./-](\d{1,2})[\./-](\d{2,4})\b", joined)
        if m:
            try:
                dd = int(m.group(1)); mm = int(m.group(2)); yy = int(m.group(3))
                if yy < 100:
                    yy += 2000
                if yy >= 2020:
                    d = _dt.date(yy, mm, dd)
                    return WOCHENTAGE[int(d.weekday())]
            except Exception:
                pass

        # 5) Deutscher Datums-Text ohne Wochentag, zum Beispiel 1. Juni 2026.
        month_map = {
            "januar": 1, "februar": 2, "märz": 3, "maerz": 3, "april": 4,
            "mai": 5, "juni": 6, "juli": 7, "august": 8, "september": 9,
            "oktober": 10, "november": 11, "dezember": 12,
        }
        m = _re.search(r"\b(\d{1,2})\.\s*(Januar|Februar|März|Maerz|April|Mai|Juni|Juli|August|September|Oktober|November|Dezember)\s+(\d{4})\b", joined, _re.IGNORECASE)
        if m:
            try:
                dd = int(m.group(1))
                mm = month_map.get(m.group(2).lower(), 0)
                yy = int(m.group(3))
                if mm and yy >= 2020:
                    d = _dt.date(yy, mm, dd)
                    return WOCHENTAGE[int(d.weekday())]
            except Exception:
                pass

        return ""

    def _read_sheets(raw: bytes):
        """Liest bevorzugt das Blatt 'Touren', fällt aber robust auf alle Blätter zurück."""
        if not raw:
            return []
        frames = []
        try:
            xl = _pd.ExcelFile(BytesIO(raw), engine=EXCEL_READ_ENGINE)
            names = list(xl.sheet_names or [])
        except Exception:
            try:
                xl = _pd.ExcelFile(BytesIO(raw))
                names = list(xl.sheet_names or [])
            except Exception:
                return []

        # Erst Touren-Blätter, danach alle übrigen. So funktioniert es auch,
        # wenn das Blatt anders heißt oder mehrere Tourenpläne enthalten sind.
        ordered = []
        for n in names:
            if str(n).strip().lower() == "touren":
                ordered.append(n)
        for n in names:
            if n not in ordered:
                ordered.append(n)

        for sheet in ordered:
            try:
                df = _pd.read_excel(BytesIO(raw), sheet_name=sheet, header=None, engine=EXCEL_READ_ENGINE)
            except Exception:
                try:
                    df = _pd.read_excel(BytesIO(raw), sheet_name=sheet, header=None)
                except Exception:
                    continue
            if df is not None and not df.empty:
                frames.append((str(sheet), df))
        return frames

    def _sheet_has_day_headers(df) -> bool:
        """Echte Tourenpläne haben Tagesüberschriften wie 'Montag, 1. Juni 2026'.

        Kundendaten-/Tourkundenlisten können ebenfalls in Spalte A Tournummern und
        in Spalte I Uhrzeiten enthalten. Diese Listen dürfen für die Verspätungstabelle
        nicht als Quelle dienen, weil das Liefer-/Kundenzeiten sind und keine
        Abfahrtszeiten. Deshalb akzeptieren wir nur Blätter mit Tagesblöcken.
        """
        hits = 0
        try:
            rows_iter = df.itertuples(index=False, name=None)
            for i, row in enumerate(rows_iter):
                if i > 500:
                    break
                if _weekday_from_header_row(row):
                    hits += 1
                    if hits >= 1:
                        return True
        except Exception:
            return False
        return False

    # {wochentag: {tour: Counter(zeiten)}} und flacher Index tour -> Counter(zeiten)
    agg: dict = {}
    flat_agg: dict = {}
    source_rows = 0
    skipped_sheets = 0

    for datei in dateien or []:
        try:
            datei.seek(0)
        except Exception:
            pass
        try:
            raw = datei.read()
        except Exception:
            raw = b""
        if isinstance(raw, str):
            raw = raw.encode("utf-8", errors="ignore")

        for _sheet_name, df in _read_sheets(raw):
            if not _sheet_has_day_headers(df):
                skipped_sheets += 1
                continue

            current_wd = ""
            # nicht pauschal Zeilen wegschneiden; Überschriften werden sowieso übersprungen.
            for row in df.itertuples(index=False, name=None):
                header_wd = _weekday_from_header_row(row)
                if header_wd:
                    current_wd = header_wd
                    continue

                # Vor der ersten Tagesüberschrift nichts übernehmen. So vermeiden wir,
                # dass Kopfzeilen oder fremde Tabellen als Tourenplan gewertet werden.
                if not current_wd:
                    continue

                tour = _tour_code(row[0] if len(row) > 0 else "")
                # Sonderfall: wenn in A ein Strich steht, aber in B 'NMS-Shuttle Tour 5045'.
                if not tour:
                    tour = _tour_code_from_text(row[1] if len(row) > 1 else "")
                if not tour:
                    continue

                zeit = _zeit(row[8] if len(row) > 8 else None)
                if not zeit:
                    continue

                source_rows += 1
                flat_agg.setdefault(tour, Counter())[zeit] += 1
                agg.setdefault(current_wd, {}).setdefault(tour, Counter())[zeit] += 1

    def _best(counter):
        if not counter:
            return ""
        # häufigste Zeit; bei Gleichstand früheste Uhrzeit
        best = sorted(counter.items(), key=lambda kv: (-kv[1], kv[0]))
        return best[0][0] if best else ""

    out: dict = {"__by_tour": {}}
    for tour, counter in flat_agg.items():
        z = _best(counter)
        if z:
            out["__by_tour"][tour] = z

    for wd, tours in agg.items():
        out[wd] = {}
        for tour, counter in tours.items():
            z = _best(counter)
            if z:
                out[wd][tour] = z

    # Diagnose im JSON, damit man im Browser/Quelltext sofort sieht, ob etwas geladen wurde.
    out["__meta"] = {
        "touren_mit_abfahrt": len(out.get("__by_tour", {})),
        "gelesene_zeilen": source_rows,
        "ignorierte_blaetter_ohne_tagesblock": skipped_sheets,
    }
    return _json.dumps(out, ensure_ascii=False)


def parse_versp_abfahrt_csv(uploaded_file) -> str:
    """Liest Tourenstart-Zeiten fuer die Verspaetungstabelle aus einer CSV.

    Standard ohne Kopfzeile:
        Tournummer;Uhrzeit
        1005;1:00
        1006;20:00

    Optional mit Kopfzeile:
        Wochentag;Tour;Uhrzeit
        Montag;1005;01:00
    """
    import csv as _csv
    import json as _json
    import datetime as _dt
    import re as _re
    from io import StringIO

    raw = read_upload_bytes(uploaded_file)
    if isinstance(raw, str):
        raw = raw.encode("utf-8", errors="ignore")
    if not raw:
        return "{}"

    text = ""
    for enc in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            pass
    if not text:
        text = raw.decode("utf-8", errors="replace")

    def _is_empty(v) -> bool:
        return str(v or "").strip().lower() in ("", "nan", "none", "nat")

    def _zeit(val):
        if _is_empty(val):
            return ""
        if isinstance(val, (_dt.datetime, _dt.time)):
            return val.strftime("%H:%M")
        v = str(val).strip()
        low = v.lower().replace(" ", "")
        if low in ("n.a.", "n.a", "na", "nan", "none", "elternzeit"):
            return ""
        m = _re.match(r"^\s*(\d{1,2})\s*[:\.]\s*(\d{1,2})(?:\s*:\s*\d{1,2})?\s*$", v)
        if m:
            h = int(m.group(1)); mi = int(m.group(2))
            if 0 <= h <= 23 and 0 <= mi <= 59:
                return f"{h:02d}:{mi:02d}"
            return ""
        digits = "".join(ch for ch in v if ch.isdigit())
        if len(digits) in (1, 2):
            h = int(digits)
            if 0 <= h <= 23:
                return f"{h:02d}:00"
        if len(digits) in (3, 4):
            h = int(digits[:-2]); mi = int(digits[-2:])
            if 0 <= h <= 23 and 0 <= mi <= 59:
                return f"{h:02d}:{mi:02d}"
        return ""

    WD = {
        "mo":"Montag", "montag":"Montag",
        "di":"Dienstag", "dienstag":"Dienstag",
        "mi":"Mittwoch", "mittwoch":"Mittwoch",
        "do":"Donnerstag", "donnerstag":"Donnerstag",
        "fr":"Freitag", "freitag":"Freitag",
        "sa":"Samstag", "samstag":"Samstag",
        "so":"Sonntag", "sonntag":"Sonntag",
    }

    def _day(v):
        k = str(v or "").strip().lower().rstrip(".")
        return WD.get(k, "")

    def _tour(v):
        s = str(v or "").strip().replace("\ufeff", "")
        if not s or s.lower() in ("nan", "none"):
            return ""
        if _re.fullmatch(r"\d+\.0+", s):
            s = s.split(".", 1)[0]
        return _re.sub(r"\s+", " ", s)

    sample = "\n".join(text.splitlines()[:20])
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except Exception:
        class _Dialect(_csv.excel):
            delimiter = ";" if text.count(";") >= max(text.count(","), text.count("\t")) else ","
        dialect = _Dialect

    rows = []
    for row in _csv.reader(StringIO(text), dialect):
        if not row or all(_is_empty(c) for c in row):
            continue
        rows.append([str(c).strip() for c in row])
    if not rows:
        return "{}"

    first = [c.lower().strip().replace("ä", "ae") for c in rows[0]]
    has_header = any(("tour" in c or "uhr" in c or "zeit" in c or "abfahrt" in c or "wochentag" in c or c == "tag") for c in first)

    tour_col = 0
    time_col = 1 if len(rows[0]) > 1 else 0
    day_col = None
    start_idx = 0
    if has_header:
        start_idx = 1
        for idx, c in enumerate(first):
            if day_col is None and ("wochentag" in c or c == "tag" or c.endswith("tag")):
                day_col = idx
            if "tour" in c:
                tour_col = idx
            if "uhr" in c or "start" in c or "abfahrt" in c or c in ("zeit", "uhrzeit"):
                time_col = idx
    elif len(rows[0]) >= 3 and _day(rows[0][0]):
        day_col = 0
        tour_col = 1
        time_col = 2

    out = {"__by_tour": {}}
    loaded = 0
    skipped = 0
    duplicates = 0
    for row in rows[start_idx:]:
        if len(row) <= max(tour_col, time_col):
            skipped += 1
            continue
        tour = _tour(row[tour_col])
        zeit = _zeit(row[time_col])
        if not tour or not zeit:
            skipped += 1
            continue
        key = tour if not _re.fullmatch(r"\d+\.0+", tour) else tour.split(".", 1)[0]
        if key in out["__by_tour"]:
            if out["__by_tour"][key] != zeit:
                duplicates += 1
            # Bei mehrfacher Tour bleibt bewusst die erste Zeile der CSV gültig.
            # So kann die CSV-Reihenfolge steuern, welche Startzeit für Sammel-/HUPA-Touren gilt.
        else:
            out["__by_tour"][key] = zeit
        digits = "".join(ch for ch in key if ch.isdigit())
        if digits and digits != key and digits not in out["__by_tour"]:
            out["__by_tour"][digits] = zeit
        if day_col is not None and len(row) > day_col:
            wd = _day(row[day_col])
            if wd:
                bucket = out.setdefault(wd, {})
                if key not in bucket:
                    bucket[key] = zeit
                if digits and digits != key and digits not in bucket:
                    bucket[digits] = zeit
        loaded += 1

    out["__meta"] = {
        "quelle": getattr(uploaded_file, "name", "") or "Tourenstart CSV",
        "touren_mit_abfahrt": len(out.get("__by_tour", {})),
        "gelesene_zeilen": loaded,
        "uebersprungene_zeilen": skipped,
        "abweichende_dubletten": duplicates,
    }
    return _json.dumps(out, ensure_ascii=False)

def parse_timerecording_csv(uploaded_file) -> str:
    """Liest die Tachograph-Schicht-Datei (CSV oder XLSX) und liefert JSON
    pro Fahrer.

    Unterstützt sowohl das neue YellowFox-Format mit getrennten Spalten
    ``Datum``/``Beginn``/``Ende`` als auch das ältere Format mit kombinierten
    Spalten ``Schichtbeginn``/``Schichtende``.

    YellowFox/timerecording_v3 ist eine Tagesaggregation. Deshalb können in
    ``Beginn`` und ``Ende`` mehrere Uhrzeiten stehen. Außerdem erzeugt YellowFox
    bei nicht mehr ausgelesenen Fahrerkarten technische Platzhalter wie
    ``00:00`` bis ``24:00`` ohne Arbeits-, Lenk- oder Schichtzeit. Diese Zeilen
    sind keine Schichten und werden hier ausdrücklich verworfen.
    """
    import json as _json
    import csv as _csv
    import datetime as _dt
    from io import StringIO, BytesIO

    raw = read_upload_bytes(uploaded_file)
    if not raw:
        return "{}"

    is_xlsx = raw[:4] == b'PK\x03\x04'
    rows = []
    header_raw = []

    if is_xlsx:
        try:
            import pandas as _pd
            df = _pd.read_excel(BytesIO(raw), header=0, dtype=str)
            header_raw = [str(c) for c in df.columns]
            for _, r in df.iterrows():
                rows.append([str(r[c]) if _pd.notna(r[c]) else "" for c in df.columns])
        except Exception:
            return "{}"
    else:
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return "{}"
        reader = _csv.reader(StringIO(text), delimiter=";", quotechar='"')
        all_rows = list(reader)
        if not all_rows or len(all_rows) < 2:
            return "{}"
        header_raw = [h.strip() for h in all_rows[0]]
        rows = all_rows[1:]

    if not rows:
        return "{}"

    header = [h.lower().strip() for h in header_raw]

    def col(name_variants):
        for v in name_variants:
            v = v.lower().strip()
            for i, h in enumerate(header):
                if v in h:
                    return i
        return -1

    def col_exact(name_variants):
        wanted = {str(v).lower().strip() for v in name_variants}
        for i, h in enumerate(header):
            if h in wanted:
                return i
        return -1

    idx_person = col(["person"])
    idx_date   = col(["datum", "date"])
    idx_beg    = col(["schichtbeginn", "beginn"])
    idx_end    = col(["schichtende", "ende"])
    idx_dauer  = col(["schichtdauer"])
    idx_profil = col(["arbeitszeit nach arbeitszeitprofil"])
    idx_lkw    = col(["fahrzeuge", "terminal"])
    idx_lenk   = col(["lenkzeit"])
    idx_bereit = col(["bereitschaft"])
    idx_arbeit = col_exact(["arbeitszeit"])
    idx_card   = col(["fahrerschlüssel", "fahrerschluessel", "driver card", "kartennummer"])
    idx_ma     = col(["ma-nummer", "ma nummer", "personalnummer", "mitarbeiternummer"])

    # YellowFox liefert zwei unterschiedliche Exportvarianten:
    # 1) separate Spalten ``Datum`` + ``Beginn`` + ``Ende``
    # 2) kombinierte Spalten ``Schichtbeginn`` + ``Schichtende`` mit Datum/Uhrzeit
    # Eine separate Datumsspalte darf deshalb NICHT zwingend vorausgesetzt werden.
    if idx_person < 0 or idx_beg < 0:
        return "{}"

    WD = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    by_driver = {}

    def split_dt(s):
        s = (s or "").strip()
        if not s or s.casefold() in ("nan", "none", "nat"):
            return ("", "")

        m_time = re.fullmatch(r"(\d{1,2}):(\d{2})(?::\d{2})?", s)
        if m_time:
            hour = int(m_time.group(1))
            minute = int(m_time.group(2))
            if 0 <= hour <= 23 and 0 <= minute <= 59:
                return ("", f"{hour:02d}:{minute:02d}")

        try:
            d_obj = _dt.datetime.fromisoformat(s)
            has_time = bool(re.search(r"[ T]\d{1,2}:\d{2}", s))
            return (
                d_obj.strftime("%d.%m.%Y"),
                d_obj.strftime("%H:%M") if has_time else "",
            )
        except (ValueError, TypeError):
            pass

        for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
            try:
                d_obj = _dt.datetime.strptime(s, fmt)
                return (
                    d_obj.strftime("%d.%m.%Y"),
                    d_obj.strftime("%H:%M") if "%H" in fmt else "",
                )
            except ValueError:
                continue
        return ("", "")

    def parse_time_list(value, allow_24=False):
        """Extrahiert einzelne Uhrzeiten, auch aus ``00:00,18:45``."""
        out = []
        for hh, mm in re.findall(r"(?<!\d)(\d{1,2}):(\d{2})(?::\d{2})?(?!\d)", str(value or "")):
            hour = int(hh)
            minute = int(mm)
            valid_hour = 0 <= hour <= 23 or (allow_24 and hour == 24 and minute == 0)
            if not valid_hour or not (0 <= minute <= 59):
                continue
            value_norm = f"{hour:02d}:{minute:02d}"
            if value_norm not in out:
                out.append(value_norm)
        return out

    def fmt_duration(s):
        s = (s or "").strip()
        if not s or s.casefold() in ("nan", "none", "nat"):
            return ""
        m = re.search(r'(\d{1,2}):(\d{2})', s)
        if m:
            return f"{int(m.group(1)):02d}:{m.group(2)}"
        return s

    def cell(r, idx):
        return (r[idx] or "").strip() if 0 <= idx < len(r) else ""

    for r in rows:
        if not r or len(r) <= idx_person:
            continue
        name = cell(r, idx_person)
        if not name or name.casefold() in ("nan", "none"):
            continue

        date_raw = cell(r, idx_date)
        beg_raw = cell(r, idx_beg)
        end_raw = cell(r, idx_end)
        date_d, _date_t = split_dt(date_raw)
        beg_d, beg_t = split_dt(beg_raw)
        end_d, end_t = split_dt(end_raw)

        if not beg_d:
            beg_d = date_d
        if not end_d:
            end_d = beg_d

        beg_times = parse_time_list(beg_raw, allow_24=False)
        end_times = parse_time_list(end_raw, allow_24=True)
        if beg_t and beg_t not in beg_times:
            beg_times.insert(0, beg_t)
        if end_t and end_t not in end_times:
            end_times.insert(0, end_t)

        dauer  = fmt_duration(cell(r, idx_dauer))
        profil = fmt_duration(cell(r, idx_profil))
        lkw    = cell(r, idx_lkw)
        lenk   = fmt_duration(cell(r, idx_lenk))
        bereit = fmt_duration(cell(r, idx_bereit))
        arbeit = fmt_duration(cell(r, idx_arbeit))
        card   = cell(r, idx_card)
        ma_nr  = cell(r, idx_ma)

        if lkw.casefold() in ("nan", "none"):
            lkw = ""
        if card.casefold() in ("nan", "none", "0"):
            card = ""
        if ma_nr.casefold() in ("nan", "none", "0"):
            ma_nr = ""

        # Technische YellowFox-Platzhalter besitzen zwar Beginn/Ende, aber
        # keinerlei echte Aktivitätswerte. Sie dürfen weder als Schicht noch als
        # Sa-/So-Einsatz gewertet werden.
        has_shift_evidence = any((dauer, profil, lkw, lenk, bereit, arbeit))
        if not has_shift_evidence:
            continue
        if not beg_d or not beg_times:
            continue

        sort_key = beg_d
        wd = ""
        try:
            d_obj = _dt.datetime.strptime(beg_d, "%d.%m.%Y")
            wd = WD[d_obj.weekday()]
            sort_key = d_obj.strftime("%Y-%m-%d") + " " + beg_times[0]
        except Exception:
            pass

        next_day = bool(end_d) and end_d != beg_d
        if not next_day and "24:00" in end_times:
            next_day = True
        if not next_day and beg_times and end_times:
            try:
                last_beg = int(beg_times[-1][:2]) * 60 + int(beg_times[-1][3:5])
                last_end = int(end_times[-1][:2]) * 60 + int(end_times[-1][3:5])
                next_day = last_end < last_beg
            except Exception:
                pass

        entry = {
            "tag": beg_d,
            "wochentag": wd,
            "beginn": ", ".join(beg_times),
            "ende": ", ".join(end_times),
            "beginn_zeiten": beg_times,
            "ende_zeiten": end_times,
            "echte_beginne": list(beg_times),
            "fortsetzung_vortag": False,
            "ende_naechster_tag": next_day,
            "schichtdauer": dauer,
            "profil": profil,
            "lkw": lkw,
            "hat_schichtdaten": True,
            "fahrerschluessel": card,
            "ma_nummer": ma_nr,
            "_sort": sort_key,
        }
        by_driver.setdefault(name, []).append(entry)

    # Eine 00:00-Zeile ist häufig nur die Fortsetzung einer am Vortag bis
    # 24:00 laufenden Schicht. Für die Anfangstagsregel ist 00:00 dann kein
    # neuer Schichtbeginn. Weitere Startzeiten desselben Tages bleiben erhalten.
    for name in by_driver:
        by_driver[name].sort(key=lambda e: e.get("_sort", ""))
        previous = None
        for entry in by_driver[name]:
            starts = list(entry.get("beginn_zeiten") or [])
            continuation = False
            if starts and starts[0] == "00:00" and previous:
                try:
                    prev_date = _dt.datetime.strptime(previous.get("tag", ""), "%d.%m.%Y").date()
                    this_date = _dt.datetime.strptime(entry.get("tag", ""), "%d.%m.%Y").date()
                    prev_ends = previous.get("ende_zeiten") or []
                    if (this_date - prev_date).days == 1 and "24:00" in prev_ends:
                        starts = starts[1:]
                        continuation = True
                except Exception:
                    pass
            entry["echte_beginne"] = starts
            entry["fortsetzung_vortag"] = continuation
            entry.pop("_sort", None)
            previous = entry

    return _json.dumps(by_driver, ensure_ascii=False)



# =============================================================================
# JAVASCRIPT-BAUSTEINE DES DASHBOARDS
# Statische Skripte liegen bewusst außerhalb von combine_html(). Dadurch bleibt
# die eigentliche Zusammenstellung der HTML-Datei kurz und besser wartbar.
# =============================================================================

# _static_payload_text("_JS_SPESEN") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_FA") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_ZULAGE") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_WASH") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_WASH_RANKING") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_FW_GRAPH") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_VERSTOSS") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

_JS_KNAPP = r""""""

# _static_payload_text("_JS_SPED") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_FABEW") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_BUS") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

_BUS_CONTACT_RUNTIME_PATCH = r"""
// ── Buskontakte: Jens Becker + Eskalationsstufe Daniel Bock ──────────────────
(function(){
  var JENS_PHONE = "0172-5829596";
  var DANIEL_PHONE = "0170-6653968";
  var DANIEL_EMAIL = "daniel.bock@123bus.de";

  function digits(value){
    return String(value || "").replace(/\D/g, "");
  }

  function isDaniel(entry){
    var text = String((entry && (entry.label || entry.kopf)) || "").toLowerCase();
    return text.indexOf("daniel bock") !== -1 ||
           digits(entry && entry.nummer) === digits(DANIEL_PHONE) ||
           String((entry && entry.email) || "").toLowerCase() === DANIEL_EMAIL;
  }

  if(Array.isArray(BUS_KONTAKTE)){
    var jensIndex = BUS_KONTAKTE.findIndex(function(entry){
      var label = String((entry && entry.label) || "").toLowerCase();
      return label.indexOf("jens becker") !== -1 || label.indexOf("becker tours") !== -1;
    });

    var jensEntry = {
      label: "Jens Becker (Becker Tours)",
      nummer: JENS_PHONE
    };
    if(jensIndex >= 0){
      jensEntry = Object.assign({}, BUS_KONTAKTE[jensIndex], jensEntry);
    }

    BUS_KONTAKTE = BUS_KONTAKTE.filter(function(entry, index){
      return index !== jensIndex && !isDaniel(entry);
    });

    var insertAt = jensIndex >= 0 ? Math.min(jensIndex, BUS_KONTAKTE.length) : Math.min(1, BUS_KONTAKTE.length);
    BUS_KONTAKTE.splice(insertAt, 0,
      jensEntry,
      {
        label: "Daniel Bock (nächsthöherer Ansprechpartner)",
        nummer: DANIEL_PHONE,
        email: DANIEL_EMAIL
      }
    );
  }

  function patchEmergencyScenario(titlePart, jensHeading){
    if(!Array.isArray(BUS_NOTFALL)) return;
    var scenario = BUS_NOTFALL.find(function(entry){
      return String((entry && entry.titel) || "").toLowerCase().indexOf(titlePart) !== -1;
    });
    if(!scenario || !Array.isArray(scenario.schritte)) return;

    var jensIndex = scenario.schritte.findIndex(function(step){
      var heading = String((step && step.kopf) || "").toLowerCase();
      var number = digits(step && step.nummer);
      return heading.indexOf("jens becker") !== -1 ||
             heading.indexOf("becker tours") !== -1 ||
             number === "01725825596" || number === digits(JENS_PHONE);
    });
    if(jensIndex < 0) return;

    scenario.schritte[jensIndex] = Object.assign({}, scenario.schritte[jensIndex], {
      kopf: jensHeading,
      nummer: JENS_PHONE
    });
    scenario.schritte = scenario.schritte.filter(function(step, index){
      return index === jensIndex || !isDaniel(step);
    });
    scenario.schritte.splice(jensIndex + 1, 0, {
      kopf: "Daniel Bock – nächsthöherer Ansprechpartner",
      nummer: DANIEL_PHONE,
      email: DANIEL_EMAIL,
      punkte: ["falls Jens Becker nicht erreicht wird"]
    });
  }

  patchEmergencyScenario("busfahrer ist nicht gekommen", "Handynummer Jens Becker");
  patchEmergencyScenario("bus hat unfall / panne", "Jens Becker von Becker Tours informieren");

  function addDanielEmailLinks(root){
    if(!root) return;
    var links = root.querySelectorAll('a[href="tel:' + digits(DANIEL_PHONE) + '"]');
    links.forEach(function(phoneLink){
      var host = phoneLink.parentElement;
      if(!host || host.querySelector('a[href="mailto:' + DANIEL_EMAIL + '"]')) return;
      var mail = document.createElement("a");
      mail.href = "mailto:" + DANIEL_EMAIL;
      mail.textContent = DANIEL_EMAIL;
      mail.style.cssText = "display:block;margin-top:3px;font-size:11.5px;font-weight:800;color:#1e6091;text-decoration:none;word-break:break-all";
      host.appendChild(mail);
    });
  }

  if(typeof busRender === "function"){
    var originalBusRender = busRender;
    busRender = function(){
      originalBusRender.apply(this, arguments);
      addDanielEmailLinks(document.getElementById("bus-content"));
    };
  }

  if(typeof busPDF === "function"){
    var originalBusPDF = busPDF;
    busPDF = function(){
      var openedWindow = null;
      var realOpen = window.open;
      window.open = function(){
        openedWindow = realOpen.apply(window, arguments);
        return openedWindow;
      };
      try {
        originalBusPDF.apply(this, arguments);
      } finally {
        window.open = realOpen;
      }
      if(openedWindow){
        setTimeout(function(){
          try {
            var doc = openedWindow.document;
            if(!doc || !doc.body || doc.body.innerHTML.indexOf(DANIEL_EMAIL) !== -1) return;
            var nodes = doc.querySelectorAll("span,.num,b,div");
            Array.prototype.forEach.call(nodes, function(node){
              if(String(node.textContent || "").indexOf(DANIEL_PHONE) === -1) return;
              var host = node.parentElement || node;
              if(host.querySelector && host.querySelector('[data-daniel-mail="1"]')) return;
              var mail = doc.createElement("div");
              mail.setAttribute("data-daniel-mail", "1");
              mail.textContent = DANIEL_EMAIL;
              mail.style.cssText = "margin-top:1mm;font-size:7pt;font-weight:800;color:#1e6091";
              host.appendChild(mail);
            });
          } catch(_err) {}
        }, 50);
      }
    };
  }
})();
"""


def _patch_bus_javascript_contacts(source: str) -> str:
    """Aktualisiert die Buskontakte in Anzeige, Notfallplan und Druckansicht."""
    source = str(source or "")
    if not source:
        return source

    # Bekannte frühere Schreibweisen/Fehleinträge der Becker-Tours-Nummer.
    for old_number in (
        "0172-5825596",
        "0172 - 5825596",
        "0172 5825596",
    ):
        source = source.replace(old_number, "0172-5829596")

    source = source.replace('label:"Becker Tours (Chef)"', 'label:"Jens Becker (Becker Tours)"')
    source = source.replace('kopf:"Handynummer Chef Becker Tours"', 'kopf:"Handynummer Jens Becker"')
    source = source.replace('kopf:"Firma Becker Tours informieren"', 'kopf:"Jens Becker von Becker Tours informieren"')

    # Daniel Bock direkt nach Jens Becker in der Kontaktleiste ergänzen.
    source = re.sub(
        r'(\{\s*label:"Jens Becker \(Becker Tours\)",\s*nummer:"0172-5829596"\s*\},)(?!\s*\{\s*label:"Daniel Bock)',
        r'\1\n  { label:"Daniel Bock (nächsthöherer Ansprechpartner)", nummer:"0170-6653968", email:"daniel.bock@123bus.de" },',
        source,
        count=1,
    )

    # Daniel als nächste Eskalationsstufe in beiden relevanten Notfallszenarien.
    source = re.sub(
        r'(\{\s*kopf:"Handynummer Jens Becker",\s*nummer:"0172-5829596",\s*punkte:\[[^\]]*\]\s*\},)(?!\s*\{\s*kopf:"Daniel Bock)',
        r'\1\n      { kopf:"Daniel Bock – nächsthöherer Ansprechpartner", nummer:"0170-6653968", email:"daniel.bock@123bus.de", punkte:["falls Jens Becker nicht erreicht wird"] },',
        source,
        count=1,
    )
    source = re.sub(
        r'(\{\s*kopf:"Jens Becker von Becker Tours informieren",\s*nummer:"0172-5829596",\s*punkte:\[[^\]]*\]\s*\},)(?!\s*\{\s*kopf:"Daniel Bock)',
        r'\1\n      { kopf:"Daniel Bock – nächsthöherer Ansprechpartner", nummer:"0170-6653968", email:"daniel.bock@123bus.de", punkte:["falls Jens Becker nicht erreicht wird"] },',
        source,
        count=1,
    )

    # E-Mail in der normalen Kontaktansicht anzeigen.
    contact_phone_line = """    html += "<a href='tel:"+busTel(k.nummer)+"' style='font-size:17px;font-weight:900;color:#1e6091;text-decoration:none;font-variant-numeric:tabular-nums'>"+busEsc(k.nummer)+"</a>";"""
    if "if(k.email)" not in source and contact_phone_line in source:
        source = source.replace(
            contact_phone_line,
            contact_phone_line + """\n    if(k.email) html += "<a href='mailto:"+busEsc(k.email)+"' style='display:block;margin-top:3px;font-size:11px;font-weight:800;color:#64748b;text-decoration:none;word-break:break-all'>"+busEsc(k.email)+"</a>";""",
            1,
        )

    step_phone_line = """      if(st.nummer) html += "<a href='tel:"+busTel(st.nummer)+"' style='display:inline-block;margin-top:3px;font-size:13px;font-weight:900;color:"+sz.farbe+";text-decoration:none;font-variant-numeric:tabular-nums'>"+busEsc(st.nummer)+"</a>";"""
    if "if(st.email)" not in source and step_phone_line in source:
        source = source.replace(
            step_phone_line,
            step_phone_line + """\n      if(st.email) html += "<a href='mailto:"+busEsc(st.email)+"' style='display:block;margin-top:2px;font-size:11px;font-weight:800;color:"+sz.farbe+";text-decoration:none;word-break:break-all'>"+busEsc(st.email)+"</a>";""",
            1,
        )

    # E-Mail auch im PDF/Druck ausgeben.
    pdf_contacts_old = """  BUS_KONTAKTE.forEach(function(k){ b += "<div><b>"+busEsc(k.label)+"</b><span>"+busEsc(k.nummer)+"</span></div>"; });"""
    pdf_contacts_new = """  BUS_KONTAKTE.forEach(function(k){ b += "<div><b>"+busEsc(k.label)+"</b><span>"+busEsc(k.nummer)+"</span>"+(k.email?"<small style='display:block;margin-top:1mm;font-size:6.5pt;font-weight:800;color:#64748b'>"+busEsc(k.email)+"</small>":"")+"</div>"; });"""
    source = source.replace(pdf_contacts_old, pdf_contacts_new, 1)

    pdf_step_phone = """      if(st.nummer) b += " <span class='num' style='color:"+sz.farbe+"'>"+busEsc(st.nummer)+"</span>";"""
    if "if(st.email) b +=" not in source and pdf_step_phone in source:
        source = source.replace(
            pdf_step_phone,
            pdf_step_phone + """\n      if(st.email) b += " <span style='display:block;margin-top:.5mm;font-size:7pt;font-weight:800;color:"+sz.farbe+"'>"+busEsc(st.email)+"</span>";""",
            1,
        )

    # Laufzeit-Sicherung: funktioniert auch dann, wenn der ausgelagerte Baustein
    # später leicht anders formatiert wird.
    if "Buskontakte: Jens Becker + Eskalationsstufe Daniel Bock" not in source:
        source += "\n" + _BUS_CONTACT_RUNTIME_PATCH
    return source

# _static_payload_text("_JS_ARZT") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_VERSP") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

# _static_payload_text("_JS_WA") wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.

def _dashboard_javascript_parts() -> dict:
    """Liefert die statischen JavaScript-Bausteine für die HTML-Erzeugung."""
    return {
        'spesen': _static_payload_text("_JS_SPESEN"),
        'fa': _static_payload_text("_JS_FA"),
        'zulage': _static_payload_text("_JS_ZULAGE"),
        'wash': _static_payload_text("_JS_WASH"),
        'wash_ranking': _static_payload_text("_JS_WASH_RANKING"),
        'fw_graph': _static_payload_text("_JS_FW_GRAPH"),
        'verstoss': _static_payload_text("_JS_VERSTOSS"),
        'knapp': _JS_KNAPP,
        'sped': _static_payload_text("_JS_SPED"),
        'fabew': _static_payload_text("_JS_FABEW"),
        'bus': _patch_bus_javascript_contacts(_static_payload_text("_JS_BUS")),
        'arzt': _static_payload_text("_JS_ARZT"),
        'versp': _static_payload_text("_JS_VERSP"),
        'wa': _static_payload_text("_JS_WA"),
    }



# =============================================================================
# HTML-TEMPLATE DES DASHBOARDS
# Dieser Bereich enthält ausschließlich das fertige HTML-Grundgerüst. Die
# Datenaufbereitung und Komprimierung bleibt getrennt in combine_html().
# =============================================================================

# =============================================================================
# DATENAUFBEREITUNG FÜR DIE HTML-ERZEUGUNG
# Kleine, getrennt testbare Helfer ersetzen den früheren monolithischen Block
# innerhalb von combine_html().
# =============================================================================

def _to_js_array(b64: str, width: int = 100) -> str:
    chunks = [b64[i:i + width] for i in range(0, len(b64), width)]
    return ",\n".join(f'"{chunk}"' for chunk in chunks)


_PDF_DOCUMENTS_JS_TEMPLATE = r"""
// ── Komprimiert eingebettete PDF-Dokumente ──────────────────────────────────
// Die PDF-Daten stecken in dieser HTML-Datei und werden erst bei Bedarf
// entpackt. Dadurch ist kein zusaetzlicher Dokumentenordner erforderlich.
var PDF_DOCUMENTS = __PDF_DOCUMENTS__;
var PDF_URL_CACHE = {};

async function documentPdfInflate(chunks){
  var b64 = (chunks || []).join("");
  if(!b64) return null;
  var bin = atob(b64);
  var bytes = new Uint8Array(bin.length);
  for(var i=0; i<bin.length; i++) bytes[i] = bin.charCodeAt(i);
  if(typeof DecompressionStream !== "function"){
    throw new Error("Dieser Browser kann die eingebettete PDF nicht entpacken.");
  }
  // Wichtig: Schreiben und Lesen muessen gleichzeitig laufen. Wenn zuerst auf
  // writer.write() gewartet wird, kann der TransformStream wegen Backpressure
  // haengen bleiben und die PDF-Vorschau bleibt leer.
  var inputStream = new Blob([bytes]).stream();
  var outputStream = inputStream.pipeThrough(new DecompressionStream("deflate"));
  return new Uint8Array(await new Response(outputStream).arrayBuffer());
}

async function documentPdfGetUrl(id){
  if(PDF_URL_CACHE[id]) return PDF_URL_CACHE[id];
  var doc = PDF_DOCUMENTS[id];
  if(!doc || !doc.z) return "";
  try{
    var pdfBytes = await documentPdfInflate(doc.z);
    if(!pdfBytes) return "";
    var url = URL.createObjectURL(new Blob([pdfBytes], {type:"application/pdf"}));
    PDF_URL_CACHE[id] = url;
    return url;
  }catch(err){
    console.error("PDF konnte nicht entpackt werden:", id, err);
    return "";
  }
}

async function documentPdfInit(id){
  var frame = document.getElementById(id + "-pdf-frame");
  var fallback = document.getElementById(id + "-pdf-fallback");
  var url = await documentPdfGetUrl(id);
  if(!frame || !url){
    if(frame) frame.style.display = "none";
    if(fallback) fallback.style.display = "flex";
    return;
  }
  if(!frame.dataset.loaded){
    frame.src = url + "#view=FitH&toolbar=1&navpanes=0";
    frame.dataset.loaded = "1";
  }
  frame.style.display = "block";
  if(fallback) fallback.style.display = "none";
}

async function documentPdfOpen(id){
  var url = await documentPdfGetUrl(id);
  if(!url){ alert("Die PDF konnte nicht geladen werden."); return; }
  var a = document.createElement("a");
  a.href = url + "#view=FitH&toolbar=1&navpanes=0";
  a.target = "_blank";
  a.rel = "noopener";
  document.body.appendChild(a);
  a.click();
  a.remove();
}

async function documentPdfDownload(id){
  var doc = PDF_DOCUMENTS[id];
  var url = await documentPdfGetUrl(id);
  if(!doc || !url){ alert("Die PDF konnte nicht geladen werden."); return; }
  var a = document.createElement("a");
  a.href = url;
  a.download = doc.filename;
  document.body.appendChild(a);
  a.click();
  a.remove();
}

// Kompatibilitaetsfunktionen fuer den bereits vorhandenen KNAPP-Bereich.
function knappGetPdfUrl(){ return documentPdfGetUrl("knapp"); }
function knappInit(){ documentPdfInit("knapp"); }
function knappOpenPdf(){ documentPdfOpen("knapp"); }
function knappDownloadPdf(){ documentPdfDownload("knapp"); }
"""


_EXCEL_DOCUMENTS_JS_TEMPLATE = r"""
// ── Eingebettete Excel-Dokumente ──────────────────────────────────────────────
var EXCEL_DOCUMENTS = __EXCEL_DOCUMENTS__;

async function documentExcelDownload(id){
  var doc = EXCEL_DOCUMENTS[id];
  if(!doc || !doc.z){ alert("Die Excel-Datei konnte nicht geladen werden."); return; }
  try{
    // Verwendet denselben zlib-Entpacker wie die eingebetteten PDF-Dokumente.
    var xlsxBytes = await documentPdfInflate(doc.z);
    if(!xlsxBytes){ alert("Die Excel-Datei konnte nicht geladen werden."); return; }
    var url = URL.createObjectURL(new Blob([xlsxBytes], {
      type:"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }));
    var a = document.createElement("a");
    a.href = url;
    a.download = doc.filename || "Reisekostenabrechnung.xlsx";
    document.body.appendChild(a);
    a.click();
    a.remove();
    setTimeout(function(){ URL.revokeObjectURL(url); }, 1500);
  }catch(err){
    console.error("Excel-Datei konnte nicht entpackt werden:", id, err);
    alert("Die Excel-Datei konnte nicht geladen werden.");
  }
}
"""


def _build_excel_documents_js() -> str:
    payload = {
        "reisekosten": {
            "z": [_REISEKOSTEN_XLSX_B64],
            "filename": "Reisekostenabrechnung.xlsx",
            "title": "Reisekostenabrechnung",
        }
    }
    payload_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return _EXCEL_DOCUMENTS_JS_TEMPLATE.replace("__EXCEL_DOCUMENTS__", payload_json)


def _build_pdf_documents_js() -> str:
    payload = {
        doc_id: {
            "z": list(meta["z"]),
            "filename": meta["filename"],
            "title": meta["title"],
        }
        for doc_id, meta in _load_embedded_pdf_documents().items()
    }
    payload_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return (
        _PDF_DOCUMENTS_JS_TEMPLATE.replace("__PDF_DOCUMENTS__", payload_json)
        + "\n"
        + _build_excel_documents_js()
    )


def _safe_json_object(value: str) -> str:
    try:
        obj = json.loads(value or "{}")
        if not isinstance(obj, dict):
            obj = {}
    except Exception:
        obj = {}
    return json.dumps(obj, ensure_ascii=False)


def _build_instances_js(instances: list) -> str:
    normal_versp_start_json = "{}"
    if instances and isinstance(instances[0], dict):
        normal_versp_start_json = instances[0].get("versp_start_json") or "{}"

    parts = []
    for idx, inst in enumerate(instances):
        search_b64 = base64.b64encode(
            zlib.compress(inst["suche_html"].encode("utf-8"), 9)
        ).decode("ascii")
        print_b64 = base64.b64encode(
            zlib.compress(inst["druck_html"].encode("utf-8"), 9)
        ).decode("ascii")
        search_js = _to_js_array(search_b64)
        print_js = _to_js_array(print_b64)
        name_escaped = inst["name"].replace('"', "&quot;").replace("'", "&#39;")
        versp_json = inst.get("versp_start_json") or "{}"
        if idx > 0 and versp_json in ("{}", "", None):
            versp_json = normal_versp_start_json
        versp_js = _safe_json_object(versp_json)
        week_obj = inst.get("woche_data") or {}
        try:
            week_js = json.dumps(week_obj, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            week_js = "{}"
        parts.append(
            f'{{name:"{name_escaped}",'
            f's:[{search_js}],'
            f'd:[{print_js}],'
            f'versp:{versp_js},'
            f'woche:{week_js}}}'
        )
    return ",\n".join(parts)


def _build_saturday_json_from_timerecording(timerec_json: str) -> str:
    sam_json = "[]"
    try:
        import json as _j
        import datetime as _dt2

        timerec = _j.loads(timerec_json) if timerec_json and timerec_json != "{}" else {}
        if not isinstance(timerec, dict):
            timerec = {}

        def _sam_excluded_driver(name):
            norm = str(name or "").strip().casefold()
            if not norm:
                return True
            if norm in ("unzugeordnet, planung", "planung, unzugeordnet"):
                return True
            return any(str(ex or "").casefold() in norm for ex in EXCLUDED_DRIVER_NAMES)

        def _sam_parse_mins(zeit_str):
            if not zeit_str or zeit_str == "n.A.":
                return None
            try:
                parts = str(zeit_str).strip().split(":")
                hour = int(parts[0])
                minute = int(parts[1])
                if 0 <= hour <= 23 and 0 <= minute <= 59:
                    return hour * 60 + minute
            except Exception:
                pass
            return None

        def _sam_start_values(shift):
            raw_values = shift.get("echte_beginne")
            if isinstance(raw_values, list):
                candidates = raw_values
            else:
                candidates = re.findall(r"(?<!\d)(\d{1,2}:\d{2})(?!\d)", str(shift.get("beginn", "") or ""))
            values = []
            for raw_value in candidates:
                mins = _sam_parse_mins(raw_value)
                if mins is None:
                    continue
                value = f"{mins // 60:02d}:{mins % 60:02d}"
                if value not in values:
                    values.append(value)
            return values

        def _sam_has_shift_evidence(shift):
            """Zweiter Sicherheitsfilter für bereits gecachte ältere Parserdaten.

            Technische YellowFox-Zeilen enthalten oft nur Datum, 00:00 und
            24:00, aber keinerlei Dauer, Fahrzeug oder Arbeitsprofil. Solche
            Zeilen dürfen auch dann nicht zählen, wenn sie aus einem älteren
            ``timerec_json`` stammen.
            """
            if shift.get("hat_schichtdaten") is True:
                return True

            evidence_keys = (
                "schichtdauer", "profil", "lkw",
                "lenkzeit", "lenk", "bereitschaft", "bereit",
                "arbeitszeit", "arbeit",
            )
            empty_values = {"", "nan", "none", "nat", "0", "0:00", "00:00", "00:00:00"}
            for key in evidence_keys:
                value = str(shift.get(key, "") or "").strip().casefold()
                if value not in empty_values:
                    return True
            return False

        if not timerec:
            sam_json = "[]"
        else:
            sam_by_name = {}
            sam_by_day = {}
            sam_active_years = {}

            def _sam_clean_name(value):
                s = str(value or "").replace("\xa0", " ")
                s = re.sub(r"[\u200b-\u200f\u202a-\u202e\u2060\ufeff]", "", s)
                s = re.sub(r"\s+", " ", s).strip()
                s = re.sub(r"\s*,\s*", ", ", s)
                return s.strip(" ,;-")

            def _sam_id(value):
                s = str(value or "").strip()
                if s.casefold() in ("", "nan", "none", "0"):
                    return ""
                return re.sub(r"\s+", "", s)

            def _sam_name_key(value):
                s = _sam_clean_name(value).casefold()
                if not s:
                    return ""
                s = (s.replace("ä", "ae").replace("ö", "oe")
                       .replace("ü", "ue").replace("ß", "ss"))
                s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
                tokens = [t for t in re.findall(r"[a-z0-9]+", s) if not t.isdigit()]
                return "|".join(sorted(tokens))

            def _sam_display_score(value):
                s = _sam_clean_name(value)
                score = 0
                if "," in s:
                    score += 100
                if s and not s.isupper() and not s.islower():
                    score += 20
                score += min(len(s), 60) / 100.0
                return score

            def _sam_update_display(driver, candidate):
                candidate = _sam_clean_name(candidate)
                if not candidate:
                    return
                current = driver.get("name", "")
                if not current or _sam_display_score(candidate) > _sam_display_score(current):
                    driver["name"] = candidate
                    parts = candidate.split(",", 1)
                    if len(parts) > 1:
                        driver["nachname"] = parts[0].strip()
                        driver["vorname"] = parts[1].strip()
                    else:
                        driver["nachname"] = candidate
                        driver["vorname"] = ""

            name_to_ma = {}
            card_to_ma = {}
            for raw_name, shifts in timerec.items():
                if not isinstance(shifts, list):
                    continue
                clean_name = _sam_clean_name(raw_name)
                ma_values = set()
                for shift in shifts:
                    if not isinstance(shift, dict):
                        continue
                    ma_val = _sam_id(shift.get("ma_nummer", ""))
                    card_val = _sam_id(shift.get("fahrerschluessel", ""))
                    if ma_val:
                        ma_values.add(ma_val)
                        if card_val:
                            card_to_ma[card_val] = ma_val
                if len(ma_values) == 1:
                    name_to_ma[_sam_name_key(clean_name)] = next(iter(ma_values))

            def _sam_person_key(raw_name, shift):
                ma_val = _sam_id((shift or {}).get("ma_nummer", ""))
                card_val = _sam_id((shift or {}).get("fahrerschluessel", ""))
                name_key = _sam_name_key(raw_name)
                if ma_val:
                    return "ma:" + ma_val
                if card_val and card_val in card_to_ma:
                    return "ma:" + card_to_ma[card_val]
                if name_key and name_key in name_to_ma:
                    return "ma:" + name_to_ma[name_key]
                if card_val:
                    return "card:" + card_val
                return "name:" + name_key if name_key else ""

            def _ensure_sam_driver(raw_name, person_key):
                display_name = _sam_clean_name(raw_name)
                if not person_key:
                    return ""
                if person_key not in sam_by_name:
                    sam_by_name[person_key] = {
                        "person_key": person_key,
                        "name": display_name,
                        "nachname": "",
                        "vorname": "",
                        "einsaetze": 0,
                        "daten": [],
                        "aktive_jahre": [],
                    }
                    sam_by_day[person_key] = {}
                    sam_active_years[person_key] = set()
                _sam_update_display(sam_by_name[person_key], display_name)
                return person_key

            # Gezählt wird nur ein echter Schichtbeginn. Technische 00:00–24:00-
            # Platzhalter und reine Fortsetzungen vom Vortag besitzen keine
            # ``echte_beginne`` und gelangen deshalb nicht in die Auswertung.
            for driver_name, shifts in timerec.items():
                if _sam_excluded_driver(driver_name) or not isinstance(shifts, list):
                    continue

                name = _sam_clean_name(driver_name)
                for shift in shifts:
                    if not isinstance(shift, dict):
                        continue

                    # Unabhängig vom Parser nochmals prüfen. Das entfernt
                    # auch Platzhalter aus einem alten Streamlit-Session-Cache.
                    if not _sam_has_shift_evidence(shift):
                        continue

                    tag_str = str(shift.get("tag", "") or "").strip()
                    lkw = str(shift.get("lkw", "") or "").strip()
                    starts = _sam_start_values(shift)
                    if not starts:
                        continue

                    try:
                        d_obj = _dt2.datetime.strptime(tag_str, "%d.%m.%Y")
                    except Exception:
                        continue

                    person_key = _sam_person_key(name, shift)
                    driver_key = _ensure_sam_driver(name, person_key)
                    if not driver_key:
                        continue
                    sam_active_years[driver_key].add(d_obj.year)

                    weekday_idx = d_obj.weekday()  # Mo=0 ... So=6
                    if weekday_idx == 5:
                        tag_label = "Sa"
                        qualifying_starts = starts
                    elif weekday_idx == 4:
                        tag_label = "Fr→Sa"
                        qualifying_starts = [s for s in starts if (_sam_parse_mins(s) or 0) >= 18 * 60]
                    elif weekday_idx == 6:
                        tag_label = "So"
                        qualifying_starts = [s for s in starts if _sam_parse_mins(s) is not None and _sam_parse_mins(s) <= 15 * 60]
                    else:
                        continue

                    if not qualifying_starts:
                        continue

                    day_key = d_obj.strftime("%Y-%m-%d")
                    kw = d_obj.isocalendar()[1]
                    day_map = sam_by_day[driver_key]
                    if day_key not in day_map:
                        day_map[day_key] = {
                            "iso": day_key,
                            "datum": f"{tag_str} (KW{kw})",
                            "tour": "",
                            "tag": tag_label,
                            "beginn": "",
                            "_lkw": set(),
                            "_starts": set(),
                        }
                    for start_value in qualifying_starts:
                        day_map[day_key]["_starts"].add(start_value)
                    if lkw and lkw.casefold() not in ("nan", "none", "0"):
                        day_map[day_key]["_lkw"].add(lkw)

            sam_list = []
            for driver_key, driver in sam_by_name.items():
                entries = []
                for day_key in sorted(sam_by_day.get(driver_key, {})):
                    entry = sam_by_day[driver_key][day_key]
                    lkw_values = sorted(entry.pop("_lkw", set()))
                    start_values = sorted(entry.pop("_starts", set()))
                    entry["beginn"] = ", ".join(start_values)
                    entry["tour"] = ("LKW " + ", ".join(lkw_values)) if lkw_values else ""
                    entries.append(entry)
                driver["daten"] = entries
                driver["einsaetze"] = len(entries)
                driver["aktive_jahre"] = sorted(sam_active_years.get(driver_key, set()), reverse=True)
                sam_list.append(driver)

            sam_list.sort(key=lambda x: str(x.get("name", "")).casefold())
            sam_json = _j.dumps(sam_list, ensure_ascii=False)
    except Exception:
        sam_json = "[]"
    return sam_json


def _json_or_default(raw_value: str, default_json: str):
    try:
        return json.loads(raw_value if raw_value not in (None, "") else default_json)
    except Exception:
        return json.loads(default_json)


def _build_embedded_data_js(
    *,
    fahrzeugwaesche_json: str,
    tanken_json: str,
    tel_json: str,
    sam_json: str,
    fa_json: str,
    fahrerbewertung_json: str,
    zulage_json: str,
    drittkunden_json: str,
    verstoss_json: str,
    spesen_json: str,
    grosskunden_json: str,
    timerec_json: str,
    spediteure_json: str,
    versp_abfahrt_json: str,
) -> str:
    data = {
        "fahrzeugwaesche": _json_or_default(fahrzeugwaesche_json, "[]"),
        "tanken": _json_or_default(tanken_json, "[]"),
        "telefon": _json_or_default(tel_json, "[]"),
        "samstag": _json_or_default(sam_json, "[]"),
        "fahrer": _json_or_default(fa_json, "[]"),
        "fahrerbewertung": _json_or_default(fahrerbewertung_json, '{"profile":"","event_types":[],"g_months":{},"g_ev":{},"drivers":[]}'),
        "zulagen": _json_or_default(zulage_json, "{}"),
        "drittkunden": _json_or_default(drittkunden_json, "[]"),
        "verstoesse": _json_or_default(verstoss_json, '{"drivers":[],"total_violations":0}'),
        "spesen": _json_or_default(spesen_json, '{"drivers":[],"months":[],"total_cost":0,"total_rows":0}'),
        "grosskunden": _json_or_default(grosskunden_json, "[]"),
        "timerecording": _json_or_default(timerec_json, "{}"),
        "spediteure": _json_or_default(spediteure_json, '{"katalog":[],"fahrten":[]}'),
        "verspaetung_abfahrt": _json_or_default(versp_abfahrt_json, "{}"),
    }
    raw = json.dumps(data, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    compressed_b64 = base64.b64encode(zlib.compress(raw, 9)).decode("ascii")
    return _to_js_array(compressed_b64)



def _tank_panels_html() -> str:
    """HTML-Panels für Tankübersicht und die erweiterte Tankauswertung."""
    return r"""
<style>
  #ddmenu-vz .dd-group-title{padding:7px 12px 5px;background:#eef3f8;color:#64748b;font-size:9px;font-weight:950;text-transform:uppercase;letter-spacing:.65px;border-bottom:1px solid #dce4ed}
  #ddmenu-vz .dd-subitem{padding-left:27px;position:relative}
  #ddmenu-vz .dd-subitem:before{content:'›';position:absolute;left:14px;color:#94a3b8;font-size:15px;line-height:1}
  #panel-tank,#panel-tank-graph{--tank:#9a5b00;--tank-dark:#704000;--tank-soft:#fff7e6;--ink:#132033;--muted:#64748b}
  .tank-shell{width:100%;max-width:1728px;margin:0 auto}
  .tank-card{background:#fff;border:1px solid #d8dee7;border-radius:13px;box-shadow:0 3px 12px rgba(15,23,42,.06);overflow:hidden}
  .tank-head{display:flex;align-items:center;gap:13px;padding:17px 20px;background:linear-gradient(180deg,#fffaf0 0%,#fff 100%);border-bottom:1px solid #eceff4;flex-wrap:wrap}
  .tank-icon{width:42px;height:42px;border-radius:11px;background:linear-gradient(135deg,#d99a22,#9a5b00);color:#fff;display:flex;align-items:center;justify-content:center;font-size:21px;box-shadow:0 5px 13px rgba(154,91,0,.24);flex-shrink:0}
  .tank-title{font-size:19px;font-weight:950;color:var(--ink);letter-spacing:-.3px}
  .tank-sub{font-size:11.5px;font-weight:650;color:var(--muted);margin-top:2px}
  .tank-controls{display:flex;gap:9px;align-items:center;flex-wrap:wrap;padding:13px 18px;background:#fbfcfe;border-bottom:1px solid #edf1f5}
  .tank-select,.tank-search{padding:9px 11px;border:1.5px solid #cfd8e3;border-radius:8px;background:#fff;color:#26374a;font:700 12px 'Segoe UI',Arial,sans-serif;outline:none;min-width:145px}
  .tank-search{min-width:240px;flex:1;max-width:430px}
  .tank-select:focus,.tank-search:focus{border-color:#b7791f;box-shadow:0 0 0 3px rgba(183,121,31,.12)}
  .tank-export-btn{padding:9px 14px;border:1.5px solid #15803d;border-radius:8px;background:linear-gradient(180deg,#22c55e 0%,#15803d 100%);color:#fff;font:900 12px 'Segoe UI',Arial,sans-serif;cursor:pointer;box-shadow:0 3px 8px rgba(21,128,61,.2);white-space:nowrap}
  .tank-export-btn:hover{filter:brightness(1.04)}
  .tank-kpis{display:grid;grid-template-columns:repeat(6,minmax(125px,1fr));gap:10px;padding:15px 18px;background:#f6f8fb}
  .tank-graph-kpis{margin:0 18px 14px;border:1px solid #d8dee7;border-radius:13px;box-shadow:0 3px 12px rgba(15,23,42,.05)}
  .tank-kpi{background:#fff;border:1px solid #e1e7ef;border-radius:11px;padding:12px 13px;min-width:0}
  .tank-kpi-label{font-size:9px;font-weight:950;color:#718096;text-transform:uppercase;letter-spacing:.55px}
  .tank-kpi-value{font-size:19px;font-weight:950;color:#172033;margin-top:4px;font-variant-numeric:tabular-nums;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .tank-kpi-note{font-size:9.5px;color:#94a3b8;font-weight:700;margin-top:2px}
  .tank-section-head{display:flex;justify-content:space-between;align-items:center;gap:10px;padding:13px 18px;background:#fff;border-top:1px solid #edf1f5;border-bottom:1px solid #edf1f5}
  .tank-section-title{font-size:12px;font-weight:950;color:#1f2937}
  .tank-section-meta{font-size:10.5px;color:#64748b;font-weight:700}
  .tank-table-wrap{overflow:auto;max-height:520px}
  .tank-table{width:100%;border-collapse:collapse;font-size:11px;min-width:900px}
  .tank-table th{position:sticky;top:0;z-index:1;background:#edf2f7;color:#536273;font-size:9px;text-transform:uppercase;letter-spacing:.42px;font-weight:950;padding:9px 10px;text-align:left;border-bottom:1px solid #d8e0e9;white-space:nowrap}
  .tank-table td{padding:8px 10px;border-bottom:1px solid #edf1f5;color:#263548;font-weight:650;vertical-align:middle;white-space:nowrap}
  .tank-table tbody tr:nth-child(even) td{background:#fafbfc}
  .tank-table tbody tr:hover td{background:#fff8e8}
  .tank-num{text-align:right!important;font-variant-numeric:tabular-nums}
  .tank-empty{color:#94a3b8;padding:58px 20px;text-align:center;font-size:14px;font-weight:650}
  .tank-badge{display:inline-flex;align-items:center;padding:2px 7px;border-radius:999px;background:#fff4d6;border:1px solid #f2d083;color:#815000;font-size:9.5px;font-weight:900}
  .tank-graph-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px;padding:0 18px 24px}
  .tank-chart-card{background:#fff;border:1px solid #d8dee7;border-radius:12px;padding:15px 16px;box-shadow:0 3px 12px rgba(15,23,42,.05);min-height:350px}
  .tank-chart-card.wide{grid-column:1/-1}
  .tank-chart-title{font-size:12.5px;font-weight:950;color:#1f2937;margin-bottom:3px}
  .tank-chart-sub{font-size:10.5px;color:#64748b;font-weight:650;margin-bottom:12px}
  .tank-canvas-wrap{position:relative;height:285px}
  .tank-canvas-wrap.tank-driver-chart{height:420px}
  @media(max-width:1250px){.tank-graph-kpis{grid-template-columns:repeat(3,minmax(120px,1fr))}}
  @media(max-width:1100px){.tank-kpis{grid-template-columns:repeat(3,minmax(120px,1fr))}.tank-graph-grid{grid-template-columns:1fr}.tank-chart-card.wide{grid-column:auto}}
  @media(max-width:650px){.tank-kpis{grid-template-columns:repeat(2,minmax(110px,1fr))}.tank-search{min-width:100%;max-width:none}.tank-select{flex:1}.tank-head{padding:14px}.tank-controls{padding:11px}.tank-graph-grid{padding:0 10px 18px}.tank-graph-kpis{margin:0 10px 12px}}
</style>

<div id="panel-tank" style="display:none;flex:1;overflow-y:auto;padding:18px 18px 30px;background:linear-gradient(180deg,#f5f7fa 0%,#edf1f5 100%);font-family:'Segoe UI',Arial,sans-serif">
  <div class="tank-shell tank-card">
    <div class="tank-head">
      <div class="tank-icon">&#9981;</div>
      <div style="min-width:0;flex:1">
        <div class="tank-title">Tanken &ndash; Übersicht</div>
        <div class="tank-sub">Monatliche Tankdateien aus dem Streamlit-Mehrfach-Upload auswerten</div>
      </div>
      <span id="tank-data-range" class="tank-badge"></span>
    </div>
    <div class="tank-controls">
      <select id="tank-year" class="tank-select" onchange="tankRenderOverview()"></select>
      <select id="tank-month" class="tank-select" onchange="tankRenderOverview()"></select>
      <select id="tank-group" class="tank-select" onchange="tankRenderOverview()">
        <option value="fahrzeug">Zusammenfassung nach LKW</option>
        <option value="fahrer">Zusammenfassung nach Fahrer</option>
        <option value="produkt">Zusammenfassung nach Produkt</option>
      </select>
      <input id="tank-search" class="tank-search" type="search" placeholder="Fahrer, Kennzeichen oder Fahrzeug suchen …" oninput="tankRenderOverview()">
    </div>
    <div id="tank-kpis" class="tank-kpis"></div>
    <div class="tank-section-head"><div class="tank-section-title" id="tank-summary-title">Zusammenfassung</div><div class="tank-section-meta" id="tank-summary-meta"></div></div>
    <div id="tank-summary-table" class="tank-table-wrap"></div>
    <div class="tank-section-head"><div class="tank-section-title">Einzelne Tankvorgänge</div><div class="tank-section-meta" id="tank-detail-meta"></div></div>
    <div id="tank-detail-table" class="tank-table-wrap"></div>
  </div>
</div>

<div id="panel-tank-graph" style="display:none;flex:1;overflow-y:auto;padding:18px 0 30px;background:linear-gradient(180deg,#f5f7fa 0%,#edf1f5 100%);font-family:'Segoe UI',Arial,sans-serif">
  <div class="tank-shell">
    <div class="tank-head" style="margin:0 18px 12px;border:1px solid #d8dee7;border-radius:13px;box-shadow:0 3px 12px rgba(15,23,42,.05)">
      <div class="tank-icon">&#128202;</div>
      <div style="min-width:220px;flex:1"><div class="tank-title">Tanken &ndash; Graph &amp; Auswertung</div><div class="tank-sub">Monatsentwicklung, Verbrauch, Datenqualität sowie Auswertung nach Fahrzeugen, Fahrern und Produkten</div></div>
      <select id="tank-graph-year" class="tank-select" onchange="tankRenderGraph()" title="Jahr"></select>
      <select id="tank-graph-product" class="tank-select" onchange="tankRenderGraph()" title="Produkt"></select>
      <select id="tank-graph-from" class="tank-select" onchange="tankRenderGraph()" title="Von Monat"></select>
      <select id="tank-graph-to" class="tank-select" onchange="tankRenderGraph()" title="Bis Monat"></select>
      <input id="tank-graph-search" class="tank-search" type="search" placeholder="LKW, Fahrer, Firma oder Zapfsäule …" oninput="tankRenderGraph()">
      <button class="tank-export-btn" type="button" onclick="tankExportExcel()" title="Aktuelle Tankauswertung als Excel-Datei exportieren">&#128190; Excel-Export</button>
      <span id="tank-graph-stats" class="tank-badge"></span>
    </div>
    <div id="tank-graph-kpis" class="tank-kpis tank-graph-kpis"></div>
    <div id="tank-graph-empty" class="tank-empty" style="display:none">Keine Tankdaten für die gewählte Auswahl vorhanden.</div>
    <div id="tank-graph-grid" class="tank-graph-grid">
      <div class="tank-chart-card wide"><div class="tank-chart-title">Getankte Liter pro Monat</div><div class="tank-chart-sub">Gesamtmenge im gewählten Zeitraum</div><div class="tank-canvas-wrap"><canvas id="tank-chart-liters"></canvas></div></div>
      <div class="tank-chart-card"><div class="tank-chart-title">Ø Verbrauch pro Monat</div><div class="tank-chart-sub">Liter je 100 km aus plausiblen Streckenangaben</div><div class="tank-canvas-wrap"><canvas id="tank-chart-consumption"></canvas></div></div>
      <div class="tank-chart-card"><div class="tank-chart-title">Plausible Strecke pro Monat</div><div class="tank-chart-sub">Summierte Kilometer aus Tankvorgängen mit 10 bis 5.000 km</div><div class="tank-canvas-wrap"><canvas id="tank-chart-distance"></canvas></div></div>
      <div class="tank-chart-card"><div class="tank-chart-title">Tankvorgänge pro Monat</div><div class="tank-chart-sub">Anzahl der erfassten Betankungen</div><div class="tank-canvas-wrap"><canvas id="tank-chart-count"></canvas></div></div>
      <div class="tank-chart-card"><div class="tank-chart-title">Ø Liter je Tankvorgang</div><div class="tank-chart-sub">Durchschnittliche Tankmenge pro Betankung</div><div class="tank-canvas-wrap"><canvas id="tank-chart-average"></canvas></div></div>
      <div class="tank-chart-card"><div class="tank-chart-title">Produktverteilung</div><div class="tank-chart-sub">Anteil der getankten Liter nach Produkt</div><div class="tank-canvas-wrap"><canvas id="tank-chart-products"></canvas></div></div>
      <div class="tank-chart-card"><div class="tank-chart-title">Top 12 LKW nach Litermenge</div><div class="tank-chart-sub">Fahrzeuge mit der höchsten getankten Gesamtmenge</div><div class="tank-canvas-wrap"><canvas id="tank-chart-vehicles"></canvas></div></div>
      <div class="tank-chart-card"><div class="tank-chart-title">Top 12 LKW nach Ø Verbrauch</div><div class="tank-chart-sub">Nur Fahrzeuge mit mindestens 200 plausiblen Kilometern</div><div class="tank-canvas-wrap"><canvas id="tank-chart-vehicle-consumption"></canvas></div></div>
      <div class="tank-chart-card wide"><div class="tank-chart-title">Getankte Liter nach Fahrer</div><div class="tank-chart-sub">Alle Fahrer der aktuellen Auswahl, absteigend nach getankter Gesamtmenge</div><div class="tank-canvas-wrap tank-driver-chart" id="tank-driver-chart-wrap"><canvas id="tank-chart-drivers"></canvas></div></div>
    </div>
  </div>
</div>
"""



def _tank_dashboard_js() -> str:
    """Browserlogik für Tankübersicht, erweiterte Graphen und Excel-Export."""
    return r"""
// ── Waschen & Tanken: gruppiertes Dropdown ──────────────────────────────────
window.ddSelectWashTank = function(area){
  showArea(area);
  document.querySelectorAll('.nav-dd').forEach(function(d){ d.classList.remove('open'); });
};
window.buildVzDdMenu = function(){
  var menu = document.getElementById('ddmenu-vz');
  if(!menu) return;
  var item = function(area,label){
    return "<div class='dd-item dd-subitem" + (currentArea===area?' active':'') + "' data-area='"+area+"' onclick='ddSelectWashTank(this.dataset.area)'>"+label+"</div>";
  };
  menu.innerHTML = "<div class='dd-group-title'>Fahrzeugwäsche</div>" + item('vz','Übersicht') + item('vz_graph','Graph') +
                   "<div class='dd-group-title'>Tanken</div>" + item('tank','Übersicht') + item('tank_graph','Graph & Auswertung');
};

// ── Tankdaten: Hilfsfunktionen ───────────────────────────────────────────────
var tankCharts = {liters:null, consumption:null, distance:null, count:null, average:null, products:null, vehicles:null, vehicleConsumption:null, drivers:null};
var TANK_MONTHS = ['Januar','Februar','März','April','Mai','Juni','Juli','August','September','Oktober','November','Dezember'];
function tankEsc(v){return String(v==null?'':v).replace(/[&<>"']/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];});}
function tankN(v){var n=Number(v);return Number.isFinite(n)?n:0;}
function tankFmt(v,d){return new Intl.NumberFormat('de-DE',{minimumFractionDigits:d||0,maximumFractionDigits:d==null?0:d}).format(tankN(v));}
function tankDateLabel(iso){if(!iso)return '';var p=String(iso).slice(0,10).split('-');return p.length===3?p[2]+'.'+p[1]+'.'+p[0]:iso;}
function tankValidDistance(r){var km=tankN(r.km);return km>=10 && km<=5000;}
function tankConsumption(rows){var liters=0,km=0;(rows||[]).forEach(function(r){if(tankValidDistance(r)){liters+=tankN(r.menge_liter);km+=tankN(r.km);}});return km>0?(liters/km*100):0;}
function tankYears(){return Array.from(new Set((TANK_DATA||[]).map(function(r){return Number(r.jahr)||0;}).filter(Boolean))).sort(function(a,b){return b-a;});}
function tankProducts(){return Array.from(new Set((TANK_DATA||[]).map(function(r){return String(r.produkt||'Ohne Angabe').trim()||'Ohne Angabe';}))).sort(function(a,b){return a.localeCompare(b,'de');});}
function tankFiltered(){
  var y=Number((document.getElementById('tank-year')||{}).value)||0;
  var m=Number((document.getElementById('tank-month')||{}).value)||0;
  var q=String((document.getElementById('tank-search')||{}).value||'').trim().toLocaleLowerCase('de-DE');
  return (TANK_DATA||[]).filter(function(r){
    if(y && Number(r.jahr)!==y) return false;
    if(m && Number(r.monat)!==m) return false;
    if(q){var hay=[r.fahrer,r.fahrzeug,r.fahrzeug_ia,r.produkt,r.zapfsaeule,r.firma].join(' ').toLocaleLowerCase('de-DE');if(hay.indexOf(q)<0)return false;}
    return true;
  });
}
function tankSetOptions(){
  var years=tankYears(), y=document.getElementById('tank-year'), gy=document.getElementById('tank-graph-year'), m=document.getElementById('tank-month');
  if(y && !y.options.length){y.innerHTML=years.map(function(v){return '<option value="'+v+'">Jahr '+v+'</option>';}).join('')+'<option value="0">Alle Jahre</option>';}
  if(gy && !gy.options.length){gy.innerHTML=years.map(function(v){return '<option value="'+v+'">Jahr '+v+'</option>';}).join('');}
  if(m && !m.options.length){m.innerHTML='<option value="0">Ganzes Jahr</option>'+TANK_MONTHS.map(function(v,i){return '<option value="'+(i+1)+'">'+v+'</option>';}).join('');}
  var gp=document.getElementById('tank-graph-product');
  if(gp){
    var products=tankProducts(), signature=products.join('|'), previous=gp.value||'';
    if(gp.dataset.signature!==signature){gp.innerHTML='<option value="">Alle Produkte</option>'+products.map(function(v){return '<option value="'+tankEsc(v)+'">'+tankEsc(v)+'</option>';}).join('');gp.dataset.signature=signature;if(products.indexOf(previous)>=0)gp.value=previous;}
  }
  var gf=document.getElementById('tank-graph-from'), gt=document.getElementById('tank-graph-to');
  if(gf && !gf.options.length){gf.innerHTML=TANK_MONTHS.map(function(v,i){return '<option value="'+(i+1)+'">Von '+v+'</option>';}).join('');gf.value='1';}
  if(gt && !gt.options.length){gt.innerHTML=TANK_MONTHS.map(function(v,i){return '<option value="'+(i+1)+'">Bis '+v+'</option>';}).join('');gt.value='12';}
  if(years.length){if(y && !y.value)y.value=String(years[0]);if(gy && !gy.value)gy.value=String(years[0]);}
  var range=document.getElementById('tank-data-range');
  if(range){var dates=(TANK_DATA||[]).map(function(r){return r.date_iso||'';}).filter(Boolean).sort();range.textContent=dates.length?tankDateLabel(dates[0])+' – '+tankDateLabel(dates[dates.length-1]):'Keine Daten';}
}
function tankKpi(label,value,note){return '<div class="tank-kpi"><div class="tank-kpi-label">'+tankEsc(label)+'</div><div class="tank-kpi-value">'+tankEsc(value)+'</div><div class="tank-kpi-note">'+tankEsc(note||'')+'</div></div>';}
function tankAggregate(rows,key){
  var map={};
  (rows||[]).forEach(function(r){
    var name=String(r[key]||'Ohne Angabe').trim()||'Ohne Angabe';
    if(!map[name])map[name]={name:name,count:0,liters:0,km:0,validKm:0,validLiters:0,validCount:0,last:'',first:'',drivers:new Set(),vehicles:new Set()};
    var a=map[name];a.count++;a.liters+=tankN(r.menge_liter);a.km+=tankN(r.km);
    if(tankValidDistance(r)){a.validCount++;a.validLiters+=tankN(r.menge_liter);a.validKm+=tankN(r.km);}
    if((r.date_iso||'')>a.last)a.last=r.date_iso||'';
    if(!a.first || (r.date_iso||'')<a.first)a.first=r.date_iso||'';
    if(r.fahrer)a.drivers.add(r.fahrer);if(r.fahrzeug)a.vehicles.add(r.fahrzeug);
  });
  return Object.values(map).map(function(a){a.consumption=a.validKm>0?(a.validLiters/a.validKm*100):0;a.avg=a.count?a.liters/a.count:0;a.validShare=a.count?(a.validCount/a.count*100):0;return a;}).sort(function(a,b){return b.liters-a.liters;});
}
function tankMonthly(rows){
  var months=Array.from({length:12},function(_,i){return {month:i+1,label:TANK_MONTHS[i],rows:[]};});
  (rows||[]).forEach(function(r){var m=Number(r.monat)||0;if(m>=1&&m<=12)months[m-1].rows.push(r);});
  return months.map(function(m){
    var liters=m.rows.reduce(function(s,r){return s+tankN(r.menge_liter);},0);
    var km=m.rows.reduce(function(s,r){return s+(tankValidDistance(r)?tankN(r.km):0);},0);
    return {month:m.month,label:m.label,count:m.rows.length,liters:liters,km:km,consumption:tankConsumption(m.rows),average:m.rows.length?liters/m.rows.length:0,vehicles:new Set(m.rows.map(function(r){return r.fahrzeug;}).filter(Boolean)).size,drivers:new Set(m.rows.map(function(r){return r.fahrer;}).filter(Boolean)).size};
  });
}
function tankRenderOverview(){
  tankSetOptions();
  var rows=tankFiltered(), totalLiters=rows.reduce(function(s,r){return s+tankN(r.menge_liter);},0), totalKm=rows.reduce(function(s,r){return s+(tankValidDistance(r)?tankN(r.km):0);},0);
  var vehicles=new Set(rows.map(function(r){return r.fahrzeug;}).filter(Boolean)), drivers=new Set(rows.map(function(r){return r.fahrer;}).filter(Boolean));
  var k=document.getElementById('tank-kpis');
  if(k)k.innerHTML=tankKpi('Tankvorgänge',tankFmt(rows.length,0),'gefilterte Datensätze')+tankKpi('Liter gesamt',tankFmt(totalLiters,2)+' l','Diesel / Produkte')+tankKpi('Strecke',tankFmt(totalKm,0)+' km','plausible km-Angaben')+tankKpi('Ø Verbrauch',tankFmt(tankConsumption(rows),2)+' l','je 100 km')+tankKpi('Fahrzeuge',tankFmt(vehicles.size,0),'unterschiedliche LKW')+tankKpi('Fahrer',tankFmt(drivers.size,0),'unterschiedliche Fahrer');
  var group=(document.getElementById('tank-group')||{value:'fahrzeug'}).value||'fahrzeug';
  var labels={fahrzeug:'LKW',fahrer:'Fahrer',produkt:'Produkt'}, ag=tankAggregate(rows,group);
  var title=document.getElementById('tank-summary-title');if(title)title.textContent='Zusammenfassung nach '+(labels[group]||group);
  var meta=document.getElementById('tank-summary-meta');if(meta)meta.textContent=ag.length+' Gruppen · '+tankFmt(totalLiters,2)+' Liter';
  var target=document.getElementById('tank-summary-table');
  if(target){
    if(!ag.length)target.innerHTML='<div class="tank-empty">Keine Tankdaten für die gewählte Auswahl.</div>';
    else target.innerHTML='<table class="tank-table"><thead><tr><th>'+(labels[group]||'Gruppe')+'</th><th class="tank-num">Tankvorgänge</th><th class="tank-num">Liter</th><th class="tank-num">plausible km</th><th class="tank-num">Ø l/100 km</th><th class="tank-num">Ø Liter/Tankung</th><th>Letzte Tankung</th><th class="tank-num">Fahrer</th><th class="tank-num">LKW</th></tr></thead><tbody>'+ag.map(function(a){return '<tr><td><b>'+tankEsc(a.name)+'</b></td><td class="tank-num">'+tankFmt(a.count,0)+'</td><td class="tank-num">'+tankFmt(a.liters,2)+'</td><td class="tank-num">'+tankFmt(a.validKm,0)+'</td><td class="tank-num">'+(a.consumption?tankFmt(a.consumption,2):'–')+'</td><td class="tank-num">'+tankFmt(a.avg,2)+'</td><td>'+tankDateLabel(a.last)+'</td><td class="tank-num">'+a.drivers.size+'</td><td class="tank-num">'+a.vehicles.size+'</td></tr>';}).join('')+'</tbody></table>';
  }
  var detail=document.getElementById('tank-detail-table'), sorted=rows.slice().sort(function(a,b){return String(b.datetime_iso||'').localeCompare(String(a.datetime_iso||''));}), shown=sorted.slice(0,500);
  var dm=document.getElementById('tank-detail-meta');if(dm)dm.textContent=sorted.length>500?'500 von '+sorted.length+' Einträgen':sorted.length+' Einträge';
  if(detail){if(!shown.length)detail.innerHTML='<div class="tank-empty">Keine einzelnen Tankvorgänge vorhanden.</div>';else detail.innerHTML='<table class="tank-table"><thead><tr><th>Datum</th><th>Uhrzeit</th><th>Kennzeichen</th><th>Fahrer</th><th>Produkt</th><th class="tank-num">Liter</th><th class="tank-num">Kilometerstand</th><th class="tank-num">km</th><th class="tank-num">l/100 km</th><th>Zapfsäule</th></tr></thead><tbody>'+shown.map(function(r){var c=tankValidDistance(r)?tankN(r.menge_liter)/tankN(r.km)*100:0;return '<tr><td>'+tankDateLabel(r.date_iso)+'</td><td>'+tankEsc(String(r.uhrzeit||'').slice(0,5))+'</td><td><b>'+tankEsc(r.fahrzeug||r.fahrzeug_ia)+'</b></td><td>'+tankEsc(r.fahrer)+'</td><td>'+tankEsc(r.produkt)+'</td><td class="tank-num">'+tankFmt(r.menge_liter,2)+'</td><td class="tank-num">'+tankFmt(r.kilometerzaehler,0)+'</td><td class="tank-num">'+tankFmt(r.km,0)+'</td><td class="tank-num">'+(c?tankFmt(c,2):'–')+'</td><td>'+tankEsc(r.zapfsaeule)+'</td></tr>';}).join('')+'</tbody></table>';}
}
window.tankInitOverview=function(){tankSetOptions();tankRenderOverview();};
function tankDestroyChart(key){if(tankCharts[key]){tankCharts[key].destroy();tankCharts[key]=null;}}
function tankGraphRows(){
  var year=Number((document.getElementById('tank-graph-year')||{}).value)||0;
  var product=String((document.getElementById('tank-graph-product')||{}).value||'');
  var from=Number((document.getElementById('tank-graph-from')||{}).value)||1;
  var to=Number((document.getElementById('tank-graph-to')||{}).value)||12;
  if(from>to){var tmp=from;from=to;to=tmp;}
  var q=String((document.getElementById('tank-graph-search')||{}).value||'').trim().toLocaleLowerCase('de-DE');
  return (TANK_DATA||[]).filter(function(r){
    var month=Number(r.monat)||0;
    if(year && Number(r.jahr)!==year)return false;
    if(month<from || month>to)return false;
    if(product && String(r.produkt||'Ohne Angabe').trim()!==product)return false;
    if(q){var hay=[r.fahrer,r.fahrzeug,r.fahrzeug_ia,r.produkt,r.zapfsaeule,r.firma].join(' ').toLocaleLowerCase('de-DE');if(hay.indexOf(q)<0)return false;}
    return true;
  });
}
function tankGraphFilterLabel(){
  var year=String((document.getElementById('tank-graph-year')||{}).value||'');
  var product=String((document.getElementById('tank-graph-product')||{}).value||'Alle Produkte');
  var from=Math.max(1,Number((document.getElementById('tank-graph-from')||{}).value)||1);
  var to=Math.min(12,Number((document.getElementById('tank-graph-to')||{}).value)||12);
  if(from>to){var tmp=from;from=to;to=tmp;}
  return {year:year,product:product||'Alle Produkte',from:from,to:to,months:TANK_MONTHS[from-1]+' bis '+TANK_MONTHS[to-1],search:String((document.getElementById('tank-graph-search')||{}).value||'').trim()};
}
function tankRenderGraph(){
  tankSetOptions();
  var rows=tankGraphRows(), filter=tankGraphFilterLabel();
  var empty=document.getElementById('tank-graph-empty'), grid=document.getElementById('tank-graph-grid');if(empty)empty.style.display=rows.length?'none':'block';if(grid)grid.style.display=rows.length?'grid':'none';
  Object.keys(tankCharts).forEach(tankDestroyChart);
  var totalLiters=rows.reduce(function(s,r){return s+tankN(r.menge_liter);},0), totalKm=rows.reduce(function(s,r){return s+(tankValidDistance(r)?tankN(r.km):0);},0), validCount=rows.filter(tankValidDistance).length;
  var vehicles=new Set(rows.map(function(r){return r.fahrzeug;}).filter(Boolean)), drivers=new Set(rows.map(function(r){return r.fahrer;}).filter(Boolean));
  var stats=document.getElementById('tank-graph-stats');if(stats)stats.textContent=tankFmt(rows.length,0)+' Vorgänge · '+tankFmt(vehicles.size,0)+' LKW · '+tankFmt(drivers.size,0)+' Fahrer';
  var kpis=document.getElementById('tank-graph-kpis');if(kpis)kpis.innerHTML=tankKpi('Tankvorgänge',tankFmt(rows.length,0),filter.months)+tankKpi('Liter gesamt',tankFmt(totalLiters,2)+' l',filter.product)+tankKpi('Plausible Strecke',tankFmt(totalKm,0)+' km','10 bis 5.000 km je Vorgang')+tankKpi('Ø Verbrauch',tankFmt(tankConsumption(rows),2)+' l/100 km','gewichteter Gesamtwert')+tankKpi('Ø Tankmenge',tankFmt(rows.length?totalLiters/rows.length:0,2)+' l','pro Tankvorgang')+tankKpi('Datenqualität',tankFmt(rows.length?validCount/rows.length*100:0,1)+' %','Tankungen mit plausibler Strecke');
  if(!rows.length)return;
  var monthly=tankMonthly(rows), rangeMonthly=monthly.filter(function(m){return m.month>=filter.from&&m.month<=filter.to;});
  var vehicleAg=tankAggregate(rows,'fahrzeug').filter(function(a){return a.name!=='Ohne Angabe';});
  var topVehicles=vehicleAg.slice(0,12);
  var topVehicleConsumption=vehicleAg.filter(function(a){return a.validKm>=200&&a.consumption>0;}).sort(function(a,b){return b.consumption-a.consumption;}).slice(0,12);
  var driverAg=tankAggregate(rows,'fahrer').filter(function(a){return a.name!=='Ohne Angabe';});
  var productAg=tankAggregate(rows,'produkt');
  var basePlugins={legend:{display:false},tooltip:{mode:'index',intersect:false}};
  var monthOptions={responsive:true,maintainAspectRatio:false,plugins:basePlugins,scales:{x:{grid:{display:false},ticks:{font:{size:10,weight:'bold'}}},y:{beginAtZero:true,ticks:{font:{size:10}}}}};
  var horizontalOptions={responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{legend:{display:false}},scales:{x:{beginAtZero:true,ticks:{font:{size:10}}},y:{grid:{display:false},ticks:{font:{size:10,weight:'bold'}}}}};
  var labels=rangeMonthly.map(function(m){return m.label;});
  var c1=document.getElementById('tank-chart-liters');if(c1)tankCharts.liters=new Chart(c1,{type:'bar',data:{labels:labels,datasets:[{label:'Liter',data:rangeMonthly.map(function(m){return m.liters;}),borderWidth:1,borderRadius:5}]},options:monthOptions});
  var c2=document.getElementById('tank-chart-consumption');if(c2)tankCharts.consumption=new Chart(c2,{type:'line',data:{labels:labels,datasets:[{label:'l/100 km',data:rangeMonthly.map(function(m){return m.consumption||null;}),tension:.28,spanGaps:true,pointRadius:4,borderWidth:2,fill:false}]},options:monthOptions});
  var c3=document.getElementById('tank-chart-distance');if(c3)tankCharts.distance=new Chart(c3,{type:'bar',data:{labels:labels,datasets:[{label:'km',data:rangeMonthly.map(function(m){return m.km;}),borderWidth:1,borderRadius:5}]},options:monthOptions});
  var c4=document.getElementById('tank-chart-count');if(c4)tankCharts.count=new Chart(c4,{type:'bar',data:{labels:labels,datasets:[{label:'Tankvorgänge',data:rangeMonthly.map(function(m){return m.count;}),borderWidth:1,borderRadius:5}]},options:monthOptions});
  var c5=document.getElementById('tank-chart-average');if(c5)tankCharts.average=new Chart(c5,{type:'line',data:{labels:labels,datasets:[{label:'Ø Liter',data:rangeMonthly.map(function(m){return m.average||null;}),tension:.28,spanGaps:true,pointRadius:4,borderWidth:2,fill:false}]},options:monthOptions});
  var c6=document.getElementById('tank-chart-products');if(c6)tankCharts.products=new Chart(c6,{type:'doughnut',data:{labels:productAg.map(function(a){return a.name;}),datasets:[{label:'Liter',data:productAg.map(function(a){return a.liters;}),borderWidth:1}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:true,position:'bottom',labels:{boxWidth:12,font:{size:10,weight:'bold'}}},tooltip:{callbacks:{label:function(ctx){return ctx.label+': '+tankFmt(ctx.raw,2)+' l';}}}}}});
  var c7=document.getElementById('tank-chart-vehicles');if(c7)tankCharts.vehicles=new Chart(c7,{type:'bar',data:{labels:topVehicles.map(function(a){return a.name;}),datasets:[{label:'Liter',data:topVehicles.map(function(a){return a.liters;}),borderWidth:1,borderRadius:4}]},options:horizontalOptions});
  var c8=document.getElementById('tank-chart-vehicle-consumption');if(c8)tankCharts.vehicleConsumption=new Chart(c8,{type:'bar',data:{labels:topVehicleConsumption.map(function(a){return a.name;}),datasets:[{label:'l/100 km',data:topVehicleConsumption.map(function(a){return a.consumption;}),borderWidth:1,borderRadius:4}]},options:horizontalOptions});
  var driverWrap=document.getElementById('tank-driver-chart-wrap');if(driverWrap)driverWrap.style.height=Math.max(420,driverAg.length*26+70)+'px';
  var c9=document.getElementById('tank-chart-drivers');if(c9)tankCharts.drivers=new Chart(c9,{type:'bar',data:{labels:driverAg.map(function(a){return a.name;}),datasets:[{label:'Liter',data:driverAg.map(function(a){return a.liters;}),borderWidth:1,borderRadius:4}]},options:horizontalOptions});
}
window.tankInitGraph=function(){tankSetOptions();tankRenderGraph();};

// ── Tankauswertung: Excel-Export ─────────────────────────────────────────────
function tankExcelSafeName(value){return String(value||'').replace(/[\\/:*?"<>|]+/g,'_').replace(/\s+/g,'_').replace(/^_+|_+$/g,'').slice(0,45)||'Auswertung';}
function tankExcelSheet(lib,data,widths,autoFilter){
  var ws=lib.utils.aoa_to_sheet(data);
  ws['!cols']=(widths||[]).map(function(w){return {wch:w};});
  if(autoFilter && data.length>1)ws['!autofilter']={ref:'A1:'+lib.utils.encode_cell({r:data.length-1,c:data[0].length-1})};
  if(data.length){
    for(var c=0;c<data[0].length;c++){
      var ref=lib.utils.encode_cell({r:0,c:c});
      if(ws[ref])ws[ref].s={font:{bold:true,color:{rgb:'FFFFFF'}},fill:{fgColor:{rgb:'9A5B00'}},alignment:{horizontal:'center',vertical:'center'}};
    }
  }
  return ws;
}
function tankExportExcel(){
  var lib=window.XLSXStyle||window.XLSX;
  if(!lib||!lib.utils){alert('Excel-Bibliothek nicht geladen. Bitte die Seite neu laden.');return;}
  var rows=tankGraphRows(), filter=tankGraphFilterLabel();
  if(!rows.length){alert('Für die aktuelle Auswahl sind keine Tankdaten vorhanden.');return;}
  var totalLiters=rows.reduce(function(s,r){return s+tankN(r.menge_liter);},0), totalKm=rows.reduce(function(s,r){return s+(tankValidDistance(r)?tankN(r.km):0);},0), validCount=rows.filter(tankValidDistance).length;
  var vehicles=new Set(rows.map(function(r){return r.fahrzeug;}).filter(Boolean)), drivers=new Set(rows.map(function(r){return r.fahrer;}).filter(Boolean));
  var wb=lib.utils.book_new();
  var overview=[
    ['Tankauswertung','Wert','Einheit / Hinweis'],
    ['Jahr',filter.year,''],
    ['Zeitraum',filter.months,''],
    ['Produkt',filter.product,''],
    ['Suchfilter',filter.search||'Kein Filter',''],
    ['Tankvorgänge',rows.length,'Anzahl'],
    ['Liter gesamt',totalLiters,'Liter'],
    ['Plausible Strecke',totalKm,'km (10 bis 5.000 km je Tankung)'],
    ['Ø Verbrauch',tankConsumption(rows),'l/100 km'],
    ['Ø Tankmenge',rows.length?totalLiters/rows.length:0,'Liter je Tankvorgang'],
    ['Datenqualität',rows.length?validCount/rows.length*100:0,'% mit plausibler Strecke'],
    ['Fahrzeuge',vehicles.size,'Anzahl'],
    ['Fahrer',drivers.size,'Anzahl']
  ];
  var wsOverview=tankExcelSheet(lib,overview,[28,22,38],false);
  ['B7','B8','B9','B10','B11'].forEach(function(ref){if(wsOverview[ref])wsOverview[ref].z='#,##0.00';});
  lib.utils.book_append_sheet(wb,wsOverview,'Übersicht');

  var monthly=tankMonthly(rows).filter(function(m){return m.month>=filter.from&&m.month<=filter.to;});
  var monthData=[['Monat','Tankvorgänge','Liter','Plausible km','Ø l/100 km','Ø Liter/Tankung','Fahrzeuge','Fahrer']].concat(monthly.map(function(m){return [m.label,m.count,m.liters,m.km,m.consumption,m.average,m.vehicles,m.drivers];}));
  var wsMonths=tankExcelSheet(lib,monthData,[16,15,14,16,15,19,12,12],true);
  lib.utils.book_append_sheet(wb,wsMonths,'Monate');

  function aggregateData(key,label){
    return [[label,'Tankvorgänge','Liter','Plausible km','Ø l/100 km','Ø Liter/Tankung','Gültige km-Daten %','Erste Tankung','Letzte Tankung','Fahrer','LKW']].concat(tankAggregate(rows,key).map(function(a){return [a.name,a.count,a.liters,a.validKm,a.consumption,a.avg,a.validShare,tankDateLabel(a.first),tankDateLabel(a.last),Array.from(a.drivers).sort().join(', '),Array.from(a.vehicles).sort().join(', ')];}));
  }
  lib.utils.book_append_sheet(wb,tankExcelSheet(lib,aggregateData('fahrzeug','LKW'),[22,14,14,16,15,18,18,14,14,35,30],true),'Fahrzeuge');
  lib.utils.book_append_sheet(wb,tankExcelSheet(lib,aggregateData('fahrer','Fahrer'),[30,14,14,16,15,18,18,14,14,35,30],true),'Fahrer');
  lib.utils.book_append_sheet(wb,tankExcelSheet(lib,aggregateData('produkt','Produkt'),[28,14,14,16,15,18,18,14,14,35,30],true),'Produkte');

  var sorted=rows.slice().sort(function(a,b){return String(a.datetime_iso||'').localeCompare(String(b.datetime_iso||''));});
  var detail=[['Datum','Uhrzeit','Firma / Spedition','Kennzeichen','Fahrzeug IA','Fahrer','Produkt','Liter','Kilometerstand','km','l/100 km','Zapfsäule','Quelldatei']].concat(sorted.map(function(r){return [tankDateLabel(r.date_iso),String(r.uhrzeit||'').slice(0,5),r.firma||'',r.fahrzeug||'',r.fahrzeug_ia||'',r.fahrer||'',r.produkt||'',tankN(r.menge_liter),tankN(r.kilometerzaehler),tankN(r.km),tankValidDistance(r)?tankN(r.menge_liter)/tankN(r.km)*100:'',r.zapfsaeule||'',r.quelle||''];}));
  lib.utils.book_append_sheet(wb,tankExcelSheet(lib,detail,[13,9,24,18,18,30,20,12,16,12,13,14,32],true),'Tankvorgänge');

  var filename='Tankauswertung_'+tankExcelSafeName(filter.year)+'_'+String(filter.from).padStart(2,'0')+'-'+String(filter.to).padStart(2,'0');
  if(filter.product!=='Alle Produkte')filename+='_'+tankExcelSafeName(filter.product);
  lib.writeFile(wb,filename+'.xlsx');
}
window.tankExportExcel=tankExportExcel;
"""



def _zulagen_graph_js() -> str:
    """Erweitert die bestehende Zulagenansicht um einen Chart.js-Graphmodus."""
    return r"""
// ── Zulagen-Graph ─────────────────────────────────────────────────────────────
(function(){
  var ZG_VIEW = 'list';
  var ZG_TAB = 'sonder';
  var ZG_CHARTS = {months:null, drivers:null};
  var ZG_TIMER = null;

  function zgNumber(value){
    var n = Number(value);
    return Number.isFinite(n) ? n : 0;
  }
  function zgMoney(value){
    return zgNumber(value).toLocaleString('de-DE',{minimumFractionDigits:2,maximumFractionDigits:2})+' €';
  }
  function zgLabel(tab){
    if(tab==='fuengers') return 'Füngers';
    if(tab==='drittkunden') return 'Drittkunden';
    return 'Sonderfahrzeuge';
  }
  function zgMonths(tab){
    if(tab==='drittkunden') return Array.isArray(DRITTKUNDEN_DATA) ? DRITTKUNDEN_DATA : [];
    var root = (ZULAGE_DATA && typeof ZULAGE_DATA==='object') ? ZULAGE_DATA : {};
    return Array.isArray(root[tab]) ? root[tab] : [];
  }
  function zgMonthTotal(month){
    return (month && Array.isArray(month.fahrer) ? month.fahrer : []).reduce(function(sum,f){return sum+zgNumber(f.gesamt);},0);
  }
  function zgMonthEntries(month){
    return (month && Array.isArray(month.fahrer) ? month.fahrer : []).reduce(function(sum,f){return sum+(Array.isArray(f.tage)?f.tage.length:0);},0);
  }
  function zgDetectTab(){
    if(ZG_TAB) return ZG_TAB;
    ['sonder','fuengers','drittkunden'].some(function(tab){
      var el=document.getElementById('ztab-'+tab);
      if(!el) return false;
      var bg=(window.getComputedStyle?getComputedStyle(el).backgroundColor:el.style.backgroundColor)||'';
      if(bg && bg!=='rgb(255, 255, 255)' && bg!=='rgba(0, 0, 0, 0)' && bg!=='transparent'){
        ZG_TAB=tab; return true;
      }
      return false;
    });
    return ZG_TAB || 'sonder';
  }
  function zgSelectedMonth(months){
    if(!months.length) return null;
    var sel=document.getElementById('zulage-month-sel');
    if(!sel) return months[months.length-1];
    var value=String(sel.value||'').trim();
    var text=sel.options && sel.selectedIndex>=0 ? String(sel.options[sel.selectedIndex].textContent||'').trim() : '';
    var found=months.find(function(m){return String(m.monat||'').trim()===value || String(m.monat||'').trim()===text;});
    if(found) return found;
    var idx=Number(value);
    if(Number.isInteger(idx)){
      if(idx>=0 && idx<months.length) return months[idx];
      if(idx>0 && idx<=months.length) return months[idx-1];
    }
    if(sel.options && sel.options.length===months.length && sel.selectedIndex>=0 && sel.selectedIndex<months.length) return months[sel.selectedIndex];
    if(sel.options && sel.options.length===months.length+1 && sel.selectedIndex>0) return months[sel.selectedIndex-1] || months[months.length-1];
    return months[months.length-1];
  }
  function zgDestroy(key){
    if(ZG_CHARTS[key]){ZG_CHARTS[key].destroy();ZG_CHARTS[key]=null;}
  }
  function zgLayout(){
    var grid=document.getElementById('zulage-graph-grid');
    if(grid) grid.style.gridTemplateColumns=(window.innerWidth<1120)?'minmax(0,1fr)':'minmax(0,1fr) minmax(0,1fr)';
  }
  function zgKpi(title,value,sub){
    return '<div style="background:#fff;border:1px solid #d8e0ea;border-radius:12px;padding:13px 15px;box-shadow:0 3px 10px rgba(15,23,42,.05);min-width:0;">'
      +'<div style="font-size:10px;font-weight:900;text-transform:uppercase;letter-spacing:.45px;color:#64748b;">'+title+'</div>'
      +'<div style="font-size:22px;font-weight:950;color:#0f172a;letter-spacing:-.45px;margin-top:3px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">'+value+'</div>'
      +'<div style="font-size:10px;font-weight:700;color:#94a3b8;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">'+sub+'</div></div>';
  }
  function zgRender(){
    if(ZG_VIEW!=='graph') return;
    zgLayout();
    var tab=zgDetectTab(), months=zgMonths(tab), selected=zgSelectedMonth(months);
    var graphGrid=document.getElementById('zulage-graph-grid');
    var empty=document.getElementById('zulage-graph-empty');
    var kpis=document.getElementById('zulage-graph-kpis');
    var hasData=months.some(function(m){return (m.fahrer||[]).length>0;});
    if(graphGrid) graphGrid.style.display=hasData?'grid':'none';
    if(empty) empty.style.display=hasData?'none':'block';
    zgDestroy('months'); zgDestroy('drivers');
    if(!hasData){if(kpis)kpis.innerHTML='';return;}

    var allTotal=months.reduce(function(sum,m){return sum+zgMonthTotal(m);},0);
    var currentTotal=zgMonthTotal(selected);
    var currentDrivers=(selected&&Array.isArray(selected.fahrer)?selected.fahrer:[]).filter(function(f){return zgNumber(f.gesamt)!==0;});
    var currentEntries=zgMonthEntries(selected);
    var avg=currentDrivers.length?currentTotal/currentDrivers.length:0;
    if(kpis) kpis.innerHTML=
      zgKpi('Auswahl',zgLabel(tab),(selected&&selected.monat)||'Kein Monat')+
      zgKpi('Monatssumme',zgMoney(currentTotal),currentEntries.toLocaleString('de-DE')+' Einsätze')+
      zgKpi('Fahrer',currentDrivers.length.toLocaleString('de-DE'),'mit Zulage im Monat')+
      zgKpi('Ø je Fahrer',zgMoney(avg),'im ausgewählten Monat')+
      zgKpi('Gesamt',zgMoney(allTotal),months.length.toLocaleString('de-DE')+' Monate');

    var monthSub=document.getElementById('zulage-month-chart-sub');
    if(monthSub) monthSub.textContent=zgLabel(tab)+' · Gesamtsumme je Monat';
    var driverSub=document.getElementById('zulage-driver-chart-sub');
    if(driverSub) driverSub.textContent=zgLabel(tab)+' · '+((selected&&selected.monat)||'Kein Monat');

    var monthCanvas=document.getElementById('zulage-chart-months');
    if(monthCanvas && window.Chart){
      ZG_CHARTS.months=new Chart(monthCanvas,{type:'bar',data:{labels:months.map(function(m){return m.monat||'';}),datasets:[{label:'Zulagen',data:months.map(zgMonthTotal),borderWidth:1,borderRadius:6}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},tooltip:{callbacks:{label:function(ctx){return ' '+zgMoney(ctx.raw);}}}},scales:{x:{grid:{display:false},ticks:{font:{size:10,weight:'bold'},maxRotation:45,minRotation:0}},y:{beginAtZero:true,ticks:{font:{size:10},callback:function(v){return Number(v).toLocaleString('de-DE')+' €';}}}}}});
    }

    currentDrivers.sort(function(a,b){return zgNumber(b.gesamt)-zgNumber(a.gesamt) || String(a.name||'').localeCompare(String(b.name||''),'de');});
    var driverWrap=document.getElementById('zulage-driver-chart-wrap');
    if(driverWrap) driverWrap.style.height=Math.max(420,currentDrivers.length*31+75)+'px';
    var driverCanvas=document.getElementById('zulage-chart-drivers');
    if(driverCanvas && window.Chart){
      ZG_CHARTS.drivers=new Chart(driverCanvas,{type:'bar',data:{labels:currentDrivers.map(function(f){return f.name||'Ohne Angabe';}),datasets:[{label:'Zulagen',data:currentDrivers.map(function(f){return zgNumber(f.gesamt);}),borderWidth:1,borderRadius:5}]},options:{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{legend:{display:false},tooltip:{callbacks:{label:function(ctx){return ' '+zgMoney(ctx.raw);}}}},scales:{x:{beginAtZero:true,ticks:{font:{size:10},callback:function(v){return Number(v).toLocaleString('de-DE')+' €';}}},y:{grid:{display:false},ticks:{font:{size:10,weight:'bold'},autoSkip:false}}}}});
    }
  }
  function zgSchedule(){
    if(ZG_TIMER) clearTimeout(ZG_TIMER);
    ZG_TIMER=setTimeout(function(){ZG_TIMER=null;zgRender();},20);
  }

  window.zulagenSetView=function(view){
    ZG_VIEW=(view==='graph')?'graph':'list';
    var list=document.getElementById('zulage-content'), graph=document.getElementById('zulage-graph-content');
    var listBtn=document.getElementById('zulage-view-list'), graphBtn=document.getElementById('zulage-view-graph');
    if(list) list.style.display=ZG_VIEW==='list'?'block':'none';
    if(graph) graph.style.display=ZG_VIEW==='graph'?'block':'none';
    if(listBtn){listBtn.style.background=ZG_VIEW==='list'?'#1b66b3':'#fff';listBtn.style.color=ZG_VIEW==='list'?'#fff':'#1b66b3';}
    if(graphBtn){graphBtn.style.background=ZG_VIEW==='graph'?'#1b66b3':'#fff';graphBtn.style.color=ZG_VIEW==='graph'?'#fff':'#1b66b3';}
    if(ZG_VIEW==='graph') requestAnimationFrame(zgSchedule);
  };

  function zgBindControls(){
    ['sonder','fuengers','drittkunden'].forEach(function(tab){
      var btn=document.getElementById('ztab-'+tab);
      if(btn && !btn.dataset.zgBound){
        btn.dataset.zgBound='1';
        btn.addEventListener('click',function(){ZG_TAB=tab;zgSchedule();});
      }
    });
    var sel=document.getElementById('zulage-month-sel');
    if(sel && !sel.dataset.zgBound){sel.dataset.zgBound='1';sel.addEventListener('change',zgSchedule);}
  }

  var originalInit=window.zulagenInit;
  if(typeof originalInit==='function') window.zulagenInit=function(){
    var result=originalInit.apply(this,arguments); ZG_TAB='sonder'; zgBindControls(); zgSchedule(); return result;
  };
  var originalTab=window.zulagenTab;
  if(typeof originalTab==='function') window.zulagenTab=function(tab){
    ZG_TAB=tab||'sonder'; var result=originalTab.apply(this,arguments); zgSchedule(); return result;
  };
  var originalRender=window.zulagenRender;
  if(typeof originalRender==='function') window.zulagenRender=function(){
    var result=originalRender.apply(this,arguments); zgSchedule(); return result;
  };
  function zgReady(){zgBindControls();zgLayout();}
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',zgReady);
  else zgReady();
  window.addEventListener('resize',function(){zgLayout();if(ZG_VIEW==='graph')zgSchedule();});
})();
// ── /Zulagen-Graph ────────────────────────────────────────────────────────────
"""



def _render_dashboard_html(
    *,
    logo_data_url: str,
    last_updated: str,
    generation_meta_json: str,
    instances_js: str,
    embedded_data_js: str,
    documents_js_code: str,
    knapp_js_code: str,
    spesen_js_code: str,
    fa_js_code: str,
    zulage_js_code: str,
    wash_js_code: str,
    wash_ranking_js_code: str,
    fw_graph_js_code: str,
    tank_panels_html: str,
    tank_js_code: str,
    verstoss_js_code: str,
    sped_js_code: str,
    fabew_js_code: str,
    bus_js_code: str,
    arzt_js_code: str,
    versp_js_code: str,
    wa_js_code: str,
    zulage_xlsx_sonder: str,
    zulage_xlsx_fuengers: str,
    zulage_xlsx_drittkunden: str,
) -> str:
    """Setzt die vorbereiteten Daten und Skriptbausteine in das HTML ein."""
    return f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Fuhrpark NFC</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<script>window.__XLSX_CE__ = window.XLSX;</script>
<script src="https://cdn.jsdelivr.net/npm/xlsx-js-style@1.2.0/dist/xlsx.bundle.js"></script>
<script>window.XLSXStyle = window.XLSX; if(window.__XLSX_CE__) window.XLSX = window.__XLSX_CE__;</script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.1/jspdf.umd.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf-autotable/3.8.2/jspdf.plugin.autotable.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%;font-family:'Segoe UI',Arial,sans-serif}}
.topnav{{
  height:56px;
  background:linear-gradient(180deg,#eef2f6 0%,#dde4eb 100%);
  display:flex;align-items:center;padding:0 12px;gap:4px;
  box-shadow:0 2px 10px rgba(15,23,42,.08);
  border-bottom:1px solid #c5ced8;
  flex-shrink:0;
  overflow-x:auto;
  scrollbar-width:none;
}}
.topnav::-webkit-scrollbar{{display:none;}}
.topnav-logo-wrap{{
  display:flex;align-items:center;flex-shrink:0;
  padding-right:4px;
}}
.topnav-logo{{
  height:28px;
  width:auto;
  display:block;
  object-fit:contain;
}}
.nav-sep{{
  width:1px;height:22px;background:#c8d1db;flex-shrink:0;margin:0 4px;
}}
.nav-btn{{
  padding:6px 9px;border-radius:8px;
  border:1px solid #bcc8d6;
  cursor:pointer;font-weight:800;font-size:12px;
  transition:all .15s ease;background:linear-gradient(180deg,#f9fbfd 0%,#edf2f7 100%);color:#334155;
  white-space:nowrap;flex-shrink:0;
  position:relative;
  box-shadow:0 1px 2px rgba(15,23,42,.05);
}}
.nav-btn:hover:not(.active){{background:linear-gradient(180deg,#ffffff 0%,#eef3f8 100%);border-color:#aeb9c8}}
.nav-btn.active{{background:linear-gradient(180deg,#4f87e8 0%,#3d72d4 100%);border-color:#3f73cf;color:#fff;box-shadow:0 3px 10px rgba(61,114,212,.25)}}
.inst-btn{{border-color:#c5cfda;color:#5b6b80}}
.inst-btn:hover:not(.active){{background:linear-gradient(180deg,#ffffff 0%,#eef3f8 100%)}}
.inst-btn.active{{background:linear-gradient(180deg,#6b7f99 0%,#55677f 100%);border-color:#55677f;color:#fff}}
/* ── Dropdown ── */
.nav-dd{{position:relative;flex-shrink:0}}
.nav-dd-btn{{
  padding:6px 9px;border-radius:8px;
  border:1px solid #bcc8d6;
  cursor:pointer;font-weight:800;font-size:12px;
  transition:all .15s ease;background:linear-gradient(180deg,#f9fbfd 0%,#edf2f7 100%);color:#334155;
  white-space:nowrap;display:flex;align-items:center;gap:4px;
  box-shadow:0 1px 2px rgba(15,23,42,.05);
}}
.nav-dd-btn:hover{{background:linear-gradient(180deg,#ffffff 0%,#eef3f8 100%);border-color:#aeb9c8}}
.nav-dd-btn.active{{background:linear-gradient(180deg,#4f87e8 0%,#3d72d4 100%);border-color:#3f73cf;color:#fff;box-shadow:0 3px 10px rgba(61,114,212,.25)}}
#btn-suche.active{{background:linear-gradient(180deg,#f6dc67 0%,#e6be22 100%);border-color:#d5ac10;color:#334155;box-shadow:0 3px 10px rgba(214,172,16,.28)}}
#btn-suche.active .dd-arrow{{filter:none;opacity:.9}}
.dd-arrow{{font-size:9px;opacity:.75;transition:transform .15s}}
.inst-label{{font-size:9px;font-weight:700;opacity:1;background:#dbe5f0;border:1px solid #c0cad8;border-radius:5px;padding:1px 5px;margin:0 2px 0 4px;white-space:nowrap;color:#4b5d73}}
.nav-dd.open .dd-arrow{{transform:rotate(180deg)}}
.dd-menu{{
  display:none;
  position:fixed;
  background:#f7f9fc;border:1px solid #c5ced8;
  border-radius:8px;min-width:170px;
  box-shadow:0 10px 24px rgba(15,23,42,.12);
  z-index:99999;overflow:hidden;
}}
.nav-dd.open .dd-menu{{display:block}}
.dd-item{{
  padding:9px 14px;cursor:pointer;font-size:12px;font-weight:700;
  color:#334155;transition:background .12s;white-space:nowrap;
  border-bottom:1px solid #e2e8f0;
}}
.dd-item:last-child{{border-bottom:none}}
.dd-item:hover{{background:#eaf3fb}}
.dd-item.active{{background:linear-gradient(180deg,#2f80b7 0%,#1e6091 100%);color:#fff}}
.topnav-meta-btn{{
  margin-left:auto;flex-shrink:0;white-space:nowrap;
  padding:5px 8px;border-radius:7px;border:1px solid #b9c5d2;
  background:linear-gradient(180deg,#ffffff 0%,#edf2f7 100%);
  color:#415269;font-size:10px;font-weight:900;cursor:pointer;
  box-shadow:0 1px 2px rgba(15,23,42,.05);
}}
.topnav-meta-btn:hover{{background:#fff;border-color:#8fa1b4}}
.topnav-stamp{{
  font-size:10px;font-weight:800;color:#5b6b80;white-space:nowrap;
  position:sticky;right:0;flex-shrink:0;padding:0 2px 0 4px;background:#e4e9ef;
}}
.build-info-overlay{{display:none;position:fixed;inset:0;z-index:200000;background:rgba(15,23,42,.55);padding:24px;align-items:center;justify-content:center}}
.build-info-overlay.open{{display:flex}}
.build-info-dialog{{width:min(920px,96vw);max-height:88vh;overflow:auto;background:#f8fafc;border:1px solid #b9c6d5;border-radius:14px;box-shadow:0 24px 70px rgba(15,23,42,.30)}}
.build-info-head{{position:sticky;top:0;z-index:2;display:flex;align-items:center;justify-content:space-between;gap:12px;padding:16px 18px;background:linear-gradient(180deg,#ffffff 0%,#eef3f8 100%);border-bottom:1px solid #d5dde7}}
.build-info-title{{font-size:17px;font-weight:950;color:#0f172a}}
.build-info-sub{{font-size:11px;color:#64748b;font-weight:700;margin-top:2px}}
.build-info-close{{width:34px;height:34px;border-radius:8px;border:1px solid #c4ceda;background:#fff;color:#334155;font-size:19px;font-weight:900;cursor:pointer}}
.build-info-body{{padding:16px 18px 20px}}
.build-info-cards{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:9px;margin-bottom:16px}}
.build-info-card{{background:#fff;border:1px solid #d7e0ea;border-radius:10px;padding:11px 12px}}
.build-info-card-label{{font-size:9px;font-weight:900;text-transform:uppercase;letter-spacing:.45px;color:#64748b}}
.build-info-card-value{{font-size:16px;font-weight:950;color:#0f172a;margin-top:4px;overflow-wrap:anywhere}}
.build-info-section{{background:#fff;border:1px solid #d7e0ea;border-radius:10px;overflow:hidden;margin-top:10px}}
.build-info-section h3{{font-size:11px;text-transform:uppercase;letter-spacing:.45px;color:#475569;padding:10px 12px;background:#f1f5f9;border-bottom:1px solid #dde5ee}}
.build-info-table{{width:100%;border-collapse:collapse;font-size:11px}}
.build-info-table th,.build-info-table td{{padding:8px 10px;text-align:left;border-bottom:1px solid #edf1f5;vertical-align:top}}
.build-info-table th{{font-size:9px;text-transform:uppercase;letter-spacing:.35px;color:#64748b;background:#fbfcfe}}
.build-info-table tr:last-child td{{border-bottom:none}}
.build-info-pill{{display:inline-flex;padding:2px 6px;border-radius:999px;font-size:9px;font-weight:900;white-space:nowrap}}
.build-info-pill.ok{{background:#dcfce7;color:#166534}}
.build-info-pill.warning{{background:#fef3c7;color:#92400e}}
.build-info-pill.error{{background:#fee2e2;color:#991b1b}}
.build-info-pill.info{{background:#e2e8f0;color:#475569}}
@media(max-width:720px){{.build-info-cards{{grid-template-columns:repeat(2,minmax(0,1fr))}}.topnav-stamp{{display:none}}}}
.frame-wrap{{height:calc(100vh - 56px);display:flex;flex-direction:column}}
iframe{{flex:1;width:100%;border:none;display:none}}
iframe.active{{display:block}}
.vz-day-btn{{padding:7px 14px;border:1.5px solid #9db9d5;background:#fff;color:#1e6091;border-radius:7px;cursor:pointer;font-weight:800;font-size:12px;font-family:'Segoe UI',Arial,sans-serif;transition:all .15s;letter-spacing:.1px}}
.vz-day-btn:hover{{background:#eaf3fb;border-color:#1e6091}}
.vz-day-btn.active{{background:linear-gradient(180deg,#2f80b7 0%,#1e6091 100%);color:#fff;border-color:#164e75;box-shadow:0 2px 7px rgba(30,96,145,.28)}}
</style>
</head>
<body>

<nav class="topnav">
  <div class="topnav-logo-wrap">
    <img class="topnav-logo" src="{logo_data_url}" alt="Nordfrische Center Logo">
  </div>
  <div class="nav-sep"></div>
  <div class="nav-dd" id="dd-suche">
    <button class="nav-dd-btn active" id="btn-suche" onclick="ddToggle('suche',event)">
      &#128269; Suche <span id="inst-label-suche"></span><span class="dd-arrow">&#9660;</span>
    </button>
    <div class="dd-menu" id="ddmenu-suche"></div>
  </div>
  <div class="nav-dd" id="dd-vz">
    <button class="nav-dd-btn" id="btn-vz" onclick="ddToggle('vz',event)">
      &#128703; Waschen &amp; Tanken <span id="inst-label-vz"></span><span class="dd-arrow">&#9660;</span>
    </button>
    <div class="dd-menu" id="ddmenu-vz"></div>
  </div>
  <div class="nav-dd" id="dd-verstoss">
    <button class="nav-dd-btn" id="btn-verstoss" onclick="ddToggle('verstoss',event)">
      &#9888;&#65039; Versto&#223;auswertung <span class="dd-arrow">&#9660;</span>
    </button>
    <div class="dd-menu" id="ddmenu-verstoss"></div>
  </div>
  <div class="nav-dd" id="dd-sam">
    <button class="nav-dd-btn" id="btn-sam" onclick="ddToggle('sam',event)">
      &#128664; Sa + So Einsätze <span class="dd-arrow">&#9660;</span>
    </button>
    <div class="dd-menu" id="ddmenu-sam"></div>
  </div>
  <div class="nav-dd" id="dd-fa">
    <button class="nav-dd-btn" id="btn-fa" onclick="ddToggle('fa',event)">
      &#128101; Fahrerauswertung <span class="dd-arrow">&#9660;</span>
    </button>
    <div class="dd-menu" id="ddmenu-fa"></div>
  </div>
  <button class="nav-btn" id="btn-zulage" onclick="showArea('zulage')">&#128176; Zulagen</button>
  <button class="nav-btn" id="btn-spesen" onclick="showArea('spesen')">&#128181; Spesen</button>
  <button class="nav-btn" id="btn-gk" onclick="showArea('gk')">&#127970; Gro&#223;kunden</button>
  <div class="nav-dd" id="dd-sped">
    <button class="nav-dd-btn" id="btn-sped" onclick="ddToggle('sped',event)">
      &#128666; Spediteure <span class="dd-arrow">&#9660;</span>
    </button>
    <div class="dd-menu" id="ddmenu-sped"></div>
  </div>
  <div class="nav-dd" id="dd-wa">
    <button class="nav-dd-btn" id="btn-wa" onclick="ddToggle('wa',event)">
      &#128202; Wochenauslastung <span class="dd-arrow">&#9660;</span>
    </button>
    <div class="dd-menu" id="ddmenu-wa"></div>
  </div>
  <div class="nav-dd" id="dd-infos">
    <button class="nav-dd-btn" id="btn-infos" onclick="ddToggle('infos',event)">
      &#128203; Infos &amp; Aush&#228;nge <span class="dd-arrow">&#9660;</span>
    </button>
    <div class="dd-menu" id="ddmenu-infos"></div>
  </div>
  </div>
  <span class="topnav-stamp">v{APP_DISPLAY_VERSION} &middot; {last_updated}</span>
</nav>

<div class="build-info-overlay" id="buildInfoOverlay" onclick="buildInfoBackdrop(event)">
  <div class="build-info-dialog" role="dialog" aria-modal="true" aria-labelledby="buildInfoTitle">
    <div class="build-info-head">
      <div><div class="build-info-title" id="buildInfoTitle">Datenstand und Version</div><div class="build-info-sub">Technische Informationen zur aktuell geöffneten Einzeldatei</div></div>
      <button class="build-info-close" type="button" onclick="closeBuildInfo()" aria-label="Schliessen">&times;</button>
    </div>
    <div class="build-info-body" id="buildInfoContent"></div>
  </div>
</div>
<script>
const APP_GENERATION_META = {generation_meta_json};
function buildInfoEsc(value){{return String(value==null?"":value).replace(/[&<>"']/g,function(ch){{return {{"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}}[ch];}});}}
function buildInfoBytes(value){{var n=Number(String(value||"0").replace(/^H/,""))||0;if(n<1024)return n+" B";if(n<1048576)return(n/1024).toFixed(1)+" KB";return(n/1048576).toFixed(1)+" MB";}}
function buildInfoDuration(value){{var n=Number(value)||0;return n<1000?n+" ms":(n/1000).toFixed(2)+" s";}}
function renderBuildInfo(){{
  var m=APP_GENERATION_META||{{}},stats=m.stats||{{}};
  var cards=[["Version","v"+(m.version||"-")],["Erstellt",m.created_at||"-"],["HTML-Datei",buildInfoBytes(m.html_bytes)],["Erzeugung",buildInfoDuration(m.generation_ms)]];
  var html='<div class="build-info-cards">'+cards.map(function(c){{return '<div class="build-info-card"><div class="build-info-card-label">'+buildInfoEsc(c[0])+'</div><div class="build-info-card-value">'+buildInfoEsc(c[1])+'</div></div>';}}).join('')+'</div>';
  var datasets=Array.isArray(m.datasets)?m.datasets:[];
  if(datasets.length)html+='<div class="build-info-section"><h3>Datenumfang</h3><table class="build-info-table"><thead><tr><th>Bereich</th><th>Umfang</th><th>Hinweis</th></tr></thead><tbody>'+datasets.map(function(d){{return '<tr><td><b>'+buildInfoEsc(d.label)+'</b></td><td>'+buildInfoEsc(d.value)+'</td><td>'+buildInfoEsc(d.detail||'')+'</td></tr>';}}).join('')+'</tbody></table></div>';
  var sources=Array.isArray(m.sources)?m.sources:[];
  if(sources.length)html+='<div class="build-info-section"><h3>Verarbeitete Quellen</h3><table class="build-info-table"><thead><tr><th>Status</th><th>Bereich</th><th>Datei</th><th>Details</th><th>Zeit</th></tr></thead><tbody>'+sources.map(function(r){{var st=r.status||'info';var label=st==='ok'?'OK':st==='warning'?'Hinweis':st==='error'?'Fehler':'Info';return '<tr><td><span class="build-info-pill '+buildInfoEsc(st)+'">'+buildInfoEsc(label)+'</span></td><td><b>'+buildInfoEsc(r.area||'')+'</b></td><td>'+buildInfoEsc(r.file||'')+'</td><td>'+buildInfoEsc(r.detail||'')+'</td><td>'+buildInfoEsc(r.time||'')+'</td></tr>';}}).join('')+'</tbody></table></div>';
  html+='<div class="build-info-section"><h3>Zusammenfassung</h3><table class="build-info-table"><tbody><tr><td><b>Wochen</b></td><td>'+buildInfoEsc((m.weeks||[]).join(', ')||'Keine')+'</td></tr><tr><td><b>Eingebettete PDFs</b></td><td>'+buildInfoEsc(stats.pdf_count||0)+' Dokumente, '+buildInfoEsc(buildInfoBytes(stats.pdf_compressed_bytes||0))+' komprimiert</td></tr><tr><td><b>Quellenstatus</b></td><td>'+buildInfoEsc(stats.ok||0)+' erfolgreich, '+buildInfoEsc(stats.warning||0)+' Hinweise, '+buildInfoEsc(stats.error||0)+' Fehler</td></tr></tbody></table></div>';
  var target=document.getElementById('buildInfoContent');if(target)target.innerHTML=html;
}}
function openBuildInfo(){{renderBuildInfo();var el=document.getElementById('buildInfoOverlay');if(el)el.classList.add('open');}}
function closeBuildInfo(){{var el=document.getElementById('buildInfoOverlay');if(el)el.classList.remove('open');}}
function buildInfoBackdrop(event){{if(event&&event.target&&event.target.id==='buildInfoOverlay')closeBuildInfo();}}
document.addEventListener('keydown',function(e){{if(e.key==='Escape')closeBuildInfo();}});
</script>

<div class="frame-wrap">
  <iframe id="frame-suche" class="active" title="Kunden-Suche"></iframe>
  <iframe id="frame-druck" title="Druckbereich" style="display:none!important;width:0;height:0;border:0"></iframe>
  <div id="panel-vz" style="display:none;flex:1;overflow-y:auto;padding:18px 18px 28px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif">
    <div style="width:100%;max-width:1728px;margin:0 auto">

      <!-- ── Header-Karte mit Titel + Aktionen ─────────────────────────────── -->
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:18px 22px;box-shadow:0 2px 10px rgba(30,96,145,.08);margin-bottom:18px">
        <div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin-bottom:16px">
          <div style="display:flex;align-items:center;gap:12px;min-width:0;flex:1">
            <div style="width:40px;height:40px;border-radius:10px;background:linear-gradient(135deg,#1e6091 0%,#2f80b7 100%);display:flex;align-items:center;justify-content:center;font-size:20px;box-shadow:0 2px 7px rgba(30,96,145,.25);flex-shrink:0">&#128703;</div>
            <div style="min-width:0">
              <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Fahrzeugw&#228;sche</h2>
              <p style="color:#64748b;font-size:12px;margin:2px 0 0 0;font-weight:500">Tag und Datum ausw&#228;hlen, dann PDF exportieren</p>
            </div>
          </div>
          <button onclick="fwExportPdf()" style="padding:10px 18px;background:linear-gradient(180deg,#ef4444 0%,#dc2626 100%);color:#fff;border:none;border-radius:8px;font-weight:800;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(220,38,38,.28);display:inline-flex;align-items:center;gap:7px;transition:all .15s" onmouseover="this.style.transform='translateY(-1px)';this.style.boxShadow='0 4px 10px rgba(220,38,38,.35)'" onmouseout="this.style.transform='';this.style.boxShadow='0 2px 6px rgba(220,38,38,.28)'">
            &#128196; Fahrzeugw&#228;schen PDF
          </button>
        </div>

        <div style="display:flex;gap:18px;align-items:flex-start;flex-wrap:wrap;padding-top:14px;border-top:1px solid #eef2f7">
          <div style="flex:1;min-width:280px">
            <div style="font-size:11px;font-weight:800;color:#1e6091;margin-bottom:8px;letter-spacing:.4px;text-transform:uppercase">Tag</div>
            <div style="display:flex;flex-wrap:wrap;gap:6px" id="fw-day-btns">
              <button class="vz-day-btn" onclick="fwSelectDay('Montag')">Montag</button>
              <button class="vz-day-btn" onclick="fwSelectDay('Dienstag')">Dienstag</button>
              <button class="vz-day-btn" onclick="fwSelectDay('Mittwoch')">Mittwoch</button>
              <button class="vz-day-btn" onclick="fwSelectDay('Donnerstag')">Donnerstag</button>
              <button class="vz-day-btn" onclick="fwSelectDay('Freitag')">Freitag</button>
              <button class="vz-day-btn" onclick="fwSelectDay('Samstag')">Samstag</button>
            </div>
          </div>
          <div style="flex-shrink:0">
            <div style="font-size:11px;font-weight:800;color:#1e6091;margin-bottom:8px;letter-spacing:.4px;text-transform:uppercase">Datum</div>
            <input id="fw-date-picker" type="date" onchange="fwSetDate(this.value)"
              style="padding:9px 12px;border:1.5px solid #9db9d5;border-radius:7px;font-size:13px;font-weight:700;font-family:inherit;outline:none;background:#fff;color:#0b1220;min-width:180px;box-shadow:inset 0 1px 2px rgba(15,23,42,.04)">
          </div>
        </div>
      </div>

      <!-- ── Gesamt-Banner ──────────────────────────────────────────────── -->
      <div id="fw-total-banner" style="margin-bottom:14px"></div>

      <!-- ── Übersicht-Karte ───────────────────────────────────────────────── -->
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;box-shadow:0 2px 10px rgba(30,96,145,.07);overflow:hidden;margin-bottom:14px">

        <div style="display:flex;justify-content:space-between;align-items:center;gap:14px;flex-wrap:wrap;padding:16px 20px;border-bottom:1px solid #eef2f7;background:linear-gradient(180deg,#f8fbff 0%,#ffffff 100%)">
          <div style="display:flex;align-items:center;gap:10px">
            <div style="width:32px;height:32px;border-radius:8px;background:#e7f1fb;color:#1e6091;display:flex;align-items:center;justify-content:center;font-size:15px">&#128203;</div>
            <div>
              <div style="font-size:13px;font-weight:900;color:#0f172a;letter-spacing:-.1px">&#220;bersicht Fahrzeugw&#228;sche</div>
              <div style="font-size:11px;color:#64748b;margin-top:1px">Wer hat wann welchen LKW gewaschen</div>
            </div>
          </div>
          <div id="fw-overview-stats" style="display:flex;gap:6px;flex-wrap:wrap"></div>
        </div>

        <div style="padding:14px 20px;border-bottom:1px solid #eef2f7;background:#fbfcfd;display:flex;gap:10px;flex-wrap:wrap;align-items:center">
          <select id="fw-overview-year" onchange="fwRenderOverview()"
            style="min-width:150px;padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:7px;font-size:12px;font-weight:600;font-family:inherit;background:#fff;color:#0f172a;outline:none;cursor:pointer;box-shadow:inset 0 1px 2px rgba(15,23,42,.03);transition:border-color .15s"
            onfocus="this.style.borderColor='#1e6091'" onblur="this.style.borderColor='#b9cce3'"></select>
          <select id="fw-overview-driver" onchange="fwRenderOverview()"
            style="min-width:260px;padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:7px;font-size:12px;font-weight:600;font-family:inherit;background:#fff;color:#0f172a;outline:none;cursor:pointer;box-shadow:inset 0 1px 2px rgba(15,23,42,.03);transition:border-color .15s"
            onfocus="this.style.borderColor='#1e6091'" onblur="this.style.borderColor='#b9cce3'"></select>
        </div>

        <div id="fw-overview-table"></div>
      </div>

      <!-- ── Rangliste ───────────────────────────────────────────────────── -->
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;box-shadow:0 2px 10px rgba(30,96,145,.07);overflow:hidden">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:14px;flex-wrap:wrap;padding:16px 20px;border-bottom:1px solid #eef2f7;background:linear-gradient(180deg,#f8fbff 0%,#ffffff 100%)">
          <div style="display:flex;align-items:center;gap:10px">
            <div style="width:32px;height:32px;border-radius:8px;background:#fff7e6;color:#9a5b00;display:flex;align-items:center;justify-content:center;font-size:17px">&#127942;</div>
            <div>
              <div style="font-size:13px;font-weight:900;color:#0f172a;letter-spacing:-.1px">Rangliste Fahrzeugw&#228;sche</div>
              <div style="font-size:11px;color:#64748b;margin-top:1px">Wer hat am meisten gewaschen &ndash; PDF-Export pro Fahrer</div>
            </div>
          </div>
        </div>
        <div id="fw-ranking-body"></div>
      </div>

      <!-- ── LKW-Rangliste ───────────────────────────────────────────────── -->
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;box-shadow:0 2px 10px rgba(30,96,145,.07);overflow:hidden;margin-top:14px">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:14px;flex-wrap:wrap;padding:16px 20px;border-bottom:1px solid #eef2f7;background:linear-gradient(180deg,#f8fbff 0%,#ffffff 100%)">
          <div style="display:flex;align-items:center;gap:10px">
            <div style="width:32px;height:32px;border-radius:8px;background:#e7f1fb;color:#1e6091;display:flex;align-items:center;justify-content:center;font-size:17px">&#128666;</div>
            <div>
              <div style="font-size:13px;font-weight:900;color:#0f172a;letter-spacing:-.1px">Rangliste LKW</div>
              <div style="font-size:11px;color:#64748b;margin-top:1px">Welche Fahrzeuge am meisten gewaschen wurden &ndash; PDF-Export pro LKW</div>
            </div>
          </div>
        </div>
        <div id="fw-lkw-ranking-body"></div>
      </div>

    </div>
  </div>

  <!-- ── Fahrzeugwäsche Graph Panel ──────────────────────────────────────── -->
  <div id="panel-vz-graph" style="display:none;flex:1;flex-direction:column;background:#f1f5f9;font-family:'Segoe UI',Arial,sans-serif;overflow:hidden;">
    <div style="width:100%;max-width:1728px;margin:0 auto;display:flex;flex-direction:column;flex:1;overflow:hidden;">
      <div style="display:flex;align-items:center;gap:10px;padding:16px 18px;flex-wrap:wrap;flex-shrink:0;">
        <h2 style="margin:0;font-size:17px;font-weight:900;color:#0f172a;">&#128703; Fahrzeugw&#228;sche &ndash; Graph pro Jahr</h2>
        <select id="fw-graph-mode" onchange="fwGraphSetMode(this.value)" title="Ansicht"
          style="padding:8px 11px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12px;font-weight:900;font-family:inherit;background:#fff;color:#1f3347;outline:none;cursor:pointer;">
          <option value="single">Ein Jahr</option>
          <option value="compare">Jahre vergleichen</option>
        </select>
        <select id="fw-graph-year" onchange="fwGraphSetJahr(this.value)" title="Jahr 1"
          style="padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12.5px;font-weight:800;font-family:inherit;outline:none;background:#fff;color:#1e6091;cursor:pointer;"></select>
        <select id="fw-graph-year-2" onchange="fwGraphSetJahr2(this.value)" title="Jahr 2"
          style="display:none;padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12.5px;font-weight:800;font-family:inherit;outline:none;background:#fff;color:#047857;cursor:pointer;"></select>
        <span id="fw-graph-stats" style="font-size:12px;font-weight:700;color:#64748b;margin-left:auto;"></span>
      </div>
      <div id="fw-graph-content" style="flex:1;overflow-y:auto;padding:4px 0 30px 18px;">
        <div style="color:#94a3b8;padding:60px;text-align:center;font-size:14px;">Keine Fahrzeugw&#228;sche-Daten &ndash; bitte Waschdatei in Streamlit hochladen.</div>
      </div>
    </div>
  </div>

{tank_panels_html}

  <div id="panel-tel" style="display:none;flex:1;overflow-y:auto;font-family:'Segoe UI',Arial,sans-serif;">
    <style>
      #panel-tel{{--tink:#0f1f33;--tacc:#1e6091;background:linear-gradient(180deg,#eef3f9 0%,#f5f8fc 100%);}}
      #panel-tel .tel-wrap{{max-width:1060px;margin:0 auto;padding:0 20px 40px;}}
      #panel-tel .tel-head{{position:sticky;top:0;z-index:5;background:linear-gradient(180deg,#f3f7fc 0%,rgba(243,247,252,.92) 100%);backdrop-filter:blur(6px);padding:18px 0 14px;margin-bottom:6px;}}
      #panel-tel .tel-title{{display:flex;align-items:center;gap:12px;margin-bottom:12px;}}
      #panel-tel .tel-ico{{width:40px;height:40px;border-radius:12px;background:linear-gradient(135deg,#1e6091,#3aa0d8);color:#fff;display:flex;align-items:center;justify-content:center;font-size:19px;box-shadow:0 8px 18px rgba(30,96,145,.25);flex-shrink:0;}}
      #panel-tel .tel-h{{font-size:19px;font-weight:950;letter-spacing:-.4px;color:var(--tink);line-height:1.05;}}
      #panel-tel .tel-sub{{font-size:11.5px;font-weight:700;color:#64748b;margin-top:2px;}}
      #panel-tel .tel-count{{margin-left:auto;font-size:12px;font-weight:850;color:var(--tacc);background:#e8f1fb;border:1px solid #cfe0f1;border-radius:999px;padding:4px 12px;white-space:nowrap;font-variant-numeric:tabular-nums;}}
      #panel-tel .tel-controls{{display:flex;gap:10px;align-items:center;flex-wrap:wrap;}}
      #panel-tel .tel-searchwrap{{position:relative;flex:1;min-width:220px;max-width:460px;}}
      #panel-tel .tel-sicon{{position:absolute;left:13px;top:50%;transform:translateY(-50%);font-size:13px;opacity:.5;pointer-events:none;}}
      #panel-tel .tel-search{{width:100%;padding:10px 14px 10px 35px;border:1.5px solid #cdddee;border-radius:11px;font-size:13px;font-family:inherit;font-weight:600;outline:none;background:#fff;color:var(--tink);box-shadow:0 1px 3px rgba(30,96,145,.05);transition:border-color .14s,box-shadow .14s;}}
      #panel-tel .tel-search:focus{{border-color:var(--tacc);box-shadow:0 0 0 3px rgba(30,96,145,.13);}}
      #panel-tel .tel-pdf{{display:inline-flex;align-items:center;gap:7px;padding:10px 16px;border:1.5px solid #cdddee;border-radius:11px;background:#fff;color:var(--tacc);font-size:12.5px;font-weight:850;font-family:inherit;cursor:pointer;white-space:nowrap;transition:background .12s,border-color .12s;}}
      #panel-tel .tel-pdf:hover{{background:#eef5fc;border-color:#bcd4ec;}}
      #panel-tel .tel-group{{margin-bottom:20px;}}
      #panel-tel .tel-group-h{{display:flex;align-items:center;gap:9px;margin-bottom:9px;padding-bottom:7px;border-bottom:1.5px solid #dbe7f3;}}
      #panel-tel .tel-group-n{{font-size:12px;font-weight:950;text-transform:uppercase;letter-spacing:.6px;color:var(--tacc);}}
      #panel-tel .tel-group-c{{font-size:10.5px;font-weight:800;color:#64748b;background:#fff;border:1px solid #dbe7f3;border-radius:999px;padding:1px 8px;font-variant-numeric:tabular-nums;}}
      #panel-tel .tel-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(268px,1fr));gap:10px;}}
      #panel-tel .tel-card{{display:flex;align-items:center;gap:12px;background:#fff;border:1px solid #e2e8f0;border-radius:13px;padding:11px 12px;box-shadow:0 3px 10px rgba(15,31,51,.04);transition:transform .12s,box-shadow .12s,border-color .12s;}}
      #panel-tel .tel-card:hover{{transform:translateY(-1px);box-shadow:0 8px 20px rgba(15,31,51,.09);border-color:#cfe0f1;}}
      #panel-tel .tel-av{{flex-shrink:0;width:38px;height:38px;border-radius:11px;display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:900;letter-spacing:.3px;}}
      #panel-tel .tel-body{{min-width:0;flex:1;display:flex;flex-direction:column;gap:2px;}}
      #panel-tel .tel-name{{font-size:13.5px;font-weight:850;color:var(--tink);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
      #panel-tel .tel-num{{display:inline-flex;align-items:center;gap:6px;font-size:13.5px;font-weight:800;color:var(--tacc);text-decoration:none;font-variant-numeric:tabular-nums;letter-spacing:.2px;width:fit-content;}}
      #panel-tel .tel-num:hover{{text-decoration:underline;}}
      #panel-tel .tel-num-i{{color:#7aa0c4;font-weight:400;font-size:12px;}}
      #panel-tel .tel-role{{display:inline-block;width:fit-content;margin-top:3px;font-size:10px;font-weight:900;text-transform:uppercase;letter-spacing:.4px;color:#92400e;background:#fef3c7;border:1px solid #fcd34d;border-radius:999px;padding:1px 8px;}}
      #panel-tel .tel-mail{{display:inline-block;width:fit-content;margin-top:2px;font-size:11px;font-weight:700;color:#64748b;text-decoration:none;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:100%;}}
      #panel-tel .tel-mail:hover{{color:var(--tacc);text-decoration:underline;}}
      #panel-tel .tel-copy{{flex-shrink:0;width:32px;height:32px;border-radius:9px;border:1px solid #e2e8f0;background:#f8fafc;color:#64748b;cursor:pointer;font-size:14px;display:flex;align-items:center;justify-content:center;transition:background .12s,color .12s,border-color .12s;}}
      #panel-tel .tel-copy:hover{{background:#eef5fc;color:var(--tacc);border-color:#cfe0f1;}}
      #panel-tel .tel-copy.copied{{background:#ecfdf5;color:#16a34a;border-color:#bbf7d0;}}
      #panel-tel .tel-empty{{color:#94a3b8;padding:60px;text-align:center;font-size:14px;}}
      @media (prefers-reduced-motion:reduce){{#panel-tel .tel-card{{transition:none;}}}}
    </style>
    <div class="tel-wrap">
      <div class="tel-head">
        <div class="tel-title">
          <span class="tel-ico">&#128222;</span>
          <div>
            <div class="tel-h">Telefonliste</div>
            <div class="tel-sub">Fachberater &amp; Kontakte &middot; klick auf die Nummer zum Anrufen</div>
          </div>
          <span id="tel-count" class="tel-count"></span>
        </div>
        <div class="tel-controls">
          <div class="tel-searchwrap">
            <span class="tel-sicon">&#128269;</span>
            <input id="tel-search" class="tel-search" placeholder="Name oder Nummer suchen..." oninput="telFilter(this.value)">
          </div>
          <button class="tel-pdf" onclick="telPDF()">&#128196; PDF</button>
        </div>
      </div>
      <div id="tel-content" class="tel-content"></div>
    </div>
  </div>

  <div id="panel-bus" style="display:none;flex:1;overflow-y:auto;padding:18px 18px 28px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif">
    <div style="width:100%;max-width:1180px;margin:0 auto">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:18px 22px;box-shadow:0 2px 10px rgba(30,96,145,.08);margin-bottom:18px;display:flex;align-items:center;gap:14px;flex-wrap:wrap">
        <div style="width:40px;height:40px;border-radius:10px;background:linear-gradient(135deg,#1e6091 0%,#2f80b7 100%);display:flex;align-items:center;justify-content:center;font-size:20px;box-shadow:0 2px 7px rgba(30,96,145,.25);flex-shrink:0">&#128652;</div>
        <div style="min-width:0;flex:1">
          <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Busfahrplan Mitarbeiterbus</h2>
          <p style="color:#64748b;font-size:12px;margin:2px 0 0 0;font-weight:500">Fahrplan &amp; Notfallplan &middot; Becker Tours</p>
        </div>
        <span style="background:linear-gradient(180deg,#fde047 0%,#facc15 100%);color:#713f12;font-weight:900;font-size:12px;padding:7px 14px;border-radius:8px;border:1px solid #eab308;box-shadow:0 2px 6px rgba(234,179,8,.25);white-space:nowrap">g&uuml;ltig ab KW&nbsp;23</span>
        <button onclick="busPDF()" style="padding:10px 18px;background:linear-gradient(180deg,#ef4444 0%,#dc2626 100%);color:#fff;border:none;border-radius:8px;font-weight:800;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(220,38,38,.28);display:inline-flex;align-items:center;gap:7px;white-space:nowrap">&#128196; PDF / Drucken</button>
      </div>
      <div id="bus-content"></div>
    </div>
  </div>

  <div id="panel-arzt" style="display:none;flex:1;overflow-y:auto;padding:18px 18px 28px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif">
    <div style="width:100%;max-width:880px;margin:0 auto">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:18px 22px;box-shadow:0 2px 10px rgba(30,96,145,.08);margin-bottom:18px;display:flex;align-items:center;gap:14px;flex-wrap:wrap">
        <div style="width:40px;height:40px;border-radius:10px;background:linear-gradient(135deg,#0e7490 0%,#0891b2 100%);display:flex;align-items:center;justify-content:center;font-size:20px;box-shadow:0 2px 7px rgba(8,145,178,.25);flex-shrink:0">&#129658;</div>
        <div style="min-width:0;flex:1">
          <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Betriebs&#228;rztin</h2>
          <p style="color:#64748b;font-size:12px;margin:2px 0 0 0;font-weight:500">Aushang &middot; ASIG / ArbSchG / DGUV Vorschrift 2</p>
        </div>
        <button onclick="arztPDF()" style="padding:10px 18px;background:linear-gradient(180deg,#ef4444 0%,#dc2626 100%);color:#fff;border:none;border-radius:8px;font-weight:800;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(220,38,38,.28);display:inline-flex;align-items:center;gap:7px;white-space:nowrap">&#128196; PDF / Drucken</button>
      </div>
      <div id="arzt-content"></div>
    </div>
  </div>

  <div id="panel-versp" style="display:none;flex:1;overflow-y:auto;padding:18px 18px 28px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif">
    <div style="width:100%;max-width:880px;margin:0 auto">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:18px 22px;box-shadow:0 2px 10px rgba(30,96,145,.08);margin-bottom:18px;display:flex;align-items:center;gap:14px;flex-wrap:wrap">
        <div style="width:40px;height:40px;border-radius:10px;background:linear-gradient(135deg,#2f80b7 0%,#1e6091 100%);display:flex;align-items:center;justify-content:center;font-size:20px;box-shadow:0 2px 7px rgba(30,96,145,.25);flex-shrink:0">&#9201;&#65039;</div>
        <div style="min-width:0;flex:1">
          <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Verspätungstabelle</h2>
          <p style="color:#64748b;font-size:12px;margin:2px 0 0 0;font-weight:500">Tag wählen &middot; Excel mit Kunden &amp; Touren zum Eintragen der Abfahrtszeiten</p>
        </div>
      </div>
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:18px 22px;box-shadow:0 2px 10px rgba(30,96,145,.07)">
        <div id="versp-inst-wrap" style="display:none;margin-bottom:16px">
          <div style="font-size:12px;font-weight:900;color:#0f172a;text-transform:uppercase;letter-spacing:.4px;margin-bottom:10px">Woche / Quelldatei</div>
          <select id="versp-inst-sel" onchange="verspSelectInst(this.value)"
            style="padding:8px 12px;border:2px solid #1e6091;border-radius:7px;font-size:13px;font-weight:700;color:#1e6091;cursor:pointer;font-family:inherit;outline:none;background:#fff;min-width:220px"></select>
        </div>
        <div style="font-size:12px;font-weight:900;color:#0f172a;text-transform:uppercase;letter-spacing:.4px;margin-bottom:10px">Liefertag</div>
        <div style="display:flex;flex-wrap:wrap;gap:6px;margin-bottom:16px" id="versp-day-btns">
          <button class="vz-day-btn" onclick="verspSelectDay('Montag')">Montag</button>
          <button class="vz-day-btn" onclick="verspSelectDay('Dienstag')">Dienstag</button>
          <button class="vz-day-btn" onclick="verspSelectDay('Mittwoch')">Mittwoch</button>
          <button class="vz-day-btn" onclick="verspSelectDay('Donnerstag')">Donnerstag</button>
          <button class="vz-day-btn" onclick="verspSelectDay('Freitag')">Freitag</button>
          <button class="vz-day-btn" onclick="verspSelectDay('Samstag')">Samstag</button>
        </div>
        <div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin-bottom:16px">
          <button id="versp-dl-btn" onclick="verspDownload()" disabled
            style="padding:11px 20px;background:linear-gradient(180deg,#16a34a 0%,#15803d 100%);color:#fff;border:none;border-radius:8px;font-weight:800;font-size:13px;cursor:default;font-family:inherit;box-shadow:0 2px 6px rgba(21,128,61,.28);display:inline-flex;align-items:center;gap:8px;white-space:nowrap;opacity:.5">&#11015;&#65039; Excel herunterladen</button>
          <div id="versp-info" style="font-size:13px;color:#64748b">Bitte einen Tag auswählen.</div>
        </div>
        <div style="display:flex;align-items:flex-start;gap:9px;background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:10px 13px;margin-bottom:16px;font-size:12.5px;color:#92400e;line-height:1.45">
          <span style="font-size:15px;line-height:1.2;flex-shrink:0">&#9888;&#65039;</span>
          <span>Die <b>reguläre Abfahrtszeit</b> wird automatisch aus den Tourenplänen ermittelt (Abendtouren über den Vortag). Bitte trotzdem einmal <b>logisch überprüfen</b>, bevor die Tabelle weiterverwendet wird.</span>
        </div>
        <div id="versp-preview"></div>
      </div>
    </div>
  </div>

  <div id="panel-knapp" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;height:auto;min-height:100%;max-width:1440px;margin:0 auto;display:flex;flex-direction:column">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);margin-bottom:14px;display:flex;align-items:center;gap:14px;flex-wrap:wrap;flex-shrink:0">
        <div style="width:40px;height:40px;border-radius:10px;background:linear-gradient(135deg,#f28c28 0%,#d96712 100%);color:#fff;display:flex;align-items:center;justify-content:center;font-size:20px;box-shadow:0 2px 7px rgba(217,103,18,.28);flex-shrink:0">&#9881;&#65039;</div>
        <div style="min-width:0;flex:1">
          <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">KNAPP</h2>
          <p style="color:#64748b;font-size:12px;margin:2px 0 0 0;font-weight:500">Servicedesk-Eskalation P1/P2 &middot; Eskalationspyramide und Verantwortlichkeiten &middot; Stand 07.07.2026</p>
        </div>
        <button onclick="knappOpenPdf()" style="padding:10px 16px;background:linear-gradient(180deg,#f28c28 0%,#d96712 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(217,103,18,.28);display:inline-flex;align-items:center;gap:7px;white-space:nowrap">&#128196; PDF öffnen</button>
        <button onclick="knappDownloadPdf()" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px;white-space:nowrap">&#11015;&#65039; Herunterladen</button>
      </div>

      <div style="background:#fff;border:1px solid #cad7e8;border-left:5px solid #d96712;border-radius:12px;padding:14px 18px;box-shadow:0 2px 10px rgba(30,96,145,.07);margin-bottom:14px;flex-shrink:0;color:#334155">
        <div style="display:flex;align-items:center;gap:9px;margin-bottom:9px">
          <span style="font-size:17px">&#9888;&#65039;</span>
          <div style="font-size:13px;font-weight:900;color:#0f172a;text-transform:uppercase;letter-spacing:.35px">Ablauf bei kritischen Störungen</div>
        </div>
        <ul style="margin:0 0 10px 23px;padding:0;font-size:13px;line-height:1.55;font-weight:600">
          <li style="margin-bottom:4px">Jegliche Tickets in kritischen Bereichen müssen bei der Hotline mit <b>Prio&nbsp;1</b> angelegt werden.</li>
          <li style="margin-bottom:4px">Ab <b>2&nbsp;Stunden Lagerstillstand</b> muss Christian Henning telefonisch informiert werden.</li>
          <li style="margin-bottom:4px">Anschließend muss alle <b>30&nbsp;Minuten</b> ein Update über die Situation erfolgen.</li>
          <li>Ab <b>4&nbsp;Stunden Lagerstillstand</b> werden weitere Schritte situativ entschieden.</li>
        </ul>
        <div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:8px;padding:9px 12px;font-size:13px;font-weight:800;color:#9a3412">
          Die Eskalation bei KNAPP an die nächsten Stufen übernimmt <b>Christian Henning</b>.
        </div>
      </div>

      <div style="position:relative;flex:none;height:clamp(760px,calc(100vh - 120px),1200px);min-height:760px;background:#fff;border:1px solid #cad7e8;border-radius:12px;box-shadow:0 2px 10px rgba(30,96,145,.08);overflow:hidden;margin-bottom:18px">
        <iframe id="knapp-pdf-frame" title="KNAPP Eskalationspyramide" style="display:none;width:100%;height:100%;min-height:760px;border:0;background:#eef2f7"></iframe>
        <div id="knapp-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterstützt.</div>
          <button onclick="knappOpenPdf()" style="padding:10px 18px;background:#d96712;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF separat öffnen</button>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Schlüsselübergabe ───────────────────────────────────────────────── -->
  <div id="panel-schluessel" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1280px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#23906a 0%,#187454 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(24,116,84,.25);flex-shrink:0">&#128273;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Schl&#252;ssel&#252;bergabe</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">Protokoll direkt anzeigen, ausdrucken oder herunterladen</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentPdfOpen('schluessel')" style="padding:10px 16px;background:linear-gradient(180deg,#23906a 0%,#187454 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(24,116,84,.25);display:inline-flex;align-items:center;gap:7px">&#128424; PDF &#246;ffnen / drucken</button>
          <button onclick="documentPdfDownload('schluessel')" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px">&#11015; PDF herunterladen</button>
        </div>
      </div>
      <div style="position:relative;flex:1;min-height:780px;background:#eef2f7;border:1px solid #cad7e8;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.08)">
        <iframe id="schluessel-pdf-frame" title="Schlüsselübergabe Protokoll" style="display:none;width:100%;height:calc(100vh - 170px);min-height:780px;border:0;background:#eef2f7"></iframe>
        <div id="schluessel-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterstützt.</div>
          <button onclick="documentPdfOpen('schluessel')" style="padding:10px 18px;background:#187454;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF &#246;ffnen / drucken</button>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Entgeltportal ───────────────────────────────────────────────────── -->
  <div id="panel-entgelt" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1280px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#4f87e8 0%,#315fb7 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(49,95,183,.25);flex-shrink:0">&#128196;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Entgeltportal</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">Hinweis zu verlorenen Zugangsdaten direkt anzeigen und ausdrucken</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentPdfOpen('entgelt')" style="padding:10px 16px;background:linear-gradient(180deg,#4f87e8 0%,#315fb7 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(49,95,183,.25);display:inline-flex;align-items:center;gap:7px">&#128424; PDF &#246;ffnen / drucken</button>
          <button onclick="documentPdfDownload('entgelt')" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px">&#11015; PDF herunterladen</button>
        </div>
      </div>
      <div style="position:relative;flex:1;min-height:780px;background:#eef2f7;border:1px solid #cad7e8;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.08)">
        <iframe id="entgelt-pdf-frame" title="Entgeltportal Zugang" style="display:none;width:100%;height:calc(100vh - 170px);min-height:780px;border:0;background:#eef2f7"></iframe>
        <div id="entgelt-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterstützt.</div>
          <button onclick="documentPdfOpen('entgelt')" style="padding:10px 18px;background:#315fb7;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF &#246;ffnen / drucken</button>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Schadenmeldung Fuhrpark ──────────────────────────────────────────── -->
  <div id="panel-schaden" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1280px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#dc2626 0%,#991b1b 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(153,27,27,.25);flex-shrink:0">&#128663;&#65039;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Schadenmeldung Fuhrpark</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">Formular direkt anzeigen, ausdrucken oder herunterladen</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentPdfOpen('schaden')" style="padding:10px 16px;background:linear-gradient(180deg,#dc2626 0%,#991b1b 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(153,27,27,.25);display:inline-flex;align-items:center;gap:7px">&#128424; PDF &#246;ffnen / drucken</button>
          <button onclick="documentPdfDownload('schaden')" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px">&#11015; PDF herunterladen</button>
        </div>
      </div>
      <div style="position:relative;flex:1;min-height:780px;background:#eef2f7;border:1px solid #cad7e8;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.08)">
        <iframe id="schaden-pdf-frame" title="Schadenmeldung Fuhrpark" style="display:none;width:100%;height:calc(100vh - 170px);min-height:780px;border:0;background:#eef2f7"></iframe>
        <div id="schaden-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterstützt.</div>
          <button onclick="documentPdfOpen('schaden')" style="padding:10px 18px;background:#991b1b;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF &#246;ffnen / drucken</button>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Mängelanzeige Fuhrpark ──────────────────────────────────────────── -->
  <div id="panel-maengel" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1280px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#e59b16 0%,#b86b08 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(184,107,8,.25);flex-shrink:0">&#9888;&#65039;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">M&#228;ngelanzeige Fuhrpark</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">Fahrzeugm&#228;ngel dokumentieren, ausdrucken oder herunterladen</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentPdfOpen('maengel')" style="padding:10px 16px;background:linear-gradient(180deg,#e59b16 0%,#b86b08 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(184,107,8,.25);display:inline-flex;align-items:center;gap:7px">&#128424; PDF &#246;ffnen / drucken</button>
          <button onclick="documentPdfDownload('maengel')" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px">&#11015; PDF herunterladen</button>
        </div>
      </div>
      <div style="position:relative;flex:1;min-height:780px;background:#eef2f7;border:1px solid #cad7e8;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.08)">
        <iframe id="maengel-pdf-frame" title="Mängelanzeige Fuhrpark" style="display:none;width:100%;height:calc(100vh - 170px);min-height:780px;border:0;background:#eef2f7"></iframe>
        <div id="maengel-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterstützt.</div>
          <button onclick="documentPdfOpen('maengel')" style="padding:10px 18px;background:#b86b08;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF &#246;ffnen / drucken</button>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Übergabeprotokoll LKW ───────────────────────────────────────────── -->
  <div id="panel-lkw_uebergabe" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1280px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#d97706 0%,#92400e 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(146,64,14,.25);flex-shrink:0">&#128666;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">&#220;bergabeprotokoll LKW</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">LKW-&#220;bergabe dokumentieren, ausdrucken oder herunterladen</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentPdfOpen('lkw_uebergabe')" style="padding:10px 16px;background:linear-gradient(180deg,#d97706 0%,#92400e 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(146,64,14,.25);display:inline-flex;align-items:center;gap:7px">&#128424; PDF &#246;ffnen / drucken</button>
          <button onclick="documentPdfDownload('lkw_uebergabe')" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px">&#11015; PDF herunterladen</button>
        </div>
      </div>
      <div style="position:relative;flex:1;min-height:780px;background:#eef2f7;border:1px solid #cad7e8;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.08)">
        <iframe id="lkw_uebergabe-pdf-frame" title="Übergabeprotokoll LKW" style="display:none;width:100%;height:calc(100vh - 170px);min-height:780px;border:0;background:#eef2f7"></iframe>
        <div id="lkw_uebergabe-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterstützt.</div>
          <button onclick="documentPdfOpen('lkw_uebergabe')" style="padding:10px 18px;background:#92400e;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF &#246;ffnen / drucken</button>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Reinigungs- und Fahrernachweis ─────────────────────────────────── -->
  <div id="panel-reinigung" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1280px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#0f766e 0%,#115e59 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(17,94,89,.25);flex-shrink:0">&#129529;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Reinigungs- und Fahrernachweis</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">Formblatt FB-FP-001-11 &middot; g&#252;ltig ab 22.08.2026 &middot; direkt anzeigen, drucken oder herunterladen</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentPdfOpen('reinigung')" style="padding:10px 16px;background:linear-gradient(180deg,#0f766e 0%,#115e59 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(17,94,89,.25);display:inline-flex;align-items:center;gap:7px">&#128424; PDF &#246;ffnen / drucken</button>
          <button onclick="documentPdfDownload('reinigung')" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px">&#11015; PDF herunterladen</button>
        </div>
      </div>
      <div style="position:relative;flex:1;min-height:780px;background:#eef2f7;border:1px solid #cad7e8;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.08)">
        <iframe id="reinigung-pdf-frame" title="Reinigungs- und Fahrernachweis" style="display:none;width:100%;height:calc(100vh - 170px);min-height:780px;border:0;background:#eef2f7"></iframe>
        <div id="reinigung-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterst&#252;tzt.</div>
          <button onclick="documentPdfOpen('reinigung')" style="padding:10px 18px;background:#115e59;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF &#246;ffnen / drucken</button>
        </div>
      </div>
    </div>
  </div>


  <!-- ── Reisekostenabrechnung Excel ─────────────────────────────────────── -->
  <div id="panel-reisekosten" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1440px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#16803a 0%,#166534 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(22,101,52,.25);flex-shrink:0">&#128202;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Reisekostenabrechnung</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">Excel-Vorlage f&#252;r die Reisekostenabrechnung &middot; Vorschau unten</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentExcelDownload('reisekosten')" style="padding:10px 16px;background:linear-gradient(180deg,#16a34a 0%,#15803d 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(21,128,61,.28);display:inline-flex;align-items:center;gap:7px">&#11015;&#65039; Excel herunterladen</button>
        </div>
      </div>

      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:14px 16px;box-shadow:0 2px 10px rgba(30,96,145,.07);color:#334155">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-bottom:10px">
          <div>
            <div style="font-size:14px;font-weight:900;color:#0f172a">Vorschau</div>
            <div style="font-size:12px;color:#64748b;margin-top:2px">Tabellenblatt &bdquo;Reisekosten&ldquo; &middot; Zum Ausf&#252;llen bitte die Excel-Datei herunterladen.</div>
          </div>
          <span style="font-size:11px;font-weight:800;color:#166534;background:#dcfce7;border:1px solid #bbf7d0;border-radius:999px;padding:5px 9px;white-space:nowrap">Reisekostenabrechnung.xlsx</span>
        </div>
        <div style="background:#eef2f7;border:1px solid #cbd5e1;border-radius:9px;padding:10px;overflow:auto;max-height:calc(100vh - 220px);min-height:520px">
          <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAABZsAAAOKCAYAAAD9VI5tAAAACXBIWXMAAAsSAAALEgHS3X78AAAAAXNSR0IArs4c6QAAAARzQklUCAgICHwIZIgAACAASURBVHic7N15dBRl9v/xTyedkAQIBBIhJiAIElYBkU0EAcUNUVQEAQHREdcZVBSUfN0F2URBcRBHEQcQBAdGkFVQEcK+b7IIYQtgEsjaSWfp+v0xSf1SJEFId2gC79c5fU5XPVVP3aouINzcvmU7fvy4Ua5cOZVFTqdTWVlZqlixordDuWpkZGTI5XKpfPny3g4FAOAFZ86cUeXKleXj4+PtUACUQFJSkgIDA+WJn//T09Pl4+OjwMBAt+fKzs5WZmam2z/XezIm/O9+KV++vPz8/LwdSpmXk5Oj9PR0VapUyduhXFY89WcfxXO5XEpOTlZISIi3Q7niJCcnKzAwUP7+/t4O5Yp09uxZVapUif93lCFHjhyRzWaTvVy5cgoKCvJ2PG4p6/GXJYZhKDs7m2sOAFep5ORkBQUF8UMfUEY5HA4FBAR4JBmbm5sreehncafTqdzcXLfn8mRMkNLS0kikeEhWVpacTif35jk89WcfxXO5XEpLS+Mal4L8f1MDAgK8HcoVKSUlRYGBgfL19fV2KLhI/E8RAAAAAAAAAOA2ks0AAAAAAAAAALeRbAYAAAAAAAAAuI1kMwAAAAAAAADAbSSbAQAAAAAAAABus3s7gKvRiRMntHHjRt1www1q1KiRt8MB8BeWL1+u9PR0c9lutys8PFzNmzeXjw+/swNwZTh48KB27dqluLg4hYaGqlWrVqpVq5a3w8Ilkv/55/P19VWNGjV0ww03qHz58hc93/r163Xy5Mkix+rWravGjRtf8FwLFy5UTk5OkWP33nuv/P39Lzo+b0lOTtaaNWt0/Phx2e12XXfddWrXrp0CAgK8HVqZcurUKa1bt073339/oZ/Fdu7cqeTkZN16661ei+9yl3/97rvvPtntpARKauHChWratKlq1Kjh7VDKtMzMTC1ZskQdOnRQlSpVLGOrVq2Sw+HQnXfeyf+7Sij/+rZv315Vq1b1djhXpcOHD2v79u2qX7++6tevX2h8//792rNnjxo3bqy6det6JcYL9dxzz2nXrl0yDEOGYZjrXS6X7Ha7Bg8erPr161PZ7A3z58/X+PHjNWXKFG+HAuACzJo1S0uXLtW+ffu0b98+bdy4UW+99ZZeffVVZWdnezs8AHDbl19+qeeff1579uxR+fLltXv3bg0YMEBLlizxdmi4RPbu3aspU6aY/9Zt3rxZI0aM0MCBA7V3796Lnm/FihWaO3euOV/BV2Ji4kXNNXnyZP38889FzuVyuS46Nm/ZvHmz+vbtqwULFsgwDGVnZ2vatGnq16+fjh075u3wypS4uDiNHz9eubm5hcY2btyohQsXeiWusiL/+hX3SxxcmMmTJys2NtbbYZR5DodD48eP19mzZy3rZ8+erTFjxuj6668n0ewGp9Op8ePH68yZM94O5ap18OBBjR8/XlOnTi1yfMaMGRo/frz2799/yWO7WLt37zbf22w28+Xr6yvDMLR48WIlJiZS2XypZWdna9myZXr00Uc1a9YsxcXF6dprr/V2WAD+wr333qtu3bqZy/Hx8erTp4/Wr19P5QyAMm3dunX673//q0mTJlmqKX7++We99957atmyJZUwV4lq1app2LBh5nJubq6mTp2q1157TTNmzFCFChUuar5WrVrpH//4h0die/DBB9WxY0ePzOUNiYmJeu+999S/f3/16NHDTJw8+eSTGjNmjN566y19+eWXstls3g4VALxu7ty5+ve//63x48erXr163g4HcFv16tW1ZcsWJSYmWn6uTklJUUxMTJnOC+bk5Mhut8tmsyk9PV05OTlUNl9q27ZtU1ZWlh5//HHVrFlTa9as8XZIAEogLCxM4eHhcjqd3g4FANzy2Wef6bHHHiv0tb3bbrtNt912m44cOeK12OBdvr6+GjBggMqXL6+ffvrJ2+GUadOnT9cNN9ygnj17Wir07Ha7nnzySUVFRV10xTcAXInmzp2rr7/+mkQzrighISHq0KGD1q9fb1m/ceNGtW7dWtWqVfNabBejYOuMfHa73VwfGBgom81GsvlSW7lypbp06aKgoCDdeeedWrx4sbdDAnCRcnNztXr1aqWmpqpZs2beDgcASiwpKUmxsbFq1apVoTEfHx+99dZbuummm7wSGy4Pfn5+uuWWW3TgwAFvh1Km7dy5U3fddVeRY/kV5aGhoZc8LgC4nMybN0+ffvopiWZckTp27KgVK1ZY1q1atapMfXOrqGRzfksrl8slm80mPz8/2mhcSmlpafrpp580duxYSVK7du30r3/9S3/88Yfq1Knj7fAAnMdHH32kTz/9VMp7yIJhGBo6dChfLQdQpp06dUqSSHLhvEJDQ7Vu3bqL3m/+/PlatmxZofVvvvlmkb/gOJ+RI0dq3LhxlnXDhg1T+/btLzquSy0zM1MHDhxQeHi4Zf3cuXMt35Bq0aJFkQ8OQvG6d+9eqPVIZmambr75Zq/FBKBk/vOf/2jTpk3y9/fX8ePHSTbjinPTTTfp3Xff1Z9//qlrrrnGbKHx8ssva8GCBd4O7y+lpaXJMIxC/+4WrGy22+3y8fEh2XwpbdiwQcr7TcDOnTtlGIZCQkK0evVqks3AZe6xxx5T586dpbzf3O3evVtffPGFfH19dc8993g7PAAoET8/Pynv4TEX248XV4+MjAwFBwdf9H4dO3ZU3759C60vyVdFn3jiCbVs2dKyLiws7KLn8Qa73S5fX185HA7L+rNnzyojI0OSFBMTo5CQEJLNF2n8+PGy263/pZ03b54SEhK8FhOAktmyZYsmTJigTZs26cMPP1SjRo3KTGsB4EKUL19enTp10tq1a/XAAw9o48aNatu2rSpVquTt0C5Ifk4zX36COb+i2WazKSUlRampqSSbL6UVK1YoODhYkydPNtcFBgZq8eLF6tevH09YBS5jYWFhqlWrlrlcp04dnTlzRj/++CPJZgBlVkREhCQpLi6uyG9qfP755woODlbv3r29EB0uF7GxsSV6cE3lypU9VlBRvXr1MlucYbfbVadOHR06dMiSMH/qqafM97t27fJSdGXb9ddfb/7SLF+VKlVINsPjjh49qtOnT1v+DBf1dXKUXHR0tKpXr657771Xq1ev1ocffqhRo0aRJ8EV5bbbbtPs2bP1wAMP6Ndff9Vtt93m7ZAu2LnfVsuvcLbZbJbK5pSUFHo2XyqJiYlavXq13njjDX3++efma9SoUYqLi9Pvv//u7RABXKT8p60CQFkVEBCgW265RT/88EOhsYyMDC1cuJAWG1e5Y8eO6ZdfflGHDh28HUqZdvvtt2vWrFlKS0srNHb06FHt37/fK3EBuDBbtmzR6NGjLZV86enp8vf393ZoV4zAwEAp75kRQ4YM0Z49e4r8+QQoy1q0aKF9+/Zp3759Wrt2bZlp+5SYmKi1a9da1kVGRkrnJJtzcnKUm5tLZfOlEhMTo9DQUDVu3NiyvlatWrrhhhv066+/qmHDhl6LD8D5JSYm6ujRo1LeD5d79uzRvHnz1LNnT2+HBgBuee655/T4448rMjJSvXr1UkBAgJKSkjR27FgFBwfrlltu8XaIuEScTqf5b11WVpb++OMPTZ48Wffcc0+Jfk5NSUkx5yvI399f1atX90jMZUX37t21cOFCvfPOOxo8eLD5H7QDBw5ozJgxPAMCuMzVqFFDf/75p3bu3Kkbb7xRmzZtkmEYuu6667wd2hUpLCxMQ4YM0YgRI9SsWTPLN0xRMqdOnVK5cuUs60JCQlSxYkWvxXQ1CggI0O23367Ro0frlltuKTMtNGJiYuTj42M+DNAwDMXFxckwDOXm5srHx0c2m00ul0vBwcEkmy+VpUuX6q677irUU0yS7rrrLs2cOVN/+9vfCn0NDMDl4auvvtJXX31lLkdEROjBBx/UI4884tW4AMBdNWvW1Lhx4zR27FhNnTpV4eHhOnHihBo3bqz3339f5cuX93aIuESOHz+uxx57TMr7GmTjxo312GOP6cEHHyzRfMuWLSvyAYFNmjTRpEmT3I63LAkICNCnn36qjz76SH369FHVqlVlGIYCAwP1xhtvaPv27d4OEcB5NG/eXPfdd59eeOEF1alTR0eOHNHf//73q+4XZ5dSfm/bUaNG6ZNPPiFX4qbo6OhC69544w116dLFK/FczW677TYtXLjQ/JmrLPjxxx+VnZ1tLvv4+MgwDDPJrLwEdOXKlRUSEiJbfHy8ERQU5MWQS87hcMjpdCokJMTboVw10tPTlZ2drcqVK3s7FACAF5w8eVLVqlWjf94VKCcnR6dOnVJiYqJCQ0MVHh7O53wFSkhIUPny5c2vK7sjNTVVkjxSFeV0OpWenq4qVapcNjGVluTkZB07dkxVq1ZV9erVZbPZlJOTI8MwLrtkyp9//qnKlSvTKsADsrKylJycXGYebHmpeOrP/qVy/PhxnTlzRuHh4WXms3S5XPrzzz9JjJeChIQEVahQQQEBAd4O5Yp06tQphYWFydfX19uh4AKcOXNGsbGxCggIoLIZAAAA/6tkjYyMNL/eD6B0VKpUqdDXZov69iOAyw//TgJA0apUqaLTp0/LZrPxgEAAAAAAAAAAgPtINgMAAAAAAAAA3EayGQAAAAAAAADgNrvT6fR2DCXmdDqVlZUlh8Ph7VCuGk6nU7m5uVxzALhKuVwuORwOHhwHlFHZ2dnKzMyUYRhuz+V0OmWz2Tzyc2F2drays7PdnsuTMUHKzc1VRkaGcnJyvB1KmZeTk6OcnBzuzXN46s8+iudyucyf3+BZ+f+mulwub4dyRcr/N4j/d5Q9tkOHDhll9enCWVlZqlChgkeepo0LZxiGbDabt8MAAHgB/wYAZZun/wx7cj5PzcXfU57DtfQsrmfRuC6lj2tcOriupetqvr4ul+uIj4/Pdd6OoyQCAgJkr1ixooKCgrwdS4k4HA4FBASoQoUK3g4FAAAAAAAAANxVpr+KQC06AAAAAAAAAMBtJJsBAAAAAAAAAG4j2QwAAAAAAAAAcBvJZgAAAAAAAACA20g2AwAAAAAAAADcRrIZAAAAAAAAAOA2ks0AAAAAAAAAALeRbAYAAAAAAAAAuI1kMwAAAAAAAADAbSSbAQAAAAAAAABuI9kMAAAAAAAAAHAbyWYAAAAAAAAAgNtINgMAAAAAAAAA3EayGQAAAAAAAADgNpLNAAAAAAAAAAC3kWwGAAAAAAAAALiNZDMAAAAAAAAAwG0kmwEAAAAAAAAAbiPZDAAAAAAAAABwG8lmAAAAAAAAAIDbSDYDAAAAAAAAANxGshkAAAAAAAAA4DaSzQAAAAAAAAAAt5FsBgAAAAAAAAC4jWQzAAAAAAAAAMBtJJsBAAAAAAAAAG6zezuAK82ff/6pb7/91lx+8MEHVbNmTa/Fs3XrVq1atcpcfv7552W387FfSgsWLNChQ4ckSeHh4erZs6e3Q/KIK/W8AAAAAAAAUDJXXNbxm2++UWxs7Hm3CQgIUEhIiKKiotSyZUsFBgZ67PgJCQl68cUXzeVWrVp5Ndm8c+dOSzyDBg0i2XyJLVy4UFOmTJEkde/e/YpJyl6p5wUAAAAAAICSueKyjgsXLtScOXMuePvIyEiNGDFC/fr1k81mK9XYcGEcDoeGDh0qwzD0yiuvqHbt2t4OCQAAAAAAAMBfKPVk82+//ab27duX9mFK7Pjx4xowYIASExP10ksvuT1fgwYN5HA4zOVy5cq5PefVZtWqVZo0aZIk6amnnvJ2OAAAAAAAAAAuQKkmm2NjYxUdHW3pGXwp1a1bt1CVs2EYSkxM1G+//aZx48aZieFhw4bp4Ycfdrvlhc1m82hbjqvRf//7X2+HAAAAAAAAAOAilUqyOTY2VmvXrtW8efNKY/oLFhgYqGbNmhU5dscdd6hGjRpm5Wx2drbWrFlTZLJ57969WrFihfbt26e0tDRdc801atOmje6+++5CieWEhATNnTvXXO7atatq1Khh2SYnJ0e//vqr1q1bp9jYWGVmZio8PFzXXXedunbtqlq1ahV7ThkZGVq+fLk2bNigU6dOyd/fX3Xr1tU999yjBg0aXPQ1Onz4sJYuXWoud+rUSVFRUZZtEhMT9dNPP2nr1q2Kj4+Xn5+frr32WrVu3VodO3Y8b/X2tm3b9Msvv2j//v1KTk5WaGioIiIi1KlTJ918882W1iVff/21MjMzLQ9YnDt3rtatW6ewsDA9/PDDlrlPnz6tpUuXavv27UpMTFTlypV14403qmvXrqpWrVqhWObNm6fTp09Lkpo1a6Y2bdooKSlJc+fO1datW5Wenq4GDRrokUce0fXXX1/k+RiGoU2bNunXX3/VH3/8IYfDoSpVqqhp06a67777FBoaet7rnX++p0+f1vz587V7925lZmaqQYMG6tGjR6F75dixY/rxxx/N5SeffFJnzpzRZ599ptjYWD366KO65557LPtczP1a0NmzZ7Vs2TJt3rxZ8fHxCgwMVFRUlO68886/vLd8fHwkScnJyZo3b562bt0qh8OhqKgo9ejRo9A9fe55PfHEE/L399f69eu1YMECHT16VCEhIerYsaO6detWqM/4zz//rH379kmSqlSpUmS/6G+++cb8ZVKTJk3Url07c8wT90L+uW7cuFFpaWmqVauW7r//frVo0UK7du3S6tWrJUlBQUHq37//ea8fAAAAAADAFSE+Pt5IT0/36GvVqlXGxx9/bERHRxstWrTw+Pz5r/j4eCM1NdUo6JFHHjEkGZKMJk2aGOdz4MABc1tJxkcffWQZz8zMNIYNG2bZpuCrefPmxr59+yz77N6927JNTEyMZfzIkSNGhw4dip1TkvHKK68YTqezULwbNmwwGjRoUOx+I0eONFwul2WfadOmWbZxOBzmWHx8vNGiRQtz7LnnnjNycnIs+3/zzTdGWFhYscds0qSJsWHDhkKxOhwO4+mnnz7ved57773GqVOnzH1q165d7Lb33XdfobhCQkKK3DYsLMxYuHBhoZj69OljbvPmm28ahw4dMpo0aVJo/0qVKhm//fZbof2TkpKMvn37FhtjtWrVjJ9++qnQfoMGDTK36dWrl7FhwwYjMjKy0P5Vq1YtdNzNmzdbtjl27Jhx6623msujR482ty3J/Zpv4cKFRkRERLH7Dhs2zMjMzCz2vPr27WscPnzYaNasWaF9Q0JCjDVr1pz3vE6dOmWMGjWqyGP37t3byMrKsuw/dOhQc7xz585FnlNUVJS5zYgRIzx6L+zfv98yf8HXZ599Znz66afmcuvWrYuMDwAAAAAAoAh7vB2AO3xKI4HdokULPfXUU+revXtpTO8xqampluWKFStall977TWNHj1ayqtOfOutt/Svf/1LvXr1kiRt3bpV999/v5KSki7oeIZhaODAgWZbkWbNmumNN97Q2LFjNXjwYFWtWlWSNG7cOH388ceWffft26cuXbpo7969kqS77rpLn3/+uUaPHq3w8HBJ0vDhwzVlypQLiiUjI0MDBw7U5s2bJUmPPPKIxo8fL19fX3ObL774Qv3791d8fLwkqWHDhnrxxRc1aNAgM9adO3eqY8eO2r59u2X+cePG6fPPP5ckRURE6OWXX9bYsWMVHR1tVskuWrRITzzxhAzDkPKqT2+++WbLPM2bN1f79u0tlbHfffed+vfvr7Nnz0p5fZ2//PJLvfrqq9L/foGiBx980Dy3fAWres+cOaNnn31Wv//+u5o3b66AgABzLDk5WU8++aScTqdl/1deeUUzZsyQ8u6H119/XR999JFZMXv69Gk9/PDDOnbsWLHXPSkpSQMGDFDlypX1/PPPW6qSExMT1bt3b/O8JMnf39+y/8yZM82K2XOV9H5dsWKF7rvvPp04cULKuy9ffvlldevWzdxm9OjRev3114s9L7vdrmeffVa7d+8udD3Pnj2rJ554wnI9zz2vuXPn6rXXXlPNmjVVp04dy9i3335rqXb3BHfuhaysLA0YMMCsrJakLl26aNiwYbrzzjv13HPPadGiRUUeCwAAAAAA4IpWGpXN+a+NGzdetpXNGRkZRv/+/S0Vidu3bzfHN2zYYBlbsmSJOZabm2up6ixYXXq+yua9e/daKiaTk5MtMcXExFiqULOzs82xXr16mWN33323pcp0x44dlnnPnDljjhVV2Zybm2s888wzlsrQc2M5fPiw4efnZ27zxBNPWKqijx49aqmy7tixo1lV7XK5jPDwcHPs3KrWhIQEo2bNmoYkIzIy0tiz5///wubc67d161bLvqmpqZa53377bcv41KlTLdepoIKfWcWKFY3WrVsbhw8fNoy8quUePXpYjv3rr7+a+55bBT9x4kTL+RSssp4wYUKxx82vAi54LadMmWIZnzJlSrHXIyIiwnjqqaeMbdu2Gb///rtx6NAhw3Djfs3MzLR8jg888IAltkmTJhX7Z+Rir+eqVauKPa+KFSsan376qZGdnW24XC5j8eLFlvF77rnHck3drWx2515YtGiRZSw6Otpy77/11luW8eLiAwAAAAAAKAKVzZerhIQETZgwwfL6+OOPNXz4cLVq1UrffPONue2TTz6pG2+80Vwu2G+6WbNmuvPOO81lHx8fDRo0yFz+8ssvzerc8ylYSZ2dna2UlBTLeJs2bXTixAk5nU5t2bLF7FN74sQJzZ4929zu6aeftvRJbtKkiXr37i3lVWIuX778vHGMHj1akydPlvIqh2fOnKng4GDLNrNnz1Z2dra5/N5771kqNGvUqKHo6Ghz+ZdfftG2bdukvJ7UJ0+eNMfOnDljmbtq1aratGmT0tLSdOzYsYvqNb1y5UrL3E888YRlvHfv3maF+pIlS/THH38UOU9qaqrGjh1rVkxXqlRJw4cPt2xTcN/Q0FCtX79eP//8sxYuXKgePXpYzqdgP+A9e/ac9xzefvtty7UcOHCg6tWrZy4XrIo91/XXX69JkyapadOmioqKUu3atSU37tdVq1aZ1fLKq94uGNsTTzyhli1bqlOnTurevbsOHz5cZFypqakaM2bMea/nwYMHiz2v2267Tc8//7zsdrtsNpvuvvtuPfbYY+Z4cdXcnnCx98JPP/1kvvfz89PLL79s9uK22WwaPHhwoW9JAAAAAAAAXA1K5QGBl4uTJ0/qxRdf/MvtnnjiCX3yySeWdb/99pv5/pprrtHGjRst4wVbEezfv1/Hjx8v9HC3c9WqVUsBAQHKzMyUw+FQs2bN1KdPH7Vt21ZNmzZVvXr1dO211xbab9euXZbllJQUbdiwwbKuYHJry5YtRT4wTZKmT59uJtIqVqyo77//vsiH6a1YscJ837lz5yLjat26tWV527Ztat68ufz8/NSxY0f98ssvkqRu3bqpd+/e6tixo5o2baomTZooLCys2Ot0Plu3brUsHz9+3JJ8lqQ6deqYie9du3YVassgSZGRkbr11lst685NeqelpZnvK1eurFatWkl57VBSUlIUFxen7Oxs5eTkWB5gd257loIaN26sunXrWtbZ7Xbdcccd2r9/vyRp7dq1xe7fr18/+fn5FVpf0vv13FYj5z5QMyAgoNC9VpSIiAi1b9/esu581/Ncjz76aKF1LVu21PTp06W8a+p0Os/7MMqSuth7If/ekqQOHTqoSpUqlm1DQkJ07733Wn5BBAAAAAAAcDW4opPNf6Vhw4aaOnWqWrZsaVYm5itYybhs2TItW7bsvHOdPHnyL5PNYWFhGjFihIYMGSLl9ej95JNPzER3ZGSk+vTpo8cff9yS7EpISLDMM2DAgPMe59ChQ8WOFaxwTU1N1cGDB83q2IIKVruemxwteD4FnT592nz/zjvv6K677lJmZqZ0Tt/doKAgPfLII+rXr586d+5c6NqfT8FjSNItt9xy3u2L65/csGHDQscNCAhQtWrVzGO4XC7LeFxcnCZNmqRvv/222Apf5SWji1O/fv0i11evXt18f/r06WITq0UlzuXG/Zrfp1mSqlWrpgoVKpx3v+I0atTooq9nQREREYXWnZvEPd/+7rjYe6HgNYuMjCxyzuLWAwAAAAAAXMmu6GRznTp19PXXX1vWffvtt/rss8+kvHYHNputyGSnw+Ew31etWrXIyt6CCracOJ+XXnpJ119/vcaNG6c1a9ZYxo4fP64xY8boo48+0vTp083q5JycHMt2UVFRhR6wVtD5xpRX0ZxffTtkyBDFxMQUSjIWrM4tqpJWkuVhgpIsD1Hr0KGDYmJi9MEHH2jOnDmW7RwOh6ZNm6Zp06ZpyJAhGjt27AUnnAteZz8/v2KTt38lKCioyPUFK5QLio2NVefOnS1J5vbt2ysiIkJ+fn6KiYkptmVHQedes+KOm5ubW+R2xT1srqT3a0ZGhrmuuM/5Qlzs9TxXUYl1Hx/3uvxcSGsblSD2gn8eL/TzBAAAAAAAuBpc0RmRoKCgQl+Pr1+/vhYuXKijR49KkgYPHqyVK1cqICDAsl1ERISSk5MlSV27dtW0adM8EpPNZlP37t3VvXt3xcbGavPmzdq6dat++uknrV+/XspLBA4cOFAdOnRQ9erVC/V/nT17tpo2bVqi47/zzjsaMGCA2rZtq5MnT2rnzp2aOHFioR61NWvW1M6dO6W8CuyinNsu4txK1ObNm+u7775TQkKCtmzZoq1bt2r16tVaunSpmez88MMPdccdd+juu+++oPgrV65svvfz89O2bdvcTkpeiA8//NBMNIeHh2vp0qVq0qSJOd6nT58LSjbHxcUVub5gmws/P79ik8rFKen9WvB6JiQkFGoJUhYU1bYkOzv7vNXn7ggLCzM/6/j4+CK3Ke5zBgAAAAAAuJKVapauYcOGWrVqVWke4qKFhoZq9OjR5vLatWvNSueCWrRoYb4/cOBAqcRSq1YtPfzww3r//fe1bt06rV69WpUqVZLyKlXz+xNff/31lv3cSWS9+uqruu666zRy5Ehz3dtvv13ooXYF+kGMOwAAIABJREFU21PExMQUqq6WpH379lmWGzZsWOQxQ0NDdeedd2rYsGFasGCB/vjjD7Vt29Yy/4Uq+CA9h8NRbLLP0xYuXGi+f/DBBy2JZpfLZenjez67du0qsndxwf3btGlzUa1F5Mb9GhUVZb7PzMzUkSNHCm3z4YcfauTIkfr4448tvby9qWD1/sGDBwvdn7///vsFf9vgYjVu3Nh8v3r1akt1uPLuy5UrV5bKsQEAAAAAAC5npV8Sehnq2bOn7rvvPnM5Ojpav//+u2WbO+64w3y/du1abd++3TK+e/duDRs2TOPGjdM333xTZDL2XD///LPGjBmjp556Sj/99FOh8Xbt2unGG280l/NbKTRo0MDSq7eoB4+NHTtW7777rj777DPt2LHjL2Pp27ev+UC37OxsvfLKK5bWDd26dTPfHz16VIsWLbLsbxiG/vWvf5nLkZGR5gMD9+7dq08++UT/+Mc/NHHixELHrlGjhqWSuWA/3HOTrAUrfpV3jQpasGCBZTk9PV1DhgzRBx98oC+++KJQj+eSyq8aVhEPuluwYIGlx3V+n+qinD17ttC13Ldvn6XHcqdOnS46vpLerx06dLBs99///teyHBsbq1deeUXR0dF66aWXLpuK3YIPtTx79qyWL19uLhuGUeiBn55U8EGIZ8+eNR9imG/mzJmWvs4AAAAAAABXi7L1fXkP8fHx0ahRo8x2DpmZmXr55Zf1ww8/mC0EHnjgAdWtW1cHDx6UJPXr108TJ05UVFSUjh49qiFDhpg9l6Ojoy+o9cCpU6c0bNgwSdJvv/2mTz75RI0aNVJgYKCSk5O1dOlS/fbbb1JeK4XmzZtLef1shw4dqqefflqSNG3aNNWuXVt9+vRRQECA5s6dq6FDh0p5/Zh37979l7H4+flp1KhRZvJ28eLFmjVrlvr27StJuuuuu3TnnXeaSdBBgwbJ4XCoZcuWSklJ0dSpUzVjxgxzvvfee8/s++zr66t//OMfUt6D1oKDg9WxY0dVqlRJGRkZ2r59u6ZOnWru26ZNG/P9ua04JkyYoEqVKiknJ0ctW7ZUw4YN1bdvX/PYr7zyioKCgtS+fXs5HA6NGzfOTIK3b99eAwcO/MtrcSFat26tJUuWSHl9v++99141aNBAa9as0bBhw3TzzTdr06ZNkqQNGzbo9OnTqlKlivz8/CzJ9KCgIL344ouy2Wy66aabdOLECfOzy/fII49cdHwlvV+joqL0t7/9zbxm0dHRCgkJUfv27RUXF6c33njDPEbdunXVvXv3Elw9z2vZsqVluV+/fnrzzTcVGRmpH374QatXr1a7du3Mc/bkwwW7du2qiIgIM6E8aNAg7d27V40aNdLOnTs1YcIENWvW7IKr3QEAAAAAAK4Y8fHxRnp6epl8xcfHG6mpqUZBjzzyiCHJkGQ0adLEOJ93333X3FaS8dVXX1nG165da1StWtWyzbmvbt26GUlJSeY+u3fvtozHxMSYYw6Hw3j00UfPO1/+a8qUKZZYsrOzjccff/y8+wQEBBjz58+37Ddt2jTLNg6HwzL+zDPPmGPh4eHGqVOnzLFjx44Zbdq0+ctYo6OjDZfLZZl37NixF3SeTz75pJGTk2Pu53K5jLZt2xbaLiIiwtzm1KlTRsuWLc87b7169Yw9e/ZYYho0aJA53r179yLviYiICHObjz/+2Fw/b968Yo/Vpk0b4/Dhw0ZQUJBl/fDhww3DMIwBAwaY6/r27Wv079+/2LneffddSzznu5/OVZL71TAMIyEhwejUqdN59wsLCzM2bdrkset5Iec1Y8aMYu9dl8tlPPjgg8XG+8MPPxiPPfaYufzmm296LHbDMIxFixYZfn5+RR779ddfN6Kjo83lzp07F/uZAQAAAAAAnGPPBWxz2boq22jke/HFF9WgQQNz+dVXX9WxY8fM5TZt2mjt2rV6+umnCz2kLyoqSp988olmzZpl9ln+K4GBgfr666/10UcfWXr+5vPz81PPnj21atUqPfXUU5Yxu92uL774QtOmTTMrngvq37+/Vq9erQceeOCCYsmXX8kqSSdPntR7771njkVGRmrJkiX68MMPLb1982O9//77tWzZMr3//vuF2l+88sor+s9//qPbb7+9yON26NBB06dP1+TJk+Xr62uut9ls+vzzzwtVrhbsW12tWjUtXbpU7733niIjIy3bhYSEaPjw4VqxYoXls3VX9+7dNXXqVEVERJjrgoKCNHjwYC1atEi1atXSF198YbkX8t+np6eb68LDwzV58mT93//9n2XbyMhITZo0qdCDGi9GSe/XqlWr6ocfftC4ceNUt25dy1hYWJheeuklbd682dIX2ttsNpu+/PJLPfvss5b1zZo106JFi9StWzfzvlYRrU/cdc8992j9+vV66qmn1KBBAzVo0EC9evXS/PnzNWLECEtbnYL3NwAAAAAAwJXMFh8fbwQFBXk7jhJxOBwKCAgw2zeUpoyMDJ04cUJZWVkKDQ1VaGiofHxKnqs3DEMJCQlKTExUVlaWgoKCFBERocDAwAva/+TJkzp79qwCAwN1zTXXqHz58iWO5ULjjYuLU3Jysvz8/FStWjUFBwdf0L4pKSk6ffq0MjIyVK5cOV1zzTWWRGBRXC6XTp48qczMTFWuXFlVq1YtcrucnBydOHFCaWlpqly5ssLCwiwPj/O0rKwsxcbGyjAMhYeHF7oGTqdTiYmJCg4O/sv7Mi0tTSdOnJCvr69q1KihcuXKeSzOkt6vhmHo+PHjSklJUfny5RUeHu7RuErD2bNndfLkSVWoUEHXXnvtBbW0KW0DBgzQN998I0l69NFH9e2333o7JAAAAAAAUDbsleS5CspLjGQzAFyEpKQk7dixQ6dOnVJcXJwefvhh1ahRwxzPyMhQo0aNdPjwYUnSiBEj3KpYBwAAAAAAV5UynWz2fgkgAJQhKSkpuuOOO5SdnS1J2rx5syZMmKAqVaooNTVV77//vplolqRu3bp5MVoAAAAAAIBL56ru2QwAF6tmzZp69913zeXp06erevXqioqKUnBwsMaMGWOOjRw5ssj+7AAAAAAAAFciks0AcJGGDRum6dOn6+abb5YkZWdna//+/eZ4kyZNNGvWLL322mtejBIAAAAAAODSomczAJSQy+XS/v37dfr0aSUlJSkoKEg1a9ZU3bp15evr6+3wAAAAAABA2UPPZgC4Gvn4+Kh+/fqqX7++t0MBAAAAAADwOtpoAAAAAAAAAADcRrIZAAAAAAAAAOA2ks0AAAAAAAAAALeRbAYAAAAAAAAAuI1kMwAAAAAAAADAbSSbAQAAAAAAAABuI9kMAAAAAAAAAHAbyWYAAAAAAAAAgNtINgMAAAAAAAAA3EayGQAAAAAAAADgNpLNAAAAAAAAAAC3kWwGAAAAAAAAALiNZDMAAAAAAAAAwG0kmwEAAAAAAAAAbiPZDAAAAAAAAABwG8lmAAAAAAAAAIDbSDYDAAAAAAAAANxmT01NldPp9HYcJZKVlaVvv/1WiYmJ3g4FAAAAAADgslO5cmUlJSV5O4wrTrly5cpsPq0sqFSpkpKTk70dhlekp6eHlS9f3tthlEifPn1k9/f3V7ly5bwdCwAAAAAAAACgDLPFx8cbQUFB3o6jRBwOhwICAlShQgVvhwIAAAAAAAAA7torqYG3gygpe2lNnJaWpkOHDikzM1N16tRR1apVS+tQAAAAAAAAAAAvK5Vk886dO/X6668rPDxcFStW1M6dOzVs2DDdcccdpXE4AAAAAAAAAICXeTzZnJWVpTfffFP9+vVTr169JEm//PKL3n//fd10002qUqWKpw8JAAAAAAAAAPAyH09PeOLECSUmJqpr167mug4dOsjlcik2NtbThwMAAAAAAAAAXAY8nmwODw/X9OnTLQ/ti42NVU5ODn2bAQAAAAAAAOAK5fFkc0BAgGrWrGkuHzlyRG+++aa6deum6667ztOHAwAAAAAAAABcBkrlAYGSlJmZqdmzZ2vGjBnq16+fHn300dI6FAAAAAAAAADAy0ol2Xz69Gm99tprioyM1NSpUxUREVEahwEAAAAAAAAAXCY8nmx2uVz6v//7P7Vr105PPvmkbDabpw8BAAAAAAAAALjMeDzZvHv3bh04cEDPPPOM9u3bZxmLjIy0PDgQAAAAAAAAAHBl8Hiy+fjx4/Lx8dGrr75aaGzChAlq0qSJpw8JAAAAAAAAAPAy2+nTp42AgABvx1EimZmZCgoKoloaAAAAAAAAwJVgr6QG3g6ipOwpKSlyOBzejqNEcnNzFRgY6O0wAAAAAAAAAOCqZ69cubKCgoK8HUeJOBwOHkAIAAAAAAAAAJcBH28HAAAAAAAAAAAo+0g2AwAAAAAAAADcRrIZAAAAAAAAAOA2ks0AAAAAAAAAALeRbAYAAAAAAAAAuI1kMwAAAAAAAABcHsK8HYA7SDYDAAAAAAAAwOUh1NsBuMNeWhP/+eefOnz4sPz8/FSvXj1VqFChtA4FAAAAAAAAAPCyUkk2L1myROPHj1ejRo0UFxcnp9OpUaNGqX79+qVxOAAAAAAAAACAl9ni4+ONoKAgj02YmpqqHj166J133lGbNm3kcrk0atQoJSQkaPz48R47jiQ5HA4FBARQNQ0AAAAAAAAAXubxns0nT55UUFCQWrdu/b8D+PioXbt2OnLkiKcPBQAAAAAAAAC4THi8jUa9evU0b948czk7O1tr1641k88AAAAAAABAPsMw5HK5ZLPZ5ONTuC4yNzdXkuTr6+uF6Dxv8eLF2rp1q2655RZ17NjRI/tnZGRowYIFOnbsmGrUqKGePXuWQuRW//znP3X27Fn16tVLderUueD9XC6XDMMwl6+UzxX/4/HK5oLef/993Xfffdq7d6+ee+650jwUAAAAAAAAyqBJkybJbrfrscceKzS2Z88e2e122e12JScnX9S8hmFozpw5OnjwoAejdd/atWsVHR2tzZs3e2z/oUOHqlevXhoxYoT+85//SJfg/L/55htFR0fr5MmT5rpffvlFMTEx592vYcOG5mdqt9sVHBysVq1aafjw4dq3b1+JYtmzZ4+l+BXeU6rJ5oEDB2rUqFGqVq2a3nvvvdI8FAAAAAAAAGDaunWrevbsWeKk7uXqnXfeUU5Ojl5++WVz3fz586W8qudZs2ZJl+D816xZo5ycHLVr107Kq65+9NFHNXv27Avav2PHjurdu7datWql/fv364MPPlCzZs30/fffX3QsI0aM0JAhQy56P3iex9touFwuKa9Xc0REhPnq0aOHEhISFBoa6ulDAgAAAAAA4CqxZs0aHTlyRK1atVLlypU1b948JSUlqXXr1urQoYMkadmyZZo5c6a5fW5urnr27Cm73S6Hw6GlS5fqwIEDCgwM1K233qrmzZsXOX96erqWLVum/v37q1q1aoqLi9Py5ct18uRJlS9fXi1atFDbtm1ls9nM/Z1Op5YtW6a9e/eqfPnyuummm9S2bdtC5+Hj46MzZ85o/vz5Sk5OVqtWrczErSR9//33cjqd6tq1q2JiYrR37169/PLL2rhxow4ePKgGDRrouuuu05IlS5SQkCBJWr16teLj4+Xv71/s+ReUnJysH3/8UQEBAXrwwQe1YMEC7d69W1FRUbr//vvlcrk0f/58xcbGqnnz5urSpYu578KFC5WWlqaOHTsqPj5eP/zwg06fPq19+/Zp5syZ6tSpk8LDw4v9HIcOHap77rlHkpSamqoRI0Zo9OjReuyxx9SkSRPVq1dPyqvQXrt2rbZt26aMjAzVrVtXd999t8qVK6eTJ09q+fLlmjNnjipXrqyZM2eqUaNGatq0qXn+Xbt2VaVKlSRJO3bs0K5du1S7dm21bdvWrfNHMeLj44309HSPvb799ltj4MCBlnUnT540WrRoYezatcujx4qPjzdSU1MNAAAAAAAAlE2ffPKJIcno3bt3obHdu3cbkgxJRlJSkmEYhvHaa68Zkozx48cbTZs2NcclGTNnzjQMwzC6d+9uWS/JcDgcRlxcnHHzzTcbkozOnTsbDRo0MCQZn3/+uXnMYcOGGZKMUaNGGSEhIYYkY8eOHcaaNWuMoKAgIygoyGjXrp1Rq1YtQ5Lx0ksvmfvGx8cbHTp0MCQZ1apVM/f/xz/+YeTm5hqGYRhvvPGGIcl45513jObNm1tinDVrljlX/rl9+OGHhiQjKCjIsv+4ceMs1yf/1b59+2LP/1zHjh0zY50wYYJl+08++cR4+eWXLeu+/fZbc982bdoYkozffvvN+PjjjwsdLyYmpsjPOyoqypBkLFq0yLI+KyvLvHbvvPOOYRiG4XK5jL///e+GJKNu3bpGixYtDElGp06dDIfDYcTExBQ67scff2wY/2sKbUgyDhw4UOheGzp0qNvnj6J5vI1G48aNtWvXLu3cuVPK++3DwoULFR4ersjISE8fDgAAAAAAAFcRPz8/Ka91wpgxY3TmzBmzpcTEiRMlSXPmzNGAAQMkSTNmzJDT6VRgYKBGjhypTZs2adSoUVqxYoU2bdqktm3bavDgwTp+/Lhl/o8++kgffPCBtmzZolq1amny5MlyOBxauXKlVq9erf379+vvf/+7zp49a1YWjxkzRqtWrdLjjz+uY8eO6fDhw2rdurUmTpyon376yXIeEydO1KhRo3TmzBm99NJLkqQJEyaY4wEBAZKkqVOn6scff9SqVasKXYsGDRrI6XQqKChIkrRr1y6tXLmy2PMv7lqePn1asbGxSkhI0FtvvSVJevfdd+Xn56fExES9/vrrkqR//etfRX4mL7zwgqZMmWK+dzqdatOmzQV+ov8/lq5du0qSNm7cKEnat2+fFi1apPDwcG3atEkbNmzQfffdp59//lmLFy9WmzZttGXLFklS7dq15XQ69cILL1zUMT1x/vj/PJ5svv766xUdHa3Bgwfrb3/7m/r06aMFCxbo7bffNj9AAAAAAAAAwB0PPfSQ7rzzToWEhJiJ1W3btsnlcslut8vX11eS5OvrK39/f+Xk5GjGjBlSXpL24MGDiouLU7t27ZSZmanVq1dLktkSo2PHjnr66afVvHlzVaxYUf8rlpUmT56s5cuXKykpSRMnTtTUqVMVGhqqnJwcff3111Lec8z8/PxUqVIlTZ8+XVu2bFHTpk2Ljf/xxx+X8vos57eozY/j2Wef1b333qsWLVoUugY2m03+/v7mcsEH7517/n/lmWeeUdWqVdWjRw/pf90Q9Nxzz6lKlSrmul27dhW5r6+vr9miw8fHR/7+/pbWIheqevXqkqSkpCRJUv369c3PKb/tSN26dSVJsbGxstlsltYg/v7+5nlfLHfOH/+fx3s2S1KXLl106623KjY2Vn5+fqpRo4bKlStXGocCAAAAAABAGZaflMxP5hZUcN25ycv8pKMkhYSESJIyMzOVk5NTZHI1KSlJZ8+elSQ98MADhcaPHTtmWb7xxhsty6+88orWrVunr7/+2kwq33777Ro+fLg6d+6s5ORkxcfHS5KuueaaIuMsSfxRUVFF7u9p+f2V8/sbS1K1atUkScHBwVJeb+XSdOLECUlSRESEJCkrK0vjxo3TP//5T7PyPF9OTo5Hj305nP+VoFSSzZIUGBioBg0alNb0AAAAAAAAuAKUL19eknT48OFCY3FxcVJeS4n87fJdbAVrwULIRYsW6YYbbrCMF0wynru9JDVt2lTbt2/XunXrtHHjRq1cuVLLli3TqlWrtGHDBl1//fXmtk6n8y/judD4L1UBZ1GVyCWpTi4pp9OpOXPmSJL5oMSvvvpK0dHRuv3227Vy5UqFhobqnXfesbQb+SsFk9JpaWnFbuft879SeLyNBgAAAAAAAHChGjduLElav369VqxYYa5PT0/Xp59+KuW1nChpe4Tc3FxJUsWKFXXzzTdLeYnNunXrqm7dusrNzZXT6TR7JBclJydHv//+u3bs2KHOnTtr2LBhWrp0qaKjo5Wdna0dO3YoODhYrVu3liTt2LHD3HfkyJG69dZbzWroSy3//C+VklQcZ2Zm6r333tPWrVsVFhamhx56SCrQu7lr16664YYbVL58ecXExEiS2W4kX1ZWlmW5Xr16kqSDBw+acS1atKiEZ4ULVWqVzQAAAAAAAMBfadGihfr27asZM2bonnvu0UMPPaQqVarol19+0d69exUSEqLhw4df9LxVqlSRJP373/9WYGCgunTposGDB6tfv34aPny4XC6XXC6XXnrpJWVkZGj79u2qWLFisfP1799f27Zt08SJE9WoUSOlpqbq119/lQokzIcMGaKePXtq+PDhstvtSktL04gRI+Tn52dW614qRZ1/hQoVSu14+ZXhCxcu1G233aaWLVuqdu3axW4/cuRITZ06VVlZWdq0aZNOnDihSpUq6d///rfZRqNmzZqSpNmzZ6tmzZpatGiRWeG+bt067du3z2xxceLECU2aNEmtWrVSy5Yt1aVLF+3fv18vvfSSDh06pJiYGPNBikW1bIFnUNkMAAAAAAAAr7HZbJoyZYo++OADhYeHa/bs2frnP/+pI0eOqH///lqzZo0aNWp00fP26dNHNWvW1JIlSzRgwAClp6erb9+++vzzz5Wenq6HH35YjzzyiBo0aKDly5ebCc6i2O12fffdd3rooYf07LPPqkOHDuratatyc3O1YMEC3XTTTZKkHj16aNq0aSpXrpz69OmjQYMGqU2bNlq2bFmhth2lrajzL02dOnXSrbfeqqNHj6pXr15mRXFxVq9erTlz5mj58uWqXr26Xn/9dW3YsEF33XWXuc0zzzyju+66S+vXr1fv3r0VGhqqOXPmqF27dvrvf/+rcePGqWbNmnr++eclSS+88IKWLFkiSXrttdfUpUsXHTx4UK+//rpatGihQYMGSXmV1Cgdtvj4eCM/q1/WOBwOBQQElOpvZQAAAAAAAFA64uLitH//fnPZ5XIpJSVFubm5Cg4Olp+fn1vz5+TkKDU1VRUqVLDMlZubq+TkZJUrV65QL+i/4nQ6lZ6eroCAABWXU3O5XEpOTpbdbj9vtXRpK3j+aWlploRzQECAx5OuhmEoLS1N/v7+qlSpUqH5IyIiStQHOSUlRXa73VKZ7HA4FBQUZM6Xmpoqm81WKE+YkpIif3//87ZJuRgBAQFq06aNR+a6EtFGAwAAAAAAAF6xefNmjR8/3tthXBWCg4Mt1cb16tWzJPo9LSoqSvv27bOsq1OnjlJTU0vtmJdCtWrVNGvWLG+HcdmyHTp0yPD39/d2HCWSlZWlsLAwKpsBAAAAAAAAwMvs/v7+KleunLfjAAAAAAAAAACUYfZy5coV218GAAAAAAAAAIAL4ePtAAAAAAAAAAAAZR/JZgAAAAAAAACA20g2AwAAAAAAAADcVurJ5szMTP3888+lfRgAAAAAAAAAgBeVerJ56tSpGjVqVGkfBgAAAAAAAADgRaWabN60aZO+//770jwEAAAAAAAAAOAyUGrJ5qSkJI0bN07PPPNMaR0CAAAAAAAAAHCZKJVks2EYmjBhgnr16qW6deuWxiEAAAAAAAAA4EqT4O0A3FEqyealS5cqLS1NDzzwQGlMDwAAAAAAAABXonhvB+AOu6cnPH78uL744gt99tln8vEp9ecPAgAAAAAAAAAuAx5PNi9ZskTBwcGaOXOmJCk+Pl5Op1MfffSR2rdvr5tvvtnThwQAAAAAAAAAeJnHk82tWrVSaGiouezr6ysfHx/Vrl1bwcHBnj4cAAAAAAAAAOAy4PFk84033qgbb7zRXN62bZt+/PFHde/e3dOHAgAAAAAAAABcJmiqDAAAAAAAAABwmy0+Pt4ICgrydhwl4nA4FBAQoAoVKng7FAAAAAAAAABw115JDbwdRElR2QwAAAAAAAAAcBvJZgAAAAAAAACA2+wul0s5OTnejqNEXC6Xt0MAAAAAAAAAAEiyp6SkyOFweDuOEsnNzVVgYKC3wwAAAAAAAACAq569cuXKKssPCLTZbN4OAwAAAAAAAACuenZvBwCUNTt27NDhw4e9HQYAwAOqVKmi9u3bezsMAAAAALgikGwGLlJycrKcTqdq167t7VAAAG74888/FR8f7+0wAAAAAOCKQbIZKIHQ0FC1bNnS22EAANywc+dOHThwwNthAAAAAMAVg2QzAAAALohhGHK5XJZ1Pj4+RT5Dw+VyyTCMIscLzuPr66vc3Fzz/YXMkb+/zWaTj4+PR88RAAAAQMnx0zkAAAAuyH/+8x/Z7XbLKygoSK1atdLw4cMVFxdnbtuwYUPZ7XYtWbLEMseePXtUq1Yt2e12ffnll0pMTDTn2rNnj7ldcnKyAgICZLfbNXDgQMsc7777rux2u1544YVLcNYAAAAALhSVzQAAALgoISEhuuOOOyRJGRkZWrt2rT744APNmTNHMTExCgsLK3K/vXv3qmvXrjp69KimTJmip556SpLUuXNnrVy5Ujt27FDDhg2lvDYn2dnZkqQlS5YoJydHdvv/fnRdu3atJPFwRwBXrfxvfqiIb4WcO57/7ZA9e/Zo/vz5uvbaa/X444+XSlyLFy/W1q1bdcstt6hjx45uzfXPf/5TZ8+eVa9evVSnTh2PxXiuyZMnKzk5WUOHDtWSJUs8Fj8AXK1KpbI5OTlZsbGxlteJEydK41DFys3NtbzO/crnlSolJUUjR47UyJEj5XQ6vR0OAAC4AkVFRem7777Td999pwULFmjv3r2qV6+eDh48qB9++KHIffbu3at7771XsbGx+vLLL81EsyR16dJFkrRx40Zz3ZYtWyRJzZo10+nTp7V//35JksPh0K+//ipJ5vMTUlJSlJycrJycnFI8awC4fOR/e8Rut2vVqlWWsZycHNWuXdscz//79MiRI4qOjtb3339vbhsfH6/Zs2crOTnZI3GtXbtW0dHR2rx5s9tzffPNN4qOjtbJkyfNdb/88osdMWGFAAAgAElEQVRiYmLcnjuf0+nUkCFDtHz5ctlsNo/GDwBXq1JJNi9evFj9+/e3vN59993SOFSRYv8fe/cdV2X9P/7/gYIiQ0FQEkeYE5y4cIuWKy1HrqLUt75Ny1HZx1GkvXOQZpZm5WqoJZqKK8XQXKAQDlRUFCFFVEZHiCVDDry+f/w4148jDpRDmD7vt9u5veUar9fzuk7vc12vHRNTZIhn+fLlcXNz4z//+Q+nTp36x2L5p2VlZeHt7Y23t7fWG0gIIYQQojRVq1ZN62WcmJhYZH9kZCT9+vUjJiaGH3/8kTFjxhjtb9++PQAHDx7Uth09ehSA//73vwDa+9vFixfJzs6mYcOGWk83Nzc37OzsjCqrhRDiaREQEGD098mTJ4mNjS1yXJ8+fdDr9ezYsUPbtnXrVkaMGIFOp/tHYn0YR48eRa/X06lTJygo644YMYJffvnFZHlcvnyZzMxM2rRpA8Ann3yCXq9n6tSpJstDCCGeNqUyjUZcXBxjx45l+PDh2ra7LRzzT3jppZewtLREr9cTHh7OmjVrWL9+PYcOHaJjx45lEpMQQgghxJMkIyOD0NBQAJ599lmjfWfPnmXixIlcuXKFBQsW3HXodqtWrbCwsODUqVP89ddf2NnZERAQgKenp/a+FhoaipeXF+Hh4QD069evzN4vhRDicdG5c2c2bNjAJ598ok019PvvvwPQqVMnreGOgp7NwcHBVK1alRdeeIFNmzbx448/AvDrr7/SpEkTevXqBcDNmzfZt28f169fx87Oji5dutC4cWMtLT8/P3JycujXrx/BwcFcuHDBqILWzMyM+Ph4du3axd9//42HhwfdunUzij0+Pp59+/YRFxeHo6MjvXv3pnbt2tr+Xbt2kZGRgaenJzqdjp07d5KYmEhkZCS+vr50796dGjVqAHD+/HkCAwNJS0vj2WefpV+/ftja2gIQHR3NsWPHcHFxoV69emzcuJEuXbrQqlUrIiMjAWjSpAkUjLCJjo7G1dUVd3d3UlNT2b17N5aWlgwePJhDhw4RGhpKtWrVGDRoEPb29lq8aWlpKKWwtrbWvgshhHgalcov4PXr12nZsiWWlpalkfxDWb16NU5OTlAwRGbs2LGsX7+eVatWaYWXzMxMAgICiIqKolKlSnTu3Bl3d3ctjaNHj3L16lXatWvHrVu32Lt3LyNHjsTJyYmMjAwCAgK4cuUKFAwrfeGFF6hUqZJ2fl5eHoGBgZw5c4bbt29Tp04devbsiYODg3aM4WHdv39/4uPj8ff3x8LCgp49e9KoUSOja4qKiiIwMJCkpCScnZ3p06cPjo6O97wHheO3s7Nj27ZtpKSk4OHhQdeuXbXjfH19oaDwVqVKFQDCw8M5d+4cdevWpUOHDkYP20GDBvHrr79y/vx5GjVqxMsvv0x+fj7bt28nJiYGd3d3bVjsndd46dIlDhw4QNWqVRk6dChVqlTRHty1a9fmlVdeoWLFitq5D/qO7vWyk5eXR0ZGBoB2TUIIIYQomatXrzJjxgwoeL/av38/586dw9XVlRdffNHo2BkzZmBhYQHA+vXrGTt2bJH3lsqVK9O7d2927drF2bNnqVatGqmpqXTt2hVXV1dsbW0JCAhAKaUNbTb0dKNgyg2lFHZ2dv/A1QshxOOje/fuzJ07l1OnTtG2bVuUUmzZsoXOnTtTq1Yto2MvXLiAl5cX/fv3p1u3bnh5eWn7pk6dysCBA+nVqxfh4eH06tWLxMREPD09OXPmDBkZGWzevJkBAwYAMHfuXM6cOcPixYt5//33sbKyMqpsjo2NpWPHjmRlZWkjXn788UetwTE4OJiXX36Z27dv06VLF06ePKnNmdyuXTsAPv30U/744w+CgoI4efIks2fPhoKe3AEBAQQHB1OjRg2+/fZbJk6cSJ06dWjSpAm///47zZs3x9/fn+rVq3P+/Hm8vLwYP3480dHR7N+/nx9++IFWrVpx/vx5AK0ifdeuXcydO5fPP/8cd3d30tPT8fLywsnJieTkZKMpoJYtW0ZwcLBW9ndzc+PGjRsEBwfToUOHUvrGhRDiX0Cn06lbt26Z9NOrVy+1du1a9dlnn6lPP/1U/fbbbyojI8Pk+eh0OpWenq7udOXKFQUoQCUkJBjt+/HHHxWgevTooZRSKi4uTrVp00bb5urqqgC1cuVK7ZwZM2YoQC1YsEDZ29srQIWHh6uEhATVtGlTBajWrVurZs2aKUB1795dpaSkKKWUSktLUy+//LIClJWVlapZs6YCVM2aNdWxY8e0PFq2bKkAtWbNGmVra6vFb2trqyIiIrTjfH19FaAcHBxU165dFaBq1KihoqKilFJKJSQkaOca7s3MmTMVoL744gvVokULbT+gfH19tbQN2wxpKaXUsmXLFKCmT5+ulFLq2rVrClBOTk5q6dKlRmktW7ZMTZ061Wjbhg0bilzjihUrlIWFhXbM0KFD1ZYtW4zOmzx5snZecb4jw3UtXrxYu9dKKXXx4kUtzczMzCL/rTyKwMBAtX//fpOkJYQQouyEh4crPz+/sg7jX+XO57XhY2FhoUaOHKmuXLmiHduoUSMFqOeff15FRUWpgQMHas/927dvF0n7iy++UIBatGiR+umnnxSgfv31V6WUUoMGDVKAunbtmnJzc1OAunr16j967UII8Tgx/Mb+9ttvClDz5s1TSil1/vx5BaiFCxeqoUOHKkCdOHFCKaWUv7+/AlT//v2VUkrl5OQoKysrBagLFy6o3NxcpZRSU6dOVTVq1FBffvmlUkqpgwcPKkC5ublp+Xt4eChANW3aVO3evVvLY9asWQpQVapUUWfPnlX5+flq5cqVClB169ZVmZmZ6vbt29pv+cGDB5VSSkVFRSkrKyvl4eGh8vPzlVJKtW/fXgEqKChI6fV6tWrVKgWoSZMmqZycHJWfn68iIyO1MrGh7L9p0yYFKG9vb6WUUrt379bKsO+++646ceKEun79ulJKqcGDBytAJScnG8X/+eefK3VH+Xr06NEqISFBhYWFqSpVqihA+fv7a/fEUNYPDg4u1e9eCPFUiCjGMY8tk8/ZnJGRQVJSEr/99hvOzs5Uq1aNRYsWsXz5clNn9dCUUtpiAobewj4+Ppw4cYIFCxawf/9+Tpw4QYcOHXjnnXe4fv06gNYb58svv+TTTz8lLCwMFxcXdu7cyblz51i2bBknTpwgPDyc5cuX88wzz3Dx4kUAFi1axM6dOxkwYADx8fFcu3aNpUuXcuPGDcaMGcPt27cBtF7g33zzDX/88Qc3btzA09OT9PR0NmzYAEBubi5Lly7F3t6enTt3cvjwYRYuXEh8fDwrV66853Ub4p8/fz6fffYZycnJWqvzV1999VD30JBWYmIiMTEx3Lx5k48//hiAOXPmYGFhQVJSEh988AEA3333nXau4Rp/+eUXLl++rM0VtnnzZtatW0dcXBxbtmyBglbiv//+u9jfkSHtH3/8kd27dxdZJEMIIYQQptO6dWsSEhJISEggMTGR9PR01q5di4uLS5Fj33//ferXr8+KFSto1KgRmzdv5tNPPy1ynGHe5tDQUG1+5hYtWkDBMHGAAwcOEBERQcuWLalTp04pX6UQQjz+mjVrhqurK7/88gtKKQ4cOACFFl69nwoVKmj/Nqx1BLB48WLi4uKYOHEiycnJ2ojciIgIbt26BYWmyXzrrbd48cUXad26tVHar7/+Ok2bNsXMzEzrQX3lyhWio6M5f/48ERERWFpa4uzsTHR0NADu7u6Ehoby559/Fom1fPnyWnzlypWjQoUKmJmZaQvGtmvXjvT0dKKjo6lfvz4UjKZRSmmx5ufnM3/+fFq3bk3NmjXJz88nKCiIFi1aGE2HcS/Tpk3DyckJd3d3Bg4cCAWjug3CwsJISEigVatWD0xLCCGeZCavbC5XrhwLFixgyZIlDBo0iFdffZUFCxawceNGEhISTJ3dA82dO5cZM2Ywbdo0XnjhBVavXo2FhQVjxoxBr9ezfv16AFxdXYmOjiYuLo5OnTqRnZ3NkSNHoNCD1NPTk/Hjx+Pu7q7N/wSwadMm/Pz8iImJYfz48fj6+uLh4UFeXp5W2Tpx4kQqV66MmZkZ48aNw8LCgnPnznH27FmjeCdNmoSbmxvOzs68/vrrUPBQpqCi948//iA5ORl3d3eSkpKoWbMmFCy88yCDBw+mV69e2NvbM2rUKABOnz5Nfn7+I93bCRMm4ODgwJAhQ6BgJeO3336bqlWratvOnTtX5Lxx48ZRq1YtXnrpJe3FZfTo0dSoUYMBAwZoFceJiYkP/R3d+bJTr149rTD8OEzrIoQQQjwJLCwscHJywsnJierVqxtNfXUvTk5O/PTTT1haWvLxxx+zfft2o/0tW7bE1taWgwcPcvjwYRo1aqQNATc811etWgVA3759jc6Njo4mMjJSqwQRQoinRfny5RkxYgRnz54lIiKCnTt30qhRI62x7lGcPHmSF198EWtraxwcHGjevLm2T6/XGx1755SPBnXr1tX+bW1trTUQ6nQ6bTHC7OxsGjVqRIMGDWjQoIE2v3RcXFyxYzXUMezYsUNLx1DZGxMTo02pSEHFvJWVlfZ3bGwsOp1Oa+x8kMLTkhSeqtOgevXqODk5FeuZKIQQTzKTz9lsZWVVZOE9V1dXbGxsiI2N5ZlnnjF1lvf1zTffGP3do0cP/ve//9GmTRtu3ryp9Z41zD1V2LVr14z+LvyQBRgyZAh+fn4EBAQQFBQEBfM0TZkyhTfffJOUlBTi4+PhjgdTpUqVeO6554iMjCxSAV94UR1DRWzhB9jmzZvx8fHh9OnTRucZekjfj6GFF9BabrOzs9Hr9Uat2sVlWIyh8FzIhodu5cqVAUhPTy9ynqGC3MzMDGdnZ23uaQpa1GvUqMGVK1fIzc0lJSXlob6jO192zM3NtZiEEEIIUbbatm3Lt99+y5gxYxg9ejRHjx7VFmWqVKkSffr0YfPmzSQlJTF58mStMdnwDmaoiLjzXdPT01PmyRRCPLV69erFxx9/zJYtW9i3bx//+9//KFfu0fqVZWRkMHjwYGJjY1m7di19+vRBp9PRtGnTux5/r4rVwmVYpZRWprO2ttb22draEhYWVuTch6kzMHQoGjZsGPPnzy+yv/BaSoUrmgEuXboEBZXQxfGo91QIIZ42Jq9s/vPPP4mMjDRaHEav15OdnU3VqlVNnd0DhYeHU716dSh4sNnY2Gj7Cj8Y/f39adCggdG5dy4od+eD1N7ent27d3Pq1ClCQ0MJCQnBz8+PCRMmYG5uzvDhw7Vjs7Ozjc5NTk6GOx5+FLRM38vJkycZNmwYTk5OhISE0KBBAwICAowWdrif+6VdWOHW6sItwXe62wrwxVkV/m4P6XudV9LvSAghhBCPl9GjRxMWFsbXX3+Nl5cX+/fv1xrYu3fvzubNmwHw8PDQzrG3t8fT05NDhw4B0KZNmzKKXgghHj9t2rShbt26fPbZZwD07t37odPIy8uDglG1sbGxUNC5ysrKSutYRUHFcXEEBgby4YcfQkGlrqETUq1atbSyX3p6OhUrVqR27dpQUN61sbF5YJmucHnV1dUVgKioKJ577jnKlStHZmYmERERODo6alNv3E1ERIRRGiUVHR1NXl4etWrVwtra2iRpCiHEv5HJm+bKly/PggULOHDgAEoplFJs3ryZZ5991qjX7j/FMJTFycnJqKKZgpZUQ2ElJyeH+vXrU79+ffLy8sjJyXngtAtXr17lyJEjNG/enIkTJ/Lzzz+zadMmAI4dO4aNjQ3du3cH4MSJE9p5Fy5cQKfTYWFhgZubW7GvJTw8HAoKX+3bt8fBwUFrCX7UqTAKa9iwIRQ8JCl4iPv7+5c43ZIo6XeUk5NDZGQkkZGRxX4xEkIIIcTdvfLKKyilCAkJeeCxFy9eRClVZMoLMzMzli1bhlKK06dPaxXNFEyHZXh/vLMx/eDBg9q+O3u9Xb9+HaWU9GoWQjyVzM3NefXVV8nMzMTFxeWhGuQMc+2vWLGCPXv24OTkpK3Ts2TJEtauXcvs2bNp2bIlAAEBAXcdvWpgKHNduHCBGTNmsH37diZPngwF8zjXrFkTZ2dnJkyYAMDkyZMJCAhg6dKltGnThrfffvueHZEMHY127drFpk2buHLlCj169KBFixacOnWKmTNn8vvvvzNt2jTatm3L2rVr73vthtHCd3ZoelSenp40btxYK7cLIcTTyuQ9m11cXPjkk0/49NNPWblyJUoprK2tmT17tvbQepy88847vPHGG3z44Yfk5+eTn5/Pe++9R1ZWFmfOnDGam/lOy5cvZ+HChbz//vv079+f/Px8bdE7wzxRH3zwAQcPHuSDDz7AzMyMKlWqsHjxYm3fwwwRMkw1cfjwYX7++WeuXbvGyZMnoWBu5ODgYOrVq/fI96Jnz55cunSJ9957j8uXLxMcHKwNNSrLitqSfEcxMTE0btwYgMzMzCI9yYUQQgghhBDi365Pnz74+Pjw6quv3rc3750++OADxowZw5IlSwgLC+Pw4cN89913TJo0CW9vb1q0aMEPP/xATEwMb7zxhjY/9L0YpsiYPn06MTExDBs2jNzcXHr37m20OOyiRYuwsbFh5cqVWhl67Nix+Pj43DP+7t2707lzZ44cOcLw4cPZu3cvdevWZdu2bUyfPp1FixaxaNEiHBwcmDdvHtOmTbvvtR89epSaNWsaTXkphBDCBHQ6nbp165bJPzdv3lQnT55U4eHhKiUlpVTy0Ol0Kj09Xd3pypUrClCASkhIKLK/sPz8fLVy5UpVp04d7ZyePXuqsLAw7ZhZs2YpQH3++edG52ZkZKjp06cre3t77VwXFxe1dOlSdfv2be243bt3qzZt2mjH1KpVSy1ZskTl5uZqx7Rv314BKigoSNu2bds2BaihQ4cqpZTS6/Xq3Xff1dIZMGCAunHjhnrnnXcUoKpUqaISEhK0/YZ7c7f4Y2NjteNycnKUUkpdu3ZN9ezZUwHKyspKffbZZ8rPz08BavLkyUopddf075ZWVFSUls79rrFZs2YKUKGhodq2unXrKkCFh4cX+zu6W9pKKXXx4kXtnMzMzPv+t1BcgYGBav/+/SZJSwghRNkJDw9Xfn5+ZR2GEEIIUaYyMjKUTqdTeXl52rbs7GyVkJCg9Hq9ti0rK+uhy1Spqanqr7/+uuf+rKwsFRcXV+x08/PzlU6nUxkZGXfNKyEhwagsfi+5ubnq4MGD6tixY8XKVwgh/mERZR1ASZjpdDp150T5/xaZmZlYWloWmR7jUej1em7evEmlSpWKzAP8IIaF7MqXL4+9vf09h/3cvHmTvLw8HBwcHqq1+U4ZGRlkZ2fj6OhotM3S0rJE6RokJydTsWLFx26eqZJ8R6YUFBREbm4uPXr0KLMYhBBClNzZs2eJiopi8ODBZR2KEEIIIYQQQhhcAEwzoXwZMPk0Gv9W5ubmDzWlRWEWFhZUq1btgccVrhwuCRsbmyIV7KaocDcoi4Uci6Mk35EQQgghhBBCCCGEEKJ0mXyBQCGEEEIIIYQQQgghhBBPH6lsFkIIIYQQQgghhBBCCFFiUtkshBBCCCGEEEIIIYQQosRKrbI5Pz+fy5cvExISQlJSUmllI4QQQgghhBBCCCGEEOIxYJ6SksKtW7dMmmhmZiZLlizh+vXr2Nracu3aNaZPn07Lli1Nmo9er6dixYomTVOI4oiNjWXHjh1lHYYQQogSyMjIwMrKqqzDEEIIIYQoVStXrmTjxo1lHYYQopiys7OftbS0LOswHomPjw/m1tbWmPoCNmzYgLW1NWvWrMHc3JwtW7awcuVKfvrpJ8qVM11n6uzsbMzMzEyWnhDFUadOHZP/f0YIIUTZsLGxKesQhBBCCCFK1XPPPUe3bt3KOgwhRDHdunUr09ra+l/ZK8bBwQEznU6nTNmrJzMzk1deeYWlS5fSsGFDKKgUvnjxIm5ublSoUMGkeVlaWkpBUQghhBBCCCGEEEII8SS4ALiWdRCPytzUCcbFxZGZmUmdOnW4cOECycnJ1K9f3+RTaAghhBBCCCGEEEIIIYR4fJi8sjktLQ1ra2vmzJlDamoqlSpV4vTp08yYMYOePXuaOjshhBBCCCGEEEIIIYQQjwGTVzbfvn2bjIwMWrduzSuvvAJAUFAQc+fOpU2bNtjb25s6SyGEEEIIIe5r48aN/P3332UdhhBPreeff16bZlEIIYQQTy7zlJQUbt26ZbIE7ezsGDp0KJ06dUKn0wHQqFEjBg0aRFxcHHq93mR56fV6KlasaLL0hBBCCCHEk8vd3Z06deqUdRhCPHX2799f1iEIIYQQ4h9ibm1tjaWlpckSrFy5Mm+++WaR7WPGjDFZHgbZ2dmYmZmZPF0hhBBCCPHksbOzw9nZuazDEOKpY25u8gG1QgghhHhMmVtYWPxrewfn5eWVdQhCCCGEEEIIIYQQQgghSmPOZiGedKmpqSadekYIIUTZqVChAo6OjmUdhhBCCCGEEEI8EaSyWYiHFB4ezrlz58o6DCGEECbg5OTE4MGDyzoMIYQQQgghhHgiSGWzEI+gUaNG9OjRo6zDEEIIUQJnz54lKiqqrMMQQgghhBBCiCdGubIOQAghhBBCCCGEEEIIIcS/n/RsFkIIIYR4BIaFisuXL19kX35+PkopzMzMKFfu39m2n5aWxtdffw3A+++/X+wFpZcvX87ff//N8OHDqVevXilHKYQQQgghhHic/DtLP0IIIYQQZSgmJgZzc3PMzc1JTEw02ufn50f58uVxcnLi2LFjZRZjSWVlZeHt7Y23tze5ubnFPm/dunV4e3sTHx9fqvEJIcS/RVpaGj4+Pvj4+JCTk3PP4yIiIvDx8WHNmjUPdd6d9uzZg4+PD4cOHTJJ/IUppcjLy0Mpdc9999r/OLrbPV6+fDk+Pj78+eefZR2eEEL8K5m8sjkjI4OYmJi7flJSUkydnRBCCCHEY2Pr1q0MGTKEatWq4e/vT/v27cs6pEfm5OSEXq9Hr9djY2NT1uEIIcS/VnEb765evYq3tzd+fn4Pdd6dQkJC8Pb25uTJkyaJv7D/+7//w9zcnI8//thoe0pKCn379sXc3Jy33nrroeJ9GIcOHSI4ONhk6d3tHkujqRBClIzJp9E4f/4806ZNu+u+qVOnMnDgQFNneVeGoa3lypXDzMzMaJ9haOvdhr0+jgzx3u1airNfCCGEEKVv27ZtjBgxgmrVqrF7927atm1rtD8zM5OAgACioqKoVKkSnTt3xt3dXdt/9OhRrl69Srt27bCzs2Pbtm2kpKTg4eFB165dteP8/PzIycmhf//+xMfH4+/vj4WFBT179qRRo0YAJCYmsn///iIxOjo60qtXL+3v8+fPExgYSFpaGs8++yz9+vXD1tZWi3f79u0ADB06FAsLCwBu3rxJQEAA169fx9HRkW7dulG/fv0ieZmZmREVFcVvv/2Gubk5PXv2LHLc/fJPTU1l9+7dWFpaMnjwYA4dOkRoaCjVqlVj0KBB2Nvba+mkpaWhlMLa2hpzc5klTgjxZDA0+nGPKZseF6mpqYwYMYKAgADGjx/PV199RYUKFUyeT1ZWFiNGjGD48OF07NjR5OkLIYQwDZO/jbdp04a9e/cabYuPj2f69OlGBaXSFBMTQ926dQEYNWqUNgzJYMCAAezatYvw8HCaNWv2j8RUElOmTOGbb77h3Xff5csvvzTal5ubS/369YmNjcXf35++ffuWWZxCCCHE02rHjh0MHz6cqlWrsnv3blq3bm20Pz4+npdffpkTJ07Qo0cP4uPjmTJlCitXruTNN98EYNeuXSxYsIAvvviCtWvXcubMGe18X19fXn31VQDmzZvH6dOnWbNmDZMnTyY9PR0AW1tbQkNDcXV15erVq3h5eRWJc+DAgVpl87fffsvEiROpU6cOTZo04ffff6d58+b4+/tTvXp10tPTtTRefvllLCwsOHXqFC+99BI3btygTp06Wq+vNWvW8NprrxnlderUKT766CNSU1MBqFKlCqGhoVqFeHHzd3JyIjk5mXHjxmlpL1u2jODgYCpVqgSAm5sbN27cIDg4mA4dOpjgGxVCPC3Cw8MJCQkhKSkJR0dHunfvToMGDbT9SilCQkI4ffo0WVlZ1K9fnz59+hSZx/7SpUscOnSIlJQUateuTZ8+fYwaxQzMzMwICgoiNDSU6tWr8/LLL2NnZ3fX2O7V6JeTk8PevXu5cOEC1tbWtGrV6q6/feXKlSM5OZnt27eTmppKu3bt6NSpk9Ex8fHx7Nu3j7i4OBwdHenduze1a9cu9v1LTU3l1VdfJSAggAkTJvDVV19pcRpcv36dAwcOEBcXh42NDW3atMHDw0PrKBUdHc2xY8dwcXGhdevW7Ny5kz///JMGDRowYMAAzM3NOXPmDDt37iQxMZHIyEh8fX3p3r07VlZWWsNkv3798PX1pUqVKgwePBgKRl4HBATw559/Ur58edzc3OjRo0ex1yEwyM/PJygoiLCwMPR6PS1atOD55583agSQhk8hhCig0+nUrVu3Su2Tnp6u3n77bRUQEGDytHU6nUpPT1d3unLligK0T2BgoNH+/v37K0CFh4cXOfdxdPDgQQWoGjVqqOzsbKN9R48eVYCqWbOmyszMLLMYnyaBgYFq//79ZR2GEEKIEgoPD1d+fn6PdG7hd41vv/1WWVhYKEAdPHjwrsdPmjRJAWrBggVKKaVu3bqlOnTooCwtLdW1a9eUUkrNmjVLAcrBwUEFBASo5ORkNXXqVAWo9u3ba2m1b99eAapt27bq/Pnz6saNG8rT01MBatasWUoppbKzs1VCQoJKSEhQly9fVi1atFCAWr16tVJKqcjISO3dIiEhQSml1KZNmxSgvL29lVJKJSQkaNeYnp6ucnNzVevWrRWgVqxYoZRS6vz588rCwkJZWFioxMTEewTLPzYAACAASURBVMbXtWtXBajZs2c/Uv6jR49WCQkJKiwsTFWpUkUByt/fX7snNWvWVIAKDg6+53e2YcMGdeHChUf6voUQJePr66siIyPLOowivvrqK+23qHPnzsrBwUEBatOmTUoppfLz89XkyZMVoOrXr6/9Bnbv3t2o7LVu3Trt96pu3bpammfOnFHqjt+zlStXGpVVW7duraXl7++vANW/f/8i5xnKvTqdTvtNdXJyUvb29gpQU6ZMUXl5eUoVep588sknyt3d3Si/jRs3anEfPXpUOTg4KFtbW/Xiiy9q6YWGht7znhmeS7NmzVIpKSmqb9++ClATJ05Ut2/fLnL8jh07lJWVlQJUnTp1tOfl+PHjVW5urnYMoN588001evRoo3gnTpyolFJqyZIlRtsNv/mxsbHa/Z43b54C1JgxY5RSSl28eFE1atRIu1eG50fnzp21Z9bd7rHhORYUFKSUUio3N1eNHDlSAaply5aqc+fOClCjRo1Ser1eu9biPIuEEKKYIso6gJIo9QUC9+3bh6WlJZ07dy7trIpwcXEBwNvbWxt+dC83b95kw4YNLFq0iNWrV3Px4kWj/X5+fvj6+pKWlsaJEyf47LPP+O6777TeOocOHWLhwoX4+voWWbwhMzOTbdu28dlnn7Fs2TJOnTr1UNfRqVMn6tevT3x8PKGhoUb7AgICABg9erTWuyc8PJzly5ezYMECvv/+e6Kjo43OOXr0KL6+vkRHR3PmzBkWLVpEYmIi0dHR+Pr6EhwcTE5ODps3b2bBggX4+fkVuX8Pul+F87h+/TpfffUVS5Ys4cqVKwBERkayePFiVqxYUWRhpfz8fA4fPsyXX37JokWL2Lt3rzYtCgVTpKSmpmr3XgghhChLb7/9tvbvH3/80eiZBaDX61m/fj0Arq6uREdHExcXR6dOncjOzubIkSNGxw8ePJhevXphb2/PqFGjADh9+jT5+flGx02aNAk3NzecnZ15/fXXAbTnbMWKFXFycsLJyYkVK1Zw5swZRo0axdixYwE4fPgwAO3atSM9PZ3o6Ghtiov169ffdWGnCxcuaPN/GnpZu7m5ERoayvHjx7X3kLvFN3LkSKP4Hjb/adOm4eTkhLu7uzYl2/Xr17X9YWFhJCQk0KpVq/t8U0II8f/T6/V8/PHH2NvbExUVRVBQEBcuXGDAgAGcOHECCsos/v7+1KhRgxMnTnDs2DH69+/PwYMH2bNnDxSMqjX8th4/fpzLly/z008/ER8fz6RJk4rkGxISQkJCAmFhYVSpUoWTJ08+1EJ+n332GYGBgYwePZpr165x5coVPDw8+Oqrr/j999+Njv3qq69YsGABycnJvPfeewAsXboUCkbIjhs3jqSkJHbu3Mnu3bs5cuQIOTk5TJky5YEL/CUnJ/Paa6+xZ88eXnrpJZYsWVKkR/ONGzd4/fXXyczMxN/fn6tXrxITE0PTpk1ZuXIlW7ZsAdDO++WXX2jSpAnJycn89NNPAHzzzTckJiYyadIkVq1aBQXPl5ycHNq3b6+dGx8fz/HjxwkJCWH69Onk5+czbtw4IiMj+eSTT7hx4wbx8fG8/vrrHDlyhLlz5xb7nm/ZsoV169YxYMAAQkNDCQwMZOrUqaxdu1b770AIIUQhpdmz+ebNm6pv374qPDy8VNJ/UM/mvn37qgkTJihA/fzzz9r+O3s2nzlzRjk5OSlAeXp6Knt7e2VhYaG2b9+undOyZUutJ4+hNRZQQ4cOVVu2bDFqYZ08ebJ2XlxcnGrTpo0CVI8ePZSrq6vWov0wfHx8FKA+/PBDbVteXp7WUnv8+HGVn5+vtWIDysXFRfv3999/r503Y8YMrXeVoSU8PDy8WC3Kxb1fM2fOVIBauHChql+/vlFMISEhWq8BQ+uwocd2cVqNL168qJ1bFr25pWezEEI8GUzVs3ny5Mnq4sWLWm+2xYsXGx2r0+mK9MYq/Pnss8+UKtQTbeHChdq5hh5bgMrJyVGqUI+rQ4cOacdt27ZNey8pbOvWrQpQTZs2VcnJydr2OXPm3DemtLS0Ir299u7dq/UOu5/ixPew+aempmppTZ8+XQFq2bJlD/WdSc9mIcrO49izOScnRysLzZ07V4WEhKiMjIx7Hp+WlqZ0Op169913jX7rf/jhB63saZCVlaX++OMPFR4ervR6vdHv2fnz57XjRo0apQC1atUqpYrRszk3N1dVq1ZNAerw4cNaOlFRUSosLEwbKWJ4nowbN0475syZMwpQlpaWKi8vT506dUr7OzIyUkVFRamoqCjVqVMnBaioqKi73gdDz+bCHysrK3X8+PEix65Zs0YrMxZm6KX86quvGl23i4uL1jtar9dr309YWJjRvZ4yZYqWVuF7dOzYMW37uXPntO03b97UtgcHBytA2draaqOAHtSz+Y033lCAmjdvnnaffvrppyJl/8TERJWQkFBkNLIQQjwC6dl8L7t376ZFixbUq1evNLO5p7S0ND744ANsbW2N5gy809q1aylXrhxffvklBw8eZOvWreTm5vLhhx9qx1haWkJBa+vly5fZsWMHAJs3b2bdunXExcVpLbPLli3j77//BsDHx4cTJ06wYMEC9u/fz4kTJ+jQoQPvvPOOUY+cBxkwYAAAGzZs0HoZnzlzhsjISFq3bk3r1q357bffmDt3Li4uLkRFRXHlyhWtt9TYsWO5dOkSFGo5/vLLL/n0008JCwvDxcWlWC3Kxb1fhrSWLFmCr68vly5dokaNGsTExPDGG29w4MABIiIisLe35/Tp0xw9ehSk1VgIIcS/kLe3N40aNeLnn3/GwsKC999/32j9isLzQvr7+xMVFWX0GT16tFF6xV0E6kHHRUdHM378eCwtLVmzZo3R3KGG95phw4YVicewgOGdDIs9paamPrDH24Pie9j8y5Ur9cF4QoinTIUKFfj222+xsrJi1qxZdOjQAUdHR0aOHKmNDL19+zY+Pj7Url2bypUrU61aNZYsWQIFPaMp6L0L8Oyzz2ppW1pa4uHhQbNmzYr8FtaqVUv7t5OTExTMwVwcqamp6HQ6AKpXr65tr1+/Pu7u7lp6hbcbGJ4B2dnZ6PV6LZ3s7GwaNWpEgwYNaNCggVYui4uLu28sdevWJSgoiIULF5KZmYmXl5c2j7/BtWvXALS5+u+8B3eOwG3cuLFWjixfvrx2XG5ubrHuT+HrNcRfrVo1HBwctO01atQAID09neTk5GKle/XqVQA++ugj7T698cYbRa6hevXqODk5PfR80EII8aQptTf3vLw8/Pz86N27d2llUawY6tSpw0cffURMTAxff/31XY9bvHgxcXFxTJw4keTkZO1hFBERwa1bt4yOHTduHLVq1eKll17Sjhs9ejQ1atRgwIABWuEpMTHxoYfN3o+bmxvdu3fnypUr2jQc+/bt0/I3MzNj69atUFCxbHjQdurUSRtueuDAAShYlALA09OT8ePH4+7urq38TsGLyDvvvIO9vT2vvvqq9mJieGA/zP0aPHgwbdu2pUGDBrz44osADBo0iObNm+Pq6qotaGhI29/fH4C2bdsSGxvLn3/+ibu7O4BWcK9Xrx4JCQkkJCRo91sIIYQoax07duSrr74CMKqssLW1pU2bNlBQoVC/fn3q169PXl4eOTk5pfIsy8rKYty4ceh0OpYuXVpkwUJXV1cAoqKieO6556hfvz7Ozs6kpKRgbm5+14WNGjZsCAUVE4bps5RSvPzyy3Tr1k0bdl4cj5L//URHRxMZGVnkPUQIIe5nxIgRxMTEsG3bNry9vWncuDE//fQT/fr1IyMjgx9++EFrULx06RLJycm88847RmkYGuIyMzOLlWdJGs8KT1NRnArq+zX6GSpEbW1t79ro96BpiV5//XU6d+7M1KlTGT58OJcuXeK///2vUVyG59udv82GvytXrmy0vaSL6hW+P4a809PTjaa3Kvw9Fff5a21tDcDChQuL3KcVK1aUKGYhhHgSlVpl85kzZ8jKynos5s6bMGEC9erVw8fHh8uXLxd5iJ08eZIXX3wRa2trHBwcaN68ubbvzrmKa9asCQUVts7OzgDa/5qbm2stpbm5uaSkpGg9nAcMGKC1gn7++edQqKW3uAzzHe7duxelFJs2bQKgf//+UKhVtXCrOsBzzz0Hd8xtCBhdZ2EPalF+mPtVOBZDpXSdOnW0bY6OjkZpF6fV2NzcXJuH0lBxLoQQQjwOxo8fz7hx40hMTOSNN97QRlUZKic+/PBDtm7dypYtW3jhhRfo1q0baWlpJo/j+++/59ChQ9ja2nLjxg3mzJmjfZKSkujRowctWrTg1KlTzJw5k99//51p06bRtm1b1q5de9c0a9SowbvvvgvA5MmT2bFjB97e3vz6669kZWXRtGnTYsf3KPnfj6enJ40bNyY8PPyhzxVCPJ1u3brF6dOnuXnzJgMHDmTevHkcP36czp07c+nSJWJjYzl+/DgA/fr1o0GDBlhbWxMcHAwF68wA2m/f0aNHtTJNXFwcrVq14vnnnycjI8NkMVeuXBkPDw8oWKfHwMfHh86dO7NmzZpip2XonJSenk7FihW1htDU1FTy8vKK3TvX3Nycr7/+mmbNmuHv78/HH3+s7WvZsiUUrG+UlZWlbTfcw06dOhU73sIetB4TBY2alpaWZGdnG90rw9oDHTp0MBrxcz+GOo2kpCTtPtnY2JCRkWFUYS0Nn0II8f8pWdPhfRw7dgwPD48St06aQuXKlZk7dy6vvfYac+fOxc7OTtuXkZHB4MGDiY2NZe3atfTp0wedTnfPAtPdWqLvVeF557DZBg0aGO2vUqXKQ13Hiy++iIWFBZs3b2bIkCGcPHmSAQMGaAshWllZQUGPo8IMLziGFtm7xVfY/b4zU9yv+yncajx48GCjfYZeA0IIIcTjyszMjM8//5yzZ8/yxx9/8O677/L999/j5eVFZmYm8+fP55VXXgGgZ8+eLFy4UGvINiXD0OD09HTmzJljtG/kyJE4ODiwbds2pk+fzqJFi1i0aBEODg7MmzePadOm3TNdHx8frK2t+frrrxk4cCAWFhaMHj2aefPmPVQPbSsrq0fKXwghTCUpKYkOHTrw7LPPMn/+fJycnLh27RqRkZHUqVNH+1AwzWCdOnXw9/fXyit//PEHkZGRdO/enbZt23L8+HHeeecd+vbty5o1azh16hTvvfceNjY2Jq18fP/99xk2bBgffvgh5ubmZGRkMH/+fCwsLB6q8tbZ2ZkJEyawYsUKJk+ezFtvvcXFixd599136dGjh7YQfXE4Ojqybt06OnXqxMKFC2nWrBleXl54enrSu3dvAgIC+O9//8vw4cM5efIkK1euxMnJSVtYsbgM5eddu3bRrVs32rZtq5WB7xaTt7c3s2bNYvz48cyYMYPk5GS8vb0BmDVrVrHzfe2111i8eDFff/01devWxcXFhUWLFnHgwAF+++03bTS3p6cnN27cIDg4mA4dOjzUtQkhxJOk1GqCg4ODGTZsWGkl/9CGDh3KihUrWLNmjTYMlIJV0WNjYwEYMmQIVlZWBAUFafuLMyfhvRiGzZ44cUIbNkvBqsZ6vV4rlIWHh3P16lWeffbZe/Y2pmAOKC8vL9asWcP3338PBQ8+g44dO+Lv709gYCDjxo2DguFVhtWNDS3LJVGa94uCVuM9e/ZorcaANmWGoQd5Tk4OMTExUDCkV3o3CyGE+Ke5uLjc85lXuXJlQkJCimx/8803GTNmDDdv3qRSpUpFGp0NPY8Lq127dpF87pb2wIEDjY6bPXs2s2fPvu811K1bl82bN5OWlkZWVhZVq1Y1GoLs5ORUJO9KlSoxb948Zs2aRXJyMra2ttjY2Dx0fI+aPwUN0gsXLjTa9jDrYAghBAWjLffs2cPs2bMZMmSItn3YsGF89NFH2NjYMGHCBEJCQggICODVV1/lvffeY+HChQwcOJAdO3ZQrVo1Vq9ejZ+fHzNnzmT58uUsX74cBwcH/ve//zF9+nSTxz1kyBDWrl3LnDlztLJgjx49+PTTT4t0bnqQRYsWYWNjw8qVK7U1icaOHYuPj89Ddxpr2bIlq1evxsvLizfffJPGjRvTunVrfH19+eSTT/juu+/w9fWFgmfCvHnzjEa8Fkf37t3p3LkzR44cYfjw4ezdu/e+5eeZM2dSuXJlFi1apH3Hbdq04ccff9SmcywONzc39uzZw/Tp03nrrbe0bZs2bSrTaUOFEOKxpdPp1K1bt/6VH51Op60YW5hhhfj27dsbbQ8MDDRaNTc8PFwlJiYqCwsLBaj58+erNWvWKDc3N9WyZUsFqI0bN6q0tLQiK9IqpVSzZs0UoEJDQ7VthpXow8PDlVJKW6XW1dVV+fn5qc2bN6tatWopBwcHdf36daUKrRY8a9asB67ouGfPHgUoCwsLZWtrq1JSUrR9165dUzVr1lSAmj17ttqxY4fy8vJSgHr++ee1VewN+X3++edGad+5+vHdrrO49+tuedxt9fgpU6YoQP3www9KKaXOnz+vLC0tlZWVlVq+fLnas2eP6tGjhwLUb7/9ppRS6uLFi9p3mJmZ+cB7ZmqBgYFq//79/3i+QgghTCs8PFz5+fmVdRjiH7JhwwZ14cKFsg5DiKeSr6+vioyMLOsw7iktLU3Fx8er7Ozsu+6/efOmSktL0/7Oz89XqampKj8/3+i49PR0FRcXd890TCkvL08lJCSo5OTkEqeVlZWl4uLiSrVslZWVpW7cuKFSU1NLlE5+fr7S6XQqIyOj2Ofk5uaqhIQEpdPpSpS3UkrpdDql0+lUXl5eidMSQoj7iCjrAEriqVrau0uXLtr8vwbVq1fnu+++w9bWFm9vb7788kt++OEHZs6ciYWFBSNGjNDmEX4UXl5erFy5klu3bvHKK68wdOhQXF1d2bdvX5Fhs4V789yLp6cntWrVIjc3l5EjRxr1iqpVqxZ79+5l4MCBzJkzhwEDBrBz506mTJnCL7/8YpJpKEr7fhlajZs0acJbb71F3759SUhIkFZjIYQQQgghRKmwtbXlmWeeuec0gw4ODkYLqpuZmVG5cuUiIyxtbGyoUaNGsec7Loly5crh5ORU7HmH78fS0pIaNWpQqVIlk8R2rzycnZ2LLAr4sMzMzHB0dCwyReT9GNb8MawXVBKOjo44OjqWaKFHIYR40pnpdDp1r3mOHneZmZlYWloWGb75KHJyckhJScHR0VFbtTc7OxullEkeunq9/p7DZgEmTpyIu7s7//3vf0ucF0BaWhoZGRlUrVq1VFa5L+37BXDz5k0Aqlat+lg9zIOCgsjNzaVHjx5lHYoQQogSOHv2LFFRUUXWCBBPpo0bN9KyZUsaN25c1qEI8dTZsGEDrVu3NprOUAghhBD3dAFwLesgHlXZr973mKhYsSJOTk5G20xZSWtubs4zzzxz1303b97Ez8+PUaNGmSy/ypUrl7jV+H5K+35R0GoshBBCCCGEEEIIIYT4d5DK5sfAjRs3WLJkCW3bti3rUIQQQgghnlinTp0iMjKyrMMQ4qmTkZFR1iEIIYQQ4h9SapXNKSkpREVFUb58eRo2bGiSqS6eVC1atKBFixZlHYYQQgghxBOradOmZGVllXUYQjyVnJ2dcXBwKOswhBBCCPEPMM/MzCQvL8+kiYaFhbFgwQKqVatGTk4O2dnZzJ492+RzdOXk5JTKfMRCCCGEEOLJ0rRp07IOQQghhCgT+/bt4+jRo2UdhhCimG7dulXtYRZCfZy89tprmOfn55Ofn2/ShJcuXcrAgQMZPnw4SilWrVrF6tWrWbhwoUnzUUqZPHYhiiMrK4u4uLiyDkMIIUQJpKamYmZmVtZhCCGEEEKUqsuXL3P48OGyDkMIUUzZ2dlW/9bOtb1798bcxsYGKysrkyWanZ1NUlIS7du3x87ODoCOHTsSFBRElSpVTJYPQGZmJuXKlTNpmkIUR2xsLLGxsWUdhhBCiBK6c7FbIYQQQognzfjx4xk/fnxZhyGEKL6rgGtZB/GoTD5ns6WlJV27dmXHjh04Ozuj1+vZuXMnL7zwgqmzEqJMdOnShS5dupR1GEIIIYQQQgghhBBCPFbMdDqdMmXPZoD09HRGjhxJUlISAC4uLqxevZqKFSuaNJ/MzEwsLS1l8UEhhBBCCCGEEEIIIcST4MK/uWezyeeguH37NrNmzaJbt274+vqybt06GjRowPz581FKmTo7IYQQQgghhBBCCCGEEI8B88zMTPLy8kyWYGRkJKmpqYwePZry5ctDwfxAkyZN4tKlSzg7O5ssr5ycHP6tE2YLIYQQQoh/zrlz58jKyirrMIR4aj333HM4ODiUdRhCCCGEKGXm+fn55OfnmyxBZ2dnpk2bhpmZmZZuhQoVmDVrFnZ2dibNSyll0vSEEEIIIcST6dy5cyilTLowthCieP766y+qVKkilc1CCCHEU8DcxsbGpC/dVapUuWvvZTs7O5PlYZCZmUm5ciafCUQIIYQQQjyB3N3dady4cVmHIcRTZ8OGDWUdghBCCCH+IVJTK4QQQgghhBBCCCGEEKLEpLJZCCGEEEIIIYQQQgghRImZl3UAQvzbXL16lb/++quswxBCCGECNjY2uLq6lnUYQgghhBBCCPFEkJ7NQjyk2NhYrly5UtZhCCGEKKG//vqLCxculHUYQgghhBBCCPHEkJ7NQjwCR0dH2rZtW9ZhCCGEKIGzZ88SFRVV1mHcVX5+Pkop7e/y5cvf89iIiAi2b9+Os7Mzo0eP/ociLLm0tDS+/vprAN5//30qVqxY1iEJIYQQQgghSkh6NgshhBBCPGbc3NwwNzfXPpUrV6Zdu3Z8+OGHREZGGh179epVvL298fPzK7N4IyIi2LZt20Odk5WVhbe3N97e3uTm5mrbDx06RHBwcClEKYQQQgghhChtpdazOSUlhUuXLmFtbU29evWwtLQsrayEEEIIIZ5Inp6e1KhRg7/++osTJ05w/PhxvvzyS37++WdeeeUVAPr06YNer8fMzKzM4pw/fz4hISEMGjSo2Oc4OTmh1+uhUM/trKwsRowYwfDhw+nYsWOpxSuEEMWVl5dn9LeZmRnlyj0dfbaWL1/O33//zfDhw6lXr16xztmzZw+nTp2iY8eOeHp6mjymO9OXUTJCCPH4KZXK5iNHjuDj44OrqyspKSno9XoWLlzIM888UxrZFWF4IbhzyKlhSGq5cuWMCmRKKfLz8x/qxaG0H6JCCCGEENOnT6dv374ApKenM3/+fBYuXMjrr79Os2bNaNiwIVevXiU4OJiqVavSp08f7dz4+Hj27dtHXFwcjo6O9O7dm9q1awOQmprK7t27sbS0ZPDgwRw6dIjQ0FCqVavGoEGDsLe319LJyMggICBAW6+gUaNGvPDCC1SqVEnLY/PmzdjZ2eHr60uTJk1o0aIFeXl5HD58mPPnz3Pr1i1q165N7969cXR0BCAzM5Pt27cDMHToUCIiIti5cyeJiYlERkbi6+tL9+7d0el0nDt3rsi9MeRDwZQcSimsra0xN5dZ4oQQphETE0PdunWLbHd1dcXDw4MpU6bg7u5eJrH9E9atW8cff/xB165di13ZHBISwty5c/n888/vW05WSrFlyxbc3d2pX79+sWO6M33DKBmAKVOmSGWzEEI8BkzeJJucnMz8+fP58MMPWbx4MatXr8bDw4PVq1ebOqu7SkpK0oacRkREaNtTU1OxtLTE3Nyc//znP0bnzJkzB3NzcyZNmlTsfEJCQvD29ubkyZMmjf9R6XQ6fvnlF1JTU8s6FCGEEEKUAltbW+bOnUvXrl3Jzs5m48aNAFy4cAEvLy+++eYb7djg4GCaNWvGpEmTCAoK4qOPPqJFixYcO3YMCiquvby8ePvtt/nuu+/o3r07M2fOZOzYsfTo0YOsrCwAEhMT6dChA0OGDGHjxo2sW7eOl19+mX79+pGamkpMTAyjRo0iNzcXnU6Hl5cXhw4dQq/XM3LkSJ5//nm+/vprdu7cyX/+8x9atmypvZ8ZYvDy8iInJ4dDhw4xe/ZsAAICAvDy8iImJobAwEDtuMKfQ4cOadfr5uaGnZ0dx48f/0e/EyHE0+Oll15i6NChDBo0iNu3b7NmzRo8PDxk2p9HdOrUKYYNG/bQ5elPPvkEvV7P1KlTSy02IYQQJWPyrh+RkZFYW1vTuXNnAMqVK8egQYMYMWIEkydPxs7OztRZGnFwcKBHjx4cOHCA8PBw3NzcoGARIMN8gL/99ht6vV7r+RISEgJAly5dSjW20rR161YmTJhAVFQUVapUKetwhBBCCFEKLCws6NevH4GBgfesWM3NzWXcuHEkJSVx8OBBPD09iY6OpkWLFkyZMoWQkBAsLCygoDL56NGjJCQkEBcXR/fu3Tl9+jSHDh2ib9++7Ny5k3PnzrFs2TKtUX7FihUEBgZy8eJF2rdvT1hYGK1ataJu3bpcvHiR8uXLc/bsWXx9fRk2bBgbN27EzMyMI0eO8MUXX3DixAnt/aywSZMmYWVlxZtvvsmkSZNYvHgxFhYWNGvWjKFDhwIQFRVFz549AejatWsp3mkhhDC2evVqnJycAMjJyWHs2LGsX7+eVatWadP+5OfnExQURFhYGHq9nhYtWvD8889rI26jo6M5duwYLi4u1KtXj40bN9KlSxdatWr1wNEgBuHh4Rw9epTU1FSqVatGt27djHoGHz16lKtXr9KuXTvs7OzYtm0bKSkpeHh4FPndjIqKIjAwkKSkJJydnenTp0+R/O6UnZ3N77//TnR0NObm5jRv3pwuXboUmcrJzMyM+Ph4du3axd9//42HhwfdunUDYO/evfj6+mrx5uXlMWzYMMzNzR94D48fP050dDSurq737FV+4MABEhISaNu2LQ0aNICC593+/fupVKmSNuWTn58fOTk59O/fn/j4ePz9/bGwsKBnz540atRIS0+v17Nr1y4iIiKoSoHHxQAAIABJREFUXr06gwYN4tKlS1y5coU2bdrQsGFDkFE2QgihMfkvoF6vN1o9nYIKZ6UUSUlJpV7ZDNCzZ08OHDjA8ePHGTFiBABhYWEAtGzZktOnT3Pp0iXc3NzIzMzk8OHDALRt21ZL40EP8cLXlpyczPbt20lNTaVdu3Z06tRJ21/cBxjA9evXOXDgAHFxcdjY2NCmTRs8PDy0B/fdXk46dOhAdHQ0P/74IwC//vorTZo0oVevXuTk5JCdnY25uTnW1talcq+FEEII8c8yTEuWkpJy1/3nz58nIiICS0tLnJ2diY6OBsDd3Z2jR4/y559/Ymtrqx0/bdo0nJyccHJyYuDAgaxdu5br168bpblp0yZq1KhB69atGT9+PBMmTND2FS5QV6hQAQqGR1PQw/r777/Hw8OD9u3bs3Xr1nteV/ny5bW0ypUrp6VlY2ODjY0NWVlZzJo1i+zsbFatWmVUyRAWFoZS6h95zxRCiIoVK/LCCy+wfv16rl27BgXl4LFjx7Ju3TpatmyJjY0N06dPZ9SoUXz//feUL1+e8+fP4+Xlxfjx44mOjmb//v388MMPNG/enFGjRuHr60vDhg1xcHDgxIkTVK9enb179+Lm5oZSio8//pi5c+cC4OLiwv9j777DorjaNoDfVKnSRFQQQcGCIihKRI29xkZ8fTVRo8YSjT1qYiF2rJhiSywxduxiQWNvIGAvKIgaQBRQEQVpS1n2++PdmW9XQBZYXMv9u65cl9lZZp4ZhjnnPHNKbGwsAGDDhg0YOnQoACAwMBCLFy/Gr7/+is2bN+PWrVti3P7+/vj6668BADt27ED//v1hZWWF+vXr48KFC6hatSouXLhQ5LQWr169Qq9evRAUFIQmTZogOTkZMTExmDZtGhYtWqT03bi4ODRv3hxZWVl49uwZAGDjxo0YMmQI/vzzT3EapZUrV2LlypViAri4axgYGChOo1FUsnnTpk3YunUrdu7cKSabHz16hAEDBsDV1VU8lq+vL27evIlNmzZh3LhxSEtLA+QjiS5duoR69eoBACZMmIA//vgDAGBnZ4e1a9fC0dERe/bswd69e8Vks4uLC+Lj4xESEgIvL69S3FlERB8HtU+jUadOHaSmpooJ3Ly8POzevRuQv+l7F5o1awYAOHv2rPjZxYsXAQDDhw8H5MN2AODevXuQSCSoXbs2atWqBZlMhlmzZsHNzQ2jR4/G2rVrMXz4cDg7O+Pvv/8ucKy0tDR06NABw4YNw6RJk9CyZUvs2rVL3O7r64sBAwYgICAATZs2xaRJkzBu3Dg0bdoUkZGR4vcOHTqEOnXqYPDgwfjzzz8xadIkeHl54fvvvxcXzxEqJ1u2bMGAAQMwceJE3LhxAwMGDMClS5cAAJMmTcKff/4JyBPd5ubmGDhwYLlcZyIiInr34uPjAQC2traFbk9KSgLkvc/q1KkDZ2dnODs7i3WhhIQEpe/b2dmJ/1bstQcAffr0QefOnREUFIQ+ffrA0dERDRo0wNq1awt0LlDk7u6OKVOm4MmTJxgxYgQaNmwIOzs7zJgxQ2zMl9ScOXNw7tw5DBo0SKzPCSpXrgwbGxvO1UlE74RMJhOnzxA6EO3duxdbtmxBr169cOnSJVy4cAGTJk3C5s2b8c8//wDy0SkAcODAAbi6uuLq1avo1KkT7ty5I44GuXfvHkJCQnDmzBl4enri6tWrgHx07vz58+Hg4IAHDx4gJiYGwcHBgDxBe//+faVjLFiwAEuXLsXLly/FKSdWrFgByEfALF++HBYWFjh06BDOnz+PJUuWIDExEWvXri3yvI8fP45Hjx6hT58+uHz5Mq5evQozMzMsXrwYjx49Uvrupk2bcPjwYaV9zps3D1lZWdizZw8GDx4MANi+fTuys7NhaGio0jVUJwMDAwDA6tWrERYWhvj4eLRp0wZpaWnYsWMHIJ+qSkg0h4SE4PHjx5g5cyb27NkDFLJOFBERlUOyuXLlypg+fToWLFiAoUOHol+/fmJjxNDQUN2HK1Tjxo2hp6eHGzdu4Pnz58jJycHx48fRpk0bcYiTkJy9ffs2AKBbt27Q0tJSuRAXrFixAosXL8bLly/xww8/AACWL18ublelAIuPj8fAgQORmZmJo0eP4tGjR4iNjRUbc3v37gWKqJx0794d2dnZMDIyAuSFoVDwERER0cclOztbLOcVR1IpEhKupqamePDgQYH/GjdurPT9ty2ObGFhgSNHjuDKlStYtWoVBgwYgOjoaIwaNarQl/ACLS0t+Pn54cGDB9i8eTPGjh0LXV1dLFq0COPGjSvxeQcEBGDp0qVo0KABfvvttwLDtYmIytv8+fMxdepU/Pjjj+jQoQPWr18PPT09sUfx0aNHAflo2bi4OPz7779iz9sTJ04A8mcj5NNtLFiwAB4eHrC1tS0wGiQ8PFwcDTJo0CBAPm0i5G1SoedxixYt4O3tDcinjlDUu3dvdOrUCRYWFmJi9+bNm8jPz4eenh7CwsLw8uVLNGrUCMnJyeILzKioqCKvwVdffYVHjx5h9+7dSElJgVQqFRdQFHp4CwYOHIgGDRpAS0sLAwYMAADExMSI028ISVodHR1xJIsq17A8jB07Fi4uLqhWrZrYUUtYFFcYId2+fXuxt3KPHj0K7VV9/fp1PH36tEA5S0T0qVF7shnyB/GePXvwww8/YNWqVeID+10NbaxYsSI6d+4MyOdqvnfvHlJTU9GqVSvUq1cPpqamOH78OGQymbgggdBgK0shPmTIEEDeazo/P1/pe28rwE6dOoW0tDS0adNGXHG+WrVqYq+dQ4cOAW+pnAiFM+RDWYUhqN7e3nj69Ck2bNig1utLRERE755EIsH8+fNx48YNWFtbo3fv3oV+T6i/pKWloUKFCnBycoKTkxNSU1MhlUpL1Pv30aNHCA4ORsOGDTFmzBhs27ZNHLEmLDYoyMnJEf/94sULhIWFwcTEBIMGDcLKlSvF7wcGBhZ7XGFUF+TTiI0cORIGBgbYtGkTLC0tC3z/4cOHiIqKQkZGhsrnRkRUEqtXr8bSpUuxbNkynDlzBu3atcPp06fRpEkTQP68BICff/5ZHFHyzTffAPJnlCJXV1exsxBUHA0i7KNGjRpK+6pZsyYgn5JRkeJUGBYWFoC8HBGer3v27EGjRo1gZGSESpUqiW1UxWf5mxITEzFq1ChYWFjA0tISlStXxs2bN4E3ntsAxCQ0ABgbG8Pe3h5QGH1TmJJcQ3VSvKZWVlaAwggfYQoQxfPR0tISp+dQxFE2RET/o/Y5m9PT03Hjxg00bdoUrq6ugHwBPhsbm2IXG1Cndu3aITAwEDdu3BDnNmzatCkMDAzQoUMHBAQEID4+XkweC/M1q7MQV0wCv60AE94CvzmHszCstbjKSVGMjIxU+h4RERG9nxYuXIiNGzciJycHV69eRXx8PMzMzLB169Yip9GoVq0aRo0ahTVr1mDcuHH4/vvvce/ePUycOBHt2rXD8ePHVT7+n3/+iSVLlmDy5Mno3r078vPzcfDgQUA+kgzyl/yQj9RavXo1PD09kZCQAG9vb/Ts2ROjR4+GkZGRuCCz0CGgMMIix4GBgWjdujWaNm2K6dOnIykpCY0aNcKRI0dw5MgRQD6XszA0vE2bNpwnk4jK1e3bt1G5cmVAnjw1MTFR2i6skbNkyZICLwMV24WQt9MUCaNBRo4ciZCQEFy5cgUBAQFYtGgREhISsGnTJvFnJBKJ0s+mp6crHV/wtukdrl27hr59+8LGxgahoaFwdnbG8ePHxR7IRZk8eTJ27NiBKVOmYMKECTAxMcHnn3+OO3fuFPiu0NaFfNqRV69eFRqnopJcw7cROmnl5uaKn73tZeTbrpXwexaus0A4HyIiKkjtPZsrVKgAPz8/HDp0CDKZDOnp6diyZQt69+79TldkFeZtvnTpkjg/s5ubGwCgZcuWgLyXckREBNzd3cU3reosxFX9njDVxpsFoPD/QiNOwAQyERHRpyE4OBh79uzByZMnUaVKFUyfPh2XL19+a8IWAPz8/DBlyhScOXMGXbp0wcSJEzFs2DDs2LGjRPWxmTNn4qeffsLff/+Ntm3bon379jh9+jSWL18uDh23t7fHmDFjAPlIrmPHjqFnz55Yu3YtIiIi0KVLF7Rq1Qpz587F5MmTlaYbe1Pbtm3RsmVLxMXFoV+/fnj48KHYq+zGjRuYPXu2+N+6detUPg8iorISeq3a2NgUSDRD4QVccnKyOKLExMQE6enpYnuvKKqMBhGmg7xw4YL4c9nZ2Th37hwg7x2tKmEqSWHxVisrK3G6iDdH6ApkMpk4zYW3tzfs7OyQkpIiJprfnMdfMc779++LPbQV1wkAAKlUKv67LNdQkdC56969e+Jnp0+fVvnnFQk9mk+fPo3U1FRA3llMuO6KOMqGiOh/1J791dPTw4wZMzB79mwEBATg5cuX6NixI/773/+q+1Bv5e7uDlNTU5w9exYODg6oU6eOWLB5eHgAgNhIEaaugLwQP3r0KC5cuIARI0YAZSjESxIrAJw7dw5ZWVni3NbCohNFzclYGMXC+tWrV3j+/DkMDQ3FZDoRERG9/xQbyMXp2rVrgUa+iYkJ/Pz8MH/+fLx69Qrm5uZKa2fY2NgUusDfkiVLsGTJEvH/jY2NsWTJEvj6+iIlJQU6OjqwsLBQmjNZS0sLq1atwvz586GlpSVOm/bdd99h+PDhSE1NRW5uLiwtLZUS3YXFYGFhgQsXLiA5ORmGhoYwNjZGx44di70Gb448IyJ61/r3749ffvkFq1atgqOjIxwcHODn54czZ87g2LFjb31JePHixWJHgwgLyW/btg01a9aEh4cHdu/ejfv376N9+/Zo3769yrFWq1YNAHD+/Hls27YNjx8/FqeXvHPnDkJCQsTktkBLSwuurq4IDg7GunXrkJiYiOXLl6NNmzY4d+4czp07h7p164rP9cjISEydOhVeXl7iAnsDBw4UR+UIUyJt3boVhoaG6NixY5muoSJh1PIvv/wCAwMDpKWl4datW0AhSfHitGrVCg0aNMCdO3fg7e2N9u3bIzAwEPXr1xenEBFwlA0R0f+Uy5zNzZo1w65duzBjxgz4+/tjypQp77RXM+SLEXbp0gXJycm4du0aOnXqJDaMGjZsCMgLdSi8JYa8ELe1tcW2bdswe/ZsHDp0SFwYsKSFuKratGmDzp0748mTJxg+fDgOHTqE2bNnY+3atbCxscGwYcOK3YeDgwMAYM2aNeJKvf/88w/q1q2LCRMmqD1mIiIiev8ZGBigatWqZV6kWU9PD9bW1rC0tCxycT4LC4sC63Noa2vDwsIClStXVrkuqKWlhUqVKr11qDUR0fvGxcUF//zzD+rXr4/vv/8eXbt2xdOnT7F79+5ik6SqjAaxs7PDiRMn4O3tjXnz5qFXr144dOgQxo8fj127dpVomokOHTpg4sSJSE1NxTfffINLly5h+/btmDBhAhITE/HFF18U+nO//vor6tSpgy1btmDs2LEYPnw4Vq5cCVtbW8ybNw8nTpwQp8/46aefkJ+fj759++LkyZPo3LkzFi1aJO6rf//+sLe3x7FjxzB48GBkZGSU6Roq6t27N8aMGQOJRIKZM2fi2bNnWLZsGVDIdBjFqVChAnbt2gVvb29ERUUhKCgIixcvRt26dQGFKTuIiOj/aSUlJck+1GkZMjMzYWBgUOgwJsjnGRw9ejQAYNu2bUpzULVt21bsrZyYmCjO6wwAERER8PHxwYEDBwD5au7ffvstZs2aJQ7JmTVrFubPn49ly5Zh8uTJgHw4jdCDODs7G/r6+vDy8kJYWBiCgoLE6TsOHDiAL7/8Ev/973/FRXZevnyJuXPn4q+//kJmZiYgH57k6+uL+vXrA/Lk8RdffIHu3bvj8OHDSue6bds2DB06FLm5uWjVqhXOnz8Pf39/DBgwAN7e3ggICFDTVaegoCDk5uaiXbt2mg6FiIjKIDw8HA8ePChyoT36uOzcuRPu7u5igoCI3p0dO3bAw8MDtWvX1nQo78SLFy8Aee9dbW3V+3fl5+cXORpE0evXr5Geng5LS8sSTS/xpvT0dEgkEqW1lYQpK4o6tlQqxYsXL2Bubi4uhJeXlweJRFJou/z169fIzs6GtbV1gW25ubni6Js3k+WlvYZvnl9ubq64tlJZpKamiusKQD4158WLF3H27Fm0adOmzPsnInpDJIB6mg6itD7qZHNZqasQLwmJRIKXL1/CxMSkwFzNxcnIyEBWVlaZCmQqHpPNREQfByabPy1MNhNpzqeWbKaPR3p6OlxdXREbG4u9e/eiQ4cOOH36NP7zn/+gatWqiIyMVEpCExGpyQedbH63c1t8YCpWrFjihG9ZGRgYiHNolZSxsTGHnBIRERERERGpgYmJCX777TcMHz4cffr0ET93c3PDqlWrmGgmIioEk81ERERERERERIUQFga8c+cO0tLSYGNjgzp16ryz0c9ERB8aJpuJiIiI6JOQkpKChIQETYdB9MnJy8vTdAhEZWJqagovLy9Nh0FE9EFgspmIiIiIPgk3btzAjRs3NB0GEREREdFHSzclJQVpaWlq2+Hly5fh6elZ4POsrCxER0dDR0cHjo6O4sq1ZSGVSmFjY1Pm/RARERHRx+2rr77SdAhEREQaMXXqVCxdulTTYRCR6j7YxQEPHz4M3YoVK8LQ0FAtO3z06BH8/Pxw9uxZpc+fPHmC6dOnw9jYGDk5OdDT08OCBQtgbW1dpuNlZWVBS0urjFETEREREREREX2cRo4cia5du2o6DCJSUVZW1iNDQ8Mamo6jNFxdXaGrra0NHR2dMu0oNjYWoaGhCAgIAIAC+9uyZQs8PT3xww8/ID8/H/PmzcP+/fsxevToMh1XW1u7TD9PVFovXrzAlStXNB0GERGVwfPnzzUdAhEREVG5q1mzJmrWrKnpMIhIdZmaDqAs1JKtTU5ORnp6Otzd3Qtse/XqFU6dOgVvb+//HVBbG7169cKBAwcgkUjUcXiid8re3h6Ojo6aDoOIiMqocuXKqFfvgx2hRkRERERE9N5RywKBHh4e8PDwQEREBI4dO6a07cWLFwCAGjX+v/e3o6MjJBIJUlNTYWBgoI4QiN6ZGjVqKN3PREREREREREREpKaezW+TkpICCwsLpSkvjI2NAQCpqanlfXgiIiIiIiIiIiIiegfKPdksk8kgk8kKfAYA+fn55X14IiIiIiIiIiIiInoHyj3ZbG5ujpSUFKXEckZGhriNiIiIiIiIiIiIiD585Z5strGxgY6ODmJjY8XPYmJiYGpqCisrq/I+PBERERERERERERG9A7qZmZmQSqVq2VlOTg5sbGyQlpYmfqatrY0ePXrgxIkTGDBgAGQyGU6ePInevXtDIpFAIpGU+njZ2dlcYJCIiIiIiIiIiIjoPaAVExMjq1Chglp2lp2djcTERDg4OCh9LpVK8ezZM2hpaUEmk0FHRwfW1tZKiwaWRk5ODiwsLFCxYsUyRk5ERERERERERESkcZEA6mk6iNLSNTExgZGRkdp2WLly5UI/t7S0VNsxBJmZmWVOWBMRERERERERERFR2elqOgCiD01QUBDu3Lmj6TCIiEgNbGxs0Lt3b02HQURERERE9FFgspmoFOzt7dGoUSNNh0FERGUQHR2NpKQkTYdBRERERET00WCymagUDA0NUa1aNU2HQUREZZCcnIznz59rOgwiIiIiIqKPBic8JiIiIiIiIiIiIqIyY7KZiIiIiIiIiIiIiMqMyWYiIiIiIiIiIiIiKjO1J5uDgoLKtL2sZDIZpFIp8vPzC90ulUohlUohk8nKNY4PyevXr7Fw4UIsXLgQ2dnZAIB//vkHCxcuxLlz5zQdHhEREREREREREX0A1Jpsjo2NhY+PT6m3q8P+/fuhq6uLFi1aKH2en5+PadOmQVdXFy1btsSzZ8/UfuykpCTs2rULqampat+3qiIiIhAQEFCin8nKyoKPjw98fHyQm5sLAAgNDYWPjw+uXbtWTpESERERERERERHRx0RXHTuJjY1FaGhokUnO4raXt/z8fPj4+GDJkiXw8vJCQEAAbGxs1H6c/fv3Y9SoUXjw4AHMzMzUvn9VLFiwAKGhofjyyy81cnwiIiIiIiIiIiL6NKkl2ZycnIz09HS4u7vj2LFjJd5envLz8zFz5kwsXrwYLVq0wL59+wokmtPT03H8+HH8+++/0NHRgYuLC9q1a4cKFSqI35FKpTh//jzu3r2LjIwMVK9eHZ07d0alSpWQl5eH3bt3Y+PGjQCAw4cPo379+ujUqRMAIDExESdPnkRCQgIqVaqEzp07o3r16uK+9+3bh+zsbHTv3h2JiYk4evQo9PT00LFjR9SpU0elGIRj7NmzB+bm5vD390f9+vXh5uamUgyqunv3Li5cuIDXr1+jRo0a6NatG0xNTcXtr1+/hkwmg7GxMXR11XJ7ERERERERERER0QdALdlADw8PeHh4ICIiotBkcnHby4tMJsOsWbOwcOFCfP7559i3bx+sra2VvhMVFYVevXohKioKNjY2kEgkSE1NRcuWLbFv3z5UrlwZeXl5GDx4MPz9/VG7dm1YWVnh6tWrqFy5Mk6cOAFHR0cMGDBA3OekSZPg7e2NTp06ISQkBD179kROTg4+//xzXLt2DT/99BOOHTsGT09PAICvry9u3ryJTZs2Ydy4cUhLSwMAmJqa4tKlS6hXr16xMaSmpmLw4MGAfDqPAQMG4Pfff4ebm5tKMajijz/+wJgxY2Bvb4/69evj1KlTaNiwIY4ePYrKlSsDAFxcXBAfH4+QkBB4eXmp5fdIRERERERERERE7z+1LxD4vpBKpZg5cyYWLFgAa2tr7N27t0CiOT8/HyNGjEBUVBTmzp2L+Ph4JCYmYuDAgQgODsb8+fMBAHfu3IG/vz/69u2Le/fuISQkBGfOnIGnpyeuXr0KQ0NDZGdnw8jICAAQGRmJPXv2IDc3FyNGjEBycjIOHTqEI0eOIDg4GNnZ2Rg/fry4SKGBgQEAYPXq1QgLC0N8fDzatGmDtLQ07NixQ6UYmjVrhuvXrwMAHB0dkZ2djbFjx6ocQ3Hu37+PMWPGoGrVqrh8+TKOHj2K7du349q1a1ixYoUaf3NERERERERERET0Ifpok81XrlzBggULoKenh6SkpELni46MjERQUBAAYMyYMdDR0YGhoSFGjx4NANi8eTOys7PFhGxISAg2bNiA8PBwNGvWDPv378egQYMAAPr6+uJ+dXV1oauri7t37yIiIgIGBgaoVq0aHj58CABo1KgRLl26hH///VcpnrFjx8LFxQXVqlXDwIEDAQAxMTGAvJf222LQ0tJSmrZCX18fOjo6JY6hKOfPnwcAeHp6Ii0tDQ8fPoSTkxMAYPv27WJ8169fx9OnT9G4cWOV9ktEREREREREREQfh4822aynp4d169aJ03aMGzdOTCwLEhISAADW1tawsrISP69atSoAIC0tDS9fvoS7uzumTJmCJ0+eYMSIEWjYsCHs7OwwY8YMccqLwiQlJQEAJBIJ6tSpA2dnZzg7O+PixYtKxxfUqFFD/LcQT3Z2NgC8sxiK8vTpUwDAwYMHxX0ICeXY2Fikp6cDACpXrgwbGxul+a6JiIiIiIiIiIjo4/fRruDm4eGBESNGAAB+//13TJw4EQMHDkRQUBDs7e0Bhekr0tLSIJVKoaOjAwDIzMwU92NgYAAtLS34+flh5MiRCAkJwZUrVxAQEIBFixYhISEBmzZtKjQGIeFqamoqTnGhqEqVKkr/Lxy/MOUZQ0ZGRpHHFQjXqm/fvliwYEGB7YaGhsXug4iIiIiIiIiIiD5eH23PZkXjxo3DoEGDEBcXhyFDhojJ1Xr16sHAwAASiQS3b98Wv3/t2jUAgJeXFywsLPDixQuEhYXBxMQEgwYNwsqVK3H58mUAQGBgYIHjSaVSABCnmUhLS0OFChXg5OQEJycnpKamQiqVlqj3b0liyMnJEf+trhjq1asHAHjw4AFq1qwJJycnVKtWDSkpKeK0IQDw8OFDREVFqZTAJiIiIiIiIiIioo/HJ5Fs1tbWxu+//w4PDw+cPXsWU6dOhUwmQ6VKleDj4wMAGDlyJPbt24f169dj8uTJAICZM2cCAC5evAgvLy+MHDkSx48fR1BQELZt2wYA6Ny5s3gcBwcHAMCaNWvwzz//oFq1ahg1ahQgT3gfP34cy5cvR5MmTTB69GhoaWmpfA6qxFCxYkUAQHx8PFavXo0rV66oLYZ27drBzc0NN27cwLRp03Dq1Cn8+OOPaNq0KTZv3ix+r02bNqhbt65S8p6IiIiIiIiIiIg+fmqdRsPFxQUXLlwo9fbyZGFhgS1btqB58+ZYvXo13NzcMGLECEybNg0VK1aEn58f+vTpAwBo0qQJNm7ciK5duwIAevbsibVr18LPzw9dunQBABgZGWHy5MmYNm2aeIzp06dj6NCh+P3333H9+nV07doVfn5+MDExwdq1a3Hw4EEAwLBhw7Bw4UKlBf2Ko0oM9vb2GDNmDFavXo2xY8di3rx5aNq0qVpiMDIyQkBAAH766Sf4+fnBz88PVlZW8PX1xY8//qjyeRAREREREREREdHHSSspKUlmZGSk6ThKJTMzEwYGBjAxMSnzvvLy8pCcnAwdHR1UqlSp0O/k5+cjNTUVubm5sLS0LDRRm5GRgaysLFhaWkJb+/87jkskErx69Qrm5uZlmt9YlRhevXoFLS0tmJubK32urhhev34tnqOenl6p9/OhCgoKQm5uLtq1a6fpUIiIqAzCw8Px4MED9O7dW9OhEBERERERCSIB1NN0EKX10S4QWFK6urqwsbF563e0tbVhYWHx1u8YGxvD2Ni4wOcGBgaoWrVqmeNUJYaitqsrhooVK4pTdhARERERERERERHhU5mzmYiIiIiIiIghuf3LAAAgAElEQVSIiIjKF5PNRERERERERERERFRmTDYTERERERERERERUZkx2UxEREREREREREREZaabkpKCtLQ0TcdRKlKptNhF/YjKQ1RUFKKiojQdBhERlRHrEUREREREROqjW7FiRRgaGqpth8HBwWjZsmWBz9PT0xETE4OsrCzUqlULVlZWZT5WVlYWtLS0yrwfopJo2LAhatWqpekwiIhIDfT19TUdAhERERER0UdDV1tbGzo6OmrZWWxsLGbOnIkLFy4ofR4eHo7p06ejatWqMDU1RXh4OKZOnYoOHTqU6Xja2pwFhN49MzMzmJmZaToMIiIiIiIiIiKi94quOnYSGxuL0NBQBAQEFNiWk5ODWbNm4ZtvvkG/fv0AAOfOnYOvry8aN24MS0tLdYRARERERERERERERBqklq7BycnJSE9Ph7u7e4Ft8fHxSE5ORrdu3cTPWrVqhfz8fMTGxqrj8ERERERERERERESkYWrp2ezh4QEPDw9ERETg2LFjStuqVq2Kbdu2wcTERPwsNjYWeXl5apm3mYiIiIiIiIiIiIg0r9wnPTYwMIC9vb34/48ePcKsWbPQo0cP1KhRo7wPT0RERERERERERETvgFp6NqtCIpFg165d2L59O7755ht89dVX7+rQRERERERERERERFTO3kmy+dmzZ5g2bRrs7OywceNG2NravovDEhEREREREREREdE7Uu7J5vz8fPz8889o0aIFhg0bBi0trfI+JBERERERERERERG9Y7r5+fnIy8tTy87y8/NhYGCgtL+7d+/i8ePHGDVqFCIjI5W+b2trC2Nj4zId78aNG5BKpWWKm0pGW1sb+fn5mg6DiIg0QFdXV231BiJ69/T09GBoaKi2/enr6yMnJ0ct+1JXHTM3NxdZWVlqielTx3q/euno6LDtWghel/LH+lv50NLSgkwm03QYH61P+dmQlZVlpM762rvk6uoKrQcPHsh0ddXTwTkjIwNRUVFo3Lix+FlycjLi4uIK/X7t2rXLlGyWSqX4z3/+g1u3bpV6H1RyDRo0wJ07dzQdBhERaUCrVq1w4cIFTYdBRKVkaWmJly9fqm1/dnZ2ePLkiVr21axZM4SFhZV5P/b29kW2P6hkGjZsiNu3b2s6jI9Gy5YtERwcrOkw3jvq+tunorVu3Rrnz5/XdBgfHZY35YvPzA/T4cOHoZWUlCQzMjLSdCylkpmZCQMDA5iYmGg6FCIiIiIiIiIiIqKyigRQT9NBlJa2pgMgIiIiIiIiIiIiog8fk81EREREREREREREVGZMNhMRERERERERERFRmak92RwUFFTo58+fP8elS5dw/fp1pKenq/uwRERERERERERERKRBuurcWWxsLHx8fAqsUn/s2DH8+uuvqF+/PhISEpCdnY3Fixejbt266jw8EREREREREREREWmIWpLNsbGxCA0NRUBAQIFtaWlp+O233zBv3jw0a9YM+fn5WLx4MdatW4dff/1VHYcnIiIiIiIiIiIiIg1TyzQaycnJSE9Ph7u7e4FtiYmJMDIywmefffa/A2pro0WLFnj06JE6Dk1ERERERERERERE7wG19Gz28PCAh4cHIiIicOzYMaVttWvXVurxnJubi9DQUDH5TEREREREREREREQfPrXO2VwcX19fBAUFoUqVKli9evW7PDQRERERERERERERlSO1TKOhqm+//RaLFy+GjY0N5s+f/y4PTURERERERERERETlqNx7Nufn5wPyuZptbW3F//r06YMXL16gUqVK5R0CEREREREREREREZWzcu/ZHBgYiPHjxyt9ZmBgAABIS0sr78MTERERERERERER0TtQ7snmBg0a4M6dOwgPDwcAyGQyBAYGomrVqrCzsyvvwxMRERERERERERHRO1Du02jUrFkTPj4+mDBhAmrWrImMjAwAwJw5c6Cnp1fehyciIiIiIiIiIiKid0ArKSlJZmRkVO4HysrKQmxsLPT09FC9enVUqFChzPvMzMyEgYEBTExM1BIjERERERERERERkQZFAqin6SBKq9x7NgsMDQ1Rr94He52IiIiIiIiIiIiI6C3Kfc5mIiIiIiIiIiIiIvr4MdlMRERERERERERERGWmm5KSIi7a96HJycnB7NmzkZiYqOlQPhk5OTkwMTFBTk6OpkMhIiINCAsLg56eHiwtLeHo6KjpcIiIlEgkEhgZGSE/P1/ToXwUsrKyIJVKuUaOGkgkEshkMhgaGmo6lPdKZmYmtLW1YWBgoOlQiOg9Y2hoiPT0dOjo6Gg6lHdOIpHU+BCfi7m5uRg0aBB0jY2NP9gHu0QiQVZWFhwcHDQdyifj4cOHcHNzQ1JSkqZDISIiDTh37hwkEgmqVKnC8peI3jvh4eGoX78+0tLSNB3KR2HLli1o0aIFn/dqcP/+fYSFhWHQoEGaDuW9cuvWLURFRaFv376aDuWjlZCQgBMnTqB///7Q19fXdDhEKtu8eTMGDx6s6TA0xUjTAZTG48ePMXPmTOjq6emhQoUKmo6nVKRSKapXr47//Oc/mg7lk3Hs2DF06dJF02EQEZGGbNq0CVWrVkXPnj3Ro0cPTYdDRKTE39+fbQM12r59O1q1aoWOHTtqOpQP3pEjR3Dp0iUMGTJE06G8V7Zu3Yro6Ghel3J0/fp1MdnMUQr0Idm6dSufDR+YkJAQREZGQlfTgRAREdGHZeHChZxCg4iIiIiIiArgAoFEREREREREREREVGZMNhMRERERERERERFRmTHZTERERERERERERERlxjmbSS3y8/Mhk8mUPtPW1oaWllap9hceHo5bt26hZs2aaN68uZqifP/JZDLk5+cXuV24pqW5PlKpFDt37oRMJkOvXr1gampaqvhev36NvLw8mJubQ0dHp8T7KA2pVKr0/0UdV7h+Wlpa0NYu+l2aqte5pDGmpqZCR0cHJiYm7+za5OTk4Pr160hMTAQAODk5oUGDBqX+2yMi0oSnT5/i1KlTMDY2xpdfflmmfX0odQjFulNh5Y5iWVWScunWrVsIDw+Hk5MTmjVrVuT3Tp48iWfPnsHLywu1atUq07kIPpRrrwrh9/O2342wTbGO5e3trZEFuIS/ISMjI/Tu3fudH784Ql2usPrRm9da1Xu4PH1M9/KH6s32ZUnr5+po+6AE7ZD3/W/wUyaTyZCeno7c3FyYmppCT09P0yFRGbzZli+u7V8cdT0r1C0rKwv79u1DtWrV0K5dO02HU2JqTzYHBQXh888/L3K7RCJBaGgo2rZtq+5DkwYNHjwYjx8/VvrMwMAAtWrVQocOHfDFF1+gQoUKKu/v7Nmz2L9/P6ZMmVIO0b6/jh49Cj8/vyK3z5kzB23atCnV9Xn69CnWr18PExMTfP311yWKKy4uDkeOHMHRo0eRlpYGADAzM0PXrl3x9ddfw8zMTKX9PHz4EElJSfDy8lL52AkJCejfv7/SZ5aWlqhVqxY6deqEtm3bQlf3f4+y8+fPY86cOXB1dcXKlSuL3OeSJUtw7NixIrf7+/ujWrVqKsUXFxeHffv24Z9//kFOTg4AwNzcHD169EC/fv2UGpylOf+3SUxMxIwZMxATE6P0+bBhw/DNN9+o5RhERO9CREQE/vrrL3zxxRdl3tf58+exd+9epTJS3c9fdVCsOy1fvhxubm7ituzsbHz33Xd49OgRAGD9+vVwdnZWab/nzp1DQEAApk6d+tbvbdmyBY8fP8Znn31WpvNQVFj95MWLF7h9+zaaN28OAwMDtR2rPKWnp6N79+4AgI0bNxZYEHX+/Pk4c+YMxo8fj969eyMxMRHr16+Hubl5ietY6iL8DXXs2FH8LDc3F8HBwahXrx6qVKmikbgA4ObNm5g4cSIsLS2xd+9epaRATk4OevbsCYlEgj/++AMuLi4q38PlqbB7+X18jnzM3mxfCm3LFi1aoHv37qhYseJbf74sbR9BSdohkZGRBf4GSbNSU1Nx+PBhBAQEIDk5GQCgq6uLDh06YMCAAahevbqmQ6RSWL16Nfbu3av0mbm5OWrUqIE2bdqgS5cuMDQ0VHl/6nhWlIfLly/jr7/+wqxZszQdSqmoNdkcGxsLHx8fXLhwocjvbNy4EQcOHGCy+SPy8uVLsSLQsmVLscB99eoVbt26hbt37yIuLg4TJkxQeZ9jx47FmDFjyvSG6kOUk5ODli1bKn0WExOD+Ph4GBgYiA3N0lyfJ0+eAAAaNmxYol63wcHBWLx4MdLT01G3bl24uLggMzMTFy9exM6dO3Hjxg0sW7as2DeA2dnZWLhwIWrXrl2iSrpwb5mbm8Pd3R0AkJycjCtXruDKlSt4+PAhvv/+ewBAdHQ0AKBevXpv3ae2tnaB63z9+nVkZmbCyckJNjY2KsX29OlTjBs3DqmpqejWrRvq16+P1NRUHDx4EFu3bsXDhw+xcOFCaGlplfr8iyKTybB06VLExMRg9OjR6NWrF1JTU9G/f3/s3bsXAwcOZO9mIvpgCC/NVE2ovk1ERAQAoEaNGkAZyp/yJNSdjIyMkJmZiZcvXyptP3fuHBITE6Gvr4+cnBzY2dmpvO83z/9txwdQon0Xp7D6ydatW3HmzBm0bt1abccpb0KdSV9fv8DLZ5lMhvDwcEDhGgvX0s3N7Z2NbHpT27Zt0bp1a6Vrf+7cOSxYsAAbNmzQSEyC2NhYAICrq2uBumtiYiIkEgkAiIkfVe7h8vbmvfw+Pkc+ZoW1LzMzM3H79m3cvXsXoaGhWLZs2Vs7M1WrVg2nT58uU6/HkrRD1FmOUdnl5uZi1qxZuHXrFjw8PDB48GBAPqrn2LFjuHbtGtatWwcLCwtNh0olJJQRzZs3R+XKlQEAKSkpCAsLw61bt0qceyptnqS8HT9+HKampvD09NR0KKWilmRzbGwsQkNDERAQ8NbvXb16Ffv27XuvfoFUdoqF8Ny5c5V+v+fOncOcOXMQEBCAIUOGiD1g4+LicPv2bWRkZKBKlSpo1qyZWFmQSCQIDg4GALRp0wa6urqQyWS4f/8+/v33X6SlpcHU1BRubm6wtbVViuXVq1e4evUqXrx4AUtLSzRt2hSWlpZiLHl5eWjbti10dHQglUpx/vx55Ofno27dumJjKzIyEvHx8ahTpw6qV6+O7OxsXL16FU+ePIGxsTFcXFxQs2ZNpeMmJyfj0qVLeP36NZydneHh4YGLFy8iKysLn332mcpDMb788kul4cNRUVH46aefoK+vj/nz58PW1rbQ61PcuUOhAlS7dm2Vf7fR0dGYN28ecnJyMH36dHTs2FGsrMXHx2P8+PGIiopCYGCg+Bbw/PnzyM3NRbNmzRAREYHo6GhUqVIF9+/fR3R0NGrUqIFTp06hVatW0NfXLzYGIe7mzZvjp59+Ej8/fvw4Fi1ahF27dmHAgAGoWLEioqKiAKDA7+dNivuBvEd5cHAwqlevjvnz56v8jAoLC0Nqaio+//xz/Pjjj+Lnn332GZYuXQpHR0dkZGTg7t27uHXrVoHzT0pKQmRkJGxsbGBra4vz58+jevXqaNKkCbKzs3H9+nXEx8dDJpPB1tYWTZo0Ea9ZdHQ0bty4AQcHB/Tp0wfa2tqQyWTIy8uDg4ODUqK5uHuDiEhd3la+p6enIywsDHp6evDy8sL58+chkUjQo0cP3Lt3DwDg4OCABw8e4ObNm9DV1YWnp6dSWV/cszEzM1NshNjZ2eHSpUuFPn/19fWLLd/v3r2LxMRE1KlTBxUqVMClS5eQmZkJFxcXuLq6qnzehRHqTo0bN0ZwcLBSsjk7Oxvbtm1D06ZNcfHiRbi6uoo9dIo7//T0dLEstLW1xaVLlxATEwMzMzN8/vnn4mgb4fjCvrOysnDx4sUCcRobG4uJtfz8fNy9excPHz6ETCaDg4MDGjVqJJY3b9ZPoqOjce/ePRw8eBAODg44e/Ys3NzcYG1tXey1j4+PF8vHWrVqITQ0FM+fP4ednR2aN28ultNSqRRZWVkAoNapK+Li4gAA9evXL/B7fP78OZKSkgAA9vb2gEJdpU6dOnj+/HmR90phdaTu3bsjLi4OCQkJBeJo2LCh2JAuriwPCQlBVlYW3N3dIZPJcOvWLWzZsgWQ122lUmmJ6oDqJHQGcHJyKrBNuNZOTk4wNTVV+R5WvEcU61D169dX+7187dq1Ip8jVD6Kal/GxsZiyJAhCA8Px507d+Dh4YHw8HA8e/YM9erVQ3Z2Nq5cuQIvLy9kZGQgPj4eNWrUgL29PYKCggAALVq0UOr1GBoaioyMDDg5OcHBwUEpjtK0Q4R95Ofn4+LFi8jOzoa1tTWSkpJgZmaGpk2bAvKE+vXr1wEAzZo1E+9toQ2p+BmV3L///otbt27BwMAAvr6+4u+8Y8eOmD59Ouzt7ZGSkgILCwvxeVKpUiU4OjoiJCQEr1+/Rq1ateDh4VGidpWqeYvi6g1BQUHIzs5Gs2bNkJiYWGS96FOjWM8bPny4Ut3h5MmTWLBgAS5fvqz0M6XJkxRXTqAEOaDStMXj4+MREhKCvn37frDPAbUkm5OTk5Geng53d/cih6anpKRg2bJlGDVqFNavX6+Ow9J7QhjiWdibIMU5APPy8iCTybBt2zZs2LABVapUgZ2dHa5evQoXFxf4+fnB2NgYjx49gq+vL+zt7dGhQwfIZDKsWrUK+/btg7W1NSwtLREVFQVdXV34+fmhUaNGgDzx5+vrCwBwcXFBZGQkdHR08Msvv6BWrVrYvHkzYmJixArGlStXMG/ePADAtGnTYGdnB6lUCj8/PyQkJMDf3x9xcXGYM2cOoqOj0bhxYzx79gzx8fHw8fERh0hFRUVhypQpSEtLg5WVFTIyMjB48GBs2LAB2traOHz4cKmua3R0NGbMmIHU1FQsWLBArJi8eX1UOXcAePDgAQAUGAr6Nhs3bkROTg4GDhyIzp07K22ztbWFt7c3NmzYgMuXL+Prr79GdnY2Zs+eDQD4/vvv8eeff8LY2BgGBgbi0KWzZ88iIiIC7du3VymGhw8fAoU0UBo0aCD+OzMzE0ZGRmJPI6Hxp4ozZ85g6dKlqFKlChYuXIiqVauq/LPCfF9PnjzB8+fPxQZhzZo1sWbNGkBeUC1durTQ8w8NDcWqVavQr18/XLt2DQ8fPsSUKVPw6tUr+Pj4ICIiAnXr1kVycjKSkpLQrl07cRjN3bt3AXliW1tbG0+fPsW8efNQpUoVjB8/XoxRlXuDiKisVCnfY2Nj4evri6ZNm+Ly5cs4cuQIOnbsiM6dO4vP72vXrmHbtm3ifq2srPD333/DzMxMpWejkKCwt7eHmZlZkc9fVcr3ffv24cyZM/jmm29w+PBhpKSkiHEJ016oct6FEepObm5uCA4OxosXL8Rtp0+fxuPHj9GtWzdcvHhRHK2jyvkLvXNq1qyJdevW4ciRI+J+z507hyVLligdX9h3dHS0WFYo6tWrF7y8vJCeno5ffvkFZ8+ehbOzM/T09BAREQFvb29MnDhR3Kdi/eTQoUMIDAwE5AkiX19fbNmyRaVrf+XKFfz+++/o3r07Hj58KL6MAIBRo0bhq6++AuTTM0yePBlOTk7466+/VLpXVSE0POvWrVtgm3CP2djYoFKlSoBCXSUlJQXfffddofdKYXUkExMT/Pe//4W/v7+Y3FQkJIuLK8tzc3OxYMECZGZmYufOnbh16xYWLVok7mfZsmX4+eefNZZsvnPnDlBEHVTo9SzU61S9h4uqQxkaGqr1Xm7Xrl2RzxEqP0W1Lx0cHGBra4v4+Hikp6cDAHbt2oXg4GB899132LZtGzIzM9G4cWPs3bsXp0+fxty5c1GrVi2sXLkSqamp+Pvvv8UEVUxMDKZPnw5zc3Ns3LixQByqtkMMDQ2V2iFSqRTr1q3Drl274O3tjT59+mDChAnw8vIS23QBAQHYunUrIG9zmZiYICEhAT4+PnB3d+c9VkZChyyJRIIHDx6gYcOGAABDQ0P8/vvvSt+9fPkyli9fjtatWyM2NhaZmZniS8UJEyaIncGKexarkrdQpd6Qm5uL+fPnIycnB0OHDsXff/8txqpYL/oUKY48ejPpLrx8rl+/vvhZafIkqpQTquaAStsWF2aLaNOmjdqu3bumljkKPDw8MGLECHh7exe6XSaTYfny5ejXr1+hb7Tpw/bvv/8ChQwZSk1NxYEDBwB5zx1LS0ucP38eGzZsgLu7OzZs2IBly5ZhwoQJiIiIwJkzZwCFHg5CT5CEhATs27cPjRo1gr+/P9auXYvly5fDw8NDrKDGx8djzpw50NLSwh9//IGlS5di6dKlSElJwc6dOwH5m3HIe8HIZDLs27dP7OEgVFZu3ryJ6Oho9O7dGyYmJli8eDGio6MxZ84c/Prrr1i7di1sbGywYsUK5ObmQiqV4tdff0VaWhomTZqE3bt3Y9euXTh+/Djy8vLg6upaormqBXFxcZgxYwaSk5Mxa9YstGjRQmmb4vVR5dxlMhlu374NlCAR++zZM7EHQKdOnQr9jpBcTU1NBeRDIQWRkZFYuXIlVqxYgW3btsHIyAiQz4fs7++v0hQPhQ1VFQjnY25uDktLSyQkJCAzMxNQGIZZnODgYMybNw/W1tZYtGhRieftatasGapUqYKYmBgMHDgQvr6+OHnypFIPNW1t7SLPX6hInzp1Ct7e3li/fj1atmyJe/fuwdTUFN9++y3WrFmDP/74A5AnxoV7Vfi7c3R0RFZWFqZOnYqkpCSsWLFCLLhUuTeIiNRBlfI9Pj4eAHDv3j2YmZlhzZo1GDJkiNIw+tTUVOzevRv+/v6wsrJCcnKy+KxU5dmoWEYW9fzNzc0ttnyXyWTiS71bt27hl19+wcGDB8UpmITkpyrnXRjhGe7k5ARbW1s8f/4ckDeUtm/fjm7duonlpJAUKcn5R0dHw8nJCQcPHhSTc5cuXRIbYsLxhX07OTlh//792L9/P3x8fAAApqam6NatGwBg06ZNOHv2LPr164c1a9Zg1apVaN++PQ4cOCA20t6sn/zwww/igjZz587FyZMnUaVKlWKvPRQSkCEhIRg6dCgCAwMxbNgwQF5XEwjHdHFxUfFOVY3w+31bcrRhw4bQ0tJSqqvcu3evyHulsDrS8uXLoaOjgylTpmD//v3YuXOnOFS/R48eYlKtuLI8MTERmZmZMDc3R+XKldGhQwdxCHG/fv1w8uRJjSWuXr9+LfZsLqyeJfQGFeouqt7DRdWh1H0vv60eR+WnqPZldHS0WJZUq1YNMpkMkZGRgPx5PHPmTKxbtw729vZKbR9tbW3UqVMHkL+4E+zYsQMA8O233xaYTqEk7RDhb9DKygqWlpZYtWoVdu3ahW+//Rbjx48X55fOy8sD5GVdQECA2BYV2jCnT58GAPTp04f3WBk5OjqK13f8+PHic1b4+1YkPNefPn2KJUuWYM+ePRg5ciQA4NChQ4CK7SpV8haq1BsSExPFtYCSkpKKrBd9ioRzNzU1xf79+7Fjxw5s3boVixcvxqpVq+Dp6Ynhw4cDZciTFFdOqJoDKm1bPDc3F4GBgXB2di70pfeHQu0LBBbm+PHjSE9PR69evcRfJH08hMbY9u3bxYnapVKp2PBxcXHB5MmToaWlhd27dwPyubdSUlKQkpIivpUT3lIJvUmEFxPCCsBxcXEIDg6Gq6sr3NzclBbSOXbsGCQSCVq1agUdHR3Ex8eLQ9uEiqNQyEulUty5cwdXrlyBr68vwsPDxViPHj0KAOjatSuuX7+OiIgI2NjYwMnJSazYVK1aFTdv3kRycjKeP3+OqKgo2NjYoGvXrtDR0YGZmRk8PT0RGxtb7NzBhUlISMD06dPx9OlTTJs2rcDKo29eH1XOPSkpCUlJSYW+ASyK4vC1opKwQsVI6A2s2Ntn8uTJ4tCRuLg4ZGZmwtraGlWrVlW58pSUlIRnz54B8sb01atXIZVK8eTJE4SEhAAARowYAX19/QLDMItz+fJlzJkzB+bm5liwYEGJenwLrKyssHz5cuzbtw/Hjx/HqVOncOrUKRgYGGD8+PHiQldJSUmFnr/wtzNw4EBxISIA8PLyEnvfvHjxQkzm6+vri79b4Wdr1KiBpKQkPHr0CDVq1FA6d1XuDSIidShJ+d6kSROMGDFCfBYKLzZr1qyJcePGic8pU1NTJCcni0keVZ6Nb5aRhT1/r1y5Umz5rqOjI5Y/U6ZMERsgQhkq9FhW5bwLIzzDq1evDjs7OzERefr0acTHx8PX1xf+/v6AQuNHlfMXknre3t5iTywhSWJiYlKgDBH2XaFCBVSoUAEPHjzAqlWroKurizlz5qB27dp4+fKlWL/z9PTE06dPAXlZD3mj2NnZucC119bWFhM1jo6O0NPTQ1hYWLHXvkqVKmJP2DFjxohzFQrnobgoWMeOHdG6detSvdgvSnZ2doHro0jo6SgkwRSn1XjbvVJUHQnyupZUKsWKFStw8+ZNtGrVCuPGjYOOjo5KZblQB3JzcxOnOxMSG05OTuJILE1Q7IX25ugxqVQq3iPC1AMlvYffrENBfj+r814uqh5H5Uf4/d69exdr166FTCbDy5cvxd5+Xbp0gZOTE54/fy72Op8wYYL44kn4u1Rs+9SsWROXL18Wvx8VFYUTJ06gZs2a6NKlS4EYStMOqV27NlasWIHAwEBMnDhR7IwnLI4qvFA7ceIEJBIJevbsKU7jkZWVhUOHDsHBwUGtC7d+qnR0dDBr1ixxlM3Vq1dx9epVQJ7M/+677wo8T4YOHSoupircS8I9oMqzWJW8hSr1BqG8KK5e9CkSns/JyclYu3at0jZnZ2f07dtX7BBXmjyJKuVEVlaWSjmg0rbFb9++jfj4eEycOPGDXsOs3JPNT548wfr16/HHH3980BeKCpeWliZWuuvVqwc9PT3k5ubixo0bAABfX194enpCX18fL1++FOfXWbVqFVatWqW0L2EuGuHttFCxtLe3RzLkUTcAACAASURBVP/+/eHv74+5c+cC8obqkCFDxCFM165dA+QF94kTJ5T2K7ylFir1UqkUBw4cgIODAzw9PWFmZob09HTEx8fj9OnT6NKlC6pXry7u59mzZxg4cGCBczcyMhJ7qzRp0kSpIi8kYYubO/hNz58/h4+PD+Lj4/HDDz8UWvF58/qocu6KczOqOr/c69evAQCVKlUqslItvDwShmUKDZv27dsrNaKECpjQC0hViitQKw6jhHx4TN++fdGqVStAoeBRHDZTlJs3b2LWrFkwMDAQF3spLRsbG4wePRrDhg1DREQETp8+jcDAQCxduhS1atVCnTp1Cj3/1NRUsUGl2HM9Ly8Pe/bswdGjR5XOH/JGpL6+vtLfXfXq1WFoaAhPT09cvnwZP//8M+bNmwdjY2OV7g0iorJStXwXysz27dsrlQVC2aE4B2paWpr4ebVq1VR6NqKQMrKw56/QqHxb+S40AlxdXZUSjsLn9vb2Kp/3m4RneJUqVWBlZYWqVavi7t27yMzMxPbt29G9e3c4ODiIZWz16tVVPn/hGjdv3lzcLiR0GzRoAB0dnQJliOL3Zs6cidTUVMyePRseHh4AgPv374vfmTx5coHzEZKpb157oQFnZGQkJhlVufaK8QkxQOF3qTjstDzmMXz27JnYo+zN5KhiT0chOapYxyrqXsFb6kiCzZs34+DBg3B3dxfX64CK9bzC5pt8W8L8XRLuv3r16hWogz59+lR8aSLci6rcw0XVoRS/r857ubT1WCodxWdAWFgYwsLCxG22trbo2bMnvL29oaWlJf79OTk5KXXyKaztI6zPI0xbJLzQGzZsWKEvrErTDgkNDRW/Y21tLf5bX18fZmZmyM3NRVZWFnbv3o3u3buL91hmZibCwsKQlJSEb7/9VqMviD4mhoaG6NevH/r06YPo6GiEhYWJHeTs7e3Rs2dPpftNcXoUYRSFsEisKs/i4vIWqtYbhPvpbfWiT5XwfJ40aZI4gig3NxexsbFYtmwZpkyZgmXLlqFJkyalypOoUk6omgMqbVtcGOFQWPn2ISn3ZPOxY8dQsWJF8WGelJSE7Oxs/Pbbb/j888/RpEmT8g6BypHwx2lkZAQ/Pz9xMb8JEybg9u3bePbsmfiAFJKXALB69eoCLx8qVaqk1JtEsQH03XffoUOHDrh+/TrCwsJw9epV3Lt3D1u3bhUn9Yd87uU3hzkJiwEID++HDx/i9OnT+PHHH6Gvrw9LS0ukp6eLw1Z69eoFyOfdA4Cvv/5arEgoMjU1FYe9Kj7wc3NzxQdLSSr4ycnJmDlzJmJiYvD999+LcSgq7Pqocu7CcJOSDMMQeg7FxcUhOzu7QCUsNjZW7I0mPOiFh/ObPbpLuzqzEHfr1q2VVpStUKFCgXkwhUZdcfMQ3717Fz///DMgfxlSlqG3KSkpMDQ0FHuENWrUCO7u7khOTkZoaCgePHiAOnXqFHr+wu/N0dFRfPsK+Ryha9euRbt27bBo0SJYWVnhwIEDWLt2rfj7E/7unJ2dxft67ty5WLRoES5cuIDAwED069dPpXuDiKisVC3fhd6qb5YFQtmh+IJWeM65uLjAyMgIu3btKvbZWFgZWdjzV5XyXSh/FMsIxXOws7NT6bwLo9iw0dLSgrW1NdLT03H48GEkJiaiT58+ePHiBZ49ewZnZ2eYmpqqfP5CfIr1D6FxKiQiFcsQIen58uVLzJ49G0+fPsWECRPQtm1b8eeFZGCTJk3EqSwUCYspv3ntFc9TaIypcu2FhqSDg4M4BRoUplt4cwEvdRNGu6GQZPbjx4/FslXo9arKvYK31JEgn7t1y5YtcHJywqxZs5SOq0pZLtSBhGtT1AsFTVDsvPAmIeni4uICc3Nzle/houpQKKd7ubT1WCodxfbl5s2bxTmb9fT0YGJiopTwF/7+mjRpUujnim0foYfzixcvEBERgfPnz6NZs2ZKLzYUlaYd8uWXX8LJyQl+fn44cuSIUrLI0tISOTk5uHDhApKSktCjRw+xh2pGRgZOnz4NKysrtG7dutTXjv5fRkYGpFIpKlasCB0dHTg7O8PZ2Rn6+vr4888/cefOHfTs2VO832rUqKH0e31zPnlV21Vvy1uoWm/4P/beOy7KK3v8f1MFHMSIBBBBQOwtrliwoFHzicaN0eiKaWpiNmuSNYlZTYxpKpqYGFtWY1kLil3UKJplgyVYUJSIiCKooCI1CIKMlGGG+f3x5d7fDAwy2I3P+/XKa9eZh6fcuc855557ijl20ZOIoXxu3bq1UZM9V1dX+vXrx7Zt24iLi8Pf3/+O/CTm6AmxqVSbD+hO1uL5+fn8+uuvDBw40GjD6nHkvocad+vWjZdeegkfHx98fHxwc3PD0tISHx8fozQ4hccTw+YNogi/hYWFTBnaunWr3BU0/L2bNWtGmzZt8PX15dKlSzg4OODi4kJGRgYajUY2XdFoNKSkpJCcnIyvry8jR47khx9+oFevXjKNFAPjwc3NjTZt2tCmTRvS09PR6/XSUBTKY/fu3Tg7O8ti605OTuTn57N79266desmFwFit6lBgwbynCUlJRQVFeHt7W1k0BgqjsjISLKysrC2tpb3lZSURHR0tMkaUVQuvr766iuSk5MZP348o0aNMnlc1fEx99nFgkOUisjMzCQ6OlpG55iiVatWODg4oNFo+Pnnn9Hr9fK7zMxMvv/+e7RaLUFBQTRv3hydTiejsKoKU/HcYsFl7j0YRs03atRI/lfVwDO89u0c/MnJyUybNo3S0lJmzpwp6yJWxZx7++CDDxg2bJgsvSIoKyuTteAMy4hUfX5hwBh2qccgIuL555+nadOmWFlZ8dtvv4GBI128d4a77/b29rKJo0j3Mmdu1DY3FRQUFGrDHP2enp6OVqvFxcXFyDlUk/wWMlI48MyRjaZ0pCn5a45+F/rHcKGXnp6ORqPBw8ODRo0amfXcphAyXDiunJ2dAVi/fj1Dhw7F29tb3rfI1jHn+cUYOzs7G42xcIKIZxHXF+dWq9XMnDmTS5cuMWbMGFm6QCAcvpaWlnK8VCoVWVlZPP3006hUKpNjL0qDGI6hOWMv7s9QPxrOE6G/1Go1aWlpMsX1XmEYdSzmAYBGo5E1FgcOHCgXuVXHFxNz5XY20sGDB1m0aBFubm7MmDGjWof62nS5qbERY+Lr6/vQnRJi/ly4cIGioiL5+Y0bN9iyZQtUlkSgDnO4Jhvqfs1lU3JE4f5hKANEk7VGjRrh6OhYLbK8pmCTqmsfDDIVbty4wc6dO6GyVrOFhQVqtZro6Giio6NlKYQ7WYcMHz5cOomio6PlRgWV86+8vJydO3fSt29fmjdvLh1OJ06c4NSpU4wcObLGxrIK5hMSEsKQIUNYtGgRFRUV8nO9Xi/LHonSFeL9VqvVssxJeXk5Bw8ehEqnI2bIYnP8FubYDebaRU8iQj5jQh6r1WrpiBay/k78JOboCUFtPiBzrl+V6OhotFotAwcOvEej9vC475HNHTt2lJ0/qUxf37t3b43NBBUeL0TzhqplCHr27ImLiws5OTlERUUxaNAgnnrqKQIDAzl06BAhISH06NGDffv2ERERwejRo5kwYUK1NDWtVsvUqVMpKSnhn//8J02aNCEnJ4dz587h4uIid5MGDBhAYmKibPxz8eJFVqxYQbt27ViwYAFWVlZSMBw5coR//OMfUpGrVCoiIiIAmDJlinyGHj16sG7dOnbt2oWHhwcajYbFixej0WhYuXIl9vb20jG9c+dObG1tUavVMkqjXbt20oBYvnw5cXFxsrN1VebPny+F49mzZ5k2bZrR97a2tkyfPt1kGp85z161ucX+/ftZtWoVr776ajVDXeDo6MjEiRP57rvvWLp0KdHR0TRv3py8vDyOHj2KVqtl0KBBjBs3DiodnIWFhUbpsgKx4XD8+HHs7Ozw9/c36x4Maz3ejpycHBkttXHjRtmYUmBpacmHH37ItGnTKCwsxNnZmZ07d0pDU9C+fXtee+01s+5twIABnDlzhkWLFnH+/Hk8PT1lCtzly5fx8/OTaZumnl+kf1ZtmiqMkN27d1NaWsqhQ4fkQjEnJ4cbN24YNXaKjY2lqKiIp556iv/+979gEDVlztyobW4qKCgo1IY5+l1E5HTq1MnIWSB0h52dnVGESNUGdubIRlM60pT8NUe/m2oKVdUJas5zm0I8m4hCFc7moqIiRo4caXQt4UAx5/kNx9gwWko8i1jYiOuLcy9btozTp0/L861du1Zec/jw4bRu3Ro3NzdOnDjBzz//jJubG6GhoZw7d47g4GD69OljcuxF08ezZ88SFRVFQECAWWNf9f4wKLfg5OQk6yYePHiQefPmMWjQIKZOnWpyrO+Epk2b0rdvX6Kiopg0aRKBgYHY2dlx5swZUlNTcXFx4c0335THG/ZQEFSdKzXZSJcvX2b27Nnyefft2ye/CwwMxMfHp1Zdfv36dQoKClCpVPIdKisrk+N24MAB6bR7GPj7++Pu7s61a9d455136N69OxqNhuPHj5Ofn4+/v7/cLDd3DtdkQ92vuWxKjijcP4QMMCcj09T7h8GcMfzcxcUFOzs74uPjyc/PZ9iwYbJpoAhI8fX1lZHOdV2HiPfbxsaGQYMGERoayv79+2WzsgYNGshSk0I/1KtXDzs7O6KiorCzs6uxKbtC3ejRowcbN25k//79FBYW0r59e3Q6HefPnyc2NhYHBwcpd8R8s7Gx4aeffqJLly4cO3aM5ORk2rVrJ/sG1CaLzfFbODg41Go3iPlUm130JCLks52dHUuXLgWgoqKCoqIi4uPjKSwsxN3dXWa03ImfxBw9Ya4PyJzrG6LX6/nll19wcXHhL3/5ywMb1/vFA2kQqPDnxTDV0RA7Oztefvllli9fzubNm+nfvz+2trZ89NFHqFQqfv75Z8LCwlCpVLz++uu88cYbYKKxj4ODA9988w3Lly9nzpw58vy9e/dm3Lhx0mH84osvUlJSwubNm6XDuH///rz77ruyjIcwPO3s7KRywWAB16pVK6PagG3btmXGjBmsWLGCr776CioXDRMmTJCCv2/fvly5coXdu3cTHh7Oyy+/zKBBg0hOTpa7jmVlZdXq+xkiDG5BTExMtWN69+6NhYVFtfEx59nz8/Ol8V41lbOqkV6VwYMH4+zszLZt2zh58iTx8fFQmbEwZMgQ+vTpIxcD4hodO3asVmfshRdeID4+nj179kgjvbZ7MLzv2pyghjXVTpw4Ue37rl27kpmZKRuCiDIXVala0+924zN06FCsra3ZuXOnUQ0mNzc33njjDYYPHy43OEw9f03vztixY0lLS+Po0aNcuHCBjz/+GAsLCy5evMjSpUvx8/Mz+ttDhw6xbds2qNyUePPNN2VTydrmRm1zU0FBQcEcLCwszNbvVdPQDXWHYT1XkUovjH9zZKMpHWlK/tam303pTUzYKOY8tymEDBe6TUT7Dhs2TH5W1SFdl+c3HOPr16+TkZFhFGljqEN0Oh1HjhyBylqD69atk3/74osvQmX0V3BwMEuWLGHhwoVQGa0zdepU+vTpY3JsALp37054eDgJCQnk5+cTGBholm1lSj+KBWaHDh3k4qwmh+PdYmFhwZQpU2jVqhXbt2+XAQkqlYqhQ4fyyiuvSIexuXOlJhspKSkJrVYLwNGjRzl69Kj8TqTS16bLDcuViCxDPz8/AgICOHbsGN988w0hISH3dIzqgpOTE/Pnz2fz5s1ERkbKgAA3NzfGjx/Pyy+/LEu13ckcFtzPuWxKjijcP8TvW5tTzfD9M4wSrOm9tLKyok2bNsTFxeHg4MCrr74qvxNRoyLj407WIYbvd9++fQkNDSU8PJzRo0ejUqnkmrNjx45GDeOefvpp0tLSeOmll+Tmo8Ld0bp1a77//ns2bdpETEyMbAyoUql44YUXGDFihNzQFE7CiRMn8uuvv8pyi3379uX999+XtkltstjW1tYsv0VtdoO5dtGTiJDPpaWlhIeHy8+dnJzw9vbG39+fwYMHy6yUO/GTmKMnzPEBmXP9qiQnJ5OYmMi4cePM7rP1KGORm5urf9jpVXdKcXExM2fOZMSIEfftGhUVFbJ8gKWl5RPfFCIiIsJk07q6otFoUKvVRl2la6O4uJjS0tLb/o1Op6OwsBB7e/t7VpNWr9dz8+ZNLC0tTTZ0KS4uxs7ODgsLCywsLPj666+JiopixowZcqFw7tw5fvzxR5YtW3bf5lBdnr24uJigoCDWrFlTY03JqpSUlFBSUoKjo+MdNa3QaDRoNBrpgL2Te3hQ1PXeSkpKuHXrFra2tiZT/DDx/LdD/JaGY63VatHpdCYbmNT229xubjyIuanw56Jfv36sWbOm1kgfhSeTO9Hv5lJX2Wh4T6bkb236vS7cz+cW3Onz30tu3bqFVqtFpVJVi8gxhU6no7i4uFqd1Xs59lXZuHGjkRPpbikqKqK8vJwGDRpIZ+7DoK42rl6v59atW9jZ2d3VfT/33HN88sknPPfcc3d8DkFFRQWFhYVYWlrSoEGDh2p31HUu18WOq4m9e/cyf/582fzpfiHs0vr169/1emj79u2UlJQwaNCg+2avh4aGEhYWxq5du+7L+RXg1KlTfPzxx+zZs+e+NFZ9lBH62cLCQtZvFpSUlDB48GConOvOzs4UFhZiY2NTYwkic2SxOX6LB2E3/BkYMGDAXcvMO/UR3U5PmOMDquv1s7Ozyc7OplmzZrdtIPioEx0dzeLFi5XI5toYO3asUdSkm5sb7dq1o3///vTs2fOOjKRLly6Rm5tLQEDAPb7bxwfRmK8uODg41Fp3zsrKqs7nrQ0LCwtZ08mQ69evM378eAoLC/nuu+/w9vbm5MmTREVF4eHhIaMeNBoN69atIygo6L4a1XV59q1btzJw4MA6GY1368AXO753cw8PirremzljU/X5b4ep39La2rrGxWJt169pbjyouamgoPDkcCf63VzqKhsN78mU/K1Jv98J9/O5BXf6/PeSutYStbKyMulMvpdjf7+5187wO6WuNq6FhcUj51SytLR8ZBbQdZ3LdbHjHgY6nY7IyEjCwsKMao137dqV0aNHG2VvAsTGxqJSqW5bqkKtVvPvf/8bKjMvFBQeR26nn9PT06Gynrc4pjbdZI4sNsdv8SDsBoX/x536iEzpCXN9QHdyfTc3N9zc3Op8n48qirP5NhiG1QcGBmJpacnVq1fZv38/+/fvZ9iwYUycONGs3XBBWVkZ33zzDS1btnyinc1/Bho3bsyYMWNYvnw5n376qfy8S5cuvPvuu1I4ZWZm8txzzxl1w36YFBYW4urqyssvv/xE30NNPMr3dq951OamgoKCgoKCgoJC3VmzZg3r16/Hy8uL9957D5VKRVxcHJGRkZw8eZJFixbJ0g1paWlMnjyZKVOm3NbZLNbBLVq0eOQ2LhQU7gWGJZqUwBsFczDXB6SgOJtvi1CwDRs25Ouvv5ZO5bi4OL788kt+/vlnunbtSq9evaAyNezcuXOkpKRQUlJC48aN6datm9wdi4mJIT4+ntTUVJo1a8a+ffsIDAwkMTGR69ev07JlS1kTqrCwkJMnT2JtbU2/fv0AOHz4MGVlZXTt2pWLFy9y8eJFXF1d6d27NzY2NsTGxpKSkoK7uzu9e/eukxNc4c4YMWIEAwcOJDs7G51OR6NGjXB1dTVSVt7e3o9UPVwnJyeZLvQk30NNPMr3dq951OamgoKCgoKCgoJC3SgvL2fr1q0ATJ06VdYMHTx4MCqVisLCQkpKStBoNBw6dIhDhw4BkJWVRUxMDN27d0ev1xMfH8/FixepV68eAQEBstGlYQ1SBYU/E61bt2bhwoVKhLFCnTDHB6SgOJtvi1CwHTt2NHLcdu7cmTfeeIOlS5dy/PhxevXqhU6nY/78+ezduxdPT09sbGxITU3Fy8uLxYsXo1Kp+P7772WDsoMHD5KYmMiAAQNYvXo1Z86ckWlKVNZRnTVrFgEBAfTr14/y8nKCg4PRaDQEBQWxZcsWeew//vEPysrKjJp/fPDBB09EZOajgJOT02OTCqqgoKCgoKCgoKCg8OfB0tISW1tbNBoNiYmJtG7dWvYa+vDDD+VxV65cYdasWfLf69evZ9iwYXTt2pWffvqJsLAwrK2tcXFxYcuWLTzzzDMAspGagsKfDQ8PD9l4VEGhLig+oNqxLigo4NatWw/7Pu6IoqIiKioq7tv5RSfwli1bVvtOdCjOzs6Gyp3hmzdvEhgYyBdffIGNjQ0TJkwgOTmZlJQUOnfuzPr16xkxYgTFxcVs3LgRV1dXSktLZQdUw063oiNuq1at5Pk1Gg1UTuzdu3ezY8cOQkJC2LBhA+PGjWP37t1s3LiRzZs3k5CQoDibFRQUFBTuOZaWlnz77be11qJTUFB4NLGwsOCTTz7B3d39rs+Vk5ODWq2+J86omzdvcvXqVTp06HBX5ykvL+fgwYNKiah7hJOT00NtjPhnwtLSUjaev5dYWVnx2muvsXz5chYvXswvv/xCYGAgXbt2pU2bNlhaWgLQrFkzfvrpJ9577z18fX1Zvnw5lpaWREdHExYWhqenJ3PnzsXNzY09e/bwww8/QJU16v3A0tKSZs2a3ddrPOk4ODhgZ2en9Gm5D+j1evR6vXzPFO4t9erVo7i4WFl3PEZYWVnRsmVLrOvXr4+dnd3Dvp87wtbW9r6+1OfOnYPKVPOqaLVaABo0aABA06ZNCQ4OprS0FLVajU6nk8eKxl25ubkUFxfj4uKCu7s7FhYWpKeno9Vq8fHxMdoZuXDhAgA+Pj5gUNKjVatWBAUFYWVlJRW/r68vI0aMwMLCQirqR6WZiYKCgoLCn4uKigratGnzyDR5UlBQqBtRUVH3xNEMEBkZyeuvv35PzpWUlES3bt3u+jznzp3jzJkzirP5HlG/fn1lLO8RzZo1u2+BUqNHj8bV1ZWdO3eSkJBAamoqISEhdOvWjSlTpuDi4oKFhQUZGRkAtG/fHhsbGwB27doFwGuvvSabU3Xs2FGeu2nTpvflngUVFRUyo1jh/lBcXExpaSmvvvrqI93o8nHk119/5f/+7/8e9m38aVm3bp3iaH7M0Ol0XLhwAWsbGxvq1av3sO/njjB06N5rioqKZCdfU7u5Fy9eNPouKiqKsLAwEhISqh0rFLQoQN+xY0e5o2hYlF5QUVHBmTNnjM5/+fJlAAICAmRJD6GUAwICqp1POKkVFBQUFBTuNcOGDVP0jILCY4qwKRUUFP48WFhY0L9/f/r3709mZianTp1i06ZNnDhxgmXLlvHll1+CwfsvsnRFnyAqHdCC4uJiANzc3HB2dn4IT6RwP3j11VeVZo/3mISEBMaNG/ewb+NPS2ho6MO+BYU7RMmJqgERSWxnZ1ct+iMvL0/uAPfq1YuEhAS+/vprfHx8WLZsGU2bNiU5OZl//etftG3bVgp0odxbtGghz1VV4QOkp6eTn5+Pra2trCFUNdLZ8DNfX1/52fnz56Fy51xBQUFBQUFBQUFBQUHhz4tGo0GtVssmZ02aNKFJkyZ4eXnxwQcfcPToUSoqKrC0tKy2VhT9hAAjp7JYC3fo0EEpu6CgoKCgUGeUwjI1YNgcUKQYURnRPH36dHJzcwkKCsLPz4/Tp08DEBgYSOvWrVGpVBw7dgyqdO8VUceGqUjiM5GOXFFRwY4dO+S1bW1t0el0MtJZGAY6nU5GUXt6egJQVlbG2bNnjT5TUFBQUFBQUFBQUFBQ+PMRExPDCy+8wPvvv09RUZHRd9evXwfAxcUFS0tLdDqdLBMpgqkMy3qI/kA6nY7//ve/YBAQpVariY6OJjo6+r5mFysoKCgo/DlQIptrQDQHzMjIYPr06eh0OjIyMkhNTQVgxIgRvPXWW1BZywzgwIEDNGvWjNTUVNngr6ioiIyMDDw8PCgpKQHg+PHj2NnZ4e/vL6OeN23aRE5ODqdPnyY9PR0MmgPm5ORQUFCAnZ0dTZo0gcqGgWq1moYNG+Lq6gqVEdFarRY3NzcaN278QMdLQUFBQUFBQUFBQUFB4cHRtm1bmjRpQlpaGhMnTqRPnz5YW1uTnp7OwYMHobJ0ApU9h4RDef/+/fj7++Pu7k7Dhg0pKCggNDSULl26cPjwYVkyUgRJJScnM23aNHx9fenZs+dDe14FBQUFhccDJbK5BkSKUUZGBr/99hsxMTHY2NgwfPhwfvzxRyZOnChrXT/33HMEBARw7do1vv/+exwcHJg0aRIeHh7873//k4r+hRdewNHRkT179nD8+HEAhg8fjqenJ4mJiaxfvx5/f3969+4NBtHJhmlMoqC/+KxTp06yhrNh/Wcl3UlBQUFBQUFBQUFBQeHPi6OjI99++y1//etfyczMJDQ0lDVr1nDw4EG6dOnCt99+y+DBgwGoV6+erC27fPlyrl27hkql4tNPP8XJyYmwsDB+/PFH+vXrR4MGDcCgtIYIpGrXrt1De1YFBQUFhccHJbK5BpYsWWL2sULJFxYWYmtri729PVQWMy8rK5P/7tmzJ9u3b0ej0ciI5hYtWhASEkJhYSEqlUo6k99++215/u7du/Pbb78ZXTMgIKDaZ88+++xD6xZdUVGBXq/H0tKymqNbr9fLOmFPshP85s2b7N69G0tLS1555ZV7PhZinC0sLLC0rL6PJFLeavr+fnD06FEuX77MX/7yF6OSMgoKCn8ehOy5HULuVFRUUFxcjK2tLba2tjLjR+jJmkhISCA+Pr7OEVXZ2dns27cPBwcHXn75ZbP/7l5dX7B7925u3brFqFGj5AaxwNT43Y2+1Ol0bN68Gb1ez0svvYSjo+Mdnedx4m5/HwWF22H4jlZ9fwESExM5deoUnp6e9O3b97bnunz5MgkJCZSWltKrVy/Zm+V2xMfHk5CQgJ+fHz169LiLJ3m4iLXC3dqhkZGR5OTkEBAQQPPmze/pWOtnHgAAIABJREFUPdZEaWkpJ06cIDs7m0aNGjFw4MAHcl1z8fDwYPLkyUyaNAm1Wo1Wq8XR0VGuKw0ZN24cQUFBWFpaysCpgIAAtm3bxs2bN2nYsCHW1tZs2LDB6O9GjBjBiBEjHtgzPUzEmsnU+y7msfjeUD7cbh1c0/dPErcbV0HV8a2N8vJybt68WeN8BygpKaGkpISGDRvWSfbo9Xpu3ryJpaXlI21L6XQ6CgsLsbGxQaVSmZxj4ph69erJrHxzMGfsioqKqKiowMnJ6Y7u297evtZ1QE3o9XrUajXl5eU4Ojoalb99lHgSbXPuh7P58OHD9OnTx+izwsJCbty4YfSZjY2NWQbW40TVF8zS0rLaiyMW2IZYWVnJhg6PK2PHjpWR3d26dTP67rvvviMiIoKpU6cyaNCgh3aPD5vU1FRWrlyJv7+/TGe7l0RFRTF9+nQ6dOjAv//9b/m5RqPhP//5D9u2bcPb25svv/zygRnnu3bt4sSJEyxYsOCBXE9BQeHBs3btWkJCQm57TK9evZg9ezZJSUm89957zJkzhx49ejBx4kTc3d0JDg6Wx5aXl3PkyBHatGmDm5sbAAcPHmTHjh1Mnjy5Tvd2/vx5Vq5cyXPPPXeHT0eN179+/TpnzpyhZ8+e2NnZ1fi32dnZzJ8/nxEjRphcOC1ZsoSwsDCjz6ytrfHw8KB37968/PLLRk2baiM7O5v//Oc/qFQqXnnlFbP/7nEmKiqKsLCwOs8PBYXaKCsr491335Vl9Hbs2FHNZj927BihoaG8++67tz1XXFwckyZNwtraGm9v72r2ck389ttv7Ny5k08//fSOnyM2NhaVSkXr1q3v+Bx3w9WrVxk/fjxarZa2bdvy008/3fG51q1bx7Vr1+jevfs9vcea0Ol0TJ8+nePHj+Pu7k6fPn0YOHCg2TrgQWJlZWWWw8eUY8fW1lYpwwhkZmbKddrPP/9Mw4YN5Xf/+9//mDt3Lra2tnz66af07dvXSIdPmDCB0aNHG51v/vz5hIeHA/Cvf/2LF1988YE+z6NARkYGa9eu5ddff8XFxYWtW7eadIhevHiRiRMnUlpairu7Oxs3bqzROa/T6dixYwebNm0iPz8fW1tbxowZw2uvvSb/pqioiNWrVxMeHo5Wq8XFxYVJkybVuimt0WiIiIggNDSU3NxcAHr37s2nn376SDkJxRhs3bpV3mfr1q15//336dChA1Q2AN28eTO7du2SJXQCAgL45z//eVtfnDljd+XKFZYuXUpMTAxUlvT55JNP8Pb2vu19FxYWynsqLi7G2tqa0aNHM378eLM3YwoLCwkPD2fnzp2yyam1tTUDBw7ktddee+j9y6rq3CfRNudeO5uvXLnC559/zqFDh4w+/+9//1vNqGjTpg3Lly+/l5dXeEjk5+fLsh6mXmzRyFA0N3xSuXz5MhjU4r7XiIWQYXpbYWEhc+fO5ciRI/To0YNPPvnkgW1saDQa2djSsCmmgoLCnwsrKysGDBgAwI0bNzh16hTW1tZGEX7+/v5QuQk7ZswY/Pz80Gg09OzZs5pM/O2335g9ezarVq2Sn/3zn//k/fffr3M0nJC7LVq0uKtnNHX90NBQDhw4UGsko7CJ+vXrZ/L7xMREqCyLJZoFl5aWcubMGTZs2EBsbCyLFi0y25nRpEkT9u/f/0CzWB42YgyfdDvjbliwYIFcsBnyzjvv4OXlVadzhYeHExMTg729PZ9//rn8/MqVK6xcuRIqoytF4zFzWbFiBWlpaXTr1o2hQ4fW6W/vlIiICFJTU3F0dKSoqIjMzMxqdlRSUhIAPj4+tz3X7t27AZgyZQrPP/+82fdwt/M7LS2NyZMnM2XKlIfmbF63bh1arRaVSkViYiKlpaV35KA1XHM8KNvy4sWLHD9+HE9PT1atWiWDhszVAQqPF2J++fj4SEezTqcjJCSE0NBQPD09+fLLL2nZsiVUvp/W1tZYW1vLhoyCK1euEB4ejoODA8XFxbU64f5saDQawsLCWLduHaWlpQB07NjRpEOxsLCQ2bNn4+DgQGlpaa1lQdeuXcu6devo3r077du3Z/fu3axcuZLWrVvj7++PRqNhxowZxMbG8uKLL6JSqdi0aRPffPMN69atu+16eOnSpezcuZOOHTsyfPhwjh07xpEjR2jfvn21zYSHyZYtW1ixYgWenp6MHTuW1NRUDh8+zPTp09m4cSO2trbMmjWLuLg4nnvuOfr06UNsbCy7d+/m5s2bNWbymzN22dnZfPzxx5SXl/PGG2+QmZnJ/v37+fHHH5k3b16Nv11xcTGfffYZiYmJDBgwAB8fH3bt2sX69evp2rUrnTp1qvW5y8vL+eqrr4iPj6dLly6MHTsWKrNeIiIi+P3331mxYoW0qR80pnTuk2ibc6+czVeuXOHYsWPs3LnT5PeZmZmMHz+eoKAg+dmTnELyZ0Mo5YYNG/L0008bfZefn09GRgYYOKIrKio4d+4cKSkplJSU0LhxY7p16yZ34m/evMmJEyewtrY2WpzHxcWRl5dHy5Yt5cInOzubhIQEbty4gZ2dHX5+frRp00bOr4yMDM6fP4+rqyvNmzfn2LFj/PHHHzRt2pSePXveNj2ntLQUrVZbLRq9sLCQ33//ndzcXBwcHGjevHm1EhEajYZjx46RmZmJi4sLvXr14tKlSwBGhkZZWRmxsbGkp6dTv3592rZti6+vr/ze8P49PDyIiorC09NTOm4MSU5OBpBRy9euXWPmzJlcvHiRkSNH8s4778jnMHeM6zJ+V65cIT4+Ho1GwzPPPIO1tTWlpaW4uLgYRUqkpaVx5swZbt26hZubGz169JBpfMXFxURHR2NjY0OfPn04efIkly9fxsnJiT59+sjyM+Yel5aWxoULF3j66afp2LEjVKbbHDx4kIqKCrp3746joyPnzp0jKyuLli1botPpOHnyJPb29gQGBuLk5ERKSgq///479evXp2/fvvL8YnwaN26Ml5cXR44coaysjO7du+Pl5UVeXh5Hjx6lvLycHj16GO0g381vX9PcVFB4GLzxxhvy/0dGRnLq1CnatWvHl19+We1YNzc3vLy8iIyMxM7Ojq5du9K+fXsAcnNziY+PZ926dVAZlazT6eS7RaXD1tr6/5kuKSkpJCQkoNVq8ff3x9XVlaNHjxrJNSEXhdytqKjg6NGjlJWV4e3tjZ+fn9RJly5dQq/X4+3tTefOnaUeKS0tNbp+amoqSUlJ7Nq1C29vbw4ePEinTp1wcXGp9rw6nY69e/fi6+trspRQcXGxdCR9+OGH1WTA2LFjSU5OJiEhga5du0KlQz82Npbr16/TqFEjunbtarRoSkpKIiMjg2bNmkknu1qt5tSpU3Ih7OHhQefOnY3kR20yKTMzU96rIZ6entjb23PhwgU5puL3uXz5MjY2NtIZI8bS8LPb6QQqZfb58+e5dOkSxcXFNGnShO7duxvpDXFfws5Qq9XExMSg1+t55plnaNy4Mbdu3UKv12Nvb29Wau6TRElJCbt27aJVq1bVokXrElUvOHr0KElJSRQUFDB58mT5W23dupWEhAQKCwv56KOP6nTO+Ph4Nm7ciKurKxYWFg/E2VxQUMDatWvx8PBg4MCBrF27loyMDCmzqLT3RFCFh4cHx44d4+rVqzg7O9O7d2/s7e3Jy8sjLi6OkydPQqVtHBsbi7+/f602sVqtlnLMw8ODmJgYk/aOKZvB3d0dtVotN7yysrKIiYl5YBHBgtOnT7N//34GDRpEQUEBx48fJzs728geNteuM+xjI6Jz09PTSUpKwsrKil69epGYmMj169eN1guFhYWcPHnSSD8cPnyYsrIyevToQVZWFqdPn8ba2ppu3bpJe+3kyZMcPXoUgMaNG3Po0CEaNmxIZmamWTpA4fFDbFKLdYNarWbhwoXs27cPf39/pk6dKtc1Qv/4+flhYWFBTk6O0bm2bdsmdeKlS5eeuOCb1NRUNm3axNtvv01mZiY7duwwufmv0+lYsGABN2/e5Pnnn2fTpk23DRK4cuUK69ato23btgQHB2Nra0uDBg1YsGAB8fHx+Pv7c/DgQWJjY/nb3/7G+++/D0BOTg4HDhzgwoULNZYkSkhIYOfOnfj6+vLdd99hb29P3759uXLlygPLDDaHiooKCgsL8fX1Zfr06Xh5eaFWqzl8+DB5eXnk5uZia2tLXFwcAB999JG07Xbv3s25c+e4deuWyZIa5ozd+vXryc/PZ8aMGfTt2xeNRsOJEyc4deoUN27cqNGZHx4eTmJiIkOHDmXSpElYWFjQuXNnSktLa92wFaSkpBAfH4+dnR2zZs2SuuC5557js88+w8vLi4KCAulsNrQjS0pKcHFxoWvXrkZR6sIX0KpVK+rVq0dMTAzFxcW0bdtWRolT6eiOj48nMzOT0tJSGjduTJcuXXByckKj0XDo0CGTOteUbV5SUkJMTAzZ2dk8/fTT9OrVi5SUFDIzM430V232uV6v58KFC6SkpFBUVISjoyOdOnV6JKpI3BNnc15eHmq1mmeeeYaIiIhq36enp/PMM888MilGCveWq1evQqVSrrqAS09Ph8poXpVKhU6nY/78+ezduxdPT09sbGxITU3Fy8uLxYsX06BBA5KTk5k1axY9e/Y0coQuW7aM5ORkWSJi//79BAcHo1KpaNq0Kenp6ajVat555x2Z/nTy5EkWLlzIX//6Vy5duiSjT6gh1cmQadOmcerUKWbNmiWbNv722298//33FBcX4+HhQU5ODlqtlpEjR/L+++9jYWFBUVERX3zxBfHx8ahUKuzt7YmJiSElJQVACo60tDSmT59Oamoqf/nLX8jJySEjI4PPP/9cpnwfO3aMxYsXExQUxO+//86lS5dMpglrtVq52PH29ub06dMEBweTl5fHxx9/zIsvvmi0wWPuGJs7fvv27WPWrFkAuLu7s2rVKl566SU5LywsLNDr9axfv55Vq1bh5uZG06ZNiY2NpW3btsydO5f69euTlpbGrFmzaNu2LSdOnGDv3r3yer/99hvfffedHDtzjjt48CBr1qzhvffek0ZjXl4eM2fOhMqsC4Dt27dz4MAB/va3vxmlGR0+fJgRI0bw+eefo9VqAYiJiZF/Hxsby4IFCxg0aBCJiYmySeeGDRv45ptvmD17ttxs2bx5M2vWrEGlUt31b29qbiooPAqIRVqbNm2qfZeUlMTUqVNRq9W0a9eO5ORkSktLmTRpEi+99BKnTp3i22+/lcf/8MMPfPHFF+j1embNmoWXl5esk7l3717mzp0LgKurKytXrmTMmDGsWLGCPn360K9fP8rLy6Vc9PLyQqfTsWLFCrZs2cKwYcMIDAxErVYzb948Dh48SIsWLbCxsSExMZFhw4ZJZ9jVq1eNrr9792727NkDlQueWbNmSQd5VRISErh69SoffPCBSQen0JGibIYhHh4eODs7Sz0DcPz4cSlr27Zty/nz57GysmLevHlyERQWFsb+/fuZMWMGLVq0IDU1lcmTJ5Ofn0+rVq3Iz88nNzeXXr16MXPmTKysrMySSdHR0SxevLjaM3z++efS4H/33XelA3/hwoUkJCTg4OAgHcvi/t98800CAwNr1QllZWUsWLCAiIgIHBwcqF+/Prm5uXTo0IE5c+ZQv3596Xxq1qwZDRs2pLCwkODgYGJjY5kyZQqNGzemqKhIpi5v3bq12sb4k44Yw379+t11eqdWq+XMmTMEBASwf/9+CgoKcHV1JSUlhUOHDtG5c2cuXbpUJye2Tqdj9erV9OnThyZNmrBv3z5Z+/d+sn37dvLz8/nkk09wcHAAg3dWIBacjo6OhIWFcejQIQoLC9FqtQQEBDB79mzi4+Ple0tlY7bXXnuNzp0712oTi+v5+vqyYsWKGu0dUzbDuHHjjEocrV+/nmHDhj1QZ7P47QD+9re/SbsrKyvLyNlsrl0n1hxi8y4lJYVp06ZRUVHBjBkzsLW1ZfXq1Zw5c8aopNy5c+eYNWsWAQEBUj8EBwej0Wh466235D1SucGyevVqGjRowMKFC6UdFxcXR1xcHAMHDmTfvn1ghg5QePwQwUEtWrQgOzub2bNnk5CQwLBhw5gwYYKRP0O8n97e3mg0GiP5kJKSwt69e3n33XdZunQpXl5eDy3S8mFhZ2dHSEgIzs7OfPLJJ1BDBsimTZvkey4cdbeLAj9w4AAAgwcPlpvmwrkpSrfu2rULgP/7v/+Tfyd0f2FhocnzXr9+XcqoXr16UVRUhEajwcPD45Fw3BliaWnJu+++a1S+KTMzEwAHBweeeuop9Hq9DNAQ0bR//PEHVJbbqKl2c21jl5+fz549e2jYsCEBAQFgUIanqKiImzdvVnM263Q6cnNzZYZPly5dyM3NRaVS1bnpqHim0tJSLl68KNf49vb2LFy40OjYgoICfvjhB44cOYKTkxO2trbk5ubi4uLC999/L+ej8AW88cYbhIeHU1BQIM+xaNEiOnXqhFqtZtq0aZw5cwYvLy8sLS25cuUKHh4eLFy4kFu3bhnpekOdW9U2Lyws5NNPPyUpKQkHBwfs7e05duwYN2/eJCYmRlaFqM0+1+v1LF68mO3bt+Pi4kKjRo1ITk7G2tqauXPn0rlz5zqN7b3mnsRwd+nShb///e8MGzbM5PeXL18mPz+fJUuW8OOPP3L48GFZ+F3h8Uc4UfPy8li9erXRf9u2bQMDozArK4ubN28SGBjIqlWrWLVqFa1atSItLU2ex1S5CcPIjqZNm6LX66VhuHbtWpYtWyZTaXJycmQDAtE5OTo6mrfeeos9e/Ywfvx4qIy0qAmNRsPZs2fBwDl8+vRppk+fjrOzMxs2bGDDhg2sXr0aBwcHwsLCOH/+PAAbN24kPj6e/v37s23bNjZv3oytra0sc+Hh4YFGo2HOnDmkpqYyffp05s+fz/Lly3F1deXHH3+kvLwcDIzqffv2MWzYMP7zn/+YdC5mZmbKmkdnz55l8uTJFBYWsmDBAoYOHVptQWbOGJs7ftnZ2cyZMweAefPmsWnTJhYvXsyWLVsA5I5+VFQUq1at4plnnmHVqlX88MMPfPjhhyQmJkqjQRhpIkpg165dcoERExMjm4mZe5x4HkPDpmpUjF6v59y5c1A5hzdt2sTs2bMBOHHiBIcOHWLr1q189dVXUJkSL34fMY4XLlxg+vTpbNq0CZVKRUFBAUuWLJHpRtbW1uTm5pKenn7Xv72puamg8Kgg5KDhjrsgLi4OHx8fgoODWbRokYx8Pnz4MAADBw7kww8/BCAoKIjIyEgGDBggN3FEZEF2djYLFizA1taWn376iS1btrBw4UJWrFgBlQY0lfqmuLgYZ2dnGjVqJOXSm2++yQcffICtrS0hISEcPHiQoKAgli1bxuLFixkwYAA///wzFy9ehEpDz/D6kyZNon///gDMmDGDyMjIGt9FIdt69epl8nvxnnfo0MEomresrIy9e/eSk5ODk5MTrVq1IiMjg+nTp2NhYcFPP/3E999/z/fff09BQQGbN2+GyugGUb5I3NPevXvJz8/n22+/Zfny5WzYsIGhQ4fi4OBAbm6u2TJp0KBB7Nixgx07djBq1Cg51t26dZOOOKF7T58+TUJCAgEBARQXF1NWVgaVZQSsra0ZNGiQWTohLCyMiIgIevTowbZt29i4cSO9evUiISGBqKioar/PjRs3+Oqrrzh79izffPMNQ4YMAQO57+LiokQfmkDMw3uR4i3skWeeeQYqF3pUOvlHjBhBdnZ2jWnUNREVFUV8fDyvvfYaTZo0IS8vj/z8/Lu+19tx7do1Nm3ahI+PDwMGDJD144UtJxBjV1JSQkBAAGFhYfznP/+BSgdweno6ffv2ZcaMGQD07duXyMhI3nrrLbNsYjG/U1NTb2vvmLIZXnrpJblg9fX1JTIykokTJ97XcatKVFQUZ86cYejQoTRv3lw6bKo67c2168S4+Pr6kpyczCeffIK9vT3z5s2jbdu2lJSUyEwHQ7ks7Flh92ZlZcnAgtzcXLZu3crGjRtxdnYmLy+Pq1evYmFhQUhIiNRnixcvJjIyks8++8xsHaDweKHX6+UmdVFREZMmTSIhIYF//OMffPjhh9UC58R75+vry9NPP01aWprUg9u2bcPT01OuQQyjI58UvL29cXZ2pry8XK5dqpbcjImJYeXKlfz973+ne/fu0oa5Xc1dkW1mWBZIOJlVKhV//PEHiYmJqFQqozWgyO4SNoshGo2GMWPG8Msvv0BlmZxRo0bx888/39UYPChu3LghN9jGjRtH/fr1UalUjB07Fq1WS3BwMGvXrmX27Nn4+fkxZcoUk+cxZ+yEjO3atat09ut0OrKysqCGmvCXLl1i9OjRcvPu66+/ZtSoUXINXhd8fHykk/uDDz5g8uTJ7NixQ+pLgU6nk47mt99+mx07drB582aCgoLIzc1lzZo1UPnei/uIj49n3rx57Nq1S/pcRLDdiRMnOHPmDG+88QZr165lzZo1fPjhhzRv3py0tDSaNWtmUueass03btxIUlISQ4YMYefOnWzevBkrKytZ/9rT09Ms+zwzM5Pt27fTuXNnNm7cyPLly1m0aBFdunSReu9hcs8bBFZFrVaTl5dHREQEQ4YMobS0lLlz55KQkMB77713vy+v8AAQL+e5c+dqFBgi4qpp06YEBwdTWlqKWq2WChkDwSQW+IbOCsPIJbFTJyK99u3bR0BAAJ6entIwFQjF9v7778tGLKLmXYMGDWp8JisrK7l4F/W6RFfm119/XRrLXl5etG/fnhMnTnDt2jU8PT3Zvn07VKaWi2fq3Lkze/fulRHex48fJzExEVdXV/z8/KTgdXd35/Tp0+Tl5eHm5ibH8/XXX+evf/1rjfcrhKtWq2XRokXyc9H9uCrmjLG543fgwAG0Wi1DhgyhS5cuUOlgdnFxITc3Vy5et27dCpUNFgoKCigoKJBpomKRIRZxw4YNY/jw4UbXU6lUUqGZc5xOp6sm2DFYcIjIy+vXr8u0twkTJuDs7GxUF3H8+PE0atRIPkejRo3kjqoYn7feekuOZf369VGr1YwZM4ZmzZrJtO2ioiLs7e05derUXf32opstBnNTQeFRoKysTM5bU7VFX3nlFYKCgigqKuL69evcunULQKaxWVlZyffTz89PdpQWmzqGG1dC5oiNzNatW+Pl5UVaWpo0joVcbNmyJT/++CN79uzho48+khvj+fn5sqlPt27dyM7OhspIaSqdES1atKh2fUtLS7kY9fHxqbHzdWFhIREREfTv31+esyri3AkJCTIbBIOoG2dnZz777DMaNWrEqlWrKC0tJTAwECsrKzIyMqRMFDI9NzdXpk4KPSX07NGjR2nUqBE+Pj58/PHH8lrm6iORxh4REcHWrVvx9PRkxowZODk5SV0nrrVjxw5atWpF7969OXbsGCUlJVy5coVTp04xfPhwXF1dpfOtJp2gVqul3n3ttddkBM4777zD6NGjpfNPjKFKpeLzzz8nLS2NefPmGUXKNG/enB07dmBlZaWUcTOBGEPhGBQMHjy4zk408d61a9cOJycnbty4waVLlzh8+DBLlixh7dq1dapXXFJSwpo1axg0aBCtW7eW70ZWVtYdlfgwl9DQULRaLePGjaNevXoyquvcuXNGUdVi7IYPHy7tJB8fH2xtbdFoNKjVapk9QKWzU8gMc2xic+2immyG2NhYANq3b1+jrLpfFBcXs2rVKqytrWUpRXd3dzBw0gnq+pw3btxg6dKluLm5ERwcLDeR0tPT0Wq1+Pj4GDXKu3DhAhgEHwib19fXl4kTJ8rzOzo6kpeXJ51RxcXF8t68vb2xsbExckjeTgcoPH7k5ubKNYFhbymtVmtSd4j339vbG0tLS7RarYz8jIiIYOrUqfJ8da1R/2dCbP43bNjQyB7KyMjgu+++o1+/fowePVrWZHd2dq4xA6mgoMAogEtgWMtd2JLt27eXazYMgpCEHDJEr9czc+ZMmUW6atUq6tevf1t/waNCYmIi3333HVevXuWNN95gxIgRUKk/GzZsiLOzM9HR0URHR+Pm5sbEiRNrLAliztgJZ7/hnP7jjz8oLS1FpVKZbDTapEkT3nvvPX766Sc6dOgg+zncSVNSKysrvvrqK5lpGBsbK3WdYenQM2fOcOTIETw8PBg1apTMMAwICGDLli3Ex8dDFV/A5MmTpd0j5pewP4VfJT4+nri4OFq2bMnw4cOlzqJyTlNF5/7xxx9GtrlarZblh0eOHCmDTQICAoiIiKBVq1Y4OjqaZZ8LuyEtLY0jR47QoUMHOnXqZFbt6wfBfXc2W1paMmfOHDp27CgXKp06dWLChAm8/PLLcrGg8HhSVFQk042Cg4OrCYyvvvqK3NxcaTCKjvHCSDNERCybaigojHRRJ8/CwoL33nuPb7/9lqVLl7J06VJ8fX0ZNWoUzz//vCxnIe5NOEENz3W7uktWVlZG6R8FBQWy1l7VnWkRXWNnZ8eFCxfQaDS4ubkZRQgVFRWBQYS3MJZzcnJ4/fXXq13fwcGBwsJCqUxriooTCGPHw8OD8ePHU1JSwty5c1mxYgX//ve/jepymjvG5o6fGBfDY8rLy+UOs6enJ/n5+XIXdPHixdVSsYVsEDuHhp1uDYW2UBLmHHft2jXUajUuLi5GBothVAxVouLEceJc3bp1k/NApCaJhhWG4yOcGvn5+VJZic9yc3MpKirC1tYWd3d3mXp5p7991bmpoPCokJGRISPFqqYb5ufns379eg4cOGCUmkaV5n1CNho6uES0tJBXwjg0lMV6vZ6bN28a/a2Qi8eOHZPHGUa1CucDlR3iqyKMy6rXFw5dBwcHkwsWQXR0NBqNRjZPNIU4d/PmzaVjRMjK8ePHM3z4cCkff//9dwB+/fVXfv31V6PziA1Cw8wNIfdfeuklTp48yZ49e9izZw/Ozs68+OJvDVvhAAAgAElEQVSLBAUFYW9vb5Y+Ehw/fpw5c+bQqFEjgoOD5aJRHKPVaklOTubIkSN88cUXcqFSXFxMZGQkAH/961/N0gnJyclycWoYvVR1I0OModiEs7a2rrZIrVevnlHkuIIx58+fR6VSUV5eLlNsqSH6qzYuX74sF1TNmjXjxo0bHDhwgFGjRsno1LpEUIeHh3Pt2jVatGjBypUrZXRV1drJ95KEhAR+/fVXrK2tycrKIiwsTGZkFhQUcP36dSlLhBwRNdWpXOALWSje36r14zHDJsZMe+d2NkPVzbIHSXh4OBkZGfj5+Uk5nJubCwbjITDnOQ3trmXLlkFlHwDDTYeqmShUOgiqBh+IcQkMDJSysqioSDpamjRpAibKAVIHHaDw+CF0qJ2dHWPGjKFTp068//77hIaG8uyzz1aruSz0j5eXF8XFxVC5Wbxlyxa8vLx49tlnWbp0KTzhDWzFe2lYcrOkpIQ5c+bg5OTEpEmTsLKyku9bx44da2yiJmSIt7e3kY4ScrRFixZSLhuOeV5eHteuXUOlUpncRD1+/Dhff/21/LfI5F28ePF90zV3i06nY/v27bIZnmF5xdLSUj755BMSEhIIDg6me/fupKam8vHHHzNlyhTWrVtn0g8n1rG3GztDp6dAyM5u3bqZLBsXGhoqA88SEhIYPXo01tbWhIeHGzm1zcXe3p6goCBGjhxJamoqx48fZ8OGDYSFheHl5cXQoUPleiEgIMDIFyKCXcRGgqHeMJwbIpBDfBYQEEDXrl05efIkH3/8Mba2tvTv358333xT2sOmdG5V2zw5ORmNRkOjRo2Mxlk00ayLv8jNzY1XX32VjRs3yiAOf39/xo0b90jM2/vubHZwcDAyHKiMKBR1SxVn8+ONeHkcHBwICAgwEhZFRUVSIXh6epKQkMDXX3+Nj48Py5Yto2nTpiQnJ/Ovf/2Ltm3bolKpuH79OtnZ2djZ2RkJMLEoNXxx+/btS9u2bTl16hSxsbEcPHiQOXPm4OLiQpcuXeS9eXt7G0WAmjL4a8OwtpNhva1bt25Jw7dp06ZSufn6+hrtgAsBLBy0wtnyyiuvEBgYWO16jo6O0oDx8fGptb6kEIavvPIK/fv3p7y8nPDwcJKSkvjll1+MStzk5eWZNcbmjJ9er5f/NnTiZGVlya7jbm5uRumSS5YsqWZANG7cmLKyMpPlIcTYia7P5h4n7t8wXVen08lGCUK4V63/Z3guQweHUB5Vz+/t7S3nhFC+Yj5TRYHZ2tre899eQeFRQcz11q1by/kvmDt3LseOHWP8+PEMHjyY+vXrM3PmTI4dOyY3fgwdCSJ90jBaWnwm3lnDd+Pq1asUFBTg6Ogo5ZqQi8OHD8fPz4+5c+eyd+9e6YwRct3f318uKgzx9PQ0eX1Do7GmiDa9Xs8vv/yCs7OzyYauVZ9t8uTJ0ukeEhJCSEgIFy5cMBpHIV+mTp1abdEqoiDF2BjKLh8fH1asWMHp06eJi4vjwIEDhISEYG1tzeuvv26WTKLS6J05cyZ2dnYEBwcb6VBxfa1WS3h4OO7u7vTu3Vs6eDIzM9mzZw/9+vWjefPmRql9NekEUaqpZcuWNY6z4Rh+9NFHZGVlsWXLFg4ePChLfSjcnrKyMlnyxFRTz7py4cIF2rdvT7169WjSpAlHjx7l1KlThIaGcuLECaglPdqQ69evExoaSocOHbC2tuaPP/6QUTxVyzDcK3Q6HatWrYLK+SycmoZkZ2fj4uJilEFlaAOJ+e3k5IS7uzs6nc6ofjyVi+3abGJz7R0hG0zZDFU3yx4U4rejMn266obSxYsXKSkpwd7evs52nYuLC1OmTCEkJITExETOnTsnncumFvrp6enk5+cbZXwY2usCcf62bdtKJ5a4B0Mb0RwdoPB4InRoYGCg7P8TFBTEli1bWL16tSyph4H+UalUuLu7y426EydOyHIr9erVq7F8xJNE1TUUlb2XEhIScHV15bPPPgOD4KwzZ84wceJE5s2bV60RuijLZehDysnJISkpCXd3d3x9faW8NQzOEevcZ5991mRz9cDAQL777js+/fRTunXrJnuIPKoNhTUaDfPnzyciIoKePXsyadIkIz0UGxtLQkIC3bt3p0+fPlBpG7Zs2ZLTp09z4cIFk344Mb63GzuxcWzoHxCO3Zp6AkyYMAEHBwdCQkJ45513CAoKwsLCosZNhdtx69YtdDodDRo0wMrKihYtWtCiRQtsbW1ZunQpZ8+eZejQoTIgsGr5NMMMLGrwBRjqJbHJVL9+febMmcPZs2eJi4vj0KFDRERE8McffzB//nyoQedWtc3z8vKgUmcbPr+Yt3XxF1GZ8Tdw4EBOnTrF8ePHiY2NJSkpidDQ0IdeJ/6e1Gy+HSkpKbL2jUCr1VJaWqpE5/0JMGwOWHVXqmrTHrFwDAwMlM4IEekgXm6hqN3d3Y1SqEVElHBwpqWlcfr0aRo3bszzzz/P559/zjvvvAMGEaiGtTAFhgsDofSTkpKIjo42qvOTmZlJWlqa3KU2TKExrBMoFk6dO3fG19dXLoLEjhmVBre4f2FEixe/QYMGtGnThjZt2lBSUkJRURHe3t5YWFhIA7e2Gl+GzyQW/zY2NtJ5smbNGhkJZO4Ymzt+FRUVMqVEKCcqDQgMUnAMx69Zs2a0adMGX19fLl26hIODAy4uLjIqsmrqVNWSH+YeJ1LiDRXMvn375KJM/P7CuWW44DD1WdVNCjE+hruG5jiu7/a3rzo3FRQeFURkXdVGH4WFhVLWDxkyhMaNG5Ofny8/E3JRvLO+vr5yoS/ed1dX12qZM2JRotfrZX+ATp06YWVlZSSrhg8fzsCBA3FxcSE6OloueoSRbGlpKd9FlUpFVlYWTz/9NCqVyuT1RU06U3WpBZcuXSIhIYEXXnihxojamiLBRUO+w4cPG0Vfi2Pc3Nzk/aanp6PX66vJM5Eqfv36dRISEqioqKBXr17885//lJEPQl6ZK5O++uoriouL+frrr6v9xsLZfOXKFSIiIhg1ahR2dnbydwwPD6e0tFRufJqjE0RkueFmb1JSEu+99550BooxdHFx4aWXXpINJA0bvWq1WtLS0oxqaSr8/4iyA4ZOgDtFvHdC37m4uHDkyBGCgoJwdnYmJSWlWrbR7diyZQs6nY4vv/ySadOmMW3aNL744gscHR2r1U6+Vxw+fJjTp0/To0cP2Z9D/CcixoQdkZmZiVqthip2X3R0NAD9+/fHxsaG7OxsCgsLjTbDzLGJxW9Tm71Tk82g0+nkZsyDjsDdvHkzarWacePGGY1haGiorHsrZL65dp2QWQEBAXTr1o3BgwdDZWkfgbDlhVyrqKhgx44dULlWMafMmqENJ+aZYTakOTpA4fFE6FBDeThq1CgcHR05cOCAzObEYN6KaF0x59avX0+zZs149tln5Sa6KRvmSUK8y4Y1gMvLyxkwYADt27fH3d0dd3d3Gdnp4+ND27ZtsbW1rbZOF/bD9evXZcZJeHg4ACNGjMDGxkZmign7oaysTJYtGDRokMl7tLCwkPLn2rVr6PV6rKysOHXqFIcPH77vfQLqgl6vZ9GiRURERDBo0CCmT59ezaFqap1469atauMYExNDdHS0dGyaM3Yim0RkFGZnZ7N3716cnZ2rBZkKLC0tpXM7KytLljXbv38/J06cMPIj3I6QkBCGDBnCokWLjMqF6vV6GeQonkH4Gg0ztjQaDb/99htU9oqhhnW/6LPk4eFBo0aNUKvVJCUlkZ2dTadOnRg3bhxLliyR/bJ0Ol2NOreqbS7mbWFhoXyGlJQUqcuEr6E2+7y8vJyUlBSSk5Px9fVl5MiR/PDDD/Tq1Qu1Wm3k/3lY3PfIZisrK+bMmYOdnR3PPvssVBbMb9as2ROdTvJnQZQkMGw0JzCs+YNBSvKBAwdo1qwZqamp8piioiIyMjKkAXr58mVWrFhBvXr1OHbsmBSYTZs2lQ07Dhw4wNixY+ncuTMlJSWyWZCIDhP3ZmggCoPfyclJpjssX76cuLg4Zs2ahZeXF1qtlrfffpvi4mJWr16Nr68vTz31FM899xyRkZEsXbqUIUOGkJuby+rVq3FxceHDDz/EwsJCCqn4+Hh27txJo0aN2Lt3r7x/IXh69OjBunXr2LVrl2wYuHjxYjQaDStXrsTe3l4auLWlPubk5EgFYbhr7u/vT+/evTly5AgbNmyQjbfMGWNzx8/KyoquXbty+PBhtm7dilar5eLFi1Lpi/F46qmnCAwM5NChQ4SEhNCjRw/27dtHREQEo0ePZsKECVL5derUyeQun3g2c48Tzo/IyEicnJz4448/ZFRzixYtZMRg1ZIipsqMmIpKMjU+4jPD36yqArub395wbq5ateq2pWAUFB40YkOm6gK8Xr16ODg4UFxczJYtW2jVqhUHDhzAw8ODjIwM0tLScHZ2loZmdnY2Bw4coEOHDkaplyJDoVu3buzcuZMVK1aQnp7O+fPnpYEpFohCLoo0ZxsbGwYNGkRoaCj79+/n7bffpnXr1ri5uXHixAl+/vln3NzcCA0N5dy5cwQHB9OnTx+T1xeLobNnzxIVFVUtPY/K9HgqM3BqQpzbMIqOSqey0De7du2STVwGDBhAYmIiGzdupLy8nIsXL7JixQratWvHggULsLKyqia7Dh8+zKJFixgwYACDBw/G0tJSluAQDuPaZJKlpSVffvkleXl5+Pn5ceHCBekEb9WqFT169JB65ejRozg6OsrmWUIOR0VF0bFjR1lDzhyd0KZNG5ydnUlOTiY0NBQvLy/Wr1/PxYsXZeM/Q31gYWGBn58f7dq149y5c8TGxtKzZ08uX77M3//+d5ycnKTTSeH/R4yhoRPgThE2gpABLi4uODg48OKLL0JlxI8oRVUbqampbNu2jfHjxxs5IC0sLGjZsmW12sn3gpKSEtl8euzYsdXKAbVo0YIjR47IMRP/6+HhwfLly/nb3/5GVlYWmzZtws7OjqFDhxodZ5hCbo5NLII2arN3bmcziE2X/fv34+/vb1S26H6RkpJCWFgYrq6ujBw5slqmS8uWLTlz5gyZmZn4+PiYbddVtbF69uzJokWL+N///se4ceNwcXGR19q0aRM5OTmcPn3aqBQGBvrBzs5OlsvARJk1DNKYTaU6304HKDyeiDlnmLnj7OzMmDFjWLJkCStWrKBjx47Uq1fPqC8EBo6toqIi2YRYrAHq2hT1z8CGDRtkZKiIet28eTN79uyhQ4cOvPrqq0a1bqnsgZObm8ukSZOk/K26Tnd3d8fPz49Lly4xf/58SktLiYyMpF27dtI2+P/Yu//oqO46/+OvmUwmQ34RSGqaJoRGqjQRCoFK3YK0SIpGlrVQFFahtBt7jNr110r4TWhLONgqLqvoVnYhVdCyW6EtxU21RhI8cIzFlQxQIEhnWiCw0ymhTJI7yWTm+0eb+fKzhdxJbn48H+dwejJ3+vm8cjM/PvO+n/l8Oi8YPf/887LZbDp8+LD+93//V7Nnz77iYvnFbr31Vt1555169dVXtWbNGqWnp2vnzp2Kj49XZWVlt52rG7V3717t2rVLeu+i58XLf+i9TfPy8/Plcrn0pz/9SZWVlcrNzdUrr7yit99+W4WFhSooKNCFCxe0aNEi6aK9la7n3I0dO1YvvPCC/uM//kNHjx7V7t27deHCBZWXl0dn217NhAkTlJycrJ07d0Y3qa6urta4ceOu+S3Ay33iE5/QL3/5S/3+97/X+fPnNWrUKHV0dOi1117Tq6++qsTExOi+EHfffbe2bNmiXbt2KTc3VzfddJNeeeUVHT16VF/4wheifV5tedHLJ92dPHlSpaWlGjlypB588EElJyfr4MGDCoVC+tSnPqW4uDgFg8Grvude3n7n5/fjx4/rySefVGZmpv76179G9yPrfM/7oPF5WlqaFi9erNbWVj366KO65ZZbdPbsWR06dEg33XTTJe9vVun2YvOtt96qxx57LLoLeiQSUVJSklauXMlXj/qBzq8KXO3q/uWD3/vuu09//vOftW/fPj355JN68MEH9e1vf1tlZWV6+eWXNWzYMH3pS1/S5z//+eiu83fddZe+/vWv65//+Z81ePDg6FXhb3zjG3K5XNq6daueeeYZ6b1B5BNPPBGdUdOZ7eIBw8VLGnS+KHS+AHQWWTs3MUhOTr7kg8ajjz6qwYMH68UXX9Qf//hHOZ1OfeYzn9E//uM/RovIBQUFevjhh7V582atX79eY8aM0dy5c1VXVyeHwxGdSVdQUKDHHntMP/vZz6JfyRo9erRKS0ujLwxXy381nR9GLt8MxWaz6aGHHtIf//hH7dixQ9OmTVN+fr7y8vKu6xxfz/mTpEceeUSNjY2qq6vTwYMHtWDBAt17773avn179CqrzWbTt771LSUnJ+v555/Xc889p+TkZM2bN0/z58+XLvqK1cUfhN566y2dOnVKDocj+re43vtNmjRJv//97/Xqq6/qP//zP/W5z31O8+fP1xNPPBEdGHZuRHHx379zmZGLv27Z+QE6OTn5ir/PxR/Qr7YD+uVFajN/+87H5sW/J9AbhEKhq35I03sXuBYtWqSnnnpK27ZtU0FBgRYvXqw//OEP2rx5s1atWqVt27bptttu09/93d9p3759WrNmjSorK6/6lej58+frnXfe0Z49e/Tyyy/rn/7pn7R3797olX1dtoxO51jjnnvu0S9+8Qvt3LlTc+fO1eDBg/XEE09ow4YN+td//VfpvaLR4sWLo185vFr/d911l3bu3Cm326233377iq+2tba26je/+Y3uuOOO970g1Nl252alF5s+fbp+97vfadeuXZozZ45yc3M1Y8YMtba26tlnn40WoD/1qU/pq1/9qpxO51Vfz2bMmKFAIKBf//rX0W+c3HTTTfrKV74SLQB+0GvSiRMnou0eP348+uFZUvSrr06nU4MHD9b58+f1wAMPRN+LOgtqem8TlM4P29fznjBkyBA98cQT+ulPfxqdyZydna2FCxfqs5/97CXnsPP9wGazqbi4WIcOHdKuXbt09913R9+3xo8f32u/DmulKVOmRCeDmJWdnR2dMSRJ//AP/xAtuOq9JVOu14c//OFL2rrYD37wA5NJr27Xrl1644039JnPfOaqz8vO993OomTnf2fPni3DMLR69Wq1tbUpLy9Pjz76aHR8cPlyELrOMXF7e7t0HeOda40ZEhIS9NBDD6mysjK6g313F5sjkUh0XL5gwYIrCs16729bX18ffV253nHd5b9nenq6PvWpT+m3v/2tdu/erc9//vOaOXOm3G63Dh8+rNOnT+uhhx7SrbfeKo/Hc8VSSJ0znTt1FsY6CwLNzc3RtaQvHtd90HsA+qaL30MvX/Li7//+7/Xiiy+qoaFBv/nNbzRz5szo47bzeZ6SkqLExERlZmZGLzR3vv8MxM0Bd+/eHZ3R3KlznHjxPj+dLly4oBMnTkSXYNRFyzzponFNfHy8Fi5cqKeeeiq63u/06dNVUlISvcCdnZ2t5cuX60c/+pG2bt2qwYMH68tf/nJ0o9Jrsdls+va3v63169erurpaeu/C/Ne//vVu3ZD2RnVeBNN7S45czOFwaOjQoUpISNDq1au1cePGaKE8LS1NDz30kGbNmnXJxZD09PToZ/brOXeTJk3SF7/4RT333HPRyQCPPfbY+06w0HsXZFatWqUNGzZo27Zt0nvj2K997WvXvZzG7bffrieffFK/+tWv9Kc//Sm6MWBycrI++9nP6oEHHoiOvT/60Y+qoqJCGzdu1A9/+EPpvefikiVLNG3aNOkatQBdZfx/++23a9WqVXrmmWeimxs6nU7NnDlTDz74oHSN99z09PQr2s/NzdXKlSu1ceNG7d69W/fee6+++MUv6sCBA5fUc66nZrBmzRo9/fTTWrt27SV/n4ceeuiSMbhVbD6fL9KVDUBuVGtrq7xer+Lj45WbmxuTQnNLS4sef/zx6I6b6H6dX9cw4/z583I6ndE3hHA4rGAwGP1ZUnRX7ouLp1fTudP3oEGDLvn/zXr55ZdVX18f/UB/sVAopHfeeUcpKSnXfBy3tLTIMAwNGTLkfa9kd25qZbfb3/dKYHe43nN8PcLhsM6dO6ekpKToDLdr6fybXby7eHfpPL/x8fFd2uioO1n5twfMuPfee7V58+YbngnZ1tam5uZmpaWlRV8Xg8GgbDZb9LUgEomoublZLpfrmhuGdL5nOJ1OxcXFqbW1VV/4whd04cIFPfvss13aC6K5uTm6zvz1FCQ7OjqiFyUvf42/cOGC/va3vyk9Pb1b1mjs6OjQ+fPnb+h9r6OjQ++8844cDsdVM8vC16TreU/o/PukpqZ2aXbYD37wA33sYx8zPX7pL8rLy6PLqZi1ZcuWq25c0xV1dXWaMGGC6XbKysrU0dHRbUXpa+ncGPBGHqfXMyaORS673d7lTTIXLFgQLSD3BZ2vkd05zny/94D3c/jwYX3ta1+75oWUgeoXv/iFnnvuOb3wwgtWR+m3/vKXv+g73/mOXnrppateBOqLLn/9vJyZ14ILFy7IZrNd17n6l3/5lx5/v7kRnZ/5b+S96XrOXef4bfDgwTd0Mb9zvJmQkPCBdQO9982+zgkTV+vfZrNF12++ls6l98yObyORiAKBQHRMerU+r+c9NxgMRj/nxMXFaceOHVq/fr3mzJmjr371q1f0+UHj887aU0/UV67H3r179eMf/7j7ZzZ3GjRo0CUb1mDgury4abfbr3iTuN43QafTGfO1v8+fP69f/epXKisru+rxziuG7ycxMfG6ips2my0mxd6uiOVAw263X/cV3+74m12Llef3g/TmbEB3cDqdVwyALh+IfdDA/k9/+pPKy8uVlZWlZcuWadCgQdq+fbsuXLig++67r8ubDt/o1f+4uLhrDvhSUlI0duzYLuW43r5v9DX04vUkr8Wq16TreU8wMzvj8OHD+stf/qJHHnmky20A16MrEx+uZ0wci1wDSVdeI7vSBxMFAGt90JjFzGtBf3p+d+Uz//Wcu65+po/VePNG+o/V39Nms31gW+/3nhuJRPStb31LBw4c0Fe+8hVNnTpVJ0+e1JYtW6SL1pG+vM8POl/XW3vqaT1WbAb6is41eS7eIAQAYL3CwkLde++9qqqq0pe//OXo7dOnT49uEgtc7K233tKqVasu2ZQQAAAA6Ek2m00PP/ywKioq9PTTT+vpp5+W3lta45vf/GaP7KvQkyg2A5d5v40DAADWcTqdWrx4sR5++GH5/X7Fx8crPT29x74tgb6H9VQBAADQG4wdO1ZbtmzRm2++KcMwlJqaqptvvrlXLH8RaxSbAQBAn5KZmanMzEyrYwAAAADAdUtISBgQG4faGhoaIrHYrM8Kzc3NWrBgQXSnYHS/9vZ2paamqrW11eooAAALBAKB993AD0Dv1rkx5/Xu/v5+QqGQXC6XQqFQTNpKSUkxPcbs3Hinqxvi4VLBYFDx8fExebwMdKFQKLqJE/6/YDAoXWUfB8ROR0eHQqGQnE5nlzbaxbW1tbXJ4XDwGtlNDMNQQkICj9s+JBQKKTc3V46kpKTr2gWyN3I6nRoyZIjGjx9vdZQB48yZMzp9+jTnHAAGqJqaGo0YMUIZGRlWRwHQBUeOHFFGRoYKCwtNt3Xw4EH5fD5NmTLFdFtut1sHDhzQvHnzTLVTXV0tn8/HWDVG9u/fr5ycHL5NEgONjY06duwYj83LeL1eeTwe3X333VZH6bfOnTun+vp6jRs3rs/Wfnort9uttLQ0DRs2zOoo/VJNTY3uuOOOfrVp4/UyDKPF5XL1vp3/PoDf71djY6NsPp8v0ht3LrweLS0t2rBhg8rLy62OMmBUVlZq3bp1qq+vtzoKAMACNptN27dv18yZM62OAqALZs+erZEjR6qiosJ0W2vWrFFtba2qqqpMt1VRUaEVK1YoHA6baqe4uFh1dXXy+/2mM0EqKCjQ0qVLTV8EgLRp0yaVlJQoEolYHaVXidVzH9dWXV2tqVOnyuv1Kjc31+o4/UpRUZGmTZumsrIyq6P0Sw6HQ3v37tWECROsjmKF1yTlWx3iRu3cuVOlpaVirj8AAAAAAAAAwDSKzQAAAAAAAAAA0yg2AwAAAAAAAABMo9gMAAAAAAAAADCNYjMAAAAAAAAAwLSYF5v37Nlz1dvD4bBOnDihffv2sTs0AAAAAAAAAPQzjlg25vF4tGzZMtXW1l5ye3Nzsx5//HF5vV6lpqbK4/GooqJCH//4x2PZPQAAAAAAAADAIjEpNns8Hu3bt087duy46vHKykq5XC5t2bJFDodDzz77rJ588klt27ZNdjsreQAAAAAAAABAXxeTSq/f71cgENDYsWOvONbS0qKXXnpJX/rSl+RwvFvbvv/++7Vs2TKFQqFYdA8AAAAAAAAAsFhMZjaPHz9e48eP1+HDh1VVVXXJsdOnT6ulpUW5ubl67bXX9Pbbb+u22267amEaAAAAAAAAANA3xXTN5qt55513lJSUpMcff1znz5/XoEGD9Ne//lWLFi3Sfffd193dAwAAAAAAAAB6QLcXm9va2hQIBDR+/Hg98MADkqQ9e/boiSee0J133qkhQ4Z0dwQAAAAAAAAAQDfr9t35UlNTJUmf/exno7dNnDhRNptNr7/+end3DwAAAAAAAADoAd1ebL7pppuk92Y4dwqHw+ro6NCgQYO6u3sAAAAAAAAAQA/okWLzvffeq61btyocDkuStm/froyMDOXl5XV39wAAAAAAAACAHtDtazZL0qOPPqpFixbpD3/4gxISEtTe3q5Vq1bJ5XL1RPcAAAAAAAAAgG4W02JzQUGBamtrr7j9Qx/6kDZu3Civ16uOjg7l5uZSaAYAAAAAAACAfqRHZjZLksPh0IgRI3qqOwAAAAAAAABAD+r2NZsBAAAAAAAAAP0fxWYAAAAAAAAAgCF22zEAACAASURBVGkUmwEAAAAAAAAAptkaGhoi8fHxVufokubmZi1YsEDHjx+3OsqAEQwGFYlE2OARAAaopqYmJSYmyul0Wh0FQBcYhqGOjg7FYvzf3t4um80mh8P8NjBtbW0KhUJKTEw01Y5hGJKk1NRU05nw7vm02+285sdAMBhUMBjU5MmTrY7Sq7zxxht64403NGnSJKuj9FsXLlzQ/v37lZqaKrud+YaxZBiGbDabEhISrI7SL7W0tMjpdMZknNHXdHR0hOPi4vrcE7a9vV3Z2dlyJCUl9dnCodPp1JAhQzRmzBirowwYZ86c0enTpznnADBA1dTUKC8vTxkZGVZHAdAFR44cUUZGhgoLC023dfDgQfl8Pk2ZMsV0W263WwcOHNCsWbNMtVNdXS2fz8dYNUb279+vnJwcZWZmWh2lz2tsbNSxY8d0zz33WB2lV6mtrdWJEyc4L93o9ddf1/79+5Wfn99naz+9ldvtVlpamoYNG2Z1lH6ppqZGo0aNUkpKitVRepxhGIbL5TJ3Bd4Cfr9fjY2Nsvl8vojZGQRWaWlp0YYNG1ReXm51lAGjsrJS69atU319vdVRAAAWsNls2r59u2bOnGl1FABdMHv2bI0cOVIVFRWm21qzZo1qa2tVVVVluq2KigqtWLFC4XDYVDvFxcWqq6uT3+83nQlSQUGBli5dqnnz5lkdpc/btGmTSkpKFIlErI7Sq8TquY9rq66u1tSpU+X1epWbm2t1nH6lqKhI06ZNU1lZmdVR+iWHw6G9e/dqwoQJVkexwmuS8q0OcaN27typ0tJS1mwGAAAAAAAAAJhHsRkAAAAAAAAAYBrFZgAAAAAAAACAaRSbAQAAAAAAAACmUWwGAAAAAAAAAJjmiHWDe/bs0Sc/+cnoz4FAQG+99dZV75uWlqa0tLRYRwAAAAAAAAAA9LCYFps9Ho+WLVum2tra6G2HDh3SwoULr3r/73znO7r//vtjGQEAAAAAAAAAYIGYFJs9Ho/27dunHTt2XHHszjvv1G9/+9tLbmtsbFRZWZkmT54ci+4BAAAAAAAAABaLyZrNfr9fgUBAY8eOveJYXFycXC5X9J/T6dRPfvITffOb39TQoUNj0T0AAAAAAAAAwGIxmdk8fvx4jR8/XocPH1ZVVdX73vd3v/udXC6XJk2aFIuuAQAAAAAAAAC9QExmNl+v1tZWbdy4UQ899FBPdgsAAAAAAAAA6GY9WmzetWuXxowZoxEjRvRktwAAAAAAAACAbtZjxeaOjg79+te/1qc//eme6hIAAAAAAAAA0EN6rNh84MABtba2aty4cT3VJQAAAAAAAACgh/RYsbmurk533XWXHI6Y7EkIAAAAAAAAAOhFeqzYvHfvXo0ePbqnugMAAAAAAAAA9KCYTjMuKChQbW3tVY/9/Oc/j2VXAAAAAAAAAIBepMdmNgMAAAAAAAAA+i+KzQAAAAAAAAAA0yg2AwAAAAAAAABMo9gMAAAAAAAAADDN1tDQEImPj7c6R5c0NzdrwYIFOn78uNVRBoxgMKhIJCKXy2V1FACABZqampSYmCin02l1FABdYBiGOjo6FIvxf3t7u2w2mxwO83uOt7W1KRQKKTEx0VQ7hmFIklJTU01nwrvn026385ofA8FgUK2trUpLS7M6Sq/S2tqqYDDIeelGoVBIgUBAqampstuZbxhLhmHIZrMpISHB6ij9UktLi5xOZ0zGGX1NR0dHOC4urs89Ydvb25WdnS1HUlJSny0cOp1ODRkyRGPGjLE6yoBx5swZnT59mnMOAANUTU2N8vLylJGRYXUUAF1w5MgRZWRkqLCw0HRbBw8elM/n05QpU0y35Xa7deDAAc2aNctUO9XV1fL5fIxVY2T//v3KyclRZmam1VH6vMbGRh07dozH5mW8Xq88Hg/npRudO3dO9fX1ys/P77O1n97K7XYrLS1Nw4YNszpKv1RTU6NRo0YpJSXF6ig9zjAMw+VymbsCbwG/36/GxkbZfD5fxOwMAqu0tLRow4YNKi8vtzrKgFFZWal169apvr7e6igAAAvYbDZt375dM2fOtDoKgC6YPXu2Ro4cqYqKCtNtrVmzRrW1taqqqjLdVkVFhVasWKFwOGyqneLiYtXV1cnv95vOBKmgoEBLly7VvHnzrI7S523atEklJSWKRCJWR+lVYvXcx7VVV1dr6tSp8nq9ys3NtTpOv1JUVKRp06aprKzM6ij9ksPh0N69ezVhwgSro1jhNUn5Voe4UTt37lRpaSlrNgMAAAAAAAAAzKPYDAAAAAAAAAAwjWIzAAAAAAAAAMA0is0AAAAAAAAAANMoNgMAAAAAAAAATIt5sXnPnj1Xvb2pqUl//vOf9Ze//EWBQCDW3QIAAAAAAAAALOSIZWMej0fLli1TbW3tJbfX1dVp1apVyszMlGEYam1t1dq1a3X77bfHsnsAAAAAAAAAgEViUmz2eDzat2+fduzYcdXjTz75pObMmaMHH3xQkUhE//Zv/6Yf//jH+vGPfxyL7gEAAAAAAAAAFovJMhp+v1+BQEBjx4694phhGPL5fBo7dqxsNpvsdrvGjx8vj8cTi64BAAAAAAAAAL1ATGY2jx8/XuPHj9fhw4dVVVV1yTGXy6XJkyfrhRde0C233KJQKKQXX3xRRUVFsegaAAAAAAAAANALxHTN5mspKyvTgw8+qAceeECSdOutt2r16tU90TUAAAAAAAAAoAfEZBmN99PW1qYVK1bonnvu0S9/+Uv9/Oc/10c+8hFVVFQoEol0d/cAAAAAAAAAgB7Q7TObDx8+rBMnTuj73/++HI53u/vOd76j6dOny+v16tZbb+3uCAAAAAAAAACAbtbtM5vPnTun5OTkaKFZkgYNGqRBgwapqampu7sHAAAAAAAAAPSAbi82FxQUyOfzqba2Nnrbrl275HQ6ddttt3V39wAAAAAAAACAHtDty2hkZmaqvLxc3/ve97R582YFg0G1tbXpscceU3Jycnd3DwAAAAAAAADoATEtNhcUFFwyg7nTpEmTVFhYqDfffFN2u125ublyuVyx7BoAAAAAAAAAYKFun9ncKSkpSbfffntPdQcAAAAAAAAA6EHdvmYzAAAAAAAAAKD/o9gMAAAAAAAAADCNYjMAAAAAAAAAwDRbQ0NDJD4+3uocXdLc3KwFCxbo+PHjVkcZMILBoCKRCBs8AsAA1dTUpMTERDmdTqujAOgCwzDU0dGhWIz/29vbZbPZ5HCY3wamra1NoVBIiYmJptoxDEOSlJqaajoT3j2fdrud1/wYCAaDam1tVVpamtVRepXW1lYFg0HOSzcKhUIKBAJKTU2V3c58w1gyDEM2m00JCQlWR+mXWlpa5HQ6YzLO6Gs6OjrCcXFxfe4J297eruzsbDmSkpL6bOHQ6XRqyJAhGjNmjNVRBowzZ87o9OnTnHMAGKBqamqUl5enjIwMq6MA6IIjR44oIyNDhYWFpts6ePCgfD6fpkyZYrott9utAwcOaNasWabaqa6uls/nY6waI/v371dOTo4yMzOtjtLnNTY26tixYzw2L+P1euXxeDgv3ejcuXOqr69Xfn5+n6399FZut1tpaWkaNmyY1VH6pZqaGo0aNUopKSlWR+lxhmEYLpfL3BV4C/j9fjU2Nsrm8/kiZmcQWKWlpUUbNmxQeXm51VEGjMrKSq1bt0719fVWRwEAWMBms2n79u2aOXOm1VEAdMHs2bM1cuRIVVRUmG5rzZo1qq2tVVVVlem2KioqtGLFCoXDYVPtFBcXq66uTn6/33QmSAUFBVq6dKnmzZtndZQ+b9OmTSopKVEkErE6Sq8Sq+c+rq26ulpTp06V1+tVbm6u1XH6laKiIk2bNk1lZWVWR+mXHA6H9u7dqwkTJlgdxQqvScq3OsSN2rlzp0pLS1mzGQAAAAAAAABgHsVmAAAAAAAAAIBpFJsBAAAAAAAAAKZRbAYAAAAAAAAAmEaxGQAAAAAAAABgWsyLzXv27Lnq7U1NTaqrq9OhQ4dkGEasuwUAAAAAAAAAWMgRy8Y8Ho+WLVum2traS27/4x//qDVr1ig/P19NTU0KhUL63ve+p5tvvjmW3QMAAAAAAAAALBKTYrPH49G+ffu0Y8eOK469/fbbqqio0LJlyzRp0iSFw2H9+7//uzZu3KgVK1bEonsAAAAAAAAAgMVisoyG3+9XIBDQ2LFjrzh29OhRJSUladKkSe92aLdr5syZeuWVV9TU1BSL7gEAAAAAAAAAFotJsXn8+PF65JFHdP/9919xLBQKKRKJXNqp3a5IJCK/3x+L7gEAAAAAAAAAFov5BoGXGzlypM6fP6+amhrpveLzf/3Xf0mS3nnnne7uHgAAAAAAAADQA2K6QeDVfOhDH9KSJUtUUVGhZ555RufPn9c999wjSRo0aFB3dw8AAAAAAAAA6AHdXmyWpKlTp+rOO+/UG2+8oYyMDCUkJOi5555TWlpaT3QPAAAAAAAAAOhm3b6MRiAQ0J49e5SQkKDRo0crKytLR48eVWZmpjIyMrq7ewAAAAAAAABAD+j2YnNCQoKeeuopvfjii4pEIgoEAvr5z3+uWbNmyeHokYnVAAAAAAAAAIBu1u3V3vj4eC1dulTl5eXasWOH3n77bd133336/Oc/391dAwAAAAAAAAB6SEyLzQUFBaqtrb3i9k984hPatm2b3nzzTd1yyy1KT0+PZbcAAAAAAAAAAIv12DoWaWlpbAgIAAAAAAAAAP1Ut6/ZDAAAAAAAAADo/yg2AwAAAAAAAABMo9gMAAAAAAAAADDN1tDQEImPj7c6R5c0NzdrwYIFOn78uNVRBoxgMKhIJCKXy2V1FACABZqampSYmCin02l1FABdYBiGOjo6FIvxf3t7u2w2mxwO89vAtLW1KRQKKTEx0VQ7hmFIklJTU01nwrvn026385ofA8FgUK2trexjdJnW1lYFg0HOSzcKhUIKBAJKTU2V3c58w1gyDEM2m00JCQlWR+mXWlpa5HQ6YzLO6Gs6OjrCcXFxfe4J297eruzsbDmSkpL6bOHQ6XRqyJAhGjNmjNVRBowzZ87o9OnTnHMAGKBqamqUl5enjIwMq6MA6IIjR44oIyNDhYWFpts6ePCgfD6fpkyZYrott9utAwcOaNasWabaqa6uls/nY6waI/v371dOTo4yMzOtjtLnNTY26tixYzw2L+P1euXxeDgv3ejcuXOqr69Xfn5+n6399FZut1tpaWkaNmyY1VH6pZqaGo0aNUopKSlWR+lxhmEYLpfL3BV4C/j9fjU2Nsrm8/kiZmcQWKWlpUUbNmxQeXm51VEGjMrKSq1bt0719fVWRwEAWMBms2n79u2aOXOm1VEAdMHs2bM1cuRIVVRUmG5rzZo1qq2tVVVVlem2KioqtGLFCoXDYVPtFBcXq66uTn6/33QmSAUFBVq6dKnmzZtndZQ+b9OmTSopKVEkErE6Sq8Sq+c+rq26ulpTp06V1+tVbm6u1XH6laKiIk2bNk1lZWVWR+mXHA6H9u7dqwkTJlgdxQqvScq3OsSN2rlzp0pLS1mzGQAAAAAAAABgHsVmAAAAAAAAAIBpFJsBAAAAAAAAAKZRbAYAAAAAAAAAmEaxGQAAAAAAAABgmiNWDQUCAZ04cUKGYWjEiBFKT0+/5HhLS4uOHj0qh8Ohj3zkI3K5XLHqGgAAAAAAAABgsZgUm91ut5YsWaKsrCylpKTI7XZr0aJFKioqkiS9+eabWrRokRITE9Xe3i6Hw6G1a9fqpptuikX3AAAAAAAAAACLmS42t7W1aeXKlZo/f77mzJkjSdq9e7dWr16tcePGaejQoaqsrNTHP/5xffvb31Y4HNbjjz+u//7v/9bXvva1WPwOAAAAAAAAAACLmV6z+dSpU/L7/Zo+fXr0tsmTJyscDsvj8ejcuXN65ZVXdP/997/bod2uz33uc3r++edlGIbZ7gEAAAAAAAAAvYDpYnNWVpa2bNmi5OTk6G0ej0ehUEjp6el66623JEnDhw+PHs/Ly5NhGDp//rzZ7gEAAAAAAAAAvYDpYrPL5VJubm70Z6/Xq5UrV2rGjBkaPny4mpqaNGTIENnt/7+rpKQkSaLYDAAAAAAAAAD9REw2CJQkwzC0bds2bd26VfPnz9fcuXMlSZFIRJFI5JL7dv4cDodj1T0AAAAAAAAAwEIxKTafPXtWixcvVk5OjjZv3qzs7OzosbS0NDU1NSkcDkdnNzc3N0ePAQAAAAAAAAD6PtPF5nA4rOXLl2vixIkqKSmRzWa75HhmZqbi4uLk8Xj04Q9/WJL0+uuvKyUlRenp6Wa7BwAAAAAAAAD0AqbXbD506JAaGhpUWFioo0eP6siRI9F/gUBAgwcP1tSpU/Xiiy9K7xWnX3jhBU2fPl3x8fGx+B0AAAAAAAAAABYzPbP55MmTstvtWrhw4RXH1q9fr9GjR6u0tFTf/e539cgjjygYDGro0KH67ne/a7ZrAAAAAAAAAEAvYbrYXFxcrOLi4ve9T3p6ujZu3CiPxyOHw6Hc3Nzo+s0AAAAAAAAAgL4vJhsEXldHDoduu+22nuoOAAAAAAAAANCDmF4MAAAAAAAAADCNYjMAAAAAAAAAwDSKzQAAAAAAAAAA02wNDQ2R+Ph4q3N0SXNzsxYsWKDjx49bHWXACAaDikQicrlcVkcBAFigqalJiYmJcjqdVkcB0AWGYaijo0OxGP+3t7fLZrPJ4TC/DUxbW5tCoZASExNNtWMYhiQpNTXVdCa8ez7tdjuv+TEQDAbV2tqqtLQ0q6P0Kq2trQoGg5yXbhQKhRQIBJSamiq7nfmGsWQYhmw2mxISEqyO0i+1tLTI6XTGZJzR13R0dITj4uL63BO2vb1d2dnZciQlJfXZwqHT6dSQIUM0ZswYq6MMGGfOnNHp06c55wAwQNXU1CgvL08ZGRlWRwHQBUeOHFFGRoYKCwtNt3Xw4EH5fD5NmTLFdFtut1sHDhzQrFmzTLVTXV0tn8/HWDVG9u/fr5ycHGVmZlodpc9rbGzUsWPHeGxexuv1yuPxcF660blz51RfX6/8/Pw+W/vprdxut9LS0jRs2DCro/RLNTU1GjVqlFJSUqyO0uMMwzBcLpe5K/AW8Pv9amxslM3n80XMziCwSktLizZs2KDy8nKrowwYlZWVWrdunerr662OAgCwgM1m0/bt2zVz5kyrowDogtmzZ2vkyJGqqKgw3daaNWtUW1urqqoq021VVFRoxYoVCofDptopLi5WXV2d/H6/6UyQCgoKtHTpUs2bN8/qKH3epk2bVFJSokgkYnWUXiVWz31cW3V1taZOnSqv16vc3Fyr4/QrRUVFmjZtmsrKyqyO0i85HA7t3btXEyZMsDqKFV6TlG91iBu1c+dOlZaWsmYzAAAAAAAAAMA8is0AAAAAAAAAANMoNgMAAAAAAAAATKPYDAAAAAAAAAAwjWIzAAAAAAAAAMC0mBWbA4GA6uvrP3D35z179sSqSwAAAAAAAABALxGTYrPb7dbcuXP1ox/9SM8++6zmzp2rV1555Yr7eTweLVu2LBZdAgAAAAAAAAB6EYfZBtra2rRy5UrNnz9fc+bMkSTt3r1bq1ev1rhx4zR06FB5PB7t27dPO3bsiEVmAAAAAAAAAEAvY3pm86lTp+T3+zV9+vTobZMnT1Y4HJbH45Ek+f1+BQIBjR071mx3AAAAAAAAAIBeyPTM5qysLG3ZskXJycnR2zwej0KhkNLT0yVJ48eP1/jx43X48GFVVVWZ7RIAAAAAAAAA0MuYntnscrmUm5sb/dnr9WrlypWaMWOGhg8fbrZ5AAAAAAAAAEAfYHpmcyfDMLRt2zZt3bpV8+fP19y5c2PVNAAAAAAAAACgl4tJsfns2bNavHixcnJytHnzZmVnZ8eiWQAAAAAAAABAH2G62BwOh7V8+XJNnDhRJSUlstlssUkGAAAAAAAAAOgzTBebDx06pIaGBpWWluro0aOXHMvJyblk40AAAAAAAAAAQP9kuth88uRJ2e12LVy48Ipj69ev1+jRo812AQAAAAAAAADo5UwXm4uLi1VcXHxd9y0oKFBtba3ZLgEAAAAAAAAAvYzd6gAAAAAAAAAAgL6PYjMAAAAAAAAAwDSKzQAAAAAAAAAA0yg2AwAAAAAAAABMszU0NETi4+OtztElzc3NWrBggY4fP251lAEjGAwqEonI5XJZHQUAYIGmpiYlJibK6XRaHQVAFxiGoY6ODsVi/N/e3i6bzSaHw/Se42pra1MoFFJiYqKpdgzDkCSlpqaazoR3z6fdbuc1PwaCwaBaW1uVlpZmdZRepbW1VcFgkPPSjUKhkAKBgFJTU2W3M98wlgzDkM1mU0JCgtVR+qWWlhY5nc6YjDP6mo6OjnBcXFyfe8K2t7crOztbjqSkpD5bOHQ6nRoyZIjGjBljdZQB48yZMzp9+jTnHAAGqJqaGuXl5SkjI8PqKAC64MiRI8rIyFBhYaHptg4ePCifz6cpU6aYbsvtduvAgQOaNWuWqXaqq6vl8/kYq8bI/v37lZOTo8zMTKuj9HmNjY06duwYj83LeL1eeTwezks3OnfunOrr65Wfn99naz+9ldvtVlpamoYNG2Z1lH6ppqZGo0aNUkpKitVRepxhGIbL5TJ3Bd4Cfr9fjY2Nsvl8vojZGQRWaWlp0YYNG1ReXm51lAGjsrJS69atU319vdVRAAAWsNls2r59u2bOnGl1FABdMHv2bI0cOVIVFRWm21qzZo1qa2tVVVVluq2KigqtWLFC4XDYVDvFxcWqq6uT3+83nQlSQUGBli5dqnnz5lkdpc/btGmTSkpKFIlErI7Sq8TquY9rq66u1tSpU+X1epWbm2t1nH6lqKhI06ZNU1lZmdVR+iWHw6G9e/dqwoQJVkexwmuS8q0OcaN27typ0tJS1mwGAAAAAAAAAJhHsRkAAAAAAAAAYBrFZgAAAAAAAACAaRSbAQAAAAAAAACmUWwGAAAAAAAAAJjmiFVDgUBAJ06ckGEYGjFihNLT02/oOAAAAAAAAACg74pJsdntdmvJkiXKyspSSkqK3G63Fi1apKKious6DgAAAAAAAADo20wXm9va2rRy5UrNnz9fc+bMkSTt3r1bq1ev1rhx45ScnPy+x4cOHWr+twAAAAAAAAAAWMr0ms2nTp2S3+/X9OnTo7dNnjxZ4XBYHo/nA48DAAAAAAAAAPo+0zObs7KytGXLFiUnJ0dv83g8CoVCSk9PV2Zm5vseBwAAAAAAAAD0faaLzS6XS7m5udGfvV6vVq5cqRkzZmj48OGS9IHHAQAAAAAAAAB9W0w2CJQkwzC0bds2bd26VfPnz9fcuXNv6DgAAAAAAAAAoO+KSbH57NmzWrx4sXJycrR582ZlZ2ff0HEAAAAAAAAAQN9mutgcDoe1fPlyTZw4USUlJbLZbDd0HAAAAAAAAADQ95kuNh86dEgNDQ0qLS3V0aNHLzmWk5Oj119//X2PX7xxIAAAAAAAAACgbzJdbD558qTsdrsWLlx4xbH169d/4PHRo0ebjQAAAAAAAAAAsJjpYnNxcbGKi4uveXz06NHvexwAAAAAAAAA0PfZrQ4AAAAAAAAAAOj7KDYDAAAAAAAAAEyj2AwAAAAAAAAAMI1iMwAAAAAAAADANFtDQ0MkPj7e6hxd0tzcrAULFuj48eNWRxkwgsGgIpGIXC6X1VEAABZoampSYmKinE6n1VEAdIFhGOro6FAsxv/t7e2y2WxyOEzvOa62tjaFQiElJiaaascwDElSamqq6Ux493za7XZe82MgGAyqtbVVaWlpVkfpVVpbWxUMBjkv3SgUCikQCCg1NVV2O/MNY8kwDNlsNiUkJFgdpV9qaWmR0+mMyTijr+no6AjHxcX1uSdse3u7srOz5UhKSuqzhUOn06khQ4ZozJgxVkcZMM6cOaPTp09zzgFggKqpqVFeXp4yMjKsjgKgC44cOaKMjAwVFhaabuvgwYPy+XyaMmWK6bbcbrcOHDigWbNmmWqnurpaPp+PsWqM7N+/Xzk5OcrMzLQ6Sp/X2NioY8eO8di8jNfrlcfj4bx0o3Pnzqm+vl75+fl9tvbTW7ndbqWlpWnYsGFWR+mXampqNGrUKKWkpFgdpccZhmG4XC5zV+At4Pf71djYKJvP54uYnUFglZaWFm3YsEHl5eVWRxkwKisrtW7dOtXX11sdBQBgAZvNpu3bt2vmzJlWRwHQBbNnz9bIkSNVUVFhuq01a9aotrZWVVVVptuqqKjQihUrFA6HTbVTXFysuro6+f1+05kgFRQUaOnSpZo3b57VUfq8TZs2qaSkRJFIxOoovUqsnvu4turqak2dOlVer1e5ublWx+lXioqKNG3aNJWVlVkdpV9yOBzau3evJkyYYHUUK7wmKd/qEDdq586dKi0tZc1mAAAAAAAAAIB5FJsBAAAAAAAAAKZRbAYAAAAAAAAAmEaxGQAAAAAAAABgGsVmAAAAAAAAAIBpjlg1FAgEdOLECRmGoREjRig9Pf2S4//3f/+n119/XfHx8froRz+q5OTkWHUNAAAAAAAAALBYTIrNbrdbS5YsUVZWllJSUuR2u7Vo0SIVFRVJkqqqqrRu3Tp97GMf0+nTpxUMBrV27VrdfvvtsegeAAAAAAAAAGAx08XmtrY2rVy5UvPnz9ecOXMkSbt379bq1as1btw4xcfH64c//KEef/xxfeITn1A4HNbatWv1s5/9TOvWrYvF7wAAAAAAAAAAsJjpNZtPnTolv9+v6dOnR2+bPHmywuGwPB6PGhsblZiYqLvuuuvdDu12TZw4UV6v12zXAAAAAAAAAIBewvTM5qysLG3ZsuWSNZg9Ho9CoZDS09M1fPhw7dixI3qsvb1d+/btixafAQAAAAAAAAB9n+lis8vlUm5ubvRnr9erlStXasaMGRo+fPgl9129erX27Nmjm2++WRs2bDDbclkbuQAAIABJREFUNQAAAAAAAACglzC9jEYnwzD0zDPP6JFHHtGnP/1pfetb37riPg8//LDWrl2rzMxMPfHEE7HqGgAAAAAAAABgMdMzmyXp7NmzWrx4sXJycrR582ZlZ2dHj4XDYem9tZqzs7Oj/2bPnq233npLGRkZsYgAAAAAAAAAALCQ6ZnN4XBYy5cv18SJE/X4449fUmiWpJdeeknf+MY3LrnN5XJJki5cuGC2ewAAAAAAAABAL2C62Hzo0CE1NDSosLBQR48e1ZEjR6L/AoGARo0apYMHD8rtdkuSIpGIXnrpJWVlZSknJycWvwMAAAAAAAAAwGKml9E4efKk7Ha7Fi5ceMWx9evXa/To0Vq2bJm++c1v6sMf/rCam5slSatWrVJ8fLzZ7gEAAAAAAAAAvYDpYnNxcbGKi4vf9z733XefJk2aJI/Ho/j4eA0bNkwJCQlmuwYAAAAAAAAA9BIx2SDwegwaNEj5+fk91R0AAAAAAAAAoAeZXrMZAAAAAAAAAACKzQAAAAAAAAAA0yg2AwAAAAAAAABMszU0NETi4+OtztElzc3NWrBggY4fP251lAEjGAwqEonI5XJZHQUAYIGmpiYlJibK6XRaHQVAFxiGoY6ODsVi/N/e3i6bzSaHw/w2MG1tbQqFQkpMTDTVjmEYkqTU1FTTmfDu+bTb7bzmx0AwGFRra6vS0tKsjtKrtLa2KhgMcl66USgUUiAQUGpqqux25hvGkmEYstlsSkhIsDpKv9TS0iKn0xmTcUZf09HREY6Li+tzT9j29nZlZ2fLkZSU1GcLh06nU0OGDNGYMWOsjjJgnDlzRqdPn+acA8AAVVNTo7y8PGVkZFgdBUAXHDlyRBkZGSosLDTd1sGDB+Xz+TRlyhTTbbndbh04cECzZs0y1U51dbV8Ph9j1RjZv3+/cnJylJmZaXWUPq+xsVHHjh3jsXkZr9crj8fDeelG586dU319vfLz8/ts7ae3crvdSktL07Bhw6yO0i/V1NRo1KhRSklJsTpKjzMMw3C5XOauwFvA7/ersbFRNp/PFzE7g8AqLS0t2rBhg8rLy62OMmBUVlZq3bp1qq+vtzoKAMACNptN27dv18yZM62OAqALZs+erZEjR6qiosJ0W2vWrFFtba2qqqpMt1VRUaEVK1YoHA6baqe4uFh1dXXy+/2mM0EqKCjQ0qVLNW/ePKuj9HmbNm1SSUmJIpGI1VF6lVg993Ft1dXVmjp1qrxer3Jzc62O068UFRVp2rRpKisrszpKv+RwOLR3715NmDDB6ihWeE1SvtUhbtTOnTtVWlrKms0AAAAAAAAAAPMoNgMAAAAAAAAATKPYDAAAAAAAAAAwjWIzAAAAAAAAAMA0is0AAAAAAAAAANNiVmwOBAKqr6//wN2fDcPQH/7wh1h1CwAAAAAAAADoBRyxaMTtdmvJkiXKyspSSkqK3G63Fi1apKKioivuu3nzZj3//POaMmVKLLoGAAAAAAAAAPQCpovNbW1tWrlypebPn685c+ZIknbv3q3Vq1dr3LhxGjp0aPS+r776qn79618rLi7ObLcAAAAAAAAAgF7E9DIap06dkt/v1/Tp06O3TZ48WeFwWB6PJ3pbU1OTvv/976u0tNRslwAAAAAAAACAXsb0zOasrCxt2bJFycnJ0ds8Ho9CoZDS09MlSZFIROvXr9ecOXOUl5dntksAAAAAAAAAQC9jemazy+VSbm5u9Gev16uVK1dqxowZGj58uCTp5ZdfViAQ0Oc+9zmz3QEAAAAAAAAAeqGYbBAoSYZhaNu2bdq6davmz5+vuXPnSpJOnjypjRs36ic/+YnsdtO1bQAAAAAAAABALxSTYvPZs2e1ePFi5eTkaPPmzcrOzo4eq6qqUmpqqn75y19Kknw+n4LBoH74wx/qk5/8pO68885YRAAAAAAAAAAAWMh0sTkcDmv58uWaOHGiSkpKZLPZLjk+YcIEZWRkRH+Oi4uT3W5XXl6eUlNTzXYPAAAAAAAAAOgFTBebDx06pIaGBpWWluro0aOXHMvJydEdd9yhO+64I3rbX//6V+3atUv333+/2a4BAAAAAAAAAL2E6WLzyZMnZbfbtXDhwiuOrV+/XqNHjzbbBQAAAAAAAACglzNdbC4uLlZxcfF133/s2LF6+eWXzXYLAAAAAAAAAOhF7FYHAAAAAAAAAAD0fRSbAQAAAAAAAACmUWwGAAAAAAAAAJhGsRkAAAAAAAAAYJqtoaEhEh8fb3WOLmlubtaCBQt0/Phxq6MMGMFgUJFIRC6Xy+ooAAALNDU1KTExUU6n0+ooALrAMAx1dHQoFuP/9vZ22Ww2ORym9xxXW1ubQqGQEhMTTbVjGIYkKTU11XQmvHs+7XY7r/kxEAwG1draqrS0NKuj9Cqtra0KBoOcl24UCoUUCASUmpoqu535hrFkGIZsNpsSEhKsjtIvtbS0yOl0xmSc0dd0dHSE4+Li+twTtr29XdnZ2XIkJSX12cKh0+nUkCFDNGbMGKujDBhnzpzR6dOnOecAMEDV1NQoLy9PGRkZVkcB0AVHjhxRRkaGCgsLTbd18OBB+Xw+TZkyxXRbbrdbBw4c0KxZs0y1U11dLZ/Px1g1Rvbv36+cnBxlZmZaHaXPa2xs1LFjx3hsXsbr9crj8XBeutG5c+dUX1+v/Pz8Plv76a3cbrfS0tI0bNgwq6P0SzU1NRo1apRSUlKsjtLjDMMwXC6XuSvwFvD7/WpsbJTN5/NFzM4gsEpLS4s2bNig8vJyq6MMGJWVlVq3bp3q6+utjgIAsIDNZtP27ds1c+ZMq6MA6ILZs2dr5MiRqqioMN3WmjVrVFtbq6qqKtNtVVRUaMWKFQqHw6baKS4uVl1dnfx+v+lMkAoKCrR06VLNmzfP6ih93qZNm1RSUqJIJGJ1lF4lVs99XFt1dbWmTp0qr9er3Nxcq+P0K0VFRZo2bZrKysqsjtIvORwO7d27VxMmTLA6ihVek5RvdYgbtXPnTpWWlrJmMwAAAAAAAADAPIrNAAAAAAAAAADTKDYDAAAAAAAAAEyj2AwAAAAAAAAAMI1iMwAAAAAAAADANEesGgoEAjpx4oQMw9CIESOUnp4ePXb+/HmdO3fukvvHx8crOzs7Vt0DAAAAAAAAACwUk2Kz2+3WkiVLlJWVpZSUFLndbi1atEhFRUWSpP/5n//RT37yk0v+n/z8fD399NOx6B4AAAAAAAAAYDHTxea2tjatXLlS8+fP15w5cyRJu3fv1urVqzVu3DgNHTpUp0+fVklJSfS4JNlsNrNdAwAAAAAAAAB6CdNrNp86dUp+v1/Tp0+P3jZ58mSFw2F5PB5J0smTJzVs2DC5XK7ov4SEBLNdAwAAAAAAAAB6CdMzm7OysrRlyxYlJydHb/N4PAqFQtF1m19//XW9/fbb2rBhgzo6OlRYWKhJkyYxuxkAAAAAAAAA+gnTM5tdLpdyc3OjP3u9Xq1cuVIzZszQ8OHDFQgE5Pf7VVVVpVtuuUU33XSTnnrqKf30pz812zUAAAAAAAAAoJeIyQaBkmQYhrZt26atW7dq/vz5mjt3riTJbrdr7dq1uuOOO6Kzn8eMGaPS0lLNmjVLN998c6wiAAAAAAAAAAAsEpNi89mzZ7V48WLl5ORo8+bNys7Ojh5LTEzU3Xfffcn98/PzlZycrDfeeINiMwAAAAAAAAD0A6aLzeFwWMuXL9fEiRNVUlJyxTrMf/vb33T06FF99rP/j737D2+qvvs//kpI09DS2tJqqS1FqIOWH6tFBTYYG1JwHWMTxobbYIyhW2+3Wzc3QdCCA4p1epWxS+7d0w3QG6fODSZU7ebWWdxg6yU6miI/yiBRbJBQWyRtT5o0+f6h5rv6YyoneNryfFxXrsuec/L+vHLaxOM7Hz/nc7Ft4XBYhmFo8ODBZocHAAAAAAAAAPQCppvN+/fvV1NTk8rKynTo0KEe+3JzczVgwABVVlbK5XJp2rRpkqTHHntMw4YN07Bhw8wODwAAAAAAAADoBUw3m48fPy673a5bbrnlHfs2bNigcePG6cc//rHuvPNO/eIXv1A0GlVycrJWrlyphIQEs8MDAAAAAAAAAHoB083m0tJSlZaW/sdjpk2bpkmTJsnr9SohIUF5eXk0mgEAAAAAAACgH4nLDQI/iIEDB6qgoOCjGg4AAAAAAAAA8BGyWx0AAAAAAAAAAND30WwGAAAAAAAAAJhGsxkAAAAAAAAAYJqtqakp2ldv1tfe3q5FixbpyJEjVkc5bwSDQUWjUblcLqujAAAs0NbWpqSkJDmdTqujADgLhmGou7s7LjfrDoVCstlscjjM3wamq6tL4XBYSUlJpuoYhiFJSk1NNZ0Jb5xPu93OZ34cBINBdXZ2Ki0tzeoovUpnZ6eCwSDn5RwKh8MKBAJKTU2V3c58w3gyDEM2m02JiYlWR+mXOjo65HQ643Kd0dd0d3dHBgwY0OfesKFQSDk5OXIkJyf32cah0+lUenq6ioqKrI5y3jhx4oSam5s55wBwnqqrq9Pw4cOVmZlpdRQAZ+HgwYPKzMxUcXGx6VqNjY3y+/2aNm2a6Vput1v79u3T3LlzTdWpra2V3+/nWjVO9u7dq9zcXGVlZVkdpc/z+Xw6fPgwf5tv4/V65fF4OC/nUGtrqxoaGlRYWNhnez+9ldvtVlpamoYOHWp1lH6prq5OY8eOVUpKitVRPnKGYRgul8vcN/AWaGlpkc/nk83v90fNziCwSkdHhzZu3KhVq1ZZHeW8sWXLFlVVVamhocHqKAAAC9hsNm3btk1z5syxOgqAszBv3jyNGjVKFRUVpmutW7dOu3btUk1NjelaFRUVKi8vVyQSMVWntLRU9fX1amlpMZ0J0ujRo7VixQotWLDA6ih93qZNm7RkyRJFo1Gro/Qq8Xrv473V1tZq+vTp8nq9ysvLszpOv1JSUqKZM2dq6dKlVkfplxwOh3bv3q0JEyZYHcUKByQVWh3iw9q5c6fKyspYsxkAAAAAAAAAYB7NZgAAAAAAAACAaTSbAQAAAAAAAACm0WwGAAAAAAAAAJhGsxkAAAAAAAAAYJojXoUCgYCOHj0qwzCUn5+vjIyMHvsjkYg8Ho9effVVjRw58h37AQAAAAAAAAB9V1yazW63W8uXL1d2drZSUlLkdru1bNkylZSUSJLa29u1evVqeb1epaamyuPxqKKiQldeeWU8hgcAAAAAAAAAWMx0s7mrq0srV67UwoULNX/+fEnSM888o7Vr12r8+PEaPHiwtmzZIpfLpa1bt8rhcOiRRx7RT37yEz366KOy21nJAwAAAAAAAAD6OtPN5ldeeUUtLS2aNWtWbNvUqVNjy2a4XC5VV1drw4YNcjjeGO6aa65RQUGBwuGwnE6n2QgAAAAAAAAAAIuZbjZnZ2dr69atGjRoUGybx+NROBxWRkaGmpub1dHRoby8PB04cECvvfaaLr30Ul122WVmhwYAAAAAAAAA9BKmm80ul0t5eXmxn71er1auXKnZs2dr2LBhev7555WcnKzVq1fr9OnTGjhwoP75z39q2bJlmjFjhtnhAQAAAAAAAAC9QFxuEChJhmHo0Ucf1UMPPaSFCxfq2muvld5c0zkQCOjyyy/Xl770JUnSs88+qzVr1uiKK65Qenp6vCIAAAAAAAAAACwSl2bzq6++qltvvVW5ubnavHmzcnJyYvtSU1MlSZ/73Odi2yZPniybzaZjx47RbAYAAAAAAACAfsB0szkSiej222/X5MmTtWTJEtlsth77L7zwQunNGc4DBw6MPae7uzv2MwAAAAAAAACgb7ObLbB//341NTWpuLhYhw4d0sGDB2OPQCCgCy+8UJ/5zGf00EMPKRKJSJK2bdumzMxMDR8+PB6vAQAAAAAAAABgMdMzm48fPy673a5bbrnlHfs2bNigcePG6Xvf+56WLVumv/zlL0pMTFQoFNIdd9whl8tldngAAAAAAAAAQC9gutlcWlqq0tLS/3jMRRddpPvvv19er1fd3d3Ky8uj0QwAAAAAAAAA/UhcbhD4gQZyOJSfn/9RDQcAAAAAAAAA+AiZXrMZAAAAAAAAAACazQAAAAAAAAAA02g2AwAAAAAAAABMszU1NUUTEhKsznFW2tvbtWjRIh05csTqKOeNYDCoaDTKDR4B4DzV1tampKQkOZ1Oq6MAOAuGYai7u1vxuP4PhUKy2WxyOMzfBqarq0vhcFhJSUmm6hiGIUlKTU01nQlvnE+73c5nfhwEg0F1dnYqLS3N6ii9Smdnp4LBIOflHAqHwwoEAkpNTZXdznzDeDIMQzabTYmJiVZH6Zc6OjrkdDrjcp3R13R3d0cGDBjQ596woVBIOTk5ciQnJ/fZxqHT6VR6erqKioqsjnLeOHHihJqbmznnAHCeqqur0/Dhw5WZmWl1FABn4eDBg8rMzFRxcbHpWo2NjfL7/Zo2bZrpWm63W/v27dPcuXNN1amtrZXf7+daNU727t2r3NxcZWVlWR2lz/P5fDp8+DB/m2/j9Xrl8Xg4L+dQa2urGhoaVFhY2Gd7P72V2+1WWlqahg4danWUfqmurk5jx45VSkqK1VE+coZhGC6Xy9w38BZoaWmRz+eTze/3R83OILBKR0eHNm7cqFWrVlkd5byxZcsWVVVVqaGhweooAAAL2Gw2bdu2TXPmzLE6CoCzMG/ePI0aNUoVFRWma61bt067du1STU2N6VoVFRUqLy9XJBIxVae0tFT19fVqaWkxnQnS6NGjtWLFCi1YsMDqKH3epk2btGTJEkWjUauj9Crxeu/jvdXW1mr69Onyer3Ky8uzOk6/UlJSopkzZ2rp0qVWR+mXHA6Hdu/erQkTJlgdxQoHJBVaHeLD2rlzp8rKylizGQAAAAAAAABgHs1mAAAAAAAAAIBpNJsBAAAAAAAAAKbRbAYAAAAAAAAAmEazGQAAAAAAAABgmiNehQKBgI4ePSrDMJSfn6+MjIzY9lOnTr3rc9LS0pSWlhavCAAAAAAAAAAAi8Sl2ex2u7V8+XJlZ2crJSVFbrdby5YtU0lJifbv369bbrnlXZ93880365prrolHBAAAAAAAAACAhUw3m7u6urRy5UotXLhQ8+fPlyQ988wzWrt2rcaPH68rrrhCf/zjH3s8x+fzaenSpZo6darZ4QEAAAAAAAAAvYDpNZtfeeUVtbS0aNasWbFtU6dOVSQSkcfj0YABA+RyuWIPp9Op//mf/9FNN92kwYMHmx0eAAAAAAAAANALmG42Z2dna+vWrRo0aFBsm8fjUTgcjq3b/O+efvppuVwuTZkyxezQAAAAAAAAAIBewnSz2eVyKS8vL/az1+vVypUrNXv2bA0bNqzHsZ2dnbr//vv1zW9+0+ywAAAAAAAAAIBeJC43CJQkwzD06KOP6qGHHtLChQt17bXXvuOYJ554QkVFRcrPz4/XsAAAAAAAAACAXiAuzeZXX31Vt956q3Jzc7V582bl5OS845ju7m797ne/0w9+8IN4DAkAAAAAAAAA6EVMN5sjkYhuv/12TZ48WUuWLJHNZnvX4/bt26fOzk6NHz/e7JAAAAAAAAAAgF7GdLN5//79ampqUllZmQ4dOtRjX25ubuzGgfX19Zo4caIcjrit3AEAAAAAAAAA6CVMd36PHz8uu92uW2655R37NmzYoHHjxkmSdu/era985StmhwMAAAAAAAAA9EKmm82lpaUqLS193+MefPBBs0MBAAAAAAAAAHopu9UBAAAAAAAAAAB9H81mAAAAAAAAAIBpNJsBAAAAAAAAAKbRbAYAAAAAAAAAmGZramqKJiQkWJ3jrLS3t2vRokU6cuSI1VHOG8FgUNFoVC6Xy+ooAAALtLW1KSkpSU6n0+ooAM6CYRjq7u5WPK7/Q6GQbDabHA7T9xxXV1eXwuGwkpKSTNUxDEOSlJqaajoT3jifdrudz/w4CAaD6uzsVFpamtVRepXOzk4Fg0HOyzkUDocVCASUmpoqu535hvFkGIZsNpsSExOtjtIvdXR0yOl0xuU6o6/p7u6ODBgwoM+9YUOhkHJycuRITk7us41Dp9Op9PR0FRUVWR3lvHHixAk1NzdzzgHgPFVXV6fhw4crMzPT6igAzsLBgweVmZmp4uJi07UaGxvl9/s1bdo007Xcbrf27dunuXPnmqpTW1srv9/PtWqc7N27V7m5ucrKyrI6Sp/n8/l0+PBh/jbfxuv1yuPxcF7OodbWVjU0NKiwsLDP9n56K7fbrbS0NA0dOtTqKP1SXV2dxo4dq5SUFKujfOQMwzBcLpe5b+At0NLSIp/PJ5vf74+anUFglY6ODm3cuFGrVq2yOsp5Y8uWLaqqqlJDQ4PVUQAAFrDZbNq2bZvmzJljdRQAZ2HevHkaNWqUKioqTNdat26ddu3apZqaGtO1KioqVF5erkgkYqpOaWmp6uvr1dLSYjoTpNGjR2vFihVasGCB1VH6vE2bNmnJkiWKRqNWR+lV4vXex3urra3V9OnT5fV6lZeXZ3WcfqWkpEQzZ87U0qVLrY7SLzkcDu3evVsTJkywOooVDkgqtDrEh7Vz506VlZWxZjMAAAAAAAAAwDyazQAAAAAAAAAA02g2AwAAAAAAAABMo9kMAAAAAAAAADCNZjMAAAAAAAAAwDRHvAoFAgEdPXpUhmEoPz9fGRkZPfa3tbWpqalJAwYM0MiRIzVo0KB4DQ0AAAAAAAAAsFhcms1ut1vLly9Xdna2UlJS5Ha7tWzZMpWUlEiS6uvrdccddygrK0uGYaizs1OVlZUqKCiIx/AAAAAAAAAAAIuZbjZ3dXVp5cqVWrhwoebPny9JeuaZZ7R27VqNHz9egwcP1k9+8hPNnz9f3/jGNxSNRvWzn/1M9957r+699954vAYAAAAAAAAAgMVMr9n8yiuvqKWlRbNmzYptmzp1qiKRiDwejwzDkN/v12WXXSabzSa73a7LL79cHo/H7NAAAAAAAAAAgF7C9Mzm7Oxsbd26tccazB6PR+FwWBkZGXK5XJo6daoef/xxXXzxxQqHw9qxY0dsiQ0AAAAAAAAAQN9nutnscrmUl5cX+9nr9WrlypWaPXu2hg0bJklaunSpvvGNb+hLX/qSJOmSSy7R2rVrzQ4NAAAAAAAAAOglTC+j8RbDMPTAAw/o+uuv19VXX63vf//70ptrOpeXl+vTn/60fv3rX+vBBx/Uxz72MVVUVCgajcZreAAAAAAAAACAhUzPbJakV199Vbfeeqtyc3O1efNm5eTkxPa9+OKLOnr0qO655x45HG8Md/PNN2vWrFnyer265JJL4hEBAAAAAAAAAGAh083mSCSi22+/XZMnT9aSJUtks9l67G9tbdWgQYNijWZJGjhwoAYOHKi2tjazwwMAAAAAAAAAegHTy2js379fTU1NKi4u1qFDh3Tw4MHYIxAIaPTo0fL7/dq1a1fsOU888YScTqcuvfRSs8MDAAAAAAAAAHoB0zObjx8/LrvdrltuueUd+zZs2KBx48Zp1apVuuuuu7R582YFg0F1dXXpxz/+sQYNGmR2eAAAAAAAAABAL2C62VxaWqrS0tL/eMyUKVNUXFysl19+WXa7XXl5eXK5XGaHBgAAAAAAAAD0EnG5QeAHkZycrIKCgo9qOAAAAAAAAADAR8j0ms0AAAAAAAAAANBsBgAAAAAAAACYRrMZAAAAAAAAAGCarampKZqQkGB1jrPS3t6uRYsW6ciRI1ZHOW8Eg0FFo1Fu8AgA56m2tjYlJSXJ6XRaHQXAWTAMQ93d3YrH9X8oFJLNZpPDYf42MF1dXQqHw0pKSjJVxzAMSVJqaqrpTHjjfNrtdj7z4yAYDKqzs1NpaWlWR+lVOjs7FQwGOS/nUDgcViAQUGpqqux25hvGk2EYstlsSkxMtDpKv9TR0SGn0xmX64y+pru7OzJgwIA+94YNhULKycmRIzk5uc82Dp1Op9LT01VUVGR1lPPGiRMn1NzczDkHgPNUXV2dhg8frszMTKujADgLBw8eVGZmpoqLi03XamxslN/v17Rp00zXcrvd2rdvn+bOnWuqTm1trfx+P9eqcbJ3717l5uYqKyvL6ih9ns/n0+HDh/nbfBuv1yuPx8N5OYdaW1vV0NCgwsLCPtv76a3cbrfS0tI0dOhQq6P0S3V1dRo7dqxSUlKsjvKRMwzDcLlc5r6Bt0BLS4t8Pp9sfr8/anYGgVU6Ojq0ceNGrVq1yuoo540tW7aoqqpKDQ0NVkcBAFjAZrNp27ZtmjNnjtVRAJyFefPmadSoUaqoqDBda926ddq1a5dqampM16qoqFB5ebkikYipOqWlpaqvr1dLS4vpTJBGjx6tFStWaMGCBVZH6fM2bdqkJUuWKBqNWh2lV4nXex/vrba2VtOnT5fX61VeXp7VcfqVkpISzZw5U0uXLrU6Sr/kcDi0e/duTZgwweooVjggqdDqEB/Wzp07VVZWxprNAAAAAAAAAADzaDYDAAAAAAAAAEyj2QwAAAAAAAAAMI1mMwAAAAAAAADANJrNAAAAAAAAAADTHPEqFAgEdPToURmGofz8fGVkZPTY39bWpsOHDys5OVn5+flyuVzxGhoAAAAAAAAAYLG4NJvdbreWL1+u7OxspaSkyO12a9myZSopKZEk/fWvf9W6detUWFiotrY2hcNh3XXXXRoyZEg8hgcAAAAAAAAAWMx0s7mrq0srV67UwoULNX/+fEnSM888o7Vr12r8+PGSpIqKCt12222aMmWKIpGI/vd//1f333+/ysvLzb8CAAAAAAAAAIDlTK/Z/Morr6ilpUWzZs2KbZs6daoikYg8Ho8OHTqk5ORkTZky5Y0B7XbNmTNHf/rTn9TW1mZ2eAAAAAAAAABAL2C62Zydna2tW7dq0KBBsW0ej0fhcFgZGRkKh8OKRqM9B7XbFY1G1dLSYnZ4AAAAAAAAAEAvYLrZ7HK5lJeXF/vZ6/V87h0FAAAgAElEQVRq5cqVmj17toYNG6ZRo0bp9OnTqqurkySFw2H95je/kSS9/vrrZocHAAAAAAAAAPQCcblBoCQZhqFHH31UDz30kBYuXKhrr71WknTRRRdp+fLlqqio0AMPPKDTp0/r05/+tCRp4MCB8RoeAAAAAAAAAGChuDSbX331Vd16663Kzc3V5s2blZOT02P/9OnTdcUVV+ill15SZmamEhMT9dvf/lZpaWnxGB4AAAAAAAAAYDHTzeZIJKLbb79dkydP1pIlS2Sz2XrsDwQCeuGFF3TllVdq3LhxkqQ9e/YoKytLmZmZZocHAAAAAAAAAPQCptds3r9/v5qamlRcXKxDhw7p4MGDsUcgEFBiYqLuvvtu7dixQ9FoVIFAQA8++KDmzp0rhyNuq3gAAAAAAAAAACxkutt7/Phx2e123XLLLe/Yt2HDBo0bN04rVqzQqlWrtH37dr322muaMWOGvvzlL5sdGgAAAAAAAADQS5huNpeWlqq0tPQ/HjNp0iQ9+uijevnll3XxxRcrIyPD7LAAAAAAAAAAgF7kI1vHIi0tjRsCAgAAAAAAAEA/ZXrNZgAAAAAAAAAAaDYDAAAAAAAAAEyj2QwAAAAAAAAAMM3W1NQUTUhIsDrHWWlvb9eiRYt05MgRq6OcN4LBoKLRqFwul9VRAAAWaGtrU1JSkpxOp9VRAJwFwzDU3d2teFz/h0Ih2Ww2ORzmbwPT1dWlcDispKQkU3UMw5Akpaamms6EN86n3W7nMz8OgsGgOjs7uY/R23R2dioYDHJezqFwOKxAIKDU1FTZ7cw3jCfDMGSz2ZSYmGh1lH6po6NDTqczLtcZfU13d3dkwIABfe4NGwqFlJOTI0dycnKfbRw6nU6lp6erqKjI6ijnjRMnTqi5uZlzDgDnqbq6Og0fPlyZmZlWRwFwFg4ePKjMzEwVFxebrtXY2Ci/369p06aZruV2u7Vv3z7NnTvXVJ3a2lr5/X6uVeNk7969ys3NVVZWltVR+jyfz6fDhw/zt/k2Xq9XHo+H83IOtba2qqGhQYWFhX2299Nbud1upaWlaejQoVZH6Zfq6uo0duxYpaSkWB3lI2cYhuFyucx9A2+BlpYW+Xw+2fx+f9TsDAKrdHR0aOPGjVq1apXVUc4bW7ZsUVVVlRoaGqyOAgCwgM1m07Zt2zRnzhyrowA4C/PmzdOoUaNUUVFhuta6deu0a9cu1dTUmK5VUVGh8vJyRSIRU3VKS0tVX1+vlpYW05kgjR49WitWrNCCBQusjtLnbdq0SUuWLFE0GrU6Sq8Sr/c+3lttba2mT58ur9ervLw8q+P0KyUlJZo5c6aWLl1qdZR+yeFwaPfu3ZowYYLVUaxwQFKh1SE+rJ07d6qsrIw1mwEAAAAAAAAA5tFsBgAAAAAAAACYRrMZAAAAAAAAAGAazWYAAAAAAAAAgGk0mwEAAAAAAAAApjniVejkyZM6duyYEhISNHLkSA0aNKjH/o6ODh06dEgOh0Mf+9jH5HK54jU0AAAAAAAAAMBicWk219TUqKqqSmPGjFFzc7OCwaAqKytVUFAgSXr55Ze1bNkyJSUlKRQKyeFwqLKyUhdeeGE8hgcAAAAAAAAAWMz0MhpnzpzR+vXrtXr1aq1fv14PP/ywJkyYoPvuuy92zJYtW3TllVfql7/8pTZv3qyhQ4fqscceMzs0AAAAAAAAAKCXMN1s9vl8SkpK0sSJE98oaLdr8uTJ8nq9kqTW1lb96U9/0jXXXBPb/8UvflG///3vZRiG2eEBAAAAAAAAAL2A6WbzyJEjtX37dtlsNklSKBTSnj17Ys3nU6dOSZKGDRsWe87w4cNlGIZOnz5tdngAAAAAAAAAQC8QtxsEStLatWv17LPPasiQIdq4caMkqa2tTenp6bLb/39fOzk5WZJ0+vRpZWVlxTMCAAAAAAAAAMACpmc2/7vFixersrJSWVlZWrNmjSQpGo0qGo32OO6tnyORSDyHBwAAAAAAAABYxPTM5rcaxna7XTk5ObHHvHnzdOrUKaWlpamtrU2RSCQ2u7m9vV2SlJaWZnZ4AAAAAAAAAEAvYHpmc3V1tW688cYe21wulyTpzJkzysrK0oABA+TxeGL7jx07ppSUFGVkZJgdHgAAAAAAAADQC5huNo8dO1aNjY1yu93Sm0tkVFdXKzs7W7m5ubrgggs0ffp07dixQ3pzJvTjjz+uWbNmKSEhwfwrAAAAAAAAAABYzvQyGiNGjNBtt92mm266SSNGjIgtkXHHHXfEmsllZWX60Y9+pOuvv17BYFCDBw/Wj370I9PhAQAAAAAAAAC9g+lmsyTNmDFDU6ZMkcfjUUJCgoYOHarExMTY/oyMDN1///3yeDxyOBzKy8uLrd8MAAAAAAAAAOj74tJslqSBAweqsLDwvQdyOHTppZfGazgAAAAAAAAAQC/C9GIAAAAAAAAAgGk0mwEAAAAAAAAAptFsBgAAAAAAAACYZmtqaoomJCRYneOstLe3a9GiRTpy5IjVUc4bwWBQ0WhULpfL6igAAAu0tbUpKSlJTqfT6igAzoJhGOru7lY8rv9DoZBsNpscDvO3genq6lI4HFZSUpKpOoZhSJJSU1NNZ8Ib59Nut/OZHwfBYFCdnZ1KS0uzOkqv0tnZqWAwyHk5h8LhsAKBgFJTU2W3M98wngzDkM1mU2JiotVR+qWOjg45nc64XGf0Nd3d3ZEBAwb0uTdsKBRSTk6OHMnJyX22ceh0OpWenq6ioiKro5w3Tpw4oebmZs45AJyn6urqNHz4cGVmZlodBcBZOHjwoDIzM1VcXGy6VmNjo/x+v6ZNm2a6ltvt1r59+zR37lxTdWpra+X3+7lWjZO9e/cqNzdXWVlZVkfp83w+nw4fPszf5tt4vV55PB7OyznU2tqqhoYGFRYW9tneT2/ldruVlpamoUOHWh2lX6qrq9PYsWOVkpJidZSPnGEYhsvlMvcNvAVaWlrk8/lk8/v9UbMzCKzS0dGhjRs3atWqVVZHOW9s2bJFVVVVamhosDoKAMACNptN27Zt05w5c6yOAuAszJs3T6NGjVJFRYXpWuvWrdOuXbtUU1NjulZFRYXKy8sViURM1SktLVV9fb1aWlpMZ4I0evRorVixQgsWLLA6Sp+3adMmLVmyRNFo1OoovUq83vt4b7W1tZo+fbq8Xq/y8vKsjtOvlJSUaObMmVq6dKnVUfolh8Oh3bt3a8KECVZHscIBSYVWh/iwdu7cqbKyMtZsBgAAAAAAAACYR7MZAAAAAAAAAGAazWYAAAAAAAAAgGk0mwEAAAAAAAAAptFsBgAAAAAAAACYFrdm88mTJ/WPf/xDzz//vAKBwHse9+yzz8ZrSAAAAAAAAABAL+GIR5GamhpVVVVpzJgxam5uVjAYVGVlpQoKCnoc5/F4dNttt2nXrl3xGBYAAAAAAAAA0EuYbjafOXNG69ev1+rVqzVp0iRFIhFVVlbqvvvuU1VVlfRmk3nPnj3avn17PDIDAAAAAAAAAHoZ08to+Hw+JSUlaeLEiW8UtNs1efJkeb3e2DEtLS0KBAK67LLLzA4HAAAAAAAAAOiFTM9sHjlyZI8Zy6FQSHv27Ik1nyXp8ssv1+WXX64XX3xRNTU1ZocEAAAAAAAAAPQycVmz+S1r167Vs88+qyFDhmjjxo3xLA0AAAAAAAAA6MVML6Px7xYvXqzKykplZWVpzZo18SwNAAAAAAAAAOjFTM9sjkQi0ptrNefk5MQe8+bN06lTp5SZmRmPnAAAAAAAAACAXsz0zObq6mrdeOONPba5XC5J0pkzZ8yWBwAAAAAAAAD0AaabzWPHjlVjY6PcbrckKRqNqrq6WtnZ2crNzY1HRgAAAAAAAABAL2d6GY0RI0botttu00033aQRI0aovb1dknTHHXcoISEhDhEBAAAAAAAAAL2d6WazJM2YMUNTpkyRx+NRQkKChg4dqsTExHccN3r0aO3atSseQwIAAAAAAAAAepG4NJslaeDAgSosLIxXOQAAAAAAAABAH2J6zWYAAAAAAAAAAGg2AwAAAAAAAABMo9kMAAAAAAAAADDN1tTUFE1ISLA6x1lpb2/XokWLdOTIEaujnDeCwaCi0ahcLpfVUQAAFmhra1NSUpKcTqfVUQCcBcMw1N3drXhc/4dCIdlsNjkc5m8D09XVpXA4rKSkJFN1DMOQJKWmpprOhDfOp91u5zM/DoLBoDo7O5WWlmZ1lF6ls7NTwWCQ83IOhcNhBQIBpaamym5nvmE8GYYhm82mxMREq6P0Sx0dHXI6nXG5zuhruru7IwMGDOhzb9hQKKScnBw5kpOT+2zj0Ol0Kj09XUVFRVZHOW+cOHFCzc3NnHMAOE/V1dVp+PDhyszMtDoKgLNw8OBBZWZmqri42HStxsZG+f1+TZs2zXQtt9utffv2ae7cuabq1NbWyu/3c60aJ3v37lVubq6ysrKsjtLn+Xw+HT58mL/Nt/F6vfJ4PJyXc6i1tVUNDQ0qLCzss72f3srtdistLU1Dhw61Okq/VFdXp7FjxyolJcXqKB85wzAMl8tl7ht4C7S0tMjn88nm9/ujZmcQWKWjo0MbN27UqlWrrI5y3tiyZYuqqqrU0NBgdRQAgAVsNpu2bdumOXPmWB0FwFmYN2+eRo0apYqKCtO11q1bp127dqmmpsZ0rYqKCpWXlysSiZiqU1paqvr6erW0tJjOBGn06NFasWKFFixYYHWUPm/Tpk1asmSJotGo1VF6lXi99/HeamtrNX36dHm9XuXl5Vkdp18pKSnRzJkztXTpUquj9EsOh0O7d+/WhAkTrI5ihQOSCq0O8WHt3LlTZWVlrNkMAAAAAAAAADCPZjMAAAAAAAAAwDSazQAAAAAAAAAA02g2AwAAAAAAAABMo9kMAAAAAAAAADDNEa9CJ0+e1LFjx5SQkKCRI0dq0KBBPfYHAgEdPXpUhmEoPz9fGRkZ8RoaAAAAAAAAAGCxuDSba2pqVFVVpTFjxqi5uVnBYFCVlZUqKCiQJLndbi1fvlzZ2dlKSUmR2+3WsmXLVFJSEo/hAQAAAAAAAAAWM91sPnPmjNavX6/Vq1dr0qRJikQiqqys1H333aeqqip1dXVp5cqVWrhwoebPny9JeuaZZ7R27VqNHz9egwcPjsfrAAAAAAAAAABYyPSazT6fT0lJSZo4ceIbBe12TZ48WV6vV5L0yiuvqKWlRbNmzYo9Z+rUqYpEIvJ4PGaHBwAAAAAAAAD0AqZnNo8cOVLbt2+P/RwKhbRnz55Y8zk7O1tbt27tsYazx+NROBxm3WYAAAAAAAAA6CfidoNASVq7dq2effZZDRkyRBs3bpQkuVwu5eXlxY7xer1auXKlZs+erWHDhsVzeAAAAAAAAACARUwvo/HvFi9erMrKSmVlZWnNmjU99hmGoQceeEDXX3+9rr76an3/+9+P59AAAAAAAAAAAAuZntkciUSkN9dqzsnJiT3mzZunU6dOKTMzU6+++qpuvfVW5ebmavPmzcrJyYlHdgAAAAAAAABAL2F6ZnN1dbVuvPHGHttcLpck6cyZM4pEIrr99ts1efJkrV69mkYzAAAAAAAAAPRDppvNY8eOVWNjo9xutyQpGo2qurpa2dnZys3N1f79+9XU1KTi4mIdOnRIBw8ejD0CgUA8XgMAAAAAAAAAwGKml9EYMWKEbrvtNt10000aMWKE2tvbJUl33HGHEhISdPz4cdntdt1yyy3veO6GDRs0btw4sxEAAAAAAAAAABYz3WyWpBkzZmjKlCnyeDxKSEjQ0KFDlZiYKEkqLS1VaWlpPIYBAAAAAAAAAPRScWk2S9LAgQNVWFgYr3IAAAAAAAAAgD7E9JrNAAAAAAAAAADQbAYAAAAAAAAAmEazGQAAAAAAAABgmq2pqSmakJBgdY6z0t7erkWLFunIkSNWRzlvBINBRaNRuVwuq6MAACzQ1tampKQkOZ1Oq6MAOAuGYai7u1vxuP4PhUKy2WxyOMzfBqarq0vhcFhJSUmm6hiGIUlKTU01nQlvnE+73c5nfhwEg0F1dnYqLS3N6ii9Smdnp4LBIOflHAqHwwoEAkpNTZXdznzDeDIMQzabTYmJiVZH6Zc6OjrkdDrjcp3R13R3d0cGDBjQ596woVBIOTk5ciQnJ/fZxqHT6VR6erqKioqsjnLeOHHihJqbmznnAHCeqqur0/Dhw5WZmWl1FABn4eDBg8rMzFRxcbHpWo2NjfL7/Zo2bZrpWm63W/v27dPcuXNN1amtrZXf7+daNU727t2r3NxcZWVlWR2lz/P5fDp8+DB/m2/j9Xrl8Xg4L+dQa2urGhoaVFhY2Gd7P72V2+1WWlqahg4danWUfqmurk5jx45VSkqK1VE+coZhGC6Xy9w38BZoaWmRz+eTze/3R83OILBKR0eHNm7cqFWrVlkd5byxZcsWVVVVqaGhweooAAAL2Gw2bdu2TXPmzLE6CoCzMG/ePI0aNUoVFRWma61bt067du1STU2N6VoVFRUqLy9XJBIxVae0tFT19fVqaWkxnQnS6NGjtWLFCi1YsMDqKH3epk2btGTJEkWjUauj9Crxeu/jvdXW1mr69Onyer3Ky8uzOk6/UlJSopkzZ2rp0qVWR+mXHA6Hdu/erQkTJlgdxQoHJBVaHeLD2rlzp8rKylizGQAAAAAAAABgHs1mAAAAAAAAAIBpNJsBAAAAAAAAAKbRbAYAAAAAAAAAmEazGQAAAAAAAABgmiNehU6ePKljx44pISFBI0eO1KBBgz7UfgAAAAAAAABA3xWXZnNNTY2qqqo0ZswYNTc3KxgMqrKyUgUFBR9oPwAAAAAAAACgbzO9jMaZM2e0fv16rV69WuvXr9fDDz+sCRMm6L777vtA+wEAAAAAAAAAfZ/pZrPP51NSUpImTpz4RkG7XZMnT5bX6/1A+wEAAAAAAAAAfZ/pZTRGjhyp7du3x34OhULas2dPrLn8fvsBAAAAAAAAAH1f3G4QKElr167Vs88+qyFDhmjjxo0fej8AAAAAAAAAoG8yvYzGv1u8eLEqKyuVlZWlNWvWfOj9AAAAAAAAAIC+yfTM5kgkIr25FnNOTk7sMW/ePJ06dUqDBw/+j/szMzPNvwoAAAAAAAAAgKVMz2yurq7WjTfe2GOby+WSJJ05c+Z99wMAAAAAAAAA+j7TzeaxY8eqsbFRbrdbkhSNRlVdXa3s7Gzl5ua+734AAAAAAAAAQN9nehmNESNG6LbbbtNNN92kESNGqL29XZJ0xx13KCEh4X33AwAAAAAAAAD6PtPNZkmaMWOGpkyZIo/Ho4SEBA0dOlSJiYkfeD8AAAAAAAAAoG+LS7NZkgYOHKjCwsKz3g8AAAAAAAAA6LtMr9kMAAAAAAAAAADNZgAAAAAAAACAaTSbAQAAAAAAAACm2ZqamqIJCQlW5zgr7e3tWrRokY4cOWJ1lPNGKBRSd3e3XC6X1VEAABZoa2tTUlKSnE6n1VEAnIWOjg5FIpG4vIdDoZBsNpscDvO3genq6lI4HFZSUpKpOsFgUJKUkpJiOhMkwzBkt9v5zI+DYDCozs5OpaWlWR2lV+ns7FQwGOS8nEPhcFiBQECpqamy25lvGE8dHR2y2WwaOHCg1VH6pUAgoMTERPXVnqUZ3d3dkQEDBvS5N2woFFJOTo4cycnJfbZx6HQ6lZ6erqKiIqujnFecTqe6urqsjgEAsEBdXZ2GDx+uzMxMq6MAOAt///vflZaWps9+9rOmazU2Nsrv92vatGmma7ndbu3bt09z5841VWf//v166aWXNHbsWNOZIO3du1e5ubnKysqyOkqf5/P5dPjwYf7b9W28Xq88Hg/n5RxqbW1VQ0ODCgsL+2zvp7f629/+ptTUVI0bN87qKP1SXV2d8vPzNWTIEKujfOQMwzBcLpe5b+At0NLSIp/PJ5vf74+anUFglY6ODm3cuFGrVq2yOgoAAOcFm82mbdu2ac6cOVZHAXAWcnNzNWzYMP3tb38zXWvdunXatWuXampqTNeqqKhQeXm5IpGIqTp33XWX/vSnP+npp582nQnS6NGjtWLFCi1YsMDqKH3epk2btGTJEkWjUauj9Crxeu/jvdXW1mr69Onyer3Ky8uzOk6/MnjwYE2aNElPPvmk1VH6JZvNpk2bNmnx4sVWR7HCAUmFVof4sHbu3KmysjLWbAYAAAAAAAAAmEezGQAAAAAAAABgGs1mAAAAAAAAAIBpNJsBAAAAAAAAAKbRbAYAAAAAAAAAmEazGQAAAAAAAABgGs1mAAAAAAAAAIBpNJsBAAAAAAAAAKbFrdl88uRJ/eMf/9Dzzz+vQCDwnscZhqG//OUv8RoWAAAAAAAAANALOOJRpKamRlVVVRozZoyam5sVDAZVWVmpgoKCdxy7efNm/f73v9e0adPiMTQAAAAAAAAAoBcwPbP5zJkzWr9+vVavXq3169fr4Ycf1oQJE3Tfffe949jnnntOv/vd78wO+YG1t7frueee0x/+8Af98Y9/1AsvvKDTp09/ZOPj3HrqqadUUFCggoICHT58+H2PX7t2rQoKCnTFFVd8JPkAAAAAAACA84npZrPP51NSUpImTpz4RkG7XZMnT5bX6+1xXFtbm+655x6VlZWZHfJ9nTx5UkuXLtVFF12kK6+8Up/97Gd19dVXa/z48Ro2bJhuvvlmtbS0nPMcVggEArr33nv1zDPPWB3lnAsGgzp06JAOHTqkcDj8vsefOXNGhw4dksfj+UjyAQAAAAAAAOcT083mkSNHavv27bLZbJKkUCikPXv2xJrPkhSNRrVhwwbNnz9fl156qdkh/6OXX35Zn/nMZ3T33Xero6ND6enpmj17tj73uc/pwgsv1OnTp7V+/XpdffXVevXVV89pFivU1dXpv//7v1VbW2t1FAAA0E8Fg0GdOHHiA33ZCwAA4qu1tVXNzc38e/gc6O7u1okTJ9TZ2Wl1lH6Ja8i+6dSpUx9qpYi43SBQby5T8PnPf14HDhzQDTfcENv+hz/8QYFAQF/84hfjOdw7RKNR3XDDDTpw4IAkad26dTp+/Lh27NihJ554QkePHtXq1aslSXv37tXdd9/9rjWOHTum3bt364UXXviPHzCtra1qaGjQX//6V7344ovq6up6z2P9fr9eeOEFPffcc+/a5A4EAmpublZzc7Oi0agikYgaGhr097//vccvtKurS88//7z27NnT40aMwWBQzc3N+u1vf9uj3uuvv/6hchiGEcthGEaPfadOnVJzc/O7Pq+zs1MHDhzQX//6V+3bt+9d/whPnjwZq/32x7vdVNLr9aq+vl4NDQ3v+0f91pcdR48e1T/+8Q81Nzf/x+Pfcvr06R7n/d+9tb21tTW2raWlRc3NzTp16pQkKRwOa9++faqvr9drr732rmNEo9HYuTl69Kj0tvMcCoU+UFYAAKzm8/l03XXXKSUlRdnZ2crNzdWvfvUrq2MBAHBeOHHihK699loNHjxY+fn5KiwsVE1NjdWx+oVQKKT169frwgsvVF5eni644AItXbpUwWDQ6mj9wqlTp3TDDTcoJSVFeXl5ysvL0+bNm62OhfdRXV2tgoICXXzxxUpLS9PcuXN1/Pjx931eXJvNixcvVmVlpbKysrRmzRpJ0vHjx3X//ffrRz/6kez2uA73Dnv37lV1dbUk6brrrtPy5cuVlJQU2z9o0CCVl5drzZo1euCBB/S9732vx/NffPFFzZgxQyNGjNDkyZM1fvx4DR8+XI888kiP4wKBgJYuXarc3FwVFRXpU5/6lMaMGaNx48bp97//fY9j29ra9J3vfEcXXXSRxo8fryuvvFJDhgzRDTfc0KM5uWPHDuXk5CgnJ0cvvfSSvvjFL6qoqEif+MQnVFhYqOeee04vv/yyPvnJT+ryyy/XJz/5SV122WU6dOiQJKmxsVE5OTnasmWLJGn9+vXKycmJvXk/aI7nnnsuluPvf/97j9dy8803KycnR7NmzYpti0aj+sUvfqGCggKNHj1an/rUp3TZZZdp+PDhuvPOO9Xd3R079qtf/Wqs9tsfO3bs6PF7uPrqq3XJJZdo4sSJKioq0rBhw1RVVdWj3r/r6urSt7/9beXn52vSpEnKycnRypUr39FAfrtf/vKXsQxvb8y/tf2ee+6JbVu5cqVycnL0ta99TV6vN/Z6J06cqBEjRrxjTfJTp07pC1/4Quzc5Ofnq6ysTE899VSs/iuvvPIfMwIA0BuEw2EtXrxYv/rVrzR79mzdeeedSkhI0HXXXcd/6AIAcI5Fo1Fdf/31cjgc8vv9OnPmjCoqKjRnzhwdPHjQ6nh93k9/+lPdf//9+vOf/6zOzk69+OKLeuqpp7Rhwwaro/V50WhU3/rWt+T3+3Xs2DEZhqH/+7//03e+8x3V1dVZHQ/vYdeuXZo7d67uuusunTlzRq+99pqSk5N1/fXXv+9zHWYHj0Qi0ptrNf9783DevHk6deqUampqlJqaql//+tfSmzNrg8Gg1q9fr0996lNxvVnbvzdHv/GNb7zncbfffvs7tjU3N6u0tFQvvfSSrrnmGn3rW9/SsWPHtHz5cn31q19VRkaGZsyYIUmqrKzU3XffrVGjRunmm29Wenq69u3bF/ugr6+v15VXXilJuuGGG/Twww8rJydHFRUV6u7u1ooVK/Tzn/9coVBI999/vyTJ5XLFspSXl0uS5s2bp9/+9rfy+XwqLy/XBRdcoLy8vFhz9l//+pcqKir04IMPKiMjQytWrNDdd9+tUCikKVOmaOrUqRozZsyHyvFhPfnkkyorK5PL5VJlZaXy8/Pl8/l05513asWKFUpNTdV3v/tdSdLEiROVlZUVe+7hw4e1d+9eSZLT6ZTenGn8hS98Qf/61780depU3XjjjTpw4IDKy8v1wx/+UBkZGVq0aNE7ctxzzz06cuSIvvvd7+rxxx/X8ePHtWbNGk2ePHcmPVAAACAASURBVFlXX331Wb22d/PWlxcnTpxQWVmZsrKytGTJEj344IM6ffq0Fi9erE984hO6+OKLJUk//OEPY1+ALFq0SFdccYW2b9+u+vr6WM2EhIS45QMA4Fx59tln9Yc//EHFxcV65JFHlJCQoKKiIj3xxBM9/i8gAAAQf263W9XV1frXv/6lzMxMSdJXvvIVPfbYY9q6davWrl1rdcQ+yzAM3X333frZz36m4uJiSdKll16qzZs3q6Ojw+p4fd4LL7ygnTt3qqmpSTk5OZKk6dOna8eOHbrwwgutjof38Itf/EJlZWWxVSoSExO1du1a7du3T11dXbE+3rsx3Wyurq7WH//4R917772xbW81Ts+cOaMJEybEPgglacCAAbLb7Ro+fLhSU1PNDt/DSy+9FPvnD7s29KZNm/TSSy8pPT1dmzZtUnp6uvTma7j99tt19913x5rNb83euf766/Xtb39bkvTlL39Z+fn5CgaDsRP++uuv64ILLtDXvvY1XXXVVbEm6UsvvaQf//jHeuCBB7RhwwYlJSX1mPVtt9v1+OOPy2az6dprr9VvfvMb1dTU6L/+67+0ceNGRaNRzZ49W08++aQeeeQRbdq0SZdccokqKir005/+VKFQSNOmTYstGfJhcnxYu3fvliSVlJRo2bJlse1jxoxRY2Njj9/DunXrYv/c2tqqT3/605KkGTNm6POf/7wk6dixY7rqqqt0xRVX6Pvf/74mTZoUO+d/+9vftH379ndtNr/++uv6y1/+IpfLpeuuuy72L4ja2tq4Npvf+j253W7Nnz9ft912m/RmI/3b3/62zpw5o/r6el1zzTV6+eWX9eCDD0qSvv71r2vz5s2y2Wz65je/2eNLlreWAAEAoLdqb2/Xn//8Z0nSpEmTdOLECUnSzJkzVVpaanE6AAD6v3/+85/Kz8/XiBEjemyfMmWKHnvsMcty9XWBQEB1dXXy+/1KS0uL/d/jgwcPjuvkyPPVyZMn9fTTTysrK0vd3d2x85uTk6PPfvazVsfDu+ju7taRI0f01FNPacWKFbHfmcvl0rBhwzRs2LD3rWG62Tx27FhVVVXJ7XZr3Lhxikajqq6ujq3jl5CQoI9//OOx4//5z3/qiSee0DXXXGN26Hf492+c/n2msN5cLuHWW299x3Nqa2v18Y9/PNZALigokM/nk8/nkyQNGTJEkvT000+rpaVFGRkZysvLi635HA6HNWnSJH384x/X4sWLe9ROTU3Vz3/+c+nNX9apU6cUCoViTd1QKKTTp0+/o8n79a9/PdbUnD59un7zm99Ikq699lrZbDbZbDaVlJToySefVCgUUktLS48Zw293tjk+iLfGra6u1s0336zp06erqKhIV111la666qp3fU40GtUPfvADud1uZWdn67777ov9vi677DLdd9990pvrUJ88eVLhcFgXXXSR9OYM9Hfz1uzqt2qMHDlShw8fltfr/dCv6YO67rrrYv9cUlIS++e3Mu7fvz+2be7cubGm8qBB/4+9O4+rKf//AP66cetKqVTT0CITERqMEIoxjKkYI2sylhljb+wzmfG1jq0MxiBjGYMZw9jJkn0JoUSlhUoLSnvpdrt1q/fvj+meX3t0o8H7+Xj0eHTvOefz+ZzPOecu7/s5748WRo0ahYULF76ytjHGGGO16Y8//sDy5csBAJs3b8bmzZuho6ODlJQU1KtXr66bxxhjjL31EhMTKxxUZ2RkhLt379ZJm94Gly5dgqenJwBg69atqF//3zDZtGnThAFyrOZ27NiB3bt3Q01NTbiLHwB+/fVXaGlp1WnbWMUePXqE7777DhkZGTh16pRwZ37v3r0xZcqUFypD5WDzBx98gPnz52PGjBn44IMPkJOTAwBYvHjxa08PoKurK/yflZUljE5GcZ7BtLS0ctso04CEhYUBAPz8/ITUE2UlJiYK6Spu3LiBpKQkIYAtFovRv39/TJs2rVTg8cqVK/j5559x8eLFCm+/qCinsDLAjeJAcUXP6+joCP9Xlse4pJq040W4urpi37598PPzw7p167Bu3TqgeNTT+PHj8dVXX5X7Erpt2zbs2rULKP4RwNzcvNS+7NmzBxs2bEBAQEC5+pTHqyzlrRglHz98+PCVJfPX1tYWAuAoczyUE/6VvKW45LoAYGJi8kraxRhjjJWUkZFR6nOCrq6u8CXqZYwbNw4XLlzA0aNHsWjRIowbN66WW8oYY4yxqsjl8nKD6lCcklImk6GwsJB/AK6Bzz//HGpqahgwYAD++uuvGg3CY5WbN28epFIpQkJChIGU7L+tZcuW2LJlC7y9vYX0tC9L5WAzitMg2NnZITY2FmKxGKamptDQ0Khw3Q4dOuDMmTO1UW05rVq1Ev4PDAwsFcT86quvMHLkSGFZ2VG3yi9enTt3xpw5cyosXxkwtLGxQVBQEI4fP46LFy8Ko56PHj2Ko0eP4tixYxg4cCD8/f3x8ccfAwCcnZ0xbtw46Orq4tSpU/Dw8Kh0P0q+QZRMsVAy1cbLpF6oaTvKBqDlcnm5dQwMDHDhwgWcOnUKFy5cwNmzZxEdHY2bN2/i5s2biImJEUZCobjvZ8yYAQCYP38+nJycSpW3ZcsWIcfzd999hz59+kBTUxMrVqyocvKhsl+ca5KaouT+5ufnV7muurp6qToqqq/k8VIGoJVyc3Nfun2MMcbYy3Jycio1p0VkZORLpxoDgJ07dwqTIC9ZsgRLliwBitNY8dwDjDHG2KvXoEGDCgeO5eXlQUdH550LNM+cORO3bt0SHnt7e5dK4fqiDh48CDc3N6A4Hajyu/2qVaswfPjwWmzxm0OhUKBnz57C4+bNmwvzsL2s77//Hlu3bkVeXl6pFDA+Pj6wtLSslfayfwUFBWHy5MnCYxcXFyH+9jKioqLQp08foDhlsPJHrsGDB+Pnn39+oTJqJdiM4hc+Kyur2iquRnr16gWxWAyFQoF169bB0dERDRo0AIoTWSsD4BX9UvXhhx/iwoULICKMGDGi2rqMjIwwYcIETJgwAQUFBbh37x5mzpyJ69ev47fffsPAgQNx8uRJYf1ffvkFZmZmQPFtGq/Ty7Sj5I8EJUeCFxQUICgoqMLyGzRogCFDhmDIkCEgIjx69AgeHh7Ytm0bVqxYAXd3dzRq1AgZGRkYN24c5HI5Pv300wonalS+gFlbWwu3sqA4j9KrUPKX4YyMDGF0fGxsrMplN2nSRPg/KioKvXv3Fh6HhISoXD5jjDFWHRsbm1J3RlU0IupFTJs2DTk5OZg3bx5mzJiBNWvWQCQSlfphlTHGGGOvjomJCR48eAAiKjXY6enTp8KcRe8SFxeXUvMzNWzYsEbl9O/fH5aWlmjfvj3+/vtvtGnTRqXy3gb16tUrlfZTGVeriYULF6Jdu3aYPXs27ty5I3x2rO053Ni/d/iXPG41vaO+WbNmCA4ORpcuXTB9+nR8+eWXQJl4YXXeqm8IZmZmmDVrFgDg2rVrcHFxwZ07d5Cfn4/CwkI8efIE//zzjzByFiVGn/bv3x8AEBAQAH9/f2G5cn0PDw8UFRUhOTkZ8+bNw6BBgxATEwMUj6q1sbERIv/KVA8lUz4ob2GNjIzEli1bhOcrGi2sCuXBV07c87LtKPmFdO/evcII3507d+Lhw4el6iIieHh4wNXVFd7e3kDxCF8LCws4OzuXWq9knmYjI6NSeZpLUraPiIT/jx49imvXrgHFEzbWppLpN5STKigUCqxZs0blstu3bw9tbW2gOL9lUlISAODWrVvYuXOnyuUzxhhj1dmwYQOOHDki/NX0Q6dIJEKXLl2A4nkacnNzhQmNu3fvLox0Zowxxtir8dFHHyE+Pl6YrEvp3LlzpVJ5vitsbW3h6Ogo/NU0INqgQQNYW1ujRYsWCAkJgY6ODnR0dFC/fn34+/vj/Pnztd72/zo1NbVSfau8U74mtLS00L17d6SlpSE2NlboX5FIhJMnTyI4OLhW2/4uMzAwKHXcrK2ta1SOWCyGjo4OPvvsM/j5+QnHTCKRID4+HocOHao2Ze1bFWwGgEWLFgkjk48fPw4bGxtoaGigfv36MDU1FQLQ2tra2LFjh9D5Y8eOFUZmf/7553B3d8eMGTMwbtw4eHl54b333oOamhoMDQ0RGhqKY8eOwdHREf/73//g6emJqVOnYuXKlQAgRP2VX8oAYPjw4XBzc0Pv3r1LTSS4atWqUre3qqpDhw5AcV7kCRMmYOfOnS/VDhMTEyG1xeHDh9GxY0d0794da9euFRKBFxQUAMVfPNXU1LB3716MHj0aM2fOhKenJ+bNmycM3Z88eTJ0dHQQFhYm5GmWy+VwcHBA69athT/lL5LKgP39+/cxePBgjBs3DiNHjsQPP/wAAHj48CGWLVuG+Pj4WukvOzs7GBoaAgDc3d1hb28Pa2trFBQUCDPPvkhO7Iro6OgIvyrdvXsXH3zwATp06AB7e3uMGjWqVtrPGGOMvS729vbo3bs3oqOjMXDgQCxZsgTTp0+Hn5/fOzmiijHGGHudWrduDRcXF8yfPx8JCQmQy+XYsmUL/P39hRgEqxmRSIRFixbhp59+wrVr15Cfn4/o6GhMnDgRgYGBdd28N16LFi0wefJkfPfdd4iMjERBQQEuX76MoUOHIj09va6bxyoxefJkHDp0CH/88QdycnKQkZGBhQsXYtu2bdWOcn7rgs2ampr466+/sHfvXnz66afllvfu3Rtr165FZGQkvvrqK+H2k8aNG+PMmTMYP348srKy4OnpiV9//RVWVlb4559/hMCsSCTCX3/9hXnz5iEzMxPLly+Hu7s7Nm/ejK5du+LgwYNCILF///5YtmwZxGIxAgICsG/fPsyaNQtLly7F4sWLgeKgcGXpKWpi8eLFwmjd7du348mTJy/VDpFIhE2bNsHBwQEonjixfv36OHz4MJo1awYU52dUmjNnDry8vGBqaor169fD3d0dHh4eEIvF+OWXX7B27VqgzOjqrKwsPHjwoNRfYmKiUJ7yjfL48eO4ffs2Dh48iAULFmDAgAEAgOXLlyM5OblW+svAwAAHDhwQ8n1fu3YN3bp1w9q1a9G4cWNAxfzKs2bNwubNm2FjY4PGjRujefPmuHDhArp37y6sU5P80owxxtjrVr9+ffz9998YPXo0Ll26hMWLFyMvLw9eXl4YOHBgXTePMcYYe+tt2LAB2traMDY2RoMGDfDbb7/h+PHjwnd1VnMjR47EzJkzMWjQIGhoaMDKygr9+vWrUc5bVt6KFSvQokULtG3bFmKxGK6urti8ebNKo6bZq9WmTRscOnQIP//8M7S0tNC4cWNkZ2dj+/bt1W4rSklJoTd1tk2ZTIZNmzZh0aJFla4jlUqRkZEBNTU16OnpvdDMojk5OUhLS0PDhg3RuHHjSoOBBQUFyMjIQG5uLho3bgwtLa0K18vNzUVqaioMDAxK3dohk8lQVFRU6XY1lZeXh9TUVGhpaUFHR6fG7UhPT0d+fj6MjIyqDYgSETIzM/H8+XM0atQIurq6KgVRs7KyIJVKYWRkVGryv6ysLEgkkpfKFfMiiAjPnj2DRCKBnp5erZZdEU9PT7i7uwPFqUFq+xxgjLFXRSQS4fDhw6XSJbF3T3Z2NmQyGQwMDN65CYnedCYmJmjWrBmuX7+uclkrVqzA1atXq5zE+UUtX74cCxYsKDVAoSY8PDxw/vx5nDt3TuU2sX+/aP744488arIW7NixA+PHjy83Cfu7rrau/XdNVlYW5HI53nvvvWq/d1+8eBF9+vRBXFycMH8Tq1x+fj7S0tKgp6dX7VwXjRs3hq2tLU6dOvXa2vemk0qlyMnJgb6+fqlYT0VEIhF27NhR6q78d0g4gLqdGK8YESE5ORkNGjSoNte2t7c3Jk+eXHsTBP5XaWlpvXQgr2HDhi+UDL5+/fpCCoaqNGjQAKampuWef1VBfg0NjVK5iGvaDuXI3hchEomgp6dXa4FaZU6Yip5/FUQiUakJ/WqLu7s77t69CwMDA+zevRv169dHdnY29u/fDwDo168fB5oZY4y9cbS1tYV5CRhjjDH2elX2fZmpTl1d/ZXEBti/ahKjY3VLJBLByMjopbZ564PNjNWlVq1awdPTEwAQHx8Pa2trXLt2Dffv34dEIhHSmDDGGGOMMcYYY4wx9qbjYDNjr9DXX38NQ0NDbN++HdevX8f169dhaGiIr7/+Gm5ubjyhEmOMMcYYY4wxxhh7a4gePXpE6urqdd2OGsnPz8ekSZOQn59f1015p6irq3Of11BRURHU1N66eTkZY++QK1euoG3btjAwMKjrpjDGasDPzw/16tWDubm5ymWlpqZCLpfDxMSkVspKSUmBlZVq6QnT0tJQUFAAa2trldvEgICAAJiamr707bOsvMTERDx8+BC9evWq66b8p8TFxSE2Npb75RXKzMxEUFAQbG1ta33uo3fd9evX0ahRI37PeUWuXLmCVq1a4f3336/rprx2crlcJpFI3rgJ9tLS0pCQkADRkydP6E19wcnLy8PChQsRHR1d1015p3CwmTHG3l16enrIyMio62YwxmqoXr16SExMrLXyJBIJ5HJ5rZTVuHFjpKenq1yOiYkJf1atJVpaWpBKpXXdjLeGrq4uMjMz67oZ/zncL68ef357NTQ0NJCXl1fXzXhr6ejoICsrq66bUSfe1GAzAPzwww8QpaSk0KuaqO5Vk8lkkEgknFycMcYYY4wxxhhjjDH2NggHoNrtXnWI7+dnjDHGGGOMMcYYY4wxprJamyAwOTkZMTExEIvFsLS0LDXaOCsrq9wtG2KxGMbGxrVVPWOMMcYYY4wxxhhjjLE6VCvBZh8fH6xduxZt27ZFQkIC8vLysGrVKrRu3RoAcPr0aXh5eZXaxsrKClu2bKmN6hljjDHGGGOMMcYYY4zVMZVzNmdnZ2Po0KFYsmQJbG1tUVRUhFWrViE1NRVr164FAKxduxb6+voYMWLE/1csEqk8EyrnbGaMMcYYY4wxxhhjjL1F3u2czYmJidDU1ETXrl3/LVBNDT169EBcXJywzpMnT2BqagqJRCL8qRpoZowxxhhjjDHGGGOMMfbfoXIaDUtLSxw5ckR4rFAo4OfnJwSfASAmJgbp6enYtGkTCgsL0bFjR9jZ2UEkEqlaPWOMMcYYY4wxxhhjjLH/AJVHNpe0bNkyDBgwAOHh4Zg6dSoAQCqVIi0tDT4+PmjatCkMDQ2xevVqbN68uTarZowxxhhjjDHGGGOMMVaHVM7ZXNLTp0+RnJyMvXv3QiQSwcPDAzKZDPfu3cOHH34o5FYOCwvD5MmTsX//frz//vs1ro9zNjPGGGOMMcYYY4wxxt4ib3TOZpXTaBQVFQHFuZqNjY2Fv6FDhyI1NRUGBgbo3r17qW2srKygpaWF+Ph4lYLNjDHGGGOMMcYYY4wxxv4bVE6jceLECUyfPr3UcxKJBACQnZ2N6OhonDp1qtTygoICyOVyNG7cWNXqGWOMMcYYY4wxxhhjjP0HqBxsbteuHe7fv4+QkBAAABHhxIkTaNKkCUxMTFCvXj2sWrUKFy9eBBGBiHDgwAE0a9YMzZo1q419YIwxxhhjjDHGGGOMMVbHaiVn87lz57By5Up88MEHyMnJAQAsXLgQVlb/phe5dOkSVq5cCT09PRARGjZsiIULF6J58+Yq1cs5mxljjDHGGGOMMcYYY2+RNzpnc61NEJibm4vY2FiIxWKYmppCQ0Oj3PK4uDiIxWKYmZlBLBarXCcHmxljjDHGGGOMMcYYY2+RNzvY/PDhQ1JXV6/rdtRIfn4+Vq5ciZiYmLpuyjvl/fffx7Nnz+q6GYwxxuqAhYUFoqOj67oZjLEaatiwYblBIarQ1tZGdnZ2rZTVtGlTJCQk1EpZ6enptVLOu65JkyZITEys62a8NT744AM8evSorpvxn9OsWTPExcXVdTPeavz57dXQ09NDRkZGXTfjrdW8efN3Nt4nl8tlEolE9ZHBdWDFihWor6mpKUzo96aRy+Xo2LEjzMzM6rop7xQtLS1IpdK6bgZjjLE6oKurCxMTk7puBmOshsRiMXJzc2utPA0NDeTl5dVKWTo6OtDR0VG5HC0tLcjl8lpp07uOP/fXLl1dXZiamtZ1M/5zdHR0YG5uXtfNeKvx57dXozbfA1l5Ojo672y8LycnR9awYcM3Mtisr69fe2k06gKn0WCMMcYYY4wxxhhjjL1F3ug0Gmp13QDGGGOMMcYYY4wxxhhjbz4ONjPGGGOMMcYYY4wxxhhTGQebGWOMMcYYY4wxxhhjjKmsfm0VlJycjJiYGIjFYlhaWpbLo1xUVITY2FgkJSXB0tIS+vr6tVU1Y4wxxhhjjDHGGGOMsTpWK8FmHx8frF27Fm3btkVCQgLy8vKwatUqtG7dGvh3FkUsXboUcXFxaNSoEWJjY7F8+XJ07ty5NqqvcxkZGXjy5AmaNGkChUKBhIQEtGrV6p2fuJCIEBgYCB0dHbRo0eK11RkdHY38/HxYWVlBJBK9lnpLioqKwvPnz/HRRx+99rr/q2QyGcLDw9GkSRM0bdr0pbaVSqV48OABmjZtiiZNmryyNj548AAymQwdO3Z8ZXVU5eHDh5BKpf/J8yYxMfGNeF17U9rJGGOMMcYYY4y9rVROo5GdnY1169Zh6dKlWLduHfbu3YsuXbpg69atwjo7d+6ERCLBX3/9ha1bt+Lrr7+Gp6cnioqKVK2+FD8/Pzg7O+PIkSOlno+Li4OzszPWrl1bq/UBwMWLF9GsWTP06dMHly5dwqZNm2BjYwOpVFrrdb1pnj59ChsbG/zzzz+vpT6FQoEhQ4agZcuWGDNmzGupsyKTJk3C5MmT66z+/6KbN2/CxsYGt27deultL1++DBsbGwQFBb2Stim5uLhgzpw5L7WNVCrFlStXhMepqano2bMnYmNj4ePjA2dn5xcua9SoUZg+ffpL1V8bpk6dCmdnZwwePBhTp07FL7/8gmfPnpVaR/m6JpPJXmlbHj9+jODg4Bpv/7rayRhjjDHGGGOMsYqpHGxOTEyEpqYmunbt+m+Bamro0aMH4uLigOIRjSdOnMCoUaNQv/6/A6kHDRqE+fPno6CgQNXqS9HW1sbRo0cRGBhY6vkNGzbg6NGjcHR0rNX6AODXX3+FgYEBYmJiMGzYMAQEBKBFixZ4//33a72uN83Dhw8BAG3atHkt9fn7++PIkSPYtm0bAgIC6mRUs0wmg6+vr3A9sH+FhYUBAFq1avXS2/bv3x8FBQX47LPPXkHL/pWRkYF79+7BxsbmpbbbvHkzVq1aJTxWKBRwdXVF06ZNoaen91I/ety6dQtXr159qfpVJZVKsXnzZjx+/Bjt2rVDYWEhli1bhjZt2uD69evCej/99BMKCgrw3nvvvbK2EBGmTZsGHx+fGpfxOtrJGGOMMcYYY4yxyqkcbLa0tMSRI0eEwJ5CoYCfn58QbEtISIBMJoOZmRnCw8Nx/fp1ZGVloUOHDlBXV1d9D0owNzcHAMTExAjPRUVF4ddff8XcuXNhZWWF4OBgTJkyBb169cKECROEIFhOTg5cXV2xe/durFmzBg4ODsJzXl5e2LhxI3r37o2xY8ciLi4OCoUCX375Jc6dO4f8/HxMmjQJUqkUvr6+sLOzE+o/efIkXFxc0KdPH8yfPx+pqalAcQ7rbdu24fPPP4eLiwvu3bsHNzc3eHh4AAB27doFV1dXZGRkCPvh6uqK48ePAwAOHjwIV1dXBAQEYMqUKejbt2+pkduRkZHC86tWrcKFCxfg6uoqBIAPHTqEMWPG4LPPPsM333wDf39/YdvFixdj5syZOH36NJycnDBu3DgkJibi6NGjcHJywoQJE5CSkgIAyMvLg5eXF/r37w8nJyds3LhR+BFB2bdisbjCNu7btw+urq64efMmBg8ejKNHjyIzMxNLlixBv3790K9fP6xcuRL5+fnV9smxY8fw/fffAwD279+PNWvWwNXVFdHR0QCA48ePw9XVFQcPHgSKR127urri3LlzVe5DSEiIsN6SJUvQt29fzJgxA8+fPwcAPH/+HEuWLMEnn3yCCRMm4MqVK1AoFPjwww8BAHPnzsXcuXOFfT516hRcXV3x4MEDyOVyuLq6YseOHdi5cyccHR2FZUoXL16Ei4sLHB0dcejQIWzduhWjR48W2lfZ+VVdu8tavHgx5s6di6tXr2LEiBFwcnLC6dOngeJArKurK3bt2iWsv3r1aqEd9+7dE+qaNm0aFixYUK78oKAgaGpqIikpCc7OznBycsK5c+eE5UlJSViwYAH69OmD4cOHC3UDwLJlyzB58uQKfzxQXgdXrlyBq6srBg4ciNu3byM4OBgjR47EoEGDcPfuXWH9yvorMjISANCsWbMK+6vsPv7www8YNWoUli9fjqioKGFEdFhYGHx9fdGnTx9s2LABpqamL3Qsw8LC8OWXX+LixYsAgMzMTCxatAj9+vWDm5sbgoOD4erqCm9v72r7q6pjWZby9WD48OFYunQptmzZgjt37kBTUxPu7u5A8TX+5ZdfwsvLS3hc2fVS3evSkydPsGDBAgwcOBBffPEFPD09kZeXh5CQEDg7O8Pb2xtHjx7Fzp07Kz2vKjuGZdtZWV3V7cPLXjuMMcYYY4wxxhgrISUlhXJycmrl74cffiA7OzsaOnQoJSUlUU5ODvn6+lKvXr1oxowZNG7cOJoyZQp169aNjh49qnJ9KSkplJ2dTSVZW1uTjY2N8NjNzY10dHTo6dOnFBgYSNra2jRlyhQKCAigr7/+miwtLUkmk9H9+/cJAHXq1Im++eYb2rhxI4WHhxMAsrKyol27dtHmzZsJAI0dO5akUil5enoSAJo6dSqdOHGCwsLCCACtX7+eiIh+//13AkAbNmygmzdvko2NDY0fP56IiLZs2UIAaP78+XTq1ClycHAgAOTh4UFERIMGyzhROgAAIABJREFUDSIzMzMqKioiIqKDBw8SADp//jwREU2ePJkA0LBhw2jbtm1kZ2dHACg8PJwyMzOpRYsW1K5dOzp27BgtW7aMWrRoQQAoMzOTzpw5QwBo06ZNdPbsWRowYAAZGxtTbm4uFRUVkbGxMZmYmNCiRYto7ty5BID69etHy5YtoxkzZgjbFhUV0fjx48nQ0JBOnjxJp06dIolEQn/88QcREU2YMIEA0NChQ8u1kYjom2++IQD0xRdf0MKFCyk0NJQmTpxIRkZGdO7cOVq5ciUBoCNHjlTbJ/7+/mRvb08SiYQOHDhAhw4dIgAUGBhI+fn51KFDB7KwsKDFixcTEdGaNWtIX1+f0tLSqtyHP//8kwCQra0teXl50ZgxYwgA/fnnn0RENGbMGJJIJLR9+3Y6cOAAWVhYEADy9fUlmUxGYrGYpk2bJpyPc+bMIQCUnp4unF+Wlpa0dOlSWrJkCQGgyZMnExFRQEAAicViGjZsGJ0+fZpGjx5NFhYW9Nlnn1V7flXX7pKUx9zMzIwmTZpEmzdvJkNDQzI2NqaioiK6ceMGAaC///5b2KZNmzb0ySefEBHRzp07CQDZ2dnRvHnz6OTJk+XqsLKyIj09PVq+fDl5e3tTq1atSF9fn7KzsykzM5Pat29P9vb25OfnR2vXriUAFBwcLLRt2LBh5cokIpo0aZJwDW7cuJEAUOfOnWnGjBn0yy+/EAAaMWJEtf21a9cuAkBdu3atsL/K7uOBAwdo0aJFBIDmzZtHly9fptDQUAJAc+bMIV9fX2rTpg1ZW1u/0LHcvXs3AaBbt24REdGoUaNIIpHQ77//Tnv37qU2bdoQALp06dIL9Vdlx7Ksv/76iwDQiRMnSj0/c+ZMAkAZGRnC69qGDRuqvearel0iIrK3t6dPP/2ULly4QL/99hsBoK1bt1J0dDS5uroKj0NDQys8r6o6hiXbWVVd1e3Dy1w7jDHGGGOMMcbYKxBW1w1QRa0Gmx8+fEjXrl2jadOmkZubG+Xk5NCFCxeoU6dOtHv3bmE9Hx8f6tGjBz158qTWg81jx44liURCCoWCIiIiCAB5enoSEdGIESNILBZTaGgoPX36lPbs2UMAKDQ0VAhOrlixQijr6NGj5YJs+vr61LVrVyIiYZuLFy8SEdHhw4cJAJ07d45ycnJIX1+f7O3t6enTp/T06VOaM2cO6evrk0KhIHNzc+rRo4cQAFIGsk+cOEEKhYL09fXpyy+/FOpVBiKfPHlCVBxUb968OWVmZhIR0YYNGwgARURECIGzU6dOERGRQqEgsVhMnTt3JiKiP/74gwDQjh07KCUlheRyOSkUCiIiiouLIwD0zTffEBEJAdFvv/2WiIiCgoIIAG3evJlu375NAOiHH34Q9rFz5840ceJEouKAZGVtVC43MjKixMREYT8LCgooPT2d0tLS6PTp0wSADhw48EJ90q5dO/r000+JiCg4OJgA0O3bt+nYsWNkaWlJCxcupJkzZ5JcLicLCwv66aefqt2H7777jgCQv78/ERHdunWLANBff/0l9MWiRYuENrm4uBAAevbsmbB869atwnJ7e3vq2LFjqfPFy8uLiIjkcjkBoClTphAR0cSJE4WyqDhgqfyBoqrzq7p2l6U85iNHjhTOx2HDhpGxsTEREW3fvl0I3BMRpaSkCH1GRMIPEhcuXChXNhFRUlISAaAxY8YIz7m7uxMAio6Opk2bNhEA2r9/Pz19+lTot127dgltK3ldlmRtbS38YEREZGRkRNbW1iSXy6moqIi0tbVp5MiR1faX8keAyvqron1U9ktQUJDwXHZ2NqWnp1NycjJ9/vnnZGlpWe2xLHm80tPThaD1ggULhHJHjhxJACghIeGF+quyY1mW8jhERkaWev7HH38kACSVSoXz9MKFC9VeL1W9LhUVFZGenh45ODhQUFAQKRQKysnJEeocO3as8PpYUZ9XdwxLtrOqulS55hljjDHGGGOMsdfgjQ42q5xGo6ioSJjoz9jYGB07dsTcuXPh5+eH1NRUNGrUCADg5OQkbNOjRw+IRKJS6S5qi5WVFeRyOZKTk7F27VqYm5tj4sSJUCgUOHz4MBQKBdq2bQtjY2OMGjUKAKCpqYnQ0FAAwLBhw4SylM99/PHHQHE+3rS0NDRp0qTUcmUe2vv37wuPQ0NDkZaWBl9fXxgbG8PY2Bhr1qzBe++9h8jISMTGxuKTTz4RUgMobwVv1aoVYmNjkZaWho8++khoy507d2Bubo6mTZsiLS0NISEhGDJkCHR0dAAAERERkEgkMDc3x40bNwBA2D4rKwsKhQLdunUDAOG28q+//hqGhoYYO3YsEhMTAUBI4aDMb608Rg4ODgCA2NhYAICFhYWQemPlypXCPvr7+0NbWxupqakICwurtI3K5cOHDxfyW8fFxWHs2LFo1aoVBgwYgKlTpwLFqVqq65PU1FTcv38fnTt3BorzdwNAQUEB1q9fj1mzZsHAwADp6ek4c+YM4uPjMXbs2Cr3AcV5oNu0aSPk8lWm5bC0tBQmu+vRo4fQprS0NFhYWMDIyAjh4eFAiZzVOTk5uHnzJrp3717q/FFeG8o8523btgUR4dixY7C1tYWRkREAID09HQBgbW1d5flVXbvLUh7zoUOHQiQSgYjg7+8v7FdwcDDEYrGwrXJ9a2trAMDt27fRokUL4TopS5mqYcCAAaX6CQB0dXWFPMXDhw+HsbEx2rdvDwBo2LChUFfbtm3Llau8DgYOHIgGDRogKSkJSUlJ+OKLL6ChoYGEhARkZ2fDysqq2v66fft2lf1V0T4q+6Vly5YAgN9//x02Njbo0aOHkBKiW7du1R5LFOdr7tixI/T09HDnzh0AEK5XFOfGt7CwQJMmTV6ovyo7lmXdvn0benp6QgoiJV9fX3Tq1AkNGzYs9bpW1fVS3euSSCTCb7/9Bn9/f7Rv3x6tW7fGvn37hDqvX7+Onj17Crn9y/Z5dcewZDurqkuVa54xxhhjjDHGGGNVUznYfOLECUyfPr3UcxKJBACQnZ0NQ0NDABDy7qI4QF1YWIgGDRqoWn05LVq0AABcvnwZW7duxeLFi6GjowO5XA6FQoGJEyciLy8PeXl5kMlkyMvLg7m5Oe7cuQNjY2NYWFgIZSmDPgYGBkCJoJmtra2w3NzcXAg+BwYGwsTEBCYmJsjNzQUA7Ny5E3l5eZDL5ZDL5QgODhYCy8rAExHh0qVL0NfXh7m5OaKiooASQcq0tDScP38ednZ2EIlEQjuUgSoUB2rs7e2hoaGBx48fAwAaN24stAuAkEe4cePGOHr0KKKjo/HHH3/g1KlTWL9+PVAiz3Lr1q2BagLqOTk5QHHQrWR/enh4VNtG5fIuXboIy3/44QdcvnwZoaGhuHHjBj755BNIJBK0bNnyhfukXbt2AAAtLS0AwNmzZxESEgIXFxfo6OggKysL27dvx5QpU2BqalrlPuTm5uL69eulcnAHBwcDAFq2bCkE6JXBrtTUVPj6+gqBPWWwuWTAUqFQCMHBO3fuoEmTJjAzMwOKA3PKfczJyUFSUpKwDABu3rwJFP+gUtX5VV27yyp7zJ8+fYrY2Fh07NhRaGfXrl3RsGFDoDgQqVxfJpPBz88Pffv2hZpaxS8nyn5Q9pNCocDZs2fRtWtXNG7cGM+fP4eRkVG563Lw4MHl2lZS2XOs7GNlf7Zr167K/pLJZLh582al/VXZPt6+fRt2dnZo0KAB7ty5g2+++QYTJkxAaGgofvvtNwCAjY3NCx1LPz8/4UeIpKQkoMTrTnJyslAXivOEv2h/lT2WJeXm5uLatWulArwA8Oeff8LX11f4sScwMBBmZmZo2rSpStc8igPkjx8/xpUrV9C5c2eMHz8esbGxSEpKQlRUFDp16gQU/7BXts+rOoZl21lVXapc84wxxhhjjDHGGKuaysHmdu3a4f79+wgJCQGKA6cnTpxAkyZNYGJiAkNDQ3z88cfYs2ePMAL68OHDMDAwQPPmzVXfgzI++OADAMCSJUvQqVMnuLi4AMUjXW1tbXH79m2kpqYiMDAQlpaW8PHxQUFBAa5du4ZevXoJI42Liorg6+sLsViMEydOICEhAWvWrIFYLMbgwYOFbZTBzoKCAvj6+qJnz54QiURo0aIFNDU1cf36deTk5GD37t3o2LEjYmNjhQD8yZMn4e/vj6VLl+LixYuwt7dH/fr1kZ2dDQC4ceMG/P39MXXqVMhkMiFgpAwoWVlZAcUTuN27d08YiacMau3btw/nzp0TJvpq06YNnj59is8//xxeXl4wMjJCz549oaWlJYzqu3v3LjQ1NYWg+71796Cvr49mzZoBxYFHY2NjmJqaCoFdPz8/SKVSzJ49G4MGDUJBQYEQYKysjWVH/aJ4wj+ZTIaUlBQcP34cZ86cQcuWLaGurl5tnyjLUwbZlIHRTZs2YebMmdDV1YW2tjZu3LgBb29vTJgwASgRnK5oHyIjI0sFh1EcYOzYsSN0dXWFYOCNGzfw+PFjLFq0CHK5XAjqZ2VlAcU/fFy8eBGzZ88W+qTs+YISgXxLS0toamrCxMQE169fx5UrV/Dnn39i/fr1wkjaqs6v6tpdVtljrhwdqzw2SUlJiI2Nxc2bN7Fnzx6sWrUKKA6+RUVFQaFQoEOHDpVek0FBQcJ1lJSUhC1btiA+Ph6TJk0CigOySUlJCAoKQlJSEj755BOsWLECIpEId+/ehY6OToWvFWWvg7LnQMnAa1X9pdyHyvqron0kIty7dw9isRgJCQl49uyZcMxjY2OxZs0aAEDTpk3RoEGDKo+l8ngpzxvlSP/Dhw/j5s2bmD59eqlzvbr+qupYlqSsV01NDYcOHcLu3bsxfvx4jBkzBt988w3GjBkjnKf29vYQiURVXi/VvS5t3LgRjo6OSElJwUcffSQEpTU1NYX+y8/PR0ZGRoV9XtUxLNvOqupS5ZpnjDHGGGOMMcZYNWojZ/PRo0epa9euNHLkSBo4cCANHDiQAgIChOUxMTE0fPhwcnJyImdnZxowYECp5bWZszktLY0AVDjpVVBQENna2hIAEovF5ObmRlKplKKioggArV69Wlg3JiZGyKnq5OREAMjY2JgOHjxIRCRs8/PPP5d6XLKMQ4cOkYmJCQEgIyMj2rhxo7DMw8ODdHR0qF27dsLkYMuWLSMiosTEROrQoQMBIBsbGyFHq7e3NxGRMEmfMi+qn58fAaA9e/YQEVFkZCTZ2dmRpqYmTZs2jaZPn04AKC0tjYiINm7cSEZGRkI/jR49mlJSUoiIqEWLFsKkZUVFRWRubk6DBg0iIqLCwkIyNDQkFxcX4fHixYtJU1OTAFC7du3o0qVLL9RG5fKSx2///v2kqalJEomEPDw8hHyvs2bNeuE+KVmetrY2SSQSIafz+fPnS00YV90+/PPPP8KkbEREMpmMJBIJTZ06lag4d3HPnj0JALVv355WrFhBAOiff/4hIiJfX18yNDQkAOTi4kLz588nAPT06VPhfFFOCElE5OzsXGoCRB8fH2revDkZGhrSpk2bqHPnztS7d+9qz6/q2l1WyWNORLR+/fpSubWVx0EikdCSJUvo888/FybhVNZ19erVCsum4lzagwYNEvoHALm7uwu5eVNTU4XJ4QCQg4MDxcTECG0bMGBAheWWPcemTZtGYrFYyM07fvx40tTUJLlcrlJ/VbaPysnwHB0dKSsri/r06SOcn7du3SJLS0shL3VVx1JZvq+vLxERSaVSGj16NEkkEnJwcBAmADx37twL9VdVx7IkZb3KPxMTE3J2dqZDhw5RYWEhUQWvc6pc8zExMcJEqCieGPPAgQNERJSZmUldu3YVcuxX1ueVHcOy7ayqLlWuecYYY4wxxhhj7DV4o3M2i1JSUkhTU1PloHVubi5iY2MhFothamoq3DatVFBQgLi4OBQWFsLMzExItaEKmUwGiUQipEx4UVlZWVBXV68yjYePjw8cHR1x4sQJODk5ITMzE40aNUK9evVeqi4iQmZmJrS1tYVb1ZOTk/Hs2TO0bt0a6urqOHDgAIYPH46rV6/C3t4eKB5ZnZWVBV1dXWHk64vIz89HfHw8NDQ0YGxsjMLCQnTu3BnGxsY4efKksF5BQQGeP38OLS0tqKurv9Q+laVQKCCVSl+6rRXJzc1FUVGRMDI5Ly8P6urqEIlENe6TF1HTfVAeXx0dnQrTSChv91fmsH1R2dnZePz4MZo0aQJdXV3ExMTAwsICq1evxty5c8vVX/L8ehWkUinU1NSg6muFVCpFvXr1Krz2ZDIZCgoKhDzvr0Jt91dOTg40NTWF/Mglz4WioiIoFArk5+e/0LFUiouLQ25uLiwtLaGmpoZZs2Zh69atSEhIKHUevY7+qowq17xUKkVhYSEaNWpUalsigkwmE679yrzMMaysLlX3gTHGGGOMMcYYe4XCAVjVdSNqqtaCzXWhpsHmF7F27VrMmTMHkZGRQh7o2vLrr79ixowZmDhxIlq2bIn169fD1tYWe/fuVTkAlpGRAWtra2hpaWHixIkIDAzE/v37ceHCBSGQzf77bt68iW7duuGzzz6Dk5MT9u/fj5SUFFy+fFnIEc7eDC97LAcPHowTJ05gwYIFyM3NxcqVK7FmzRohDQtjjDHGGGOMMcbeahxsriuvMti8e/duhIWFYdmyZbU+YrSgoADe3t4IDg6GWCxGmzZt4ODgUCujvQHg0aNH8Pb2RmpqKoyMjNCvXz9hkjr25rh58yYuXbokTGI5YMAAIU80e7O8zLHMzMzE4cOHERMTAx0dHdja2qJHjx48+pYxxhhjjDHGGHs3vNnB5oSEBKqtIOfrJpfLoa2t/UqCzYwxxhhjjDHGGGOMMfaavdHB5vpyuRyFhYV13Y4aKSgo4EAzY4wxxhhjjDHGGGOM/QfU19bWVnnSr7oik8n41nLGGGOMMcYYY4wxxhj7D1Cr6wYwxhhjjDHGGGOMMcYYe/NxsJkxxhhjjDHGGGOMMcaYyjjYzBhjjDHGGGOMMcYYY0xl9WuroOTkZMTExEAsFsPS0lKYuE8qlSI1NbXCbXR1daGrq1tbTWCMMcYYY4wxxhhjjDFWR2ol2Ozj44O1a9eibdu2SEhIQF5eHlatWoXWrVsjNDQU3333XYXbzZ49G4MGDaqNJjDGGGOMMcYYY4wxxhirQyqn0cjOzsa6deuwdOlSrFu3Dnv37kWXLl2wdetWAICNjQ3Onj1b6m/Xrl0wMjJCz549a2MfSklMTER8fPxLb5eQkIDHjx+/1jrrilQqRWho6Etvl52djbCwMJXrDwwMhEKhULmct010dHSldwEwxhirO3X1vlXT9+u6VlufF95kr+uz4b179yCXy195PYwxxhhjjL0olYPNiYmJ0NTURNeuXf8tUE0NPXr0QFxcHACgXr16kEgkwp+6ujq8vLwwY8YMNG7cWPU9KGPTpk04f/78S2938uRJ3L17t8p19u7dC2dnZ2RkZNRKnXXlypUr+Pnnn196u5CQEBw9ehQAQER4+PBhpesePnwYhw4dQkpKCiZNmiQ8n5CQgH79+kEkEtWw9ZULDQ3FuHHj0LFjR0yZMqVU4Hb//v1wdnZGnz59MGnSJDx69KjSci5evIixY8fC1tYW3bt3x5w5c4TzGcXnfHZ2dq23383Nrco+fVdFR0fD2dm5wj/lNfu///0PV69eLbft6tWrcfr0aaD4Gi+5rYuLCxYuXIjY2Ngq609ISMDs2bOxe/fuStc5ffo0hgwZUmkwqqpzMy0tDfPmzYONjQ2cnZ0RHBz8wn3DGKsYEcHFxaXC1w1/f/8KtwkLC8PSpUsBADNmzMDTp0+B4tf8mrxvSaVSJCQkCI/z8/MxatQo5OXlYfv27bh06VK1ZdT0/drDw6PCfa/qdexFyy35nq5UWFiIqVOnwt3dHajm88Kreg9VUh77zZs3l1uWmJiIwYMHY8+ePbVWX11+Nnz+/Dm6dOmCoqKiF96msvYyxhhjjDFWW1ROo2FpaYkjR44IjxUKBfz8/ITgc1nnzp2DRCKBnZ2dqlVXKCAgAAMGDHjp7SZMmFDl8oSEBKxbtw4mJiaIiIhAt27dVK6zroSGhqJjx44vvV337t3RvXt3AMCtW7fg7e2N5cuXV7hu8+bNoaenBzU1Nbi6ugrPP3z4ED179kT9+rWWLhwAEBUVBRcXF2zfvh0dOnTA77//jsWLF2Pjxo04ffo0fv/9d+zbtw96eno4ePAgxowZg2vXrpUr59KlS/jxxx+xa9cutGzZEgqFAn///Te++OILBAQEoH79+pg+fTrWr18PbW3tWmt/Xl4erl69CktLy1or820RHByM/Px8bNq0qdyyJk2aoKioCDt27ICLi0u55f/88w/s7e2B4qCNjY0NxowZAxQHfi5fvoy+ffvC398fenp6FW6/Z88epKeno0ePHhW2786dO/jqq6/QqlUriMXicsurOjdzc3MxePBgTJ48GYsXL8a9e/fw5Zdfwt/fHxoaGjXqL8YY8PjxYxw+fBj379+Hurp6qWXvv/9+hdtoa2vDyckJRIRPP/0UhoaGAIAHDx7U6H3rjz/+gLm5OZo2bQoUfz4aPXo0NDQ00LZtWzRv3rzaMkJDQ/HRRx+9VL0AsGfPHixbtgwffvhhqedVnSdjz549kMlkeP78ORo1aiQ8f/jwYfj4+GD48OFANZ8XXsV7aEmPHz9GQEAAsrKyMGXKlFLLPD09ce/ePYwdO7ZW6qrrz4YPHjxAt27doKmp+ULrV9VexhhjjDHGaovKI5tLWrZsGQYMGIDw8HBMnTq13PLc3Fxs27YN48aNq81qBXl5efD19cWzZ88wZMgQODo64tSpU8Ly8PBwTJkyBQ4ODvj222+FEY1yuRyjRo1Cbm5ulfs2f/58dOvWDffv33+hOvft24fjx48L6/7888/Ytm2b8HjPnj04ffo0CgoKsHXrVgwcOBCDBw/G7t27QURA8e2R06ZNwxdffFEuQFrV/owePRr+/v4YNWoUHB0dceLECWG7wMBA6Orq4ttvv0Xfvn3h6ekp1JeWloZFixbB0dERo0aNwvXr14XtFi9ejODgYJw5cwZTp07F2bNn8dtvv5XrK3d3d2hqasLc3Bz6+vpISEgQRjiFhYXB0tISixYtQq9evTBz5kxkZmYCQJX9EBwcDE9PT2zYsAETJ04sV6enpyfmz5+Prl27QkNDA0OGDMGxY8eEbfv37y8EEz/66COEhYUJZZfk6+uL/v37o1WrVlBTU4OGhgbGjRuHc+fOoV69ehg9ejQOHjyIOXPmIDExEe7u7njw4IGw/d69e4V9PXjwIA4dOgRPT0/07dsXbm5uyMrKAopvMf7pp5/g5OSEqVOn4ubNmzA3N4eBgQFQ/KOMi4sL+vfvj+XLl0MqlQLFo8U8PT1x4MAB9O/fHyNGjEB4eHhFp+xbIzQ0FN26dYO5uXm5Pw0NDcTHxyMjIwMWFhaltsvKysKdO3eEAL6/vz+6du0KU1NTmJqawsLCAuPHj4empmapY1hSo0aNcPjwYTx//hxWVlbllsfGxuK7776Dm5sbunTpUmEZVZ2bBw4cwIcffoiRI0dCIpHA1tYWDRs2LDWSnjH28h48eAA7OztYWlqWe92QSCRYuHAhzpw5A2dnZ9y8eRN//vkn7t69CxsbG4hEIhgbG2P16tVA8ftWu3btsGrVKvTt2xfffvut8FqelJSEBQsWwNnZGUOGDBHet9atWwcPDw9s2rQJV65cgb+/PzZs2AAHBwcAgLW1NebMmSPcDbF//34MGjQI/fr1g5eXl/D+FBgYiDZt2gDFgcKpU6fiwYMHiIqKwvjx4yt8H0tLS0NISAjs7OzK7bsqwea0tDRkZWXByckJkZGRwvPZ2dnw9PSEg4MDrK2tAVT8eWHz5s3l3kNTUlLwww8/oG/fvvjiiy/g4+MjlJuVlYUFCxbgs88+g5ubG4KDgzFt2jSg+Ee8yvb/wYMHcHBwQExMDPLz84Xng4KCEB0dDRMTE7Ru3brG/VBSTT4boorPb8q+u3r1KqZPn17uM1pMTAzc3Nzg4OCAFStW4P79+8J7T1V9Ul17GWOMMcYYq1UpKSmUk5NTK38PHz6ka9eu0bRp08jNza3c8l27dtG8efNqrb6UlBTKzs4mpdDQUNLU1KRdu3ZRdnY2BQQEkI6ODmVmZtKDBw+oU6dOFBoaSkREZ8+epW7dulFRURGFhIRQp06dqDJXr14lR0dHKiwsJG9vb5oyZcoL1bl161ZavXo1ERHFx8dTv379aPLkyURElJOTQ+3ataP09HSaM2cO/fTTTySXyyk7O5ucnZ3p1KlTlJeXR8bGxhQREUFERFFRUTRixAjKz8+vcn/CwsLI0NCQNm/eTDKZjC5fvkxNmjShoqIiKioqImNjY5o1axalpqZSUlISGRkZUXh4OMlkMrK3tydvb28qLCykuLg4srS0pCdPnlBhYSEZGhrS06dPKS8vj7p27Uo3btwguVxeqq/kcjlpampSWlqa8Jyrqyv5+PgQEdHEiRPJ0dGRHj58SLm5uTR+/Hhas2YNEVGl/UBEtGvXLrKxsaHz58/T8+fPS9WZmZlJ2tralJmZKTz37Nkz0tfXJyKi27dvk5OTE0VFRdGzZ89o1qxZwnEp6/z582RmZkb79u2jpKSkcssPHz5MI0eOpOzsbMrNza1yXydNmkROTk4UERFBubm5NHToUPrzzz+FfvDw8CCFQkExMTHUqVMnmjhxIhERHTp0iIYNG0YpKSlUWFhIy5cvpwULFhAR0e7du6lNmzZ0/vx5ys/Pp9WrVwvn1Ntq0KBBdOTIkUqXnzlzhj755JNyz9++fZusra2JiEgmk5FYLKb4+PhS68THx5O2tjY9fvy40vJTUlJIU1Oz3LmelpZGffr0oYiICJozZw7t2rWr3LbVnZuOjo50+fLlUtt069YqGJA9AAAgAElEQVRNuK4ZYzXzyy+/0Lx58ypcVlhYSEZGRvTjjz9SQkIC5efn07Bhw+jcuXPCOr/99hstXryYqPi13NHRUXgtHzx4sPBaPnbsWNq7dy8VFRVRVlYWTZo0iSIjIyk2Nlb4LKBQKGjDhg20bNkyofw7d+6QnZ0dERH9+eef5OLiQs+ePaP09HT6+OOP6eLFi8L79ePHj+nhw4dkZ2dHfn5+REQUERFBe/furXD/bty4QR06dKjF3vz/cl1cXGjLli30119/Cc97enrS1q1baeDAgRQQEFDl54WS76FFRUX0zTff0JYtW6igoIBiYmJIW1ubUlNTiYjo66+/pjVr1pBMJqOYmBiysbGh8ePHV7v/v/zyC3l5edGoUaOEz09FRUX0+eef09WrV0lHR4fy8vJU7o+afjas6vOb8pjPnj2bUlJSKDk5WfiMJpVKydrams6dO0dFRUXk4+ND5ubmwntPVX1SXXsZY4wxxth/TlhdN0AVKucyUOaJU1NTg7GxsfA3dOhQpKamCiM1CwsLcejQIcyaNUv1CHklIiIi4ODgINwm36lTJ+DfgDrWr1+P77//Xhgh1LdvXwwcOBAZGRmIiIiAra1thWXm5eXh+++/h5eXF9TU1NCyZUtcuXLlheps3Lgxnjx5AhTfTvvtt9/i77//BgAcO3YMo0ePRnJyMs6cOYO7d++ifv360NDQgJOTEwIDA9G3b1+IxWJERkbCwsICFhYW2LdvHwBUuz82NjaYNGkSRCIRTE1NoaamBpFIhPj4eEilUixatAg6OjpA8a3DIpEIR44cwUcffSTc9mlmZgZbW1uEh4cjPz8fDRs2RJMmTZCTk4PAwEBYW1uXu9U/KioKLVu2FPJxExGuXbsm3D5748YNbNu2DS1btgQA2Nvb486dO3jw4EGl/eDo6IiQkBAh53JZ4eHh6Ny5s7A/KB5tpqxDJBJBIpFg6tSpCA0NhZubG+bOnVvh8e7Tpw/27t2LP/74A1OmTMHHH3+M2bNnC2lfHj16hE6dOkFLSwuhoaHV7uuvv/6KVq1aAQBMTEygpqaG2NhYHDt2DDExMahfvz7Mzc1hbW2N9u3bo6CgAO7u7vDx8RGunf79+wvXTUhICMaOHSv0Q7NmzRATE1PhvrwNCgsL4evri1u3bsHNza3UssuXL6NFixYIDQ1F586dy20bHh4upL6IjIyEQqHArl27oKb27w0daWlpOHHiBLZu3QoTE5NK2xAZGQk7O7tS57pcLseECROwcOFCtGrVCv7+/hgxYkSFbajs3MzNzcX58+dL5Q4tKCjAw4cPoa+v/9J9xRj7f/fu3YNUKsWaNWtKPe/k5ARNTU1kZ2dj1qxZMDAwQFFREa5evVpq3eDgYHzyySdA8Wu5l5eX8FpuZmYmvI6oq6vj0aNHyMnJQaNGjYS7fa5duwZ7e3vh2g8KCiqVUiEiIgJdu3ZFQUEBFi9eDB8fHxgZGQHFd8hoaWnh8ePHqFevHpKTkzFnzhz8/vvvwp0arVq1EtpTVnh4OEQiUbl9//DDD/Hpp5/WuE/Dw8PRsWNHWFlZ4eTJkwCA+Ph4HDt2DD4+PnB3d0fLli0RFxdX6eeFku+hALBt2zYUFhbi+fPnaNCgAdTV1aFQKPDo0SOcPn0aXl5e0NDQgLm5OT788EN06NCh2v1XpsnIyclBREQEWrVqBW9vb7Ro0QJaWlro3r17udQqL0uVz4ZVfX6TSqXIyMjA//73P+FuLOVnNB8fH1hbW6Nv374AgH79+iE5OVm466aqPqmuvYwxxhhjjNUmlYPNJ06cwNmzZ7Fx40bhOYlEAhTfWqkMmAUFBSE3N7dGuQdfVGhoqHCLKgDIZDJkZWVBV1cXhw4dwt27d0u1s0uXLtDS0kJISAjat29fYZk7duzA8+fPERAQgICAAOTl5SEsLAxJSUkwMjKqsk59fX1kZ2cjIyMDgYGBmD17NjZt2oTCwkJ4eXnh4MGDOHv2LKRSqfClVmnSpEkQi8Xw8fGBp6cn5syZgylTpsDNzQ3169evcn9CQ0MxYMAAYTKj2NhY4TbLBw8eYMCAAcIX4IyMDMTFxcHMzAzr1q3DjRs30LNnz1Jtee+99xAREQE7OzuIRCJERkYKdZUVEREh5GkEgKdPnyIvLw9mZmZITU3Fo0ePYGNjIyxPT0+HgYEBbt++XWk/AMDt27eFW5rLSkpKKvcFKyQkBL169YK/vz9mzZoFb29v6OnpISsrC71794aTk1O5XJZKylyTv/zyC86ePYuhQ4fC29sbnTt3xt27d/Hll19Wu69paWmIjo4ulQ8xJCQEI0aMwJ07d9CvXz80aNBAWJaYmIg2bdogMjISKSkp+Oqrr0q1SRlM9ff3h6enp/B8VFRUpV8u3waxsbHIzc1FQkJCpcGBiIiICvMpBwUFCYEJ5Rd/5fEICQmBt7c3goKC/q+9Ow+Lslz/AP6dYdhXlxxRFhVFRcUFEkHBjUxEU8J+opaWG7mUlZaJmqThgqWpmWaXkHZcOGlmmLlwobKagqgjwzKggwqI7IsMAzM8vz8O81xMA2hCcfTcn+vij3m3Z5mXmXfu93nvh39eNUcqlWoFsxlj+Pjjj+Hv7w8vLy8oFAokJibymxuNtXRuam7GNc4VrXnEWxN0IoQ8m7i4OLz77rs6N5I6duzIUytprk/u3bsHfX19rW0TEhKwdOlSFBUV4d69e1ppcjSf5WiYMO+rr76Cs7Mzpk2bhqCgIHTu3BlSqZQHGAEgPj5e6yanRCKBs7MzMjMzIRKJ0Lt3b75Ok1M6MTERenp6mDp1KpYuXfrUOf1v3boFd3d3nbY3l6v6ad26dQve3t5wdHREcHAwAGDTpk0IDg5Gfn4+HB0dYWFhgfj4+GavFxp/h8rlcnz22WeorKxEjx49UFZWBgMDA4jFYkRGRmLMmDFaN/mKi4sxYMCAJ9ZTc9NXoVAgJSUFEyZMwObNm3Hy5ElcunRJ6315Vq25Nmzp+i05ORlTpkzh3wtlZWX8Gu3777/HmDFj+D4VFRWorq5+qvPiSfUlhBBCCCGkLbU62Dxw4EBs374dEokEgwYNAmMMp0+fhrW1tdYPnatXr8LNza3NJ4ZrLDk5GfPnz+evJRIJXFxcYGFhgYKCAmRlZTUZIL1+/XqTo31yc3OxZcsWbN26VWviLzc3N6Snp0MsFjdbpiaIVFFRgSNHjmDhwoUwMTFBaWkpoqOjMX78eIjFYpSXl2Pp0qXNjrTt27cvDhw4gIKCAqxcuRLh4eGYO3dui+1JTk7G8uXL+evGP3qlUqnW5ICavJbGxsYoKSnBnj17+IRqjV24cIHvl5aW1mx+2rS0ND5aBw0T5Hh5eUEoFCIzMxP9+/fXOgcSEhIwZ84c5OTkNNsP1dXVSExMbDaoKhKJdAKRx48fxwcffICzZ89i+vTp/IebpaUlBg8eDLlcrhNsvnv3LqysrPi2pqam8PPzw8mTJ3H79m24uroiPj4eGzdufKq2enp68h/KSqUS8fHxcHR0RFJSEuzs7Ph+xcXFiI2NRVhYGB48eICRI0fyUWONKRQKfgyNlJSUJnNYvyg0E0q2NApNoVBojRxGw8hjzc0cALh9+zbGjRvHR4SPGjUK33zzDTIyMpq90aShCdxoxMXF4ZtvvkFSUhJ2796Nuro61NXVYfLkyYiKitIKXrd0bopEIp1JnX7//XetCTUJIX+d5vtx4cKFWpPYaaSmpmrd9JTJZHB3d+c3aO/fv4/09HT06tULKSkp8PT05P/HNTU1Wp/DHTp0wBdffIGgoCDs2bMHH374IX788UfcvHkTY8eOBRo+47OystCjRw+g4YaVJl90eXk57O3teV3q6+uhUChgamoKqVSKsWPHYt26dRg7dizmzp0La2vrJ7b/jz/+wJYtWzB69OhW9+Wfj7ts2TKIxWI8fPgQ8fHxqKiogLe3NyIjI/kTYo2vMxpfLzDGtL5DP/roI0ydOpVP1vfzzz9DoVBAIBAgNzcXtra2vOySkhJcuHAB3377bYt1LCgogFqthrW1NZRKJcLCwnDo0CHMmTMHXbt2bXFwwdNqzbXhk65HpVKp1qCMzMxMfo127949fk6h4ZrD1dVV5/vvWepLCCGEEEJIW2r1BIG9evXCmjVrsHz5cixYsACzZs1CZGQkgoODtS5qExIS+MQxfweVSoWYmBg+WRpjDOHh4QgMDISBgQFGjhzJJ0PRPLaak5ODuro6xMbGNjkyZMOGDVi5ciUCAgLg7+/P/8aOHQuJRNJimWj4Efro0SOcOnUKEyZMgFAohFqtxr59+7BgwQIAQO/evXHt2jWejiQ5ORm7d+8GAOzduxd5eXkAALFYjGHDhqGioqLF9qhUKsTGxmoFZm/evImBAwcCDcHJxiODGv8QdHJywo0bN/i6H3/8EdHR0Xw/TWBVJpOhV69eTb4PJSUlPJhcVFSE0NBQ/qNTKpVCLpejuLgYaAgAJicnY9y4cS32Q1ZWFoYMGdLsD6p+/fohJiaGT/AYEREBgUAALy8vdO3aFampqXzCnDt37uDSpUt8xGtj3333HTZt2qQ1oVB2djYSEhIwZMgQFBUVQS6Xo3v37k/V1sbBDM0I5I4dO8LCwoJPAMcYw86dOyEWi2FtbQ17e3tIpVI+aWJRURE+/vhj1NbWQiaTafWD5tHvF3lksybI3xIPDw9ERETwSRSrqqoQFBSESZMm8b5JSkrSOu8NDQ2xePFiHDhw4Il1+OOPP7QmB/T09ARjDImJiUhMTMTKlSuxfPlyxMXFwcjICCkpKXzyrJbOzS5dusDQ0BCZmZlAwznzww8/8MALIeTZZGRkwM3NrclAMxrSLDT+PCgvL+ef5bW1tdi6dStGjhwJIyMjnRHKMpmMf5ZHR0cjNjYWAGBiYgIPDw/+2X3r1i1+U7GiokLrxtLBgweRkpKCPn36wMbGBlKpFBUVFUDDzShNyqCUlBT83//9H3r06IHly5dj586dvL5Xr15tsm2VlZW4cuVKm38vVFZWQiaT8YC5u7s7PvzwQz7COTU1lV/jNXe98OfvUJlMhp49ewINkx/u27ePX6u89NJLSEtLQ319PVQqFbZt28a/J1tqf0ZGBjw9PSEQCGBnZwepVIoTJ07wp4WuXbvW6skBW3Nt2NL1G55wjWZlZYXc3FwAwOPHj/HVV1/xG6Et9cmT6ksIIYQQQkhba5Nhxq+88gpGjRoFuVwOfX192Nra6uTyPXToUFsU1azs7Gw4ODjA3NwcAQEBePToEVxcXPgPjJ07d2LZsmVwcHBAbm4ufH19YWdnB5lMhq5du6JLly5ax7t8+TISExOxY8cOnbKcnJwQHR39xDKtrKwQGRmJAwcO8MC7SqWCg4MDH/X9yiuv4OLFixgzZgxsbGxQWlqKr7/+GgDQs2dPTJs2DcOGDUNxcTHMzMywffv2Fttz584dWFlZaY1+avz4bnx8PP9xiIZAtCZFwJIlS7Bo0SJER0dDpVKhW7dufBb0+Ph4fPHFFwCAbt26YfPmzTA1NdVJ9zBt2jTMnj0bFy5cgL29PZydnfmP3pSUFGzcuBELFiyAhYUFMjMzcezYMZiamrbYD+np6VojS//MwcEB8+bN44+hOjs74/vvv4dAIMCsWbNw9epVjBw5EjY2Nnj48CHCwsK0RhZrBAUFYe3atejXrx969+6NiooKqNVqbNu2DUOHDkVVVRUcHBwwZcoUREREtNhWiUSiVefGKTemTJmC8PBw+Pr6QiQSYfjw4XxUnVgsRkhICLy9veHk5AS5XI7Vq1fDwMBApx+aevT7RXPjxg388ssv2Lt3r9ZykUiEu3fvwtjYGO+88w4ePHiA/v37w8bGBgUFBVi4cCHef/99AEBdXR3i4uKwa9curWPMmDEDffv2xapVq3jwQyMuLg4BAQFAw6gwHx8fCIVCxMbG8uCIhlQq5QESAFi7di0CAwPRp0+fFs9NPT097N27F35+fhCLxTA0NMThw4dplBkhrZSWlgY3N7dm18fHx2Pt2rX8tYeHBzZs2ABfX19YWVlh8ODBPM2RRCLRetonLS2Nf5bb2dlh4cKFsLW15bmVt27dCjR8fy9btgy7du2Cq6srJk6ciDFjxqBbt27w8fGBk5MTOnTogA4dOmD16tXw8vKCtbU1jIyMeN7nxt/X8+bNw7BhwzBv3jykp6dj3759OHPmjE7bNHM8tDZlRlPHHTVqFA/KDxkyBLa2tvxGfUpKCj744IMWrxfeeOMNre/QtWvXYv78+XB2dkb//v0xb948hISEYNKkSfD19cXZs2fh5eWF3r17w9nZGV5eXhAIBLh8+XKz7dfklQYAPT09ODk5YcGCBTAyMoJSqcSVK1e0Upb8VW1xbdjc9RuaODclEglvz+LFixEQEICoqCgYGhqiV69ePDDdXJ88TX0JIYQQQghpa4LCwkL250e5nxfV1dUwMjLSeRTx8ePHEAqFWjlx0TC6pKysDGZmZlqjrttCc2U+LaVSiZqaGp3Ru2q1GuXl5TA1NdUJ4P9d7amsrIRIJGqxLWVlZbC0tOSPHT9NWzTq6+t57sI/7/+kfVtSU1MDNMoZ/ud1dXV1MDc3f+JxVCoVysvLYWBgoLN9TU0NGGO8b561vpr3zsLCAnp6ejrrNRMmWVpa8omoSMtUKhWqqqqaPS//KTt37kRAQIBW0Lilc1OtVkOhUDT5SDUh5J+h+cxt6nupJYwxPjK68f+wWq3mn0do9Jlvbm7eZDqx2tpa1NTUNDsa+3nW+Hrhz9+hCoUCarWa911tbS0MDAz49R0aJqBevXo1BgwYwPM9Py/a+npUpVKhsrJSK9c/IYQQQgh5IaUB6P8U2/1XeiGDzYQQ0h4UCgXOnDkDf3//9q4KIYQ8t2bOnAk3Nzd4e3vjxo0b2LVrFy5cuPBMN6IJIYQQQggh/ywKNhNCCCGEkP8apaWl+Pe//43c3FzY2dnB39+fRvMSQgghhBDynBBkZmayph7pfB6o1WqIxeKnSo1ACCGEEEIIIYQQQggh5O8jMjc3f+Y8w+1NoVC0a25WQgghhBBCCCGEEEIIIf8hEolEbT5Z3j+lrq6uvatACCGEEEIIIYQQQgghBICwvStACCGEEEIIIYQQQggh5PlHwWZCCCGEEEIIIYQQQgghrUbBZkIIIYQQQgghhBBCCCGtJmqrAz169Ah3796Fvr4+HB0dYWZmprW+rKwMMpkMenp6Ta4nhBBCCCGEEEIIIYQQ8vxqk2Dz2bNnsX37dgwYMAB5eXlQKpXYsmUL+vXrBwC4evUqgoODIRaLUVNTA4VCobW+LeXn56Ourg52dnZ/ab+8vDyo1WrY2tr+Y2W2l6qqKuTk5GDAgAF/ab/Kykrcv38fTk5OrSr/+vXrGDRo0HM7MeXfJTs7G5aWlujcuXN7V4UQQgghhBBCCCGEkL+s1Wk0KisrsWPHDmzYsAE7duzA0aNHMXz4cOzfv59vExoaihkzZiAsLAyHDx/GmDFj8M0337S26Cbt2bMHUVFRf3m/3377DSkpKS1uc/ToUfj5+aG0tLRNymwvly9fxpdffvmX95NIJPjll18AAIwxZGZmNrvtzz//jBMnTqCwsBCBgYF8eV5eHiZMmACBQPCMtW/euXPnMHXqVAwbNgzLly/Ho0ePdLaprq7Gu+++ixMnTjR7nOjoaMydOxcjRoyAh4cHVqxYgZycHL4+Pz8flZWVbV7/ZcuWtdin/6uys7Ph5+fX5J/mf3bt2rWIiYnR2Xfbtm34/fffgYb/8cb7BgQE4LPPPoNcLm+27LS0NMyfPx9Dhw7FjBkzcPPmTb7u/v37WLlyJV599VVMnDgRBw4cAGOsyeOkpqbi7bffxtChQ7F48WIUFRXxdcXFxfj000/h6uoKPz8/3Lp1q1X9RQghhBBCCCGEENJeWh1szs/Ph4mJCdzc3P5zQKEQI0eO5MG5mpoaFBYWYsiQIRAIBBAKhXBxcWkxwNMaSUlJzzTyduHChXjttdeaXZ+Xl4cdO3ZAIBAgPT29TcpsL6mpqRg6dOhf3s/DwwNBQUEAgD/++AMHDx5sdtuePXvCxcUFQqEQs2bN4sszMzPh5eUFkajNMrgAAOLj47F+/Xps374dV65cQadOnbBmzRqtbdRqNZYvX44jR47AwsKiyeNcvHgRQUFBCAoKQkJCAi5evIiBAwdi6tSpUKlUAID333+/zYPNSqUSMTExcHR0bNPjvghu3bqF2tpa7NixQ+fPyckJ9fX1CAsLQ8eOHXX2jYiIQIcOHYCGmyyurq7YtWsXdu3ahZCQENjb28Pb21vnBhIAFBUVYdq0aZgzZw6uXbuGOXPmYNq0aaitrQUAzJw5EyNGjMC5c+cQERGBw4cP49y5czrHycrKQkBAABYvXowrV65g0KBBCA4OBgAoFAq8/vrrGDx4MOLi4rBq1Sq8+eabUCqVf0NPEkIIIYQQQgghhPy9Wh1sdnR0xMmTJ/lI1bq6OiQmJvLgs5GREby8vHDq1CkUFhYiPz8fv/76K7y9vVtf+z9RKpWIjY3Fw4cP4e/vDx8fH5w5c4avT0tLw+LFizFx4kS89957POBdU1OD2bNnQ6FQNHvsL774AmvWrIG7uztu3779VGUeO3YMv/76K9/2yy+/xPfff89fHz58GL///jtUKhX279+P1157Da+//joOHTrER0jeuHEDS5cuxdSpUzFnzhzExcU9VXveeustXLt2DbNnz4aPjw9Onz7N97t+/TqsrKzw3nvvwdvbG6Ghoby84uJirF+/Hj4+Ppg9ezbi4+P5fsHBwbh16xbOnTuHJUuW4Pz589i3b59OX61atQomJibo0aMHOnXqhLy8PD4iWiqVwtHREevXr8fo0aPxwQcfoKysDABa7Idbt24hNDQUu3fvxqJFi3TK7Ny5M8LCwuDg4AADAwN4e3tDIpFobRMSEgJnZ2cYGRk1G9SNjY2Fr68v+vbtC6FQCENDQ7z99tu4cOEC9PT08NZbb+H48eNYsWIF8vPzsWrVKmRkZPD9jx49ytt6/PhxnDhxAqGhofD29sayZctQXl4ONDwRsHHjRkyaNAlLlizBlStX0KNHD55C48KFCwgICICvry9CQkJQVVUFNIwuDw0NxU8//QRfX1/MmDEDaWlpTbblRZGamgp3d3f06NFD58/Q0BD37t1DaWkpHBwctPYrLy9HcnIyf6+vXbsGNzc32NrawtbWFg4ODpg/fz5MTEy03kMNxhgOHDiA0aNHQyQSYdy4cZDL5SgpKUF5eTni4+MxefJkAIClpSUGDx6M/Px8neOEhoZizZo1cHNzg6GhIfz9/XHq1CkAwE8//QRnZ2fMnDkTRkZGGDFiBExNTbVG0hNCCCGEEEIIIYQ8NwoLC9njx4/b5G/16tVs1KhRbPr06aygoIAvf/jwIZswYQJzcXFhLi4uzN/fn5WUlLS6vMLCQlZZWck0UlNTmYmJCTt48CCrrKxkSUlJzNLSkpWVlbGMjAzm4uLCUlNTGWOMnT9/nrm7u7P6+nomkUiYi4sLa05MTAzz8fFharWaRUZGssWLFz9Vmfv372fbtm1jjDF27949NmHCBPbuu+8yxhh7/PgxGzhwICspKWErVqxgGzduZDU1NayyspL5+fmxM2fOMKVSybp3787S09MZY4xlZWWxGTNmsNra2hbbI5VK2UsvvcT27t3Lqqur2aVLl5i1tTWrr69n9fX1rHv37uzDDz9kRUVFrKCggInFYpaWlsaqq6uZp6cni4yMZGq1muXk5DBHR0f24MEDplar2UsvvcRyc3OZUqlkbm5uLCEhgdXU1Gj1VU1NDTMxMWHFxcV82axZs9jZs2cZY4wtWrSI+fj4sMzMTKZQKNj8+fPZV199xRhjzfYDY4wdPHiQubq6sqioKFZRUdHse6Wxbt06tnnzZv46PDycrVy5kt29e5dZW1sztVrd5H5RUVHMzs6OHTt2jBUUFOis//nnn9nMmTNZZWUlUygULbY1MDCQTZo0iaWnpzOFQsGmT5/OfvzxR94PW7duZXV1dezu3bvMxcWFLVq0iDHG2IkTJ9gbb7zBCgsLmVqtZiEhIWzdunWMMcYOHTrEnJycWFRUFKutrWXbtm3j59SLatq0aezkyZPNrj937hwbN26czvKrV6+yQYMGMcYYq66uZvr6+uzevXta29y7d4+Zm5uz+/fvP7EeZ86cYRMmTGD19fWMMcbmzJnDDh48yIqLi1lCQgIbOnQoy8vL09qnrKyMmZubs7KyMr7s4cOHrFOnTowxxnx8fNilS5e09nF3d+f/14QQQgghhBBCCCHPk1aPbG7snXfewZYtWyAWi7Fx40YAQG1tLdatW4fRo0fjyJEjOHToEPr06YOQkJBm85s+q/T0dEycOBFz5syBmZkZXFxcgP8E1LFz50588sknPN2Ft7c3UlJSUFpaivT0dIwYMaLJYyqVSnzyyScICQmBUChEnz59cPny5acqs2PHjjzdQnh4ON577z0+svXUqVN466238OjRI5w7dw6ffvopDA0NYWZmhkmTJuH69esQCATQ19eHTCaDSqWCg4MDjh07Bn19/Se2x9XVFYGBgTA2NoatrS2EQiEEAgHu37+PqqoqrF+/Hp06dUKXLl1gbm4OgUCAkydPYtiwYZg8eTKEQiHs7OwwYsQIpKWlIScnB6amprC2tkZtbS2f5M/Q0FCrv7KystCnTx+e0oAxhri4OPTt2xcAkJCQgM8++wx9+vSBkZERPD09IZfLkZGR0Ww/oGFEr5+fH8aPHw9zc/Nmz4G6ujps27YNGRkZeP/99wEAUVFROHv2LEJCQiCTyeDp6QmhsOlTf/z48Th69CiioqLQr18/vP7661qjye/cuQMXFxeYmZkhOzv7iW39+OOP0bdvXxgZGcHGxgZCoRByuRynTp3Ce++9B5FIhB49emDQoEEYPHgwVO0YnqsAAAxPSURBVCoVVq1ahc2bN6Nz584QCoXw9fXldZBIJJg7dy7Gjx8PfX192NvbN9uWF4FarUZsbCyWLFkCGxsbrb+srCygYeTzyy+/rLNvWloaRo4cCQCQyWSoq6vDwYMHsWnTJmzatAkrVqyAt7c39u/fDxsbmxbrcebMGXz++efYt28fBAIBqqur4erqim+//RavvfYaAgMDcfz4cVhbW+vU4eWXX4alpSVfVlBQgD59+kChUCAqKgrOzs58nUqlQmZmJjp16tTqviOEEEIIIYQQQgj5p7U6cW59fT3QkKu5e/fu/G/69OkoKirCgwcPcOfOHXz55Zc8T+9HH30EX19f5OTkoEePHq1vRYPU1FRMnDiRv66urkZ5eTmsrKxw4sQJpKSkaE1MOHz4cJiZmUEikWDw4MFNHjMsLAwVFRVISkpCUlISlEolpFIpCgoKIBaLWyyzU6dOqKysRGlpKa5fv46PPvoIe/bsgVqtxrfffovjx4/j/PnzqKqqwrhx47TKDQwMhL6+Ps6ePYvQ0FCsWLECixcvxrJlyyASiVpsT2pqKiZPnsxTm8jlcgwfPhwAkJGRgcmTJ/PgV2lpKXJycmBnZ4cdO3YgISEBXl5eWnXp0qUL0tPTMWrUKAgEAshkMl7Wn6Wnp8PDw4O/zs3NhVKphJ2dHYqKinDnzh24urry9SUlJejcuTOuXr3abD8AwNWrV7Ft27Zm33s0pOjQTNj2r3/9C/r6+pBIJNiyZQuOHz8OAwMDpKWlPTFftYeHBzw8PPD111/j/PnzmD59OiIjI/Hyyy8jJSUFb7755hPbWlxcjOzsbLi7u/P1EokEM2bMQHJyMiZMmABjY2O+Lj8/H05OTpDJZCgsLMQ777yjVSdNMPXatWsIDQ3ly7Oysnhw+0Ukl8uhUCiQl5cHAwODJrdJT0/nQeXGbt68iSFDhvBtJk6cyN8PiUSCyMhI3Lx5E0ZGRs2WX1xcjODgYKhUKpw5c4bfWJg/fz58fHxw5coVAMAPP/yADRs24IcfftDav6CgQOf9kUgkGD16NIqKitC5c2eeUxoNkyHa2NhALBb/hV4ihBBCCCGEEEII+e/Q6mDz6dOncf78ea2gpyZ4owm0mpmZaU0IZ2xsDGNjY56rt60kJydj/vz5/LVEIoGLiwssLCxQUFCArKysJgOk169fxyuvvKKzPDc3F1u2bMHWrVuhr6/Pl7u5uSE9PR1isbjZMjVBpIqKChw5cgQLFy6EiYkJSktLER0djfHjx0MsFqO8vBxLly7FypUrm2xT3759ceDAARQUFGDlypUIDw/H3LlzW2xPcnIyli9fzl9LpVI+4loqlWoFWzMyMjBq1CgYGxujpKQEe/bsgaenp84xL1y4wPdLS0vjwes/S0tL05osMSkpCV5eXhAKhcjMzET//v21zoWEhATMmTMHOTk5zfZDdXU1EhMTWwyqRkREYP/+/di9e7dW+Z9++inkcjl8fHwAAPfv34eRkRGsra0xd+5crWPcvXsXVlZWPPhnamoKPz8/nDx5Erdv34arqyvi4+P5qP0ntdXT05OP/FYqlYiPj4ejoyOSkpJgZ2fH9ysuLkZsbCzCwsLw4MEDjBw5Er/99ptOGxUKBT+GRkpKSpM5rF8Umgklmws0o6FfGo8cRkPecs3NHAC4ffs2xo0bh/HjxwMARo0ahW+++QYZGRnN3mi6efMm3n33Xaxbtw6TJk3iy4uKinDs2DGEhYXxZcOGDcOuXbt0jiESiXTqfvz4cXzwwQcQiUQwMTHRWvf7779rTahJCCGEEEIIIYQQ8jxp9fP3AwcOxO3bt/lkbIwxnD59GtbW1rCxsYGTkxMKCwsRExPD9/ntt99gYGCA3r17t7Z4TqVSISYmhk+WxhhDeHg4AgMDYWBggJEjR/KJ/VQqFYKDg5GTk4O6ujrExsY2OWHchg0bsHLlSgQEBMDf35//jR07FhKJpMUyAaBDhw549OgRTp06hQkTJkAoFEKtVmPfvn1YsGABAKB37964du0aHyGenJyM3bt3AwD27t2LvLw8AIBYLMawYcNQUVHRYntUKhViY2O1ArM3b97EwIEDgYbg5IABA/i6xoFjJycn3Lhxg6/78ccfER0dzffTBFZlMhl69erV5PtQUlLCg8lFRUUIDQ3lQWqpVAq5XI7i4mKgIQCYnJyMcePGtdgPWVlZGDJkiE5AUePXX3/FsWPHcOrUKa3gLxrOtaysLCQmJiIxMRFGRkY4d+6cTqAZAL777jts2rQJtbW1fFl2djYSEhIwZMgQFBUVQS6Xo3v37k/V1sYjuDUjkDt27AgLCws+ARxjDDt37oRYLIa1tTXs7e0hlUr5jZiioiJ8/PHHqK2thUwm0+qH+vp6xMTEvNAjmzVB/pZ4eHggIiKCT6JYVVWFoKAgTJo0ifdNUlKS1nlvaGiIxYsX48CBA00e886dO5g/fz4OHz6sFWgGAAsLC3Tq1Imn8aivr8dPP/3Eb2ikpKRAJpMBAPr164eYmBg++WhERAQEAgG8vLzQpUsXGBoaIjMzE2g4Z3744Ycmz01CCCGEEEIIIYSQ50GrRzb36tULa9aswfLly9GrVy88fvwYABAcHAx9fX2IxWKsX78eW7duRXh4OJRKJWpra/H55583OSr3WWVnZ8PBwQHm5uYICAjAo0eP4OLiwtMR7Ny5E8uWLYODgwNyc3Ph6+sLOzs7yGQydO3aFV26dNE63uXLl5GYmIgdO3bolOXk5ITo6OgnlmllZYXIyEgcOHCAj4zW5F7W5Ih95ZVXcPHiRYwZMwY2NjYoLS3F119/DQDo2bMnpk2bhmHDhqG4uBhmZmbYvn17i+25c+cOrKystHLHxsfH8xHD8fHxCA4O5utu3rzJUzQsWbIEixYtQnR0NFQqFbp164bQ0FAwxhAfH48vvvgCANCtWzds3rwZpqamOukepk2bhtmzZ+PChQuwt7eHs7MzD/ilpKRg48aNWLBgASwsLJCZmYljx47B1NS0xX5IT0/XSkfRWH19PZYsWYKamhr069ePL3d1dcUvv/yitW1BQQGKiopgb2/f5LGCgoKwdu1a9OvXD71790ZFRQXUajW2bduGoUOHoqqqCg4ODpgyZQoiIiJabKtEItGqc+OUG1OmTEF4eDh8fX0hEokwfPhwuLu7QyAQQCwWIyQkBN7e3nBycoJcLsfq1athYGCg0w/37t2Dvr7+E/MNP89u3LiBX375BXv37tVaLhKJcPfuXRgbG+Odd97BgwcP0L9/f9jY2KCgoAALFy7kObvr6uoQFxenM/J4xowZ6Nu3L1atWsVvIGhs3bqVj6pu7Pbt27CyssKRI0fw1ltvoV+/fsjNzYWbmxvWr18PAFi7di0CAwPRp08fODg4YN68eTztjLOzM77//nsIBALo6elh79698PPzg1gshqGhIQ4fPkwpNAghhBBCCCGEEPLcEhQWFrI/P8r9LBQKBeRyOfT19WFra6szcdzjx49x//59PvFcS3lSn1Z1dTWMjIx0gtaPHz+GUCjUyomLhlGkZWVlMDMz00qL0RaaK/NpKZVK1NTU6IzeVavVKC8vh6mpqU6f/l3tqayshEgkarEtZWVlsLS05Hmhn6YtGvX19Tyv9Z/3f9K+/wSVSoXy8nIYGBjoTEZYU1MDxhjvm2etr+a9s7CwgJ6ens56tVqNiooKWFpavtATALYllUqFqqqqZs/LtqZ5D83NzbVSw+zcuRMBAQFaQeOamhqgUYqhxtRqNRQKRZvefCOEEEIIIYQQQghpD20WbG4PzQWbCSGkPSgUCpw5cwb+/v7tXRVCCCGEEEIIIYSQfxwFmwkhhBBCCCGEEEIIIYS0mujx48eoq6tr73o8k9raWhw9epRPOEf+GWZmZnwyNkIIIf9brKys+CSmhJDnj76+fpummzIwMNCaXLk1zM3NUVlZ2erj6Onp8cl5SevQdX/bou/QpllaWqK8vLy9q/FCo3Pv72FoaAilUtne1Xhh0WfD82nWrFkQCYXCJnPGPg/09PSQkpKCu3fvtndV/qd07doVDx8+bO9qEEIIaQcODg7Izs5u72oQQp6RiYkJ6uvr2+x4bflD0NbWFvfv32/1cYyNjVFaWtomdfpfZ21tjfz8/PauxgujV69euHPnTntX47+Ovb09cnJy2rsaLzS6fvt7dOjQgb5v/kY9e/akeN9z6NVXX8X/AzuPcy1PASGgAAAAAElFTkSuQmCC" alt="Vorschau Reisekostenabrechnung" style="display:block;width:100%;height:auto;min-width:920px;background:#fff;border:1px solid #d7dee8;border-radius:4px;box-shadow:0 1px 4px rgba(15,23,42,.08)">
        </div>
      </div>
    </div>
  </div>

  <!-- ── Balzer Informationen ─────────────────────────────────────────────── -->
  <div id="panel-balzer" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1280px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#2563eb 0%,#1e3a8a 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(30,58,138,.25);flex-shrink:0">&#128230;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Balzer</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">Adresse, Anfahrt und Abladehinweise direkt anzeigen, ausdrucken oder herunterladen</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentPdfOpen('balzer')" style="padding:10px 16px;background:linear-gradient(180deg,#2563eb 0%,#1e3a8a 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(30,58,138,.25);display:inline-flex;align-items:center;gap:7px">&#128424; PDF &#246;ffnen / drucken</button>
          <button onclick="documentPdfDownload('balzer')" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px">&#11015; PDF herunterladen</button>
        </div>
      </div>
      <div style="position:relative;flex:1;min-height:780px;background:#eef2f7;border:1px solid #cad7e8;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.08)">
        <iframe id="balzer-pdf-frame" title="Balzer Informationen" style="display:none;width:100%;height:calc(100vh - 170px);min-height:780px;border:0;background:#eef2f7"></iframe>
        <div id="balzer-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterstützt.</div>
          <button onclick="documentPdfOpen('balzer')" style="padding:10px 18px;background:#1e3a8a;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF &#246;ffnen / drucken</button>
        </div>
      </div>
    </div>
  </div>

  <!-- ── Termine Fahrer ─────────────────────────────────────────────── -->
  <div id="panel-termine" style="display:none;flex:1;min-height:0;overflow-y:auto;overflow-x:hidden;padding:18px;background:linear-gradient(180deg,#f3f7fb 0%,#e8f0f7 100%);font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;scrollbar-gutter:stable">
    <div style="width:100%;max-width:1280px;margin:0 auto;display:flex;flex-direction:column;gap:14px;min-height:100%">
      <div style="background:#fff;border:1px solid #cad7e8;border-radius:12px;padding:16px 20px;box-shadow:0 2px 10px rgba(30,96,145,.08);display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:12px;min-width:0">
          <div style="width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#7c3aed 0%,#5b21b6 100%);display:flex;align-items:center;justify-content:center;font-size:21px;color:#fff;box-shadow:0 2px 7px rgba(91,33,182,.25);flex-shrink:0">&#128197;</div>
          <div style="min-width:0">
            <h2 style="color:#0f172a;font-size:18px;font-weight:900;margin:0;letter-spacing:-.2px">Termine Fahrer</h2>
            <p style="color:#64748b;font-size:12px;margin:3px 0 0 0;font-weight:500">Terminübersicht direkt anzeigen, ausdrucken oder herunterladen</p>
          </div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <button onclick="documentPdfOpen('termine')" style="padding:10px 16px;background:linear-gradient(180deg,#7c3aed 0%,#5b21b6 100%);color:#fff;border:none;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;box-shadow:0 2px 6px rgba(91,33,182,.25);display:inline-flex;align-items:center;gap:7px">&#128424; PDF &#246;ffnen / drucken</button>
          <button onclick="documentPdfDownload('termine')" style="padding:10px 16px;background:#fff;color:#334155;border:1.5px solid #cbd5e1;border-radius:8px;font-weight:850;font-size:12px;cursor:pointer;font-family:inherit;display:inline-flex;align-items:center;gap:7px">&#11015; PDF herunterladen</button>
        </div>
      </div>
      <div style="position:relative;flex:1;min-height:780px;background:#eef2f7;border:1px solid #cad7e8;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.08)">
        <iframe id="termine-pdf-frame" title="Termine Fahrer" style="display:none;width:100%;height:calc(100vh - 170px);min-height:780px;border:0;background:#eef2f7"></iframe>
        <div id="termine-pdf-fallback" style="display:none;position:absolute;inset:0;align-items:center;justify-content:center;flex-direction:column;gap:12px;padding:30px;text-align:center;color:#64748b">
          <div style="font-size:42px">&#128196;</div>
          <div style="font-size:14px;font-weight:800;color:#334155">Die PDF-Vorschau wird von diesem Browser nicht unterstützt.</div>
          <button onclick="documentPdfOpen('termine')" style="padding:10px 18px;background:#5b21b6;color:#fff;border:none;border-radius:8px;font-weight:850;cursor:pointer">PDF &#246;ffnen / drucken</button>
        </div>
      </div>
    </div>
  </div>

  <div id="panel-sam" style="display:none;flex:1;overflow-y:auto;padding:14px;background:#e8ecf1;font-family:Segoe UI,Arial,sans-serif">
    <div style="width:100%;max-width:none;margin:0">
      <h2 id="sam-panel-title" style="color:#1b66b3;font-size:18px;font-weight:900;margin:0 0 4px 0">&#128664; Sa + So Einsätze</h2>
      <div style="display:inline-flex;align-items:flex-start;gap:6px;margin-bottom:9px;background:#fffbeb;border:1px solid #e2e8f0;border-radius:4px;padding:7px 12px;font-size:12px;color:#92400e;line-height:1.45;"><span>&#9888;&#65039;</span><span>Grundlage sind ausschließlich die tatsächlichen Schichten von der Fahrerkarte. Hofdienste sind NICHT berücksichtigt.<br><br>Die Zuordnung erfolgt primär über die MA-Nummer. Gezählt wird der Anfangstag: Samstag immer, Freitag ab 18&nbsp;Uhr als Fr&#8594;Sa und Sonntag bis einschließlich 15&nbsp;Uhr.</span></div>

      <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;flex-wrap:wrap">
        <input id="sam-search" placeholder="Fahrer suchen..." oninput="samFilter(this.value)"
          style="flex:1;min-width:180px;max-width:280px;padding:7px 14px;border:2px solid #1b66b3;
                 border-radius:5px;font-size:13px;font-family:inherit;outline:none">
        <select id="sam-year-sel" onchange="samYearChange(this.value)"
          style="padding:7px 12px;border:2px solid #1b66b3;border-radius:5px;font-size:12px;font-weight:700;
                 color:#1b66b3;cursor:pointer;font-family:inherit;outline:none;background:#fff;">
        </select>
        <div id="sam-list-sort-buttons" style="display:flex;gap:5px;">
          <button onclick="samSort('status')" id="sam-sort-status"
            style="padding:6px 12px;border:2px solid #1b66b3;border-radius:5px;background:#fff;color:#1b66b3;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit;">
            Status &#8593;
          </button>
          <button onclick="samSort('name')" id="sam-sort-name"
            style="padding:6px 12px;border:2px solid #1b66b3;border-radius:5px;background:#fff;color:#1b66b3;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit;">
            Name
          </button>
          <button onclick="samSort('count')" id="sam-sort-count"
            style="padding:6px 12px;border:2px solid #1b66b3;border-radius:5px;background:#1b66b3;color:#fff;font-size:11px;font-weight:700;cursor:pointer;font-family:inherit;">
            Eins&#228;tze &#8595;
          </button>
        </div>
        <button onclick="samExportExcel()" title="Aktuell angezeigte Einsätze als Excel herunterladen"
          style="padding:7px 13px;border:2px solid #15803d;border-radius:5px;background:#16a34a;color:#fff;font-size:11.5px;font-weight:850;cursor:pointer;font-family:inherit;white-space:nowrap;">&#11015;&#65039; Excel</button>
      </div>

      <div id="sam-chart-tabs" style="display:none;gap:5px;flex-wrap:wrap;margin:0 0 11px 0;">
        <button id="sam-chart-matrix" onclick="samSetChartMode('matrix')" style="padding:6px 11px;border:1px solid #1b66b3;border-radius:5px;background:#1b66b3;color:#fff;font-size:11px;font-weight:800;cursor:pointer;font-family:inherit;">Einsatzmatrix</button>
        <button id="sam-chart-drivers" onclick="samSetChartMode('drivers')" style="padding:6px 11px;border:1px solid #1b66b3;border-radius:5px;background:#fff;color:#1b66b3;font-size:11px;font-weight:800;cursor:pointer;font-family:inherit;">Einsätze je Fahrer</button>
        <button id="sam-chart-months" onclick="samSetChartMode('months')" style="padding:6px 11px;border:1px solid #1b66b3;border-radius:5px;background:#fff;color:#1b66b3;font-size:11px;font-weight:800;cursor:pointer;font-family:inherit;">Monatsverteilung</button>
      </div>

      <div id="sam-stats" style="margin-bottom:12px;"></div>
      <div id="sam-content"></div>
    </div>
  </div>
  <style>/* fahrerauswertung-neutral-schema */
    #panel-fa {{ background:#f8fafc !important; }}
    #panel-fa > div:first-child {{ background:#f1f5f9 !important; border-bottom-color:#cbd5e1 !important; }}
    #panel-fa h2 {{ color:#334155 !important; }}
    #fa-search, #fa-year-sel {{ border-color:#334155 !important; color:#334155 !important; }}
    #fa-sort-name {{ border-color:#94a3b8 !important; background:linear-gradient(180deg,#e2e8f0 0%,#cbd5e1 100%) !important; color:#334155 !important; box-shadow:0 1px 0 rgba(255,255,255,.55) inset !important; }}
    #fa-sort-name:hover {{ background:linear-gradient(180deg,#fef3c7 0%,#cbd5e1 100%) !important; color:#334155 !important; }}
    #fa-btn-10h {{ border-color:#dc2626 !important; background:#fff !important; color:#dc2626 !important; }}
    #fa-btn-10h:hover {{ background:#fee2e2 !important; color:#991b1b !important; }}
    #panel-fa > div:nth-child(2) > div:first-child {{ background:#f8fafc !important; border-right-color:#cbd5e1 !important; }}
    #fa-detail-panel {{ background:#f8fafc !important; }}
    #fa-detail-panel thead tr {{ background:linear-gradient(180deg,#f8fafc 0%,#cbd5e1 100%) !important; }}
    #fa-detail-panel .fa-tabs button {{ color:#334155; }}
    /* graue Texte in der Fahrerauswertung deutlich dunkler */
    #panel-fa, #panel-fa td, #panel-fa th, #panel-fa div, #panel-fa span {{ color:#111827; }}
    #panel-fa [style*="#94a3b8"],
    #panel-fa [style*="#64748b"],
    #panel-fa [style*="#475569"],
    #panel-fa [style*="#6b7280"],
    #panel-fa [style*="#9ca3af"],
    #panel-fa [style*="#8b94a7"],
    #panel-fa [style*="#718096"] {{ color:#1f2937 !important; }}
    #panel-fa small, #panel-fa .muted {{ color:#374151 !important; }}
    #panel-fa thead th, #panel-fa thead th * {{ color:#111827 !important; }}
    #panel-fa button[style*="background:#cbd5e1"],
    #panel-fa button[style*="background:linear-gradient"] {{ color:#111827 !important; }}
    #fa-stats, #fa-stats * {{ color:#1f2937 !important; }}
  </style>
  <!-- ── Fahrerauswertung Panel ───────────────────────────────────── -->
  <div id="panel-fa" style="display:none;flex:1;overflow:hidden;background:#f8fbff;font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;">
    <div style="display:flex;align-items:center;gap:10px;padding:10px 16px;background:#f1f5f9;border-bottom:1.5px solid #cbd5e1;flex-wrap:wrap;flex-shrink:0;">
      <h2 style="color:#334155;font-size:16px;font-weight:900;margin:0;">&#128101; Schichten / Tachograph</h2>
      <input id="fa-search" placeholder="Fahrer suchen..." oninput="faRender(this.value);"
        style="flex:1;min-width:130px;max-width:200px;padding:5px 12px;border:2px solid #cbd5e1;border-radius:5px;font-size:12px;font-family:inherit;outline:none">
      <select id="fa-year-sel" onchange="faYearChange(this.value)"
        style="padding:5px 10px;border:2px solid #cbd5e1;border-radius:5px;font-size:12px;font-weight:700;color:#334155;cursor:pointer;font-family:inherit;outline:none;background:#fff;">
        </select>
      <div style="display:flex;gap:4px;">
        <button onclick="faShowFahrerUebersicht()" id="fa-sort-name"
          style="padding:5px 12px;border:2px solid #94a3b8;border-radius:5px;background:linear-gradient(180deg,#e2e8f0 0%,#cbd5e1 100%);color:#334155;font-size:11px;font-weight:900;cursor:pointer;font-family:inherit;white-space:nowrap;">Fahrer Übersicht</button>
      </div>
      <button onclick="faShow10hTours('')" id="fa-btn-10h" title="Alle Schichten mit mehr als 10:00 Netto-Arbeitszeit anzeigen"
        style="padding:5px 12px;border:2px solid #dc2626;border-radius:5px;background:#fff;color:#dc2626;font-size:11px;font-weight:900;cursor:pointer;font-family:inherit;white-space:nowrap;">10H Touren</button>
      <div id="fa-stats" style="font-size:11px;color:#1f2937;margin-left:auto;font-weight:700;"></div>
    </div>
    <div style="display:flex;flex:1;overflow:hidden;">
      <div style="width:220px;flex-shrink:0;border-right:1.5px solid #cbd5e1;overflow-y:auto;background:#f1f5f9;">
        <div id="fa-sidebar-list"></div>
      </div>
      <div id="fa-detail-panel" style="flex:1;overflow-y:auto;padding:20px 24px;background:#f8fbff;"></div>
    </div>
  </div>

  <!-- ── Fahrerbewertung Panel (Dashboard) ───────────────────────────────── -->
  <div id="panel-fa-bewertung" style="display:none;flex:1;flex-direction:column;background:linear-gradient(180deg,#e8eef6 0%,#f3f7fb 100%);font-family:'Segoe UI',Arial,sans-serif;overflow:hidden;">
    <div style="width:100%;max-width:none;margin:0;display:flex;flex-direction:column;flex:1;overflow:hidden;">
      <div style="display:flex;align-items:center;gap:12px;padding:14px 18px;flex-wrap:wrap;flex-shrink:0;background:rgba(255,255,255,.78);border-bottom:1px solid #dbe4ef;box-shadow:0 8px 24px rgba(15,23,42,.055);backdrop-filter:blur(8px);">
        <div style="display:flex;align-items:center;gap:10px;margin-right:8px;">
          <div style="width:36px;height:36px;border-radius:12px;background:linear-gradient(135deg,#1e3a5f,#2563eb);color:#fff;display:flex;align-items:center;justify-content:center;font-size:18px;box-shadow:0 8px 18px rgba(37,99,235,.25);">&#11088;</div>
          <div>
            <h2 style="margin:0;font-size:18px;font-weight:950;color:#0f172a;letter-spacing:-.35px;line-height:1.1;">Fahrerbewertung</h2>
            <div style="font-size:11px;font-weight:750;color:#64748b;margin-top:2px;">Dashboard · Note, Ereignisse und Verbrauch</div>
          </div>
        </div>
        <select id="fabew-month" onchange="faBewMonthChange(this.value)" title="Zeitraum auswählen (ganzes Jahr oder Monat)"
          style="padding:8px 12px;border:1px solid #cbd5e1;border-radius:10px;font-size:12.5px;font-weight:900;font-family:inherit;outline:none;background:#fff;color:#1d4ed8;cursor:pointer;box-shadow:0 2px 8px rgba(15,23,42,.04);"></select>
        <input id="fabew-search" placeholder="Fahrer suchen..." oninput="faBewFilter(this.value)"
          style="flex:1;min-width:220px;max-width:520px;padding:8px 14px;border:1px solid #cbd5e1;border-radius:10px;font-size:13px;font-family:inherit;font-weight:750;outline:none;background:#fff;color:#0f172a;box-shadow:0 2px 8px rgba(15,23,42,.04);">
        <button onclick="faBewExportExcel()" title="Bewertung als Excel exportieren"
          style="padding:8px 14px;border:1px solid #15803d;border-radius:10px;font-size:12px;font-weight:950;font-family:inherit;background:linear-gradient(180deg,#16a34a,#15803d);color:#fff;outline:none;cursor:pointer;box-shadow:0 8px 16px rgba(22,163,74,.18);">&#128190; Excel-Export</button>
        <span id="fabew-stats" style="font-size:12px;font-weight:750;color:#64748b;margin-left:auto;"></span>
      </div>
      <div id="fabew-content" style="flex:1;overflow-y:auto;padding:16px 18px 34px 18px;">
        <div style="color:#94a3b8;padding:60px;text-align:center;font-size:14px;">Keine Fahrerbewertungs-Daten &ndash; bitte d_rohdaten.json in Streamlit hochladen.</div>
      </div>
    </div>
  </div>

  <div id="panel-zulage" style="display:none;flex:1;flex-direction:column;background:#e8ecf1;font-family:'Segoe UI',Arial,sans-serif;overflow:hidden;">
    <div style="display:flex;align-items:center;gap:10px;padding:12px 20px;background:#f0f2f6;border-bottom:1.5px solid #c8cfd9;flex-wrap:wrap;flex-shrink:0;">
      <h2 style="margin:0;font-size:16px;font-weight:900;color:#1b66b3;">&#128176; Zulagen</h2>
      <div style="display:flex;gap:4px;">
        <button id="ztab-sonder" onclick="zulagenTab('sonder')" style="padding:3px 9px;border-radius:3px;border:1.5px solid #1b66b3;cursor:pointer;font-weight:700;font-size:12px;background:#1b66b3;color:#fff;">Sonderfahrzeuge</button>
        <button id="ztab-fuengers" onclick="zulagenTab('fuengers')" style="padding:3px 9px;border-radius:3px;border:1.5px solid #1b66b3;cursor:pointer;font-weight:700;font-size:12px;background:#fff;color:#1b66b3;">F&#252;ngers</button>
        <button id="ztab-drittkunden" onclick="zulagenTab('drittkunden')" style="padding:3px 9px;border-radius:3px;border:1.5px solid #1b66b3;cursor:pointer;font-weight:700;font-size:12px;background:#fff;color:#1b66b3;">Drittkunden</button>
      </div>
      <select id="zulage-month-sel" onchange="zulagenRender()" style="padding:5px 12px;border:2px solid #1b66b3;border-radius:5px;font-size:12px;outline:none;cursor:pointer;"></select>
      <div style="display:flex;gap:4px;margin-left:2px;">
        <button id="zulage-view-list" onclick="zulagenSetView('list')" style="padding:5px 11px;border-radius:5px;border:1.5px solid #1b66b3;cursor:pointer;font-weight:800;font-size:11px;background:#1b66b3;color:#fff;">&#9776; Liste</button>
        <button id="zulage-view-graph" onclick="zulagenSetView('graph')" style="padding:5px 11px;border-radius:5px;border:1.5px solid #1b66b3;cursor:pointer;font-weight:800;font-size:11px;background:#fff;color:#1b66b3;">&#128200; Graph</button>
      </div>
      <button onclick="zulagenExportExcel()" style="padding:5px 12px;background:#1d6f42;color:#fff;border:none;border-radius:5px;font-weight:700;font-size:12px;cursor:pointer;">&#128196; Excel</button>
      <span id="zulage-stats" style="font-size:12px;color:#64748b;margin-left:auto;font-weight:600;"></span>
    </div>
    <div id="zulage-content" style="flex:1;overflow-y:auto;padding:20px;"></div>
    <div id="zulage-graph-content" style="display:none;flex:1;overflow-y:auto;padding:18px 20px 32px;background:linear-gradient(180deg,#eef3f8 0%,#f8fafc 100%);">
      <div style="max-width:1700px;margin:0 auto;">
        <div id="zulage-graph-kpis" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin-bottom:14px;"></div>
        <div style="display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr);gap:14px;align-items:start;" id="zulage-graph-grid">
          <div style="background:#fff;border:1px solid #d8e0ea;border-radius:13px;padding:15px 16px;box-shadow:0 3px 12px rgba(15,23,42,.06);min-width:0;">
            <div style="font-size:14px;font-weight:950;color:#0f172a;">Monatsverlauf</div>
            <div id="zulage-month-chart-sub" style="font-size:10.5px;font-weight:700;color:#64748b;margin-top:2px;margin-bottom:10px;">Gesamtsumme je Monat</div>
            <div style="height:390px;position:relative;"><canvas id="zulage-chart-months"></canvas></div>
          </div>
          <div style="background:#fff;border:1px solid #d8e0ea;border-radius:13px;padding:15px 16px;box-shadow:0 3px 12px rgba(15,23,42,.06);min-width:0;">
            <div style="font-size:14px;font-weight:950;color:#0f172a;">Zulagen nach Fahrer</div>
            <div id="zulage-driver-chart-sub" style="font-size:10.5px;font-weight:700;color:#64748b;margin-top:2px;margin-bottom:10px;">Ausgewählter Monat</div>
            <div id="zulage-driver-chart-wrap" style="height:420px;position:relative;"><canvas id="zulage-chart-drivers"></canvas></div>
          </div>
        </div>
        <div id="zulage-graph-empty" style="display:none;background:#fff;border:1px dashed #cbd5e1;border-radius:13px;padding:52px 20px;text-align:center;color:#64748b;font-size:13px;font-weight:750;">Keine Zulagendaten für diese Auswahl vorhanden.</div>
      </div>
    </div>
  </div>




  <!-- ── Spesen Panel ───────────────────────────────────── -->
  <div id="panel-spesen" style="display:none;flex:1;overflow:hidden;background:#f8fbff;font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;">
    <div style="display:flex;align-items:center;gap:10px;padding:10px 16px;background:#f1f5f9;border-bottom:1.5px solid #cbd5e1;flex-wrap:wrap;flex-shrink:0;">
      <h2 style="color:#334155;font-size:16px;font-weight:900;margin:0;">&#128181; Spesen</h2>
      <input id="spesen-search" placeholder="Fahrer suchen..." oninput="spesenFilter(this.value)"
        style="flex:1;min-width:130px;max-width:220px;padding:5px 12px;border:2px solid #cbd5e1;border-radius:5px;font-size:12px;font-family:inherit;outline:none">
      <select id="spesen-month-sel" onchange="spesenMonthChange(this.value)"
        style="padding:5px 10px;border:2px solid #cbd5e1;border-radius:5px;font-size:12px;font-weight:700;color:#334155;cursor:pointer;font-family:inherit;outline:none;background:#fff;">
      </select>
      <button onclick="spesenResetView()"
        style="padding:5px 12px;border:2px solid #cbd5e1;border-radius:5px;background:#cbd5e1;color:#111827;font-size:11px;font-weight:900;cursor:pointer;font-family:inherit;white-space:nowrap;box-shadow:0 1px 3px rgba(15,23,42,.08);">Reset</button>
      <div id="spesen-stats" style="font-size:11px;color:#64748b;margin-left:auto;font-weight:700;"></div>
    </div>
    <div style="display:flex;flex:1;overflow:hidden;">
      <div style="width:240px;flex-shrink:0;border-right:1.5px solid #cbd5e1;overflow-y:auto;background:#f8fbff;">
        <div id="spesen-sidebar-list"></div>
      </div>
      <div id="spesen-detail-panel" style="flex:1;overflow-y:auto;padding:20px 24px;background:#f8fbff;">
        <div style="color:#94a3b8;padding:60px;text-align:center;font-size:14px;">Keine Spesendaten &ndash; bitte Reisekosten-CSV in Streamlit hochladen.</div>
      </div>
    </div>
  </div>

  <!-- ── Verstoßauswertung Panel ───────────────────────────────────── -->
  <div id="panel-verstoss" style="display:none;flex:1;flex-direction:column;background:#f1f5f9;font-family:'Segoe UI',Arial,sans-serif;overflow:hidden;">
    <div style="width:100%;margin:0 auto;display:flex;flex-direction:column;flex:1;overflow:hidden;">
      <div id="verstoss-toolbar" style="display:flex;align-items:center;gap:10px;padding:16px 18px;flex-wrap:wrap;flex-shrink:0;">
        <h2 style="margin:0;font-size:17px;font-weight:900;color:#0f172a;">&#9888;&#65039; Verstoßauswertung</h2>
        <input id="verstoss-search" placeholder="Fahrer suchen..." oninput="verstossFilter(this.value)"
          style="flex:1;min-width:160px;max-width:300px;padding:8px 14px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:13px;font-family:inherit;outline:none;background:#fff;transition:border .15s;color:#0f172a;"
          onfocus="this.style.borderColor='#1e3a5f'" onblur="this.style.borderColor='#cbd5e1'">
        <select id="verstoss-list-year" onchange="verstossListYearChange(this.value)"
          style="padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:13px;font-weight:800;font-family:inherit;outline:none;background:#fff;color:#0f172a;cursor:pointer;min-width:116px;">
          <option value="all">Alle Jahre</option>
        </select>
        <button id="verstoss-search-all" onclick="verstossResetSearch()"
          style="padding:8px 14px;border:1.5px solid #1e3a5f;border-radius:8px;background:#1e3a5f;color:#fff;font-size:12px;font-weight:800;cursor:pointer;font-family:inherit;white-space:nowrap;">
          Alle anzeigen</button>
        <span id="verstoss-stats" style="font-size:12px;font-weight:700;color:#64748b;margin-left:auto;"></span>
      </div>
      <div id="verstoss-body" style="flex:1;overflow-y:auto;padding:8px 20px 30px 20px;">
        <div style="color:#94a3b8;padding:60px;text-align:center;font-size:14px;">Keine Verstoßdaten &ndash; bitte Verstoß-CSV in Streamlit hochladen.</div>
      </div>
    </div>
  </div>

  <!-- ── Verstoßauswertung Graph Panel ───────────────────────────────────── -->
  <div id="panel-verstoss-graph" style="display:none;flex:1;flex-direction:column;background:#f1f5f9;font-family:'Segoe UI',Arial,sans-serif;overflow:hidden;">
    <div style="width:100%;max-width:1728px;margin:0 auto;display:flex;flex-direction:column;flex:1;overflow:hidden;">
      <div style="display:flex;align-items:center;gap:10px;padding:16px 18px;flex-wrap:wrap;flex-shrink:0;">
        <h2 style="margin:0;font-size:17px;font-weight:900;color:#0f172a;">&#9888;&#65039; Versto&#223;auswertung &ndash; Graph pro Jahr</h2>
        <select id="verstoss-graph-mode" onchange="verstossGraphSetMode(this.value)" title="Ansicht"
          style="padding:8px 11px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12px;font-weight:900;font-family:inherit;background:#fff;color:#1f3347;outline:none;cursor:pointer;">
          <option value="single">Ein Jahr</option>
          <option value="compare">Jahre vergleichen</option>
        </select>
        <select id="verstoss-graph-year" onchange="verstossGraphYearChange(this.value)" title="Jahr 1"
          style="padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12.5px;font-weight:800;font-family:inherit;outline:none;background:#fff;color:#991b1b;cursor:pointer;"></select>
        <select id="verstoss-graph-year-2" onchange="verstossGraphYear2Change(this.value)" title="Jahr 2"
          style="display:none;padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12.5px;font-weight:800;font-family:inherit;outline:none;background:#fff;color:#b45309;cursor:pointer;"></select>
        <span style="width:1px;height:24px;background:#cbd5e1;"></span>
        <select id="verstoss-graph-type" onchange="verstossGraphTypeChange(this.value)" title="Verstoßart filtern"
          style="padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12.5px;font-weight:800;font-family:inherit;outline:none;background:#fff;color:#0f172a;cursor:pointer;max-width:280px;">
          <option value="all">Alle Verstoßarten</option>
        </select>
        <select id="verstoss-graph-mon-from" onchange="verstossGraphMonFromChange(this.value)" title="Zeitraum: Von-Monat"
          style="padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12.5px;font-weight:800;font-family:inherit;outline:none;background:#fff;color:#1f3347;cursor:pointer;"></select>
        <select id="verstoss-graph-mon-to" onchange="verstossGraphMonToChange(this.value)" title="Zeitraum: Bis-Monat"
          style="padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12.5px;font-weight:800;font-family:inherit;outline:none;background:#fff;color:#1f3347;cursor:pointer;"></select>
        <button onclick="verstossGraphResetFilters()" title="Verstoßart- und Zeitraum-Filter zurücksetzen"
          style="padding:8px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:12px;font-weight:800;font-family:inherit;background:#fff;color:#475569;outline:none;cursor:pointer;">&#10005; Filter</button>
        <button onclick="verstossExportExcel()" title="Graph-Daten als Excel exportieren"
          style="padding:8px 14px;border:1.5px solid #16a34a;border-radius:8px;font-size:12px;font-weight:900;font-family:inherit;background:#16a34a;color:#fff;outline:none;cursor:pointer;">&#128190; Excel-Export</button>
        <span id="verstoss-graph-stats" style="font-size:12px;font-weight:700;color:#64748b;margin-left:auto;"></span>
      </div>
      <div id="verstoss-graph-content" style="flex:1;overflow-y:auto;padding:4px 0 30px 18px;">
        <div style="color:#94a3b8;padding:60px;text-align:center;font-size:14px;">Keine Versto&#223;daten &ndash; bitte Versto&#223;-CSV in Streamlit hochladen.</div>
      </div>
    </div>
  </div>

  <!-- ── Großkunden Panel ───────────────────────────────────────────────────── -->
  <div id="panel-wa-heat" style="display:none;flex:1;flex-direction:column;overflow:hidden;padding:16px 18px 18px;background:linear-gradient(180deg,#eef1f5 0%,#e5e9ef 100%);font-family:'Segoe UI',Arial,sans-serif">
    <div id="wa-heat-body" style="flex:1;display:flex;flex-direction:column;min-height:0;"></div>
  </div>

  <div id="panel-wa-gruppe" style="display:none;flex:1;flex-direction:column;overflow:hidden;padding:16px 18px 18px;background:linear-gradient(180deg,#eef1f5 0%,#e5e9ef 100%);font-family:'Segoe UI',Arial,sans-serif">
    <div id="wa-gruppe-body" style="flex:1;display:flex;flex-direction:column;min-height:0;"></div>
  </div>

  <div id="panel-wa-kurve" style="display:none;flex:1;flex-direction:column;overflow:hidden;padding:16px 18px 18px;background:linear-gradient(180deg,#eef1f5 0%,#e5e9ef 100%);font-family:'Segoe UI',Arial,sans-serif">
    <div id="wa-kurve-body" style="flex:1;display:flex;flex-direction:column;min-height:0;"></div>
  </div>

  <div id="panel-wa-rhythm" style="display:none;flex:1;flex-direction:column;overflow:hidden;padding:16px 18px 18px;background:linear-gradient(180deg,#eef1f5 0%,#e5e9ef 100%);font-family:'Segoe UI',Arial,sans-serif">
    <div id="wa-rhythm-body" style="flex:1;display:flex;flex-direction:column;min-height:0;"></div>
  </div>

  <div id="panel-gk" style="display:none;flex:1;overflow:hidden;font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;">
    <style>
      #panel-gk{{--ink:#0f1f33;background:linear-gradient(180deg,#eef3f9 0%,#f5f8fc 100%);}}
      #panel-gk .gk-head{{flex-shrink:0;background:linear-gradient(180deg,#fbfdff 0%,#edf3fa 100%);border-bottom:1px solid #d4e0ee;padding:13px 18px 12px;}}
      #panel-gk .gk-titlerow{{display:flex;align-items:center;gap:13px;flex-wrap:wrap;margin-bottom:11px;}}
      #panel-gk .gk-title{{display:flex;align-items:center;gap:9px;font-size:16.5px;font-weight:900;letter-spacing:-.3px;color:var(--ink);white-space:nowrap;}}
      #panel-gk .gk-dot{{width:9px;height:9px;border-radius:50%;background:linear-gradient(135deg,#1e6091,#3aa0d8);box-shadow:0 0 0 4px rgba(30,96,145,.12);}}
      #panel-gk .gk-count{{font-size:11.5px;font-weight:800;color:#1e6091;background:#e8f1fb;border:1px solid #cfe0f1;border-radius:999px;padding:2px 9px;font-variant-numeric:tabular-nums;}}
      #panel-gk .gk-searchwrap{{position:relative;flex:1;min-width:200px;max-width:380px;}}
      #panel-gk .gk-sicon{{position:absolute;left:12px;top:50%;transform:translateY(-50%);font-size:13px;opacity:.5;pointer-events:none;}}
      #panel-gk .gk-search{{width:100%;padding:9px 13px 9px 33px;border:1.5px solid #cdddee;border-radius:10px;font-size:13px;font-family:inherit;font-weight:600;outline:none;background:#fff;color:var(--ink);box-shadow:0 1px 3px rgba(30,96,145,.05);transition:border-color .14s,box-shadow .14s;}}
      #panel-gk .gk-search:focus{{border-color:#1e6091;box-shadow:0 0 0 3px rgba(30,96,145,.13);}}
      #panel-gk .gk-stat{{margin-left:auto;font-size:11.5px;font-weight:700;color:#7c8ca0;white-space:nowrap;font-variant-numeric:tabular-nums;}}
      #panel-gk .gk-tiles{{display:flex;flex-wrap:wrap;gap:7px;max-height:146px;overflow-y:auto;padding:2px 4px 2px 0;}}
      #panel-gk .gk-tile{{display:flex;align-items:center;gap:9px;cursor:pointer;font-family:inherit;text-align:left;border:1.5px solid #dbe4ef;background:#fff;border-radius:11px;padding:7px 12px 7px 8px;min-width:150px;max-width:250px;user-select:none;transition:transform .12s,box-shadow .12s,border-color .12s;}}
      #panel-gk .gk-tile:hover{{transform:translateY(-1px);box-shadow:0 6px 16px rgba(15,31,51,.10);border-color:#c3d4e7;}}
      #panel-gk .gk-tile:focus-visible{{outline:none;border-color:var(--gk);box-shadow:0 0 0 3px color-mix(in srgb,var(--gk) 28%,transparent);}}
      #panel-gk .gk-tile.is-active{{border-color:var(--gk);box-shadow:0 8px 20px color-mix(in srgb,var(--gk) 22%,rgba(15,31,51,.10));}}
      #panel-gk .gk-tile-txt{{min-width:0;display:flex;flex-direction:column;gap:3px;}}
      #panel-gk .gk-tile-name{{font-size:12.5px;font-weight:800;color:var(--ink);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;line-height:1.2;}}
      #panel-gk .gk-tile-meta{{font-size:10.5px;font-weight:700;color:#7c8ca0;}}
      #panel-gk .gk-mono{{flex-shrink:0;width:30px;height:30px;border-radius:9px;display:flex;align-items:center;justify-content:center;font-size:11.5px;font-weight:900;letter-spacing:.3px;background:var(--gks);color:var(--gkd);}}
      #panel-gk .gk-detail{{flex:1;overflow-y:auto;padding:18px 20px 34px;}}
      #panel-gk .gk-fade{{animation:gkFade .22s ease-out;}}
      @keyframes gkFade{{from{{opacity:0;transform:translateY(6px);}}to{{opacity:1;transform:none;}}}}
      @media (prefers-reduced-motion:reduce){{#panel-gk .gk-fade{{animation:none;}}#panel-gk .gk-tile{{transition:none;}}}}
      #panel-gk .gk-hero{{display:flex;align-items:center;gap:16px;background:linear-gradient(135deg,var(--gks) 0%,#fff 72%);border:1px solid var(--gkb);border-left:5px solid var(--gk);border-radius:16px;padding:16px 20px;margin-bottom:14px;box-shadow:0 10px 26px rgba(15,31,51,.06);flex-wrap:wrap;}}
      #panel-gk .gk-hero-mono{{flex-shrink:0;width:54px;height:54px;border-radius:15px;display:flex;align-items:center;justify-content:center;font-size:20px;font-weight:900;letter-spacing:.4px;background:var(--gk);color:#fff;box-shadow:0 6px 16px color-mix(in srgb,var(--gk) 40%,transparent);}}
      #panel-gk .gk-hero-txt{{min-width:0;flex:1;}}
      #panel-gk .gk-hero-name{{font-size:23px;font-weight:900;letter-spacing:-.5px;color:var(--ink);line-height:1.1;}}
      #panel-gk .gk-hero-sub{{font-size:12.5px;font-weight:700;color:var(--gkd);margin-top:3px;}}
      #panel-gk .gk-actions{{display:flex;gap:8px;flex-wrap:wrap;margin-left:auto;}}
      #panel-gk .gk-act{{display:inline-flex;align-items:center;gap:7px;background:var(--gk);color:#fff;border:none;border-radius:9px;padding:8px 14px;font-size:12.5px;font-weight:800;text-decoration:none;white-space:nowrap;cursor:pointer;box-shadow:0 4px 12px color-mix(in srgb,var(--gk) 32%,transparent);transition:filter .12s,transform .12s;}}
      #panel-gk .gk-act:hover{{filter:brightness(1.07);transform:translateY(-1px);}}
      #panel-gk .gk-act b{{font-weight:900;opacity:.85;}}
      #panel-gk .gk-act-sm{{background:#fff;color:var(--gkd);border:1.5px solid var(--gkb);box-shadow:none;}}
      #panel-gk .gk-act-sm:hover{{background:var(--gks);filter:none;}}
      #panel-gk .gk-card{{background:#fff;border:1px solid #e2e8f0;border-radius:14px;overflow:hidden;box-shadow:0 4px 14px rgba(15,31,51,.045);margin-bottom:14px;}}
      #panel-gk .gk-card-h{{display:flex;align-items:center;gap:10px;padding:11px 16px;border-bottom:1px solid #eef2f7;background:linear-gradient(180deg,#fbfdff,#fff);}}
      #panel-gk .gk-eyebrow{{display:flex;align-items:center;gap:8px;font-size:11px;font-weight:900;text-transform:uppercase;letter-spacing:.8px;color:var(--gkd);}}
      #panel-gk .gk-eyebrow::before{{content:"";width:7px;height:7px;border-radius:50%;background:var(--gk);}}
      #panel-gk .gk-cnt{{margin-left:auto;font-size:11px;font-weight:800;color:#94a3b8;font-variant-numeric:tabular-nums;}}
      #panel-gk .gk-card-b{{padding:4px 16px 10px;}}
      #panel-gk .gk-loc{{display:flex;align-items:flex-start;gap:18px;flex-wrap:wrap;padding:12px 0;border-top:1px solid #f1f5f9;}}
      #panel-gk .gk-loc:first-child{{border-top:none;}}
      #panel-gk .gk-loc-name{{min-width:220px;flex:0 0 auto;display:flex;align-items:center;gap:10px;flex-wrap:wrap;font-size:13.5px;font-weight:800;color:var(--ink);}}
      #panel-gk .gk-loc-addr{{flex:1;font-size:12.5px;color:#475569;line-height:1.6;}}
      #panel-gk .gk-loc-chips{{display:flex;flex-wrap:wrap;gap:8px;padding:10px 0 4px;}}
      #panel-gk .gk-loc-chip{{display:inline-flex;align-items:center;gap:9px;padding:7px 12px;background:#f7fafd;border:1px solid #e2e8f0;border-radius:10px;font-size:13px;font-weight:800;color:var(--ink);}}
      #panel-gk .gk-knr{{display:inline-flex;align-items:baseline;gap:5px;background:var(--gks);border:1px solid var(--gkb);border-radius:6px;padding:2px 8px;}}
      #panel-gk .gk-knr-l{{font-size:8.5px;font-weight:900;text-transform:uppercase;letter-spacing:.5px;color:var(--gkd);}}
      #panel-gk .gk-knr-v{{font-size:12.5px;font-weight:800;color:var(--ink);font-variant-numeric:tabular-nums;}}
      #panel-gk .gk-grid{{display:grid;grid-template-columns:minmax(340px,1.1fr) minmax(260px,.9fr);gap:14px;align-items:start;}}
      @media (max-width:880px){{#panel-gk .gk-grid{{grid-template-columns:1fr;}}}}
      #panel-gk .gk-c-sec{{margin:13px 0 6px;display:flex;align-items:center;gap:10px;flex-wrap:wrap;padding-bottom:6px;border-bottom:1px solid #e8eef5;}}
      #panel-gk .gk-c-sec-l{{font-size:10.5px;font-weight:900;text-transform:uppercase;letter-spacing:.7px;color:var(--gkd);}}
      #panel-gk .gk-c-row{{display:flex;align-items:center;gap:12px;padding:8px 0;border-bottom:1px solid #f4f7fa;}}
      #panel-gk .gk-c-row:last-child{{border-bottom:none;}}
      #panel-gk .gk-c-left{{flex:1;min-width:0;}}
      #panel-gk .gk-mail{{color:var(--gkd);font-size:12.5px;font-weight:700;text-decoration:none;}}
      #panel-gk .gk-mail:hover{{text-decoration:underline;}}
      #panel-gk .gk-c-label{{font-size:12.5px;font-weight:700;color:var(--ink);}}
      #panel-gk .gk-c-dash{{color:#cbd5e1;font-size:12px;}}
      #panel-gk .gk-tel{{display:inline-flex;align-items:center;gap:6px;font-size:12.5px;font-weight:800;color:var(--gkd);font-variant-numeric:tabular-nums;text-decoration:none;white-space:nowrap;}}
      #panel-gk .gk-tel-i{{color:var(--gk);font-weight:400;}}
      #panel-gk .gk-c-other{{display:flex;gap:12px;align-items:baseline;padding:7px 0;border-bottom:1px solid #f4f7fa;}}
      #panel-gk .gk-c-other-l{{font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.5px;color:#94a3b8;min-width:90px;flex-shrink:0;}}
      #panel-gk .gk-c-other-v{{font-size:12.5px;color:#334155;line-height:1.5;}}
      #panel-gk .gk-hint{{display:flex;align-items:flex-start;gap:12px;padding:10px 16px;border-top:1px solid #f1f5f9;}}
      #panel-gk .gk-hint:first-child{{border-top:none;}}
      #panel-gk .gk-hint-n{{flex-shrink:0;width:20px;height:20px;border-radius:6px;background:var(--gks);color:var(--gkd);font-size:11px;font-weight:900;display:flex;align-items:center;justify-content:center;font-variant-numeric:tabular-nums;margin-top:1px;}}
      #panel-gk .gk-hint-t{{font-size:12.5px;color:#334155;line-height:1.55;}}
      #panel-gk .gk-ff-row{{padding:9px 16px;border-top:1px solid #f4f7fa;font-size:12.5px;color:#334155;}}
      #panel-gk .gk-ff-row:first-child{{border-top:none;}}
      #panel-gk .gk-ff-long{{background:#f8fbff;line-height:1.6;}}
      #panel-gk .gk-empty{{color:#94a3b8;padding:70px;text-align:center;font-size:14px;}}
    </style>
    <!-- Kopf / Suche -->
    <div class="gk-head">
      <div class="gk-titlerow">
        <div class="gk-title"><span class="gk-dot"></span>Gro&#223;kunden <span id="gk-count" class="gk-count"></span></div>
        <div class="gk-searchwrap">
          <span class="gk-sicon">&#128269;</span>
          <input id="gk-search" class="gk-search" placeholder="Gro&#223;kunden suchen..." oninput="gkFilter(this.value)">
        </div>
        <span id="gk-stats" class="gk-stat"></span>
      </div>
      <div id="gk-tiles" class="gk-tiles"></div>
    </div>
    <!-- Detail unten -->
    <div id="gk-detail" class="gk-detail">
      <div class="gk-empty">Keine Gro&#223;kundendaten &ndash; bitte Excel in Streamlit hochladen.</div>
    </div>
  </div>

  <!-- ── Spediteure Panel ───────────────────────────────────────────────────── -->
  <div id="panel-sped" style="display:none;flex:1;overflow:hidden;background:#f3f7fb;font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;">
    <div style="flex-shrink:0;background:linear-gradient(180deg,#f8fbff 0%,#edf4fb 100%);border-bottom:1.5px solid #c6d6e8;padding:12px 18px;box-shadow:0 1px 5px rgba(30,96,145,.08);">
      <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
        <div style="display:flex;align-items:center;gap:10px;">
          <div style="width:34px;height:34px;border-radius:9px;background:linear-gradient(135deg,#1e6091 0%,#2f80b7 100%);display:flex;align-items:center;justify-content:center;font-size:17px;">&#128666;</div>
          <div>
            <div style="font-size:14px;font-weight:900;color:#0f172a;letter-spacing:-.2px;">Spediteure</div>
            <div style="font-size:11px;color:#64748b;">Tourenpl&auml;ne nach Spedition &middot; Jahr / Monat / Untername</div>
          </div>
        </div>
        <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-left:8px;">
          <select id="sped-year" onchange="spedSetJahr(this.value)" title="Jahr" style="padding:8px 11px;border:1.5px solid #b9cce3;border-radius:8px;font-size:12px;font-weight:800;font-family:inherit;background:#fff;color:#1f3347;outline:none;"></select>
          <select id="sped-month" onchange="spedSetMonat(this.value)" title="Monat" style="padding:8px 11px;border:1.5px solid #b9cce3;border-radius:8px;font-size:12px;font-weight:800;font-family:inherit;background:#fff;color:#1f3347;outline:none;"></select>
          <select id="sped-filter" onchange="spedSetSped(this.value)" title="Spedition" style="padding:8px 11px;border:1.5px solid #b9cce3;border-radius:8px;font-size:12px;font-weight:800;font-family:inherit;background:#fff;color:#1f3347;outline:none;max-width:240px;"></select>
        </div>
      </div>
      <div id="sped-stats" style="display:flex;gap:7px;flex-wrap:wrap;margin-top:10px;"></div>
    </div>
    <div id="sped-content" style="flex:1;overflow-y:auto;padding:16px 18px 28px;background:#f3f7fb;">
      <div style="color:#94a3b8;padding:70px;text-align:center;font-size:14px;">Keine Spediteur-Daten &ndash; bitte Touren-Dateien (Excel) in Streamlit hochladen.</div>
    </div>
  </div>

  <!-- ── Spediteure Graph Panel ─────────────────────────────────────────────── -->
  <div id="panel-sped-graph" style="display:none;flex:1;overflow:hidden;background:#f3f7fb;font-family:'Segoe UI',Arial,sans-serif;flex-direction:column;">
    <div style="flex-shrink:0;background:linear-gradient(180deg,#f8fbff 0%,#edf4fb 100%);border-bottom:1.5px solid #c6d6e8;padding:12px 18px;box-shadow:0 1px 5px rgba(30,96,145,.08);">
      <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
        <div style="display:flex;align-items:center;gap:10px;">
          <div style="width:34px;height:34px;border-radius:9px;background:linear-gradient(135deg,#1e6091 0%,#2f80b7 100%);display:flex;align-items:center;justify-content:center;font-size:17px;">&#128202;</div>
          <div>
            <div style="font-size:14px;font-weight:900;color:#0f172a;letter-spacing:-.2px;">Spediteure Graph</div>
            <div style="font-size:11px;color:#64748b;">Grafische Auswertung der Fahrten</div>
          </div>
        </div>
        <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-left:8px;">
          <select id="sped-graph-mode" onchange="spedGraphSetMode(this.value)" title="Ansicht" style="padding:8px 11px;border:1.5px solid #b9cce3;border-radius:8px;font-size:12px;font-weight:900;font-family:inherit;background:#fff;color:#1f3347;outline:none;">
            <option value="single">Ein Jahr</option>
            <option value="compare">Jahre vergleichen</option>
          </select>
          <select id="sped-graph-year" onchange="spedGraphSetJahr(this.value)" title="Jahr 1" style="padding:8px 11px;border:1.5px solid #b9cce3;border-radius:8px;font-size:12px;font-weight:800;font-family:inherit;background:#fff;color:#1f3347;outline:none;"></select>
          <select id="sped-graph-year-2" onchange="spedGraphSetJahr2(this.value)" title="Jahr 2" style="display:none;padding:8px 11px;border:1.5px solid #b9cce3;border-radius:8px;font-size:12px;font-weight:800;font-family:inherit;background:#fff;color:#1f3347;outline:none;"></select>
          <select id="sped-graph-month" onchange="spedGraphSetMonat(this.value)" title="Monat" style="padding:8px 11px;border:1.5px solid #b9cce3;border-radius:8px;font-size:12px;font-weight:800;font-family:inherit;background:#fff;color:#1f3347;outline:none;"></select>
          <select id="sped-graph-filter" onchange="spedGraphSetSped(this.value)" title="Spedition" style="padding:8px 11px;border:1.5px solid #b9cce3;border-radius:8px;font-size:12px;font-weight:800;font-family:inherit;background:#fff;color:#1f3347;outline:none;max-width:240px;"></select>
          <button onclick="spedGraphExportExcel()" title="Aktuellen Spediteursgraphen als Excel exportieren" style="padding:8px 13px;border:1.5px solid #16a34a;border-radius:8px;font-size:12px;font-weight:900;font-family:inherit;background:linear-gradient(180deg,#ecfdf5 0%,#dcfce7 100%);color:#166534;cursor:pointer;box-shadow:0 2px 7px rgba(22,163,74,.12);">&#128190; Excel Export</button>
        </div>
      </div>
      <div id="sped-graph-stats" style="display:flex;gap:7px;flex-wrap:wrap;margin-top:10px;"></div>
    </div>
    <div id="sped-graph-content" style="flex:1;overflow-y:auto;padding:16px 18px 28px;background:#f3f7fb;">
      <div style="color:#94a3b8;padding:70px;text-align:center;font-size:14px;">Keine Spediteur-Daten &ndash; bitte Touren-Dateien (Excel) in Streamlit hochladen.</div>
    </div>
  </div>

</div>
</div>


<script>
// ── Base64-Chunks → UTF-8-String (zlib-compressed) ──────────────────────────
async function b64ChunksToString(chunks) {{
  var b64 = chunks.join("");
  var bin = atob(b64);
  var bytes = new Uint8Array(bin.length);
  for (var i = 0; i < bin.length; i++) {{ bytes[i] = bin.charCodeAt(i); }}
  var ds = new DecompressionStream("deflate");
  var writer = ds.writable.getWriter();
  writer.write(bytes);
  writer.close();
  return new Response(ds.readable).text();
}}

// ── Komprimierte Zusatzdaten ────────────────────────────────────────────────
var EMBEDDED_DATA_B64 = [
{embedded_data_js}
];

// Sichere Startwerte, bis der komprimierte Datenblock entpackt ist.
var FAHRZEUGWAESCHE_DATA = [];
var TANK_DATA = [];
var TEL_DATA = [];
var SAM_DATA = [];
var FA_DATA = [];
var FAHRERBEWERTUNG_DATA = {{profile:"",event_types:[],g_months:{{}},g_ev:{{}},drivers:[]}};
var ZULAGE_DATA = {{}};
var DRITTKUNDEN_DATA = [];
var VERSTOSS_DATA = {{drivers:[],total_violations:0}};
var SPESEN_DATA = {{drivers:[],months:[],total_cost:0,total_rows:0}};
var GK_DATA = [];
var TIMEREC_DATA = {{}};
var SPED_DATA = {{katalog:[],fahrten:[]}};
var VERSP_ABFAHRT = {{}};
var _embeddedDataPromise = null;

function loadEmbeddedData() {{
  if(_embeddedDataPromise) return _embeddedDataPromise;
  _embeddedDataPromise = (async function() {{
    var raw = await b64ChunksToString(EMBEDDED_DATA_B64);
    var data = JSON.parse(raw);
    FAHRZEUGWAESCHE_DATA = data.fahrzeugwaesche || [];
    TANK_DATA = data.tanken || [];
    TEL_DATA = data.telefon || [];
    SAM_DATA = data.samstag || [];
    FA_DATA = data.fahrer || [];
    FAHRERBEWERTUNG_DATA = data.fahrerbewertung || {{profile:"",event_types:[],g_months:{{}},g_ev:{{}},drivers:[]}};
    ZULAGE_DATA = data.zulagen || {{}};
    DRITTKUNDEN_DATA = data.drittkunden || [];
    VERSTOSS_DATA = data.verstoesse || {{drivers:[],total_violations:0}};
    SPESEN_DATA = data.spesen || {{drivers:[],months:[],total_cost:0,total_rows:0}};
    GK_DATA = data.grosskunden || [];
    TIMEREC_DATA = data.timerecording || {{}};
    SPED_DATA = data.spediteure || {{katalog:[],fahrten:[]}};
    VERSP_ABFAHRT = data.verspaetung_abfahrt || {{}};
    EMBEDDED_DATA_B64 = [];
    return true;
  }})().catch(function(err) {{
    console.error("Zusatzdaten konnten nicht geladen werden:", err);
    _embeddedDataPromise = null;
    throw err;
  }});
  return _embeddedDataPromise;
}}

// ── Instanzen ─────────────────────────────────────────────────────────────────
var INSTANCES = [
{instances_js}
];
var currentInst = 0;

// switchInst: ersetzt durch ddSelect()

// ── Initiales Laden ───────────────────────────────────────────────────────────
var currentInst = 0;
var currentArea = "suche";

async function loadInst(i) {{
  await loadEmbeddedData();
  currentInst = i;
  var inst = INSTANCES[i];
  document.getElementById("frame-suche").srcdoc = await b64ChunksToString(inst.s);
  document.getElementById("frame-druck" ).srcdoc = await b64ChunksToString(inst.d);
  vzAllData = null;
  // Dropdown-Items aktualisieren
  ["suche","vz"].forEach(function(area) {{
    var menu = document.getElementById("ddmenu-"+area);
    if(!menu) return;
    if(area === "vz") {{ if(typeof buildVzDdMenu === "function") buildVzDdMenu(); }}
    else buildDdMenu(area);
  }});
  // Wochenauslastung neu berechnen lassen (Daten haengen an der Instanz)
  var _wh = document.getElementById("panel-wa-heat");   if(_wh) _wh.dataset.loaded = "";
  var _wg = document.getElementById("panel-wa-gruppe"); if(_wg) _wg.dataset.loaded = "";
  var _wk = document.getElementById("panel-wa-kurve");  if(_wk) _wk.dataset.loaded = "";
  var _wr = document.getElementById("panel-wa-rhythm"); if(_wr) _wr.dataset.loaded = "";
  if(currentArea === "wa_heat"   && typeof waInitHeat   === "function") {{ waInitHeat();   if(_wh) _wh.dataset.loaded="1"; }}
  if(currentArea === "wa_gruppe" && typeof waInitGruppe === "function") {{ waInitGruppe(); if(_wg) _wg.dataset.loaded="1"; }}
  if(currentArea === "wa_kurve"  && typeof waInitKurve  === "function") {{ waInitKurve();  if(_wk) _wk.dataset.loaded="1"; }}
  if(currentArea === "wa_rhythm" && typeof waInitRhythm === "function") {{ waInitRhythm(); if(_wr) _wr.dataset.loaded="1"; }}
}}

function buildSamDdMenu() {{
  var menu = document.getElementById("ddmenu-sam");
  if(!menu) return;
  var items = [
    {{ id: "sam", label: "Liste" }},
    {{ id: "sam_graph", label: "Graph" }}
  ];
  menu.innerHTML = items.map(function(it){{
    var active = (currentArea === it.id) ? " active" : "";
    return "<div class='dd-item" + active + "' data-area='" + it.id + "' onclick='ddSelectSam(this.dataset.area)'>" + it.label + "</div>";
  }}).join("");
}}

function ddSelectSam(area) {{
  showArea(area);
  document.querySelectorAll(".nav-dd").forEach(function(d){{ d.classList.remove("open"); }});
}}

function buildDdMenu(area) {{
  var menu = document.getElementById("ddmenu-"+area);
  if(!menu) return;
  var html = "";
  INSTANCES.forEach(function(inst, i) {{
    var cls = "dd-item" + (i===currentInst ? " active" : "");
    html += "<div class='" + cls + "'"
          + " data-area='" + area + "'"
          + " data-idx='" + i + "'"
          + " onclick='ddSelect(this.dataset.area,+this.dataset.idx)'>"
          + inst.name + "</div>";
  }});
  menu.innerHTML = html;
}}

function ddToggle(area, e) {{
  e.stopPropagation();
  var dd   = document.getElementById("dd-"+area);
  var menu = document.getElementById("ddmenu-"+area);
  var wasOpen = dd.classList.contains("open");
  document.querySelectorAll(".nav-dd").forEach(function(d){{d.classList.remove("open");}});
  if(!wasOpen) {{
    if(area === "verstoss") buildVerstossDdMenu();
    else if(area === "sam") buildSamDdMenu();
    else if(area === "sped") buildSpedDdMenu();
    else if(area === "infos") buildInfosDdMenu();
    else if(area === "vz") buildVzDdMenu();
    else if(area === "fa") buildFaDdMenu();
    else if(area === "wa") buildWaDdMenu();
    else buildDdMenu(area);
    dd.classList.add("open");
    // Position unter dem Button berechnen (fixed, ignoriert iframe)
    var btn  = document.getElementById("btn-"+area);
    var rect = btn.getBoundingClientRect();
    menu.style.top  = (rect.bottom + 4) + "px";
    menu.style.left = rect.left + "px";
  }}
}}

function updateInstLabels() {{
  if(INSTANCES.length <= 1) return;  // nur anzeigen wenn mehrere Instanzen
  var name = INSTANCES[currentInst].name;
  ["suche","vz"].forEach(function(area) {{
    var el = document.getElementById("inst-label-"+area);
    if(el) {{
      el.textContent = name;
      el.className = "inst-label";
    }}
  }});
}}

function ddSelect(area, instIdx) {{
  if(instIdx !== currentInst) loadInst(instIdx);
  updateInstLabels();
  showArea(area);
  document.querySelectorAll(".nav-dd").forEach(function(d){{d.classList.remove("open");}});
}}

// Klick außerhalb schließt Dropdown
document.addEventListener("click", function() {{
  document.querySelectorAll(".nav-dd").forEach(function(d){{d.classList.remove("open");}});
}});

// ── Navigation ────────────────────────────────────────────────────────────────
function showArea(s) {{
  currentArea = s;
  // iframes
  document.getElementById("frame-suche").className = (s==="suche")?"active":"";
  // Dropdown-Buttons aktiv/inaktiv
  ["suche","vz"].forEach(function(id) {{
    var btn = document.getElementById("btn-"+id);
    if(btn) btn.className = "nav-dd-btn" + ((id===s || (id==="vz" && (s==="vz_graph" || s==="tank" || s==="tank_graph")))?" active":"");
  }});
  // Telefonliste-Button
  var telBtn = document.getElementById("btn-tel");
  if(telBtn) telBtn.className = "nav-btn" + (s==="tel"?" active":"");
  // Sa + So Einsätze (Dropdown: Liste + Graph)
  var samBtn = document.getElementById("btn-sam");
  if(samBtn) samBtn.className = "nav-dd-btn" + ((s==="sam" || s==="sam_graph")?" active":"");
  if(typeof buildSamDdMenu === "function") buildSamDdMenu();
  // Fahrerauswertung-Button (Dropdown: Schichten + Fahrerbewertung)
  var faBtn = document.getElementById("btn-fa");
  if(faBtn) faBtn.className = "nav-dd-btn" + ((s==="fa" || s==="fa_bewertung")?" active":"");
  if(typeof buildFaDdMenu === "function") buildFaDdMenu();
  // Panels
  var vzPanel       = document.getElementById("panel-vz");
  var telPanel      = document.getElementById("panel-tel");
  var samPanel      = document.getElementById("panel-sam");
  vzPanel.style.display  = (s==="vz")  ? "block" : "none";
  var vzGraphPanel = document.getElementById("panel-vz-graph");
  if(vzGraphPanel) vzGraphPanel.style.display = (s==="vz_graph") ? "flex" : "none";
  var tankPanel = document.getElementById("panel-tank");
  if(tankPanel) tankPanel.style.display = (s==="tank") ? "block" : "none";
  var tankGraphPanel = document.getElementById("panel-tank-graph");
  if(tankGraphPanel) tankGraphPanel.style.display = (s==="tank_graph") ? "block" : "none";
  telPanel.style.display = (s==="tel") ? "block" : "none";
  if(samPanel)      samPanel.style.display      = (s==="sam" || s==="sam_graph") ? "block" : "none";
  var faPanel = document.getElementById("panel-fa");
  if(faPanel) faPanel.style.display = (s==="fa") ? "flex" : "none";
  var zulagePanel = document.getElementById("panel-zulage");
  if(zulagePanel) zulagePanel.style.display = (s==="zulage") ? "flex" : "none";
  var zuBtn = document.getElementById("btn-zulage");
  if(zuBtn) zuBtn.className = "nav-btn" + (s==="zulage" ? " active" : "");
  var spesenBtn = document.getElementById("btn-spesen");
  if(spesenBtn) spesenBtn.className = "nav-btn" + (s==="spesen" ? " active" : "");
  var spesenPanel = document.getElementById("panel-spesen");
  if(spesenPanel) spesenPanel.style.display = (s==="spesen") ? "flex" : "none";
  var verstossPanel = document.getElementById("panel-verstoss");
  if(verstossPanel) verstossPanel.style.display = (s==="verstoss") ? "flex" : "none";
  var verstossGraphPanel = document.getElementById("panel-verstoss-graph");
  if(verstossGraphPanel) verstossGraphPanel.style.display = (s==="verstoss_graph") ? "flex" : "none";
  var verstossBtn = document.getElementById("btn-verstoss");
  if(verstossBtn) verstossBtn.className = "nav-dd-btn" + ((s==="verstoss" || s==="verstoss_graph") ? " active" : "");
  if(typeof buildVerstossDdMenu === "function") buildVerstossDdMenu();

  if(s==="vz") {{
    fwInitDatePicker();
    if(vzPanel && !vzPanel.dataset.loaded) {{ fwInitOverview(); vzPanel.dataset.loaded="1"; }}
    else {{ fwRenderOverview(); }}
  }}
  if(s==="vz_graph") {{
    if(vzGraphPanel && !vzGraphPanel.dataset.loaded) {{ fwInitGraph(); vzGraphPanel.dataset.loaded="1"; }}
    else {{ fwRenderGraph(); }}
  }}
  if(s==="tank") {{
    if(tankPanel && !tankPanel.dataset.loaded) {{ tankInitOverview(); tankPanel.dataset.loaded="1"; }}
    else {{ tankRenderOverview(); }}
  }}
  if(s==="tank_graph") {{
    if(tankGraphPanel && !tankGraphPanel.dataset.loaded) {{ tankInitGraph(); tankGraphPanel.dataset.loaded="1"; }}
    else {{ tankRenderGraph(); }}
  }}
  if(typeof buildVzDdMenu === "function") buildVzDdMenu();
  if(s==="tel" && !telPanel.dataset.loaded) {{ telRender(""); telPanel.dataset.loaded="1"; }}
  if((s==="sam" || s==="sam_graph") && samPanel) {{
    samViewMode = s === "sam_graph" ? "charts" : "list";
    var samTitle = document.getElementById("sam-panel-title");
    if(samTitle) samTitle.innerHTML = s === "sam_graph"
      ? "&#128202; Graph – Sa + So Einsätze"
      : "&#128664; Sa + So Einsätze";
    var samSortBtns = document.getElementById("sam-list-sort-buttons");
    var samChartTabs = document.getElementById("sam-chart-tabs");
    if(samSortBtns) samSortBtns.style.display = samViewMode === "list" ? "flex" : "none";
    if(samChartTabs) samChartTabs.style.display = samViewMode === "charts" ? "flex" : "none";
    var samQuery = (document.getElementById("sam-search") || {{value:""}}).value;
    samRender(samQuery);
    samPanel.dataset.loaded = "1";
  }}
  if(s==="zulage" && zulagePanel && !zulagePanel.dataset.loaded) {{ zulagenInit(); zulagePanel.dataset.loaded="1"; }}
  if(s==="spesen") {{
    if(spesenPanel && !spesenPanel.dataset.loaded) {{ spesenInit(); spesenPanel.dataset.loaded="1"; }}
    else {{ spesenUpdateSidebarHighlight(); }}
  }}
  if(s==="verstoss") {{
    if(verstossPanel && !verstossPanel.dataset.loaded) {{ verstossInit(); verstossPanel.dataset.loaded="1"; }}
    else {{ verstossRender(); }}
  }}
  if(s==="verstoss_graph") {{
    if(verstossGraphPanel && !verstossGraphPanel.dataset.loaded) {{ verstossInitGraph(); verstossGraphPanel.dataset.loaded="1"; }}
    else {{ verstossRenderGraph(); }}
  }}
  if(s==="fa") {{ if(faPanel) faPanel.scrollTop = 0; if(faPanel && !faPanel.dataset.loaded) {{ faRender(""); faPanel.dataset.loaded="1"; }} }}
  var faBewPanel = document.getElementById("panel-fa-bewertung");
  if(faBewPanel) faBewPanel.style.display = (s==="fa_bewertung") ? "flex" : "none";
  if(s==="fa_bewertung") {{
    if(faBewPanel && !faBewPanel.dataset.loaded) {{ faBewInit(); faBewPanel.dataset.loaded="1"; }}
    else {{ faBewRender(); }}
  }}
  var gkPanel = document.getElementById("panel-gk");
  if(gkPanel) gkPanel.style.display = (s==="gk") ? "flex" : "none";
  var gkBtn = document.getElementById("btn-gk");
  if(gkBtn) gkBtn.className = "nav-btn" + (s==="gk" ? " active" : "");
  if(s==="gk" && gkPanel && !gkPanel.dataset.loaded) {{ gkRender(); gkPanel.dataset.loaded="1"; }}
  var schluesselPanel = document.getElementById("panel-schluessel");
  if(schluesselPanel) schluesselPanel.style.display = (s==="schluessel") ? "flex" : "none";
  if(s==="schluessel" && schluesselPanel && !schluesselPanel.dataset.loaded) {{ documentPdfInit("schluessel"); schluesselPanel.dataset.loaded="1"; }}
  var entgeltPanel = document.getElementById("panel-entgelt");
  if(entgeltPanel) entgeltPanel.style.display = (s==="entgelt") ? "flex" : "none";
  if(s==="entgelt" && entgeltPanel && !entgeltPanel.dataset.loaded) {{ documentPdfInit("entgelt"); entgeltPanel.dataset.loaded="1"; }}
  var schadenPanel = document.getElementById("panel-schaden");
  if(schadenPanel) schadenPanel.style.display = (s==="schaden") ? "flex" : "none";
  if(s==="schaden" && schadenPanel && !schadenPanel.dataset.loaded) {{ documentPdfInit("schaden"); schadenPanel.dataset.loaded="1"; }}
  var maengelPanel = document.getElementById("panel-maengel");
  if(maengelPanel) maengelPanel.style.display = (s==="maengel") ? "flex" : "none";
  if(s==="maengel" && maengelPanel && !maengelPanel.dataset.loaded) {{ documentPdfInit("maengel"); maengelPanel.dataset.loaded="1"; }}
  var lkwUebergabePanel = document.getElementById("panel-lkw_uebergabe");
  if(lkwUebergabePanel) lkwUebergabePanel.style.display = (s==="lkw_uebergabe") ? "flex" : "none";
  if(s==="lkw_uebergabe" && lkwUebergabePanel && !lkwUebergabePanel.dataset.loaded) {{ documentPdfInit("lkw_uebergabe"); lkwUebergabePanel.dataset.loaded="1"; }}
  var reinigungPanel = document.getElementById("panel-reinigung");
  if(reinigungPanel) reinigungPanel.style.display = (s==="reinigung") ? "flex" : "none";
  if(s==="reinigung" && reinigungPanel && !reinigungPanel.dataset.loaded) {{ documentPdfInit("reinigung"); reinigungPanel.dataset.loaded="1"; }}
  var reisekostenPanel = document.getElementById("panel-reisekosten");
  if(reisekostenPanel) reisekostenPanel.style.display = (s==="reisekosten") ? "flex" : "none";
  var balzerPanel = document.getElementById("panel-balzer");
  if(balzerPanel) balzerPanel.style.display = (s==="balzer") ? "flex" : "none";
  if(s==="balzer" && balzerPanel && !balzerPanel.dataset.loaded) {{ documentPdfInit("balzer"); balzerPanel.dataset.loaded="1"; }}
  var terminePanel = document.getElementById("panel-termine");
  if(terminePanel) terminePanel.style.display = (s==="termine") ? "flex" : "none";
  if(s==="termine" && terminePanel && !terminePanel.dataset.loaded) {{ documentPdfInit("termine"); terminePanel.dataset.loaded="1"; }}
  var busPanel = document.getElementById("panel-bus");
  if(busPanel) busPanel.style.display = (s==="bus") ? "block" : "none";
  var busBtn = document.getElementById("btn-bus");
  if(busBtn) busBtn.className = "nav-btn" + (s==="bus" ? " active" : "");
  if(s==="bus" && busPanel && !busPanel.dataset.loaded) {{ busRender(); busPanel.dataset.loaded="1"; }}
  var arztPanel = document.getElementById("panel-arzt");
  if(arztPanel) arztPanel.style.display = (s==="arzt") ? "block" : "none";
  if(s==="arzt" && arztPanel && !arztPanel.dataset.loaded) {{ arztRender(); arztPanel.dataset.loaded="1"; }}
  var verspPanel = document.getElementById("panel-versp");
  if(verspPanel) verspPanel.style.display = (s==="versp") ? "block" : "none";
  if(s==="versp" && verspPanel && !verspPanel.dataset.loaded) {{ verspInit(); verspPanel.dataset.loaded="1"; }}
  var knappPanel = document.getElementById("panel-knapp");
  if(knappPanel) knappPanel.style.display = (s==="knapp") ? "flex" : "none";
  if(s==="knapp" && knappPanel && !knappPanel.dataset.loaded) {{ knappInit(); knappPanel.dataset.loaded="1"; }}
  var infosBtn = document.getElementById("btn-infos");
  if(infosBtn) infosBtn.className = "nav-dd-btn" + ((s==="tel" || s==="bus" || s==="arzt" || s==="versp" || s==="knapp" || s==="schluessel" || s==="entgelt" || s==="schaden" || s==="maengel" || s==="lkw_uebergabe" || s==="reinigung" || s==="reisekosten" || s==="balzer" || s==="termine") ? " active" : "");
  if(typeof buildInfosDdMenu === "function") buildInfosDdMenu();
  var spedPanel = document.getElementById("panel-sped");
  if(spedPanel) spedPanel.style.display = (s==="sped") ? "flex" : "none";
  var spedGraphPanel = document.getElementById("panel-sped-graph");
  if(spedGraphPanel) spedGraphPanel.style.display = (s==="sped_graph") ? "flex" : "none";
  var spedBtn = document.getElementById("btn-sped");
  if(spedBtn) spedBtn.className = "nav-dd-btn" + ((s==="sped" || s==="sped_graph") ? " active" : "");
  if(typeof buildSpedDdMenu === "function") buildSpedDdMenu();
  if(s==="sped" && spedPanel && !spedPanel.dataset.loaded) {{ spedInit(); spedPanel.dataset.loaded="1"; }}
  if(s==="sped_graph") {{
    if(spedGraphPanel && !spedGraphPanel.dataset.loaded) {{ spedInitGraph(); spedGraphPanel.dataset.loaded="1"; }}
    else {{ spedRenderGraph(); }}
  }}
  // ── Wochenauslastung ──
  var waHeatPanel = document.getElementById("panel-wa-heat");
  if(waHeatPanel) waHeatPanel.style.display = (s==="wa_heat") ? "flex" : "none";
  var waGruppePanel = document.getElementById("panel-wa-gruppe");
  if(waGruppePanel) waGruppePanel.style.display = (s==="wa_gruppe") ? "flex" : "none";
  var waKurvePanel = document.getElementById("panel-wa-kurve");
  if(waKurvePanel) waKurvePanel.style.display = (s==="wa_kurve") ? "flex" : "none";
  var waRhythmPanel = document.getElementById("panel-wa-rhythm");
  if(waRhythmPanel) waRhythmPanel.style.display = (s==="wa_rhythm") ? "flex" : "none";
  var waBtn = document.getElementById("btn-wa");
  if(waBtn) waBtn.className = "nav-dd-btn" + ((s==="wa_heat" || s==="wa_gruppe" || s==="wa_kurve" || s==="wa_rhythm") ? " active" : "");
  if(typeof buildWaDdMenu === "function") buildWaDdMenu();
  if(s==="wa_heat") {{
    if(waHeatPanel && !waHeatPanel.dataset.loaded) {{ waInitHeat(); waHeatPanel.dataset.loaded="1"; }}
    else {{ waRenderHeat(); }}
  }}
  if(s==="wa_gruppe") {{
    if(waGruppePanel && !waGruppePanel.dataset.loaded) {{ waInitGruppe(); waGruppePanel.dataset.loaded="1"; }}
    else {{ waRenderGruppe(); }}
  }}
  if(s==="wa_kurve") {{
    if(waKurvePanel && !waKurvePanel.dataset.loaded) {{ waInitKurve(); waKurvePanel.dataset.loaded="1"; }}
    else {{ waRenderKurve(); }}
  }}
  if(s==="wa_rhythm") {{
    if(waRhythmPanel && !waRhythmPanel.dataset.loaded) {{ waInitRhythm(); waRhythmPanel.dataset.loaded="1"; }}
    else {{ waRenderRhythm(); }}
  }}
}}

// showKundenListeTop removed — Kunden Liste is now a standalone panel

if(INSTANCES.length > 0) {{
  (async function() {{
    await loadEmbeddedData();
    document.getElementById("frame-suche").srcdoc = await b64ChunksToString(INSTANCES[0].s);
    document.getElementById("frame-druck" ).srcdoc = await b64ChunksToString(INSTANCES[0].d);
    updateInstLabels();
    fwInitDatePicker();
  }})();
}}

var normalInstData = null;  // ALL_DATA der Normalwochen (Instanz 0)
window.addEventListener("message", function(e) {{
  if (e.data === "show-suche") showArea("suche");
  if (e.data && e.data.type === "vz-init-data") {{
    vzAllData = e.data.data;
    // Erste Instanz = Normalwochen → als Referenz speichern
    if(currentInst === 0) normalInstData = e.data.data;
    // Pro Woche/Instanz cachen, damit die Verspätungstabelle ohne Neuladen wechseln kann
    try {{ verspInstCache[currentInst] = e.data.data; }} catch(err) {{}}
    if(typeof currentArea !== "undefined" && currentArea === "versp" && typeof verspUpdateInfo === "function") verspUpdateInfo();
  }}
  if (e.data && e.data.type === "request-normal-data") {{
    // Druck-iframe fragt nach Normalwochen-Daten
    var fd = document.getElementById("frame-druck");
    if(normalInstData && fd && fd.contentWindow) {{
      try {{ fd.contentWindow.postMessage({{type:"normal-data",data:normalInstData}},"*"); }} catch(e) {{}}
    }}
  }}
}});

// ── Fahrzeugwäsche ────────────────────────────────────────────────────────────
var fwSelectedDay = null;
var fwSelectedDate = null;
var vzAllData = null;
var FW_EXCLUDED_SUFFIXES = ["998","999","2221","2222","2223","4444","7773","7778","7779"];

function fwSelectDay(day) {{
  fwSelectedDay = day;
  document.querySelectorAll("#fw-day-btns .vz-day-btn").forEach(function(b) {{
    b.classList.toggle("active", b.textContent.trim()===day);
  }});
}}

function fwSetDate(value) {{
  fwSelectedDate = value || null;
}}

function fwIsExcludedNumber(value) {{
  var s = (value == null ? "" : String(value)).replace(/\\D/g, "");
  if(!s) return false;
  return FW_EXCLUDED_SUFFIXES.some(function(suffix) {{
    if(s.endsWith(suffix)) return true;
    if(s.length >= suffix.length + 1 && /^[1-6]$/.test(s.charAt(0))) {{
      return s.slice(1).endsWith(suffix);
    }}
    return false;
  }});
}}

function vzCollectToursByDay(allData, day) {{
  var AREAS = ["direkt","mk","nms","malchow"];
  var seen = {{}};
  AREAS.forEach(function(area) {{
    var areaData = allData[area] || {{}};
    Object.keys(areaData).forEach(function(knr) {{
      var c = areaData[knr];
      var refNum = c && (c.sap_nummer || c.kunden_nr || knr);
      if(fwIsExcludedNumber(refNum)) return;
      if(!c || !c.tours || !c.tours[day]) return;
      var t = c.tours[day].toString().trim();
      if(!t || t === "\u2014" || t === "-" || fwIsExcludedNumber(t)) return;
      seen[t] = 1;
    }});
  }});
  return Object.keys(seen).sort(function(a,b) {{
    return (parseInt(a,10)||0) - (parseInt(b,10)||0) || a.localeCompare(b, "de");
  }});
}}

// ── Telefonliste ──────────────────────────────────────────────────────────────
function fwTodayLabel() {{
  return new Date().toLocaleDateString("de-DE", {{ day:"2-digit", month:"2-digit", year:"numeric" }});
}}

function fwInitDatePicker() {{
  var el = document.getElementById("fw-date-picker");
  if(!el) return;
  if(!fwSelectedDate) {{
    var now = new Date();
    var month = String(now.getMonth() + 1).padStart(2, "0");
    var day = String(now.getDate()).padStart(2, "0");
    fwSelectedDate = now.getFullYear() + "-" + month + "-" + day;
  }}
  el.value = fwSelectedDate;
}}

function fwDisplayDate() {{
  if(fwSelectedDate) {{
    var parts = fwSelectedDate.split("-");
    if(parts.length === 3) return parts[2] + "." + parts[1] + "." + parts[0];
  }}
  return fwTodayLabel();
}}

function fwGetSelectedDay() {{
  if(!fwSelectedDay) {{
    alert("Bitte zuerst einen Fahrzeugwaesche-Tag auswaehlen.");
    return null;
  }}
  return fwSelectedDay;
}}

function fwFilterTours(tours) {{
  return (tours || []).filter(function(t) {{
    return !fwIsExcludedNumber(t);
  }});
}}

function fwExportPdf() {{
  if(!vzAllData) {{
    alert("Die Wochendaten sind noch nicht bereit. Bitte kurz warten und erneut versuchen.");
    try {{ document.getElementById("frame-druck").contentWindow.postMessage({{type:"request-vz-data"}}, "*"); }} catch(e) {{}}
    return;
  }}
  var day = fwGetSelectedDay();
  if(!day) return;
  if(!window.jspdf || !window.jspdf.jsPDF || typeof window.jspdf.jsPDF !== "function") {{
    alert("PDF-Bibliothek ist nicht geladen.");
    return;
  }}

  var jsPDF = window.jspdf.jsPDF;
  var doc = new jsPDF({{ orientation:"portrait", unit:"mm", format:"a4" }});
  var label = (INSTANCES[currentInst] && INSTANCES[currentInst].name) ? INSTANCES[currentInst].name : "Woche";
  var tours = fwFilterTours(vzCollectToursByDay(vzAllData, day));

  doc.setFillColor(27, 102, 179);
  doc.roundedRect(12, 10, 186, 18, 3, 3, "F");
  doc.setTextColor(255, 255, 255);
  doc.setFont("helvetica", "bold");
  doc.setFontSize(17);
  doc.text("Fahrzeugwaeschen", 16, 18);
  doc.setFontSize(10);
  doc.text(label + " - " + day, 16, 24);
  doc.text("Stand: " + fwDisplayDate(), 155, 24);

  var rows = tours.map(function(t) {{
    return ["", t, "", "", ""];
  }});
  if(!rows.length) rows = [["", "-", "", "", ""]];

  // ── Jeden-Tag-Touren (feste Fahrzeuge) ──
  rows.push([{{content:"Extra", colSpan:5, styles:{{fillColor:[232,240,251], textColor:[27,102,179], fontStyle:"bold", halign:"left"}}}}]);
  var jedenTag = [
    [10, "z.b.V."],
    [3,  "Popp"],
    [1,  "Fuengers"],
    [1,  "Pfeifer"],
    [1,  "Picnic"],
    [3,  "Umfuhr NMS"],
    [5,  "Umfuhr Malchow"],
    [1,  "Langbein"],
    [1,  "Balzer"]
  ];
  jedenTag.forEach(function(item) {{
    for(var i = 0; i < item[0]; i++) {{
      rows.push(["", item[1], "", "", ""]);
    }}
  }});

  doc.autoTable({{
    startY: 33,
    head: [["Fahrer Name", "Tour", "Reinigung ja/nein", "Grund warum nicht gewaschen wurde", "Uhrzeit / Feierabend"]],
    body: rows,
    theme: "grid",
    styles: {{
      font: "helvetica",
      fontSize: 10,
      cellPadding: 2.6,
      lineColor: [205, 213, 225],
      lineWidth: 0.2,
      minCellHeight: 10,
      valign: "middle",
      overflow: "linebreak"
    }},
    headStyles: {{
      fillColor: [232, 240, 251],
      textColor: [27, 102, 179],
      fontStyle: "bold",
      halign: "left",
      lineColor: [27, 102, 179],
      lineWidth: 0.25
    }},
    columnStyles: {{
      0: {{ cellWidth: 33 }},
      1: {{ cellWidth: 23, halign: "center" }},
      2: {{ cellWidth: 33 }},
      3: {{ cellWidth: 56 }},
      4: {{ cellWidth: 41 }}
    }},
    margin: {{ left: 12, right: 12, top: 10, bottom: 12 }},
    didDrawPage: function(data) {{
      var pageSize = doc.internal.pageSize;
      var pageHeight = pageSize.height || pageSize.getHeight();
      doc.setFontSize(8);
      doc.setTextColor(100, 116, 139);
      doc.text(day + "  |  " + label, 12, pageHeight - 6);
      doc.text("Seite " + doc.getCurrentPageInfo().pageNumber, pageSize.width - 24, pageHeight - 6);
    }}
  }});

  doc.save("Fahrzeugwaeschen_" + day + "_" + label.replace(/[\\\\/:*?\"<>|]+/g, "_") + ".pdf");
}}

{wash_js_code}
{wash_ranking_js_code}
{fw_graph_js_code}
{tank_js_code}
// Zusatzdaten wurden oben komprimiert eingebettet und durch loadEmbeddedData() geladen.
var ZULAGE_XLSX_SONDER      = "{zulage_xlsx_sonder}";
var ZULAGE_XLSX_FUENGERS    = "{zulage_xlsx_fuengers}";
var ZULAGE_XLSX_DRITTKUNDEN = "{zulage_xlsx_drittkunden}";
var faActiveTab              = "schichten";

{spesen_js_code}

{sped_js_code}

{fabew_js_code}

// ── Großkunden ────────────────────────────────────────────────────────────────
var gkSelected = 0;
var gkSearchQ = "";

function gkEsc(v) {{
  return String(v == null ? "" : v)
    .replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
}}

function gkAttr(v) {{
  return gkEsc(v).replace(/"/g,"&quot;").replace(/'/g,"&#39;");
}}

// ── Deterministische Akzentfarbe + Monogramm pro Kunde ──────────────────────
function gkHue(name) {{
  var s = String(name || ""), h = 0, i;
  for (i = 0; i < s.length; i++) {{ h = (h * 31 + s.charCodeAt(i)) >>> 0; }}
  return h % 360;
}}

function gkAccentStyle(name) {{
  var h = gkHue(name);
  return "--gk:hsl(" + h + ",54%,42%);--gks:hsl(" + h + ",62%,95%);"
       + "--gkb:hsl(" + h + ",46%,86%);--gkd:hsl(" + h + ",55%,30%);";
}}

function gkInitials(name) {{
  var raw = String(name || "");
  var words = raw.replace(/[^0-9A-Za-zÄÖÜäöüß ]/g, " ").split(/\\s+/).filter(Boolean)
    .filter(function(w){{ return !/^(gmbh|mbh|kg|co|ag|ohg|ek|und|der|die|das)$/i.test(w); }});
  if (!words.length) words = raw.split(/\\s+/).filter(Boolean);
  var a = words[0] ? words[0].charAt(0) : "?";
  var b = words.length > 1 ? words[1].charAt(0) : "";
  return ((a + b).toUpperCase()) || "?";
}}

function gkMonoHtml(name, cls) {{
  return "<span class='" + (cls || "gk-mono") + "'>" + gkEsc(gkInitials(name)) + "</span>";
}}

function gkExtractEmails(v) {{
  var s = String(v == null ? "" : v);
  return s.match(/[A-Z0-9._%+-]+@[A-Z0-9.-]+\\.[A-Z]{{2,}}/ig) || [];
}}

function gkUniqEmails(list) {{
  var out = [];
  (list || []).forEach(function(e) {{
    var mail = String(e || "").trim();
    if (mail && out.indexOf(mail) === -1) out.push(mail);
  }});
  return out;
}}

function gkMailto(emails, subject) {{
  var list = gkUniqEmails(emails);
  var href = "mailto:" + list.join(";");
  if (subject) href += "?subject=" + encodeURIComponent(subject);
  return href;
}}

function gkRastingNfcEmailsFromRows(customer, rows) {{
  if (!customer || !/rasting/i.test(customer.name || "")) return [];
  var mails = [];
  var section = "";
  (rows || []).forEach(function(cr) {{
    if (cr && cr.isLabel) {{
      section = String(cr.text || "").replace(/:$/, "").toLowerCase().trim();
      return;
    }}
    var belongsToNfc = section.indexOf("nfc") >= 0;
    if (!belongsToNfc && cr && cr.labelText && /(^|\b)nfc(\b|$)/i.test(cr.labelText)) belongsToNfc = true;
    if (!belongsToNfc) return;
    (cr.emails || []).forEach(function(em) {{
      gkExtractEmails(em).forEach(function(addr) {{ mails.push(addr); }});
    }});
  }});
  return gkUniqEmails(mails);
}}

function gkRastingNfcEmailsFromLines(customer) {{
  if (!customer || !/rasting/i.test(customer.name || "")) return [];
  var mails = [];
  var section = "";
  (customer.lines || []).forEach(function(line) {{
    var t = String(line || "").trim();
    if (!t) return;
    if (gkIsSection(t)) section = t.replace(/:$/, "").toLowerCase().trim();
    var belongsToNfc = section.indexOf("nfc") >= 0 || /(^|\b)nfc(\b|$)/i.test(t);
    if (belongsToNfc) {{
      gkExtractEmails(t).forEach(function(addr) {{ mails.push(addr); }});
    }}
  }});
  return gkUniqEmails(mails);
}}

function gkDistributorLinkHtml(emails, label, subject) {{
  emails = gkUniqEmails(emails);
  if (!emails.length) return "";
  return "<a class='gk-act gk-act-sm' href='" + gkAttr(gkMailto(emails, subject || label)) + "'>"
       + "&#9993; " + gkEsc(label) + " <b>" + emails.length + "</b></a>";
}}

var GK_EDEKA_LAGER_NMS_EMAILS = [
  "Daniel.Kriegel@edeka.de",
  "Eugen.Kifer@edeka.de",
  "Jerome.Gitzel@EDEKA.de",
  "Maike.Linde@edeka.de",
  "thomas.manzke@edeka.de",
  "marco.doerfert@edeka.de",
  "stephan.bruhn@edeka.de",
  "alexa.offertaler@edeka.de",
  "alona.tymoshevska@edeka.de"
];

function gkIsEdekaLager(customer) {{
  var n = String(customer && customer.name ? customer.name : "").toLowerCase();
  return n.indexOf("edeka") >= 0 && n.indexOf("lager") >= 0;
}}

function gkEdekaLagerNmsLink(customer) {{
  if (!gkIsEdekaLager(customer)) return "";
  return gkDistributorLinkHtml(GK_EDEKA_LAGER_NMS_EMAILS, "NMS", "Edeka Lager / NMS");
}}

// Spaltentyp aus Header-Name ermitteln
function gkColType(header) {{
  var h = (header || "").toLowerCase();
  if (/mail|e-mail|email/.test(h))              return "email";
  if (/tel|telefon|mobil|fax|phone/.test(h))    return "tel";
  if (/adress|strasse|street|plz|ort|city/.test(h)) return "addr";
  if (/hinweis|info|bemerkung|anmerkung|notiz|note/.test(h)) return "hint";
  if (/ansprechpartner|kontakt|contact/.test(h)) return "contact";
  return "text";
}}

// Ist ein Zellwert ein Abschnitts-Header? (endet mit ":" und kurz)
function gkIsSection(v) {{
  var t = (v || "").trim();
  return t.endsWith(":") && t.length < 60 && !t.includes("@");
}}

function gkCustomerText(k) {{
  var parts = [k && k.name ? k.name : ""];
  if (k && k.type === "structured") {{
    (k.entries || []).forEach(function(e) {{
      parts.push(e.name || "");
      parts.push(e.kundennummer || "");
    }});
  }} else if (k && k.lines) {{
    parts = parts.concat(k.lines.slice(0, 60));
  }}
  return parts.join(" ").toLowerCase();
}}

function gkVisibleIndexes() {{
  var q = (gkSearchQ || "").toLowerCase().trim();
  var out = [];
  (GK_DATA || []).forEach(function(k, i) {{
    if (!q || gkCustomerText(k).indexOf(q) >= 0) out.push(i);
  }});
  return out;
}}

function gkFilter(q) {{
  gkSearchQ = q || "";
  var visible = gkVisibleIndexes();
  if (visible.indexOf(gkSelected) < 0) gkSelected = visible.length ? visible[0] : -1;
  gkBuildTiles(gkSelected);
  if (gkSelected >= 0) gkShow(gkSelected, true);
  else {{
    var detail = document.getElementById("gk-detail");
    if (detail) detail.innerHTML = "<div class='gk-empty'>Kein Gro&#223;kunde f&#252;r diese Suche.</div>";
  }}
}}

function gkBuildTiles(activeIdx) {{
  var tilesEl = document.getElementById("gk-tiles");
  var countEl = document.getElementById("gk-count");
  var statsEl = document.getElementById("gk-stats");
  if (!tilesEl) return;
  var total = (GK_DATA || []).length;
  var visible = gkVisibleIndexes();
  if (countEl) countEl.textContent = "(" + total + ")";
  if (statsEl) statsEl.textContent = visible.length + " angezeigt";
  var html = "";
  visible.forEach(function(realIdx) {{
    var k = GK_DATA[realIdx];
    var active = realIdx === activeIdx;
    var singleKnr = "";
    var multiCount = 0;
    if (k.type === "structured" && k.entries && k.entries.length) {{
      if (k.entries.length === 1) singleKnr = k.entries[0].kundennummer || "";
      else multiCount = k.entries.length;
    }}
    var meta = singleKnr
      ? "<span class='gk-knr'><span class='gk-knr-l'>KNr</span><span class='gk-knr-v'>" + gkEsc(singleKnr) + "</span></span>"
      : (multiCount ? "<span class='gk-tile-meta'>" + multiCount + " Standorte</span>" : "");
    html += "<button type='button' onclick='gkShow(" + realIdx + ")'"
          + " class='gk-tile" + (active ? " is-active" : "") + "'"
          + " style='" + gkAccentStyle(k.name) + "'>"
          + gkMonoHtml(k.name, "gk-mono")
          + "<span class='gk-tile-txt'>"
          + "<span class='gk-tile-name'>" + gkEsc(k.name) + "</span>"
          + (meta ? meta : "")
          + "</span>"
          + "</button>";
  }});
  tilesEl.innerHTML = html || "<div class='gk-tile-meta' style='padding:8px 4px;'>Keine Treffer.</div>";
}}

function gkRender() {{
  var detail = document.getElementById("gk-detail");
  if (!detail) return;
  if (!GK_DATA || !GK_DATA.length) {{
    gkBuildTiles(-1);
    detail.innerHTML = "<div class='gk-empty'>Keine Gro\u00dfkundendaten \u2013 bitte Excel in Streamlit hochladen.</div>";
    return;
  }}
  var visible = gkVisibleIndexes();
  gkShow(visible.length ? visible[0] : 0);
}}

function gkShow(idx, skipTiles) {{
  gkSelected = idx;
  if (!skipTiles) gkBuildTiles(idx);
  var detail = document.getElementById("gk-detail");
  if (!detail) return;
  var customer = GK_DATA[idx];
  if (!customer) {{
    detail.innerHTML = "<div class='gk-empty'>Kein Gro&#223;kunde ausgew&#228;hlt.</div>";
    return;
  }}
  if (customer.type === "structured") gkRenderStructured(customer, detail);
  else gkRenderFreeform(customer, detail);
  detail.scrollTop = 0;
}}

// ── Structured Renderer ────────────────────────────────────────────────────────
function gkRenderStructured(customer, detail) {{
  var headers = customer.content_headers || [];
  var colTypes = headers.map(gkColType);

  function idxOf(type) {{
    return headers.map(function(h,i){{ return colTypes[i]===type?i:-1; }}).filter(function(i){{return i>=0;}});
  }}
  var addrIdx  = idxOf("addr");
  var emailIdx = idxOf("email");
  var telIdx   = idxOf("tel");
  var hintIdx  = idxOf("hint");
  var otherIdx = headers.map(function(h,i){{
    return ["addr","email","tel","hint"].indexOf(colTypes[i])===-1 ? i : -1;
  }}).filter(function(i){{return i>=0;}});

  function isLabel(v) {{ return v && !v.includes("@") && v.trim().length > 0; }}
  function isPhone(v) {{ return v && /[\\d]{{4,}}/.test(v.trim()); }}

  // Sheet-weite Hinweise über alle Entries
  var allHints = [];
  customer.entries.forEach(function(e) {{
    (e.rows||[]).forEach(function(row) {{
      hintIdx.forEach(function(ci) {{
        var v = (row[ci]||"").trim();
        if (v && allHints.indexOf(v)===-1) allHints.push(v);
      }});
    }});
  }});

  var entryCount = customer.entries.length;
  var subline = entryCount === 1 ? "1 Standort" : entryCount + " Standorte";

  // Alle E-Mail-Adressen des Kunden für "Mail an alle"
  var gkAllEmails = [];
  customer.entries.forEach(function(entry) {{
    (entry.rows||[]).forEach(function(row) {{
      emailIdx.forEach(function(ci) {{
        gkExtractEmails(row[ci]||"").forEach(function(a){{ gkAllEmails.push(a); }});
      }});
    }});
  }});
  gkAllEmails = gkUniqEmails(gkAllEmails);

  var html = "<div class='gk-fade' style='" + gkAccentStyle(customer.name) + "width:100%;'>";

  // ── Hero ─────────────────────────────────────────────────────────────────────
  html += "<div class='gk-hero'>"
        + gkMonoHtml(customer.name, "gk-hero-mono")
        + "<div class='gk-hero-txt'>"
        + "<div class='gk-hero-name'>" + gkEsc(customer.name) + "</div>"
        + "<div class='gk-hero-sub'>" + subline + "</div>"
        + "</div>";
  var heroActs = "";
  if (gkAllEmails.length) heroActs += "<a class='gk-act' href='" + gkAttr(gkMailto(gkAllEmails, customer.name)) + "'>&#9993; Mail an alle <b>" + gkAllEmails.length + "</b></a>";
  heroActs += gkEdekaLagerNmsLink(customer);
  if (heroActs) html += "<div class='gk-actions'>" + heroActs + "</div>";
  html += "</div>";

  // ════════════════════════════════════════════════════════════════════════════
  // BLOCK 1: Standorte
  // ════════════════════════════════════════════════════════════════════════════
  // Daten vorab sammeln, dann Layout entscheiden
  var standorte = customer.entries.map(function(entry) {{
    var addrLines = [];
    if (addrIdx.length) {{
      (entry.rows||[]).forEach(function(row) {{
        addrIdx.forEach(function(ci) {{
          var v = (row[ci]||"").trim();
          if (v && addrLines.indexOf(v)===-1) addrLines.push(v);
        }});
      }});
    }}
    return {{ name: entry.name, kundennummer: entry.kundennummer, addrLines: addrLines }};
  }});
  var anyHasAddr = standorte.some(function(s) {{ return s.addrLines.length > 0; }});

  // KNr-Badge-Helper (kundenfarben getönt)
  function gkKnrBadge(knr) {{
    if (!knr) return "";
    return "<span class='gk-knr'><span class='gk-knr-l'>KNr</span>"
         + "<span class='gk-knr-v'>" + gkEsc(knr) + "</span></span>";
  }}

  html += "<div class='gk-card'>";
  html += "<div class='gk-card-h'><span class='gk-eyebrow'>Standorte</span>"
        + "<span class='gk-cnt'>" + entryCount + "</span></div>";

  if (anyHasAddr) {{
    html += "<div class='gk-card-b'>";
    standorte.forEach(function(s) {{
      html += "<div class='gk-loc'>";
      html += "<div class='gk-loc-name'>" + gkEsc(s.name) + gkKnrBadge(s.kundennummer) + "</div>";
      if (s.addrLines.length) {{
        html += "<div class='gk-loc-addr'>" + s.addrLines.map(gkEsc).join("<br>") + "</div>";
      }}
      html += "</div>";
    }});
    html += "</div>";
  }} else {{
    html += "<div class='gk-loc-chips' style='padding-left:16px;padding-right:16px;'>";
    standorte.forEach(function(s) {{
      html += "<div class='gk-loc-chip'>" + gkEsc(s.name) + gkKnrBadge(s.kundennummer) + "</div>";
    }});
    html += "</div>";
  }}
  html += "</div>";

  // ════════════════════════════════════════════════════════════════════════════
  // BLOCK 2+3: Kontakte links, Hinweise rechts — zweispaltiges Layout
  // ════════════════════════════════════════════════════════════════════════════
  var hasAnyContact = emailIdx.length || telIdx.length || otherIdx.length;

  // Alle Kontaktzeilen über alle Entries sammeln
  var allContactRows = [];
  customer.entries.forEach(function(entry) {{
    (entry.rows||[]).forEach(function(row) {{
      var labelVal = null;
      emailIdx.forEach(function(ci) {{
        var v = (row[ci]||"").trim();
        if (v && isLabel(v)) labelVal = v;
      }});
      if (!labelVal) {{
        otherIdx.forEach(function(ci) {{
          var v = (row[ci]||"").trim();
          if (v && isLabel(v) && !isPhone(v)) labelVal = labelVal || v;
        }});
      }}
      // Tel dieser Zeile immer zuerst bestimmen
      var rowTels = [];
      telIdx.forEach(function(ci) {{
        var v = (row[ci]||"").trim();
        if (v && isPhone(v)) rowTels.push(v);
      }});

      if (labelVal) {{
        if (rowTels.length) {{
          // Text + Tel auf gleicher Zeile → Kontaktzeile mit Beschriftung
          allContactRows.push({{isLabel:false, labelText:labelVal, emails:[], tels:rowTels}});
        }} else {{
          // Reiner Text ohne Tel → Section-Chip
          allContactRows.push({{isLabel:true, text:labelVal}});
        }}
        return;
      }}
      var rowEmails = [];
      emailIdx.forEach(function(ci) {{
        var v = (row[ci]||"").trim();
        if (v && !isLabel(v)) rowEmails.push(v);
      }});
      if (rowEmails.length || rowTels.length) {{
        // Tel-only-Zeile ohne Email → zur letzten Kontaktzeile ohne Tel mergen
        if (!rowEmails.length && rowTels.length) {{
          var last = allContactRows.length ? allContactRows[allContactRows.length-1] : null;
          if (last && !last.isLabel && !last.isOther && (!last.tels || !last.tels.length)) {{
            last.tels = rowTels;
            return;
          }}
        }}
        allContactRows.push({{isLabel:false, emails:rowEmails, tels:rowTels}});
      }}
    }});
    // Sonstige Spalten dieser Entry
    otherIdx.forEach(function(ci) {{
      var vals = [];
      (entry.rows||[]).forEach(function(row) {{
        var v = (row[ci]||"").trim();
        if (v && !isLabel(v) && vals.indexOf(v)===-1) vals.push(v);
      }});
      if (vals.length) {{
        allContactRows.push({{isOther:true, header:headers[ci], vals:vals}});
      }}
    }});
  }});

  var rastingNfcEmails = gkRastingNfcEmailsFromRows(customer, allContactRows);

  // Zwei-Spalten-Layout: Kontakte links, Hinweise rechts
  html += "<div class='gk-grid'>";

  // ── LINKE SPALTE: Kontakte ─────────────────────────────────────────────────
  html += "<div style='min-width:0;'>";
  if (hasAnyContact && allContactRows.length) {{
    html += "<div class='gk-card'>";
    html += "<div class='gk-card-h'><span class='gk-eyebrow'>Kontakte</span></div>";
    html += "<div class='gk-card-b'>";
    allContactRows.forEach(function(cr, ri) {{
      if (cr.isLabel) {{
        var sectionName = cr.text.replace(/:$/, "");
        var sKey = sectionName.toLowerCase().trim();
        html += "<div class='gk-c-sec'><span class='gk-c-sec-l'>" + gkEsc(sectionName) + "</span>";
        if (sKey === "nfc" && rastingNfcEmails.length) {{
          html += gkDistributorLinkHtml(rastingNfcEmails, "Mail an den Verteiler", "Rasting / NFC");
        }}
        html += "</div>";
      }} else if (cr.isOther) {{
        html += "<div class='gk-c-other'>"
              + "<span class='gk-c-other-l'>" + gkEsc(cr.header) + "</span>"
              + "<span class='gk-c-other-v'>" + cr.vals.map(gkEsc).join("<br>") + "</span>"
              + "</div>";
      }} else {{
        var hasEmail     = cr.emails && cr.emails.length;
        var hasLabelText = cr.labelText && cr.labelText.length;
        var hasTel       = cr.tels && cr.tels.length;
        html += "<div class='gk-c-row'><span class='gk-c-left'>";
        if (hasEmail) {{
          html += cr.emails.map(function(em) {{
            var isMailAddr = em.match(/^[^\\s@]+@[^\\s@]+/) !== null;
            return isMailAddr
              ? "<a class='gk-mail' href='mailto:" + gkEsc(em) + "'>" + gkEsc(em) + "</a>"
              : "<span class='gk-c-label'>" + gkEsc(em) + "</span>";
          }}).join(" ");
        }} else if (hasLabelText) {{
          html += "<span class='gk-c-label'>" + gkEsc(cr.labelText) + "</span>";
        }} else {{
          html += "<span class='gk-c-dash'>&ndash;</span>";
        }}
        html += "</span>";
        if (hasTel) {{
          html += cr.tels.map(function(t) {{
              var href = "tel:" + t.replace(/[^\\d\\+]/g,"");
              return "<a class='gk-tel' href='" + href + "'>"
                   + "<span class='gk-tel-i'>&#9742;</span>" + gkEsc(t) + "</a>";
            }}).join("");
        }}
        html += "</div>";
      }}
    }});
    html += "</div></div>"; // body + card
  }}
  html += "</div>"; // linke Spalte

  // ── RECHTE SPALTE: Hinweise ────────────────────────────────────────────────
  html += "<div style='min-width:0;'>";
  if (allHints.length) {{
    html += "<div class='gk-card'>";
    html += "<div class='gk-card-h'><span class='gk-eyebrow'>Hinweise</span>"
          + "<span class='gk-cnt'>" + allHints.length + "</span></div>";
    html += "<div>";
    allHints.forEach(function(h, i) {{
      html += "<div class='gk-hint'><span class='gk-hint-n'>" + (i+1) + "</span>"
            + "<span class='gk-hint-t'>" + gkEsc(h) + "</span></div>";
    }});
    html += "</div></div>";
  }}
  html += "</div>"; // rechte Spalte

  html += "</div>"; // grid

  html += "</div>"; // root
  detail.innerHTML = html;
}}

// ── Freeform Renderer ─────────────────────────────────────────────────────────
function gkRenderFreeform(customer, detail) {{
  var emailRe = /^[^\\s@]+@[^\\s@]+\\.[^\\s@]+$/;
  var rastingNfcEmails = gkRastingNfcEmailsFromLines(customer);
  var allMails = gkUniqEmails((customer.lines || []).reduce(function(acc, l){{ return acc.concat(gkExtractEmails(l)); }}, []));
  var html = "<div class='gk-fade' style='" + gkAccentStyle(customer.name) + "max-width:780px;'>";

  html += "<div class='gk-hero'>"
        + gkMonoHtml(customer.name, "gk-hero-mono")
        + "<div class='gk-hero-txt'><div class='gk-hero-name'>" + gkEsc(customer.name) + "</div></div>";
  var ffActs = "";
  if (allMails.length) ffActs += "<a class='gk-act' href='" + gkAttr(gkMailto(allMails, customer.name)) + "'>&#9993; Mail an alle <b>" + allMails.length + "</b></a>";
  ffActs += gkEdekaLagerNmsLink(customer);
  if (ffActs) html += "<div class='gk-actions'>" + ffActs + "</div>";
  html += "</div>";

  html += "<div class='gk-card'>";

  (customer.lines || []).forEach(function(line) {{
    var t = line.trim();
    if (!t) return;
    var isEmail = emailRe.test(t);
    var isSection = gkIsSection(t);
    var hasTelNum = !isEmail && /\\d{{5,}}/.test(t) && t.length < 60;
    var isLong = t.length > 90;

    if (isSection) {{
      var sectionName = t.replace(/:$/, "");
      var sectionKey = sectionName.toLowerCase().trim();
      html += "<div class='gk-c-sec' style='margin:0;padding:11px 16px;border-bottom:1px solid #eef2f7;background:linear-gradient(180deg,#fbfdff,#fff);'>"
            + "<span class='gk-c-sec-l'>" + gkEsc(sectionName) + "</span>";
      if (sectionKey === "nfc" && rastingNfcEmails.length) {{
        html += gkDistributorLinkHtml(rastingNfcEmails, "Mail an den Verteiler", "Rasting / NFC");
      }}
      html += "</div>";
    }} else if (isEmail) {{
      html += "<div class='gk-ff-row'><a class='gk-mail' href='mailto:" + gkEsc(t) + "'>" + gkEsc(t) + "</a></div>";
    }} else if (hasTelNum) {{
      html += "<div class='gk-ff-row'><a class='gk-tel' href='tel:" + gkEsc(t.replace(/[^\\d\\+]/g,"")) + "'>"
            + "<span class='gk-tel-i'>&#9742;</span>" + gkEsc(t) + "</a></div>";
    }} else if (isLong) {{
      html += "<div class='gk-ff-row gk-ff-long'>" + gkEsc(t) + "</div>";
    }} else {{
      html += "<div class='gk-ff-row'>" + gkEsc(t) + "</div>";
    }}
  }});

  html += "</div></div>";
  detail.innerHTML = html;
}}


function telPDF() {{
  var w = window.open("","_blank","width=900,height=700");
  var css = [
    "body{{font-family:'Segoe UI',Arial,sans-serif;font-size:8pt;margin:10mm;color:#000}}",
    "h1{{font-size:13pt;color:#1b66b3;margin:0 0 4mm 0;border-bottom:2px solid #1b66b3;padding-bottom:2mm}}",
    ".gruppe{{font-size:8pt;font-weight:900;text-transform:uppercase;color:#1b66b3;margin:4mm 0 1.5mm 0;letter-spacing:.3px;border-bottom:1px solid #1b66b3;padding-bottom:0.5mm}}",
    ".grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:1mm 3mm}}",
    ".item{{padding:0.8mm 0;border-bottom:1px solid #eee;line-height:1.3}}",
    ".iname{{font-weight:800;font-size:7.5pt}}",
    ".itel{{color:#1b66b3;font-size:7.5pt}}",
    ".imail{{color:#888;font-size:6.5pt;font-style:italic}}",
    ".irole{{font-size:6.5pt;color:#dc2626;font-weight:700}}",
    "@media print{{@page{{size:A4 portrait;margin:10mm}}body{{margin:0}}}}"
  ].join("");
  var body = "<h1>&#128222; Telefonliste</h1>";
  TEL_DATA.forEach(function(g) {{
    if(!g.personen.length) return;
    body += "<div class='gruppe'>" + g.gruppe + " (" + g.personen.length + ")</div>";
    body += "<div class='grid'>";
    g.personen.forEach(function(p) {{
      body += "<div class='item'>";
      body += "<div class='iname'>" + p.name + "</div>";
      body += "<div class='itel'>&#128222; " + p.tel + "</div>";
      if(p.mail) {{
        var isRole = (p.mail==="Disponent"||p.mail==="Chef");
        body += isRole
          ? "<div class='irole'>" + p.mail + "</div>"
          : "<div class='imail'>&#9993; " + p.mail + "</div>";
      }}
      body += "</div>";
    }});
    body += "</div>";
  }});
  w.document.write("<!DOCTYPE html><html><head><meta charset='utf-8'>");
  w.document.write("<title>Telefonliste NordFrischeCenter</title>");
  w.document.write("<style>" + css + "</style></head><body>");
  w.document.write(body);
  w.document.write("</body></html>");
  w.document.close();
  w.focus();
  setTimeout(function(){{ w.print(); }}, 400);
}}

function telCopy(tel) {{
  if(navigator.clipboard) navigator.clipboard.writeText(tel);
}}
function telCopyBtn(btn) {{
  var t = btn.getAttribute("data-tel") || "";
  if(navigator.clipboard) navigator.clipboard.writeText(t);
  btn.classList.add("copied");
  btn.innerHTML = "&#10003;";
  setTimeout(function(){{ btn.classList.remove("copied"); btn.innerHTML = "&#128203;"; }}, 1100);
}}
function telAvatarHtml(name) {{
  var h = gkHue(name);
  return "<span class='tel-av' style='background:hsl(" + h + ",58%,93%);color:hsl(" + h + ",48%,32%);'>"
       + gkEsc(gkInitials(name)) + "</span>";
}}
function telHover(el, on) {{
  el.style.background = on ? "#eff6ff" : "#fff";
}}
function telFilter(q) {{ telRender(q); }}

function telRender(q) {{
  q = (q || "").toLowerCase().trim();
  var html = "";
  var totalShown = 0;
  TEL_DATA.forEach(function(g) {{
    var hits = g.personen.filter(function(p) {{
      return !q || p.name.toLowerCase().includes(q) || p.tel.includes(q);
    }});
    if(!hits.length) return;
    totalShown += hits.length;
    html += "<div class='tel-group'>";
    html += "<div class='tel-group-h'><span class='tel-group-n'>" + gkEsc(g.gruppe) + "</span>"
          + "<span class='tel-group-c'>" + hits.length + "</span></div>";
    html += "<div class='tel-grid'>";
    hits.forEach(function(p) {{
      var safetel = String(p.tel).replace(/'/g,"&#39;");
      var telClean = String(p.tel).replace(/[^\\d+]/g,"");
      var isRole = (p.mail === "Disponent" || p.mail === "Chef");
      html += "<div class='tel-card'>";
      html += telAvatarHtml(p.name);
      html += "<div class='tel-body'>";
      html += "<div class='tel-name' title='" + gkAttr(p.name) + "'>" + gkEsc(p.name) + "</div>";
      html += "<a class='tel-num' href='tel:" + telClean + "' title='Anrufen: " + gkAttr(p.tel) + "'>"
            + "<span class='tel-num-i'>&#128222;</span>" + gkEsc(p.tel) + "</a>";
      if(p.mail) {{
        html += isRole
          ? "<span class='tel-role'>" + gkEsc(p.mail) + "</span>"
          : "<a class='tel-mail' href='mailto:" + gkAttr(p.mail) + "' title='" + gkAttr(p.mail) + "'>&#9993; " + gkEsc(p.mail) + "</a>";
      }}
      html += "</div>";
      html += "<button class='tel-copy' type='button' title='Nummer kopieren' data-tel='" + safetel + "' onclick='telCopyBtn(this)'>&#128203;</button>";
      html += "</div>";
    }});
    html += "</div></div>";
  }});
  if(!html) html = "<div class='tel-empty'>Keine Treffer.</div>";
  document.getElementById("tel-content").innerHTML = html;
  var cnt = document.getElementById("tel-count");
  if(cnt) cnt.textContent = totalShown ? (totalShown + (totalShown === 1 ? " Kontakt" : " Kontakte")) : "";
}}
// ── Samstags Fahrer ───────────────────────────────────────────────────────────
var samCurrentSort  = "count";
var samYearFilter   = String(new Date().getFullYear());
var samStatusFilter = "all";
var samViewMode     = "list";
var samChartMode    = "matrix";
var samLastFiltered = [];
var SAM_ZIEL        = 12;

function samSort(mode) {{
  samCurrentSort = mode;
  ["name","count","status"].forEach(function(m) {{
    var btn = document.getElementById("sam-sort-"+m);
    if(!btn) return;
    btn.style.background = mode===m ? "#1b66b3" : "#fff";
    btn.style.color      = mode===m ? "#fff"    : "#1b66b3";
  }});
  samRender((document.getElementById("sam-search")||{{value:""}}).value);
}}

function samFilter(q) {{ samRender(q); }}

function samYearChange(yr) {{
  samYearFilter = String(yr || new Date().getFullYear());
  samStatusFilter = "all";
  samRender((document.getElementById("sam-search")||{{value:""}}).value);
}}

function samSetStatusFilter(status) {{
  samStatusFilter = status || "all";
  samRender((document.getElementById("sam-search")||{{value:""}}).value);
}}

function samSetView(mode) {{
  samViewMode = mode === "charts" ? "charts" : "list";
  var listBtn = document.getElementById("sam-view-list");
  var chartBtn = document.getElementById("sam-view-charts");
  if(listBtn) {{
    listBtn.style.background = samViewMode === "list" ? "#1b66b3" : "#fff";
    listBtn.style.color = samViewMode === "list" ? "#fff" : "#1b66b3";
  }}
  if(chartBtn) {{
    chartBtn.style.background = samViewMode === "charts" ? "#1b66b3" : "#fff";
    chartBtn.style.color = samViewMode === "charts" ? "#fff" : "#1b66b3";
  }}
  var sortBtns = document.getElementById("sam-list-sort-buttons");
  var chartTabs = document.getElementById("sam-chart-tabs");
  if(sortBtns) sortBtns.style.display = samViewMode === "list" ? "flex" : "none";
  if(chartTabs) chartTabs.style.display = samViewMode === "charts" ? "flex" : "none";
  samRender((document.getElementById("sam-search")||{{value:""}}).value);
}}

function samSetChartMode(mode) {{
  samChartMode = ["matrix","drivers","months"].indexOf(mode) >= 0 ? mode : "matrix";
  ["matrix","drivers","months"].forEach(function(m) {{
    var btn = document.getElementById("sam-chart-"+m);
    if(!btn) return;
    btn.style.background = samChartMode === m ? "#1b66b3" : "#fff";
    btn.style.color = samChartMode === m ? "#fff" : "#1b66b3";
  }});
  samRender((document.getElementById("sam-search")||{{value:""}}).value);
}}

function samStatBtn(status, html, bg, color, title) {{
  var active = samStatusFilter === status;
  return "<button type='button' onclick='samSetStatusFilter(&quot;"+status+"&quot;)' title='"+samEsc(title||"")+"' " +
    "style='display:inline-flex;align-items:center;gap:3px;border-radius:4px;padding:2px 8px;" +
    "font-family:inherit;font-size:11px;font-weight:800;line-height:1.4;cursor:pointer;" +
    "background:"+bg+";color:"+color+";border:1px solid "+(active?color:"transparent")+";" +
    "box-shadow:"+(active?"0 0 0 2px rgba(15,23,42,.10) inset":"none")+";'>" + html + "</button>";
}}

function samSaturdaysElapsed(year) {{
  var today = new Date();
  var d = new Date(year, 0, 1);
  while(d.getDay() !== 6) d.setDate(d.getDate()+1);
  var count = 0;
  while(d.getFullYear() === year && d <= today) {{
    count++;
    d.setDate(d.getDate()+7);
  }}
  return count;
}}

function samTotalSaturdays(year) {{
  var count = 0;
  var d = new Date(year, 0, 1);
  while(d.getDay() !== 6) d.setDate(d.getDate()+1);
  while(d.getFullYear() === year) {{ count++; d.setDate(d.getDate()+7); }}
  return count;
}}

function samParseYear(datum) {{
  var m = String(datum||"").match(/(\\d{{2}}\\.\\d{{2}}\\.(\\d{{4}}))/);
  return m ? parseInt(m[2],10) : null;
}}

function samEsc(v) {{
  return String(v == null ? "" : v)
    .replace(/&/g, "&amp;")
    .replace(/</g, "&lt;")
    .replace(/>/g, "&gt;")
    .replace(/"/g, "&quot;");
}}

function samAttr(v) {{
  return samEsc(v).replace(/'/g, "&#39;");
}}

function samNameKey(v) {{
  var s = String(v||"").toLowerCase().normalize("NFD").replace(/[\u0300-\u036f]/g,"");
  s = s.replace(/ß/g,"ss").replace(/[^a-z0-9]+/g," ").trim();
  return s.split(/\\s+/).filter(Boolean).sort().join("|");
}}

function samMergeText(a, b) {{
  var parts = [];
  String(a||"").split(/\\s*,\\s*/).concat(String(b||"").split(/\\s*,\\s*/)).forEach(function(x) {{
    x = x.trim();
    if(x && parts.indexOf(x) === -1) parts.push(x);
  }});
  return parts.join(", ");
}}

function samPrepareDrivers() {{
  var merged = {{}};
  (SAM_DATA||[]).forEach(function(d) {{
    var key = String(d.person_key||"").trim();
    if(!key) key = "name:" + samNameKey(d.name||"");
    if(!key || key === "name:") return;
    if(!merged[key]) {{
      merged[key] = {{
        person_key:key,
        name:String(d.name||"").trim(),
        nachname:d.nachname||"",
        vorname:d.vorname||"",
        daten:[],
        aktive_jahre:[]
      }};
    }}
    var out = merged[key];
    if(String(d.name||"").trim().length > String(out.name||"").trim().length) out.name = String(d.name||"").trim();
    (d.aktive_jahre||[]).forEach(function(y) {{
      y = String(y);
      if(out.aktive_jahre.indexOf(y) === -1) out.aktive_jahre.push(y);
    }});
    var eventMap = out._eventMap || (out._eventMap = {{}});
    (d.daten||[]).forEach(function(e) {{
      var eventKey = String(e.iso||"") + "|" + String(e.tag||"") + "|" + String(e.datum||"");
      if(!eventMap[eventKey]) {{
        eventMap[eventKey] = Object.assign({{}}, e);
        out.daten.push(eventMap[eventKey]);
      }} else {{
        eventMap[eventKey].tour = samMergeText(eventMap[eventKey].tour, e.tour);
        eventMap[eventKey].beginn = samMergeText(eventMap[eventKey].beginn, e.beginn);
      }}
    }});
  }});
  return Object.keys(merged).map(function(k) {{
    var d = merged[k];
    delete d._eventMap;
    d.aktive_jahre.sort().reverse();
    d.daten.sort(function(a,b) {{ return String(a.iso||a.datum||"").localeCompare(String(b.iso||b.datum||"")); }});
    return d;
  }});
}}

function samDateFromEntry(e) {{
  var iso = String((e||{{}}).iso||"");
  var m = iso.match(/^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})$/);
  if(m) return new Date(parseInt(m[1],10),parseInt(m[2],10)-1,parseInt(m[3],10));
  m = String((e||{{}}).datum||"").match(/(\\d{{2}})\\.(\\d{{2}})\\.(\\d{{4}})/);
  return m ? new Date(parseInt(m[3],10),parseInt(m[2],10)-1,parseInt(m[1],10)) : null;
}}

function samISODate(d) {{
  if(!d) return "";
  return d.getFullYear()+"-"+String(d.getMonth()+1).padStart(2,"0")+"-"+String(d.getDate()).padStart(2,"0");
}}

function samShortDate(d) {{
  return String(d.getDate()).padStart(2,"0")+"."+String(d.getMonth()+1).padStart(2,"0")+".";
}}

function samISOWeek(d) {{
  var x = new Date(Date.UTC(d.getFullYear(), d.getMonth(), d.getDate()));
  var day = x.getUTCDay() || 7;
  x.setUTCDate(x.getUTCDate() + 4 - day);
  var y0 = new Date(Date.UTC(x.getUTCFullYear(),0,1));
  return Math.ceil((((x-y0)/86400000)+1)/7);
}}

function samWeekendSaturday(e) {{
  var d = samDateFromEntry(e);
  if(!d) return null;
  var tag = String(e.tag||"");
  if(tag === "Fr→Sa" || tag === "Fr->Sa") d.setDate(d.getDate()+1);
  else if(tag === "So") d.setDate(d.getDate()-1);
  return d;
}}

function samRender(q) {{
  q = (q||"").toLowerCase().trim();
  var content = document.getElementById("sam-content");
  var statsEl = document.getElementById("sam-stats");
  if(!content) return;

  var prepared = samPrepareDrivers();
  if(!prepared.length) {{
    samLastFiltered = [];
    if(statsEl) statsEl.innerHTML = "";
    content.innerHTML =
      "<div style='color:#64748b;background:#fff;border:1px solid #cbd5e1;border-radius:6px;padding:34px;text-align:center;font-size:14px;line-height:1.55;'>" +
      "Keine Tachograph-Schichten vorhanden.<br><b>Bitte unter Zusatzdateien die Datei timerecording_v3*.csv hochladen.</b></div>";
    return;
  }}

  var today   = new Date();
  var curYear = today.getFullYear();

  var driverData = prepared.map(function(d) {{
    var byYear = {{}};
    (d.daten||[]).forEach(function(e) {{
      var yr = samParseYear(e.datum||"") || (samDateFromEntry(e)||{{getFullYear:function(){{return null;}}}}).getFullYear();
      if(!yr) return;
      var key = String(yr);
      if(!byYear[key]) byYear[key] = [];
      byYear[key].push(e);
    }});
    var activeYears = (d.aktive_jahre||[]).map(function(y){{ return String(y); }});
    return Object.assign({{}}, d, {{_byYear:byYear, _activeYears:activeYears}});
  }});

  var allYears = [String(curYear)];
  driverData.forEach(function(d) {{
    Object.keys(d._byYear).concat(d._activeYears||[]).forEach(function(yr) {{
      if(allYears.indexOf(String(yr)) === -1) allYears.push(String(yr));
    }});
  }});
  allYears.sort(function(a,b){{ return parseInt(b,10)-parseInt(a,10); }});

  if(allYears.indexOf(String(samYearFilter)) === -1) {{
    samYearFilter = allYears.indexOf(String(curYear)) !== -1 ? String(curYear) : allYears[0];
  }}
  var yrSel = document.getElementById("sam-year-sel");
  if(yrSel) {{
    yrSel.innerHTML = allYears.map(function(y){{
      return "<option value='"+samEsc(y)+"'>"+samEsc(y)+"</option>";
    }}).join("");
    yrSel.value = String(samYearFilter);
  }}

  var selectedYear = parseInt(samYearFilter,10) || curYear;
  var satTotal = samTotalSaturdays(selectedYear);
  var satElapsed;
  if(selectedYear < curYear) satElapsed = satTotal;
  else if(selectedYear > curYear) satElapsed = 0;
  else satElapsed = samSaturdaysElapsed(selectedYear);

  var soll = selectedYear < curYear
    ? SAM_ZIEL
    : (selectedYear > curYear ? 0 : Math.round(SAM_ZIEL * satElapsed / Math.max(1,satTotal)));
  var sollText = selectedYear === curYear ? "Soll heute" : "Jahressoll";

  driverData.forEach(function(d) {{
    d._selectedCount = (d._byYear[String(selectedYear)]||[]).length;
    var diff = d._selectedCount - soll;
    if(d._selectedCount >= SAM_ZIEL) d._status = "done";
    else if(diff >= 0)               d._status = "ok";
    else if(diff >= -1)              d._status = "warn";
    else                             d._status = "crit";
    d._diff = diff;
  }});

  var baseFiltered = driverData.filter(function(d) {{
    if(d._activeYears.length && d._activeYears.indexOf(String(selectedYear)) === -1) return false;
    return !q || String(d.name||"").toLowerCase().includes(q);
  }});
  var filtered = baseFiltered.filter(function(d) {{
    return samStatusFilter === "all" || d._status === samStatusFilter;
  }});

  var statusOrder = {{crit:0, warn:1, ok:2, done:3}};
  if(samCurrentSort === "status") {{
    filtered.sort(function(a,b) {{
      var sd = statusOrder[a._status] - statusOrder[b._status];
      return sd !== 0 ? sd : String(a.name||"").localeCompare(String(b.name||""),"de");
    }});
  }} else if(samCurrentSort === "count") {{
    filtered.sort(function(a,b) {{
      return b._selectedCount !== a._selectedCount
        ? b._selectedCount - a._selectedCount
        : String(a.name||"").localeCompare(String(b.name||""),"de");
    }});
  }} else {{
    filtered.sort(function(a,b) {{ return String(a.name||"").localeCompare(String(b.name||""),"de"); }});
  }}

  // Für den Excel-Export genau die aktuell angezeigte Auswahl merken.
  samLastFiltered = filtered.slice();

  var nDone = baseFiltered.filter(function(d){{return d._status==="done";}}).length;
  var nOk   = baseFiltered.filter(function(d){{return d._status==="ok";}}).length;
  var nWarn = baseFiltered.filter(function(d){{return d._status==="warn";}}).length;
  var nCrit = baseFiltered.filter(function(d){{return d._status==="crit";}}).length;

  if(statsEl) statsEl.innerHTML =
    "<div style='display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px;font-size:11px;'>" +
    samStatBtn("all", baseFiltered.length+" Fahrer", "#eef2f7", "#64748b", "Alle Fahrer anzeigen") +
    samStatBtn("done", "&#10003; Ziel erreicht: "+nDone, "#dcfce7", "#16a34a", "Nur Fahrer mit erreichtem Jahresziel anzeigen") +
    samStatBtn("ok", "&#8679; Im Soll: "+nOk, "#dbeafe", "#1b66b3", "Nur Fahrer im Soll anzeigen") +
    samStatBtn("warn", "&#9888; Leicht hinter: "+nWarn, "#fef3c7", "#d97706", "Nur Fahrer leicht hinter Soll anzeigen") +
    samStatBtn("crit", "&#8595; Rückstand: "+nCrit, "#fee2e2", "#dc2626", "Nur Fahrer mit Rückstand anzeigen") +
    "<span style='margin-left:auto;color:#64748b;font-size:10px;'>"+sollText+": <b style='color:#1b66b3;'>"+soll+"</b> / "+SAM_ZIEL+
      " &nbsp;("+satElapsed+" von "+satTotal+" Samstagen)</span></div>";

  if(samViewMode === "charts") {{
    samRenderCharts(filtered, selectedYear, soll, satTotal, satElapsed);
    return;
  }}

  var statusCfg = {{
    done: {{border:"#16a34a",bg:"#f0fdf4",badge:"#16a34a",icon:"✓",label:"Jahresziel erreicht"}},
    ok:   {{border:"#1b66b3",bg:"#eff6ff",badge:"#1b66b3",icon:"↑",label:"Im Soll"}},
    warn: {{border:"#d97706",bg:"#fffbeb",badge:"#d97706",icon:"⚠",label:"Leicht hinter Soll"}},
    crit: {{border:"#dc2626",bg:"#fff1f2",badge:"#dc2626",icon:"↓",label:"Rückstand"}}
  }};

  var html = "<div style='display:flex;flex-direction:column;gap:8px;'>";
  filtered.forEach(function(d, idx) {{
    var cfg = statusCfg[d._status] || statusCfg.crit;
    var einsaetzeThisYear = d._selectedCount;
    var pct = Math.min(100, Math.round(einsaetzeThisYear / SAM_ZIEL * 100));
    var sollPct = Math.min(100, Math.round(soll / SAM_ZIEL * 100));

    var sortedDaten = (d._byYear[String(selectedYear)]||[]).slice().sort(function(a,b){{
      return String(a.iso||a.datum||"").localeCompare(String(b.iso||b.datum||""));
    }});

    var datesHtml = sortedDaten.length ? sortedDaten.map(function(e) {{
      var infoTxt = String(e.tour||"").trim();
      var startTxt = String(e.beginn||"").trim();
      var infoHtml = infoTxt ? " <b style='color:#1b66b3;'>"+samEsc(infoTxt)+"</b>" : "";
      if(startTxt) infoHtml += " <span style='color:#64748b;'>"+samEsc(startTxt)+" Uhr</span>";
      return "<span style='display:inline-flex;align-items:center;gap:5px;background:#fff;border:1px solid #d8dee8;border-radius:5px;padding:4px 8px;margin:3px;font-size:10.5px;color:#334155;'>" +
        "<b style='color:"+cfg.badge+";'>"+samEsc(e.tag||"Sa")+"</b> "+samEsc(e.datum||"")+infoHtml+"</span>";
    }}).join("") : "<span style='display:inline-block;color:#64748b;font-size:11px;padding:4px 0;'>Keine Einsätze im gewählten Jahr.</span>";

    var otherYears = Object.keys(d._byYear).filter(function(y){{ return y !== String(selectedYear); }}).sort().reverse();
    var prevHtml = "";
    if(otherYears.length) {{
      prevHtml = "<div style='margin-top:7px;display:flex;gap:4px;flex-wrap:wrap;'>";
      otherYears.forEach(function(y) {{
        var n = d._byYear[y].length;
        var metTarget = n >= SAM_ZIEL;
        prevHtml += "<span style='font-size:9px;padding:2px 7px;border-radius:5px;font-weight:800;background:"+
          (metTarget?"#dcfce7":"#fee2e2")+";color:"+(metTarget?"#16a34a":"#dc2626")+";'>"+
          samEsc(y)+": "+n+(metTarget?" ✓":" ✗")+"</span>";
      }});
      prevHtml += "</div>";
    }}

    html +=
      "<div onclick='samToggle(this)' style='background:"+cfg.bg+";border:2px solid "+cfg.border+";border-left-width:8px;border-radius:6px;cursor:pointer;overflow:hidden;box-shadow:0 1px 4px rgba(15,23,42,.06);'>" +
        "<div style='display:grid;grid-template-columns:42px minmax(180px,1fr) 150px 130px 34px;gap:10px;align-items:center;padding:10px 12px;'>" +
          "<div style='font-size:12px;font-weight:900;color:"+cfg.badge+";'>#"+(idx+1)+"</div>" +
          "<div style='min-width:0;'><div style='font-weight:900;font-size:14px;color:#0b1220;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>"+samEsc(d.name)+"</div>" +
            "<div style='display:inline-flex;align-items:center;gap:5px;margin-top:4px;background:"+cfg.badge+";color:#fff;border-radius:5px;padding:2px 8px;font-size:10px;font-weight:900;'>"+cfg.icon+" "+cfg.label+"</div></div>" +
          "<div style='min-width:120px;'><div style='height:12px;background:#e4e9f0;border-radius:999px;position:relative;overflow:hidden;'>" +
            "<div style='position:absolute;left:0;top:0;height:100%;width:"+pct+"%;background:"+cfg.badge+";border-radius:999px;'></div>" +
            "<div style='position:absolute;left:"+sollPct+"%;top:0;width:2px;height:100%;background:#334155;opacity:.45;'></div></div>" +
            "<div style='margin-top:3px;font-size:9px;color:#64748b;font-weight:700;'>"+sollText+": "+soll+" / Ziel: "+SAM_ZIEL+"</div></div>" +
          "<div style='text-align:right;'><div style='font-size:26px;font-weight:950;color:"+cfg.badge+";line-height:1;'>"+einsaetzeThisYear+"</div>" +
            "<div style='font-size:9px;color:#64748b;font-weight:800;'>Einsätze</div></div>" +
          "<div class='sam-arrow' style='font-size:18px;font-weight:900;color:"+cfg.badge+";text-align:right;'>⌄</div></div>" +
        "<div class='sam-dates' style='display:none;background:rgba(255,255,255,.72);border-top:1px solid rgba(148,163,184,.35);padding:9px 12px 10px 64px;'>" +
          "<div style='font-size:10px;text-transform:uppercase;letter-spacing:.35px;font-weight:900;color:#64748b;margin-bottom:5px;'>Einsätze im Jahr "+samEsc(selectedYear)+"</div>" +
          datesHtml + prevHtml + "</div></div>";
  }});

  if(!filtered.length) {{
    html += "<div style='background:#fff;border:1px solid #cbd5e1;border-radius:6px;padding:28px;text-align:center;color:#64748b;font-size:13px;'>Keine Fahrer für diesen Filter.</div>";
  }}
  html += "</div>";
  content.innerHTML = html;
}}

function samExportExcel() {{
  if(typeof XLSX === "undefined") {{
    alert("Excel-Bibliothek nicht geladen. Bitte die Seite neu laden.");
    return;
  }}

  var year = String(samYearFilter || new Date().getFullYear());
  var rows = [];
  (samLastFiltered || []).forEach(function(driver) {{
    var entries = ((driver._byYear || {{}})[year] || []).slice();
    entries.forEach(function(e) {{
      var datumText = String(e.datum || "");
      var dateMatch = datumText.match(/\\d{{2}}\\.\\d{{2}}\\.\\d{{4}}/);
      var kwMatch = datumText.match(/KW\\s*(\\d{{1,2}})/i);
      var tag = String(e.tag || "Sa");
      var art = tag === "So" ? "Sonntag bis 15 Uhr" :
                (tag.indexOf("Fr") === 0 ? "Freitag ab 18 Uhr" : "Samstag");
      var lkw = String(e.tour || "").replace(/^LKW\\s*/i, "").trim();
      rows.push({{
        _iso: String(e.iso || ""),
        Datum: dateMatch ? dateMatch[0] : datumText,
        KW: kwMatch ? parseInt(kwMatch[1], 10) : "",
        Einsatzart: art,
        Fahrer: String(driver.name || ""),
        Beginn: String(e.beginn || ""),
        LKW: lkw
      }});
    }});
  }});

  rows.sort(function(a,b) {{
    var d = String(a._iso).localeCompare(String(b._iso));
    return d !== 0 ? d : String(a.Fahrer).localeCompare(String(b.Fahrer), "de");
  }});

  if(!rows.length) {{
    alert("Für die aktuelle Auswahl sind keine Einsätze vorhanden.");
    return;
  }}

  var headers = ["Datum", "KW", "Einsatzart", "Fahrer", "Beginn", "LKW"];
  var data = [headers].concat(rows.map(function(r) {{
    return [r.Datum, r.KW, r.Einsatzart, r.Fahrer, r.Beginn, r.LKW];
  }}));
  var ws = XLSX.utils.aoa_to_sheet(data);
  ws["!cols"] = [
    {{wch:12}}, {{wch:6}}, {{wch:21}}, {{wch:28}}, {{wch:14}}, {{wch:18}}
  ];
  ws["!autofilter"] = {{ref:"A1:F" + data.length}};

  headers.forEach(function(_, idx) {{
    var ref = XLSX.utils.encode_cell({{r:0,c:idx}});
    if(ws[ref]) ws[ref].s = {{
      font: {{bold:true,color:{{rgb:"FFFFFF"}}}},
      fill: {{fgColor:{{rgb:"1B66B3"}}}},
      alignment: {{horizontal:"center",vertical:"center"}}
    }};
  }});

  for(var r=1; r<data.length; r++) {{
    var kwRef = XLSX.utils.encode_cell({{r:r,c:1}});
    if(ws[kwRef]) ws[kwRef].s = {{alignment:{{horizontal:"center"}}}};
  }}

  var wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, "Samstagseinsätze");
  var suffix = samStatusFilter !== "all" ? "_" + samStatusFilter : "";
  XLSX.writeFile(wb, "Samstagseinsaetze_" + year + suffix + ".xlsx");
}}

function samRenderCharts(drivers, selectedYear, soll, satTotal, satElapsed) {{
  if(samChartMode === "drivers") return samRenderDriverBars(drivers, selectedYear, soll);
  if(samChartMode === "months") return samRenderMonths(drivers, selectedYear);
  return samRenderMatrix(drivers, selectedYear, satTotal, satElapsed);
}}

function samLegend() {{
  return "<div style='display:flex;gap:12px;align-items:center;flex-wrap:wrap;font-size:10px;font-weight:800;color:#475569;'>"+
    "<span><i style='display:inline-block;width:11px;height:11px;border-radius:3px;background:#7c3aed;margin-right:4px;vertical-align:-1px;'></i>Fr→Sa</span>"+
    "<span><i style='display:inline-block;width:11px;height:11px;border-radius:3px;background:#1b66b3;margin-right:4px;vertical-align:-1px;'></i>Samstag</span>"+
    "<span><i style='display:inline-block;width:11px;height:11px;border-radius:3px;background:#d97706;margin-right:4px;vertical-align:-1px;'></i>Sonntag</span>"+
    "<span><i style='display:inline-block;width:11px;height:11px;border-radius:3px;background:#0f766e;margin-right:4px;vertical-align:-1px;'></i>mehrere Einsätze</span>"+
    "</div>";
}}

// Klickbare Arbeitstage in der Einsatzmatrix
var samWorkdayPopupData = {{}};
var samWorkdayBodyOverflow = "";

function samWorkdayTypeLabel(tag) {{
  tag = String(tag || "Sa");
  if(tag === "So") return "Sonntag bis 15 Uhr";
  if(tag.indexOf("Fr") === 0) return "Freitag ab 18 Uhr";
  return "Samstag";
}}

function samWorkdayBadgeColor(tag) {{
  tag = String(tag || "Sa");
  if(tag === "So") return "#d97706";
  if(tag.indexOf("Fr") === 0) return "#7c3aed";
  return "#1b66b3";
}}

function samCloseWorkdayPopup() {{
  var overlay = document.getElementById("sam-workday-popup");
  if(overlay) overlay.style.display = "none";
  document.body.style.overflow = samWorkdayBodyOverflow || "";
}}

function samOpenWorkdayPopup(key) {{
  var data = samWorkdayPopupData[key];
  if(!data) return;

  var overlay = document.getElementById("sam-workday-popup");
  if(!overlay) {{
    overlay = document.createElement("div");
    overlay.id = "sam-workday-popup";
    overlay.setAttribute("role", "presentation");
    overlay.onclick = function(ev) {{
      if(ev.target === overlay) samCloseWorkdayPopup();
    }};
    document.body.appendChild(overlay);
  }}

  var entries = Array.isArray(data.entries) ? data.entries : [];
  var cards = entries.map(function(e, idx) {{
    var tag = String(e.tag || "Sa");
    var badgeColor = samWorkdayBadgeColor(tag);
    var tour = String(e.tour || "").trim() || "—";
    var beginn = String(e.beginn || "").trim() || "—";
    var datum = String(e.datum || "").trim() || "—";
    return "<div style='border:1px solid #d8e0ea;border-left:5px solid "+badgeColor+";border-radius:8px;background:#fff;padding:12px 14px;box-shadow:0 1px 3px rgba(15,23,42,.06);'>" +
      "<div style='display:flex;align-items:center;justify-content:space-between;gap:10px;margin-bottom:10px;'>" +
        "<span style='display:inline-flex;align-items:center;border-radius:5px;background:"+badgeColor+";color:#fff;padding:4px 9px;font-size:11px;font-weight:900;'>" +
          samEsc(samWorkdayTypeLabel(tag)) + "</span>" +
        (entries.length > 1 ? "<span style='font-size:10px;font-weight:800;color:#64748b;'>Eintrag "+(idx+1)+" von "+entries.length+"</span>" : "") +
      "</div>" +
      "<div style='display:grid;grid-template-columns:110px minmax(0,1fr);gap:8px 12px;font-size:12px;line-height:1.4;'>" +
        "<div style='font-weight:800;color:#64748b;'>Fahrer</div><div style='font-weight:900;color:#0f172a;'>"+samEsc(data.driver)+"</div>" +
        "<div style='font-weight:800;color:#64748b;'>Arbeitstag</div><div style='font-weight:800;color:#0f172a;'>"+samEsc(datum)+"</div>" +
        "<div style='font-weight:800;color:#64748b;'>Beginn</div><div style='font-weight:800;color:#0f172a;'>"+samEsc(beginn)+(beginn !== "—" ? " Uhr" : "")+"</div>" +
        "<div style='font-weight:800;color:#64748b;'>LKW</div><div style='font-weight:900;color:#1b66b3;'>"+samEsc(tour)+"</div>" +
      "</div>" +
    "</div>";
  }}).join("");

  if(!cards) {{
    cards = "<div style='padding:24px;text-align:center;color:#64748b;'>Für diesen Arbeitstag sind keine Detaildaten vorhanden.</div>";
  }}

  overlay.style.cssText = "display:flex;position:fixed;inset:0;z-index:99999;align-items:center;justify-content:center;padding:22px;background:rgba(15,23,42,.58);backdrop-filter:blur(2px);";
  overlay.innerHTML =
    "<div role='dialog' aria-modal='true' aria-label='Arbeitstag Details' onclick='event.stopPropagation()' " +
      "style='width:min(660px,96vw);max-height:88vh;display:flex;flex-direction:column;background:#f8fafc;border:1px solid #cbd5e1;border-radius:12px;box-shadow:0 24px 70px rgba(15,23,42,.35);overflow:hidden;font-family:Segoe UI,Arial,sans-serif;'>" +
      "<div style='display:flex;align-items:flex-start;justify-content:space-between;gap:16px;padding:16px 18px;background:linear-gradient(180deg,#eef6ff 0%,#fff 100%);border-bottom:1px solid #dbe3ed;'>" +
        "<div style='min-width:0;'>" +
          "<div style='font-size:18px;font-weight:950;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>Arbeitstag – "+samEsc(data.driver)+"</div>" +
          "<div style='margin-top:4px;font-size:11px;font-weight:750;color:#64748b;'>Wochenende KW "+samEsc(data.kw)+" · Samstag, "+samEsc(data.weekendDate)+"</div>" +
        "</div>" +
        "<button type='button' onclick='samCloseWorkdayPopup()' aria-label='Fenster schließen' " +
          "style='flex:0 0 auto;width:34px;height:34px;border:1px solid #cbd5e1;border-radius:7px;background:#fff;color:#334155;font-size:21px;line-height:1;cursor:pointer;font-weight:700;'>&times;</button>" +
      "</div>" +
      "<div style='overflow-y:auto;padding:14px;display:flex;flex-direction:column;gap:10px;'>" + cards + "</div>" +
      "<div style='padding:10px 14px;border-top:1px solid #dbe3ed;background:#fff;display:flex;justify-content:flex-end;'>" +
        "<button type='button' onclick='samCloseWorkdayPopup()' style='padding:8px 16px;border:1px solid #1b66b3;border-radius:6px;background:#1b66b3;color:#fff;font-size:11px;font-weight:900;cursor:pointer;font-family:inherit;'>Schließen</button>" +
      "</div>" +
    "</div>";

  samWorkdayBodyOverflow = document.body.style.overflow || "";
  document.body.style.overflow = "hidden";
}}

document.addEventListener("keydown", function(ev) {{
  if(ev.key === "Escape") samCloseWorkdayPopup();
}});

function samRenderMatrix(drivers, year, satTotal, satElapsed) {{
  var content = document.getElementById("sam-content");
  if(!content) return;
  samWorkdayPopupData = {{}};
  var sats = [];
  var d = new Date(year,0,1);
  while(d.getDay() !== 6) d.setDate(d.getDate()+1);
  while(d.getFullYear() === year) {{ sats.push(new Date(d)); d.setDate(d.getDate()+7); }}

  var html = "<div style='background:#fff;border:1px solid #cbd5e1;border-radius:7px;box-shadow:0 1px 5px rgba(15,23,42,.06);overflow:hidden;'>"+
    "<div style='display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap;padding:11px 13px;border-bottom:1px solid #e2e8f0;background:#f8fafc;'>"+
      "<div style='font-size:14px;font-weight:900;color:#0f172a;'>Einsatzmatrix "+samEsc(year)+"</div>"+
      samLegend()+"</div>";

  if(!drivers.length) {{
    content.innerHTML = html + "<div style='padding:30px;text-align:center;color:#64748b;'>Keine Fahrer für diesen Filter.</div></div>";
    return;
  }}

  html += "<div style='overflow-x:auto;overflow-y:hidden;width:100%;max-height:none;'>"+
    "<table style='border-collapse:separate;border-spacing:0;table-layout:fixed;min-width:"+(210+sats.length*39)+"px;width:max-content;font-size:9px;'>"+
    "<thead><tr><th style='position:sticky;left:0;top:0;z-index:5;width:210px;min-width:210px;background:#e2e8f0;border-right:2px solid #94a3b8;border-bottom:1px solid #94a3b8;padding:7px 9px;text-align:left;font-size:10px;color:#334155;'>Fahrer</th>";
  sats.forEach(function(s, idx) {{
    var future = year === new Date().getFullYear() && idx >= satElapsed;
    html += "<th title='Samstag, "+samAttr(samShortDate(s)+year)+"' style='position:sticky;top:0;z-index:3;width:39px;min-width:39px;max-width:39px;background:"+(future?"#f8fafc":"#e2e8f0")+";border-right:1px solid #cbd5e1;border-bottom:1px solid #94a3b8;padding:4px 1px;text-align:center;color:"+(future?"#94a3b8":"#334155")+";'>"+
      "<div style='font-size:8px;font-weight:900;'>KW"+samISOWeek(s)+"</div><div style='font-size:7.5px;'>"+samShortDate(s)+"</div></th>";
  }});
  html += "<th style='position:sticky;right:0;top:0;z-index:5;width:54px;min-width:54px;background:#e2e8f0;border-left:2px solid #94a3b8;border-bottom:1px solid #94a3b8;text-align:center;color:#334155;'>Σ</th></tr></thead><tbody>";

  drivers.slice().sort(function(a,b){{return String(a.name||"").localeCompare(String(b.name||""),"de");}}).forEach(function(driver, rowIdx) {{
    var weekendMap = {{}};
    var entries = driver._byYear[String(year)]||[];
    entries.forEach(function(e) {{
      var sat = samWeekendSaturday(e);
      if(!sat || sat.getFullYear() !== year) return;
      var key = samISODate(sat);
      if(!weekendMap[key]) weekendMap[key] = [];
      weekendMap[key].push(e);
    }});
    var rowBg = rowIdx%2 ? "#f8fafc" : "#fff";
    html += "<tr><td style='position:sticky;left:0;z-index:2;width:210px;max-width:210px;background:"+rowBg+";border-right:2px solid #94a3b8;border-bottom:1px solid #e2e8f0;padding:6px 9px;font-size:10.5px;font-weight:800;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;' title='"+samAttr(driver.name)+"'>"+samEsc(driver.name)+"</td>";
    sats.forEach(function(s) {{
      var list = weekendMap[samISODate(s)]||[];
      if(!list.length) {{
        html += "<td style='width:39px;height:30px;background:"+rowBg+";border-right:1px solid #e2e8f0;border-bottom:1px solid #e2e8f0;text-align:center;color:#cbd5e1;'>·</td>";
        return;
      }}
      var tags = list.map(function(e){{return String(e.tag||"Sa");}});
      var color = list.length > 1 ? "#0f766e" : (tags[0] === "So" ? "#d97706" : (tags[0].indexOf("Fr") === 0 ? "#7c3aed" : "#1b66b3"));
      var label = list.length > 1 ? String(list.length)+"×" : (tags[0] === "So" ? "So" : (tags[0].indexOf("Fr") === 0 ? "Fr" : "Sa"));
      var title = driver.name + "\\n" + list.map(function(e){{
        return String(e.tag||"Sa")+" · "+String(e.datum||"")+(e.beginn?" · "+e.beginn+" Uhr":"")+(e.tour?" · "+e.tour:"");
      }}).join("\\n");
      var workdayKey = "sam-workday-"+rowIdx+"-"+samISODate(s);
      samWorkdayPopupData[workdayKey] = {{
        driver: String(driver.name || ""),
        weekendDate: samShortDate(s) + year,
        kw: samISOWeek(s),
        entries: list.map(function(e) {{
          return {{
            tag: String(e.tag || "Sa"),
            datum: String(e.datum || ""),
            beginn: String(e.beginn || ""),
            tour: String(e.tour || "")
          }};
        }})
      }};
      html += "<td title='"+samAttr(title)+"\\nKlicken für Details' onclick='samOpenWorkdayPopup(&quot;"+workdayKey+"&quot;);event.stopPropagation();' " +
        "style='width:39px;height:30px;background:"+rowBg+";border-right:1px solid #e2e8f0;border-bottom:1px solid #e2e8f0;text-align:center;padding:2px;cursor:pointer;'>"+
        "<span style='display:inline-flex;align-items:center;justify-content:center;width:28px;height:22px;border-radius:4px;background:"+color+";color:#fff;font-size:8px;font-weight:950;box-shadow:0 1px 2px rgba(15,23,42,.15);transition:transform .12s,box-shadow .12s;'>"+label+"</span></td>";
    }});
    html += "<td style='position:sticky;right:0;z-index:2;width:54px;background:"+rowBg+";border-left:2px solid #94a3b8;border-bottom:1px solid #e2e8f0;text-align:center;font-size:12px;font-weight:950;color:#1b66b3;'>"+entries.length+"</td></tr>";
  }});
  html += "</tbody></table></div></div>";
  content.innerHTML = html;
}}

function samRenderDriverBars(drivers, year, soll) {{
  var content = document.getElementById("sam-content");
  if(!content) return;
  if(!drivers.length) {{
    content.innerHTML = "<div style='background:#fff;border:1px solid #cbd5e1;border-radius:7px;padding:30px;text-align:center;color:#64748b;'>Keine Fahrer für diesen Filter.</div>";
    return;
  }}
  var list = drivers.slice().sort(function(a,b){{return b._selectedCount-a._selectedCount || String(a.name).localeCompare(String(b.name),"de");}});
  var maxVal = Math.max.apply(null, list.map(function(d){{return d._selectedCount;}}).concat([SAM_ZIEL,1]));
  var html = "<div style='background:#fff;border:1px solid #cbd5e1;border-radius:7px;padding:13px;box-shadow:0 1px 5px rgba(15,23,42,.06);'>"+
    "<div style='display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap;margin-bottom:13px;'>"+
    "<div><div style='font-size:14px;font-weight:900;color:#0f172a;'>Einsätze je Fahrer "+samEsc(year)+"</div><div style='font-size:10.5px;color:#64748b;margin-top:2px;'>Gestapelt nach Einsatzart; die dunkle Markierung zeigt das Jahresziel "+SAM_ZIEL+".</div></div>"+samLegend()+"</div>";
  list.forEach(function(d) {{
    var entries = d._byYear[String(year)]||[];
    var fr=0,sa=0,so=0;
    entries.forEach(function(e){{var t=String(e.tag||"Sa"); if(t==="So")so++; else if(t.indexOf("Fr")===0)fr++; else sa++;}});
    var frPct=fr/maxVal*100, saPct=sa/maxVal*100, soPct=so/maxVal*100, zielPct=SAM_ZIEL/maxVal*100;
    html += "<div style='display:grid;grid-template-columns:minmax(170px,260px) minmax(280px,1fr) 52px;gap:10px;align-items:center;padding:7px 4px;border-top:1px solid #eef2f7;'>"+
      "<div style='font-size:11.5px;font-weight:850;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;' title='"+samAttr(d.name)+"'>"+samEsc(d.name)+"</div>"+
      "<div><div style='height:18px;background:#eef2f7;border-radius:4px;position:relative;overflow:hidden;display:flex;'>"+
        "<span title='Fr→Sa: "+fr+"' style='height:100%;width:"+frPct+"%;background:#7c3aed;'></span>"+
        "<span title='Samstag: "+sa+"' style='height:100%;width:"+saPct+"%;background:#1b66b3;'></span>"+
        "<span title='Sonntag: "+so+"' style='height:100%;width:"+soPct+"%;background:#d97706;'></span>"+
        "<i style='position:absolute;left:"+zielPct+"%;top:0;width:2px;height:100%;background:#0f172a;opacity:.65;'></i>"+
      "</div><div style='display:flex;gap:8px;margin-top:2px;font-size:8.5px;color:#64748b;font-weight:700;'><span>Fr "+fr+"</span><span>Sa "+sa+"</span><span>So "+so+"</span><span style='margin-left:auto;'>Soll aktuell "+soll+"</span></div></div>"+
      "<div style='font-size:20px;font-weight:950;text-align:right;color:"+(entries.length>=SAM_ZIEL?"#16a34a":"#1b66b3")+";'>"+entries.length+"</div></div>";
  }});
  html += "</div>";
  content.innerHTML = html;
}}

function samRenderMonths(drivers, year) {{
  var content = document.getElementById("sam-content");
  if(!content) return;
  var months = Array.from({{length:12}},function(){{return {{fr:0,sa:0,so:0,total:0,drivers:{{}}}};}});
  drivers.forEach(function(d) {{
    (d._byYear[String(year)]||[]).forEach(function(e) {{
      var dt = samDateFromEntry(e); if(!dt) return;
      var m = months[dt.getMonth()];
      var t = String(e.tag||"Sa");
      if(t === "So") m.so++; else if(t.indexOf("Fr")===0) m.fr++; else m.sa++;
      m.total++;
      m.drivers[d.person_key||d.name] = 1;
    }});
  }});
  var maxVal = Math.max.apply(null, months.map(function(m){{return m.total;}}).concat([1]));
  var names = ["Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"];
  var html = "<div style='background:#fff;border:1px solid #cbd5e1;border-radius:7px;padding:13px;box-shadow:0 1px 5px rgba(15,23,42,.06);'>"+
    "<div style='display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap;margin-bottom:12px;'>"+
    "<div><div style='font-size:14px;font-weight:900;color:#0f172a;'>Monatsverteilung "+samEsc(year)+"</div><div style='font-size:10.5px;color:#64748b;margin-top:2px;'>Anzahl aller Einsätze der aktuell gefilterten Fahrer pro Monat.</div></div>"+samLegend()+"</div>"+
    "<div style='height:310px;display:flex;align-items:flex-end;gap:8px;border-left:1px solid #cbd5e1;border-bottom:1px solid #cbd5e1;padding:14px 10px 0 10px;background:linear-gradient(to top,#f8fafc,#fff);'>";
  months.forEach(function(m,idx) {{
    var h = Math.round(m.total/maxVal*245);
    var hFr = m.total ? Math.round(h*m.fr/m.total) : 0;
    var hSa = m.total ? Math.round(h*m.sa/m.total) : 0;
    var hSo = Math.max(0,h-hFr-hSa);
    var title = names[idx]+" "+year+": "+m.total+" Einsätze · "+Object.keys(m.drivers).length+" Fahrer · Fr "+m.fr+" · Sa "+m.sa+" · So "+m.so;
    html += "<div title='"+samAttr(title)+"' style='flex:1;min-width:48px;height:285px;display:flex;flex-direction:column;justify-content:flex-end;align-items:center;'>"+
      "<div style='font-size:11px;font-weight:950;color:#0f172a;margin-bottom:4px;'>"+m.total+"</div>"+
      "<div style='width:min(42px,80%);height:"+h+"px;min-height:"+(m.total?4:0)+"px;display:flex;flex-direction:column-reverse;border-radius:5px 5px 0 0;overflow:hidden;box-shadow:"+(m.total?"0 1px 3px rgba(15,23,42,.18)":"none")+";'>"+
        "<span style='height:"+hFr+"px;background:#7c3aed;'></span>"+
        "<span style='height:"+hSa+"px;background:#1b66b3;'></span>"+
        "<span style='height:"+hSo+"px;background:#d97706;'></span></div>"+
      "<div style='font-size:10px;font-weight:850;color:#475569;margin-top:5px;'>"+names[idx]+"</div>"+
      "<div style='font-size:8px;color:#94a3b8;'>"+Object.keys(m.drivers).length+" F.</div></div>";
  }});
  html += "</div></div>";
  content.innerHTML = html;
}}

function samToggle(el) {{
  var dates = el.querySelector(".sam-dates");
  if(!dates) return;
  var open = dates.style.display !== "none";
  dates.style.display = open ? "none" : "block";
  var arrow = el.querySelector(".sam-arrow");
  if(arrow) arrow.textContent = open ? "⌄" : "⌃";
}}


{fa_js_code}

{wa_js_code}

{bus_js_code}

{arzt_js_code}

{knapp_js_code}

{documents_js_code}

// VERSP_ABFAHRT wird aus dem komprimierten Datenblock geladen.
{versp_js_code}

// ── Fahrerauswertung: Schichten-Tab (Tachograph-Daten) ─────────────────────────
(function() {{
  var _faShowDetailOrig = window.faShowDetail;
  if (typeof _faShowDetailOrig !== "function") return;

  var faShiftMonthFilterByDriver = {{}};

  function faHasShifts(name) {{
    return TIMEREC_DATA && TIMEREC_DATA[name] && TIMEREC_DATA[name].length > 0;
  }}

  function faTabBar(driverName, hasShifts, hasUebersicht) {{
    // Immer beide Reiter anzeigen, damit der Umschalter an derselben Stelle bleibt.
    // Fehlende Daten werden nur deaktiviert dargestellt.
    var tabs = [
      ["uebersicht", "Übersicht", !!hasUebersicht, "Keine Übersichtsdaten für diesen Fahrer"],
      ["schichten",  "Schichten",  !!hasShifts,     "Keine Schichten / Tachograph-Daten für diesen Fahrer"]
    ];
    var html = "<div class='fa-tabs' style='display:flex;gap:2px;margin-bottom:14px;border-bottom:1px solid #e2e8f0;'>";
    tabs.forEach(function(t) {{
      var active = faActiveTab === t[0] && t[2];
      var enabled = t[2];
      html += "<button type='button' "
            + (enabled ? "onclick=\\\"faSwitchTab('" + t[0] + "')\\\"" : "disabled title='" + faEsc(t[3]) + "'")
            + " style='padding:9px 18px;border:none;background:none;font-family:inherit;"
            + "cursor:" + (enabled ? "pointer" : "not-allowed") + ";"
            + "font-size:13px;font-weight:" + (active ? "800" : "600") + ";"
            + "opacity:" + (enabled ? "1" : ".42") + ";"
            + "color:" + (active ? "#334155" : "#64748b") + ";"
            + "border-bottom:2px solid " + (active ? "#334155" : "transparent") + ";"
            + "margin-bottom:-1px;'>" + t[1] + "</button>";
    }});
    html += "</div>";
    return html;
  }}

  function faEsc(v) {{
    return String(v == null ? "" : v)
      .replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;")
      .replace(/"/g,"&quot;").replace(/'/g,"&#39;");
  }}

  function _toMin(t) {{
    if (t == null) return 0;
    var s = String(t).trim();
    if (!s || s === "-" || s === "—") return 0;

    var m = s.match(/(-?\\d{{1,4}})\\s*:\\s*(\\d{{1,2}})/);
    if (m) {{
      var h = parseInt(m[1], 10) || 0;
      var mi = parseInt(m[2], 10) || 0;
      return h * 60 + mi;
    }}

    var h2 = 0, m2 = 0;
    var hm = s.match(/(\\d+(?:[\\.,]\\d+)?)\\s*(?:std|stunde|stunden|h)\b/i);
    var mm = s.match(/(\\d+)\\s*(?:min|minute|minuten|m)\b/i);
    if (hm || mm) {{
      if (hm) h2 = parseFloat(hm[1].replace(",", ".")) || 0;
      if (mm) m2 = parseInt(mm[1], 10) || 0;
      return Math.round(h2 * 60 + m2);
    }}

    var dec = parseFloat(s.replace(",", "."));
    if (!isNaN(dec)) return Math.round(dec * 60);
    return 0;
  }}

  function _fmtMin(m) {{
    m = Math.round(Number(m) || 0);
    var neg = m < 0;
    if (neg) m = Math.abs(m);
    var h = Math.floor(m / 60);
    var mm = m % 60;
    return (neg ? "-" : "") + h + ":" + (mm < 10 ? "0" + mm : mm);
  }}

  // ── Download-Cutoff-Erkennung ────────────────────────────────────────────
  // Tachograph-Downloads erzeugen für noch laufende Schichten einen
  // künstlichen Schichtende-Zeitpunkt (= Auslesezeitpunkt).  Wenn ≥10
  // Schichten denselben Ende-Zeitpunkt (Datum+Uhrzeit) teilen, handelt es
  // sich mit Sicherheit um einen solchen Cutoff.
  var _faCutoffSet = null;

  function _faComputeEndDate(s) {{
    var d = s.tag || "";
    if (!s.ende_naechster_tag) return d;
    var m = d.match(/^(\\d{{2}})\\.(\\d{{2}})\\.(\\d{{4}})$/);
    if (!m) return d;
    var dt = new Date(parseInt(m[3],10), parseInt(m[2],10)-1, parseInt(m[1],10)+1);
    return (dt.getDate()<10?"0":"") + dt.getDate() + "."
         + ((dt.getMonth()+1)<10?"0":"") + (dt.getMonth()+1) + "."
         + dt.getFullYear();
  }}

  function _faGetCutoffSet() {{
    if (_faCutoffSet) return _faCutoffSet;
    var freq = {{}};
    Object.keys(TIMEREC_DATA || {{}}).forEach(function(name) {{
      (TIMEREC_DATA[name] || []).forEach(function(s) {{
        if (!s.ende) return;
        var key = _faComputeEndDate(s) + " " + s.ende;
        freq[key] = (freq[key] || 0) + 1;
      }});
    }});
    _faCutoffSet = {{}};
    Object.keys(freq).forEach(function(k) {{
      if (freq[k] >= 10) _faCutoffSet[k] = true;
    }});
    return _faCutoffSet;
  }}

  function _faIsShiftCutoff(s) {{
    if (!s.ende) return false;
    var key = _faComputeEndDate(s) + " " + s.ende;
    return !!_faGetCutoffSet()[key];
  }}

  function _faIsShiftInvalid(s) {{
    // 1) Noch laufend: Schichtende = Tachograph-Download-Zeitpunkt
    if (_faIsShiftCutoff(s)) return true;
    // 2) Tachograph-Fehler: Schichtdauer > 24 Stunden
    if (_toMin(s.schichtdauer) > 1440) return true;
    return false;
  }}

  function _vacationCreditMin(rows) {{
    return (rows || []).length * 480;
  }}

  function _monthInfo(s) {{
    var MONATE = ["Januar","Februar","März","April","Mai","Juni","Juli","August","September","Oktober","November","Dezember"];
    var m = (s.tag || "").match(/^(\\d{{2}})\\.(\\d{{2}})\\.(\\d{{4}})$/);
    if (!m) return {{ key: "0000-00", label: "Unbekannt" }};
    return {{ key: m[3] + "-" + m[2], label: (MONATE[parseInt(m[2],10)-1] || m[2]) + " " + m[3] }};
  }}


  function _monthLabelFromKey(key) {{
    var MONATE = ["Januar","Februar","März","April","Mai","Juni","Juli","August","September","Oktober","November","Dezember"];
    var m = String(key || "").match(/^(\\d{{4}})-(\\d{{2}})$/);
    if (!m) return "Unbekannt";
    return (MONATE[parseInt(m[2],10)-1] || m[2]) + " " + m[1];
  }}

  var _faPlanDriverIndex = null;
  // Manuelle Ausnahmezuordnung, falls Schreibweisen trotz Normalisierung nicht eindeutig passen.
  // Links: Name aus Tachograph-CSV, rechts: Name aus Tourenplanung/Fahrerauswertung.
  var FA_NAME_ALIASES = {{
    // Beispiel: "Mustermann, H." : "Mustermann, Hans"
  }};

  function _faNameNorm(value) {{
    var s = String(value == null ? "" : value).toLowerCase();
    try {{ s = s.normalize("NFKD").replace(/[\u0300-\u036f]/g, ""); }} catch(e) {{}}
    s = s.replace(/ß/g, "ss").replace(/æ/g, "ae").replace(/œ/g, "oe");
    s = s.replace(/\\([^)]*\\)/g, " ");
    s = s.replace(/[_.\\/,;:|]+/g, " ");
    s = s.replace(/[-–—]+/g, " ");
    s = s.replace(/\\s+/g, " ").trim();
    return s;
  }}

  function _faNameTokens(value) {{
    return _faNameNorm(value).split(" ").filter(function(t) {{
      return t && t !== "fahrer" && t !== "fahrerin" && t !== "herr" && t !== "frau";
    }});
  }}

  function _faNameCombos(value) {{
    var raw = String(value == null ? "" : value);
    var combos = [];
    if (raw.indexOf(",") >= 0) {{
      var parts = raw.split(",");
      var last = _faNameNorm(parts.shift() || "");
      var first = _faNameNorm(parts.join(" ") || "");
      if (last || first) combos.push({{ last:last, first:first }});
    }}
    var t = _faNameTokens(raw);
    if (t.length >= 2) {{
      // Variante Vorname Nachname
      combos.push({{ last:t[t.length-1], first:t.slice(0, -1).join(" ") }});
      // Variante Nachname Vorname
      combos.push({{ last:t[0], first:t.slice(1).join(" ") }});
    }} else if (t.length === 1) {{
      combos.push({{ last:t[0], first:"" }});
    }}
    var seen = {{}};
    return combos.filter(function(c) {{
      var k = c.last + "|" + c.first;
      if (seen[k]) return false;
      seen[k] = true;
      return !!(c.last || c.first);
    }});
  }}

  function _faNameKeys(value, initialOnly) {{
    var out = {{}};
    var norm = _faNameNorm(value);
    if (norm && !initialOnly) out[norm] = true;
    var tokens = _faNameTokens(value);
    if (tokens.length > 1 && !initialOnly) {{
      out[tokens.join(" ")] = true;
      out[tokens.slice().reverse().join(" ")] = true;
    }}
    _faNameCombos(value).forEach(function(c) {{
      if (!c.last) return;
      if (!initialOnly) {{
        out[c.last + "|" + c.first] = true;
        if (c.first) {{
          out[c.last + " " + c.first] = true;
          out[c.first + " " + c.last] = true;
        }}
      }}
      var firstInitial = (c.first || "").charAt(0);
      if (firstInitial) out[c.last + "|" + firstInitial] = true;
    }});
    return Object.keys(out);
  }}

  function _faIndexAdd(bucket, key, driver) {{
    if (!key) return;
    if (!bucket[key]) bucket[key] = [];
    if (bucket[key].indexOf(driver) < 0) bucket[key].push(driver);
  }}

  function _faBuildPlanDriverIndex() {{
    var idx = {{ exact:{{}}, initial:{{}}, last:{{}} }};
    (FA_DATA || []).forEach(function(d) {{
      if (!d || !d.name) return;
      _faNameKeys(d.name, false).forEach(function(k) {{ _faIndexAdd(idx.exact, k, d); }});
      _faNameKeys(d.name, true).forEach(function(k) {{ _faIndexAdd(idx.initial, k, d); }});
      _faNameCombos(d.name).forEach(function(c) {{ _faIndexAdd(idx.last, c.last, d); }});
    }});
    return idx;
  }}

  function _faPickUniqueFromIndex(bucket, keys) {{
    for (var i=0; i<keys.length; i++) {{
      var arr = bucket[keys[i]] || [];
      if (arr.length === 1) return arr[0];
    }}
    return null;
  }}

  function _faPlanDriver(name) {{
    if (!Array.isArray(FA_DATA) || !FA_DATA.length) return null;

    // 1) Direkter Treffer bleibt bevorzugt.
    var direct = FA_DATA.find(function(d) {{ return d && d.name === name; }});
    if (direct) return direct;

    // 1b) Manuelle Ausnahmen, wenn ein Name in den Dateien wirklich unterschiedlich geschrieben ist.
    var alias = FA_NAME_ALIASES[name] || FA_NAME_ALIASES[_faNameNorm(name)] || "";
    if (alias) {{
      var aliasNorm = _faNameNorm(alias);
      var aliasDriver = FA_DATA.find(function(d) {{
        return d && (d.name === alias || _faNameNorm(d.name) === aliasNorm);
      }});
      if (aliasDriver) return aliasDriver;
    }}

    // 2) Robuster Namensabgleich: Komma egal, Reihenfolge egal, Umlaute egal, Bindestriche egal.
    if (!_faPlanDriverIndex) _faPlanDriverIndex = _faBuildPlanDriverIndex();

    var exact = _faPickUniqueFromIndex(_faPlanDriverIndex.exact, _faNameKeys(name, false));
    if (exact) return exact;

    // 3) Fallback: gleicher Nachname + eindeutiger Vorname-Anfangsbuchstabe.
    var initial = _faPickUniqueFromIndex(_faPlanDriverIndex.initial, _faNameKeys(name, true));
    if (initial) return initial;

    // 4) Sehr vorsichtiger Fallback: gleicher Nachname und eindeutiger Kandidat.
    var combos = _faNameCombos(name);
    for (var c=0; c<combos.length; c++) {{
      var candidates = _faPlanDriverIndex.last[combos[c].last] || [];
      if (candidates.length === 1) return candidates[0];
    }}

    return null;
  }}

  function _faPlanDateKey(entry) {{
    var raw = String((entry && entry.datum) || "");
    var m = raw.match(/(\\d{{2}})\\.(\\d{{2}})\\.(\\d{{4}})/);
    if (!m) return "";
    return m[3] + "-" + m[2] + "-" + m[1];
  }}

  function _faPlanMonthKey(entry) {{
    var dk = _faPlanDateKey(entry);
    return dk ? dk.slice(0, 7) : "";
  }}

  function _faAbsenceEntries(name, yr, monthFilter, kind) {{
    var driver = _faPlanDriver(name);
    if (!driver || !driver.years) return [];
    var needle = String(kind || "krank").toLowerCase();
    var label = needle === "urlaub" ? "Urlaub" : "Krank";
    var years = yr === "all" ? Object.keys(driver.years || {{}}) : [yr];
    var seen = {{}};
    var out = [];
    years.forEach(function(y) {{
      var data = driver.years[y];
      if (!data) return;
      (data.eintraege || []).forEach(function(e) {{
        var tour = String(e.tour || "").toLowerCase();
        if (tour.indexOf(needle) < 0) return;
        var dk = _faPlanDateKey(e);
        var mk = _faPlanMonthKey(e);
        if (!dk || !mk) return;
        if (monthFilter && monthFilter !== "all" && mk !== monthFilter) return;
        if (seen[dk]) return;
        seen[dk] = true;
        out.push({{ dateKey: dk, monthKey: mk, datum: e.datum || dk, tour: e.tour || label }});
      }});
    }});
    out.sort(function(a,b) {{ return a.dateKey.localeCompare(b.dateKey); }});
    return out;
  }}

  function _faSickEntries(name, yr, monthFilter) {{
    return _faAbsenceEntries(name, yr, monthFilter, "krank");
  }}

  function _faVacationEntries(name, yr, monthFilter) {{
    return _faAbsenceEntries(name, yr, monthFilter, "urlaub");
  }}

  function _faAddAbsenceMonths(name, yr, byMonth) {{
    _faSickEntries(name, yr, "all").concat(_faVacationEntries(name, yr, "all")).forEach(function(e) {{
      if (!byMonth[e.monthKey]) byMonth[e.monthKey] = {{ label: _monthLabelFromKey(e.monthKey), shifts: [] }};
    }});
  }}

  function _renderAbsenceBox(rows, cfg) {{
    if (!rows || !rows.length) return "";
    cfg = cfg || {{}};
    var title = cfg.title || "Abwesenheit";
    var border = cfg.border || "#cbd5e1";
    var bg = cfg.bg || "#f8fafc";
    var text = cfg.text || "#334155";
    var badgeBg = cfg.badgeBg || bg;
    var badgeText = cfg.badgeText || text;
    var html = "<details style='background:#fff;border:1px solid " + border + ";border-radius:5px;margin-bottom:10px;overflow:hidden;'>";
    html += "<summary style='cursor:pointer;padding:7px 10px;display:flex;align-items:center;gap:8px;font-size:10px;font-weight:900;text-transform:uppercase;letter-spacing:.45px;color:" + text + ";'>"
          + "<span>" + title + "</span>"
          + "<span style='display:inline-flex;align-items:center;justify-content:center;min-width:22px;height:18px;border-radius:4px;background:" + badgeBg + ";border:1px solid " + border + ";color:" + badgeText + ";font-size:10px;font-weight:900;'>" + rows.length + "</span>"
          + "<span style='margin-left:auto;color:#64748b;font-size:10px;font-weight:700;text-transform:none;letter-spacing:0;'>anklicken zum Anzeigen</span>"
          + "</summary>";
    html += "<div style='border-top:1px solid " + border + ";padding:7px 10px;display:grid;grid-template-columns:repeat(auto-fill,132px);gap:4px;max-height:88px;overflow:auto;'>";
    rows.forEach(function(e) {{
      html += "<span style='width:132px;min-height:22px;display:inline-flex;align-items:center;justify-content:center;background:" + bg + ";border:1px solid " + border + ";border-radius:4px;padding:2px 5px;font-size:10.5px;font-weight:800;color:" + text + ";font-variant-numeric:tabular-nums;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>" + faEsc(e.datum || e.dateKey) + "</span>";
    }});
    html += "</div></details>";
    return html;
  }}

  function _renderSickBox(sickRows) {{
    return _renderAbsenceBox(sickRows, {{ title:"Kranktage", border:"#fecaca", bg:"#fee2e2", text:"#991b1b", badgeBg:"#fee2e2", badgeText:"#991b1b" }});
  }}

  function _renderVacationBox(vacationRows) {{
    return _renderAbsenceBox(vacationRows, {{ title:"Urlaubstage", border:"#bae6fd", bg:"#f1f5f9", text:"#075985", badgeBg:"#f1f5f9", badgeText:"#075985" }});
  }}

  function _renderLkwBox(lkwEntries) {{
    if (!lkwEntries || !lkwEntries.length) return "";
    var total = lkwEntries.reduce(function(sum, e) {{ return sum + (parseInt(e[1], 10) || 0); }}, 0);
    var html = "<details style='background:#fff;border:1px solid #dbe4ef;border-radius:5px;margin-bottom:10px;overflow:hidden;'>";
    html += "<summary style='cursor:pointer;padding:7px 10px;display:flex;align-items:center;gap:8px;font-size:10px;font-weight:900;text-transform:uppercase;letter-spacing:.45px;color:#1e3a5f;'>"
          + "<span>LKW in Auswahl</span>"
          + "<span style='display:inline-flex;align-items:center;justify-content:center;min-width:22px;height:18px;border-radius:4px;background:#f1f5f9;border:1px solid #dbe4ef;color:#1e3a5f;font-size:10px;font-weight:900;'>" + lkwEntries.length + "</span>"
          + "<span style='color:#94a3b8;font-size:10px;font-weight:700;text-transform:none;letter-spacing:0;'>" + total + " Einsätze</span>"
          + "<span style='margin-left:auto;color:#64748b;font-size:10px;font-weight:700;text-transform:none;letter-spacing:0;'>anklicken zum Anzeigen</span>"
          + "</summary>";
    html += "<div style='border-top:1px solid #dbe4ef;padding:7px 10px;display:grid;grid-template-columns:repeat(auto-fill,minmax(112px,1fr));gap:4px;max-height:92px;overflow:auto;'>";
    lkwEntries.forEach(function(e) {{
      html += "<span style='min-height:22px;display:inline-flex;align-items:center;justify-content:space-between;gap:6px;background:#f8fbff;border:1px solid #dbe4ef;border-radius:4px;padding:2px 7px;font-size:10.5px;font-weight:800;color:#1e3a5f;font-variant-numeric:tabular-nums;white-space:nowrap;overflow:hidden;'>"
            + "<span style='overflow:hidden;text-overflow:ellipsis;'>" + faEsc(e[0]) + "</span>"
            + "<span style='font-size:10px;color:#94a3b8;font-weight:700;flex:0 0 auto;'>" + e[1] + "x</span></span>";
    }});
    html += "</div></details>";
    return html;
  }}

  // Freitag ab 18:00 zählt als Samstags-Einsatz
  function _isFrAbend(s) {{
    return s.wochentag === "Fr" && s.beginn && s.beginn >= "18:00";
  }}

  function _summarizeShifts(shifts) {{
    var out = {{ count: shifts.length, dauer: 0, netto: 0, over10: 0, samstage: 0, sonntage: 0, lkwSet: {{}} }};
    shifts.forEach(function(s) {{
      var invalid = _faIsShiftInvalid(s);
      if (!invalid) {{
        var netto = _toMin(s.profil);
        out.dauer += _toMin(s.schichtdauer);
        out.netto += netto;
        if (netto > 600) out.over10 += 1;
      }}
      if (s.wochentag === "Sa" || _isFrAbend(s)) out.samstage += 1;
      if (s.wochentag === "So") out.sonntage += 1;
      (s.lkw || "").split(",").forEach(function(l) {{
        l = l.trim();
        if (l) out.lkwSet[l] = (out.lkwSet[l] || 0) + 1;
      }});
    }});
    return out;
  }}

  function _printShiftStyles() {{
    return "<style>"
      + "@page{{size:A4 landscape;margin:9mm;}}"
      + "*{{box-sizing:border-box;}}body{{font-family:Arial,Helvetica,sans-serif;margin:0;color:#111827;font-size:11px;}}"
      + "h1{{font-size:20px;margin:0 0 2px 0;}}.sub{{font-size:11px;color:#475569;margin-bottom:10px;}}"
      + ".cards{{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin:10px 0;}}"
      + ".card{{border:1px solid #cbd5e1;border-radius:4px;padding:7px 8px;break-inside:avoid;}}"
      + ".label{{font-size:9px;color:#64748b;text-transform:uppercase;font-weight:700;letter-spacing:.4px;}}"
      + ".value{{font-size:17px;font-weight:800;font-variant-numeric:tabular-nums;line-height:1.15;margin-top:2px;}}"
      + "table{{width:100%;border-collapse:collapse;font-size:10.5px;}}th,td{{border:1px solid #cbd5e1;padding:4px 5px;text-align:left;vertical-align:top;}}"
      + "th{{background:#f1f5f9;font-size:9px;text-transform:uppercase;letter-spacing:.3px;}}.num{{text-align:right;font-variant-numeric:tabular-nums;}}.over{{font-weight:800;color:#991b1b;}}"
      + ".month-block{{break-inside:avoid;margin-top:10px;border:1px solid #cbd5e1;border-radius:5px;overflow:hidden;}}"
      + ".month-head{{display:flex;align-items:center;gap:8px;flex-wrap:wrap;background:#f8fbff;border-bottom:1px solid #cbd5e1;padding:6px 8px;}}"
      + ".month-title{{font-size:12px;font-weight:900;text-transform:uppercase;letter-spacing:.45px;color:#1e3a5f;min-width:110px;}}"
      + ".month-meta{{font-size:9.5px;color:#475569;font-weight:700;display:flex;gap:8px;flex-wrap:wrap;}}"
      + ".mini{{border-bottom:1px solid #e2e8f0;padding:5px 8px;font-size:9.5px;line-height:1.35;}}"
      + ".chip{{display:inline-block;border:1px solid #cbd5e1;border-radius:3px;padding:1px 4px;margin:1px 2px;font-weight:700;white-space:nowrap;}}"
      + ".empty{{padding:7px 8px;color:#94a3b8;font-weight:700;text-align:center;}}"
      + "</style>";
  }}

  function _renderShiftRows(rows) {{
    var html = "";
    rows.forEach(function(s, i) {{
      var netto = _toMin(s.profil);
      var invalid = _faIsShiftInvalid(s);
      var over = !invalid && netto > 600;
      var weekend = s.wochentag === "Sa" || s.wochentag === "So" || _isFrAbend(s);
      var rowBg = invalid ? "#f9fafb" : (over ? "#fff1f2" : (weekend ? "#fff7ed" : (i % 2 === 0 ? "#fff" : "#fafbfc")));
      var tagColor = invalid ? "#b0b8c4" : (weekend ? (s.wochentag === "So" ? "#dc2626" : "#b45309") : "#0f172a");
      var cellDim = invalid ? "color:#b0b8c4;text-decoration:line-through;" : "";
      var endeStr = faEsc(s.ende || "");
      if (endeStr && s.ende_naechster_tag) endeStr += " <span style='color:#94a3b8;font-size:9.5px;font-weight:600;'>+1</span>";
      var tagDisplay = (s.wochentag ? "<span style='display:inline-block;width:22px;color:" + (invalid ? "#cbd5e1" : "#94a3b8") + ";font-weight:700;font-size:10.5px;'>" + faEsc(s.wochentag) + "</span> " : "") + faEsc(s.tag);
      var cutoffLabel = "";
      if (invalid) {{
        var reason = _faIsShiftCutoff(s) ? "noch laufend" : "Tachofehler";
        cutoffLabel = " <span style='font-size:8.5px;font-weight:800;background:#fef3c7;color:#92400e;border:1px solid #fde68a;border-radius:3px;padding:0 4px;vertical-align:middle;text-decoration:none;display:inline-block;'>" + reason + "</span>";
      }}

      html += "<tr style='background:" + rowBg + ";border-bottom:1px solid #f1f5f9;" + (invalid ? "opacity:.7;" : "") + "'>";
      html += "<td style='padding:5px 8px;color:" + tagColor + ";font-weight:" + (weekend ? "700" : "600") + ";font-variant-numeric:tabular-nums;white-space:nowrap;'>" + tagDisplay + cutoffLabel + "</td>";
      html += "<td style='padding:5px 6px;font-variant-numeric:tabular-nums;white-space:nowrap;" + (invalid ? cellDim : "color:#475569;") + "'>" + faEsc(s.beginn || "") + "</td>";
      html += "<td style='padding:5px 6px;font-variant-numeric:tabular-nums;white-space:nowrap;" + (invalid ? cellDim : "color:#475569;") + "'>" + endeStr + "</td>";
      html += "<td style='padding:5px 7px;text-align:right;font-weight:800;font-variant-numeric:tabular-nums;white-space:nowrap;" + (invalid ? cellDim : "color:#1e3a5f;") + "'>" + faEsc(s.schichtdauer || "") + "</td>";
      html += "<td style='padding:5px 7px;text-align:right;font-weight:" + (over ? "900" : "700") + ";font-variant-numeric:tabular-nums;white-space:nowrap;" + (invalid ? cellDim : ("color:" + (over ? "#be123c" : "#0f172a") + ";")) + "'>" + faEsc(s.profil || "") + "</td>";
      html += "<td style='padding:5px 8px;font-weight:600;font-size:11px;white-space:normal;" + (invalid ? cellDim : "color:#166534;") + "'>" + faEsc(s.lkw || "") + "</td>";
      html += "</tr>";
    }});
    return html;
  }}

  function _renderShiftTable(rows) {{
    var thBase = "padding:6px 8px;font-weight:800;font-size:9.5px;text-transform:uppercase;letter-spacing:.35px;border-bottom:1px solid #e2e8f0;white-space:nowrap;";
    var html = "<table style='width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed;'>";
    html += "<colgroup>"
          + "<col style='width:128px;'>"
          + "<col style='width:68px;'>"
          + "<col style='width:68px;'>"
          + "<col style='width:94px;'>"
          + "<col style='width:108px;'>"
          + "<col>"
          + "</colgroup>";
    html += "<thead><tr style='background:#fafbfc;color:#64748b;'>"
          + "<th style='" + thBase + "text-align:left;'>Tag</th>"
          + "<th style='" + thBase + "text-align:left;'>Beginn</th>"
          + "<th style='" + thBase + "text-align:left;'>Ende</th>"
          + "<th style='" + thBase + "text-align:right;'>Schichtdauer</th>"
          + "<th style='" + thBase + "text-align:right;'>Netto</th>"
          + "<th style='" + thBase + "text-align:left;'>LKW</th>"
          + "</tr></thead><tbody>";
    html += _renderShiftRows(rows);
    html += "</tbody></table>";
    return html;
  }}

  function faRenderShifts(name, panel) {{
    var all = (TIMEREC_DATA[name] || []).slice();
    var yr = faYearFilter;
    var yearShifts = (yr === "all") ? all : all.filter(function(s) {{
      var m = (s.tag || "").match(/(\\d{{4}})$/);
      return m && m[1] === yr;
    }});

    var byMonth = {{}};
    yearShifts.forEach(function(s) {{
      var mi = _monthInfo(s);
      if (!byMonth[mi.key]) byMonth[mi.key] = {{ label: mi.label, shifts: [] }};
      byMonth[mi.key].shifts.push(s);
    }});
    _faAddAbsenceMonths(name, yr, byMonth);
    var monthKeys = Object.keys(byMonth).sort(function(a,b) {{ return String(b).localeCompare(String(a)); }});

    // Standardansicht beim Öffnen: Jahresübersicht / alle Monate.
    // Nur wenn der Nutzer bewusst einen Monat gewählt hat und dieser Monat nicht mehr existiert,
    // wird wieder auf die Jahresübersicht zurückgestellt.
    if (!faShiftMonthFilterByDriver[name]) {{
      faShiftMonthFilterByDriver[name] = "all";
    }} else if (faShiftMonthFilterByDriver[name] !== "all" && !byMonth[faShiftMonthFilterByDriver[name]]) {{
      faShiftMonthFilterByDriver[name] = "all";
    }}
    var monthFilter = faShiftMonthFilterByDriver[name] || "all";
    var shifts = monthFilter === "all" ? yearShifts : ((byMonth[monthFilter] && byMonth[monthFilter].shifts) ? byMonth[monthFilter].shifts : []);
    var selectedLabel = monthFilter === "all" ? "Alle Monate" : (byMonth[monthFilter] ? byMonth[monthFilter].label : "Monat");
    var stats = _summarizeShifts(shifts);
    var sickRows = _faSickEntries(name, yr, monthFilter);
    var vacationRows = _faVacationEntries(name, yr, monthFilter);
    var sickCount = sickRows.length;
    var vacationCount = vacationRows.length;
    var vacationCredit = _vacationCreditMin(vacationRows);
    var nettoWithVacation = stats.netto + vacationCredit;

    var html = "";

    html += "<div style='background:#fff;border:1.5px solid #e2e8f0;border-radius:5px;padding:14px 18px;margin-bottom:12px;'>"
          + "<div style='display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;'>"
          + "<div>"
          + "<div style='font-size:20px;font-weight:900;color:#0b1220;'>" + faEsc(name) + "</div>"
          + "<div style='font-size:11.5px;color:#64748b;font-weight:700;margin-top:3px;'>Schichten / Tachograph · Netto-Arbeitszeit aus Arbeitszeitprofil · Urlaub +8:00 je Tag</div>"
          + "</div>"
          + "<div style='display:flex;align-items:center;gap:8px;flex-wrap:wrap;'>"
          + "<label style='font-size:10.5px;color:#64748b;font-weight:800;text-transform:uppercase;letter-spacing:.5px;'>Monat</label>"
          + "<select onchange=\\\"faSetShiftMonth(this.value)\\\" style='padding:7px 10px;border:1px solid #cbd5e1;border-radius:5px;background:#fff;font-size:12px;font-weight:700;color:#0f172a;'>";
    html += "<option value='all'" + (monthFilter === "all" ? " selected" : "") + ">Alle Monate</option>";
    monthKeys.forEach(function(mk) {{
      html += "<option value='" + faEsc(mk) + "'" + (monthFilter === mk ? " selected" : "") + ">" + faEsc(byMonth[mk].label) + "</option>";
    }});
    html += "</select>"
          + "<button type='button' onclick='faPrintShiftMonth()' title='Aktuelle Monatsauswahl drucken' style='padding:8px 12px;border:1px solid #dc2626;border-radius:5px;background:#dc2626;color:#fff;font-size:12px;font-weight:900;cursor:pointer;font-family:inherit;white-space:nowrap;'>PDF Druck</button>"
          + "</div></div></div>";

    if (!yearShifts.length && !sickCount && !vacationCount) {{
      html += "<div style='color:#94a3b8;padding:40px;text-align:center;font-size:14px;background:#fff;border:1px solid #e2e8f0;border-radius:5px;'>"
            + "Keine Schichten, Kranktage oder Urlaubstage in diesem Zeitraum.</div>";
      panel.innerHTML = html;
      panel.scrollTop = 0;
      return;
    }}

    html += "<div style='background:#fff;border:1px solid #e2e8f0;border-radius:5px;padding:9px 12px;margin-bottom:10px;display:flex;gap:16px;flex-wrap:wrap;'>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Auswahl</div>"
          + "<div style='font-size:17px;font-weight:900;color:#0f172a;line-height:1.25;'>" + faEsc(selectedLabel) + "</div></div>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Schichten</div>"
          + "<div style='font-size:17px;font-weight:800;color:#0f172a;font-variant-numeric:tabular-nums;line-height:1.15;'>" + stats.count + "</div></div>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Kranktage</div>"
          + "<div style='font-size:17px;font-weight:900;color:" + (sickCount ? "#be123c" : "#166534") + ";font-variant-numeric:tabular-nums;line-height:1.15;'>" + sickCount + "</div></div>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Urlaubstage</div>"
          + "<div style='font-size:17px;font-weight:900;color:" + (vacationCount ? "#075985" : "#166534") + ";font-variant-numeric:tabular-nums;line-height:1.15;'>" + vacationCount + "</div></div>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Σ Netto Schichten</div>"
          + "<div style='font-size:17px;font-weight:900;color:#1e3a5f;font-variant-numeric:tabular-nums;line-height:1.15;'>" + _fmtMin(stats.netto) + "</div></div>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Urlaub +8h</div>"
          + "<div style='font-size:17px;font-weight:900;color:" + (vacationCredit ? "#075985" : "#166534") + ";font-variant-numeric:tabular-nums;line-height:1.15;'>" + _fmtMin(vacationCredit) + "</div></div>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Σ Netto inkl. Urlaub</div>"
          + "<div style='font-size:17px;font-weight:900;color:#0f172a;font-variant-numeric:tabular-nums;line-height:1.15;'>" + _fmtMin(nettoWithVacation) + "</div></div>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Schichten &gt; 10:00 Netto</div>"
          + "<div style='font-size:17px;font-weight:900;color:" + (stats.over10 ? "#be123c" : "#166534") + ";font-variant-numeric:tabular-nums;line-height:1.15;'>" + stats.over10 + "</div></div>";
    html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Σ Schichtdauer</div>"
          + "<div style='font-size:17px;font-weight:800;color:#475569;font-variant-numeric:tabular-nums;line-height:1.15;'>" + _fmtMin(stats.dauer) + "</div></div>";
    if (stats.samstage) html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Samstage</div>"
          + "<div style='font-size:17px;font-weight:800;color:#b45309;font-variant-numeric:tabular-nums;line-height:1.15;'>" + stats.samstage + "</div></div>";
    if (stats.sonntage) html += "<div><div style='font-size:9.5px;color:#64748b;text-transform:uppercase;letter-spacing:.45px;font-weight:700;'>Sonntage</div>"
          + "<div style='font-size:17px;font-weight:800;color:#dc2626;font-variant-numeric:tabular-nums;line-height:1.15;'>" + stats.sonntage + "</div></div>";
    html += "</div>";
    html += _renderSickBox(sickRows);
    html += _renderVacationBox(vacationRows);

    var lkwEntries = Object.keys(stats.lkwSet).map(function(k){{return [k, stats.lkwSet[k]];}}).sort(function(a,b){{return b[1]-a[1];}});
    html += _renderLkwBox(lkwEntries);

    if (monthFilter !== "all") {{
      html += "<div style='background:#fff;border:1px solid #e2e8f0;border-radius:6px;overflow:hidden;margin-bottom:12px;'>";
      html += "<div style='padding:9px 14px;background:#f8fbff;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;gap:14px;flex-wrap:wrap;'>";
      html += "<span style='font-size:12px;font-weight:800;color:#1e3a5f;text-transform:uppercase;letter-spacing:.6px;'>" + faEsc(selectedLabel) + "</span>";
      html += "<span style='font-size:11px;color:#94a3b8;font-weight:600;'>" + stats.count + " Schichten" + (sickCount ? " · Krank " + sickCount : "") + (vacationCount ? " · Urlaub " + vacationCount : "") + "</span>";
      html += "<span style='margin-left:auto;font-size:11px;color:#64748b;font-weight:600;'>Σ Netto <b style='color:#1e3a5f;font-variant-numeric:tabular-nums;'>" + _fmtMin(stats.netto) + "</b> &nbsp;·&nbsp; Urlaub +8h <b style='color:#075985;font-variant-numeric:tabular-nums;'>" + _fmtMin(vacationCredit) + "</b> &nbsp;·&nbsp; Σ inkl. Urlaub <b style='color:#0f172a;font-variant-numeric:tabular-nums;'>" + _fmtMin(nettoWithVacation) + "</b> &nbsp;·&nbsp; &gt; 10:00 Netto <b style='color:#be123c;font-variant-numeric:tabular-nums;'>" + stats.over10 + "</b></span>";
      html += "</div>" + _renderShiftTable(shifts) + "</div>";
    }} else {{
      monthKeys.forEach(function(mk) {{
        var grp = byMonth[mk];
        var gs = _summarizeShifts(grp.shifts);
        var gsSick = _faSickEntries(name, yr, mk).length;
        var gsVacation = _faVacationEntries(name, yr, mk).length;
        var gsVacationCredit = gsVacation * 480;
        var gsNettoWithVacation = gs.netto + gsVacationCredit;
        html += "<div style='background:#fff;border:1px solid #e2e8f0;border-radius:6px;overflow:hidden;margin-bottom:12px;'>";
        html += "<div style='padding:9px 14px;background:#f8fbff;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;gap:14px;flex-wrap:wrap;'>";
        html += "<span style='font-size:12px;font-weight:800;color:#1e3a5f;text-transform:uppercase;letter-spacing:.6px;'>" + faEsc(grp.label) + "</span>";
        html += "<span style='font-size:11px;color:#94a3b8;font-weight:600;'>" + gs.count + " Schichten" + (gsSick ? " · Krank " + gsSick : "") + (gsVacation ? " · Urlaub " + gsVacation : "") + "</span>";
        html += "<span style='margin-left:auto;font-size:11px;color:#64748b;font-weight:600;'>Σ Netto <b style='color:#1e3a5f;font-variant-numeric:tabular-nums;'>" + _fmtMin(gs.netto) + "</b> &nbsp;·&nbsp; Urlaub +8h <b style='color:#075985;font-variant-numeric:tabular-nums;'>" + _fmtMin(gsVacationCredit) + "</b> &nbsp;·&nbsp; Σ inkl. Urlaub <b style='color:#0f172a;font-variant-numeric:tabular-nums;'>" + _fmtMin(gsNettoWithVacation) + "</b> &nbsp;·&nbsp; &gt; 10:00 Netto <b style='color:#be123c;font-variant-numeric:tabular-nums;'>" + gs.over10 + "</b></span>";
        html += "</div>" + _renderShiftTable(grp.shifts) + "</div>";
      }});
    }}

    panel.innerHTML = html;
    panel.scrollTop = 0;
  }}

  window.faSetShiftMonth = function(value) {{
    if (!faSelectedName) return;
    faShiftMonthFilterByDriver[faSelectedName] = value || "all";
    window.faShowDetail(faSelectedName);
  }};

  window.faPrintShiftMonth = function() {{
    var name = faSelectedName;
    if (!name) return;

    var all = (TIMEREC_DATA[name] || []).slice();
    var yr = faYearFilter;
    var yearShifts = (yr === "all") ? all : all.filter(function(s) {{
      var m = (s.tag || "").match(/(\\d{{4}})$/);
      return m && m[1] === yr;
    }});

    var byMonth = {{}};
    yearShifts.forEach(function(s) {{
      var mi = _monthInfo(s);
      if (!byMonth[mi.key]) byMonth[mi.key] = {{ label: mi.label, shifts: [] }};
      byMonth[mi.key].shifts.push(s);
    }});
    _faAddAbsenceMonths(name, yr, byMonth);

    var monthKeys = Object.keys(byMonth).sort(function(a,b) {{ return String(b).localeCompare(String(a)); }});
    var monthFilter = faShiftMonthFilterByDriver[name] || "all";
    if (monthFilter !== "all" && !byMonth[monthFilter]) monthFilter = "all";

    var shifts = monthFilter === "all" ? yearShifts : ((byMonth[monthFilter] && byMonth[monthFilter].shifts) ? byMonth[monthFilter].shifts : []);
    var label = monthFilter === "all" ? "Alle Monate" : (byMonth[monthFilter] ? byMonth[monthFilter].label : "Monat");
    var stats = _summarizeShifts(shifts);
    var sickRows = _faSickEntries(name, yr, monthFilter);
    var vacationRows = _faVacationEntries(name, yr, monthFilter);
    var vacationCredit = _vacationCreditMin(vacationRows);
    var nettoWithVacation = stats.netto + vacationCredit;
    var lkwEntries = Object.keys(stats.lkwSet || {{}}).map(function(k) {{ return [k, stats.lkwSet[k]]; }}).sort(function(a,b) {{ return b[1]-a[1]; }});
    var printedAt = new Date().toLocaleString("de-DE");

    function _printRows(rowList) {{
      var out = "";
      rowList.forEach(function(s) {{
        var netto = _toMin(s.profil);
        var invalid = _faIsShiftInvalid(s);
        var cls = invalid ? "invalid" : (netto > 600 ? "over" : "");
        out += "<tr" + (invalid ? " style='opacity:.5;'" : "") + ">"
           + "<td>" + faEsc((s.wochentag || "") + " " + (s.tag || "")) + (invalid ? " <span style='font-size:5pt;color:#92400e;font-weight:800;'>" + (_faIsShiftCutoff(s) ? "laufend" : "Fehler") + "</span>" : "") + "</td>"
           + "<td>" + faEsc(s.beginn || "") + "</td>"
           + "<td>" + faEsc(s.ende || "") + (s.ende_naechster_tag ? " +1" : "") + "</td>"
           + "<td class='num'" + (invalid ? " style='text-decoration:line-through;'" : "") + ">" + faEsc(s.schichtdauer || "") + "</td>"
           + "<td class='num " + cls + "'" + (invalid ? " style='text-decoration:line-through;'" : "") + ">" + faEsc(s.profil || "") + "</td>"
           + "<td>" + faEsc(s.lkw || "") + "</td>"
           + "</tr>";
      }});
      if (!out) out = "<tr><td colspan='6' class='empty'>Keine Schichten in diesem Monat</td></tr>";
      return out;
    }}

    function _printAbsenceLine(title, rows, styleAttr, suffixFn) {{
      if (!rows || !rows.length) return "";
      suffixFn = suffixFn || function() {{ return ""; }};
      return "<div class='mini' style='" + styleAttr + "'><b>" + title + ":</b> "
        + rows.map(function(e) {{ return faEsc(e.datum || e.dateKey) + suffixFn(e); }}).join(", ")
        + "</div>";
    }}

    function _printLkwLine(entries) {{
      if (!entries || !entries.length) return "";
      return "<div class='mini'><b>LKW:</b> "
        + entries.map(function(e) {{
            return "<span class='chip'>" + faEsc(e[0]) + " <span style='color:#64748b;'>" + e[1] + "x</span></span>";
          }}).join(" ")
        + "</div>";
    }}

    function _printMonthSection(mk, monthLabel, rowList) {{
      rowList = rowList || [];
      var ms = _summarizeShifts(rowList);
      var mSick = _faSickEntries(name, yr, mk);
      var mVacation = _faVacationEntries(name, yr, mk);
      var mVacationCredit = _vacationCreditMin(mVacation);
      var mNettoWithVacation = ms.netto + mVacationCredit;
      var mLkwEntries = Object.keys(ms.lkwSet || {{}}).map(function(k) {{ return [k, ms.lkwSet[k]]; }}).sort(function(a,b) {{ return b[1]-a[1]; }});
      var sec = "<section class='month-block'>";
      sec += "<div class='month-head'><div class='month-title'>" + faEsc(monthLabel) + "</div>";
      sec += "<div class='month-meta'>"
          + "<span>Schichten <b>" + ms.count + "</b></span>"
          + "<span>Krank <b>" + mSick.length + "</b></span>"
          + "<span>Urlaub <b>" + mVacation.length + "</b></span>"
          + "<span>Netto <b>" + _fmtMin(ms.netto) + "</b></span>"
          + "<span>Urlaub +8h <b>" + _fmtMin(mVacationCredit) + "</b></span>"
          + "<span>Σ inkl. Urlaub <b>" + _fmtMin(mNettoWithVacation) + "</b></span>"
          + "<span>&gt;10:00 <b style='color:#991b1b;'>" + ms.over10 + "</b></span>"
          + "</div></div>";
      sec += _printAbsenceLine("Kranktage", mSick, "border-color:#fecaca;background:#fff7f7;color:#991b1b;");
      sec += _printAbsenceLine("Urlaubstage (+8:00 je Tag / Σ " + _fmtMin(mVacationCredit) + ")", mVacation, "border-color:#bae6fd;background:#f0f9ff;color:#075985;", function() {{ return " (+8:00)"; }});
      sec += _printLkwLine(mLkwEntries);
      sec += "<table><thead><tr><th>Tag</th><th>Beginn</th><th>Ende</th><th class='num'>Schichtdauer</th><th class='num'>Netto-Arbeitszeit</th><th>LKW</th></tr></thead><tbody>" + _printRows(rowList) + "</tbody></table>";
      sec += "</section>";
      return sec;
    }}

    var doc = "<!doctype html><html><head><meta charset='utf-8'><title>Schichten " + faEsc(name) + " " + faEsc(label) + "</title>" + _printShiftStyles() + "</head><body>";
    doc += "<h1>Schichten / Tachograph</h1>";
    doc += "<div class='sub'><b>Fahrer:</b> " + faEsc(name) + " &nbsp; · &nbsp; <b>Auswahl:</b> " + faEsc(label) + " &nbsp; · &nbsp; <b>Ausdruck:</b> " + faEsc(printedAt) + "</div>";
    doc += "<div class='cards'>"
        + "<div class='card'><div class='label'>Schichten</div><div class='value'>" + stats.count + "</div></div>"
        + "<div class='card'><div class='label'>Kranktage</div><div class='value'>" + sickRows.length + "</div></div>"
        + "<div class='card'><div class='label'>Urlaubstage</div><div class='value'>" + vacationRows.length + "</div></div>"
        + "<div class='card'><div class='label'>Σ Netto Schichten</div><div class='value'>" + _fmtMin(stats.netto) + "</div></div>"
        + "<div class='card'><div class='label'>Urlaub +8h</div><div class='value'>" + _fmtMin(vacationCredit) + "</div></div>"
        + "<div class='card'><div class='label'>Σ Netto inkl. Urlaub</div><div class='value'>" + _fmtMin(nettoWithVacation) + "</div></div>"
        + "<div class='card'><div class='label'>Schichten &gt; 10:00 Netto</div><div class='value'>" + stats.over10 + "</div></div>"
        + "<div class='card'><div class='label'>Σ Schichtdauer</div><div class='value'>" + _fmtMin(stats.dauer) + "</div></div>"
        + "</div>";

    if (monthFilter === "all") {{
      monthKeys.forEach(function(mk) {{
        var grp = byMonth[mk] || {{ label: _monthLabelFromKey(mk), shifts: [] }};
        doc += _printMonthSection(mk, grp.label, grp.shifts || []);
      }});
    }} else {{
      var grp = byMonth[monthFilter] || {{ label: label, shifts: shifts }};
      doc += _printMonthSection(monthFilter, label, grp.shifts || shifts);
    }}

    doc += "<script>window.onload=function(){{setTimeout(function(){{window.print();}},150);}}<\\/script>";
    doc += "</body></html>";

    var w = window.open("", "_blank");
    if (!w) {{
      alert("Der PDF-Druck konnte nicht geöffnet werden. Bitte Pop-up-Blocker prüfen.");
      return;
    }}
    w.document.open();
    w.document.write(doc);
    w.document.close();
  }};

  window.faSwitchTab = function(tab) {{
    faActiveTab = tab;
    if (faSelectedName) window.faShowDetail(faSelectedName);
  }};

  // Override
  window.faShowDetail = function(name) {{
    faSelectedName = name;
    if (window.FA_10H_MODE) {{
      window.FA_10H_SELECTED_DRIVER = name || "";
      faBuildSidebarHighlight(name);
      window.faShow10hTours(name);
      return;
    }}
    window.FA_10H_MODE = false;
    _faSet10hButton(false);
    faBuildSidebarHighlight(name);
    var panel = document.getElementById("fa-detail-panel");
    if (!panel) return;

    var hasShifts = faHasShifts(name);
    var driver = FA_DATA.find(function(d){{ return d.name === name; }});
    var hasUebersicht = !!driver;

    // Wenn nur Schichten-Daten vorhanden, automatisch dorthin wechseln
    if (!hasUebersicht && hasShifts && faActiveTab !== "schichten") {{
      faActiveTab = "schichten";
    }}
    if (!hasShifts && faActiveTab === "schichten") {{
      faActiveTab = "uebersicht";
    }}

    if (faActiveTab === "schichten" && hasShifts) {{
      faRenderShifts(name, panel);
    }} else if (hasUebersicht) {{
      _faShowDetailOrig(name);
      // Tab-Bar oben einfuegen, falls beide Tabs verfuegbar
      var tabBar = faTabBar(name, hasShifts, true);
      if (tabBar) panel.insertAdjacentHTML('afterbegin', tabBar);
    }} else {{
      // Kein Driver in FA_DATA und keine Schichten
      panel.innerHTML = "<div style='color:#94a3b8;padding:40px;text-align:center;font-size:14px;'>Keine Daten f&uuml;r diesen Fahrer.</div>";
    }}
  }};

  // faBuildSidebarHighlight ergänzen: Fahrer aus TIMEREC_DATA hinzufügen, die nicht in FA_DATA sind
  var _faGetFilteredOrig = window.faGetFiltered;
  if (typeof _faGetFilteredOrig === "function") {{
    window.faGetFiltered = function() {{
      var list = _faGetFilteredOrig();
      // Schichten-only Fahrer mergen
      if (TIMEREC_DATA && typeof TIMEREC_DATA === "object") {{
        var have = {{}};
        list.forEach(function(d) {{ have[d.name] = true; }});
        var q = (faSearchQuery || "").toLowerCase().trim();
        Object.keys(TIMEREC_DATA).forEach(function(n) {{
          if (have[n]) return;
          if (q && n.toLowerCase().indexOf(q) < 0) return;
          // Year-Filter prüfen
          if (faYearFilter !== "all") {{
            var hasYr = (TIMEREC_DATA[n] || []).some(function(s) {{
              var m = (s.tag || "").match(/(\\d{{4}})$/);
              return m && m[1] === faYearFilter;
            }});
            if (!hasYr) return;
          }}
          list.push({{ name: n, years: {{}}, _shiftsOnly: true }});
        }});
        if (faCurrentSort === "name") {{
          list.sort(function(a,b){{ return a.name.localeCompare(b.name, "de"); }});
        }}
      }}
      return list;
    }};
  }}

  // faPopulateYears erweitern: Jahre aus TIMEREC_DATA hinzufuegen
  var _faPopulateYearsOrig = window.faPopulateYears;
  if (typeof _faPopulateYearsOrig === "function") {{
    window.faPopulateYears = function() {{
      _faPopulateYearsOrig();
      var yrSel = document.getElementById("fa-year-sel");
      if (!yrSel) return;
      var existing = {{}};
      Array.prototype.slice.call(yrSel.options).forEach(function(o){{ existing[o.value] = true; }});
      var newYears = [];
      Object.keys(TIMEREC_DATA || {{}}).forEach(function(n) {{
        (TIMEREC_DATA[n] || []).forEach(function(s) {{
          var m = (s.tag || "").match(/(\\d{{4}})$/);
          if (m && !existing[m[1]] && m[1] !== "2024") {{
            existing[m[1]] = true;
            newYears.push(m[1]);
          }}
        }});
      }});
      if (newYears.length) {{
        var current = yrSel.value;
        var allValues = Array.prototype.slice.call(yrSel.options).map(function(o){{return o.value;}}).concat(newYears);
        allValues = allValues.filter(function(v,i,a){{ return a.indexOf(v) === i; }});
        allValues.sort().reverse();
        yrSel.innerHTML = allValues.map(function(y){{
          return "<option value='" + y + "'>" + y + "</option>";
        }}).join("");
        if (current && allValues.indexOf(current) !== -1) yrSel.value = current;
        else yrSel.value = allValues[0];
        faYearFilter = yrSel.value;
      }}
    }};
  }}

  // faRender erweitern: auch bei leerem FA_DATA arbeiten, sofern TIMEREC_DATA da ist
  var _faRenderOrig = window.faRender;
  if (typeof _faRenderOrig === "function") {{
    window.faRender = function(q) {{
      faSearchQuery = q || "";
      var hasFA = FA_DATA && FA_DATA.length;
      var hasTR = TIMEREC_DATA && Object.keys(TIMEREC_DATA).length > 0;
      if (!hasFA && !hasTR) {{
        var c = document.getElementById("fa-detail-panel");
        if (c) c.innerHTML = "<div style='color:#94a3b8;padding:40px;text-align:center;font-size:14px;'>Keine Daten vorhanden.<br>Bitte Fahrerauswertungs- oder Schicht-Dateien hochladen.</div>";
        return;
      }}
      window.faPopulateYears();
      window.faBuildSidebarHighlight(null);
    }};
  }}

  // CSV-only Fahrerauswertung: keine alte Übersicht, kein Switcher.
  window.faSort = function(mode) {{
    faCurrentSort = mode;
    ["name","arbeit"].forEach(function(m) {{
      var btn = document.getElementById("fa-sort-" + m);
      if (!btn) return;
      btn.style.background = mode === m ? "#334155" : "#fff";
      btn.style.color = mode === m ? "#111827" : "#334155";
    }});
    faBuildSidebarHighlight(null);
  }};

  window.faYearChange = function(yr) {{
    faYearFilter = yr || "all";
    if (window.FA_10H_MODE) {{
      window.faShow10hTours(window.FA_10H_SELECTED_DRIVER || "");
    }} else {{
      faBuildSidebarHighlight(null);
    }}
  }};

  function _faAllShiftNames() {{
    return Object.keys(TIMEREC_DATA || {{}}).filter(function(n) {{
      return n && (TIMEREC_DATA[n] || []).length > 0;
    }});
  }}

  function _faShiftRowsForYear(name) {{
    var rows = (TIMEREC_DATA && TIMEREC_DATA[name] ? TIMEREC_DATA[name] : []).slice();
    if (faYearFilter === "all") return rows;
    return rows.filter(function(s) {{
      var m = (s.tag || "").match(/(\\d{{4}})$/);
      return m && m[1] === faYearFilter;
    }});
  }}

  function _faShiftStatsForName(name) {{
    var rows = _faShiftRowsForYear(name);
    return _summarizeShifts(rows);
  }}

  function _faSet10hButton(active) {{
    var btn = document.getElementById("fa-btn-10h");
    if (!btn) return;
    btn.textContent = "10H Touren";
    // Wichtig: CSS aus dem neutralen Schema arbeitet mit !important.
    // Deshalb auch hier setProperty(..., "important"), sonst wird der Text
    // beim aktiven Button farblich und wirkt verschwunden.
    btn.style.setProperty("background", active ? "#dc2626" : "#fff", "important");
    btn.style.setProperty("color", active ? "#fff" : "#dc2626", "important");
    btn.style.setProperty("border-color", "#dc2626", "important");
    var btnDriver = document.getElementById("fa-sort-name");
    if (btnDriver) {{
      btnDriver.textContent = "Fahrer Übersicht";
      btnDriver.style.setProperty("background", "linear-gradient(180deg,#e2e8f0 0%,#cbd5e1 100%)", "important");
      btnDriver.style.setProperty("color", "#334155", "important");
      btnDriver.style.setProperty("border-color", "#94a3b8", "important");
    }}
  }}

  function _faShiftDateKey(s) {{
    var raw = String((s && s.tag) || "");
    var m = raw.match(/^(\\d{{2}})\\.(\\d{{2}})\\.(\\d{{4}})$/);
    if (!m) return "";
    return m[3] + "-" + m[2] + "-" + m[1];
  }}

  function _faDateLabelFromKey(key) {{
    var m = String(key || "").match(/^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})$/);
    if (!m) return key || "";
    return m[3] + "." + m[2] + "." + m[1];
  }}

  // ── 10H "Unfertig"-Erkennung ─────────────────────────────────────────────
  // Wenn die Fahrerkarte nach der Tour noch nicht erneut ausgelesen wurde, setzt
  // der Tachograph-Download als Schichtende den Auslesezeitpunkt ein. Dadurch wird
  // die Schicht künstlich lang. Erkennbar daran, dass viele lange Schichten exakt
  // im selben Zeitfenster "enden" (= Moment der Massen-Auslesung) ODER dass eine
  // Schichtdauer völlig unrealistisch ist (> 16:00). Eine echte 10H-Tour (z.B.
  // Ende 13:27 / Dauer 10:57) wird so NICHT fälschlich markiert.
  var _faCutoffData = null;

  function _faEndDateKey(s) {{
    var m = String(_faComputeEndDate(s)).match(/^(\\d{{2}})\\.(\\d{{2}})\\.(\\d{{4}})$/);
    return m ? (m[3] + "-" + m[2] + "-" + m[1]) : "";
  }}

  function _faEndMinutes(s) {{
    var m = String((s && s.ende) || "").match(/^(\\d{{1,2}}):(\\d{{2}})/);
    if (!m) return -1;
    return parseInt(m[1], 10) * 60 + parseInt(m[2], 10);
  }}

  // Schwellen
  var FA_CUTOFF_LONG_MIN   = 720;  // ab 12:00 Schichtdauer gilt eine Schicht als "lang/auffällig"
  var FA_CUTOFF_ABSURD_MIN = 960;  // ab 16:00 Schichtdauer = einzeln schon unmöglich
  var FA_CUTOFF_WINDOW     = 12;   // Minuten-Fenster für gemeinsames Ende
  var FA_CUTOFF_CLUSTER    = 3;    // so viele lange Schichten im Fenster = Massen-Auslesung

  function _faGetCutoffData() {{
    if (_faCutoffData) return _faCutoffData;
    var byDay = {{}};
    Object.keys(TIMEREC_DATA || {{}}).forEach(function(name) {{
      (TIMEREC_DATA[name] || []).forEach(function(s) {{
        if (!s.ende) return;
        if (_toMin(s.schichtdauer) <= FA_CUTOFF_LONG_MIN) return;
        var dk = _faEndDateKey(s);
        var em = _faEndMinutes(s);
        if (!dk || em < 0) return;
        (byDay[dk] = byDay[dk] || []).push(em);
      }});
    }});
    _faCutoffData = byDay;
    return _faCutoffData;
  }}

  function _faIsShiftReadoutCutoff(s) {{
    if (!s.ende) return false;
    var gross = _toMin(s.schichtdauer);
    // a) einzelne, völlig unrealistische Schichtdauer
    if (gross > FA_CUTOFF_ABSURD_MIN) return true;
    // b) lange Schicht, die mit weiteren im selben Fenster endet (Massen-Auslesung)
    if (gross <= FA_CUTOFF_LONG_MIN) return false;
    var em = _faEndMinutes(s);
    if (em < 0) return false;
    var list = _faGetCutoffData()[_faEndDateKey(s)] || [];
    var n = 0;
    for (var i = 0; i < list.length; i++) {{
      if (Math.abs(list[i] - em) <= FA_CUTOFF_WINDOW) n++;
    }}
    return n >= FA_CUTOFF_CLUSTER;
  }}

  function _faIs10hUnfertig(s) {{
    if (_faIsShiftReadoutCutoff(s)) return true;  // fuzzy: Ende-Fenster / unmögliche Dauer
    if (_faIsShiftCutoff(s)) return true;          // exakt identisches Ende (Altlogik)
    return false;
  }}

  function _faNormLkw(value) {{
    var s = String(value == null ? "" : value).toLowerCase();
    try {{ s = s.normalize("NFKD").replace(/[̀-ͯ]/g, ""); }} catch(e) {{}}
    s = s.replace(/[^a-z0-9]+/g, "");
    return s;
  }}

  function _faLkwMatch(a, b) {{
    var an = _faNormLkw(a), bn = _faNormLkw(b);
    if (!an || !bn) return false;
    return an === bn || an.indexOf(bn) >= 0 || bn.indexOf(an) >= 0;
  }}

  function _faEntryTimeDiffMin(planTime, shiftTime) {{
    var a = _toMin(planTime), b = _toMin(shiftTime);
    if (!a || !b) return 99999;
    return Math.abs(a - b);
  }}

  function _faAddDaysKey(key, days) {{
    var m = String(key || "").match(/^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})$/);
    if (!m) return "";
    var d = new Date(parseInt(m[1],10), parseInt(m[2],10)-1, parseInt(m[3],10));
    d.setDate(d.getDate() + (days || 0));
    var y = d.getFullYear();
    var mo = String(d.getMonth()+1).padStart(2,"0");
    var da = String(d.getDate()).padStart(2,"0");
    return y + "-" + mo + "-" + da;
  }}

  function _faCandidateDateKeys(shift) {{
    var dk = _faShiftDateKey(shift);
    if (!dk) return [];
    // Schichtbeginn prüfen: bei spätem Start (>= 20:00) auch Folgetag suchen,
    // weil der Fahrer kurz vor Mitternacht seine Karte steckt, die Tour aber
    // im Tourenplan unter dem nächsten Tag steht.
    var beginn = _toMin(shift && shift.beginn);
    if (beginn && beginn >= 1200) {{
      // 1200 min = 20:00 Uhr → auch Folgetag als Kandidat
      var nextDk = _faAddDaysKey(dk, 1);
      return nextDk ? [dk, nextDk] : [dk];
    }}
    return [dk];
  }}

  function _faPlanEntriesForShift(name, shift) {{
    var driver = _faPlanDriver(name);
    if (!driver || !driver.years) return [];
    var candidateKeys = _faCandidateDateKeys(shift);
    if (!candidateKeys.length) return [];
    var keyRank = {{}};
    candidateKeys.forEach(function(k, i) {{ keyRank[k] = i; }});
    var years = faYearFilter === "all" ? Object.keys(driver.years || {{}}) : [faYearFilter];
    // Bei Schichten über Jahreswechsel auch angrenzende Jahre durchsuchen.
    candidateKeys.forEach(function(k) {{
      var y = k.slice(0,4);
      if (years.indexOf(y) < 0) years.push(y);
    }});
    var out = [];
    years.forEach(function(y) {{
      var data = driver.years[y];
      if (!data) return;
      (data.eintraege || []).forEach(function(e) {{
        var planKey = _faPlanDateKey(e);
        if (keyRank[planKey] === undefined) return;
        var t = String(e.tour || "");
        if (/krank|urlaub|ausgleich/i.test(t)) return;
        e._matchDateRank = keyRank[planKey];
        e._matchDriverName = driver.name || name;
        e._matchSource = "Fahrer + Anfangstag";
        out.push(e);
      }});
    }});
    return out;
  }}

  function _faAllPlanEntriesForShift(shift) {{
    var candidateKeys = _faCandidateDateKeys(shift);
    if (!candidateKeys.length || !Array.isArray(FA_DATA)) return [];
    var keyRank = {{}};
    candidateKeys.forEach(function(k, i) {{ keyRank[k] = i; }});
    var years = faYearFilter === "all" ? null : {{}};
    if (years) {{
      years[faYearFilter] = true;
      candidateKeys.forEach(function(k) {{ years[k.slice(0,4)] = true; }});
    }}
    var out = [];
    FA_DATA.forEach(function(driver) {{
      if (!driver || !driver.years) return;
      Object.keys(driver.years || {{}}).forEach(function(y) {{
        if (years && !years[y]) return;
        var data = driver.years[y];
        if (!data) return;
        (data.eintraege || []).forEach(function(e) {{
          var planKey = _faPlanDateKey(e);
          if (keyRank[planKey] === undefined) return;
          var t = String(e.tour || "");
          if (/krank|urlaub|ausgleich/i.test(t)) return;
          e._matchDateRank = keyRank[planKey];
          e._matchDriverName = driver.name || "";
          e._matchSource = "Fallback Anfangstag";
          out.push(e);
        }});
      }});
    }});
    return out;
  }}

  function _faScorePlanCandidate(e, shift, strictDriver) {{
    var score = strictDriver ? 500 : 900;
    var dateRank = e._matchDateRank || 0;
    score += dateRank * 120;

    var planLkw = String(e.lkw || "").trim();
    var shiftLkw = String(shift && shift.lkw || "").trim();
    var hasPlanLkw = !!_faNormLkw(planLkw);
    var hasShiftLkw = !!_faNormLkw(shiftLkw);
    var lkwOk = _faLkwMatch(planLkw, shiftLkw);
    if (lkwOk) score -= strictDriver ? 700 : 850;
    else if (hasPlanLkw && hasShiftLkw) score += strictDriver ? 180 : 420;
    else score += 80;

    var diff = _faEntryTimeDiffMin(e.zeit, shift.beginn);
    // Bei Folgetag-Zuordnung (dateRank>0, Schicht >= 20:00): Zeitdiff über
    // Mitternacht berechnen, z.B. Schicht 23:44, Plan 00:00 → 16 min
    if (dateRank > 0) {{
      var shiftMin = _toMin(shift && shift.beginn);
      var planMin  = _toMin(e.zeit);
      if (shiftMin >= 1200 && planMin !== null) {{
        var midnightDiff = (1440 - shiftMin) + planMin;
        if (midnightDiff < diff) diff = midnightDiff;
      }}
    }}
    if (diff <= 30) score -= 360;
    else if (diff <= 90) score -= 260;
    else if (diff <= 180) score -= 120;
    else if (diff <= 360) score -= 40;
    else score += strictDriver ? 40 : 140;

    var tourText = String(e.tour || "").trim();
    if (tourText) score -= 120;
    if (/z\\.?\\s*b\\.?\\s*v|sonder/i.test(tourText)) score -= 50;
    return score;
  }}

  function _faFindPlanEntryForShift(name, shift) {{
    function _candidateIsUsable(e, strictDriver, candidateCount) {{
      var lkwOk = _faLkwMatch(e && e.lkw, shift && shift.lkw);
      var diff = _faEntryTimeDiffMin(e && e.zeit, shift && shift.beginn);
      var hasPlanLkw = !!_faNormLkw(e && e.lkw);
      var hasShiftLkw = !!_faNormLkw(shift && shift.lkw);
      var hasPlanTime = !!String((e && e.zeit) || "").trim();

      // Bei exakt gleichem Fahrer + Anfangstag reicht ein einzelner Eintrag aus.
      // Gibt es mehrere Eintraege am Tag, muss LKW oder Startzeit helfen.
      if (strictDriver && candidateCount === 1) return true;
      if (strictDriver && lkwOk) return true;
      if (strictDriver && hasPlanTime && diff <= 180) return true;

      // Fallback ohne sauberen Fahrernamen nur verwenden, wenn LKW und Zeit stimmig sind.
      // Reines "gleicher Tag" ist zu unsicher und fuehrte zu falschen Zuordnungen.
      if (!strictDriver && lkwOk && diff <= 360) return true;
      if (!strictDriver && lkwOk && !hasPlanTime) return true;
      if (!strictDriver && !hasShiftLkw && hasPlanTime && diff <= 45) return true;
      return false;
    }}

    var candidates = _faPlanEntriesForShift(name, shift);
    var best = null;
    var bestScore = 999999;

    candidates.forEach(function(e) {{
      if (!_candidateIsUsable(e, true, candidates.length)) return;
      var score = _faScorePlanCandidate(e, shift, true);
      if (score < bestScore) {{ bestScore = score; best = e; }}
    }});
    if (best) {{
      var lkwOkStrict = _faLkwMatch(best.lkw, shift.lkw);
      var diffStrict = _faEntryTimeDiffMin(best.zeit, shift.beginn);
      var nextDay = (best._matchDateRank || 0) > 0 ? " (Folgetag)" : "";
      if (lkwOkStrict && diffStrict <= 360) best._matchSource = "Fahrer + Anfangstag + LKW + Startzeit" + nextDay;
      else if (lkwOkStrict) best._matchSource = "Fahrer + Anfangstag + LKW" + nextDay;
      else if (diffStrict <= 180) best._matchSource = "Fahrer + Anfangstag + Startzeit" + nextDay;
      else best._matchSource = "Fahrer + Anfangstag" + nextDay;
      return best;
    }}

    // Fallback nur noch bei eindeutigen harten Treffern.
    // Keine Tour wird besser ausgeblendet als falsch angezeigt.
    var allCandidates = _faAllPlanEntriesForShift(shift).filter(function(e) {{
      return _candidateIsUsable(e, false, 0);
    }});
    allCandidates.forEach(function(e) {{
      var score = _faScorePlanCandidate(e, shift, false);
      if (score < bestScore) {{ bestScore = score; best = e; }}
    }});
    if (best) {{
      var lkwOk = _faLkwMatch(best.lkw, shift.lkw);
      var diff = _faEntryTimeDiffMin(best.zeit, shift.beginn);
      var nextDay = (best._matchDateRank || 0) > 0 ? " (Folgetag)" : "";
      if (lkwOk && diff <= 360) best._matchSource = "Fallback: LKW + Anfangstag + Startzeit" + nextDay;
      else if (lkwOk) best._matchSource = "Fallback: LKW + Anfangstag" + nextDay;
      else best._matchSource = "Fallback: Anfangstag + Startzeit" + nextDay;
    }}
    return best;
  }}

  function _faBuild10hRows(driverFilter) {{
    var q = (faSearchQuery || "").toLowerCase().trim();
    var selectedDriver = String(driverFilter || "").trim();
    var rows = [];
    Object.keys(TIMEREC_DATA || {{}}).forEach(function(name) {{
      if (selectedDriver && name !== selectedDriver) return;
      if (q && name.toLowerCase().indexOf(q) < 0) return;
      (TIMEREC_DATA[name] || []).forEach(function(s) {{
        var yrOk = true;
        if (faYearFilter !== "all") {{
          var m = (s.tag || "").match(/(\\d{{4}})$/);
          yrOk = !!(m && m[1] === faYearFilter);
        }}
        if (!yrOk) return;
        if (_faIsShiftInvalid(s)) return;
        var netto = _toMin(s.profil);
        if (netto <= 600) return;
        var plan = _faFindPlanEntryForShift(name, s);
        var tourText = plan && plan.tour ? String(plan.tour).trim() : "";
        // 10H-Schichten immer anzeigen. Wenn keine Tour gefunden wurde,
        // bleibt die Zeile sichtbar und wird in der Tour-Spalte markiert.
        if (!tourText) tourText = "nicht gefunden";
        var dk = _faShiftDateKey(s);
        rows.push({{
          fahrer: name,
          dateKey: dk,
          datum: _faDateLabelFromKey(dk),
          wochentag: s.wochentag || "",
          beginn: s.beginn || "",
          ende: (s.ende || "") + (s.ende_naechster_tag ? " +1" : ""),
          schichtdauer: s.schichtdauer || "",
          nettoText: s.profil || _fmtMin(netto),
          nettoMin: netto,
          lkw: s.lkw || "",
          tour: tourText,
          planLkw: plan && plan.lkw ? String(plan.lkw).trim() : "",
          planZeit: plan && plan.zeit ? String(plan.zeit).trim() : "",
          planDatum: plan ? _faDateLabelFromKey(_faPlanDateKey(plan)) : "",
          planFahrer: plan && plan._matchDriverName ? String(plan._matchDriverName).trim() : "",
          matchSource: plan && plan._matchSource ? String(plan._matchSource).trim() : "",
          unfertig: _faIs10hUnfertig(s)
        }});
      }});
    }});
    rows.sort(function(a,b) {{
      return (b.dateKey || "").localeCompare(a.dateKey || "") || a.fahrer.localeCompare(b.fahrer, "de") || (a.beginn || "").localeCompare(b.beginn || "");
    }});
    return rows;
  }}

  function _faCountKeys(obj) {{
    return Object.keys(obj || {{}}).length;
  }}

  function _faRender10hFrequency(rows) {{
    if (!rows.length) return "";
    var tourMap = {{}}, fahrerMap = {{}};
    rows.forEach(function(r) {{
      var tour = String(r.tour || "").trim();
      if (!tour) return;
      if (!tourMap[tour]) tourMap[tour] = {{ name: tour, count: 0, netto: 0, fahrer: {{}} }};
      tourMap[tour].count += 1;
      tourMap[tour].netto += r.nettoMin || 0;
      tourMap[tour].fahrer[r.fahrer || ""] = true;

      var fahrer = String(r.fahrer || "").trim() || "unbekannt";
      if (!fahrerMap[fahrer]) fahrerMap[fahrer] = {{ name: fahrer, count: 0, netto: 0, touren: {{}} }};
      fahrerMap[fahrer].count += 1;
      fahrerMap[fahrer].netto += r.nettoMin || 0;
      fahrerMap[fahrer].touren[tour] = true;
    }});

    function sortedList(map) {{
      return Object.keys(map).map(function(k) {{ return map[k]; }}).sort(function(a,b) {{
        return (b.count - a.count) || a.name.localeCompare(b.name, "de");
      }});
    }}

    function renderBox(title, sub, headers, rowsHtml, countLabel) {{
      return "<details style='background:#fff;border:1px solid #e2e8f0;border-radius:6px;overflow:hidden;min-width:0;'>"
        + "<summary style='list-style:none;cursor:pointer;padding:10px 12px;background:#f8fbff;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;justify-content:space-between;gap:10px;'>"
        + "<span><span style='font-size:13px;font-weight:950;color:#0f172a;'>" + title + "</span><span style='display:block;font-size:10.5px;font-weight:700;color:#64748b;margin-top:2px;'>" + sub + "</span></span>"
        + "<span style='font-size:10px;font-weight:950;color:#be123c;background:#fff1f2;border:1px solid #fecdd3;border-radius:999px;padding:3px 8px;white-space:nowrap;'>" + countLabel + "</span>"
        + "</summary>"
        + "<div style='max-height:260px;overflow:auto;'>"
        + "<table style='width:100%;border-collapse:collapse;font-size:11.5px;'>"
        + "<thead><tr style='background:#f1f5f9;'>" + headers.map(function(h) {{ return "<th style='position:sticky;top:0;background:#f1f5f9;padding:7px 9px;border-bottom:1px solid #cbd5e1;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.35px;color:#475569;font-weight:950;'>" + h + "</th>"; }}).join("") + "</tr></thead>"
        + "<tbody>" + rowsHtml + "</tbody></table></div></details>";
    }}

    var tourRows = sortedList(tourMap).map(function(t, i) {{
      return "<tr style='background:" + (i % 2 ? "#fff" : "#f8fafc") + ";border-bottom:1px solid #e2e8f0;'>"
        + "<td style='padding:7px 9px;font-weight:900;color:#0f172a;white-space:normal;line-height:1.25;'>" + faEsc(t.name) + "</td>"
        + "<td style='padding:7px 9px;text-align:right;font-weight:950;color:#be123c;font-variant-numeric:tabular-nums;'>" + t.count + "</td>"
        + "<td style='padding:7px 9px;text-align:right;font-weight:850;color:#0f172a;font-variant-numeric:tabular-nums;'>" + _faCountKeys(t.fahrer) + "</td>"
        + "<td style='padding:7px 9px;text-align:right;font-weight:850;color:#0f172a;font-variant-numeric:tabular-nums;'>" + _fmtMin(t.netto) + "</td>"
        + "</tr>";
    }}).join("");

    var fahrerRows = sortedList(fahrerMap).map(function(f, i) {{
      return "<tr style='background:" + (i % 2 ? "#fff" : "#f8fafc") + ";border-bottom:1px solid #e2e8f0;'>"
        + "<td style='padding:7px 9px;font-weight:900;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>" + faEsc(f.name) + "</td>"
        + "<td style='padding:7px 9px;text-align:right;font-weight:950;color:#be123c;font-variant-numeric:tabular-nums;'>" + f.count + "</td>"
        + "<td style='padding:7px 9px;text-align:right;font-weight:850;color:#0f172a;font-variant-numeric:tabular-nums;'>" + _faCountKeys(f.touren) + "</td>"
        + "<td style='padding:7px 9px;text-align:right;font-weight:850;color:#0f172a;font-variant-numeric:tabular-nums;'>" + _fmtMin(f.netto) + "</td>"
        + "</tr>";
    }}).join("");

    return "<div style='display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px;'>"
      + renderBox("Touren-Häufigkeit", "Welche Tournummer / welcher Tourtext kommt in den 10H-Schichten wie oft vor", ["Tournummer", "Anzahl", "Fahrer", "Netto"], tourRows, sortedList(tourMap).length + " Touren")
      + renderBox("Fahrer-Häufigkeit", "Welche Fahrer haben wie viele 10H-Touren", ["Fahrer", "Anzahl", "Touren", "Netto"], fahrerRows, sortedList(fahrerMap).length + " Fahrer")
      + "</div>";
  }}

  function _faRender10hRows(rows) {{
    if (!rows.length) {{
      return "<div style='background:#fff;border:1px solid #e2e8f0;border-radius:5px;padding:40px;text-align:center;color:#94a3b8;font-size:14px;'>Keine 10H Touren in der aktuellen Auswahl.</div>";
    }}
    var html = "<div style='background:#fff;border:1px solid #e2e8f0;border-radius:6px;overflow:auto;'>";
    html += "<table style='width:100%;min-width:1380px;border-collapse:collapse;font-size:12px;table-layout:fixed;'>";
    html += "<colgroup><col style='width:185px;'><col style='width:104px;'><col style='width:70px;'><col style='width:70px;'><col style='width:82px;'><col style='width:82px;'><col style='width:120px;'><col style='width:340px;'><col style='width:240px;'></colgroup>";
    html += "<thead><tr style='background:#fff1f2;color:#334155;'>";
    ["Fahrer","Datum","Beginn","Ende","Schicht","Netto","LKW","Tournummer","Zuordnung"].forEach(function(h) {{
      html += "<th style='padding:8px 9px;border-bottom:1px solid #fecdd3;text-align:left;font-size:9.5px;text-transform:uppercase;letter-spacing:.35px;font-weight:900;'>" + h + "</th>";
    }});
    html += "</tr></thead><tbody>";
    rows.forEach(function(r, i) {{
      var bg = i % 2 ? "#fff" : "#fff7f8";
      if (r.unfertig) bg = i % 2 ? "#fffbeb" : "#fef3c7";
      var found = r.tour && r.tour !== "nicht gefunden";
      var unfertigBadge = r.unfertig
        ? "<div style='margin-top:3px;'><span title='Karte nach der Tour noch nicht ausgelesen – Zeit unvollständig' style='display:inline-block;padding:1px 7px;border-radius:999px;background:#f59e0b;color:#fff;font-size:8.5px;font-weight:950;text-transform:uppercase;letter-spacing:.4px;'>Unfertig</span></div>"
        : "";
      html += "<tr style='background:" + bg + ";border-bottom:1px solid #f1f5f9;'>";
      html += "<td style='padding:7px 9px;font-weight:900;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>" + faEsc(r.fahrer) + "</td>";
      html += "<td style='padding:7px 9px;font-weight:800;color:#0f172a;font-variant-numeric:tabular-nums;white-space:nowrap;'>" + faEsc((r.wochentag ? r.wochentag + " " : "") + r.datum) + unfertigBadge + "</td>";
      html += "<td style='padding:7px 9px;color:#475569;font-variant-numeric:tabular-nums;white-space:nowrap;'>" + faEsc(r.beginn) + "</td>";
      html += "<td style='padding:7px 9px;color:#475569;font-variant-numeric:tabular-nums;white-space:nowrap;'>" + faEsc(r.ende) + "</td>";
      html += "<td style='padding:7px 9px;text-align:right;color:#1e3a5f;font-weight:800;font-variant-numeric:tabular-nums;white-space:nowrap;'>" + faEsc(r.schichtdauer) + "</td>";
      html += "<td style='padding:7px 9px;text-align:right;color:#be123c;font-weight:950;font-variant-numeric:tabular-nums;white-space:nowrap;'>" + faEsc(r.nettoText) + "</td>";
      html += "<td style='padding:7px 9px;color:#166534;font-weight:800;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>" + faEsc(r.lkw) + "</td>";
      html += "<td style='padding:7px 9px;font-weight:950;color:" + (found ? "#0f172a" : "#be123c") + ";white-space:normal;overflow:visible;text-overflow:clip;line-height:1.25;'>" + faEsc(r.tour) + "</td>";
      html += "<td style='padding:7px 9px;color:#64748b;font-size:10.5px;font-weight:700;white-space:normal;line-height:1.25;overflow:visible;text-overflow:clip;'>" + (found ? ("Excel: " + faEsc((r.planDatum ? r.planDatum + " · " : "") + (r.planZeit ? r.planZeit + " · " : "") + (r.planLkw ? r.planLkw + " · " : "") + (r.planFahrer ? r.planFahrer + " · " : "") + (r.matchSource || ""))) : "kein passender Excel-Eintrag") + "</td>";
      html += "</tr>";
    }});
    html += "</tbody></table></div>";
    return html;
  }}

  function _fa10hAggRows(rows, mode) {{
    var map = {{}};
    (rows || []).forEach(function(r) {{
      var tour = String(r.tour || "").trim();
      if (!tour) return;
      var driver = String(r.fahrer || "unbekannt").trim() || "unbekannt";
      if (mode === "tour") {{
        if (!map[tour]) map[tour] = {{ name: tour, count: 0, netto: 0, fahrer: {{}} }};
        map[tour].count += 1;
        map[tour].netto += r.nettoMin || 0;
        map[tour].fahrer[driver] = true;
      }} else {{
        if (!map[driver]) map[driver] = {{ name: driver, count: 0, netto: 0, touren: {{}} }};
        map[driver].count += 1;
        map[driver].netto += r.nettoMin || 0;
        map[driver].touren[tour] = true;
      }}
    }});
    return Object.keys(map).map(function(k) {{ return map[k]; }}).sort(function(a,b) {{
      return (b.count - a.count) || a.name.localeCompare(b.name, "de");
    }});
  }}

  function _faSafePdfName(v) {{
    return String(v || "")
      .replace(/[\\/:*?"<>|]+/g, "_")
      .replace(/\\s+/g, "_")
      .slice(0, 80) || "Auswertung";
  }}

  window.faExport10hPdf = function() {{
    var rows = (window.FA_LAST_10H_ALL_ROWS || window.FA_LAST_10H_ROWS || []).slice();
    if (!rows.length) rows = _faBuild10hRows("");
    if (!rows.length) {{ alert("Keine 10H Touren in der aktuellen Auswahl."); return; }}
    if(!window.jspdf || !window.jspdf.jsPDF || typeof window.jspdf.jsPDF !== "function") {{
      alert("PDF-Bibliothek ist noch nicht geladen. Bitte kurz warten und erneut versuchen.");
      return;
    }}
    var jsPDF = window.jspdf.jsPDF;
    var doc = new jsPDF({{ orientation:"portrait", unit:"mm", format:"a4" }});
    var yearLabel = faYearFilter === "all" ? "Alle Jahre" : faYearFilter;
    var sumNetto = rows.reduce(function(s, r) {{ return s + (r.nettoMin || 0); }}, 0);
    var title = "10H Touren - Touren-Häufigkeit - " + yearLabel;

    doc.setFont("helvetica", "bold");
    doc.setFontSize(15);
    doc.text(title, 12, 14);
    doc.setFont("helvetica", "normal");
    doc.setFontSize(9);
    doc.text("Treffer: " + rows.length + "   Netto gesamt: " + _fmtMin(sumNetto) + "   Zuordnung nach Anfangstag", 12, 20);

    var tourAgg = _fa10hAggRows(rows, "tour").map(function(t) {{
      return [t.name, String(t.count), String(_faCountKeys(t.fahrer)), _fmtMin(t.netto)];
    }});
    doc.autoTable({{
      startY: 27,
      head: [["Tournummer / Tourtext", "Anzahl", "Fahrer", "Netto"]],
      body: tourAgg,
      styles: {{ fontSize: 8, cellPadding: 1.8, overflow: "linebreak", valign: "top" }},
      headStyles: {{ fillColor: [109, 40, 217], textColor: 255, fontStyle: "bold" }},
      columnStyles: {{ 0: {{ cellWidth: 112 }}, 1: {{ halign: "right", cellWidth: 22 }}, 2: {{ halign: "right", cellWidth: 22 }}, 3: {{ halign: "right", cellWidth: 28 }} }},
      margin: {{ left: 12, right: 12 }},
      didDrawPage: function(data) {{
        var pageCount = doc.internal.getNumberOfPages();
        doc.setFontSize(7);
        doc.setTextColor(120);
        doc.text("Seite " + pageCount, 198, 288, {{ align: "right" }});
      }}
    }});
    doc.save("10H_Touren_Haeufigkeit_Tour_" + _faSafePdfName(yearLabel) + ".pdf");
  }};

  window.faExport10hDriverPdf = function(driverName) {{
    var rows = (window.FA_LAST_10H_ALL_ROWS || window.FA_LAST_10H_ROWS || []).slice();
    if (!rows.length) rows = _faBuild10hRows("");
    if (!driverName) {{
      var sel = document.getElementById("fa10hDriverPdfSel");
      driverName = sel ? sel.value : "";
    }}
    driverName = String(driverName || "").trim();
    if (!driverName) {{ alert("Bitte Fahrer auswählen."); return; }}
    rows = rows.filter(function(r) {{ return String(r.fahrer || "").trim() === driverName; }});
    if (!rows.length) {{ alert("Keine 10H Touren für " + driverName + " in der aktuellen Auswahl."); return; }}
    if(!window.jspdf || !window.jspdf.jsPDF || typeof window.jspdf.jsPDF !== "function") {{
      alert("PDF-Bibliothek ist noch nicht geladen. Bitte kurz warten und erneut versuchen.");
      return;
    }}
    var jsPDF = window.jspdf.jsPDF;
    var doc = new jsPDF({{ orientation:"landscape", unit:"mm", format:"a4" }});
    var yearLabel = faYearFilter === "all" ? "Alle Jahre" : faYearFilter;
    var sumNetto = rows.reduce(function(s, r) {{ return s + (r.nettoMin || 0); }}, 0);
    doc.setFont("helvetica", "bold");
    doc.setFontSize(15);
    doc.text("10H Touren - " + driverName, 10, 14);
    doc.setFont("helvetica", "normal");
    doc.setFontSize(9);
    doc.text("Auswahl: " + yearLabel + "   Treffer: " + rows.length + "   Netto gesamt: " + _fmtMin(sumNetto) + "   Zuordnung nach Anfangstag", 10, 20);

    var tourAgg = _fa10hAggRows(rows, "tour").map(function(t) {{
      return [t.name, String(t.count), _fmtMin(t.netto)];
    }});
    doc.autoTable({{
      startY: 27,
      head: [["Tournummer / Tourtext", "Anzahl", "Netto"]],
      body: tourAgg,
      styles: {{ fontSize: 7.5, cellPadding: 1.5, overflow: "linebreak" }},
      headStyles: {{ fillColor: [109, 40, 217], textColor: 255, fontStyle: "bold" }},
      columnStyles: {{ 0: {{ cellWidth: 105 }}, 1: {{ halign: "right", cellWidth: 20 }}, 2: {{ halign: "right", cellWidth: 25 }} }},
      margin: {{ left: 10, right: 140 }}
    }});
    var startRowsY = doc.lastAutoTable ? Math.max(doc.lastAutoTable.finalY + 8, 27) : 27;
    if (startRowsY > 72) {{ doc.addPage(); startRowsY = 14; }}

    doc.autoTable({{
      startY: startRowsY,
      head: [["Datum", "Beginn", "Ende", "Schicht", "Netto", "LKW", "Tournummer", "Zuordnung"]],
      body: rows.map(function(r) {{
        var found = r.tour && r.tour !== "nicht gefunden";
        return [
          (r.wochentag ? r.wochentag + " " : "") + (r.datum || ""),
          r.beginn || "",
          r.ende || "",
          r.schichtdauer || "",
          r.nettoText || "",
          r.lkw || "",
          r.tour || "",
          (r.planDatum ? r.planDatum + " · " : "") + (r.planZeit ? r.planZeit + " · " : "") + (r.planLkw ? r.planLkw + " · " : "") + (r.matchSource || "")
        ];
      }}),
      styles: {{ fontSize: 7.2, cellPadding: 1.5, overflow: "linebreak", valign: "top" }},
      headStyles: {{ fillColor: [124, 58, 237], textColor: 255, fontStyle: "bold" }},
      columnStyles: {{
        0: {{ cellWidth: 28 }},
        1: {{ cellWidth: 16 }},
        2: {{ cellWidth: 17 }},
        3: {{ cellWidth: 18, halign: "right" }},
        4: {{ cellWidth: 18, halign: "right" }},
        5: {{ cellWidth: 25 }},
        6: {{ cellWidth: 92 }},
        7: {{ cellWidth: 58 }}
      }},
      margin: {{ left: 10, right: 10 }},
      didDrawPage: function(data) {{
        var pageCount = doc.internal.getNumberOfPages();
        doc.setFontSize(7);
        doc.setTextColor(120);
        doc.text("Seite " + pageCount, 286, 202, {{ align: "right" }});
      }}
    }});
    doc.save("10H_Touren_" + _faSafePdfName(driverName) + "_" + _faSafePdfName(yearLabel) + ".pdf");
  }};

  window.FA_10H_MODE = false;
  window.FA_10H_SELECTED_DRIVER = "";

  window.faShow10hTours = function(driverFilter) {{
    var panel = document.getElementById("fa-detail-panel");
    if (!panel) return;
    _faSet10hButton(true);
    window.FA_10H_MODE = true;
    var selectedDriver = (driverFilter === undefined || driverFilter === null) ? String(window.FA_10H_SELECTED_DRIVER || "").trim() : String(driverFilter || "").trim();
    window.FA_10H_SELECTED_DRIVER = selectedDriver;
    if (selectedDriver) {{
      faSelectedName = selectedDriver;
      faBuildSidebarHighlight(selectedDriver);
    }} else {{
      faSelectedName = null;
    }}

    var allRows = _faBuild10hRows("");
    var rows = selectedDriver ? allRows.filter(function(r) {{ return String(r.fahrer || "") === selectedDriver; }}) : allRows;
    window.FA_LAST_10H_ALL_ROWS = allRows;
    window.FA_LAST_10H_ROWS = rows;
    var sumNetto = rows.reduce(function(s, r) {{ return s + (r.nettoMin || 0); }}, 0);
    var yearLabel = faYearFilter === "all" ? "Alle Jahre" : faYearFilter;
    var html = "";
    html += "<div style='background:#fff;border:1.5px solid #cbd5e1;border-radius:5px;padding:14px 18px;margin-bottom:12px;'>";
    html += "<div style='display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;'>";
    html += "<div><div style='font-size:20px;font-weight:950;color:#334155;'>10H Touren" + (selectedDriver ? " · " + faEsc(selectedDriver) : "") + "</div>";
    html += "<div style='font-size:11.5px;color:#64748b;font-weight:700;margin-top:3px;'>Alle Schichten mit mehr als 10:00 Netto-Arbeitszeit. Schichten ohne passende Tour werden mit „nicht gefunden“ angezeigt. Zuordnung nach Fahrer, Anfangsdatum, LKW und Startzeit.</div></div>";
    html += "<div style='display:flex;gap:8px;flex-wrap:wrap;align-items:center;'>";
    var driverOptions = Object.keys(allRows.reduce(function(o, r) {{ if (r.fahrer) o[r.fahrer] = true; return o; }}, {{}})).sort(function(a,b) {{ return a.localeCompare(b, "de"); }}).map(function(n) {{
      return "<option value='" + faEsc(n) + "'" + (n === selectedDriver ? " selected" : "") + ">" + faEsc(n) + "</option>";
    }}).join("");
    html += "<button type='button' onclick='faExport10hPdf()' style='border:2px solid #cbd5e1;background:#cbd5e1;color:#111827;border-radius:5px;padding:5px 12px;font-size:12px;font-weight:950;cursor:pointer;font-family:inherit;'>PDF Tour-Häufigkeit</button>";
    html += "<select id='fa10hDriverPdfSel' onchange='window.FA_10H_SELECTED_DRIVER=this.value; faShow10hTours(this.value);' style='border:1px solid #cbd5e1;background:#fff;color:#0f172a;border-radius:5px;padding:5px 8px;font-size:12px;font-weight:800;font-family:inherit;max-width:260px;'><option value=''>Alle Fahrer anzeigen</option>" + driverOptions + "</select>";
    html += "<button type='button' onclick='faExport10hDriverPdf()' style='border:2px solid #cbd5e1;background:#cbd5e1;color:#111827;border-radius:5px;padding:5px 12px;font-size:12px;font-weight:950;cursor:pointer;font-family:inherit;'>PDF Fahrer</button>";
    html += "<span style='background:#f1f5f9;color:#334155;border:1px solid #cbd5e1;border-radius:5px;padding:5px 12px;font-size:12px;font-weight:900;'>" + rows.length + " Treffer</span>";
    html += "<span style='background:#f8fbff;color:#0f172a;border:1px solid #e2e8f0;border-radius:5px;padding:5px 12px;font-size:12px;font-weight:800;'>Auswahl: " + faEsc(yearLabel) + "</span>";
    html += "<span style='background:#f8fbff;color:#0f172a;border:1px solid #e2e8f0;border-radius:5px;padding:5px 12px;font-size:12px;font-weight:800;'>Σ Netto " + _fmtMin(sumNetto) + "</span>";
    html += "</div></div></div>";
    html += _faRender10hFrequency(rows);
    html += _faRender10hRows(rows);
    panel.innerHTML = html;
    panel.scrollTop = 0;
  }};

  window.faGetFiltered = function() {{
    var q = (faSearchQuery || "").toLowerCase().trim();
    var list = _faAllShiftNames().map(function(n) {{
      return {{ name: n, years: {{}}, _shiftsOnly: true }};
    }}).filter(function(d) {{
      if (q && d.name.toLowerCase().indexOf(q) < 0) return false;
      if (faYearFilter !== "all" && !_faShiftRowsForYear(d.name).length) return false;
      return true;
    }});

    if (faCurrentSort === "arbeit") {{
      list.sort(function(a,b) {{
        var sa = _faShiftStatsForName(a.name);
        var sb = _faShiftStatsForName(b.name);
        return (sb.netto - sa.netto) || (sb.count - sa.count) || a.name.localeCompare(b.name, "de");
      }});
    }} else {{
      list.sort(function(a,b) {{ return a.name.localeCompare(b.name, "de"); }});
    }}
    return list;
  }};

  window.faBuildSidebarHighlight = function(activeName) {{
    var sidebar = document.getElementById("fa-sidebar-list");
    if (!sidebar) return;
    var filtered = window.faGetFiltered();

    var statsEl = document.getElementById("fa-stats");
    if (statsEl) {{
      var totalShifts = 0, totalNetto = 0, totalOver10 = 0, totalSick = 0, totalVacation = 0, totalVacationCredit = 0;
      filtered.forEach(function(d) {{
        var s = _faShiftStatsForName(d.name);
        var sick = _faSickEntries(d.name, faYearFilter, "all").length;
        var vacation = _faVacationEntries(d.name, faYearFilter, "all").length;
        totalShifts += s.count || 0;
        totalNetto += s.netto || 0;
        totalOver10 += s.over10 || 0;
        totalSick += sick || 0;
        totalVacation += vacation || 0;
        totalVacationCredit += vacation * 480;
      }});
      statsEl.innerHTML = "<b>" + filtered.length + "</b> Fahrer &nbsp;&middot;&nbsp; "
        + "<b>" + totalShifts + "</b> Schichten &nbsp;&middot;&nbsp; Krank <b style='color:#be123c;'>" + totalSick + "</b>"
        + " &nbsp;&middot;&nbsp; Urlaub <b style='color:#075985;'>" + totalVacation + "</b>"
        + " &nbsp;&middot;&nbsp; Σ Netto Schichten <b>" + _fmtMin(totalNetto) + "</b>"
        + " &nbsp;&middot;&nbsp; Urlaub +8h <b style='color:#075985;'>" + _fmtMin(totalVacationCredit) + "</b>"
        + " &nbsp;&middot;&nbsp; Σ inkl. Urlaub <b>" + _fmtMin(totalNetto + totalVacationCredit) + "</b>"
        + " &nbsp;&middot;&nbsp; &gt;10:00 Netto <b style='color:#be123c;'>" + totalOver10 + "</b>";
    }}

    var html = "";
    filtered.forEach(function(d) {{
      var s = _faShiftStatsForName(d.name);
      var sick = _faSickEntries(d.name, faYearFilter, "all").length;
      var vacation = _faVacationEntries(d.name, faYearFilter, "all").length;
      var active = d.name === activeName;
      var bg = active ? "#334155" : "#fff";
      var fg = active ? "#fff" : "#0b1220";
      var sub = active ? "rgba(255,255,255,.88)" : "#64748b";
      var badgeBg = active ? "rgba(255,255,255,.25)" : "#f1f5f9";
      var badgeFg = active ? "#fff" : "#334155";
      var overBg = active ? "#991b1b" : "#fee2e2";
      var overFg = active ? "#fecaca" : "#be123c";
      html += "<div data-fa-driver='" + faEsc(d.name) + "'"
        + " style='padding:10px 14px;cursor:pointer;border-bottom:1px solid #f1f5f9;background:" + bg + ";'>"
        + "<div style='font-weight:800;font-size:13px;color:" + fg + ";white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>" + faEsc(d.name) + "</div>"
        + "<div style='font-size:10px;color:" + sub + ";font-weight:700;margin-top:3px;font-variant-numeric:tabular-nums;'>" + s.count + " Schichten" + (sick ? " · Krank " + sick : "") + (vacation ? " · Urlaub " + vacation + " (+" + _fmtMin(vacation * 480) + ")" : "") + " · Netto " + _fmtMin(s.netto + (vacation * 480)) + "</div>"
        + "<div style='display:flex;gap:3px;flex-wrap:wrap;margin-top:4px;'>"
        + "<span style='font-size:8px;font-weight:800;padding:1px 4px;border-radius:2px;background:" + badgeBg + ";color:" + badgeFg + ";'>" + s.count + " S</span>"
        + (sick ? "<span style='font-size:8px;font-weight:800;padding:1px 4px;border-radius:2px;background:" + (active ? "#991b1b" : "#fee2e2") + ";color:" + (active ? "#fecaca" : "#be123c") + ";'>K " + sick + "</span>" : "")
        + (vacation ? "<span style='font-size:8px;font-weight:800;padding:1px 4px;border-radius:2px;background:" + (active ? "#075985" : "#f1f5f9") + ";color:" + (active ? "#bae6fd" : "#075985") + ";'>U " + vacation + "</span>" : "")
        + (s.over10 ? "<span style='font-size:8px;font-weight:800;padding:1px 4px;border-radius:2px;background:" + overBg + ";color:" + overFg + ";'>&gt;10h " + s.over10 + "</span>" : "")
        + "</div></div>";
    }});
    sidebar.innerHTML = html || "<div style='padding:20px;color:#94a3b8;font-size:12px;text-align:center;'>Keine Tachograph-Schichten gefunden</div>";
    Array.prototype.slice.call(sidebar.querySelectorAll("[data-fa-driver]")).forEach(function(item) {{
      item.onclick = function() {{ window.faShowDetail(item.getAttribute("data-fa-driver")); }};
    }});

    if (!activeName && filtered.length) {{
      faSelectedName = filtered[0].name;
      window.faShowDetail(filtered[0].name);
    }} else if (activeName) {{
      faSelectedName = activeName;
    }}
  }};

  window.faPopulateYears = function() {{
    var allYears = [];
    Object.keys(TIMEREC_DATA || {{}}).forEach(function(n) {{
      (TIMEREC_DATA[n] || []).forEach(function(s) {{
        var m = (s.tag || "").match(/(\\d{{4}})$/);
        if (m && m[1] !== "2024" && allYears.indexOf(m[1]) === -1) allYears.push(m[1]);
      }});
    }});
    allYears.sort().reverse();
    var yrSel = document.getElementById("fa-year-sel");
    if (!yrSel) return;
    var current = yrSel.value;
    yrSel.innerHTML = allYears.map(function(y) {{ return "<option value='" + y + "'>" + y + "</option>"; }}).join("");
    if (allYears.length > 1) yrSel.insertAdjacentHTML("beforeend", "<option value='all'>Alle Jahre</option>");
    var curYr = String(new Date().getFullYear());
    if (current && (allYears.indexOf(current) !== -1 || current === "all")) yrSel.value = current;
    else if (allYears.indexOf(curYr) !== -1) yrSel.value = curYr;
    else if (allYears.length) yrSel.value = allYears[0];
    else yrSel.innerHTML = "<option value='all'>Keine Daten</option>";
    faYearFilter = yrSel.value || "all";
  }};

  window.faShowDetail = function(name) {{
    faSelectedName = name;
    if (window.FA_10H_MODE) {{
      window.FA_10H_SELECTED_DRIVER = name || "";
      faBuildSidebarHighlight(name);
      window.faShow10hTours(name);
      return;
    }}
    window.FA_10H_MODE = false;
    _faSet10hButton(false);
    faBuildSidebarHighlight(name);
    var panel = document.getElementById("fa-detail-panel");
    if (!panel) return;
    if (!faHasShifts(name)) {{
      panel.innerHTML = "<div style='color:#94a3b8;padding:40px;text-align:center;font-size:14px;'>Keine Schichten / Tachograph-Daten f&uuml;r diesen Fahrer.</div>";
      return;
    }}
    faRenderShifts(name, panel);
  }};

  window.faRender = function(q) {{
    faSearchQuery = q || "";
    if (!TIMEREC_DATA || !Object.keys(TIMEREC_DATA).length) {{
      var c = document.getElementById("fa-detail-panel");
      if (c) c.innerHTML = "<div style='color:#94a3b8;padding:40px;text-align:center;font-size:14px;'>Keine Tachograph-Daten vorhanden.<br>Bitte CSV <b>timerecording_v3*.csv</b> hochladen.</div>";
      var side = document.getElementById("fa-sidebar-list");
      if (side) side.innerHTML = "";
      var stats = document.getElementById("fa-stats");
      if (stats) stats.innerHTML = "";
      return;
    }}
    window.faPopulateYears();
    if (window.FA_10H_MODE) window.faShow10hTours(window.FA_10H_SELECTED_DRIVER || "");
    else window.faBuildSidebarHighlight(null);

  }};

  // ── Fahrer Übersicht: normale Timerecording-/Tachograph-Auswertung ──
  // Der Button darf NICHT auf die alte Excel-Fahrerauswertung zurückschalten.
  // Er schaltet nur aus der 10H-Ansicht zurück auf die normale Schichtenansicht.
  window.faShowFahrerUebersicht = function() {{
    window.FA_10H_MODE = false;
    window.FA_10H_SELECTED_DRIVER = "";
    _faSet10hButton(false);
    faBuildSidebarHighlight(null);
  }};

}})();
// ── /Schichten-Tab ─────────────────────────────────────────────────────────────

{zulage_js_code}
{_zulagen_graph_js()}
{verstoss_js_code}


// ── Infos & Aushänge: Reinigungsnachweis ergänzen ─────────────────────────────
(function() {{
  var _buildInfosDdMenuOriginal = (typeof window.buildInfosDdMenu === "function") ? window.buildInfosDdMenu : null;

  function _infosCloseDropdowns() {{
    document.querySelectorAll(".nav-dd").forEach(function(d) {{ d.classList.remove("open"); }});
  }}

  window.buildInfosDdMenu = function() {{
    var menu = document.getElementById("ddmenu-infos");
    if(!menu) return;

    if(_buildInfosDdMenuOriginal) {{
      _buildInfosDdMenuOriginal();
    }} else {{
      var fallbackItems = [
        ["tel", "Telefonliste"],
        ["bus", "Mitarbeiterbus"],
        ["arzt", "Betriebsärztin"],
        ["versp", "Verspätungstabelle"],
        ["knapp", "KNAPP"],
        ["schluessel", "Schlüsselübergabe"],
        ["entgelt", "Entgeltportal"],
        ["schaden", "Schadenmeldung Fuhrpark"],
        ["maengel", "Mängelanzeige Fuhrpark"],
        ["lkw_uebergabe", "Übergabeprotokoll LKW"],
        ["reinigung", "🧼 Reinigungsnachweis LKW"],
        ["reisekosten", "📊 Reisekostenabrechnung"],
        ["balzer", "Balzer"],
        ["termine", "Termine Fahrer"]
      ];
      menu.innerHTML = fallbackItems.map(function(it) {{
        var active = currentArea === it[0] ? " active" : "";
        return "<div class='dd-item" + active + "' data-area='" + it[0] + "'>" + it[1] + "</div>";
      }}).join("");
      Array.prototype.slice.call(menu.querySelectorAll("[data-area]")).forEach(function(el) {{
        el.onclick = function() {{ showArea(el.getAttribute("data-area")); _infosCloseDropdowns(); }};
      }});
    }}

    if(!menu.querySelector('[data-area="reinigung"]')) {{
      var item = document.createElement("div");
      item.className = "dd-item" + (currentArea === "reinigung" ? " active" : "");
      item.setAttribute("data-area", "reinigung");
      item.textContent = "🧼 Reinigungsnachweis LKW";
      item.onclick = function() {{ showArea("reinigung"); _infosCloseDropdowns(); }};
      menu.appendChild(item);
    }} else {{
      var existing = menu.querySelector('[data-area="reinigung"]');
      existing.className = "dd-item" + (currentArea === "reinigung" ? " active" : "");
      existing.textContent = "🧼 Reinigungsnachweis LKW";
    }}

    if(!menu.querySelector('[data-area="reisekosten"]')) {{
      var excelItem = document.createElement("div");
      excelItem.className = "dd-item" + (currentArea === "reisekosten" ? " active" : "");
      excelItem.setAttribute("data-area", "reisekosten");
      excelItem.textContent = "📊 Reisekostenabrechnung";
      excelItem.onclick = function() {{ showArea("reisekosten"); _infosCloseDropdowns(); }};
      menu.appendChild(excelItem);
    }} else {{
      var existingExcel = menu.querySelector('[data-area="reisekosten"]');
      existingExcel.className = "dd-item" + (currentArea === "reisekosten" ? " active" : "");
      existingExcel.textContent = "📊 Reisekostenabrechnung";
    }}
  }};
}})();


</script>

</body>
</html>"""


def combine_html(instances: list, tel_json: str = "[]", sam_json: str = "[]", fa_json: str = "[]", zulage_json: str = "{}", zulage_xlsx_sonder: str = "", zulage_xlsx_fuengers: str = "", drittkunden_json: str = "[]", zulage_xlsx_drittkunden: str = "", fahrzeugwaesche_json: str = "[]", tanken_json: str = "[]", verstoss_json: str = '{"drivers":[],"total_violations":0}', spesen_json: str = '{"drivers":[],"months":[],"total_cost":0,"total_rows":0}', grosskunden_json: str = "[]", timerec_json: str = "{}", spediteure_json: str = '{"katalog":[],"fahrten":[]}', fahrerbewertung_json: str = '{"profile":"","event_types":[],"g_months":{},"g_ev":{},"drivers":[]}', versp_abfahrt_json: str = "{}", last_updated: str = "", generation_meta: dict | None = None) -> str:
    _combine_started = time.perf_counter()
    try:
        _logo_up = st.session_state.get("g_logo")
    except Exception:
        _logo_up = None
    logo_data_url = logo_file_to_data_uri(_logo_up) or load_logo_data_uri()

    """
    Bettet beliebig viele Suche+Druck-Paare (Instanzen) in eine HTML ein.
    Instanz-Wechsler im Topnav. BLP Druck ist intern (hidden iframe für FW-Daten).
    """
    js_parts = _dashboard_javascript_parts()
    spesen_js_code = js_parts['spesen']
    fa_js_code = js_parts['fa']
    zulage_js_code = js_parts['zulage']
    wash_js_code = js_parts['wash']
    wash_ranking_js_code = js_parts['wash_ranking']
    fw_graph_js_code = js_parts['fw_graph']
    tank_panels_html = _tank_panels_html()
    tank_js_code = _tank_dashboard_js()
    verstoss_js_code = js_parts['verstoss']
    knapp_js_code = js_parts['knapp']
    sped_js_code = js_parts['sped']
    fabew_js_code = js_parts['fabew']
    bus_js_code = js_parts['bus']
    arzt_js_code = js_parts['arzt']
    versp_js_code = js_parts['versp']
    wa_js_code = js_parts['wa']




    # Rangliste + PDF pro Fahrer (separate Variable, Triple-Quote → keine Escape-Hölle)

    # ── Fahrzeugwäsche Graph (eigener raw-string, analog zur Verstoßauswertung) ─

    # ── Verstoßauswertung (eigener raw-string, damit keine Escape-Hölle) ─────

    documents_js_code = _build_pdf_documents_js()



    # Alle Instanzen als JS-Array vorbereiten
    instances_js = _build_instances_js(instances)


    # Sa-/So-Auswertung ausschließlich aus tatsächlichen TIMEREC-Schichten aufbauen.
    # Entscheidend ist immer der Anfangstag der Schicht. Touren-Excel ist hierfür
    # bewusst keine Datenquelle, damit geplante und tatsächlich gefahrene Einsätze
    # nicht vermischt werden.
    sam_json = _build_saturday_json_from_timerecording(timerec_json)








    # Große Zusatzdaten nicht mehr als unkomprimiertes JSON in die HTML schreiben.
    # Stattdessen werden alle Daten gemeinsam per zlib komprimiert und als Base64
    # eingebettet. Im Browser erfolgt das Entpacken einmalig beim Start.
    embedded_data_js = _build_embedded_data_js(
        fahrzeugwaesche_json=fahrzeugwaesche_json,
        tanken_json=tanken_json,
        tel_json=tel_json,
        sam_json=sam_json,
        fa_json=fa_json,
        fahrerbewertung_json=fahrerbewertung_json,
        zulage_json=zulage_json,
        drittkunden_json=drittkunden_json,
        verstoss_json=verstoss_json,
        spesen_json=spesen_json,
        grosskunden_json=grosskunden_json,
        timerec_json=timerec_json,
        spediteure_json=spediteure_json,
        versp_abfahrt_json=versp_abfahrt_json,
    )


    _meta = dict(generation_meta or {})
    _meta.setdefault("version", APP_DISPLAY_VERSION)
    _meta.setdefault("created_at", datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S"))
    _meta["html_bytes"] = "00000000000000000"
    _meta["generation_ms"] = "0000000000"
    generation_meta_json = json.dumps(_meta, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")

    _html = _render_dashboard_html(
        logo_data_url=logo_data_url,
        last_updated=last_updated,
        generation_meta_json=generation_meta_json,
        instances_js=instances_js,
        embedded_data_js=embedded_data_js,
        documents_js_code=documents_js_code,
        knapp_js_code=knapp_js_code,
        spesen_js_code=spesen_js_code,
        fa_js_code=fa_js_code,
        zulage_js_code=zulage_js_code,
        wash_js_code=wash_js_code,
        wash_ranking_js_code=wash_ranking_js_code,
        fw_graph_js_code=fw_graph_js_code,
        tank_panels_html=tank_panels_html,
        tank_js_code=tank_js_code,
        verstoss_js_code=verstoss_js_code,
        sped_js_code=sped_js_code,
        fabew_js_code=fabew_js_code,
        bus_js_code=bus_js_code,
        arzt_js_code=arzt_js_code,
        versp_js_code=versp_js_code,
        wa_js_code=wa_js_code,
        zulage_xlsx_sonder=zulage_xlsx_sonder,
        zulage_xlsx_fuengers=zulage_xlsx_fuengers,
        zulage_xlsx_drittkunden=zulage_xlsx_drittkunden,
    )
    _generation_ms = max(0, int(round((time.perf_counter() - _combine_started) * 1000)))
    _html_size = len(_html.encode("utf-8"))
    _html = _html.replace(
        '"html_bytes":"00000000000000000"',
        f'"html_bytes":"{_html_size:017d}"',
        1,
    ).replace(
        '"generation_ms":"0000000000"',
        f'"generation_ms":"{_generation_ms:010d}"',
        1,
    )
    return _html


def build_plane_zulagen_json(zulage_json_str: str, drittkunden_json_str: str, generated_at: str = "") -> bytes:
    """Erzeugt eine flache JSON-Datei fuer plane.php aus den bereits berechneten Zulagen.

    Die Plane muss dadurch keine Excel-Dateien lesen. Sie bekommt nur noch eine Datei:
    csv/zulagen.json
    """
    import json as _json
    import re as _re
    import datetime as _dt

    def _loads(value, fallback):
        try:
            return _json.loads(value or "")
        except Exception:
            return fallback

    def _amount(value) -> float:
        try:
            if value is None or value == "":
                return 0.0
            if isinstance(value, str):
                value = value.replace("€", "").replace(".", "").replace(",", ".").strip()
            return float(value)
        except Exception:
            return 0.0

    def _date_parts(raw: str):
        raw = str(raw or "").strip()
        m = _re.search(r"(\d{2})\.(\d{2})\.(\d{4})", raw)
        if not m:
            return "", "", None, None
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            dt = _dt.date(y, mo, d)
            return dt.strftime("%d.%m.%Y"), dt.isoformat(), dt.year, dt.month
        except Exception:
            return m.group(0), "", None, None

    def _kw(raw: str, iso: str) -> str:
        raw = str(raw or "")
        m = _re.search(r"KW\s*([0-9]{1,2})", raw, flags=_re.IGNORECASE)
        if m:
            return "KW" + m.group(1).zfill(2)
        if iso:
            try:
                return "KW" + str(_dt.date.fromisoformat(iso).isocalendar()[1]).zfill(2)
            except Exception:
                pass
        return ""

    entries = []
    zulage_data = _loads(zulage_json_str, {})
    dk_data = _loads(drittkunden_json_str, [])

    def _add_common(typ: str, fahrer: dict, tag: dict, betrag_key: str, extra: dict):
        raw_date = tag.get("datum") or tag.get("datum_raw") or ""
        datum, iso, jahr, monat = _date_parts(raw_date)
        if not datum:
            return
        amount = _amount(tag.get(betrag_key, tag.get("betrag", 0)))
        item = {
            "typ": typ,
            "fahrer": fahrer.get("name", ""),
            "persnr": fahrer.get("persnr", ""),
            "datum": datum,
            "datum_iso": iso,
            "datum_raw": raw_date,
            "jahr": jahr,
            "monat": monat,
            "kw": tag.get("kw") or _kw(raw_date, iso),
            "betrag": amount,
        }
        item.update({k: ("" if v is None else v) for k, v in extra.items()})
        entries.append(item)

    for typ_key, typ_label in (("sonder", "Sonderfahrzeug"), ("fuengers", "Füngers")):
        for monat in (zulage_data.get(typ_key, []) if isinstance(zulage_data, dict) else []):
            for fahrer in monat.get("fahrer", []):
                for tag in fahrer.get("tage", []):
                    if typ_key == "sonder":
                        _add_common(typ_label, fahrer, tag, "verdienst", {
                            "tour": tag.get("tour", ""),
                            "lkw": tag.get("lkw", ""),
                            "art": tag.get("art", ""),
                            "kommentar": tag.get("kommentar", ""),
                            "info": tag.get("info", ""),
                        })
                    else:
                        _add_common(typ_label, fahrer, tag, "verdienst", {
                            "tour": tag.get("tour", ""),
                            "lkw": tag.get("lkw", ""),
                            "art": tag.get("art", ""),
                            "kommentar": tag.get("kommentar", ""),
                            "info": tag.get("info", ""),
                        })

    if isinstance(dk_data, list):
        for monat in dk_data:
            for fahrer in monat.get("fahrer", []):
                for tag in fahrer.get("tage", []):
                    _add_common("Drittkunden", fahrer, tag, "zulage", {
                        "tour": tag.get("tour", ""),
                        "lkw": tag.get("lkw", ""),
                        "art": tag.get("art", ""),
                        "kommentar": tag.get("kommentar", ""),
                        "info": tag.get("info", ""),
                    })

    entries.sort(key=lambda e: (e.get("datum_iso") or "", e.get("fahrer") or "", e.get("typ") or ""), reverse=True)
    payload = {
        "schema": "plane_zulagen_v1",
        "generated_at": generated_at or _dt.datetime.now().strftime("%d.%m.%Y %H:%M"),
        "source": "Streamlit Touren-Dateien",
        "entries": entries,
        "totals": {
            "entries": len(entries),
            "betrag": round(sum(_amount(e.get("betrag")) for e in entries), 2),
        },
    }
    return _json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

def generate_zulage_excel(zulage_json_str: str, tab: str = "sonder") -> bytes:
    """Formatierte Excel mit openpyxl – inkl. Persnr. und Zusammenfassung."""
    import io as _io, json as _j
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    try:
        data = _j.loads(zulage_json_str)
    except Exception:
        return None
    months = data.get(tab, [])
    if not months:
        return None
    wb = Workbook(); wb.remove(wb.active)
    def _s(st="thin",c="CCCCCC"): return Side(style=st, color=c)
    thin = Border(left=_s(),right=_s(),top=_s(),bottom=_s())
    med  = Border(left=_s("medium","1F4E78"),right=_s("medium","1F4E78"),top=_s("medium","1F4E78"),bottom=_s("medium","1F4E78"))
    gmed = Border(left=_s("medium","70AD47"),right=_s("medium","70AD47"),top=_s("medium","70AD47"),bottom=_s("medium","70AD47"))
    FT=PatternFill("solid",fgColor="1F4E78"); FH=PatternFill("solid",fgColor="D9E2F3")
    FN=PatternFill("solid",fgColor="4472C4"); FG=PatternFill("solid",fgColor="70AD47")
    FL=PatternFill("solid",fgColor="1F4E78"); FA=PatternFill("solid",fgColor="F8F9FA")
    FW=PatternFill("solid",fgColor="FFFFFF"); FS=PatternFill("solid",fgColor="EEF4FF")

    # Track summary: {(name, persnr): total}
    summary = {}

    for monat in months:
        ws = wb.create_sheet(title=monat["monat"][:31])
        is_s = (tab == "sonder")
        # Columns: Name | Persnr. | Datum | Tour | LKW | Art | Verdienst
        #      or: Name | Persnr. | Datum | Kommentar | Verdienst
        hdrs   = ["Name","Persnr.","Datum","Tour","LKW","Art","Verdienst"] if is_s else ["Name","Persnr.","Datum","Kommentar","Verdienst"]
        widths = [24,13,32,14,10,14,14]                                    if is_s else [24,13,32,40,14]
        nc = len(hdrs)

        # Title
        ws.append([monat["monat"]]+[""]*( nc-1)); tr=ws.max_row
        ws.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=nc)
        c=ws.cell(tr,1); c.font=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
        c.fill=FT; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=med
        ws.row_dimensions[tr].height=28

        # Header
        ws.append(hdrs); hr=ws.max_row
        for col in range(1,nc+1):
            c=ws.cell(hr,col); c.font=Font(name="Calibri",bold=True,size=10,color="1F4E78")
            c.fill=FH; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin
        ws.row_dimensions[hr].height=22

        alt=False
        for fahrer in monat["fahrer"]:
            r0=ws.max_row+1
            # Accumulate for summary
            key=(fahrer["name"], fahrer["persnr"])
            summary[key] = summary.get(key, 0) + fahrer["gesamt"]

            for ti,tag in enumerate(fahrer["tage"]):
                if is_s:
                    tour=tag.get("tour","")
                    if not tour or tour=="zbv": tour="z.b.v."
                    row=[fahrer["name"] if ti==0 else "",
                         fahrer["persnr"] if ti==0 else "",
                         tag["datum"],tour,tag.get("lkw",""),tag.get("art",""),tag["verdienst"]]
                else:
                    row=[fahrer["name"] if ti==0 else "",
                         fahrer["persnr"] if ti==0 else "",
                         tag["datum"],tag.get("kommentar",""),tag["verdienst"]]
                ws.append(row); r=ws.max_row; fill=FA if alt else FW
                for col in range(1,nc+1):
                    if col in (1,2) and ti==0: continue   # styled after merge
                    c=ws.cell(r,col); iv=(col==nc)
                    c.font=Font(name="Calibri",size=10,color="16A34A" if iv else "2C3E50",bold=iv)
                    c.fill=fill; c.border=thin
                    c.alignment=Alignment(horizontal="right" if iv else ("center" if col==2 else "left"),vertical="center")
                    if iv and isinstance(c.value,(int,float)): c.number_format='#,##0.00 "€"'
                ws.row_dimensions[r].height=20; alt=not alt

            # Merge Name column over all tage
            ws.merge_cells(start_row=r0,start_column=1,end_row=ws.max_row,end_column=1)
            nc2=ws.cell(r0,1); nc2.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
            nc2.fill=FN; nc2.alignment=Alignment(horizontal="left",vertical="center"); nc2.border=med

            # Merge Persnr column over all tage
            ws.merge_cells(start_row=r0,start_column=2,end_row=ws.max_row,end_column=2)
            pc=ws.cell(r0,2); pc.font=Font(name="Calibri",bold=False,size=10,color="FFFFFF")
            pc.fill=FN; pc.alignment=Alignment(horizontal="center",vertical="center"); pc.border=med

            # Gesamt row
            gv=[""]*nc; gv[nc-2]="Gesamt"; gv[nc-1]=fahrer["gesamt"]
            ws.append(gv); gr=ws.max_row
            ws.merge_cells(start_row=gr,start_column=1,end_row=gr,end_column=nc-1)
            for col in range(1,nc+1):
                c=ws.cell(gr,col); c.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
                c.fill=FG; c.alignment=Alignment(horizontal="right",vertical="center"); c.border=gmed
                if col==nc and isinstance(c.value,(int,float)): c.number_format='#,##0.00 "€"'
            ws.row_dimensions[gr].height=20; alt=False

            # Spacer
            ws.append([""]*nc); ws.row_dimensions[ws.max_row].height=6

        # Monatsgesamt
        tv=[""]*nc; tv[nc-2]="Monatsgesamt"; tv[nc-1]=sum(f["gesamt"] for f in monat["fahrer"])
        ws.append(tv); tr2=ws.max_row
        ws.merge_cells(start_row=tr2,start_column=1,end_row=tr2,end_column=nc-1)
        for col in range(1,nc+1):
            c=ws.cell(tr2,col); c.font=Font(name="Calibri",bold=True,size=13,color="FFFFFF")
            c.fill=FL; c.alignment=Alignment(horizontal="right",vertical="center"); c.border=med
            if col==nc and isinstance(c.value,(int,float)): c.number_format='#,##0.00 "€"'
        ws.row_dimensions[tr2].height=26

        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes="B3"   # freeze Name+Persnr columns + title+header rows

    # ── Zusammenfassung sheet ─────────────────────────────────────────────
    if summary:
        ws_sum = wb.create_sheet(title="Zusammenfassung", index=0)
        tab_label = "Sonderfahrzeuge" if tab=="sonder" else "F\u00fcngers"
        # Title
        ws_sum.append([f"Zusammenfassung \u2013 {tab_label}","",""])
        tr=ws_sum.max_row
        ws_sum.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=3)
        c=ws_sum.cell(tr,1); c.font=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
        c.fill=FT; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=med
        ws_sum.row_dimensions[tr].height=30

        # Header
        ws_sum.append(["Name","Personalnummer","Gesamt"]); hr=ws_sum.max_row
        for col in range(1,4):
            c=ws_sum.cell(hr,col); c.font=Font(name="Calibri",bold=True,size=11,color="1F4E78")
            c.fill=FH; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin
        ws_sum.row_dimensions[hr].height=24

        # Data rows sorted by name
        total_all = 0
        for ri,(( name, persnr), total) in enumerate(sorted(summary.items(), key=lambda x:x[0][0])):
            fill = FA if ri%2==0 else FW
            ws_sum.append([name, persnr, total]); r=ws_sum.max_row
            total_all += total
            for col in range(1,4):
                c=ws_sum.cell(r,col); iv=(col==3)
                c.font=Font(name="Calibri",size=11,color="16A34A" if iv else "2C3E50",bold=iv)
                c.fill=fill; c.border=thin
                c.alignment=Alignment(horizontal="right" if iv else ("center" if col==2 else "left"),vertical="center")
                if iv: c.number_format='#,##0.00 "€"'
            ws_sum.row_dimensions[r].height=21

        # Gesamtsumme
        ws_sum.append(["","Gesamtsumme", total_all]); gr=ws_sum.max_row
        for col in range(1,4):
            c=ws_sum.cell(gr,col); c.font=Font(name="Calibri",bold=True,size=13,color="FFFFFF")
            c.fill=FL; c.border=med
            c.alignment=Alignment(horizontal="right",vertical="center")
            if col==3: c.number_format='#,##0.00 "€"'
        ws_sum.row_dimensions[gr].height=28

        ws_sum.column_dimensions["A"].width=26
        ws_sum.column_dimensions["B"].width=18
        ws_sum.column_dimensions["C"].width=16
        ws_sum.freeze_panes="A3"

    buf=_io.BytesIO(); wb.save(buf); return buf.getvalue()


_ZULAGE_PERSNR = {
    "Adler":{"Philipp":"00041450"},"Auer":{"Frank":"00020795"},
    "Batkowski":{"Tilo":"00046601"},"Benabbes":{"Badr":"00048980"},
    "Biebow":{"Thomas":"00042004"},"Blaesing":{"Elmar":"00049093"},
    "Bursian":{"Ronny":"00025714"},"Buth":{"Sven":"00046673"},
    "Boehnke":{"Marcel":"00020833"},"Carstensen":{"Martin":"00042412"},
    "Chege":{"Moses Gichuru":"00046106"},"Dammasch":{"Bernd":"00019297"},
    "Demuth":{"Harry":"00020796"},"Doroszkiewicz":{"Bogumil":"00049132"},
    "Duerr":{"Holger":"00039164"},"Effenberger":{"Sven":"00030807"},
    "Engel":{"Raymond":"00033429"},
    "Fechner":{"Danny":"00043696","Klaus":"00038278"},
    "Findeklee":{"Bernd":"00020804"},"Flint":{"Henryk":"00042414"},
    "Fuhlbruegge":{"Justin":"00046289"},"Gehrmann":{"Rayk":"00046702"},
    "Gheonea":{"Daniel-Costel":"00054489"},"Glanz":{"Bjoern":"00041914"},
    "Gnech":{"Torsten":"00018613"},"Greve":{"Nicole":"00040760"},
    "Guthmann":{"Fred":"00018328"},"Hagen":{"Andy":"00020271"},
    "Hartig":{"Sebastian":"00044120"},"Haus":{"David":"00046101"},
    "Heeser":{"Bernd":"00041916"},"Helm":{"Philipp":"00046685"},
    "Henkel":{"Bastian":"00048187"},"Holtz":{"Torsten":"00021159"},
    "Hirdina":{"Christopher":"00053400"},"Hintz":{"Leif":"00054808"},
    "Huebner":{"Christian":"00054531"},"Janikiewicz":{"Radoslaw":"00042159"},
    "Kelling":{"Jonas Ole":"00044140"},"Kleiber":{"Lutz":"00026255"},
    "Klemkow":{"Ralf":"00040634"},"Kollmann":{"Steffen":"00040988"},
    "Koenig":{"Heiko":"00036341"},"Krazewski":{"Cezary":"00039463"},
    "Krieger":{"Christian":"00049092"},"Krull":{"Benjamin":"00044192"},
    "Lange":{"Michael":"00035407"},"Lewandowski":{"Kamil":"00041044"},
    "Likoonski":{"Vladimir":"00044766"},"Linke":{"Erich":"00048377"},
    "Lefkih":{"Houssni":"00052293"},"Ludolf":{"Michel":"00048814"},
    "Laemmel":{"Patrick":"00052946"},"Marouni":{"Ayyoub":"00048986"},
    "Mintel":{"Mario":"00046686"},"Ohlenroth":{"Nadja":"00042114"},
    "Ohms":{"Torsten":"00019300"},"Okoth":{"Tedy Omondi":"00046107"},
    "Oszmian":{"Jacub":"00039464"},"Paul":{"Toralf":"00010490"},
    "Pabst":{"Torsten":"00021976"},"Pawlak":{"Bartosz":"00036381"},
    "Piepke":{"Torsten":"00021390"},"Plinke":{"Killian":"00044137"},
    "Pogodski":{"Enrico":"00046668"},"Postu":{"Mihai":"00051391"},
    "Quint":{"Stefan":"00035718"},"Rimba":{"Rimba Gona":"00046108"},
    "Rheinschmitt":{"Ronald":"00053356"},"Rudert":{"Kevin":"00052858"},
    "Rudolph":{"Yves":"00052855"},"Ruge":{"Fabian":"00054705"},
    "Sarwatka":{"Heiko":"00028747"},"Swietoslawski":{"Jacek":"00052955"},
    "Seredynski":{"Ireneusz":"00053452"},
    "Scheil":{"Eric-Rene":"00038579","Rene":"00020851"},
    "Schlichting":{"Michael":"00021452"},
    "Schlutt":{"Hubert":"00020880","Rene":"00042932"},
    "Schmieder":{"Steffen":"00046286"},"Schneider":{"Matthias":"00045495"},
    "Schulz":{"Julian":"00049130","Stephan":"00041558"},
    "Singh":{"Jagtah":"00040902"},"Stoltz":{"Thorben":"00040991"},
    "Thal":{"Jannic":"00046006"},"Tumanow":{"Vasilli":"00045019"},
    "Wachnowski":{"Klaus":"00026019"},"Wendel":{"Danilo":"00048994"},
    "Waschitschek":{"Detlef":"00020436"},"Wille":{"Rene":"00021393"},
    "Wisniewski":{"Krzysztof":"00046550"},"Zander":{"Jan":"00042454"},
    "Zosel":{"Ingo":"00026303"},"Strehlow":{"Yves":"00052855"},
}
_ZP_MONTHS_DE = ["","Januar","Februar","März","April","Mai","Juni",
                  "Juli","August","September","Oktober","November","Dezember"]
_ZP_DAYS = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"]


def _zp_norm(s):
    import unicodedata as _uc
    s=(s or "").strip().lower().replace(" "," ")
    s=_uc.normalize("NFKC",s).replace("-"," ")
    for a,b in [("ä","ae"),("ö","oe"),("ü","ue"),("ß","ss")]: s=s.replace(a,b)
    return " ".join(s.split())


def _zp_persnr(nn,vn):
    nk,vk=_zp_norm(nn),_zp_norm(vn)
    for ln,inner in _ZULAGE_PERSNR.items():
        if _zp_norm(ln)==nk:
            for fn,pn in inner.items():
                if _zp_norm(fn)==vk: return pn
            for fn,pn in inner.items():
                fn2=_zp_norm(fn)
                if vk.startswith(fn2) or fn2.startswith(vk): return pn
            return "Unbekannt"
    return "Unbekannt"


def _zp_art(val):
    try:
        v=int(str(val).split("-")[-1].strip())
        if v in [602,156]: return "Gigaliner"
        if v in [350,620]: return "Tandem"
        if v in [520,266,458,548,541,542,543,558]: return "Gliederzug"
    except Exception: pass
    return "Unbekannt"


def _zp_verdienst(lkw1,lkw2):
    total=0
    for v in [lkw1,lkw2]:
        try:
            if v is None: continue
            vi=int(str(v).split("-")[-1].strip())
            if vi in [602,156]: total+=40
            elif vi in [350,620,520,266,458,548,541,542,543,558]: total+=20
        except Exception: pass
    return total


def parse_zulage_excel(dateien: list) -> str:
    import json as _j
    from io import BytesIO
    import datetime as _dt
    sonder_map={}; fuengers_map={}
    cur_year=_dt.datetime.now().year
    for datei in dateien:
        try:
            datei.seek(0); df_h=pd.read_excel(BytesIO(datei.read()),sheet_name="Touren",header=0)
            datei.seek(0); df_nh=pd.read_excel(BytesIO(datei.read()),sheet_name="Touren",header=None)
            df_nh=df_nh.iloc[4:].reset_index(drop=True); df_nh.columns=range(df_nh.shape[1])
        except Exception: continue
        try:
            mask=df_h.iloc[:,13].astype(str).str.strip().str.upper()=="AZ"
            for _,row in df_h[mask].iterrows():
                datum=pd.to_datetime(row.iloc[14],errors="coerce")
                if pd.isna(datum) or datum.year < 2026: continue
                mk=f"{datum.month:02d}-{datum.year}"; ml=f"{_ZP_MONTHS_DE[datum.month]} {datum.year}"
                kw=int(datum.strftime("%W"))+1
                ds=f"{_ZP_DAYS[datum.weekday()]}, {datum.strftime('%d.%m.%Y')} (KW{kw})"
                def _c(x): return str(x).replace(" "," ").strip() if pd.notnull(x) else ""
                nn=_c(row.iloc[3] if len(row)>3 else ""); vn=_c(row.iloc[4] if len(row)>4 else "")
                if not nn or not vn or nn in("0","nan") or vn in("0","nan"):
                    nn=_c(row.iloc[6] if len(row)>6 else ""); vn=_c(row.iloc[7] if len(row)>7 else "")
                if not nn or not vn or nn in("0","nan") or vn in("0","nan"): continue
                lkw1=row.iloc[10] if len(row)>10 and pd.notnull(row.iloc[10]) else None
                lkw2=row.iloc[11] if len(row)>11 and pd.notnull(row.iloc[11]) else None
                tour=str(row.iloc[0]).strip() if pd.notnull(row.iloc[0]) else ""
                verd=_zp_verdienst(lkw1,lkw2)
                if verd==0: continue
                if mk not in sonder_map: sonder_map[mk]={"label":ml,"fahrer":{}}
                sonder_map[mk]["fahrer"].setdefault((nn,vn),[]).append(
                    {"datum":ds,"tour":tour,"lkw":str(lkw2 or lkw1 or "").strip(),"art":_zp_art(lkw2 or lkw1),"verdienst":verd})
        except Exception: pass
        try:
            for _,row in df_nh.iterrows():
                komm=str(row[15]).strip() if 15 in row and pd.notnull(row[15]) else ""
                if "füngers" not in komm.lower() and "fuengers" not in komm.lower(): continue
                nn=str(row[3]).replace(" "," ").strip() if 3 in row and pd.notnull(row[3]) else ""
                vn=str(row[4]).replace(" "," ").strip() if 4 in row and pd.notnull(row[4]) else ""
                if not nn or not vn or nn in("0","nan") or vn in("0","nan"): continue
                datum=pd.to_datetime(row[14] if 14 in row else None,errors="coerce")
                if pd.isna(datum) or datum.year < 2026: continue
                mk=f"{datum.month:02d}-{datum.year}"; ml=f"{_ZP_MONTHS_DE[datum.month]} {datum.year}"
                kw=datum.isocalendar()[1]; ds=datum.strftime("%d.%m.%Y")+f" (KW {kw})"
                if mk not in fuengers_map: fuengers_map[mk]={"label":ml,"fahrer":{}}
                fuengers_map[mk]["fahrer"].setdefault((nn,vn),[]).append(
                    {"datum":ds,"kommentar":komm,"verdienst":20})
        except Exception: pass
    def _build(dm):
        res=[]
        for mk in sorted(dm.keys()):
            e=dm[mk]; fl=[]
            for (nn,vn),tage in sorted(e["fahrer"].items()):
                fl.append({"name":f"{vn} {nn}","persnr":_zp_persnr(nn,vn),
                           "gesamt":sum(t["verdienst"] for t in tage),"tage":tage})
            fl.sort(key=lambda x:x["name"]); res.append({"monat":e["label"],"fahrer":fl})
        return res
    return _j.dumps({"sonder":_build(sonder_map),"fuengers":_build(fuengers_map)},ensure_ascii=False)



DRITTKUNDEN_KEYWORDS = [
    "ahaus", "borkholzhausen", "glandorf", "optifair", "opti fair",
    "edv", "edv fleisch", "elfering", "elfering ahaus"
]


def _dk_check(comment):
    if isinstance(comment, str):
        c = comment.lower()
        return any(k in c for k in DRITTKUNDEN_KEYWORDS)
    return False


def parse_drittkunden_excel(dateien: list) -> str:
    """Liest Touren-Excels auf Drittkunden-Zulage (Ahaus etc.)."""
    import json as _j
    from io import BytesIO
    import datetime as _dt
    entries = []
    _months_de = ["","Januar","Februar","März","April","Mai","Juni",
                  "Juli","August","September","Oktober","November","Dezember"]
    for datei in dateien:
        try:
            datei.seek(0)
            df = pd.read_excel(BytesIO(datei.read()), sheet_name=0, header=None)
            df = df.iloc[4:].reset_index(drop=True)
            df.columns = range(df.shape[1])
        except Exception:
            continue
        for _, row in df.iterrows():
            kommentar = str(row[15]).strip() if 15 in row and pd.notna(row[15]) else ""
            if not _dk_check(kommentar):
                continue
            datum = pd.to_datetime(row[14] if 14 in row else None, errors="coerce")
            if pd.isna(datum) or datum.year < 2026:
                continue
            lkw = str(row[11]).strip() if 11 in row and pd.notna(row[11]) else ""
            info = str(row[1]).strip() if 1 in row and pd.notna(row[1]) else ""
            kw = f"KW{datum.isocalendar()[1]}"
            monat_label = f"{_months_de[datum.month]} {datum.year}"
            datum_str = datum.strftime("%d.%m.%Y")
            # Two driver pairs: D/E (3/4) and G/H (6/7)
            fahrer_paare = []
            if len(row) > 4: fahrer_paare.append((row[3], row[4]))
            if len(row) > 7: fahrer_paare.append((row[6], row[7]))
            seen = set()
            for nn_raw, vn_raw in fahrer_paare:
                if pd.isna(nn_raw) and pd.isna(vn_raw):
                    continue
                nn = str(nn_raw).strip() if pd.notna(nn_raw) else ""
                vn = str(vn_raw).strip() if pd.notna(vn_raw) else ""
                if not nn:
                    continue
                name = f"{nn}, {vn}".strip().rstrip(",")
                key = name.lower()
                if key in seen:
                    continue
                seen.add(key)
                zulage = 0 if "zippel" in nn.lower() else 20
                persnr = _zp_persnr(nn, vn)
                entries.append({
                    "name": name,
                    "persnr": persnr,
                    "datum": datum_str,
                    "kw": kw,
                    "lkw": lkw,
                    "zulage": zulage,
                    "info": info,
                    "monat": monat_label,
                    "monat_sort": f"{datum.year}{datum.month:02d}",
                })
    # Group by monat
    monat_map = {}
    for e in entries:
        mk = e["monat_sort"]
        monat_map.setdefault(mk, {"monat": e["monat"], "eintraege": []})
        monat_map[mk]["eintraege"].append(e)
    # Per month: group by name
    result = []
    for mk in sorted(monat_map.keys()):
        m = monat_map[mk]
        name_map = {}
        for e in m["eintraege"]:
            name_map.setdefault(e["name"], []).append(e)
        fahrer = []
        for name in sorted(name_map.keys()):
            tage = sorted(name_map[name], key=lambda x: x["datum"])
            gesamt = sum(t["zulage"] for t in tage)
            fahrer.append({"name": name, "persnr": tage[0].get("persnr", "Unbekannt") if tage else "Unbekannt", "gesamt": gesamt, "tage": tage})
        result.append({"monat": m["monat"], "fahrer": fahrer})
    return _j.dumps(result, ensure_ascii=False)


def generate_drittkunden_excel(drittkunden_json_str: str) -> bytes:
    """Formatierte Excel für Drittkunden-Zulage."""
    import io as _io, json as _j
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    try:
        months = _j.loads(drittkunden_json_str)
    except Exception:
        return None
    if not months:
        return None
    wb = Workbook(); wb.remove(wb.active)
    def _s(st="thin",c="CCCCCC"): return Side(style=st, color=c)
    thin = Border(left=_s(),right=_s(),top=_s(),bottom=_s())
    med  = Border(left=_s("medium","1F4E78"),right=_s("medium","1F4E78"),top=_s("medium","1F4E78"),bottom=_s("medium","1F4E78"))
    gmed = Border(left=_s("medium","70AD47"),right=_s("medium","70AD47"),top=_s("medium","70AD47"),bottom=_s("medium","70AD47"))
    FT=PatternFill("solid",fgColor="1F4E78"); FH=PatternFill("solid",fgColor="D9E2F3")
    FN=PatternFill("solid",fgColor="4472C4"); FG=PatternFill("solid",fgColor="70AD47")
    FL=PatternFill("solid",fgColor="1F4E78"); FA=PatternFill("solid",fgColor="F8F9FA")
    FW=PatternFill("solid",fgColor="FFFFFF")
    hdrs   = ["Name","Datum","KW","LKW","Zulage","Info"]
    widths = [26,14,8,14,12,30]
    nc = len(hdrs)
    # Summary sheet first
    summary = {}
    for monat in months:
        for f in monat["fahrer"]:
            summary[f["name"]] = summary.get(f["name"], 0) + f["gesamt"]
    ws_sum = wb.create_sheet(title="Zusammenfassung", index=0)
    ws_sum.append([f"Zusammenfassung – Drittkunden","",""])
    tr=ws_sum.max_row; ws_sum.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=3)
    c=ws_sum.cell(tr,1); c.font=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
    c.fill=FT; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=med
    ws_sum.row_dimensions[tr].height=30
    ws_sum.append(["Name","Personalnummer","Gesamt"]); hr=ws_sum.max_row
    for col in range(1,4):
        c=ws_sum.cell(hr,col); c.font=Font(name="Calibri",bold=True,size=11,color="1F4E78")
        c.fill=FH; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin
    ws_sum.row_dimensions[hr].height=24
    total_all=0
    for ri,(name,total) in enumerate(sorted(summary.items())):
        fill=FA if ri%2==0 else FW; ws_sum.append([name,"",total]); r=ws_sum.max_row; total_all+=total
        for col in range(1,4):
            c=ws_sum.cell(r,col); iv=(col==3)
            c.font=Font(name="Calibri",size=11,color="16A34A" if iv else "2C3E50",bold=iv)
            c.fill=fill; c.border=thin
            c.alignment=Alignment(horizontal="right" if iv else "left",vertical="center")
            if iv: c.number_format='#,##0.00 "€"'
        ws_sum.row_dimensions[r].height=21
    ws_sum.append(["","Gesamtsumme",total_all]); gr=ws_sum.max_row
    for col in range(1,4):
        c=ws_sum.cell(gr,col); c.font=Font(name="Calibri",bold=True,size=13,color="FFFFFF")
        c.fill=FL; c.border=med; c.alignment=Alignment(horizontal="right",vertical="center")
        if col==3: c.number_format='#,##0.00 "€"'
    ws_sum.row_dimensions[gr].height=28
    ws_sum.column_dimensions["A"].width=26; ws_sum.column_dimensions["B"].width=18; ws_sum.column_dimensions["C"].width=16
    # Month sheets
    for monat in months:
        ws = wb.create_sheet(title=monat["monat"][:31])
        msum = sum(f["gesamt"] for f in monat["fahrer"])
        ws.append([monat["monat"]]+[""]*( nc-1)); tr=ws.max_row
        ws.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=nc)
        c=ws.cell(tr,1); c.font=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
        c.fill=FT; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=med
        ws.row_dimensions[tr].height=28
        ws.append(hdrs); hr=ws.max_row
        for col in range(1,nc+1):
            c=ws.cell(hr,col); c.font=Font(name="Calibri",bold=True,size=10,color="1F4E78")
            c.fill=FH; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin
        ws.row_dimensions[hr].height=22
        alt=False
        for fahrer in monat["fahrer"]:
            r0=ws.max_row+1
            for ti,tag in enumerate(fahrer["tage"]):
                row=[fahrer["name"] if ti==0 else "",tag["datum"],tag["kw"],tag["lkw"],tag["zulage"],tag.get("info","")]
                ws.append(row); r=ws.max_row; fill=FA if alt else FW
                for col in range(1,nc+1):
                    if col==1 and ti==0: continue
                    c=ws.cell(r,col); iv=(col==5)
                    c.font=Font(name="Calibri",size=10,color="16A34A" if iv else "2C3E50",bold=iv)
                    c.fill=fill; c.border=thin
                    c.alignment=Alignment(horizontal="right" if iv else "left",vertical="center")
                    if iv and isinstance(c.value,(int,float)): c.number_format='#,##0.00 "€"'
                ws.row_dimensions[r].height=20; alt=not alt
            ws.merge_cells(start_row=r0,start_column=1,end_row=ws.max_row,end_column=1)
            nc2=ws.cell(r0,1); nc2.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
            nc2.fill=FN; nc2.alignment=Alignment(horizontal="left",vertical="center"); nc2.border=med
            gv=[""]*nc; gv[nc-2]="Gesamt"; gv[nc-1]=fahrer["gesamt"]
            ws.append(gv); gr=ws.max_row
            ws.merge_cells(start_row=gr,start_column=1,end_row=gr,end_column=nc-1)
            for col in range(1,nc+1):
                c=ws.cell(gr,col); c.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
                c.fill=FG; c.alignment=Alignment(horizontal="right",vertical="center"); c.border=gmed
                if col==nc and isinstance(c.value,(int,float)): c.number_format='#,##0.00 "€"'
            ws.row_dimensions[gr].height=20; alt=False
            ws.append([""]*nc); ws.row_dimensions[ws.max_row].height=6
        tv=[""]*nc; tv[nc-2]="Monatsgesamt"; tv[nc-1]=msum
        ws.append(tv); tr2=ws.max_row
        ws.merge_cells(start_row=tr2,start_column=1,end_row=tr2,end_column=nc-1)
        for col in range(1,nc+1):
            c=ws.cell(tr2,col); c.font=Font(name="Calibri",bold=True,size=13,color="FFFFFF")
            c.fill=FL; c.alignment=Alignment(horizontal="right",vertical="center"); c.border=med
            if col==nc and isinstance(c.value,(int,float)): c.number_format='#,##0.00 "€"'
        ws.row_dimensions[tr2].height=26
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes="A3"
    buf=_io.BytesIO(); wb.save(buf); return buf.getvalue()


def parse_telefon_excel(up) -> str:
    """Liest Telefonnummern.xlsx (Sheet 'aktuell') und gibt JSON-String zurück."""
    import json as _json
    try:
        df = pd.read_excel(up, sheet_name="aktuell", dtype=str, engine=EXCEL_READ_ENGINE)
        df.columns = ["name","vorname","vorwahl","nummer","mail","gruppe"]
        df = df.fillna("")
        groups, current_group, current_entries = [], "Eigene Fahrer", []
        # itertuples + Positions-Zugriff statt iterrows mit r["..."] —
        # bei langen Telefonlisten klar schneller, gleiche Logik.
        for tup in df.itertuples(index=False, name=None):
            name    = tup[0].strip()
            vorname = tup[1].strip()
            vorwahl = tup[2].strip()
            nummer  = tup[3].strip()
            mail    = tup[4].strip()
            gruppe  = tup[5].strip()
            if not name and not vorname and not vorwahl and not nummer:
                if current_entries:
                    groups.append({"gruppe": current_group, "personen": current_entries})
                current_entries = []
                continue
            if gruppe:
                if current_entries:
                    groups.append({"gruppe": current_group, "personen": current_entries})
                    current_entries = []
                current_group = gruppe
            tel = ""
            if vorwahl and vorwahl.lower() not in ("nan","n.a.",""):
                tel = vorwahl.strip()
                if nummer and nummer.lower() not in ("nan","n.a.",""):
                    tel += " " + nummer.strip()
            elif nummer and nummer.lower() not in ("nan","n.a.",""):
                tel = nummer.strip()
            else:
                tel = "n.a."
            vname = " ".join(filter(None, [vorname, name]))
            if not vname.strip(): continue
            current_entries.append({
                "name": vname,
                "tel":  tel,
                "mail": mail if mail.lower() not in ("nan","") else ""
            })
        if current_entries:
            groups.append({"gruppe": current_group, "personen": current_entries})
        return _json.dumps(groups, ensure_ascii=False, separators=(",", ":"))
    except Exception as e:
        st.warning(f"Telefonliste konnte nicht gelesen werden: {e}")
        return "[]"




def parse_spesen_csv(uploaded_file) -> str:
    """Parst die Reisekosten-/Spesen-CSV und aggregiert nur Steuerfreiheit + steuerpflichtigen Betrag pro Fahrer."""
    import csv as _csv
    from io import StringIO as _SIO

    empty = json.dumps({"drivers": [], "months": [], "total_cost": 0, "total_tax_free": 0, "total_taxable": 0, "total_rows": 0}, ensure_ascii=False)
    payload = read_upload_bytes(uploaded_file)
    if not payload:
        return empty

    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = payload.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        return empty
    if text.startswith("\ufeff"):
        text = text[1:]

    def _clean(v):
        s = "" if v is None else str(v).strip()
        return "" if s.lower() in ("nan", "none", "null") else s

    def _money(v) -> float:
        s = _clean(v)
        if not s:
            return 0.0
        s = s.replace("€", "").replace(" ", "")
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        try:
            return round(float(s), 2)
        except Exception:
            return 0.0

    def _duration(v) -> int:
        s = _clean(v)
        if not s:
            return 0
        if ":" in s:
            parts = s.split(":")
            try:
                h = int(float(parts[0] or 0))
                m = int(float(parts[1] or 0)) if len(parts) > 1 else 0
                return h * 60 + m
            except Exception:
                return 0
        try:
            return int(round(float(s.replace(".", "").replace(",", "."))))
        except Exception:
            return 0

    def _dt(v):
        s = _clean(v).strip('"')
        if not s:
            return None
        s19 = s[:19]
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
            try:
                return datetime.datetime.strptime(s19, fmt)
            except Exception:
                pass
        return None

    def _date_from_value(v):
        s = _clean(v).strip('"')
        if not s:
            return None
        s16 = s[:16]
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"):
            try:
                return datetime.datetime.strptime(s16, fmt).date()
            except Exception:
                pass
        return None

    def _time_label(dt):
        return dt.strftime("%H:%M") if dt else ""

    def _duration_label(minutes):
        minutes = int(minutes or 0)
        if minutes <= 0:
            return "0 Minuten"
        h, m = divmod(minutes, 60)
        if h and m:
            return f"{h} Std. {m:02d} Min."
        if h:
            return f"{h} Std."
        return f"{m} Min."

    def _month_label(value: str) -> str:
        names = ["", "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
        try:
            y, m = value.split("-", 1)
            return f"{names[int(m)]} {y}"
        except Exception:
            return value

    def _weekday(d):
        names = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
        try:
            return names[d.weekday()]
        except Exception:
            return ""

    try:
        reader = _csv.DictReader(_SIO(text), delimiter=";", quotechar='"')
    except Exception as e:
        st.warning(f"Spesen-CSV konnte nicht gelesen werden: {e}")
        return empty

    by_driver = {}
    total_rows = 0
    months_seen = set()

    for row in reader:
        total_rows += 1
        norm = {(k or "").strip().lstrip("\ufeff").upper(): v for k, v in row.items()}
        name = _clean(norm.get("DRIVER"))
        if not name:
            continue

        start_dt = _dt(norm.get("START"))
        end_dt = _dt(norm.get("END"))
        date_obj = _date_from_value(norm.get("DATE"))
        if date_obj is None and start_dt is not None:
            date_obj = start_dt.date()
        if date_obj is None:
            continue

        date_iso = date_obj.strftime("%Y-%m-%d")
        month = date_obj.strftime("%Y-%m")
        months_seen.add(month)
        duration_minutes = _duration(norm.get("DURATION"))
        # Nur diese beiden Spalten werden fuer die Spesenberechnung verwendet:
        # Steuerfreiheit -> TAX_FREE_AMOUNT
        # Steuerpflichtiger Betrag -> TAXABLE_AMOUNT
        tax_free = _money(norm.get("TAX_FREE_AMOUNT"))
        taxable = _money(norm.get("TAXABLE_AMOUNT"))
        total = round(tax_free + taxable, 2)

        entry = {
            "date_iso": date_iso,
            "date_sort": date_iso + " " + (_time_label(start_dt) or "00:00"),
            "date_label": date_obj.strftime("%d.%m.%Y"),
            "weekday": _weekday(date_obj),
            "month": month,
            "start": _clean(norm.get("START")),
            "end": _clean(norm.get("END")),
            "start_time": _time_label(start_dt),
            "end_time": _time_label(end_dt),
            "duration_minutes": duration_minutes,
            "duration_label": _duration_label(duration_minutes),
            "vehicles": _clean(norm.get("VEHICLES")),
            "region": _clean(norm.get("REGION")),
            "tax_free": tax_free,
            "taxable": taxable,
            # Legacy-Aliase fuer vorhandene Frontend-Funktionen:
            # meal = steuerfrei, night = steuerpflichtig
            "meal": tax_free,
            "night": taxable,
            "total": total,
            "pos_start": _clean(norm.get("POS_START")),
            "pos_end": _clean(norm.get("POS_END")),
            "absence": _clean(norm.get("ABSENCE")),
            "employee_nr": _clean(norm.get("EMPLOYEE_NR")),
        }

        if name not in by_driver:
            by_driver[name] = {
                "name": name,
                "employee_nr": _clean(norm.get("EMPLOYEE_NR")),
                "rows": [],
                "tax_free": 0.0,
                "taxable": 0.0,
                # Legacy-Aliase fuer vorhandene Frontend-Funktionen:
                "meal": 0.0,
                "night": 0.0,
                "gesamt": 0.0,
                "duration_minutes": 0,
                "fahrtage": 0,
            }
        d = by_driver[name]
        if not d.get("employee_nr") and entry.get("employee_nr"):
            d["employee_nr"] = entry["employee_nr"]
        d["rows"].append(entry)
        d["tax_free"] = round(d["tax_free"] + tax_free, 2)
        d["taxable"] = round(d["taxable"] + taxable, 2)
        d["meal"] = d["tax_free"]      # Legacy-Alias: steuerfrei
        d["night"] = d["taxable"]      # Legacy-Alias: steuerpflichtig
        d["gesamt"] = round(d["gesamt"] + total, 2)
        d["duration_minutes"] += duration_minutes
        if duration_minutes > 0 or entry.get("start_time") or entry.get("end_time") or entry.get("vehicles"):
            d["fahrtage"] += 1

    drivers = []
    for d in by_driver.values():
        d["rows"].sort(key=lambda r: r.get("date_sort", ""))
        drivers.append(d)
    drivers.sort(key=lambda d: (-float(d.get("gesamt", 0)), d.get("name", "")))

    months = [{"value": m, "label": _month_label(m)} for m in sorted(months_seen, reverse=True)]
    total_tax_free = round(sum(float(d.get("tax_free", 0)) for d in drivers), 2)
    total_taxable = round(sum(float(d.get("taxable", 0)) for d in drivers), 2)
    total_cost = round(total_tax_free + total_taxable, 2)

    return json.dumps({
        "drivers": drivers,
        "months": months,
        "total_cost": total_cost,
        "total_tax_free": total_tax_free,
        "total_taxable": total_taxable,
        # Legacy-Aliase fuer vorhandene Frontend-Funktionen:
        "total_meal": total_tax_free,
        "total_night": total_taxable,
        "total_rows": total_rows,
    }, ensure_ascii=False)


def parse_fahrerbewertung_json(uploaded_file) -> str:
    """Liest die Fahrerbewertungs-Rohdaten (d_rohdaten.json) und verdichtet sie zu einer
    kompakten Zusammenfassung pro Fahrer.

    Die Roh-JSON ist sehr groß (zehntausende Einzel-Events). Damit das erzeugte
    suche.html nicht aufgebläht wird, werden hier NUR Aggregate eingebettet:
    Noten, Ereigniszähler je Art und Monat sowie Kennzahlen. Die Einzel-Events
    selbst werden NICHT übernommen.
    """
    import json as _json

    try:
        raw = uploaded_file.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8-sig", errors="replace")
        data = _json.loads(raw)
    except Exception:
        return _json.dumps({"profile": "", "event_types": [], "g_months": {},
                            "g_ev": {}, "drivers": []}, ensure_ascii=False)

    drivers = data.get("drivers", []) if isinstance(data, dict) else []
    EVT = ["BRAKE", "CURVE", "OVERSPEED", "SPEEDUP"]
    out_drivers = []
    g_months: dict = {}
    g_ev = {t: 0 for t in EVT}
    profile = ""

    def _r1(x):
        return round(x, 1) if isinstance(x, (int, float)) else x

    for d in drivers:
        if not isinstance(d, dict):
            continue
        g = d.get("grades") or {}
        rp = d.get("rating_profiles") or []
        if rp and not profile:
            profile = rp[0]
        sc = g.get("sub_critical_grades") or {}
        se = g.get("sub_economic_grades") or {}
        sd = g.get("sub_difficulty_grades") or {}

        ev = {t: 0 for t in EVT}
        months: dict = {}
        for e in (d.get("critical_events") or []):
            if not isinstance(e, dict):
                continue
            t = e.get("type")
            if t in ev:
                ev[t] += 1
            st = str(e.get("start_time") or "")
            mk = st[:7] if len(st) >= 7 and st[4] == "-" else None
            if mk:
                mm = months.setdefault(mk, {"_t": 0})
                if t:
                    mm[t] = mm.get(t, 0) + 1
                mm["_t"] += 1
                gm = g_months.setdefault(mk, {"_t": 0})
                if t:
                    gm[t] = gm.get(t, 0) + 1
                gm["_t"] += 1
        for t in EVT:
            g_ev[t] += ev[t]

        fuel = None
        for fi in (d.get("fms_info") or []):
            if isinstance(fi, dict) and fi.get("vehicle") == "__ALL__":
                fuel = fi.get("avg_used_fuel")
                break

        dt = d.get("driving_time") or 0
        subs = {k: v for k, v in {
            "overspeed": sc.get("overspeed"), "brake": sc.get("brake"),
            "speedup": sc.get("speedup"), "curves": sc.get("curves"),
            "foresight": sc.get("foresight_driving"),
            "idle": se.get("idle"), "avg_speed": se.get("avg_speed"),
            "wearfree_brake": se.get("wearfree_brake"), "cruise": se.get("cruisecontrol"),
            "altitude": sd.get("altitude"), "street_type": sd.get("street_type"),
            "stopps": sd.get("count_stopps"),
        }.items() if v is not None}

        out_drivers.append({
            "name": d.get("name", ""),
            "persnr": str(d.get("employee_number", "") or ""),
            "grade": g.get("main_grade"),
            "g_crit": g.get("main_critical_grade"),
            "g_eco": g.get("main_economic_grade"),
            "g_diff": g.get("main_difficulty_grade"),
            "subs": subs,
            "dist": d.get("distance", 0) or 0,
            "hours": round(dt / 3600.0, 1) if dt else 0,
            "trips": d.get("count_evaluated_trips", 0) or 0,
            "fuel": _r1(fuel),
            "ev": ev,
            "evt": sum(ev.values()),
            "brake": d.get("total_brake_count", 0) or 0,
            "curve": d.get("total_curve_count", 0) or 0,
            "speedup": d.get("total_speedup_count", 0) or 0,
            "months": months,
            "vehicles": d.get("vehicles") or [],
        })

    out_drivers.sort(key=lambda r: (r.get("name") or "").lower())
    return _json.dumps({
        "profile": profile,
        "event_types": EVT,
        "g_months": g_months,
        "g_ev": g_ev,
        "drivers": out_drivers,
    }, ensure_ascii=False, separators=(",", ":"))


def parse_verstoss_csv(uploaded_file) -> str:
    """Parst die Digitacho-Verstoßauswertungs-CSV und aggregiert Verstöße und Bußgelder pro Fahrer.

    Die Bußgeld-Spalten werden robust gelesen. Akzeptiert werden zum Beispiel
    160, 160 €, 160,00, 1.520 € oder 1.520,50 €.
    """
    import csv as _csv
    from io import StringIO as _SIO

    empty = json.dumps({
        "drivers": [],
        "total_violations": 0,
        "total_driver_penalty": 0,
        "total_company_penalty": 0,
        "first_violation_month": "",
        "last_violation_month": "",
    }, ensure_ascii=False)

    payload = read_upload_bytes(uploaded_file)
    if not payload:
        return empty

    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = payload.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        return empty

    if text.startswith("\ufeff"):
        text = text[1:]

    try:
        reader = _csv.DictReader(_SIO(text), delimiter=";", quotechar='"')
    except Exception as e:
        st.warning(f"Verstoß-CSV konnte nicht gelesen werden: {e}")
        return empty

    def _clean(v):
        return str(v or "").replace("\xa0", " ").strip()

    def _norm_key(v: str) -> str:
        s = _clean(v).lower().replace("ß", "ss")
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        return re.sub(r"[^a-z0-9]+", "", s)

    def _row_get(row_norm: dict, *aliases: str) -> str:
        for alias in aliases:
            key = _norm_key(alias)
            if key in row_norm:
                return row_norm.get(key)
        return ""

    def _to_int(v):
        """Robust fuer Minuten/Werte: '48', '48 Min.', '1 Std. 15 Min.' -> Minuten."""
        s = _clean(v)
        if not s or s in ("—", "-", "–"):
            return 0
        h = re.search(r"(\d+)\s*(?:std|stunde|stunden|h)\b", s, flags=re.I)
        m = re.search(r"(\d+)\s*(?:min|minute|minuten|m)\b", s, flags=re.I)
        if h or m:
            return (int(h.group(1)) * 60 if h else 0) + (int(m.group(1)) if m else 0)
        m = re.search(r"-?\d+", s.replace(".", ""))
        return int(m.group(0)) if m else 0

    def _to_money(v):
        """Robust fuer Euro-Werte: '160', '160 €', '160,00', '1.520 €'."""
        s = _clean(v)
        if not s or s in ("—", "-", "–"):
            return 0
        s = re.sub(r"[^0-9,\.\-]", "", s)
        if not s or s in ("-", ",", "."):
            return 0
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            if s.count(".") > 1:
                s = s.replace(".", "")
            elif re.match(r"^-?\d{1,3}\.\d{3}$", s):
                s = s.replace(".", "")
        try:
            value = float(s)
        except Exception:
            return 0
        return int(value) if value.is_integer() else round(value, 2)

    by_driver = {}
    date_re = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})(?:\s+(\d{1,2}):(\d{2}))?")

    for row in reader:
        row_norm = {_norm_key(k): v for k, v in (row or {}).items()}

        name = _clean(_row_get(row_norm, "DRIVER", "Fahrer", "Name", "Driver Name"))
        if not name:
            continue

        start = _clean(_row_get(row_norm, "START_DATE", "Start Date", "Beginn", "Start", "Von"))
        end = _clean(_row_get(row_norm, "END_DATE", "End Date", "Ende", "Bis"))
        target = _to_int(_row_get(row_norm, "TARGET", "Soll", "Grenzwert", "Limit"))
        ist = _to_int(_row_get(row_norm, "IS", "Ist", "Tatsaechlich", "Tatsächlich"))
        diff = _to_int(_row_get(row_norm, "DIFF", "Differenz", "Ueberschreitung", "Überschreitung"))
        viol = _clean(_row_get(row_norm, "VIOLATION", "Verstoß", "Verstoss", "Verstoßart", "Verstossart"))
        law = _clean(_row_get(row_norm, "LAW", "Gesetz", "Rechtsgrundlage", "Paragraph"))
        dp = _to_money(_row_get(
            row_norm,
            "DRIVER_PENALTY", "Driver Penalty", "Bußgeld Fahrer", "Bussgeld Fahrer",
            "Fahrer Bußgeld", "Fahrer Bussgeld", "Bußgeld Fahrer EUR", "Bussgeld Fahrer EUR"
        ))
        cp = _to_money(_row_get(
            row_norm,
            "COMPANY_PENALTY", "Company Penalty", "Bußgeld Firma", "Bussgeld Firma",
            "Firma Bußgeld", "Firma Bussgeld", "Unternehmen Bußgeld", "Unternehmen Bussgeld",
            "Pauschalbetrag", "Bußgeld", "Bussgeld", "Company Fine"
        ))
        idate = _clean(_row_get(row_norm, "INSTRUCTION_DATE", "Instruction Date", "Unterweisungsdatum", "Belehrungsdatum"))
        iby = _clean(_row_get(row_norm, "INSTRUCTION_BY", "Instruction By", "Unterwiesen durch", "Belehrt durch"))
        remark = _clean(_row_get(row_norm, "REMARK_TEXT", "Remark Text", "Bemerkung", "Kommentar", "Notiz"))

        date_sort = ""
        m = date_re.match(start)
        if m:
            dd, mm, yyyy = f"{int(m.group(1)):02d}", f"{int(m.group(2)):02d}", m.group(3)
            hh = f"{int(m.group(4) or 0):02d}"
            mi = f"{int(m.group(5) or 0):02d}"
            date_sort = f"{yyyy}-{mm}-{dd} {hh}:{mi}"

        instructed = bool(idate)

        entry = {
            "start": start,
            "end": end,
            "date_sort": date_sort,
            "target": target,
            "ist": ist,
            "diff": diff,
            "violation": viol,
            "law": law,
            "driver_penalty": dp,
            "company_penalty": cp,
            "instruction_date": idate,
            "instruction_by": iby,
            "remark": remark,
            "instructed": instructed,
        }

        if name not in by_driver:
            by_driver[name] = {
                "name": name,
                "verstoesse": [],
                "count": 0,
                "count_instructed": 0,
                "count_open": 0,
                "sum_driver_penalty": 0,
                "sum_company_penalty": 0,
                "sum_diff": 0,
                "types": {},
            }

        d = by_driver[name]
        d["verstoesse"].append(entry)
        d["count"] += 1
        d["sum_driver_penalty"] = round(d["sum_driver_penalty"] + dp, 2)
        d["sum_company_penalty"] = round(d["sum_company_penalty"] + cp, 2)
        d["sum_diff"] += diff
        if instructed:
            d["count_instructed"] += 1
        else:
            d["count_open"] += 1
        t = viol or "—"
        d["types"][t] = d["types"].get(t, 0) + 1

    drivers = []
    for d in by_driver.values():
        d["verstoesse"].sort(key=lambda x: x.get("date_sort", ""), reverse=True)
        d["types"] = sorted(d["types"].items(), key=lambda kv: (-kv[1], kv[0]))
        drivers.append(d)

    def _last_sort(d):
        return d["verstoesse"][0].get("date_sort", "") if d["verstoesse"] else ""
    drivers.sort(key=lambda x: (_last_sort(x), x["name"]), reverse=True)

    total = sum(d["count"] for d in drivers)
    total_driver_penalty = round(sum(float(d.get("sum_driver_penalty", 0) or 0) for d in drivers), 2)
    total_company_penalty = round(sum(float(d.get("sum_company_penalty", 0) or 0) for d in drivers), 2)

    _all_dates = [
        e.get("date_sort", "")[:10]
        for d in drivers
        for e in d.get("verstoesse", [])
        if e.get("date_sort")
    ]

    def _month_label(iso_date: str) -> str:
        try:
            yyyy, mm, _dd = iso_date.split("-", 2)
            return f"{mm}/{yyyy}"
        except Exception:
            return ""

    first_violation_month = _month_label(min(_all_dates)) if _all_dates else ""
    last_violation_month = _month_label(max(_all_dates)) if _all_dates else ""

    return json.dumps({
        "drivers": drivers,
        "total_violations": total,
        "total_driver_penalty": total_driver_penalty,
        "total_company_penalty": total_company_penalty,
        "first_violation_month": first_violation_month,
        "last_violation_month": last_violation_month,
    }, ensure_ascii=False)


def parse_tanken_excel(uploaded_files) -> str:
    """Liest mehrere monatliche Tank-Excel-Dateien speicherschonend ein.

    Einige Exporte enthalten bis zum Excel-Zeilenlimit formatierte Leerzeilen.
    Deshalb wird im Read-only-Modus nach 100 aufeinanderfolgenden Leerzeilen
    hinter den echten Daten abgebrochen.
    """
    import openpyxl as _opxl
    from openpyxl.utils.datetime import from_excel as _from_excel

    def _norm(value) -> str:
        value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
        return re.sub(r"[^a-z0-9]+", "", value.strip().lower())

    def _clean(value) -> str:
        if value is None:
            return ""
        s = str(value).replace("\xa0", " ").strip()
        if s.lower() in {"nan", "none"}:
            return ""
        return re.sub(r"\s+", " ", s).replace("_", " ").strip()

    def _number(value) -> float:
        if value in (None, ""):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = _clean(value).replace("l", "").replace("km", "").replace(" ", "")
        if not s:
            return 0.0
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    def _date(value, epoch):
        dt = None
        if isinstance(value, datetime.datetime):
            dt = value
        elif isinstance(value, datetime.date):
            dt = datetime.datetime.combine(value, datetime.time())
        elif isinstance(value, (int, float)) and 20000 <= float(value) <= 80000:
            try:
                dt = _from_excel(float(value), epoch)
            except Exception:
                dt = None
        if dt is None:
            s = _clean(value)
            for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
                try:
                    dt = datetime.datetime.strptime(s[:10], fmt)
                    break
                except Exception:
                    pass
        if dt is None:
            return "", "", 0, 0
        return dt.strftime("%d.%m.%Y"), dt.strftime("%Y-%m-%d"), dt.year, dt.month

    def _time(value):
        if isinstance(value, datetime.datetime):
            return value.strftime("%H:%M:%S")
        if isinstance(value, datetime.time):
            return value.strftime("%H:%M:%S")
        if isinstance(value, (int, float)) and 0 <= float(value) < 1:
            seconds = int(round(float(value) * 86400)) % 86400
            return f"{seconds // 3600:02d}:{(seconds % 3600) // 60:02d}:{seconds % 60:02d}"
        s = _clean(value)
        m = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
        if m:
            return f"{int(m.group(1)):02d}:{m.group(2)}:{m.group(3) or '00'}"
        return ""

    aliases = {
        "firma": {"fahrzeugkategorie", "firmaspedition", "firma", "spedition"},
        "datum": {"datumdertransaktion", "datum"},
        "uhrzeit": {"zeitpunktdertransaktion", "zeitpunkt", "uhrzeit", "zeit"},
        "fahrzeug_ia": {"fahrzeugia", "fahrzeug"},
        "fahrzeug": {"fahrzeugkfzkennzeichen", "kfzkennzeichen", "kennzeichen"},
        "fahrer": {"fahrer"},
        "produkt": {"produkt"},
        "menge_liter": {"menge", "mengeinliter", "liter", "tankmenge"},
        "kilometerzaehler": {"kilometerzahler", "kilometerstand", "kmstand"},
        "transaktions_typ": {"transaktionstyp", "typ"},
        "zapfsaeule": {"zapfsaule", "zapfsaeule"},
        "km": {"km", "gefahrenekm", "strecke"},
    }
    required = {"datum", "uhrzeit", "fahrzeug", "fahrer", "menge_liter"}
    rows = []
    seen = set()

    for uploaded_file in uploaded_files or []:
        payload = read_upload_bytes(uploaded_file)
        if not payload:
            continue
        try:
            wb = _opxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        except Exception:
            continue
        source_name = getattr(uploaded_file, "name", "") or ""
        try:
            for ws in wb.worksheets:
                mapping = None
                data_started = False
                empty_streak = 0
                iterator = ws.iter_rows(values_only=True)
                for row_no, values in enumerate(iterator, 1):
                    values = tuple(values or ())
                    if mapping is None:
                        if row_no > 25:
                            break
                        candidate = {}
                        for idx, value in enumerate(values):
                            key = _norm(value)
                            if not key:
                                continue
                            for target, names in aliases.items():
                                if key in names and target not in candidate:
                                    candidate[target] = idx
                        if required.issubset(candidate):
                            mapping = candidate
                        continue

                    if not any(value not in (None, "") for value in values):
                        empty_streak += 1
                        if data_started and empty_streak >= 100:
                            break
                        continue
                    empty_streak = 0

                    def get(key):
                        idx = mapping.get(key)
                        return values[idx] if idx is not None and idx < len(values) else None

                    datum, date_iso, jahr, monat = _date(get("datum"), wb.epoch)
                    uhrzeit = _time(get("uhrzeit"))
                    fahrzeug = _clean(get("fahrzeug"))
                    fahrzeug_ia = _clean(get("fahrzeug_ia"))
                    fahrer = _clean(get("fahrer"))
                    produkt = _clean(get("produkt"))
                    menge = _number(get("menge_liter"))
                    transaktions_typ = _clean(get("transaktions_typ"))
                    if transaktions_typ and "betank" not in transaktions_typ.casefold():
                        continue
                    if not datum or not fahrzeug or menge <= 0:
                        continue
                    data_started = True
                    kilometerzaehler = _number(get("kilometerzaehler"))
                    km = _number(get("km"))
                    item = {
                        "firma": _clean(get("firma")),
                        "datum": datum,
                        "date_iso": date_iso,
                        "uhrzeit": uhrzeit,
                        "datetime_iso": f"{date_iso} {uhrzeit or '00:00:00'}",
                        "jahr": jahr,
                        "monat": monat,
                        "fahrzeug_ia": fahrzeug_ia,
                        "fahrzeug": fahrzeug,
                        "fahrer": fahrer,
                        "produkt": produkt,
                        "menge_liter": round(menge, 3),
                        "kilometerzaehler": round(kilometerzaehler, 1),
                        "km": round(km, 1),
                        "transaktions_typ": transaktions_typ,
                        "zapfsaeule": _clean(get("zapfsaeule")),
                        "quelle": source_name,
                    }
                    dedup = (
                        item["date_iso"], item["uhrzeit"], item["fahrzeug"], item["fahrzeug_ia"],
                        item["fahrer"], item["produkt"], item["menge_liter"], item["kilometerzaehler"],
                    )
                    if dedup in seen:
                        continue
                    seen.add(dedup)
                    rows.append(item)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    rows.sort(key=lambda x: (x.get("datetime_iso", ""), x.get("fahrzeug", "")), reverse=True)
    return json.dumps(rows, ensure_ascii=False, separators=(",", ":"))


def parse_fahrzeugwaesche_excel(uploaded_files) -> str:
    """Verarbeitet mehrere Fahrzeugwäsche-Excel-Dateien zu JSON für die Übersicht."""
    def _norm_header(value: str) -> str:
        value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
        return re.sub(r"[^a-z0-9]+", "", value.strip().lower())

    def _clean_text(value) -> str:
        if pd.isna(value):
            return ""
        s = str(value).strip()
        if s.lower() == "nan":
            return ""
        s = re.sub(r"\s+", " ", s)
        return s.replace("_", " ").strip()

    def _parse_date(value):
        if pd.isna(value):
            return "", ""
        if isinstance(value, (datetime.datetime, datetime.date, pd.Timestamp)):
            dt = pd.Timestamp(value)
            return dt.strftime("%d.%m.%Y"), dt.strftime("%Y-%m-%d")
        dt = pd.to_datetime(value, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%d.%m.%Y"), dt.strftime("%Y-%m-%d")
        s = _clean_text(value)
        m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
        if m:
            d, mo, y = m.groups()
            return f"{int(d):02d}.{int(mo):02d}.{y}", f"{y}-{int(mo):02d}-{int(d):02d}"
        return s, ""

    def _normalize_kennzeichen(value: str) -> str:
        """Normalisiert Spalte E (Fahrzeug-Kennzeichen) auf 'LWL - E XXX'.

        Extrahiert die erste zusammenhängende Ziffernfolge aus dem Eingabewert
        und baut damit das einheitliche Zielformat. Enthält der Wert keine
        Ziffern, wird er unverändert zurückgegeben (damit Freitext nicht verloren geht).
        """
        s = _clean_text(value)
        if not s:
            return ""
        m = re.search(r"\d+", s)
        if not m:
            return s
        return f"LWL - E {m.group(0)}"

    def _parse_time(value):
        if pd.isna(value):
            return ""
        if isinstance(value, datetime.datetime):
            return value.strftime("%H:%M:%S")
        if isinstance(value, datetime.time):
            return value.strftime("%H:%M:%S")
        if isinstance(value, pd.Timestamp):
            return value.strftime("%H:%M:%S")
        if isinstance(value, (int, float)) and 0 <= float(value) < 1:
            total_seconds = int(round(float(value) * 24 * 60 * 60))
            hours = (total_seconds // 3600) % 24
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        s = _clean_text(value)
        m = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
        if m:
            hh, mm, ss = m.group(1), m.group(2), m.group(3) or "00"
            return f"{int(hh):02d}:{mm}:{ss}"
        return s

    header_aliases = {
        "fahrzeug_kategorie": ["fahrzeugkategorie", "kategorie"],
        "datum": ["datumdertransaktion", "datum"],
        "uhrzeit": ["zeitpunktdertransaktion", "zeitpunkt", "uhrzeit", "zeit"],
        "fahrzeug_ia": ["fahrzeugia"],
        "fahrzeug": ["fahrzeug"],
        "fahrer": ["fahrer"],
        "produkt": ["produkt"],
        "transaktions_typ": ["transaktionstyp"],
        "zapfsaeule": ["zapfsaule", "zapfsaeule"],
    }
    # Fahrzeug-IA ist optional, weil einige Waschdateien diese Spalte nicht enthalten.
    # Wenn sie hier Pflicht bleibt, werden komplette Dateien bzw. Blätter übersprungen.
    required_fields = ["datum", "uhrzeit", "fahrer", "fahrzeug", "produkt"]

    rows = []
    seen = set()

    for uploaded_file in uploaded_files or []:
        payload = read_upload_bytes(uploaded_file)
        if not payload:
            continue
        try:
            xls = pd.ExcelFile(io.BytesIO(payload), engine=EXCEL_READ_ENGINE)
        except Exception:
            continue

        for sheet_name in xls.sheet_names:
            try:
                df = xls.parse(sheet_name=sheet_name)
            except Exception:
                continue
            if df is None or df.empty:
                continue

            norm_cols = {_norm_header(col): col for col in df.columns}
            resolved = {}
            for target, aliases in header_aliases.items():
                for alias in aliases:
                    if alias in norm_cols:
                        resolved[target] = norm_cols[alias]
                        break

            # Kennzeichen steht IMMER in Spalte E (0-basiert: Index 4) — positional übersteuern,
            # weil die Header in den Quell-Dateien uneinheitlich sind und das Header-Matching
            # sonst teils Spalte D (Fahrzeug-IA) liefert.
            if len(df.columns) > 4:
                resolved["fahrzeug"] = df.columns[4]

            if not all(field in resolved for field in required_fields):
                continue

            selected = pd.DataFrame()
            for target in header_aliases:
                if target in resolved:
                    selected[target] = df[resolved[target]]
                else:
                    selected[target] = ""
            selected = selected.dropna(how="all")
            if selected.empty:
                continue

            # itertuples statt iterrows — bei tausenden Zeilen ist iterrows
            # mit Series-Bauen pro Zeile mit Abstand der teuerste Anteil.
            sel_cols = list(selected.columns)
            i_datum   = sel_cols.index("datum")              if "datum" in sel_cols else None
            i_uhrzeit = sel_cols.index("uhrzeit")            if "uhrzeit" in sel_cols else None
            i_fahrer  = sel_cols.index("fahrer")             if "fahrer" in sel_cols else None
            i_fzg     = sel_cols.index("fahrzeug")           if "fahrzeug" in sel_cols else None
            i_fzg_ia  = sel_cols.index("fahrzeug_ia")        if "fahrzeug_ia" in sel_cols else None
            i_prod    = sel_cols.index("produkt")            if "produkt" in sel_cols else None
            i_kat     = sel_cols.index("fahrzeug_kategorie") if "fahrzeug_kategorie" in sel_cols else None
            i_typ     = sel_cols.index("transaktions_typ")   if "transaktions_typ" in sel_cols else None
            i_zaps    = sel_cols.index("zapfsaeule")         if "zapfsaeule" in sel_cols else None

            quelle_name = getattr(uploaded_file, "name", "")

            def _at(row, idx):
                if idx is None: return None
                try: return row[idx]
                except IndexError: return None

            for row in selected.itertuples(index=False, name=None):
                datum, date_iso = _parse_date(_at(row, i_datum))
                uhrzeit = _parse_time(_at(row, i_uhrzeit))
                fahrer = _clean_text(_at(row, i_fahrer))
                fahrzeug = _clean_text(_at(row, i_fzg))
                fahrzeug_ia = _clean_text(_at(row, i_fzg_ia))
                if not fahrzeug and fahrzeug_ia:
                    fahrzeug = fahrzeug_ia
                fahrzeug = _normalize_kennzeichen(fahrzeug)
                produkt = _clean_text(_at(row, i_prod))
                fahrzeug_kategorie = _clean_text(_at(row, i_kat))
                transaktions_typ = _clean_text(_at(row, i_typ))
                zapfsaeule = _clean_text(_at(row, i_zaps))
                if not any([datum, uhrzeit, fahrer, fahrzeug, fahrzeug_ia, produkt]):
                    continue
                datetime_iso = (date_iso + " " + (uhrzeit or "00:00:00")).strip() if date_iso else ""
                item = {
                    "datum": datum,
                    "date_iso": date_iso,
                    "uhrzeit": uhrzeit,
                    "datetime_iso": datetime_iso,
                    "fahrer": fahrer,
                    "fahrzeug": fahrzeug,
                    "fahrzeug_ia": fahrzeug_ia,
                    "produkt": produkt,
                    "fahrzeug_kategorie": fahrzeug_kategorie,
                    "transaktions_typ": transaktions_typ,
                    "zapfsaeule": zapfsaeule,
                    "quelle": quelle_name,
                }
                key = (
                    item["datum"], item["uhrzeit"], item["fahrer"], item["fahrzeug"],
                    item["fahrzeug_ia"], item["produkt"], item["fahrzeug_kategorie"],
                    item["transaktions_typ"], item["zapfsaeule"], item["quelle"],
                )
                if key in seen:
                    continue
                seen.add(key)
                rows.append(item)

    rows.sort(key=lambda x: (x.get("datetime_iso", ""), x.get("fahrer", ""), x.get("fahrzeug", "")), reverse=True)
    return json.dumps(rows, ensure_ascii=False, separators=(",", ":"))


def parse_grosskunden_excel(uploaded_file) -> str:
    """Liest alle Blätter der Großkunden-Excel aus und liefert JSON.

    Performance: read_only=True und Streaming über die Zeilen, damit große
    Großkunden-Dateien nicht komplett als Liste im Speicher landen.
    """
    import openpyxl as _opxl

    def _clean(v):
        if v is None:
            return ""
        return str(v).replace("\xa0", " ").strip()

    HEADER_KEYWORDS = {
        "name", "kundennummer", "mail", "telefon", "hinweise",
        "e-mail", "email", "adresse", "ansprechpartner", "bemerkung",
        "info", "kontakt", "anmerkung", "lieferzeit", "hinweis",
        "knr", "strasse", "straße", "plz", "ort",
    }

    def _is_structured(first_row):
        cells = {_clean(c).lower().strip() for c in first_row}
        return bool(HEADER_KEYWORDS & cells)

    def _row_has_value(row):
        return any(_clean(c) for c in row)

    empty = "[]"
    payload = read_upload_bytes(uploaded_file)
    if not payload:
        return empty
    try:
        wb = _opxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    except Exception:
        return empty

    result = []
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            row_iter = ws.iter_rows(values_only=True)
            first_row = None
            for row in row_iter:
                if _row_has_value(row):
                    first_row = row
                    break
            if first_row is None:
                continue

            if _is_structured(first_row):
                headers = [_clean(c) for c in first_row]

                name_col = next(
                    (i for i, h in enumerate(headers) if h.lower().strip() == "name"),
                    -1
                )
                knr_col = next(
                    (i for i, h in enumerate(headers)
                     if "kundennummer" in h.lower() or h.lower().strip() == "knr"),
                    -1
                )
                content_idx = [i for i, h in enumerate(headers)
                               if i != name_col and i != knr_col and h.strip()]
                content_headers = [headers[i] for i in content_idx]

                def _get(row, idx):
                    if idx < 0 or idx >= len(row):
                        return ""
                    return _clean(row[idx])

                entries = []
                current = None
                for row in row_iter:
                    if not _row_has_value(row):
                        continue
                    name = _get(row, name_col)
                    knr = _get(row, knr_col)
                    content_row = [_get(row, i) for i in content_idx]

                    if name:
                        current = {
                            "name": name,
                            "kundennummer": knr,
                            "rows": [content_row] if any(content_row) else [],
                        }
                        entries.append(current)
                    elif current and any(content_row):
                        current["rows"].append(content_row)

                if entries:
                    result.append({
                        "name": sname,
                        "type": "structured",
                        "content_headers": content_headers,
                        "entries": entries,
                    })
            else:
                lines = []
                for cell in first_row:
                    val = _clean(cell)
                    if val and val != "None":
                        lines.append(val)
                for row in row_iter:
                    if not _row_has_value(row):
                        continue
                    for cell in row:
                        val = _clean(cell)
                        if val and val != "None":
                            lines.append(val)
                if lines:
                    result.append({"name": sname, "type": "freeform", "lines": lines})
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return json.dumps(result, ensure_ascii=False)


# =============================================================================
# STREAMLIT UI
# =============================================================================

def _empty_inst(name="Normalwochen"):
    return {
        "name": name,
        "suche_html": None,
        "druck_html": None,
        "source_sig": None,
        "versp_start_json": "{}",
        "versp_start_sig": None,
        "woche_data": {},
        "excel_content_hash": "",
        "excel_filename": "",
        "quality_blocked": False,
    }

# Session-State robust initialisieren.
# Wichtig: nicht per Attributzugriff lesen, bevor der Key sicher existiert.
st.session_state.setdefault("instances", [_empty_inst("Normalwochen")])


# Zentrale Verarbeitungsanzeige. Die Eintraege bleiben waehrend der Session
# erhalten und werden bei einem neuen erfolgreichen Lauf desselben Bereichs
# aktualisiert. Fehler werden zusaetzlich in einer kompakten Historie bewahrt.
st.session_state.setdefault("_processing_status", {})
st.session_state.setdefault("_processing_errors", [])
st.session_state.setdefault("_quality_checks", {})


def _status_now() -> str:
    return datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")


def _uploaded_name(uploaded) -> str:
    return str(getattr(uploaded, "name", "") or "")


def _set_processing_status(key: str, area: str, status: str,
                           detail: str = "", filename: str = "") -> None:
    """Aktualisiert genau einen sichtbaren Status-Eintrag."""
    status = status if status in {"ok", "warning", "error", "info"} else "info"
    st.session_state.setdefault("_processing_status", {})[key] = {
        "bereich": area,
        "status": status,
        "datei": filename,
        "detail": str(detail or ""),
        "zeit": _status_now(),
    }


def _clear_errors_for_key(key: str) -> None:
    errors = st.session_state.setdefault("_processing_errors", [])
    st.session_state["_processing_errors"] = [e for e in errors if e.get("key") != key]


def _record_processing_error(key: str, area: str, error: Exception,
                             filename: str = "", detail: str = "") -> None:
    """Zeigt einen Fehler im Status und legt eine deduplizierte Historie an."""
    err_type = type(error).__name__
    err_text = str(error) or repr(error)
    visible_detail = detail or f"{err_type}: {err_text}"
    _set_processing_status(key, area, "error", visible_detail, filename)

    errors = st.session_state.setdefault("_processing_errors", [])
    signature = f"{key}|{filename}|{err_type}|{err_text}"
    errors[:] = [e for e in errors if e.get("signature") != signature]
    errors.append({
        "key": key,
        "bereich": area,
        "datei": filename,
        "fehlertyp": err_type,
        "fehler": err_text,
        "zeit": _status_now(),
        "signature": signature,
    })
    del errors[:-50]


def _record_processing_success(key: str, area: str, detail: str = "",
                               filename: str = "") -> None:
    _clear_errors_for_key(key)
    _set_processing_status(key, area, "ok", detail, filename)


def _quality_result(status: str, area: str, filename: str, size: int,
                    summary: str, details=None, warnings=None, errors=None,
                    content_hash: str = "", required: bool = False,
                    source_signature: str = "") -> dict:
    return {
        "status": status,
        "area": area,
        "filename": filename,
        "size": int(size or 0),
        "summary": str(summary or ""),
        "details": [str(x) for x in (details or []) if str(x)],
        "warnings": [str(x) for x in (warnings or []) if str(x)],
        "errors": [str(x) for x in (errors or []) if str(x)],
        "content_hash": str(content_hash or ""),
        "required": bool(required),
        "source_signature": str(source_signature or ""),
        "checked_at": _status_now(),
    }


def _decode_text_sample(payload: bytes) -> tuple[str, str]:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return payload.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace"), "utf-8 (Ersatzzeichen)"


def _stream_upload_hash(uploaded_file, chunk_size: int = 1024 * 1024) -> str:
    """SHA-256 ohne den gesamten Upload ein zweites Mal in den RAM zu kopieren."""
    if uploaded_file is None:
        return ""
    old_pos = None
    try:
        old_pos = uploaded_file.tell()
    except Exception:
        pass
    digest = hashlib.sha256()
    try:
        uploaded_file.seek(0)
        while True:
            chunk = uploaded_file.read(chunk_size)
            if not chunk:
                break
            if isinstance(chunk, str):
                chunk = chunk.encode("utf-8", errors="ignore")
            digest.update(chunk)
    finally:
        try:
            uploaded_file.seek(old_pos if old_pos is not None else 0)
        except Exception:
            pass
    return digest.hexdigest()


def _read_upload_sample(uploaded_file, limit: int = 131072) -> bytes:
    """Liest nur einen kleinen Anfangsbereich und stellt die Dateiposition wieder her."""
    if uploaded_file is None:
        return b""
    old_pos = None
    try:
        old_pos = uploaded_file.tell()
    except Exception:
        pass
    try:
        uploaded_file.seek(0)
        data = uploaded_file.read(limit)
    finally:
        try:
            uploaded_file.seek(old_pos if old_pos is not None else 0)
        except Exception:
            pass
    if data is None:
        return b""
    if isinstance(data, str):
        return data.encode("utf-8", errors="ignore")
    return bytes(data)


def _validate_uploaded_file(uploaded_file, expected_types, quality_key: str,
                            *, area: str = "Datei", kind: str = "auto",
                            required: bool = False) -> dict:
    """Speicherschonende Grundprüfung eines Uploads.

    Anders als Version 34 wird die Datei hier nicht komplett per ``getvalue``
    dupliziert und nicht vorsorglich mit pandas geoeffnet. Die eigentlichen
    Parser pruefen den Inhalt spaeter ohnehin nochmals. Dadurch bleibt der
    Start auch mit vielen grossen Uploads auf Streamlit Cloud stabil.
    """
    checks = st.session_state.setdefault("_quality_checks", {})
    filename = _uploaded_name(uploaded_file)
    source_sig = upload_signature(uploaded_file)
    cached = checks.get(quality_key)
    if cached and cached.get("source_signature") == source_sig and cached.get("kind") == kind:
        return cached

    try:
        size = int(getattr(uploaded_file, "size", 0) or 0)
    except Exception:
        size = 0
    ext = Path(filename).suffix.lower().lstrip(".")
    allowed = {str(x).lower().lstrip(".") for x in (expected_types or [])}
    details: list[str] = []
    warnings: list[str] = []
    errors: list[str] = []

    if not filename:
        errors.append("Dateiname fehlt")
    if allowed and ext not in allowed:
        errors.append(f"Falscher Dateityp .{ext or '?'}; erwartet: " + ", ".join(sorted(allowed)))
    if size <= 0:
        errors.append("Datei ist leer")
    else:
        details.append(_human_size(size))
        if size > 75 * 1024 * 1024:
            warnings.append("Sehr grosse Datei; die Verarbeitung kann auf Streamlit Cloud viel Speicher benoetigen")

    actual_kind = kind
    if actual_kind == "auto":
        if ext == "xlsx":
            actual_kind = "xlsx"
        elif ext == "csv":
            actual_kind = "csv"
        elif ext == "json":
            actual_kind = "json"
        elif ext in {"png", "jpg", "jpeg", "svg"}:
            actual_kind = "logo"

    sample = _read_upload_sample(uploaded_file) if size > 0 else b""
    if sample and not errors:
        if actual_kind in {"xlsx", "week_excel", "key_excel"}:
            if not sample.startswith(b"PK"):
                errors.append("Datei besitzt keine gueltige XLSX-Signatur")
            else:
                details.append("XLSX-Signatur gueltig")
                if actual_kind == "week_excel":
                    details.append("Wocheninhalt wird beim Verarbeiten geprueft")
                elif actual_kind == "key_excel":
                    details.append("Schluesselzuordnungen werden beim Verarbeiten geprueft")
        elif actual_kind == "csv":
            text, encoding = _decode_text_sample(sample)
            lines = [line for line in text.splitlines() if line.strip()]
            if not lines:
                errors.append("CSV enthaelt im Dateianfang keine Daten")
            else:
                probe = lines[:20]
                delimiters = {sep: sum(line.count(sep) for line in probe) for sep in (";", ",", "\t", "|")}
                delimiter = max(delimiters, key=delimiters.get)
                col_count = max(1, probe[0].count(delimiter) + 1) if delimiters[delimiter] else 1
                details.append(f"Dateianfang lesbar · ca. {col_count} Spalten · {encoding}")
                if col_count == 1:
                    warnings.append("Kein eindeutiges CSV-Trennzeichen im Dateianfang erkannt")
        elif actual_kind == "json":
            stripped = sample.lstrip()
            if not stripped.startswith((b"{", b"[")):
                errors.append("Dateianfang sieht nicht wie JSON aus")
            else:
                details.append("JSON-Dateianfang erkannt; Vollpruefung erfolgt beim Verarbeiten")
        elif actual_kind == "logo":
            lower = sample[:500].lstrip().lower()
            valid = (
                sample.startswith(b"\x89PNG\r\n\x1a\n")
                or sample.startswith(b"\xff\xd8\xff")
                or (ext == "svg" and (lower.startswith(b"<svg") or b"<svg" in lower))
            )
            if not valid:
                errors.append("Bildinhalt passt nicht zum ausgewaehlten PNG/JPG/SVG-Format")
            else:
                details.append("Bildsignatur gueltig")

    # Nur Wochen-Dateien werden inhaltlich gehasht, damit auch umbenannte
    # Duplikate erkannt werden. Das Hashing erfolgt gestreamt ohne Vollkopie.
    if size > 0 and actual_kind == "week_excel" and not errors:
        content_hash = _stream_upload_hash(uploaded_file)
    else:
        content_hash = hashlib.sha256(f"{filename}|{size}|{source_sig}".encode("utf-8", errors="ignore")).hexdigest()

    status = "error" if errors else ("warning" if warnings else "ok")
    if status == "ok":
        summary = "Grundprüfung bestanden"
    elif status == "warning":
        summary = f"Grundprüfung bestanden · {len(warnings)} Hinweis(e)"
    else:
        summary = f"Dateiprüfung fehlgeschlagen · {len(errors)} Fehler"
    result = _quality_result(
        status, area, filename, size, summary, details, warnings, errors,
        content_hash=content_hash, required=required, source_signature=source_sig,
    )
    result["kind"] = kind
    checks[quality_key] = result
    return result


def _render_quality_result(result: dict, *, compact: bool = True) -> None:
    if not result:
        return
    detail_text = " · ".join(result.get("details", [])[:3])
    prefix = {"ok": "✓", "warning": "⚠", "error": "✗"}.get(result.get("status"), "•")
    line = f"{prefix} Dateiprüfung: {result.get('summary', '')}"
    if detail_text:
        line += f" · {detail_text}"
    if result.get("status") == "error":
        st.error(line + (" · " + " | ".join(result.get("errors", [])[:3]) if result.get("errors") else ""))
    elif result.get("status") == "warning":
        st.warning(line + (" · " + " | ".join(result.get("warnings", [])[:3]) if result.get("warnings") else ""))
    else:
        st.caption(line)


def _quality_table_rows() -> list[dict]:
    rows = []
    for key, result in sorted((st.session_state.get("_quality_checks", {}) or {}).items(), key=lambda item: str(item[1].get("area", "")).lower()):
        status = result.get("status", "info")
        rows.append({
            "Status": {"ok": "✓ OK", "warning": "⚠ Hinweis", "error": "✗ Fehler"}.get(status, "• Info"),
            "Bereich": result.get("area", key),
            "Datei": result.get("filename", ""),
            "Groesse": _human_size(result.get("size", 0)),
            "Ergebnis": result.get("summary", ""),
            "Details": " | ".join((result.get("errors") or result.get("warnings") or result.get("details") or [])[:3]),
        })
    return rows


def _duplicate_week_groups() -> list[list[dict]]:
    by_hash: dict[str, list[dict]] = {}
    for idx, inst in enumerate(st.session_state.get("instances", []) or []):
        digest = str(inst.get("excel_content_hash", "") or "")
        if not digest:
            continue
        by_hash.setdefault(digest, []).append({
            "index": idx,
            "name": str(inst.get("name", f"Woche {idx + 1}") or f"Woche {idx + 1}"),
            "filename": str(inst.get("excel_filename", "") or ""),
        })
    return [group for group in by_hash.values() if len(group) > 1]


def _clear_generated_html_file() -> None:
    """Entfernt eine zuvor erzeugte temporäre HTML-Datei dieser Session."""
    old_path = st.session_state.pop("_generated_app_html_path", None)
    st.session_state.pop("_generated_app_html_meta", None)
    if old_path:
        try:
            Path(old_path).unlink(missing_ok=True)
        except Exception:
            pass


def _read_generated_html_file(path: str) -> bytes:
    """Wird von st.download_button erst beim tatsächlichen Download aufgerufen."""
    return Path(path).read_bytes()


def _generation_source_signature(ready_instances: list) -> str:
    """Kompakte Signatur aller Quellen, die den HTML-Export beeinflussen.

    Es werden nur bereits vorhandene Upload-/Parser-Signaturen und kleine
    Metadaten gehasht. Die großen JSON- und HTML-Inhalte werden bewusst nicht
    bei jedem Streamlit-Rerun erneut kopiert.
    """
    parts = [APP_CACHE_VERSION, EXTRA_CACHE_VERSION]
    for inst in ready_instances or []:
        parts.extend([
            str(inst.get("name", "")),
            str(inst.get("source_sig", "")),
            str(inst.get("versp_start_sig", "")),
        ])

    for key in sorted(st.session_state.keys()):
        if str(key).endswith("_sig"):
            value = st.session_state.get(key)
            if value not in (None, ""):
                parts.append(f"{key}={value}")

    for key in (
        "g_logo", "g_key", "g_fach", "g_fcsb", "g_lh_csv",
        "g_rahmen_csv", "g_kundenart_csv",
    ):
        parts.append(f"{key}={upload_signature(st.session_state.get(key))}")

    return combine_signatures(*parts)


def _human_size(num_bytes: int) -> str:
    try:
        value = max(0, int(num_bytes))
    except Exception:
        value = 0
    if value < 1024:
        return f"{value} B"
    if value < 1024 * 1024:
        return f"{value / 1024:.1f} KB"
    return f"{value / 1024 / 1024:.1f} MB"


def _safe_state_json(key: str, fallback):
    try:
        value = json.loads(st.session_state.get(key, "") or "")
        return value
    except Exception:
        return fallback


def _embedded_pdf_stats() -> tuple[int, int]:
    """Anzahl und komprimierte Bytegroesse der eingebetteten PDFs."""
    meta = _load_asset_meta()
    return int(meta.get("pdf_count", 0) or 0), int(meta.get("pdf_compressed_bytes", 0) or 0)


def _build_generation_metadata(ready_instances: list, generated_at: datetime.datetime) -> dict:
    """Erstellt eine kompakte, browserfreundliche Datenstandsuebersicht."""
    status_map = st.session_state.get("_processing_status", {}) or {}
    source_rows = []
    status_counts = {"ok": 0, "warning": 0, "error": 0, "info": 0}
    for key, row in sorted(status_map.items(), key=lambda item: str(item[1].get("bereich", "")).lower()):
        status = str(row.get("status", "info") or "info")
        if key.startswith("export_") or key == "download_html":
            continue
        status_counts[status] = status_counts.get(status, 0) + 1
        source_rows.append({
            "status": status,
            "area": str(row.get("bereich", "") or ""),
            "file": str(row.get("datei", "") or ""),
            "detail": str(row.get("detail", "") or ""),
            "time": str(row.get("zeit", "") or ""),
        })

    tel = _safe_state_json("tel_json", [])
    wash = _safe_state_json("fahrzeugwaesche_json", [])
    tank = _safe_state_json("tanken_json", [])
    timerec = _safe_state_json("timerec_json", {})
    violations = _safe_state_json("verstoss_json", {})
    expenses = _safe_state_json("spesen_json", {})
    big_customers = _safe_state_json("grosskunden_json", [])
    carriers = _safe_state_json("spediteure_json", {})
    driver_rating = _safe_state_json("fahrerbewertung_json", {})

    shift_count = sum(len(v) for v in timerec.values() if isinstance(v, list)) if isinstance(timerec, dict) else 0
    pdf_count, pdf_bytes = _embedded_pdf_stats()
    weeks = [str(inst.get("name", "Woche")) for inst in ready_instances]

    datasets = [
        {"label": "Wochen / Suche", "value": str(len(weeks)), "detail": ", ".join(weeks)},
        {"label": "Telefon / Fachberater", "value": str(len(tel) if isinstance(tel, list) else 0), "detail": "Einträge"},
        {"label": "Fahrzeugwaesche", "value": str(len(wash) if isinstance(wash, list) else 0), "detail": "Datensaetze"},
        {"label": "Tanken", "value": str(len(tank) if isinstance(tank, list) else 0), "detail": f"{sum(float(r.get('menge_liter', 0) or 0) for r in tank):,.0f} Liter" if isinstance(tank, list) else "0 Liter"},
        {"label": "Schichten / Tachograph", "value": str(shift_count), "detail": f"{len(timerec) if isinstance(timerec, dict) else 0} Fahrer"},
        {"label": "Verstöße", "value": str(violations.get("total_violations", 0) if isinstance(violations, dict) else 0), "detail": f"{len(violations.get('drivers', [])) if isinstance(violations, dict) else 0} Fahrer"},
        {"label": "Spesen", "value": str(expenses.get("total_rows", 0) if isinstance(expenses, dict) else 0), "detail": f"{len(expenses.get('drivers', [])) if isinstance(expenses, dict) else 0} Fahrer"},
        {"label": "Großkunden", "value": str(len(big_customers) if isinstance(big_customers, list) else 0), "detail": "Kunden"},
        {"label": "Spediteure", "value": str(len(carriers.get("fahrten", [])) if isinstance(carriers, dict) else 0), "detail": "Fahrten"},
        {"label": "Fahrerbewertung", "value": str(len(driver_rating.get("drivers", [])) if isinstance(driver_rating, dict) else 0), "detail": "Fahrer"},
        {"label": "PDF-Dokumente", "value": str(pdf_count), "detail": f"{_human_size(pdf_bytes)} komprimiert"},
    ]

    return {
        "version": APP_DISPLAY_VERSION,
        "created_at": generated_at.strftime("%d.%m.%Y %H:%M:%S"),
        "weeks": weeks,
        "datasets": datasets,
        "sources": source_rows[-40:],
        "stats": {
            "ok": status_counts.get("ok", 0),
            "warning": status_counts.get("warning", 0),
            "error": status_counts.get("error", 0),
            "info": status_counts.get("info", 0),
            "pdf_count": pdf_count,
            "pdf_compressed_bytes": pdf_bytes,
        },
    }


def _estimate_export_size(ready_instances: list) -> int:
    """Grobe Exportprognose ohne die große HTML vorab zu erzeugen.

    Wochen-HTML und Zusatzdaten werden beim Export stark komprimiert; die
    eingebetteten PDF-Blöcke liegen dagegen bereits als Base64 vor. Die
    Schätzung dient nur als Orientierung in der Startprüfung.
    """
    week_chars = 0
    for inst in ready_instances or []:
        week_chars += len(str(inst.get("suche_html", "") or ""))
        week_chars += len(str(inst.get("druck_html", "") or ""))

    extra_chars = 0
    for key in (
        "tel_json", "sam_json", "fa_json", "zulage_json",
        "drittkunden_json", "fahrzeugwaesche_json", "tanken_json", "verstoss_json",
        "spesen_json", "grosskunden_json", "timerec_json",
        "spediteure_json", "fahrerbewertung_json",
    ):
        extra_chars += len(str(st.session_state.get(key, "") or ""))

    asset_meta = _load_asset_meta()
    pdf_b64_chars = int(asset_meta.get("pdf_b64_chars", 0) or 0)
    static_js_chars = int(asset_meta.get("dashboard_js_chars", 0) or 0)

    # Erfahrungswerte: Wochen-HTML ca. 35–45 %, JSON ca. 45–60 % nach
    # Deflate + Base64. Ein Sicherheitsaufschlag deckt HTML/CSS/Metadaten ab.
    estimated = (
        pdf_b64_chars
        + int(week_chars * 0.43)
        + int(extra_chars * 0.56)
        + static_js_chars
        + 700_000
    )
    return max(0, int(estimated))


def _build_export_preflight(ready_instances: list) -> tuple[list[dict], list[str], int]:
    """Erstellt die sichtbare Startprüfung und nennt echte Blocker."""
    rows: list[dict] = []
    blockers: list[str] = []

    asset_error = _asset_installation_error()
    if asset_error:
        blockers.append("Programmdateien / nfc_assets")
        rows.append({"Status": "✗ Fehlt", "Bereich": "Programmdateien", "Details": asset_error})

    def add(label: str, ok: bool, detail: str, *, required: bool = False,
            warning: bool = False) -> None:
        if ok:
            status = "✓ Bereit"
        elif required:
            status = "✗ Fehlt"
            blockers.append(label)
        elif warning:
            status = "⚠ Hinweis"
        else:
            status = "– Optional"
        rows.append({
            "Status": status,
            "Prüfung": label,
            "Details": detail,
            "Pflicht": "Ja" if required else "Nein",
        })

    logo = st.session_state.get("g_logo")
    key_file = st.session_state.get("g_key")
    add("Logo", bool(logo), _uploaded_name(logo) or "Noch nicht hochgeladen", required=True)
    add("Schlüsseldatei", bool(key_file), _uploaded_name(key_file) or "Noch nicht hochgeladen", required=True)

    ready_names = [str(inst.get("name", "Woche") or "Woche") for inst in ready_instances or []]
    add(
        "Verarbeitete Woche", bool(ready_names),
        ", ".join(ready_names) if ready_names else "Noch keine Woche vollständig verarbeitet",
        required=True,
    )

    all_instances = st.session_state.get("instances", []) or []
    incomplete = [
        str(inst.get("name", "Woche") or "Woche")
        for inst in all_instances
        if not (inst.get("suche_html") and inst.get("druck_html"))
    ]
    add(
        "Weitere angelegte Wochen", not incomplete,
        "Alle angelegten Wochen sind bereit" if not incomplete else "Nicht bereit: " + ", ".join(incomplete),
        warning=bool(incomplete),
    )

    pdf_count, pdf_bytes = _embedded_pdf_stats()
    add(
        "Eingebettete PDFs", pdf_count > 0,
        f"{pdf_count} Dokumente · {_human_size(pdf_bytes)} komprimiert" if pdf_count else "Keine PDFs eingebettet",
        warning=(pdf_count == 0),
    )

    optional_keys = (
        "tel_json", "zulage_json", "drittkunden_json", "fahrzeugwaesche_json", "tanken_json",
        "verstoss_json", "spesen_json", "grosskunden_json", "timerec_json",
        "spediteure_json", "fahrerbewertung_json",
    )
    loaded_optional = sum(
        1 for key in optional_keys
        if str(st.session_state.get(key, "") or "") not in ("", "[]", "{}", '{"drivers":[],"total_violations":0}', '{"drivers":[],"months":[],"total_cost":0,"total_rows":0}')
    )
    add(
        "Zusatzdaten", loaded_optional > 0,
        f"{loaded_optional} von {len(optional_keys)} Datenbereichen geladen" if loaded_optional else "Keine Zusatzdaten geladen",
    )

    error_count = sum(
        1 for row in (st.session_state.get("_processing_status", {}) or {}).values()
        if row.get("status") == "error"
    )
    add(
        "Verarbeitungsfehler", error_count == 0,
        "Keine aktuellen Fehler" if error_count == 0 else f"{error_count} Fehler im Verarbeitungsstatus – Export bleibt möglich",
        warning=(error_count > 0),
    )

    quality_checks = list((st.session_state.get("_quality_checks", {}) or {}).values())
    required_quality_errors = [q for q in quality_checks if q.get("required") and q.get("status") == "error"]
    optional_quality_errors = [q for q in quality_checks if not q.get("required") and q.get("status") == "error"]
    quality_warnings = [q for q in quality_checks if q.get("status") == "warning"]
    quality_detail = (
        f"{len(quality_checks)} Dateien geprueft · {len(quality_warnings)} Hinweis(e)"
        if quality_checks else "Noch keine Dateien geprueft"
    ) + (f" · {len(optional_quality_errors)} optionale Fehler" if optional_quality_errors else "")
    if required_quality_errors:
        rows.append({
            "Status": "✗ Fehler",
            "Prüfung": "Dateiqualität",
            "Details": quality_detail + f" · {len(required_quality_errors)} Pflichtdatei(en) fehlerhaft",
            "Pflicht": "Ja",
        })
        blockers.append("Dateiqualität")
    elif quality_warnings or optional_quality_errors:
        rows.append({
            "Status": "⚠ Hinweis",
            "Prüfung": "Dateiqualität",
            "Details": quality_detail,
            "Pflicht": "Nein",
        })
    else:
        rows.append({
            "Status": "✓ Bereit",
            "Prüfung": "Dateiqualität",
            "Details": quality_detail,
            "Pflicht": "Nein",
        })

    duplicate_groups = _duplicate_week_groups()
    duplicate_detail = "; ".join(
        " = ".join(f"{x['name']} ({x['filename'] or 'Datei'})" for x in group)
        for group in duplicate_groups
    )
    add(
        "Doppelte Wochen", not duplicate_groups,
        "Keine identischen Wochen-Dateien erkannt" if not duplicate_groups else duplicate_detail,
        required=True,
    )

    estimate = _estimate_export_size(ready_instances)
    rows.append({
        "Status": "• Prognose",
        "Prüfung": "Erwartete HTML-Größe",
        "Details": f"ca. {_human_size(estimate)}",
        "Pflicht": "Nein",
    })
    return rows, blockers, estimate


def _reset_all_app_data() -> None:
    """Leert Uploads, Sessiondaten, Caches und die temporäre Exportdatei."""
    _clear_generated_html_file()
    try:
        st.cache_data.clear()
    except Exception:
        pass
    try:
        st.cache_resource.clear()
    except Exception:
        pass
    for key in list(st.session_state.keys()):
        try:
            del st.session_state[key]
        except Exception:
            pass


def _safe_cached_export_b64(cache_key: str, source_value: str, builder,
                            area: str) -> str:
    """Erzeugt einen Export, ohne dass ein Fehler den gesamten Download-Tab stoppt."""
    status_key = f"export_{cache_key}"
    try:
        result = get_cached_export_b64(cache_key, source_value, builder)
        _record_processing_success(
            status_key, area,
            "Export vorbereitet" if result else "Keine Exportdaten vorhanden",
        )
        return result
    except Exception as exc:
        _record_processing_error(status_key, area, exc)
        return ""


# -----------------------------------------------------------------------------
# UI-Helpers
# -----------------------------------------------------------------------------
def _global_uploader(label, types, ss_key, widget_key):
    """Stammdaten-Uploader mit echter Inhaltspruefung."""
    up = st.file_uploader(label, type=types, key=widget_key)
    required = ss_key in {"g_logo", "g_key"}
    kind = "logo" if ss_key == "g_logo" else ("key_excel" if ss_key == "g_key" else "auto")
    if up:
        quality = _validate_uploaded_file(
            up, types, f"quality_global_{ss_key}", area=label, kind=kind, required=required
        )
        _render_quality_result(quality)
        if quality.get("status") != "error":
            st.session_state[ss_key] = up
            _record_processing_success(
                f"upload_{ss_key}", label, quality.get("summary", "Datei ausgewaehlt"), _uploaded_name(up)
            )
        else:
            _set_processing_status(
                f"upload_{ss_key}", label, "error", quality.get("summary", "Dateiprüfung fehlgeschlagen"), _uploaded_name(up)
            )
    elif st.session_state.get(ss_key):
        stored = st.session_state.get(ss_key)
        _set_processing_status(
            f"upload_{ss_key}", label, "info", "In der Session vorhanden", _uploaded_name(stored)
        )


def _extra_single_upload(label, types, key_prefix, parser, summary_fn=None,
                         spinner_text="Verarbeite ..."):
    """Generische Zusatzdatei mit Fehler- und Statusanzeige."""
    up = st.file_uploader(label, type=types, key=f"{key_prefix}_widget")
    json_key = f"{key_prefix}_json"
    sig_key  = f"{key_prefix}_sig"
    status_key = f"extra_{key_prefix}"
    if up:
        filename = _uploaded_name(up)
        quality = _validate_uploaded_file(
            up, types, f"quality_extra_{key_prefix}", area=label, kind="auto", required=False
        )
        _render_quality_result(quality)
        if quality.get("status") == "error":
            _set_processing_status(status_key, label, "error", quality.get("summary", "Dateiprüfung fehlgeschlagen"), filename)
            return
        sig = combine_signatures(EXTRA_CACHE_VERSION, key_prefix, upload_signature(up))
        if st.session_state.get(sig_key) != sig:
            try:
                with st.spinner(spinner_text):
                    parsed = parser(up)
                st.session_state[json_key] = parsed
                st.session_state[sig_key] = sig
                _record_processing_success(status_key, label, "Erfolgreich verarbeitet", filename)
            except Exception as exc:
                _record_processing_error(status_key, label, exc, filename)
                st.error(f"{label}: {type(exc).__name__}: {exc}")
        elif st.session_state.get(json_key):
            _set_processing_status(status_key, label, "ok", "Bereits verarbeitet", filename)

        if st.session_state.get(json_key):
            if summary_fn:
                try:
                    summary = summary_fn(st.session_state[json_key])
                    st.caption(summary)
                    if st.session_state.get(sig_key) == sig:
                        _record_processing_success(status_key, label, summary, filename)
                except Exception as exc:
                    st.caption(f"{label}: geladen")
                    _set_processing_status(
                        status_key, label, "warning",
                        f"Datei verarbeitet, Zusammenfassung nicht moeglich: {type(exc).__name__}",
                        filename,
                    )
            else:
                st.caption(f"{label}: geladen")
    elif st.session_state.get(json_key):
        st.caption(f"{label}: geladen")
        current = st.session_state.get("_processing_status", {}).get(status_key)
        if not current:
            _set_processing_status(status_key, label, "info", "Daten in der Session vorhanden")


def _extra_multi_upload(label, types, key_prefix, parsers, summary_fn=None,
                        spinner_text="Verarbeite ..."):
    """Multi-Upload mit getrenntem Status je Parser und Gesamtstatus."""
    ups = st.file_uploader(label, type=types, accept_multiple_files=True,
                           key=f"{key_prefix}_widget")
    sig_key = f"{key_prefix}_sig"
    group_status_key = f"extra_{key_prefix}"
    if ups:
        filenames = ", ".join(_uploaded_name(up) for up in ups)
        quality_results = [
            _validate_uploaded_file(
                up, types, f"quality_extra_{key_prefix}_{idx}",
                area=f"{label} / {_uploaded_name(up)}", kind="auto", required=False,
            )
            for idx, up in enumerate(ups)
        ]
        for quality in quality_results:
            _render_quality_result(quality)
        content_hashes = [q.get("content_hash") for q in quality_results if q.get("content_hash")]
        duplicate_count = len(content_hashes) - len(set(content_hashes))
        if duplicate_count:
            st.warning(f"{duplicate_count} doppelte Datei(en) im Mehrfach-Upload erkannt; bitte entfernen.")
        if any(q.get("status") == "error" for q in quality_results) or duplicate_count:
            _set_processing_status(group_status_key, label, "error", "Dateiprüfung oder Duplikatpruefung fehlgeschlagen", filenames)
            return
        sig = combine_signatures(EXTRA_CACHE_VERSION, key_prefix, uploads_signature(ups))
        if st.session_state.get(sig_key) != sig:
            failed = []
            with st.spinner(spinner_text):
                for state_key, parser in parsers.items():
                    parser_status_key = f"extra_{key_prefix}_{state_key}"
                    try:
                        parsed = parser(ups)
                        st.session_state[state_key] = parsed
                        _record_processing_success(
                            parser_status_key, f"{label} / {state_key}",
                            "Erfolgreich verarbeitet", filenames,
                        )
                    except Exception as exc:
                        failed.append(state_key)
                        _record_processing_error(
                            parser_status_key, f"{label} / {state_key}", exc, filenames
                        )
            if failed:
                _set_processing_status(
                    group_status_key, label, "error",
                    "Fehler in: " + ", ".join(failed), filenames,
                )
                st.error(f"{label}: Fehler in {', '.join(failed)}")
            else:
                st.session_state[sig_key] = sig
                _record_processing_success(
                    group_status_key, label,
                    f"{len(ups)} Datei(en) vollstaendig verarbeitet", filenames,
                )
        else:
            _set_processing_status(
                group_status_key, label, "ok",
                f"{len(ups)} Datei(en) bereits verarbeitet", filenames,
            )

        if any(st.session_state.get(k) for k in parsers):
            if summary_fn:
                try:
                    summary = summary_fn(ups)
                    st.caption(summary)
                    if st.session_state.get(sig_key) == sig:
                        _record_processing_success(group_status_key, label, summary, filenames)
                except Exception as exc:
                    st.caption(f"{len(ups)} Dateien geladen")
                    if st.session_state.get(sig_key) == sig:
                        _set_processing_status(
                            group_status_key, label, "warning",
                            f"Dateien verarbeitet, Zusammenfassung nicht moeglich: {type(exc).__name__}",
                            filenames,
                        )
            else:
                st.caption(f"{len(ups)} Dateien geladen")
    elif any(st.session_state.get(k) for k in parsers):
        st.caption(f"{label}: geladen")
        if not st.session_state.get("_processing_status", {}).get(group_status_key):
            _set_processing_status(group_status_key, label, "info", "Daten in der Session vorhanden")


def _format_eur(value: float) -> str:
    return f"{value:,.2f} EUR".replace(",", "X").replace(".", ",").replace("X", ".")


# -----------------------------------------------------------------------------
# UI-Layout
# -----------------------------------------------------------------------------
_boot_log("06 Funktionsdefinitionen geladen; UI-Aufbau beginnt")
# region Eingebettete PDF-Dokumente (komprimierte Base64-Daten)
# PDF-Dokumente: zlib-komprimiert + Base64. Dadurch bleiben sie in der einzigen
# Ausgabe-Datei enthalten, ohne dass ein separater Ordner benoetigt wird.
# Im Browser werden sie erst beim Oeffnen des jeweiligen Menuepunkts entpackt.
# EMBEDDED_PDF_DOCUMENTS wurde fuer Cloud-Stabilitaet nach nfc_assets ausgelagert.
# endregion

st.title(f"{APP_DISPLAY_NAME} · Version {APP_DISPLAY_VERSION}")
_boot_log("07 Titel gerendert")
st.caption("Modularer Einzeldatei-Generator mit speicherschonender Dateiprüfung und sicherem Export")

tab_stamm, tab_wochen, tab_extra, tab_dl = st.tabs(
    ["Stammdaten", "Wochen", "Zusatzdateien", "Download"]
)

# === Tab: Stammdaten =========================================================
_boot_log("08 Tab Stammdaten wird aufgebaut")
with tab_stamm:
    col_a, col_b = st.columns(2)
    with col_a:
        _global_uploader("Logo",                       ["png","jpg","jpeg","svg"], "g_logo",      "global_up_logo_v2")
        _global_uploader("Schluesseldatei (Pflicht)",  ["xlsx"],                   "g_key",       "global_up_key_v2")
        _global_uploader("Telefonnummern Fachberater", ["xlsx"],                   "g_fach",      "global_up_fach_v2")
    with col_b:
        _global_uploader("Kundenliste Original",       ["xlsx"], "g_fcsb",       "global_up_fcsb_v2")
        _global_uploader("Lieferhinweise CSV",         ["csv"],  "g_lh_csv",     "global_up_lh_csv_v2")
        _global_uploader("Rahmentourprofil CSV",       ["csv"],  "g_rahmen_csv", "global_up_rahmen_csv_v2")
        _global_uploader("Kundenart Absetzer/Rampe CSV", ["csv"], "g_kundenart_csv", "global_up_kundenart_csv_v1")

    _items = [
        ("g_logo","Logo"), ("g_key","Schluessel"), ("g_fach","FB-Tel"),
        ("g_fcsb","Kundenliste"), ("g_lh_csv","Lieferhinweise"), ("g_rahmen_csv","Rahmentour"),
        ("g_kundenart_csv","Kundenart")
    ]
    _ok   = [lbl for k, lbl in _items if st.session_state.get(k)]
    _miss = [lbl for k, lbl in _items if not st.session_state.get(k)]
    st.divider()
    if _ok:
        st.caption("✓ " + ", ".join(_ok))
    if _miss:
        st.caption("Fehlt: " + ", ".join(_miss))

# === Tab: Wochen =============================================================
_boot_log("09 Tab Wochen wird aufgebaut")
with tab_wochen:
    for i, inst in enumerate(st.session_state["instances"]):
        _is_normal = (i == 0)
        _label = "Normalwoche" if _is_normal else f"Woche {i+1}"

        with st.expander(f"{_label}: {inst['name']}", expanded=_is_normal):
            new_name = st.text_input("Bezeichnung", value=inst["name"], key=f"inst_name_{i}")
            st.session_state["instances"][i]["name"] = new_name

            # Bestehende Instanzen aus alten Sessions haben diese Keys eventuell noch nicht.
            st.session_state["instances"][i].setdefault("versp_start_json", "{}")
            st.session_state["instances"][i].setdefault("versp_start_sig", None)

            _excel_label = "Normalwochen-Excel (Pflicht)" if _is_normal else "Wochen-Excel (Pflicht)"
            excel = st.file_uploader(_excel_label, type=["xlsx"], key=f"excel_{i}")
            _week_quality = None
            if excel:
                _week_quality = _validate_uploaded_file(
                    excel, ["xlsx"], f"quality_week_{i}",
                    area=f"{_label} / Wochen-Excel", kind="week_excel", required=True,
                )
                _render_quality_result(_week_quality)
                st.session_state["instances"][i]["excel_content_hash"] = _week_quality.get("content_hash", "")
                st.session_state["instances"][i]["excel_filename"] = _uploaded_name(excel)
                st.session_state["instances"][i]["quality_blocked"] = (_week_quality.get("status") == "error")
            else:
                st.session_state["instances"][i].setdefault("excel_content_hash", "")
                st.session_state["instances"][i].setdefault("excel_filename", "")
                st.session_state["instances"][i].setdefault("quality_blocked", False)

            _start_label = "Normaltouren-Start CSV (für Verspätung)" if _is_normal else "Sonderwochen-Tourenstart CSV (optional, sonst Normaltouren)"
            start_csv = st.file_uploader(_start_label, type=["csv"], key=f"versp_start_csv_{i}")
            if start_csv:
                _start_quality = _validate_uploaded_file(
                    start_csv, ["csv"], f"quality_week_{i}_tourstart",
                    area=f"{_label} / Tourenstart", kind="csv", required=False,
                )
                _render_quality_result(_start_quality)
                if _start_quality.get("status") == "error":
                    start_csv = None
            if start_csv:
                _start_sig = combine_signatures(EXTRA_CACHE_VERSION, "versp_start_csv", upload_signature(start_csv))
                if st.session_state["instances"][i].get("versp_start_sig") != _start_sig:
                    _status_key = f"week_{i}_tourstart"
                    try:
                        with st.spinner("Lese Tourenstart-CSV ..."):
                            _parsed_start = parse_versp_abfahrt_csv(start_csv)
                        st.session_state["instances"][i]["versp_start_json"] = _parsed_start
                        st.session_state["instances"][i]["versp_start_sig"] = _start_sig
                        _record_processing_success(
                            _status_key, f"{_label} / Tourenstart",
                            "Erfolgreich verarbeitet", _uploaded_name(start_csv),
                        )
                    except Exception as exc:
                        _record_processing_error(
                            _status_key, f"{_label} / Tourenstart", exc, _uploaded_name(start_csv)
                        )
                        st.error(f"Tourenstart-CSV: {type(exc).__name__}: {exc}")
                try:
                    _vsp_obj = json.loads(st.session_state["instances"][i].get("versp_start_json") or "{}")
                    _vsp_n = len((_vsp_obj.get("__by_tour") or {})) if isinstance(_vsp_obj, dict) else 0
                    st.caption(f"Tourenstart: {_vsp_n} Abfahrten geladen")
                except Exception:
                    st.caption("Tourenstart-CSV geladen")
            elif st.session_state["instances"][i].get("versp_start_json") not in (None, "", "{}"):
                try:
                    _vsp_obj = json.loads(st.session_state["instances"][i].get("versp_start_json") or "{}")
                    _vsp_n = len((_vsp_obj.get("__by_tour") or {})) if isinstance(_vsp_obj, dict) else 0
                    st.caption(f"Tourenstart: {_vsp_n} Abfahrten geladen")
                except Exception:
                    st.caption("Tourenstart-CSV geladen")
            elif (not _is_normal) and st.session_state["instances"][0].get("versp_start_json") not in (None, "", "{}"):
                st.caption("Tourenstart: nutzt Normaltouren-CSV als Fallback")

            _logo       = st.session_state.get("g_logo")
            _key        = st.session_state.get("g_key")
            _fach       = st.session_state.get("g_fach")
            _fcsb       = st.session_state.get("g_fcsb")
            _lh_csv     = st.session_state.get("g_lh_csv")
            _rahmen_csv = st.session_state.get("g_rahmen_csv")
            _kundenart_csv = st.session_state.get("g_kundenart_csv")

            if excel and _logo and _key and not st.session_state["instances"][i].get("quality_blocked", False):
                current_source_sig = combine_signatures(
                    APP_CACHE_VERSION,
                    upload_signature(excel),  upload_signature(_logo),
                    upload_signature(_key),   upload_signature(_fach),
                    upload_signature(_fcsb),  upload_signature(_lh_csv),
                    upload_signature(_rahmen_csv), upload_signature(_kundenart_csv),
                )
                if (inst.get("source_sig") != current_source_sig
                        or not inst.get("suche_html")
                        or not inst.get("druck_html")):
                    try:
                        with st.spinner("Generiere Suche + Druck ..."):
                            st.session_state["instances"][i]["suche_html"] = generate_suche_html(
                                excel, _key, _logo, _fach, _fcsb,
                                lieferhinweis_csv=_lh_csv, rahmentour_csv=_rahmen_csv,
                                kundenart_csv=_kundenart_csv
                            )
                            st.session_state["instances"][i]["druck_html"] = generate_druck_html(
                                excel, _logo, _fcsb, lieferhinweis_csv=_lh_csv
                            )
                            try:
                                st.session_state["instances"][i]["woche_data"] = compute_woche_data(excel)
                            except Exception as exc:
                                st.session_state["instances"][i]["woche_data"] = {}
                                _record_processing_error(
                                    f"week_{i}_data", f"{_label} / Wochendaten",
                                    exc, _uploaded_name(excel),
                                )
                            else:
                                _record_processing_success(
                                    f"week_{i}_data", f"{_label} / Wochendaten",
                                    "Wochendaten berechnet", _uploaded_name(excel),
                                )
                            st.session_state["instances"][i]["source_sig"] = current_source_sig
                        kb_s = len(st.session_state["instances"][i]["suche_html"]) // 1024
                        kb_d = len(st.session_state["instances"][i]["druck_html"]) // 1024
                        _record_processing_success(
                            f"week_{i}_html", f"{_label} / HTML",
                            f"Suche {kb_s} KB, Druck {kb_d} KB", _uploaded_name(excel),
                        )
                        st.success(f"Fertig: Suche {kb_s} KB, Druck {kb_d} KB")
                    except Exception as e:
                        _record_processing_error(
                            f"week_{i}_html", f"{_label} / HTML", e, _uploaded_name(excel)
                        )
                        st.error(f"Fehler: {e}")
                else:
                    kb_s = len(inst["suche_html"]) // 1024
                    kb_d = len(inst["druck_html"]) // 1024
                    st.caption(f"Bereit: Suche {kb_s} KB, Druck {kb_d} KB")
            elif inst["suche_html"] and inst["druck_html"]:
                kb_s = len(inst["suche_html"]) // 1024
                kb_d = len(inst["druck_html"]) // 1024
                st.caption(f"Bereit: Suche {kb_s} KB, Druck {kb_d} KB")
            else:
                missing = []
                if not excel: missing.append("Wochen-Excel")
                if excel and st.session_state["instances"][i].get("quality_blocked", False): missing.append("gueltige Wochen-Excel")
                if not _logo: missing.append("Logo")
                if not _key:  missing.append("Schluesseldatei")
                if missing: st.caption("Fehlt: " + ", ".join(missing))

            if i > 0:
                if st.button("Entfernen", key=f"del_inst_{i}"):
                    st.session_state["instances"].pop(i)
                    st.rerun()

    _week_duplicates = _duplicate_week_groups()
    if _week_duplicates:
        for _dup_group in _week_duplicates:
            st.error("Doppelte Wochen-Datei erkannt: " + " = ".join(
                f"{x['name']} ({x['filename'] or 'ohne Dateiname'})" for x in _dup_group
            ))

    if st.button("Woche hinzufuegen"):
        n = len(st.session_state["instances"])
        st.session_state["instances"].append(_empty_inst(f"Sonderwoche {n}"))
        st.rerun()

# === Tab: Zusatzdateien ======================================================
_boot_log("10 Tab Zusatzdateien wird aufgebaut")
with tab_extra:
    col_l, col_r = st.columns(2)

    with col_l:
        _extra_single_upload(
            "Telefonliste (Excel)", ["xlsx"], "tel",
            parse_telefon_excel,
            summary_fn=lambda j: f"Telefonliste: {len(json.loads(j))} Gruppen"
        )

        def _touren_summary(ups):
            _zd  = json.loads(st.session_state.zulage_json)
            _zns = sum(len(m["fahrer"]) for m in _zd.get("sonder",   []))
            _znf = sum(len(m["fahrer"]) for m in _zd.get("fuengers", []))
            _dkd = json.loads(st.session_state.drittkunden_json)
            _fan = len(json.loads(st.session_state.fa_json))
            _spd = json.loads(st.session_state.get("spediteure_json", '{"katalog":[],"fahrten":[]}'))
            return (f"{len(ups)} Dateien: Sonder {_zns}, Fuengers {_znf}, "
                    f"Drittkunden {len(_dkd)}, Fahrer {_fan}, "
                    f"Spediteure {len(_spd.get('katalog', []))} ({len(_spd.get('fahrten', []))} Fahrten)")
        _extra_multi_upload(
            "Touren-Dateien (Zulagen, Drittkunden, Fahrerauswertung)",
            ["xlsx"], "touren",
            {
                "zulage_json":      parse_zulage_excel,
                "drittkunden_json": parse_drittkunden_excel,
                "fa_json":          parse_fahrer_excel,
                "spediteure_json":  parse_spediteure_excel,
            },
            summary_fn=_touren_summary,
        )

        def _fw_summary(ups):
            rows  = json.loads(st.session_state.get("fahrzeugwaesche_json", "[]"))
            drv   = len({(r.get("fahrer") or "").strip() for r in rows if (r.get("fahrer") or "").strip()})
            lkw   = len({((r.get("fahrzeug") or r.get("fahrzeug_ia") or "").strip())
                         for r in rows if ((r.get("fahrzeug") or r.get("fahrzeug_ia") or "").strip())})
            return f"{len(rows)} Waschungen, {drv} Fahrer, {lkw} LKW"
        _extra_multi_upload(
            "Fahrzeugwaesche (Excel)", ["xlsx", "xls"], "fahrzeugwaesche",
            {"fahrzeugwaesche_json": parse_fahrzeugwaesche_excel},
            summary_fn=_fw_summary,
        )

        def _tank_summary(ups):
            rows = json.loads(st.session_state.get("tanken_json", "[]") or "[]")
            liters = sum(float(r.get("menge_liter", 0) or 0) for r in rows)
            vehicles = len({str(r.get("fahrzeug", "") or "").strip() for r in rows if str(r.get("fahrzeug", "") or "").strip()})
            drivers = len({str(r.get("fahrer", "") or "").strip() for r in rows if str(r.get("fahrer", "") or "").strip()})
            years = sorted({int(r.get("jahr", 0) or 0) for r in rows if int(r.get("jahr", 0) or 0)})
            year_text = ", ".join(str(y) for y in years) if years else "kein Jahr"
            return (f"{len(rows)} Tankvorgaenge, {liters:,.2f} Liter, {vehicles} LKW, "
                    f"{drivers} Fahrer, {year_text}").replace(",", "X").replace(".", ",").replace("X", ".")
        _extra_multi_upload(
            "Tanken (Excel-Monatsdateien)", ["xlsx"], "tanken",
            {"tanken_json": parse_tanken_excel},
            summary_fn=_tank_summary,
            spinner_text="Verarbeite Tank-Monatsdateien ...",
        )

    with col_r:
        def _spesen_summary(j):
            sp     = json.loads(j or "{}")
            drv    = sp.get("drivers", [])
            total  = float(sp.get("total_cost", 0) or 0)
            rows   = int(sp.get("total_rows", 0) or 0)
            return f"{rows} Zeilen, {len(drv)} Fahrer, {_format_eur(total)}"
        _extra_single_upload(
            "Spesen / Reisekosten (CSV)", ["csv"], "spesen",
            parse_spesen_csv, summary_fn=_spesen_summary,
        )

        def _verstoss_summary(j):
            vs    = json.loads(j or '{"drivers":[]}')
            drv   = vs.get("drivers", [])
            total = vs.get("total_violations", 0)
            dp    = sum(d.get("sum_driver_penalty",  0) for d in drv)
            cp    = sum(d.get("sum_company_penalty", 0) for d in drv)
            return (f"{len(drv)} Fahrer, {total} Verstoesse, "
                    f"Fahrer {dp:,} EUR, Firma {cp:,} EUR").replace(",", ".")
        _extra_single_upload(
            "Verstossauswertung Digitacho (CSV)", ["csv"], "verstoss",
            parse_verstoss_csv, summary_fn=_verstoss_summary,
        )

        _extra_single_upload(
            "Grosskunden (Excel)", ["xlsx"], "grosskunden",
            parse_grosskunden_excel,
            summary_fn=lambda j: f"{len(json.loads(j))} Kunden geladen",
            spinner_text="Verarbeite Grosskunden-Excel ...",
        )

        def _timerec_summary(j):
            tr  = json.loads(j or "{}")
            return f"{len(tr)} Fahrer, {sum(len(v) for v in tr.values())} Schichten"
        _extra_single_upload(
            "Schichten / Tachograph (CSV: timerecording_v3*.csv)", ["csv"],
            "timerec", parse_timerecording_csv,
            summary_fn=_timerec_summary,
            spinner_text="Verarbeite Schichten-CSV ...",
        )

        def _fahrerbewertung_summary(j):
            fb  = json.loads(j or "{}")
            drv = fb.get("drivers", [])
            grd = sum(1 for d in drv if d.get("grade") is not None)
            ev  = sum(d.get("evt", 0) for d in drv)
            mon = len(fb.get("g_months", {}))
            return f"{len(drv)} Fahrer ({grd} bewertet), {ev} Ereignisse, {mon} Monate"
        _extra_single_upload(
            "Fahrerbewertung Rohdaten (JSON: d_rohdaten.json)", ["json"],
            "fahrerbewertung", parse_fahrerbewertung_json,
            summary_fn=_fahrerbewertung_summary,
            spinner_text="Verarbeite Fahrerbewertungs-JSON ...",
        )


# === Tab: Download ===========================================================
# Die große Einzeldatei wird nicht mehr bei jedem Streamlit-Rerun neu gebaut.
# Das ist besonders auf Streamlit Cloud wichtig, weil ein String plus mehrere
# UTF-8-Kopien der HTML kurzzeitig sehr viel RAM belegen können.
_boot_log("11 Tab Download wird aufgebaut")
with tab_dl:
    import gc as _gc

    instances_state = st.session_state.get("instances")
    if not isinstance(instances_state, list) or not instances_state:
        instances_state = [_empty_inst("Normalwochen")]
        st.session_state["instances"] = instances_state

    ready = [inst for inst in instances_state
             if inst.get("suche_html") and inst.get("druck_html")
             and not inst.get("quality_blocked", False)]

    _quality_rows = _quality_table_rows()
    if _quality_rows:
        _quality_has_issue = any(row.get("Status") != "✓ OK" for row in _quality_rows)
        with st.expander("Dateiqualität", expanded=_quality_has_issue):
            st.dataframe(_quality_rows, width="stretch", hide_index=True)

    st.markdown("##### Startprüfung")
    _preflight_rows, _preflight_blockers, _estimated_html_size = _build_export_preflight(ready)
    st.dataframe(_preflight_rows, width="stretch", hide_index=True)
    if _preflight_blockers:
        st.warning("Vor der HTML-Erstellung fehlt noch: " + ", ".join(_preflight_blockers))
    else:
        st.success(f"Startprüfung bestanden · erwartete Größe ca. {_human_size(_estimated_html_size)}")

    if ready:
        zulage_json_state      = st.session_state.get("zulage_json", "{}")
        drittkunden_json_state = st.session_state.get("drittkunden_json", "[]")
        zulage_xlsx_sonder = (
            _safe_cached_export_b64(
                "zulage_sonder", zulage_json_state,
                lambda v: generate_zulage_excel(v, tab="sonder"),
                "Excel-Export Sonderzulagen",
            )
            if zulage_json_state not in ("{}", "") else ""
        )
        zulage_xlsx_fuengers = (
            _safe_cached_export_b64(
                "zulage_fuengers", zulage_json_state,
                lambda v: generate_zulage_excel(v, tab="fuengers"),
                "Excel-Export Fuengers",
            )
            if zulage_json_state not in ("{}", "") else ""
        )
        zulage_xlsx_drittkunden = (
            _safe_cached_export_b64(
                "drittkunden", drittkunden_json_state, generate_drittkunden_excel,
                "Excel-Export Drittkunden",
            )
            if drittkunden_json_state not in ("[]", "") else ""
        )

        current_generation_signature = _generation_source_signature(ready)
        stored_generation_meta = st.session_state.get("_generated_app_html_meta", {}) or {}
        stored_html_path = st.session_state.get("_generated_app_html_path")
        if stored_html_path and not Path(stored_html_path).is_file():
            _clear_generated_html_file()
            stored_html_path = None
        if (stored_html_path
                and stored_generation_meta.get("source_signature") != current_generation_signature):
            _clear_generated_html_file()
            _gc.collect()

        st.markdown("##### HTML-Einzeldatei")
        st.caption(
            "Die große suche.html wird bewusst nur auf Knopfdruck erstellt. "
            "Dadurch wird sie bei Uploads, Tabwechseln oder anderen Streamlit-Neuläufen "
            "nicht mehrfach im Arbeitsspeicher aufgebaut."
        )

        build_clicked = st.button(
            "suche.html erstellen / aktualisieren",
            type="primary",
            width="stretch",
            key="build_suche_html",
            disabled=bool(_preflight_blockers),
        )

        if build_clicked:
            # Alte Exportdaten zuerst freigeben, damit während der Neuerstellung
            # nicht zwei große HTML-Dateien gleichzeitig im Speicher liegen.
            _clear_generated_html_file()
            _gc.collect()

            generated_at = datetime.datetime.now()
            generation_meta = _build_generation_metadata(ready, generated_at)
            _build_progress = st.progress(5, text="1/5 Startprüfung abgeschlossen")
            try:
                _generation_started = time.perf_counter()
                _build_progress.progress(20, text="2/5 Zusatzdaten und Metadaten vorbereiten")
                _build_progress.progress(35, text="3/5 Wochen, Auswertungen und PDFs zusammenführen")
                app_html = combine_html(
                        instances=ready,
                        tel_json=st.session_state.get("tel_json", "[]"),
                        sam_json=st.session_state.get("sam_json", "[]"),
                        fa_json=st.session_state.get("fa_json", "[]"),
                        zulage_json=zulage_json_state,
                        zulage_xlsx_sonder=zulage_xlsx_sonder,
                        zulage_xlsx_fuengers=zulage_xlsx_fuengers,
                        drittkunden_json=drittkunden_json_state,
                        zulage_xlsx_drittkunden=zulage_xlsx_drittkunden,
                        fahrzeugwaesche_json=st.session_state.get("fahrzeugwaesche_json", "[]"),
                        tanken_json=st.session_state.get("tanken_json", "[]"),
                        verstoss_json=st.session_state.get("verstoss_json", '{"drivers":[],"total_violations":0}'),
                        spesen_json=st.session_state.get("spesen_json", '{"drivers":[],"months":[],"total_cost":0,"total_rows":0}'),
                        grosskunden_json=st.session_state.get("grosskunden_json", "[]"),
                        timerec_json=st.session_state.get("timerec_json", "{}"),
                        spediteure_json=st.session_state.get("spediteure_json", '{"katalog":[],"fahrten":[]}'),
                        fahrerbewertung_json=st.session_state.get("fahrerbewertung_json", '{"profile":"","event_types":[],"g_months":{},"g_ev":{},"drivers":[]}'),
                        versp_abfahrt_json="{}",
                        last_updated=generated_at.strftime("Stand: %d.%m.%Y %H:%M"),
                        generation_meta=generation_meta,
                    )

                _build_progress.progress(78, text="4/5 HTML-Datei auf dem Server speichern")

                # Direkt auf die Platte schreiben. TextIOWrapper kodiert
                # schrittweise, sodass keine zweite vollständige Byte-Kopie
                # der großen HTML im RAM entsteht.
                import tempfile as _tempfile
                with _tempfile.NamedTemporaryFile(
                    mode="w",
                    encoding="utf-8",
                    newline="",
                    suffix=".html",
                    prefix="nfc_suche_",
                    delete=False,
                ) as _tmp_html:
                    _tmp_html.write(app_html)
                    generated_html_path = _tmp_html.name
                del app_html

                _build_progress.progress(92, text="5/5 Dateigröße und Download vorbereiten")
                generation_seconds = time.perf_counter() - _generation_started
                generated_html_size = Path(generated_html_path).stat().st_size
                _pdf_count, _pdf_bytes = _embedded_pdf_stats()
                generated_meta = {
                    "created_at": generated_at,
                    "generation_seconds": generation_seconds,
                    "html_size": generated_html_size,
                    "pdf_count": _pdf_count,
                    "pdf_bytes": _pdf_bytes,
                    "week_names": [str(i.get("name", "")) for i in ready],
                    "source_signature": current_generation_signature,
                }
                st.session_state["_generated_app_html_path"] = generated_html_path
                st.session_state["_generated_app_html_meta"] = generated_meta
                _record_processing_success(
                    "download_html", "Gesamtdatei suche.html",
                    f"{_human_size(generated_html_size)}, {len(ready)} Woche(n), {generation_seconds:.2f} s",
                )
                _build_progress.progress(100, text=f"Fertig · {_human_size(generated_html_size)} in {generation_seconds:.2f} s")
                st.success("suche.html wurde erfolgreich erstellt und steht zum Download bereit.")
                _gc.collect()
            except Exception as exc:
                try:
                    if 'generated_html_path' in locals():
                        Path(generated_html_path).unlink(missing_ok=True)
                except Exception:
                    pass
                _clear_generated_html_file()
                _record_processing_error("download_html", "Gesamtdatei suche.html", exc)
                try:
                    _build_progress.progress(100, text="Erstellung abgebrochen – Fehlerdetails stehen unten im Status")
                except Exception:
                    pass
                st.error(f"suche.html konnte nicht erzeugt werden: {type(exc).__name__}: {exc}")
                _gc.collect()

        generated_html_path = st.session_state.get("_generated_app_html_path")
        generated_meta = st.session_state.get("_generated_app_html_meta", {}) or {}

        if generated_html_path and Path(generated_html_path).is_file():
            # Callable: Die Datei wird erst beim Klick gelesen und nicht dauerhaft
            # als zusätzlicher Bytes-Block im Streamlit-Session-Speicher gehalten.
            st.download_button(
                label="suche.html herunterladen",
                data=lambda p=generated_html_path: _read_generated_html_file(p),
                file_name="suche.html",
                mime="text/html",
                type="primary",
                width="stretch",
                on_click="ignore",
                key="download_suche_html",
            )

            st.caption(
                "Eigenständige Einzeldatei: Alle PDFs sind komprimiert in suche.html enthalten. "
                "Es wird kein zusätzlicher Ordner benötigt."
            )

            created_at = generated_meta.get("created_at")
            generation_seconds = float(generated_meta.get("generation_seconds", 0.0) or 0.0)
            html_size = int(generated_meta.get("html_size", Path(generated_html_path).stat().st_size) or Path(generated_html_path).stat().st_size)
            pdf_count = int(generated_meta.get("pdf_count", 0) or 0)
            pdf_bytes = int(generated_meta.get("pdf_bytes", 0) or 0)
            week_names = generated_meta.get("week_names", []) or []

            st.markdown("##### Generierungsstatistik")
            _g1, _g2, _g3, _g4 = st.columns(4)
            _g1.metric("Version", f"v{APP_DISPLAY_VERSION}")
            _g2.metric("HTML-Datei", _human_size(html_size))
            _g3.metric("Erzeugungszeit", f"{generation_seconds:.2f} s")
            _g4.metric("Eingebettete PDFs", f"{pdf_count} · {_human_size(pdf_bytes)}")
            if isinstance(created_at, datetime.datetime):
                created_label = created_at.strftime("%d.%m.%Y %H:%M:%S")
            else:
                created_label = str(created_at or "-")
            st.caption(
                f"Datenstand: {created_label} · "
                f"{len(week_names)} Woche(n): {', '.join(week_names)}"
            )
        else:
            st.info("Nach dem letzten Daten-Upload wurde noch keine aktuelle suche.html erstellt.")

        plane_zulagen_json = build_plane_zulagen_json(
            zulage_json_state,
            drittkunden_json_state,
            generated_at=datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
        )
        st.download_button(
            label="zulagen.json herunterladen",
            data=plane_zulagen_json,
            file_name="zulagen.json",
            mime="application/json",
            width="stretch",
            on_click="ignore",
        )
    else:
        st.button(
            "suche.html erstellen / aktualisieren",
            type="primary",
            width="stretch",
            key="build_suche_html_disabled",
            disabled=True,
        )
        st.info("Mindestens Logo, Schlüsseldatei und eine vollständig verarbeitete Wochen-Excel hochladen.")
        zulage_json_state      = st.session_state.get("zulage_json", "{}")
        drittkunden_json_state = st.session_state.get("drittkunden_json", "[]")
        if zulage_json_state not in ("{}", "") or drittkunden_json_state not in ("[]", ""):
            plane_zulagen_json = build_plane_zulagen_json(
                zulage_json_state,
                drittkunden_json_state,
                generated_at=datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
            )
            st.download_button(
                label="zulagen.json herunterladen",
                data=plane_zulagen_json,
                file_name="zulagen.json",
                mime="application/json",
                width="stretch",
                on_click="ignore",
            )

# === Vollständiger Reset =====================================================
st.divider()
with st.expander("App vollständig zurücksetzen", expanded=False):
    st.warning(
        "Dabei werden alle Uploads, verarbeiteten Daten, Statusmeldungen, Caches "
        "und die temporär erzeugte suche.html dieser Sitzung entfernt."
    )
    _reset_confirmed = st.checkbox(
        "Ich möchte wirklich alle Daten dieser Sitzung löschen.",
        key="_confirm_full_app_reset",
    )
    if st.button(
        "Alle Daten und Uploads zurücksetzen",
        width="stretch",
        disabled=not _reset_confirmed,
        key="full_app_reset",
    ):
        _reset_all_app_data()
        st.rerun()

# === Zentrale Verarbeitungsanzeige ===========================================
st.divider()
_status_entries = list(st.session_state.get("_processing_status", {}).values())
_error_entries = list(st.session_state.get("_processing_errors", []))
_error_count = sum(1 for row in _status_entries if row.get("status") == "error")
_warning_count = sum(1 for row in _status_entries if row.get("status") == "warning")
_ok_count = sum(1 for row in _status_entries if row.get("status") == "ok")

with st.expander(
    f"Verarbeitungsstatus: {_ok_count} erfolgreich, {_warning_count} Hinweise, {_error_count} Fehler",
    expanded=bool(_error_count),
):
    if not _status_entries:
        st.info("Noch keine Dateien verarbeitet.")
    else:
        _status_label = {
            "ok": "✓ Erfolgreich",
            "warning": "⚠ Hinweis",
            "error": "✗ Fehler",
            "info": "• Vorhanden",
        }
        _rows = []
        for row in sorted(_status_entries, key=lambda x: (x.get("bereich", "").lower(), x.get("zeit", ""))):
            _rows.append({
                "Status": _status_label.get(row.get("status"), row.get("status", "")),
                "Bereich": row.get("bereich", ""),
                "Datei": row.get("datei", ""),
                "Details": row.get("detail", ""),
                "Zeit": row.get("zeit", ""),
            })
        st.dataframe(_rows, width="stretch", hide_index=True)

    if _error_entries:
        with st.expander("Technische Fehlerdetails", expanded=False):
            for err in reversed(_error_entries[-20:]):
                st.markdown(f"**{err.get('bereich', 'Unbekannter Bereich')}** — {err.get('zeit', '')}")
                if err.get("datei"):
                    st.caption(f"Datei: {err['datei']}")
                st.code(f"{err.get('fehlertyp', 'Fehler')}: {err.get('fehler', '')}")

        _error_export = json.dumps(
            [
                {k: v for k, v in err.items() if k != "signature"}
                for err in _error_entries
            ],
            ensure_ascii=False,
            indent=2,
        )
        st.download_button(
            "Fehlerprotokoll herunterladen",
            data=_error_export.encode("utf-8"),
            file_name="nfc_generator_fehlerprotokoll.json",
            mime="application/json",
            width="stretch",
        )

    _status_col1, _status_col2 = st.columns(2)
    with _status_col1:
        if st.button("Fehlerhistorie leeren", width="stretch"):
            st.session_state["_processing_errors"] = []
            st.rerun()
    with _status_col2:
        if st.button("Statusanzeige zuruecksetzen", width="stretch"):
            st.session_state["_processing_status"] = {}
            st.session_state["_processing_errors"] = []
            st.rerun()



_boot_log("99 App-Skript vollständig ausgeführt")
